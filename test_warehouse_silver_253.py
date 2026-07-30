from datetime import datetime
from pathlib import Path
import zipfile

import pytest
from openpyxl import Workbook
from PIL import Image

from src.warehouse_management.silver import (
    SILVER_CIF_PERCENT,
    SILVER_DEFAULT_USD_RMB,
    is_silver_invoice,
    parse_silver_invoice,
)


RAW_CODES = [
    "2.6m皮绳管+9m水滴扣",
    "3m弯刀橄榄珠",
    "1.5*20车花s管",
    "1.5*20车花s管",
    "1.5*10车花直管",
    "1.5*10车花直管",
    "5.0轻生圈",
    "3x7舌形牌",
    "60x4开口圈",
    "5mm西瓜珠",
    "5mm猫眼（大孔）",
    "275车侧身-半成品",
    "275车侧身小正心万能",
    "30c车十字包方珠滴胶",
]


def _inject_wps_cell_images(path: Path, tmp_path: Path) -> None:
    first = tmp_path / "first.png"
    second = tmp_path / "second.png"
    Image.new("RGB", (60, 40), "white").save(first)
    Image.new("RGB", (60, 40), "gray").save(second)
    cell_images = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<etc:cellImages xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">
 <etc:cellImage><xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="ID_TEST_ONE"/></xdr:nvPicPr><xdr:blipFill><a:blip r:embed="rId1"/></xdr:blipFill></xdr:pic></etc:cellImage>
 <etc:cellImage><xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="ID_TEST_TWO"/></xdr:nvPicPr><xdr:blipFill><a:blip r:embed="rId2"/></xdr:blipFill></xdr:pic></etc:cellImage>
</etc:cellImages>'''
    rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
 <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/test1.png"/>
 <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/test2.png"/>
</Relationships>'''
    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/cellimages.xml", cell_images)
        archive.writestr("xl/_rels/cellimages.xml.rels", rels)
        archive.write(first, "xl/media/test1.png")
        archive.write(second, "xl/media/test2.png")


def _raw_fixture(path: Path, tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "TIAN YI DA JEWELLERY Co.,LTD"
    ws["I6"] = "DATE:"
    ws["J6"] = datetime(2026, 6, 30)
    headers = [
        "NO", "Photo", "Code", "Plating", "Size", "Quantity", "Weight",
        "silver/g", "labour/g", "price/g", "Amount",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(7, col, value)
    for line, code in enumerate(RAW_CODES, 1):
        row = 7 + line
        quantity = "1000 meters" if line == 12 else 100 + line
        weight = 100.0 + line
        price = 20.0 + line / 10
        values = [line, None, code, "rhodium", "5mm", quantity, weight, 15.4, price - 15.4, price, weight * price]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    ws["B8"] = '=_xlfn.DISPIMG("ID_TEST_ONE",1)'
    ws["B9"] = '=_xlfn.DISPIMG("ID_TEST_TWO",1)'
    wb.save(path)
    _inject_wps_cell_images(path, tmp_path)


def test_raw_30_june_invoice_is_classified_but_not_received(tmp_path: Path) -> None:
    path = tmp_path / "silver_raw.xlsx"
    _raw_fixture(path, tmp_path)
    assert is_silver_invoice(path)

    products, meta = parse_silver_invoice(path, tmp_path / "images")

    assert len(products) == 14
    assert meta.invoice_date == "2026-06-30"
    assert meta.auto_receive is False
    assert meta.purchase_price_calculated is True
    assert all(not product.received for product in products)
    assert all(product.actual_manual is None for product in products)
    assert all(product.waiting_qty == product.qty_document for product in products)
    assert products[0].silver_category == "Замки и концевики"
    assert products[7].silver_category == "Бирки / декоративные элементы"
    assert products[8].silver_category == "Кольца и соединители"
    assert products[11].qty_document == 1000
    assert products[11].unit_label == "м"
    assert products[13].silver_category == "Готовые цепочки / основы"

    expected = products[0].amount_rmb / SILVER_DEFAULT_USD_RMB / products[0].qty_document
    assert products[0].purchase_usd_per_unit == pytest.approx(expected)


def test_wps_dispimg_photos_are_extracted(tmp_path: Path) -> None:
    path = tmp_path / "silver_raw.xlsx"
    _raw_fixture(path, tmp_path)
    products, _ = parse_silver_invoice(path, tmp_path / "images")
    assert Path(products[0].image_path).is_file()
    assert Path(products[1].image_path).is_file()
