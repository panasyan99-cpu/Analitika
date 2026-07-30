from pathlib import Path

import pytest

from openpyxl import Workbook

from src.warehouse_management.silver import (
    SILVER_DEFAULT_COEFFICIENT,
    SILVER_DEFAULT_USD_VND,
    is_silver_invoice,
    parse_silver_invoice,
    silver_sale_vnd,
)


def _silver_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "TIAN YI DA JEWELLERY Co.,LTD"
    ws["L1"] = "USD/RMB"
    ws["M1"] = "K"
    ws["N1"] = "USD/VND"
    ws["L2"] = 6.71
    ws["M2"] = 10
    ws["N2"] = 26500
    ws["I6"] = "DATE:"
    ws["J6"] = 46221
    headers = [
        "NO", "Photo", "Code", "Plating", "Size", "Quantity", "Weight",
        "silver/g", "labour/g", "price/g", "Amount", "CIF price usd",
        "Sell price USD", "Sell price VND",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(7, col, value)
    for line in range(1, 19):
        row = 7 + line
        quantity = 5173 if line == 2 else 100 + line
        purchase = 2.4197229510213383 if line == 2 else 1.0 + line / 10
        values = [
            line, None, f"original-{line}", "rhodium", "15mm", quantity, 100.0 + line,
            15.4, 8.6, 24.0, 1000.0 + line, purchase, purchase * 10,
            silver_sale_vnd(purchase, 26500, 10),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    wb.save(path)


def test_silver_invoice_parser_knows_pairs_and_fixed_purchase(tmp_path: Path) -> None:
    path = tmp_path / "silver.xlsx"
    _silver_fixture(path)
    assert is_silver_invoice(path)
    products, meta = parse_silver_invoice(path, tmp_path / "images")
    assert len(products) == 18
    puset = products[1]
    assert puset.qty_document == 5173
    assert puset.unit_label == "пара"
    assert puset.silver_925 is True
    assert puset.purchase_usd_per_unit == pytest.approx(2.4197229510213383)
    assert meta.usd_vnd == SILVER_DEFAULT_USD_VND
    assert meta.coefficient == SILVER_DEFAULT_COEFFICIENT
    assert silver_sale_vnd(puset.purchase_usd_per_unit, 26500, 10) == 642000


def test_silver_invoice_business_classification(tmp_path: Path) -> None:
    path = tmp_path / "silver.xlsx"
    _silver_fixture(path)
    products, _ = parse_silver_invoice(path, tmp_path / "images")
    assert products[10].name == "Серебряная цепочка 50 см"
    assert products[10].sellable is True
    assert all(product.silver_925 for product in products)
    assert products[11].silver_category == "Основы для браслетов"
    assert products[14].silver_category == "Основы для ожерелий"


def test_silver_schema_and_ui_are_additive() -> None:
    schema = Path("src/warehouse_management/schema.py").read_text(encoding="utf-8")
    ui = Path("src/warehouse_management/ui.py").read_text(encoding="utf-8")
    app = Path("streamlit_app.py").read_text(encoding="utf-8")
    assert 'component("Закупка USD/ед."' in schema
    assert 'line("Сумма RMB"' in schema
    assert 'line("Курс USD/VND при импорте"' in schema
    assert '"Серебро 925"' in ui
    assert 'key="warehouse_silver_usd_vnd"' in ui
    assert 'key="warehouse_silver_coefficient"' in ui
    settings_body = app[app.index("def render_report_settings"):app.index("def render_mode_help_page")]
    assert '"Сувениры и касты на складе"' in settings_body
