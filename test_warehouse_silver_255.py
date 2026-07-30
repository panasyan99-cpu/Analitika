from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook

from src.warehouse_management.packing import export_master
from src.warehouse_management.silver import (
    SILVER_DEFAULT_COEFFICIENT,
    SILVER_DEFAULT_USD_RMB,
    SILVER_DEFAULT_USD_VND,
    parse_silver_invoice,
    silver_sale_vnd,
)


RAW_PATH = Path("/mnt/data/инвойс на серебро 30.06.2026г(1).xlsx")

EXPECTED_RAW = [
    (1, 200, 2415.504, 477_000),
    (2, 1550, 2715.57, 70_000),
    (3, 2068, 7416.0, 142_000),
    (4, 990, 3590.168, 144_000),
    (5, 1944, 3155.526, 65_000),
    (6, 986, 1542.42, 62_000),
    (7, 3183, 4117.509, 52_000),
    (8, 2055, 2905.301, 56_000),
    (9, 2973, 1769.604, 24_000),
    (10, 2624, 9062.56, 137_000),
    (11, 1061, 2453.724, 92_000),
    (12, 1000, 43891.2, 1_734_000),
    (13, 1999, 52385.06, 1_035_000),
    (14, 200, 12133.88, 2_397_000),
]


def _enriched_fixture(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "TIAN YI DA JEWELLERY Co.,LTD"
    ws["L1"] = "USD/RMB"
    ws["M1"] = "K"
    ws["N1"] = "USD/VND"
    ws["L2"] = SILVER_DEFAULT_USD_RMB
    ws["M2"] = SILVER_DEFAULT_COEFFICIENT
    ws["N2"] = SILVER_DEFAULT_USD_VND
    ws["I6"] = "DATE:"
    ws["J6"] = 46221
    headers = [
        "NO", "Photo", "Code", "Plating", "Size", "Quantity", "Weight",
        "silver/g", "labour/g", "price/g", "Amount", "CIF price usd",
        "Sell price USD", "Sell price VND",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(7, col, value)
    for line in range(1, 19):
        row = 7 + line
        quantity = 5173 if line == 2 else 100 + line
        purchase = 2.4197229510213383 if line == 2 else 1.0 + line / 10
        sale_vnd = silver_sale_vnd(purchase, SILVER_DEFAULT_USD_VND, SILVER_DEFAULT_COEFFICIENT)
        values = [
            line, None, f"original-{line}", "rhodium", "15mm", quantity, 100.0 + line,
            15.4, 8.6, 24.0, 1000.0 + line, purchase,
            purchase * SILVER_DEFAULT_COEFFICIENT, sale_vnd,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    wb.save(path)


@pytest.mark.skipif(not RAW_PATH.exists(), reason="real 30.06 supplier invoice is not mounted")
def test_real_raw_invoice_price_formula_and_master(tmp_path: Path) -> None:
    products, meta = parse_silver_invoice(RAW_PATH, tmp_path / "raw_images")

    assert len(products) == 14
    assert sum(p.qty_document for p in products) == 22_833
    assert sum(bool(p.image_path) for p in products) == 14
    assert sum(p.total_weight_g or 0 for p in products) == pytest.approx(7_318.51)
    assert sum(p.amount_rmb or 0 for p in products) == pytest.approx(149_554.026)
    assert meta.cif_percent == 0
    assert meta.auto_receive is False

    for product, (line, qty, amount, expected_sale_vnd) in zip(products, EXPECTED_RAW):
        assert product.number == line
        assert product.qty_document == qty
        assert product.amount_rmb == pytest.approx(amount)
        expected_purchase = amount / SILVER_DEFAULT_USD_RMB / qty
        assert product.purchase_usd_per_unit == pytest.approx(expected_purchase)
        assert product.invoice_sale_usd == pytest.approx(expected_purchase * 10)
        assert product.invoice_sale_vnd == expected_sale_vnd
        assert product.received is False
        assert product.actual_manual is None

    master = tmp_path / "raw_master.xlsx"
    export_master(master, products)
    assert master.is_file() and master.stat().st_size > 0


def test_enriched_invoice_preserves_fixed_purchase_and_stays_expected(tmp_path: Path) -> None:
    source = tmp_path / "silver_18_07.xlsx"
    _enriched_fixture(source)
    products, meta = parse_silver_invoice(source, tmp_path / "enriched_images")

    assert len(products) == 18
    assert meta.source_variant == "enriched_2026_07_18"
    assert meta.purchase_price_calculated is False
    assert meta.auto_receive is False
    assert products[1].purchase_usd_per_unit == pytest.approx(2.4197229510213383)
    assert products[1].invoice_sale_vnd == 642_000
    assert all(p.received is False and p.actual_manual is None for p in products)

    master = tmp_path / "enriched_master.xlsx"
    export_master(master, products)
    assert master.is_file() and master.stat().st_size > 0
