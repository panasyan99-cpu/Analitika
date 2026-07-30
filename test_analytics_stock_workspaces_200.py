from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import streamlit_app as app


def _base_book(with_stock: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"
    ws.cell(1, 1, "Отчет о продажах товаров за период 01.07.2026 - 30.07.2026")
    ws.cell(4, 1, "Магазин; Камень/вставка; Проба; Номенклатурная группа; Поставщик")
    ws.cell(4, 8, "Продано")
    ws.cell(4, 10, "Возврат")
    ws.cell(5, 8, "Кол-во")
    ws.cell(5, 9, "Сумма")
    ws.cell(5, 10, "Кол-во")
    ws.cell(5, 11, "Сумма")
    if with_stock:
        ws.cell(5, 12, "Остаток")
    return wb


def _row(ws, row, text, indent, qty=0, amount=0, stock=None, bold=False):
    cell = ws.cell(row, 1, text)
    cell.alignment = Alignment(indent=indent)
    cell.font = Font(bold=bold)
    ws.cell(row, 8, qty)
    ws.cell(row, 9, amount)
    if stock is not None:
        ws.cell(row, 12, stock)


def test_new_sales_hierarchy_uses_supplier_leaves_once(tmp_path: Path):
    wb = _base_book(False)
    ws = wb.active
    _row(ws, 7, "AB", 0, 3, 300, bold=True)
    _row(ws, 8, "BLUE SAPPHIRE", 2, 3, 300)
    _row(ws, 9, "B 925", 4, 3, 300)
    _row(ws, 10, "Earrings", 6, 3, 300)
    _row(ws, 11, "Поставщики", 5, 3, 300, bold=True)
    _row(ws, 12, "Taiwan", 6, 2, 220)
    _row(ws, 13, "Princess Jewelry", 6, 1, 80)
    path = tmp_path / "sales.xlsx"
    wb.save(path)

    detail, period = app.parse_supplier_report_with_period(path)
    assert period[0].strftime("%d.%m.%Y") == "01.07.2026"
    assert detail["Количество"].sum() == 3
    assert detail["Выручка"].sum() == 300
    assert set(detail["Поставщик"]) == {"Taiwan", "Princess Jewelry"}
    assert set(detail["Номенклатурная группа"]) == {"Серьги"}
    assert not app.report_has_stock(detail)


def test_new_stock_hierarchy_uses_product_stock_without_subtotal_duplicates(tmp_path: Path):
    wb = _base_book(True)
    ws = wb.active
    _row(ws, 7, "AB", 0, 5, 500, 100, bold=True)
    _row(ws, 8, "BLUE SAPPHIRE", 2, 5, 500, 100)
    _row(ws, 9, "B 925", 4, 5, 500, 100)
    _row(ws, 10, "Earrings", 6, 3, 330, 40)
    _row(ws, 11, "Ring", 6, 2, 170, 60)
    path = tmp_path / "stock.xlsx"
    wb.save(path)

    detail, _ = app.parse_supplier_report_with_period(path)
    assert detail["Количество"].sum() == 5
    assert detail["Выручка"].sum() == 500
    assert detail["Остаток"].sum() == 100
    assert app.report_has_stock(detail)
    summary = app.sales_stock_summary(detail, ["Номенклатурная группа"])
    assert set(summary["Номенклатурная группа"]) == {"Серьги", "Кольца"}
    assert summary["Остаток"].sum() == 100


def test_compact_workspaces_and_instruction_images_are_in_repository():
    source = Path("streamlit_app.py").read_text(encoding="utf-8")
    assert '("Сводка", "Магазины", "Ассортимент", "Остатки", "Поставщики")' in source
    assert '("Итог", "Магазины", "Ассортимент", "Остатки", "Поставщики")' in source
    assert '("Камни", "Пробы", "Номенклатурные группы")' in source
    assert '("По магазинам", "По номенклатурным группам", "По камням")' in source
    assert "analytics_store_setup.png" in source
    assert "analytics_hierarchy_setup.png" in source
    assert "sonu_report_setup.png" in source
    for name in (
        "assets/analytics_store_setup.png",
        "assets/analytics_hierarchy_setup.png",
        "assets/sonu_report_setup.png",
    ):
        assert Path(name).is_file()


def test_user_facing_model_wording_is_present():
    source = Path("streamlit_app.py").read_text(encoding="utf-8")
    order = Path("src/order_workflow.py").read_text(encoding="utf-8")
    warehouse = Path("src/warehouse.py").read_text(encoding="utf-8")
    assert '"SKU": "Модель"' in source
    assert "Моделей в блоке" in order
    assert "Заказано моделей" in order
    assert "Найдено моделей" in warehouse
