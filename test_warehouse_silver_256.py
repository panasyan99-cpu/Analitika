from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.warehouse_management.models import Product, SupplySummary
from src.warehouse_management.service import WarehouseService
from src.warehouse_management.silver import (
    SILVER_DEFAULT_COEFFICIENT,
    SILVER_DEFAULT_USD_VND,
    parse_silver_invoice,
    refresh_calculated_silver_prices,
)


RAW_PATH = Path("/mnt/data/инвойс на серебро 30.06.2026г(1).xlsx")


@dataclass
class FakeConfig:
    souvenirs_table_id: int = 642
    components_table_id: int = 643
    operations_table_id: int = 644
    supplies_table_id: int = 645
    supply_lines_table_id: int = 646


class FakeClient:
    def __init__(self) -> None:
        self.config = FakeConfig()
        self.rows = {642: [], 643: [], 644: [], 645: [], 646: []}
        self.counter = 1000

    def _id(self) -> int:
        self.counter += 1
        return self.counter

    def list_rows(self, table_id: int):
        return [dict(row) for row in self.rows[table_id]]

    def batch_id(self, prefix: str) -> str:
        return f"{prefix}-TEST"

    def create_row(self, table_id: int, payload):
        row = {"id": self._id(), **payload}
        self.rows[table_id].append(row)
        return dict(row)

    def batch_create(self, table_id: int, payloads):
        return [self.create_row(table_id, payload) for payload in payloads]

    def batch_update(self, table_id: int, payloads):
        for payload in payloads:
            row_id = int(payload["id"])
            target = next(row for row in self.rows[table_id] if int(row["id"]) == row_id)
            target.update(payload)
        return payloads

    def upload_file(self, path: Path):
        return {"name": path.name, "visible_name": path.name}

    def create_operations(self, payloads, *, batch_id: str, command_id: str = ""):
        return [self.create_row(self.config.operations_table_id, payload) for payload in payloads]

    def mark_operations_status(self, rows, status: str):
        for row in rows:
            row["Статус операции"] = status


@pytest.mark.skipif(not RAW_PATH.exists(), reason="real 30.06 invoice is not mounted")
def test_real_raw_invoice_creates_expected_supply_then_receives_by_recount(tmp_path: Path) -> None:
    products, _ = parse_silver_invoice(RAW_PATH, tmp_path / "images")
    for index, product in enumerate(products, start=1):
        product.sku = f"SIL{index:06d}"

    client = FakeClient()
    service = WarehouseService(client)
    result = service.create_supply_from_products(
        supply_id="SIL-TEST-001",
        supplier="TIAN YI DA JEWELLERY Co.,LTD",
        invoice=RAW_PATH.name,
        comment="товар ожидается",
        products=products,
        section="Комплектующие",
        command_id="IMPORT-TEST",
    )

    assert result["received"] == 0
    assert result["waiting"] == 22_833
    assert client.rows[644] == []
    assert len(client.rows[643]) == 14
    assert len(client.rows[646]) == 14
    assert all(line["Принято, шт."] == 0 and line["Статус"] == "Ожидается" for line in client.rows[646])
    assert all("Категория" not in row for row in client.rows[643])
    assert client.rows[646][0]["Закупка USD/ед."] == pytest.approx(2415.504 / 6.71 / 200)
    assert client.rows[646][0]["Продажа VND при импорте"] == 477_000

    supply_row = client.rows[645][0]
    summary = SupplySummary(
        row_id=int(supply_row["id"]),
        supply_id="SIL-TEST-001",
        date="2026-07-30",
        supplier="TIAN YI DA JEWELLERY Co.,LTD",
        status="Ожидается",
        sku_total=14,
        sku_received=0,
        qty_document=22_833,
        qty_received=0,
        qty_waiting=22_833,
        raw=supply_row,
    )
    first_line = client.rows[646][0]
    second_line = client.rows[646][1]
    receipt = service.receive_supply(
        summary,
        {int(first_line["id"]): 199, int(second_line["id"]): 1550},
        command_id="REC-TEST",
    )
    assert receipt["quantity"] == 1749
    assert len(client.rows[644]) == 2
    assert first_line["Принято, шт."] == 199
    assert first_line["Статус"] == "Частично получена"
    assert second_line["Принято, шт."] == 1550
    assert second_line["Статус"] == "Получена полностью"


def test_raw_price_history_refreshes_after_operator_purchase_edit() -> None:
    product = Product(
        number=1,
        boxes="",
        sku="SIL000001",
        qty_document=1,
        description="test",
        category="",
        material="Silver",
        stone="",
        color="",
        unit_weight_kg=None,
        image_path="",
        purchase_usd_per_unit=2.0,
        invoice_sale_usd=999.0,
        invoice_sale_vnd=999,
    )
    refresh_calculated_silver_prices(
        [product],
        usd_vnd=SILVER_DEFAULT_USD_VND,
        coefficient=SILVER_DEFAULT_COEFFICIENT,
    )
    assert product.invoice_sale_usd == pytest.approx(20.0)
    assert product.invoice_usd_vnd_rate == 26_500
    assert product.invoice_coefficient == 10
    assert product.invoice_sale_vnd == 530_000


def test_raw_invoice_blocks_broken_supplier_price_arithmetic(tmp_path: Path) -> None:
    path = tmp_path / "broken_raw.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "TIAN YI DA JEWELLERY Co.,LTD"
    headers = [
        "NO", "Photo", "Code", "Plating", "Size", "Quantity", "Weight",
        "silver/g", "labour/g", "price/g", "Amount",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(7, col, value)
    codes = [
        "2.6m皮绳管+9m水滴扣", "3m弯刀橄榄珠", "1.5*20车花s管", "1.5*20车花s管",
        "1.5*10车花直管", "1.5*10车花直管", "5.0轻生圈", "3x7舌形牌",
        "60x4开口圈", "5mm西瓜珠", "5mm猫眼（大孔）", "275车侧身-半成品",
        "275车侧身小正心万能", "30c车十字包方珠滴胶",
    ]
    for line, code in enumerate(codes, start=1):
        row = 7 + line
        values = [line, None, code, "rhodium", "5mm", 100, 10.0, 15.4, 4.6, 20.0, 200.0]
        if line == 4:
            values[-1] = 201.0
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value)
    wb.save(path)

    with pytest.raises(ValueError, match=r"строка 4: Amount 201"):
        parse_silver_invoice(path, tmp_path / "images")
