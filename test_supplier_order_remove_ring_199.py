from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    _clear_item_order_state,
    build_supplier_excel,
)


def _ring() -> OrderItem:
    return OrderItem(
        row=12,
        set_id="Set# 1",
        sku="SKR25N212B-BS",
        stone="Blue Sapphire",
        group="Ring",
        sales=5,
        stock_63=0,
        stock_20=0,
        stores={},
        total_stock=5,
        working_stock=5,
        ntr2_stock=0,
        ntr2_calculated=False,
        tvp_raw=0,
        stock_tt=0,
    )


def _parsed(tmp_path: Path, item: OrderItem) -> ParsedOrderWorkbook:
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    return ParsedOrderWorkbook(
        source_name=source.name,
        source_hash="hash",
        upload_path=str(source),
        period="",
        supplier="Y&J",
        store_columns=(),
        has_actual_ntr2=True,
        items=(item,),
    )


def test_remove_ring_clears_order_and_excludes_it_from_main_excel(tmp_path: Path) -> None:
    item = _ring()
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders[item.key] = 6
    draft.sizes[item.key] = {"18": 2, "19": 2, "20": 2}
    draft.stock_checked[item.key] = True
    draft.manual_edit[item.key] = True

    _clear_item_order_state(draft, item)

    assert draft.orders[item.key] == 0
    assert item.key not in draft.sizes
    assert item.key not in draft.stock_checked
    assert item.key not in draft.manual_edit

    workbook = load_workbook(io.BytesIO(build_supplier_excel(_parsed(tmp_path, item), [item], draft)))
    sheet = workbook["Order"]
    assert sheet.max_row == 1
    assert item.sku not in [cell.value for row in sheet.iter_rows() for cell in row]


def test_ring_size_card_has_remove_button_and_immediate_rerun() -> None:
    source = Path(workflow.__file__).read_text(encoding="utf-8")
    assert '"Удалить из заказа"' in source
    assert '_order_action_key("remove_from_order", item, mode, parsed.source_hash)' in source
    assert "_clear_item_order_state(draft, item)" in source
    assert "_save_session_draft(draft)" in source
    assert "st.rerun()" in source
