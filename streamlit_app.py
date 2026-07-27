from __future__ import annotations

import gc
import hashlib
import re
import tempfile
import threading
from html import escape
from pathlib import Path
from typing import Iterable

import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from openpyxl import load_workbook
from src.warehouse import render_warehouse_dashboard
from src.sonu import render_sonu_order_dashboard
from src.order_workflow import render_supplier_order_dashboard
from src.app_meta import APP_VERSION
from src.currency import get_vnd_per_usd, render_global_fx_control, vnd_to_usd
from src.product_info import REPORT_MODES, feature_cards_html, release_history_html

from src.report import (
    COLORED_ORDER,
    PEARL_ORDER,
    PRODUCT_ORDER,
    SEG_ORDER,
    TOP_ORDER,
    StoreData,
    SKIP_PRODUCTS,
    build_report_units,
    classify,
    extract_period,
    norm_product,
    normalize_store_from_report,
    totals_for,
)

SEGMENT_LABELS = {
    "TOP STONES": "Top Stones",
    "PEARLS": "Pearls",
    "COLORED STONES": "Other Stones",
}
SEGMENT_COLORS = {
    "TOP STONES": "#7030A0",
    "PEARLS": "#D3A338",
    "COLORED STONES": "#548235",
}
LIGHT_COLORS = {
    "TOP STONES": "#E9DDF1",
    "PEARLS": "#F5E7B8",
    "COLORED STONES": "#DDE8D4",
}
STONE_ORDERS = {
    "TOP STONES": TOP_ORDER,
    "PEARLS": PEARL_ORDER,
    "COLORED STONES": COLORED_ORDER,
}
PRODUCT_LABELS = {
    "Earrings": "Серьги",
    "Ring": "Кольца",
    "Pendant": "Подвески",
    "Bracelet": "Браслеты",
    "Necklace": "Ожерелья",
    "Brooch": "Броши",
    "Pearl Necklace": "Жемчужные нити",
    "Pearl Bracelet": "Жемчужные браслеты",
    "Pearl Chain": "Жемчуг на цепочке",
    "Stone": "Камни",
    "Other": "Другое",
}

METAL_GROUPS: tuple[str, ...] = ("Серебро", "Золото и платина", "Другое")
METAL_GROUP_COLORS = {
    "Серебро": "#aeb7c2",
    "Золото и платина": "#b7893f",
    "Другое": "#7d6f61",
}


# Plotly remains informative but cannot be accidentally changed on touch devices.
# Hover/tap tooltips stay enabled; zooming, panning, selection, editing and export
# controls are disabled. Streamlit dataframes intentionally remain interactive.
LOCKED_CHART_CONFIG = {
    "displayModeBar": False,
    "displaylogo": False,
    "scrollZoom": False,
    "doubleClick": False,
    "showTips": False,
    "editable": False,
    "staticPlot": False,
    "responsive": True,
    "showAxisDragHandles": False,
    "showAxisRangeEntryBoxes": False,
}


def lock_chart_interactions(fig: go.Figure) -> go.Figure:
    """Return a view-only Plotly figure while preserving hover/tap tooltips."""
    fig.update_layout(
        dragmode=False,
        clickmode="event",
        hovermode="closest",
        legend_itemclick=False,
        legend_itemdoubleclick=False,
    )

    cartesian_types = {
        "bar", "scatter", "scattergl", "box", "violin", "histogram",
        "histogram2d", "heatmap", "contour", "waterfall", "funnel",
        "candlestick", "ohlc",
    }
    if any(getattr(trace, "type", "") in cartesian_types for trace in fig.data):
        fig.update_xaxes(fixedrange=True)
        fig.update_yaxes(fixedrange=True)
    return fig


def polish_chart_surface(fig: go.Figure) -> go.Figure:
    """Apply one calm visual language to every chart without changing its data."""
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,.72)",
        font=dict(family="Inter, Arial, sans-serif", color="#2b261f", size=12),
        title_font=dict(family="Georgia, Times New Roman, serif", color="#211a12", size=18),
        margin=dict(l=18, r=18, t=58, b=24),
        hoverlabel=dict(bgcolor="#18130d", bordercolor="#c99545", font_color="#fffaf1"),
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(85,68,46,.08)", zeroline=False, linecolor="rgba(85,68,46,.14)")
    fig.update_yaxes(showgrid=False, zeroline=False, linecolor="rgba(85,68,46,.14)")
    return fig


def locked_plotly_chart(fig: go.Figure, *, width: str = "stretch", key: str | None = None) -> None:
    """Render a locked chart without changing dataframe/table behaviour."""
    st.plotly_chart(
        lock_chart_interactions(polish_chart_surface(fig)),
        width=width,
        key=key,
        config=LOCKED_CHART_CONFIG,
    )


class StoredUpload:
    """Persistent in-session representation of an uploaded file.

    Streamlit removes widget-owned values when a file uploader is no longer
    rendered. Keeping immutable bytes under a separate session key lets users
    navigate across pages without uploading the file again.
    """

    def __init__(self, name: str, data: bytes) -> None:
        self.name = name
        self._data = data

    def getvalue(self) -> bytes:
        return self._data


def persist_uploads(uploaded_files) -> None:
    if uploaded_files:
        payloads = [
            {"name": item.name, "data": bytes(item.getvalue())}
            for item in uploaded_files
        ]
        previous = st.session_state.get("uploaded_payloads", [])
        previous_signature = [(x.get("name"), len(x.get("data", b""))) for x in previous]
        new_signature = [(x["name"], len(x["data"])) for x in payloads]
        if previous_signature != new_signature:
            st.session_state["uploaded_payloads"] = payloads


def saved_uploads() -> list[StoredUpload]:
    return [
        StoredUpload(item["name"], item["data"])
        for item in st.session_state.get("uploaded_payloads", [])
    ]


def clear_saved_uploads() -> None:
    st.session_state.pop("uploaded_payloads", None)
    st.session_state.pop("upload_widget", None)
    st.session_state.pop("report_cache_signature", None)
    st.session_state.pop("report_cache_stores", None)
    st.session_state.pop("report_cache_errors", None)
    st.session_state.pop("report_cache_suppliers", None)


def uploads_signature(uploaded_files: list[StoredUpload]) -> str:
    """Stable content signature used to reuse parsed report data across reruns."""
    digest = hashlib.sha256()
    for uploaded in uploaded_files:
        data = uploaded.getvalue()
        digest.update(uploaded.name.encode("utf-8", errors="replace"))
        digest.update(len(data).to_bytes(8, "big"))
        digest.update(data)
    return digest.hexdigest()


def persist_comparison_upload(slot: int, uploaded_file) -> None:
    """Keep one comparison workbook in session state without mixing it with the base report."""
    key = f"comparison_payload_{slot}"
    ready_key = "comparison_ready"
    if uploaded_file is None:
        return
    payload = {"name": uploaded_file.name, "data": bytes(uploaded_file.getvalue())}
    previous = st.session_state.get(key)
    previous_signature = (previous or {}).get("name"), len((previous or {}).get("data", b""))
    current_signature = payload["name"], len(payload["data"])
    if previous_signature != current_signature or (previous or {}).get("data") != payload["data"]:
        st.session_state[key] = payload
        st.session_state[ready_key] = False


def saved_comparison_upload(slot: int) -> StoredUpload | None:
    payload = st.session_state.get(f"comparison_payload_{slot}")
    if not payload:
        return None
    return StoredUpload(payload["name"], payload["data"])


def clear_comparison_uploads() -> None:
    for key in [
        "comparison_payload_1", "comparison_payload_2", "comparison_upload_1",
        "comparison_upload_2", "comparison_ready", "comparison_processing",
    ]:
        st.session_state.pop(key, None)
    # The cache is session-scoped, so this releases only the current user's
    # parsed workbooks and never disrupts another viewer.
    try:
        parse_report_bundle.clear()
    except NameError:
        pass


def single_upload_payload(upload: StoredUpload) -> tuple[tuple[str, bytes], ...]:
    return ((upload.name, upload.getvalue()),)


st.set_page_config(
    page_title="Analitika — Princess Jewelry",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="collapsed",
)


def _css() -> str:
    return """
<style>
:root {
  --gold: #b7893f;
  --gold-soft: #ead8b8;
  --ink: #111111;
  --muted: #6c6c6c;
  --line: #e9e4dc;
  --paper: #fbfaf8;
}
html, body, [class*="css"] { font-family: Inter, Arial, sans-serif; }
/* Remove Streamlit service chrome/status popovers from the product UI. */
[data-testid="stStatusWidget"],
[data-testid="stConnectionStatus"],
[data-testid="stAppDeployButton"],
[data-testid="stDecoration"],
[data-testid="stHeaderActionElements"],
[data-testid="stToolbarActions"],
[data-testid="stMainMenu"],
.stDeployButton,
div[class*="StatusWidget"],
div[class*="ConnectionStatus"] { display:none !important; }
#MainMenu, footer { visibility:hidden !important; }

/* Streamlit 1.59 uses stExpandSidebarButton / stSidebarCollapseButton.
   The toolbar itself must remain mounted; otherwise the user can collapse the
   navigation and permanently lose the native reopen control. */
[data-testid="stHeader"],
[data-testid="stToolbar"] {
  visibility:visible !important;
  opacity:1 !important;
  pointer-events:none !important;
  background:transparent !important;
}
/* The collapsed sidebar control must remain impossible to miss.
   Streamlit has used several test IDs/ARIA labels across releases, so the
   selector deliberately covers both current and legacy DOM variants. */
[data-testid="stExpandSidebarButton"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
  display:flex !important;
  visibility:visible !important;
  opacity:1 !important;
  pointer-events:auto !important;
  position:fixed !important;
  top:.75rem !important;
  left:.75rem !important;
  z-index:1000000 !important;
  width:auto !important;
  min-width:0 !important;
  background:transparent !important;
}
[data-testid="stExpandSidebarButton"] button,
[data-testid="stExpandSidebarButton"] [role="button"],
[data-testid="stSidebarCollapsedControl"] button,
[data-testid="stSidebarCollapsedControl"] [role="button"],
[data-testid="collapsedControl"] button,
[data-testid="collapsedControl"] [role="button"],
button[aria-label*="sidebar" i],
button[title*="sidebar" i],
[data-testid="stSidebarCollapseButton"] button,
[data-testid="stSidebarCollapseButton"] [role="button"] {
  display:flex !important;
  align-items:center !important;
  justify-content:center !important;
  gap:8px !important;
  width:auto !important;
  min-width:54px !important;
  height:50px !important;
  min-height:50px !important;
  padding:0 15px !important;
  border:2px solid #f1c774 !important;
  border-radius:14px !important;
  background:linear-gradient(135deg,#d7a94d 0%,#aa7025 56%,#75410d 100%) !important;
  color:#ffffff !important;
  box-shadow:0 0 0 4px rgba(215,169,77,.18),0 12px 30px rgba(78,43,8,.38) !important;
  pointer-events:auto !important;
  opacity:1 !important;
  filter:none !important;
  transition:transform .16s ease,box-shadow .16s ease,background .16s ease !important;
}
[data-testid="stExpandSidebarButton"] button::after,
[data-testid="stExpandSidebarButton"] [role="button"]::after,
[data-testid="stSidebarCollapsedControl"] button::after,
[data-testid="stSidebarCollapsedControl"] [role="button"]::after,
[data-testid="collapsedControl"] button::after,
[data-testid="collapsedControl"] [role="button"]::after,
button[aria-label*="open sidebar" i]::after,
button[title*="open sidebar" i]::after {
  content:"МЕНЮ" !important;
  display:inline-block !important;
  color:#ffffff !important;
  font-size:12px !important;
  font-weight:800 !important;
  letter-spacing:.08em !important;
  line-height:1 !important;
}
[data-testid="stExpandSidebarButton"] button:hover,
[data-testid="stExpandSidebarButton"] [role="button"]:hover,
[data-testid="stSidebarCollapsedControl"] button:hover,
[data-testid="stSidebarCollapsedControl"] [role="button"]:hover,
[data-testid="collapsedControl"] button:hover,
[data-testid="collapsedControl"] [role="button"]:hover,
button[aria-label*="sidebar" i]:hover,
button[title*="sidebar" i]:hover,
[data-testid="stSidebarCollapseButton"] button:hover,
[data-testid="stSidebarCollapseButton"] [role="button"]:hover {
  background:linear-gradient(135deg,#e4bb66 0%,#bd8434 56%,#87531a 100%) !important;
  border-color:#ffe0a0 !important;
  box-shadow:0 0 0 5px rgba(228,187,102,.24),0 15px 34px rgba(78,43,8,.44) !important;
  transform:translateY(-1px) scale(1.02) !important;
}
[data-testid="stExpandSidebarButton"] svg,
[data-testid="stExpandSidebarButton"] span,
[data-testid="stSidebarCollapsedControl"] svg,
[data-testid="stSidebarCollapsedControl"] span,
[data-testid="collapsedControl"] svg,
[data-testid="collapsedControl"] span,
button[aria-label*="sidebar" i] svg,
button[title*="sidebar" i] svg,
[data-testid="stSidebarCollapseButton"] svg,
[data-testid="stSidebarCollapseButton"] span {
  color:#ffffff !important;
  fill:#ffffff !important;
  stroke:#ffffff !important;
  opacity:1 !important;
}
[data-testid="stSidebarCollapseButton"] {
  visibility:visible !important;
  opacity:1 !important;
  pointer-events:auto !important;
}
@media (max-width:640px) {
  [data-testid="stExpandSidebarButton"],
  [data-testid="stSidebarCollapsedControl"],
  [data-testid="collapsedControl"] { top:.55rem !important; left:.55rem !important; }
  [data-testid="stExpandSidebarButton"] button,
  [data-testid="stExpandSidebarButton"] [role="button"],
  [data-testid="stSidebarCollapsedControl"] button,
  [data-testid="stSidebarCollapsedControl"] [role="button"],
  [data-testid="collapsedControl"] button,
  [data-testid="collapsedControl"] [role="button"],
  button[aria-label*="open sidebar" i],
  button[title*="open sidebar" i] {
    min-width:92px !important;
    height:46px !important;
    min-height:46px !important;
    padding:0 12px !important;
  }
}
.stApp {
  background:
    radial-gradient(circle at 72% 18%, rgba(230,212,183,.20), transparent 24%),
    linear-gradient(135deg, #ffffff 0%, #fbfaf8 72%, #f6f1e9 100%);
  color: var(--ink);
}
[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #090806 0%, #15110b 100%);
  border-right: 1px solid #3a2b16;
  color: #f5ead8;
}
[data-testid="stSidebar"] * { color: #f5ead8; }
[data-testid="stSidebar"] > div:first-child { padding-top: 1.2rem; }
.block-container { padding-top: 1.2rem; padding-bottom: 3rem; max-width: 1500px; }
.brand-card {
  border: 1px solid var(--line); border-radius: 18px; background: rgba(255,255,255,.92);
  padding: 22px 24px; box-shadow: 0 10px 35px rgba(34,24,9,.05); margin-bottom: 18px;
}
.brand-kicker { color: var(--gold); font-size: 12px; letter-spacing: .12em; text-transform: uppercase; font-weight: 700; }
.brand-title { font-family: Georgia, serif; font-size: 44px; margin: 4px 0 4px; color: #171411; }
.brand-subtitle { color: var(--muted); font-size: 15px; }
.upload-panel {
  border: 1px dashed #c9aa72; border-radius: 18px; background: rgba(255,255,255,.78);
  padding: 18px 22px; margin: 6px 0 20px;
}
.kpi-card {
  border: 1px solid var(--line); border-radius: 14px; background: rgba(255,255,255,.95);
  padding: 18px 18px 16px; min-height: 118px; box-shadow: 0 8px 25px rgba(34,24,9,.045);
  overflow: visible;
}
.kpi-label { color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .06em; }
.kpi-value {
  font-family: Georgia, serif; font-size: clamp(18px, 2vw, 29px); line-height: 1.15;
  font-weight: 700; color: #16120d; margin-top: 9px; white-space: normal;
  overflow-wrap: anywhere; word-break: normal;
}
.kpi-note { color: var(--gold); font-size: 12px; margin-top: 6px; }
.kpi-leader-metric {
  color:#6f4b16; font-size:clamp(16px, 1.45vw, 21px); line-height:1.2;
  font-weight:700; margin-top:8px; white-space:normal; overflow-wrap:anywhere;
}
.section-title { font-family: Georgia, serif; font-size: 30px; margin: 22px 0 10px; }
.section-divider {
  margin: 38px 0 18px; padding: 18px 22px; border-radius: 16px;
  background: linear-gradient(90deg, rgba(183,137,63,.14), rgba(255,255,255,.96) 45%, rgba(183,137,63,.08));
  border-top: 1px solid rgba(183,137,63,.55); border-bottom: 1px solid rgba(183,137,63,.28);
  box-shadow: 0 10px 28px rgba(34,24,9,.045);
}
.section-divider-kicker { color: var(--gold); font-size: 11px; font-weight: 800; letter-spacing: .16em; text-transform: uppercase; }
.section-divider-title { font-family: Georgia, serif; color: #17120c; font-size: 28px; margin-top: 4px; }
.section-divider-copy { color: var(--muted); font-size: 13px; margin-top: 5px; }
.analysis-panel {
  margin: 16px 0 26px; padding: 18px 20px; border-radius: 15px;
  background: rgba(255,255,255,.94); border: 1px solid var(--line);
  box-shadow: 0 9px 26px rgba(34,24,9,.04);
}
.analysis-panel-title { font-family: Georgia, serif; font-size: 20px; color: #6f4b16; margin-bottom: 8px; }
.analysis-line { padding: 8px 0; border-bottom: 1px solid #f0ece5; color: #28231d; }
.analysis-line:last-child { border-bottom: none; }
.insight {
  border-left: 4px solid var(--gold); background: rgba(255,255,255,.93); border-radius: 0 12px 12px 0;
  padding: 13px 15px; margin: 8px 0; border-top: 1px solid var(--line); border-right: 1px solid var(--line); border-bottom: 1px solid var(--line);
}
.filter-panel {
  border: 1px solid var(--line); border-radius: 15px; background: rgba(255,255,255,.92);
  padding: 14px 16px 4px; margin: 8px 0 14px; box-shadow: 0 8px 22px rgba(34,24,9,.035);
}
.st-key-global_metal_filter {
  border:1px solid rgba(183,137,63,.55); border-radius:16px;
  background:linear-gradient(135deg,rgba(255,253,249,.98),rgba(242,225,193,.88));
  padding:16px 18px 14px; margin:-3px 0 20px;
  box-shadow:0 10px 28px rgba(95,61,15,.10),inset 0 1px 0 rgba(255,255,255,.8);
}
.global-metal-filter-note { display:flex; flex-direction:column; gap:4px; margin:0 0 11px; color:#2c2114; }
.global-metal-filter-note b { font-family:Georgia,serif; font-size:20px; }
.global-metal-filter-note span { color:#6e5b42; font-size:13px; line-height:1.45; }
.global-metal-filter-active {
  margin-top:11px; padding:9px 11px; border-radius:10px;
  border:1px solid rgba(183,137,63,.34); background:rgba(255,255,255,.72);
  color:#4d3a21; font-size:13px;
}
.detected-purities { margin-top:9px; color:#5f4b31; font-size:12px; line-height:1.55; }
.detected-purities b { color:#2f2417; }
.detected-purity-chip { display:inline-block; margin:4px 5px 0 0; padding:4px 8px; border-radius:999px; border:1px solid rgba(183,137,63,.36); background:rgba(255,255,255,.78); color:#5e431d; font-weight:700; }
.st-key-global_metal_filter [data-testid="stPills"] { margin:.15rem 0 .25rem; }
.st-key-global_metal_filter [data-testid="stPills"] [data-baseweb="button-group"] {
  display:grid !important; grid-template-columns:repeat(3,minmax(0,1fr)); gap:9px !important; width:100% !important;
}
.st-key-global_metal_filter [data-testid="stPills"] button {
  width:100% !important; min-height:48px !important; white-space:normal !important;
}
.small-muted { color: var(--muted); font-size: 12px; }
div[data-testid="stFileUploader"] section {
  border: 1px dashed #c9aa72; border-radius: 14px; background: #fffdf9;
}
/* Main-page actions use the same warm gold family as the site header.
   Sidebar and mobile navigation are overridden below and remain dark. */
div.stButton > button,
div.stDownloadButton > button,
[data-testid="stFormSubmitButton"] button,
[data-testid="stFileUploader"] button {
  min-height:44px !important; border-radius:10px !important;
  border:1px solid #b57b28 !important;
  background:linear-gradient(135deg,#e0bd78 0%,#c99545 48%,#b67827 100%) !important;
  color:#ffffff !important; font-weight:750 !important;
  text-shadow:0 1px 1px rgba(83,48,5,.28) !important;
  box-shadow:0 7px 18px rgba(126,80,18,.18),inset 0 1px 0 rgba(255,255,255,.28) !important;
  transition:transform .16s ease,border-color .16s ease,box-shadow .16s ease,filter .16s ease !important;
}
div.stButton > button *,
div.stDownloadButton > button *,
[data-testid="stFormSubmitButton"] button *,
[data-testid="stFileUploader"] button * { color:#ffffff !important; }
div.stButton > button:hover,
div.stDownloadButton > button:hover,
[data-testid="stFormSubmitButton"] button:hover,
[data-testid="stFileUploader"] button:hover {
  border-color:#a66b19 !important;
  background:linear-gradient(135deg,#ebcb8b 0%,#d7a553 48%,#c1842f 100%) !important;
  box-shadow:0 10px 24px rgba(154,99,25,.27),inset 0 1px 0 rgba(255,255,255,.34) !important;
  transform:translateY(-1px); filter:saturate(1.04);
}
div.stButton > button[kind="primary"],
[data-testid="stFormSubmitButton"] button[kind="primary"] {
  border-color:#9f6416 !important;
  background:linear-gradient(135deg,#dcb66d 0%,#c68d39 52%,#aa691d 100%) !important;
  color:#ffffff !important;
  box-shadow:0 8px 20px rgba(126,80,18,.26),inset 0 1px 0 rgba(255,255,255,.28) !important;
}
div.stButton > button:active,
div.stDownloadButton > button:active,
[data-testid="stFormSubmitButton"] button:active,
[data-testid="stFileUploader"] button:active {
  transform:translateY(0); border-color:#925914 !important;
  box-shadow:0 4px 12px rgba(126,80,18,.22),inset 0 1px 0 rgba(255,255,255,.20) !important;
}
div.stButton > button:focus,
div.stDownloadButton > button:focus,
[data-testid="stFormSubmitButton"] button:focus,
[data-testid="stFileUploader"] button:focus {
  outline:2px solid rgba(183,137,63,.38) !important; outline-offset:2px !important;
}
div.stButton > button:disabled,
div.stDownloadButton > button:disabled,
[data-testid="stFormSubmitButton"] button:disabled,
[data-testid="stFileUploader"] button:disabled {
  opacity:.5 !important; color:#ffffff !important; transform:none !important;
  background:linear-gradient(135deg,#d8c7a7 0%,#bea273 100%) !important;
  border-color:#b9a179 !important; box-shadow:none !important;
}

/* Streamlit otherwise paints segmented controls blue. Outside navigation,
   every segment is a light gold button with white text. */
[data-testid="stSegmentedControl"] button,
[data-testid="stSegmentedControl"] [role="radio"],
button[data-testid="stBaseButton-segmented_control"],
button[data-testid="stBaseButton-segmented_controlActive"],
button[kind="segmented_control"],
button[kind="segmented_controlActive"],
div[data-baseweb="button-group"] button {
  min-height:44px !important; border-radius:10px !important;
  border:1px solid #b57b28 !important;
  background:linear-gradient(135deg,#dfbc77 0%,#ca9644 55%,#b87928 100%) !important;
  color:#ffffff !important; font-weight:750 !important;
  text-shadow:0 1px 1px rgba(83,48,5,.28) !important;
  box-shadow:0 5px 14px rgba(126,80,18,.16),inset 0 1px 0 rgba(255,255,255,.25) !important;
  transition:transform .16s ease,border-color .16s ease,box-shadow .16s ease,filter .16s ease !important;
}
[data-testid="stSegmentedControl"] button *,
[data-testid="stSegmentedControl"] [role="radio"] *,
button[data-testid="stBaseButton-segmented_control"] *,
button[data-testid="stBaseButton-segmented_controlActive"] *,
button[kind="segmented_control"] *,
button[kind="segmented_controlActive"] *,
div[data-baseweb="button-group"] button * { color:#ffffff !important; }
[data-testid="stSegmentedControl"] button:hover,
[data-testid="stSegmentedControl"] [role="radio"]:hover,
button[data-testid="stBaseButton-segmented_control"]:hover,
button[data-testid="stBaseButton-segmented_controlActive"]:hover,
button[kind="segmented_control"]:hover,
button[kind="segmented_controlActive"]:hover,
div[data-baseweb="button-group"] button:hover {
  border-color:#9f6416 !important;
  background:linear-gradient(135deg,#ebcb8b 0%,#d6a351 55%,#bf7e2b 100%) !important;
  box-shadow:0 8px 20px rgba(154,99,25,.24),inset 0 1px 0 rgba(255,255,255,.30) !important;
  transform:translateY(-1px); filter:saturate(1.04);
}
[data-testid="stSegmentedControl"] button[aria-pressed="true"],
[data-testid="stSegmentedControl"] [role="radio"][aria-checked="true"],
button[data-testid="stBaseButton-segmented_controlActive"],
button[kind="segmented_controlActive"],
[data-testid="stSegmentedControl"] button[data-active="true"],
div[data-baseweb="button-group"] button[aria-pressed="true"] {
  border-color:#8f5510 !important;
  background:linear-gradient(135deg,#d3a85d 0%,#b97927 54%,#925612 100%) !important;
  color:#ffffff !important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.13),0 8px 20px rgba(126,80,18,.28) !important;
}
[data-testid="stSegmentedControl"] button[aria-pressed="true"] *,
[data-testid="stSegmentedControl"] [role="radio"][aria-checked="true"] *,
button[data-testid="stBaseButton-segmented_controlActive"] *,
button[kind="segmented_controlActive"] *,
[data-testid="stSegmentedControl"] button[data-active="true"] *,
div[data-baseweb="button-group"] button[aria-pressed="true"] * {
  color:#ffffff !important;
}
.block-navigation-title {
  margin: 18px 0 7px; color: #3f3529; font-size: 14px; font-weight: 800;
}
[data-testid="stSegmentedControl"] [data-baseweb="button-group"] {
  display:flex !important; flex-wrap:wrap !important; gap:7px !important;
}
[data-testid="stSegmentedControl"] button {
  border-radius:9px !important; flex:1 1 auto !important;
}
[data-testid="stSegmentedControl"] svg { fill:currentColor !important; color:currentColor !important; }
[data-testid="stMetric"] { border: 1px solid var(--line); padding: 12px; border-radius: 12px; background: #fff; }
hr { border-color: var(--line); }
[data-testid="stSidebar"] [role="radiogroup"] { gap: 0.35rem; }
[data-testid="stSidebar"] [role="radiogroup"] label {
  border-radius: 10px; padding: 0.62rem 0.72rem; border: 1px solid transparent;
  transition: all .15s ease; background: transparent;
}
[data-testid="stSidebar"] [role="radiogroup"] label:hover {
  background: rgba(183,137,63,.14); border-color: rgba(183,137,63,.35);
}
[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
  background: linear-gradient(90deg, rgba(183,137,63,.32) 0%, rgba(183,137,63,.10) 100%);
  border-color: #b7893f; color: #f2cf8c; font-weight: 700;
}
.side-nav { display:flex; flex-direction:column; gap:7px; margin:.15rem 0 1rem; }
.side-nav a,
.side-nav a:visited,
[data-testid="stSidebar"] .side-nav a,
[data-testid="stSidebar"] .side-nav a:visited {
  display:block; color:#f5ead8 !important; text-decoration:none !important;
  border-left:2px solid transparent; border-radius:0 10px 10px 0;
  padding:.66rem .78rem; font-size:.94rem; line-height:1.25;
  transition:background-color .16s ease, color .16s ease, border-color .16s ease, transform .16s ease;
}
.side-nav a:hover,
[data-testid="stSidebar"] .side-nav a:hover {
  color:#f1cc85 !important; text-decoration:none !important;
  background:linear-gradient(90deg, rgba(183,137,63,.24), rgba(183,137,63,.07));
  border-left-color:#b7893f; transform:translateX(2px);
}
.side-nav a:focus,
.side-nav a:active,
[data-testid="stSidebar"] .side-nav a:focus,
[data-testid="stSidebar"] .side-nav a:active {
  color:#ffe2a8 !important; text-decoration:none !important; outline:none;
  background:linear-gradient(90deg, rgba(183,137,63,.32), rgba(183,137,63,.10));
  border-left-color:#d4a95c;
}
.nav-hint { color:#cdbb9b; font-size:12px; margin:.2rem 0 .8rem; }
.report-anchor { position:relative; height:1px; scroll-margin-top:88px; }
html { scroll-behavior:smooth; }
.sonu-side-nav a { border:1px solid rgba(183,137,63,.26); border-left:3px solid transparent; }
.sonu-side-nav a:hover,
.sonu-side-nav a:focus,
.sonu-side-nav a:active { border-color:rgba(183,137,63,.45); border-left-color:#d4a95c; }
.executive-banner {
  position:relative; overflow:hidden; margin:4px 0 18px; padding:24px 26px;
  border-radius:18px; border:1px solid rgba(183,137,63,.46);
  background:
    radial-gradient(circle at 88% 18%, rgba(207,166,92,.24), transparent 30%),
    linear-gradient(135deg, #12100c 0%, #21190f 62%, #342511 100%);
  box-shadow:0 18px 45px rgba(38,25,7,.16); color:#fff7e8;
}
.executive-banner:after {
  content:""; position:absolute; inset:0; pointer-events:none;
  background:linear-gradient(115deg, transparent 0%, rgba(255,255,255,.05) 47%, transparent 72%);
}
.executive-banner-content { position:relative; z-index:2; max-width:920px; }
.executive-eyebrow { color:#e7c98e; font-size:11px; font-weight:800; letter-spacing:.17em; text-transform:uppercase; }
.executive-title { font-family:Georgia,serif; font-size:34px; line-height:1.12; margin:7px 0 7px; color:#fffaf0; }
.executive-copy { color:#ddcfb7; font-size:14px; line-height:1.55; }
.executive-note {
  margin:8px 0 14px; padding:11px 14px; border-radius:11px;
  border:1px solid #eadfcd; background:rgba(255,255,255,.78); color:#5e5549; font-size:12px;
}
.about-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:14px; margin:12px 0 20px; }
.about-card { border:1px solid var(--line); border-radius:14px; background:rgba(255,255,255,.94); padding:18px 19px; box-shadow:0 8px 24px rgba(34,24,9,.035); }
.about-card h4 { font-family:Georgia,serif; color:#6f4b16; font-size:19px; margin:0 0 8px; }
.about-card p { color:#4f4941; font-size:14px; line-height:1.55; margin:0; }
.about-card ul { color:#4f4941; font-size:14px; line-height:1.58; margin:.25rem 0 0; padding-left:1.1rem; }
.about-step { border-left:3px solid #b7893f; padding-left:12px; margin:9px 0; color:#302a23; }
.about-note {
  margin-top:14px; padding:11px 12px; border-radius:10px;
  border:1px solid rgba(183,137,63,.34); background:#fbf5e9;
  color:#5e4825; font-size:13px; line-height:1.5;
}
.updates-card { align-self:start; }
.updates-scroll {
  max-height:270px; overflow-y:auto; overscroll-behavior:contain;
  -webkit-overflow-scrolling:touch; touch-action:pan-y;
  padding-right:10px; scrollbar-gutter:stable;
}
.updates-scroll::-webkit-scrollbar { width:8px; }
.updates-scroll::-webkit-scrollbar-track { background:#f3ecdf; border-radius:999px; }
.updates-scroll::-webkit-scrollbar-thumb { background:#c49a55; border-radius:999px; }
.updates-scroll::-webkit-scrollbar-thumb:hover { background:#a9782f; }
@media (max-width: 780px) {
  .about-grid { grid-template-columns:1fr; }
  .updates-scroll { max-height:320px; }
}

.luxury-hero {
  position: relative; overflow: hidden; min-height: 220px; border-radius: 22px;
  border: 1px solid #eadfcd; margin-bottom: 18px; padding: 30px 34px;
  background:
    radial-gradient(circle at 84% 20%, rgba(183,137,63,.24), transparent 26%),
    radial-gradient(circle at 72% 76%, rgba(234,216,184,.42), transparent 32%),
    linear-gradient(135deg, #fffdf9 0%, #f7f0e4 58%, #efe0c5 100%);
  box-shadow: 0 24px 65px rgba(56,36,10,.12);
}
.luxury-hero:after {
  content:""; position:absolute; inset:0; pointer-events:none;
  background: linear-gradient(135deg, rgba(183,137,63,.08), transparent 45%);
}
.luxury-hero-content { position:relative; z-index:2; max-width:620px; }
.luxury-eyebrow { color:#9d6f29; font-size:12px; font-weight:800; letter-spacing:.17em; text-transform:uppercase; }
.luxury-title { font-family: Georgia, 'Times New Roman', serif; font-size: clamp(38px, 4.3vw, 54px); line-height:1.02; margin:10px 0 12px; color:#17120c; }
.luxury-title span { color:#a8742a; }
.luxury-copy { color:#5e5549; font-size:15px; line-height:1.55; max-width:560px; }
.luxury-badges { display:flex; flex-wrap:wrap; gap:10px; margin-top:22px; }
.luxury-badge { border:1px solid rgba(183,137,63,.32); background:rgba(255,255,255,.78); color:#6f4b16; border-radius:999px; padding:8px 12px; font-size:12px; font-weight:700; }
.luxury-divider { width:70px; height:2px; background:linear-gradient(90deg,#b7893f,transparent); margin:18px 0; }

[data-testid="stSidebar"]:before {
  content:""; display:block; height:6px; background:linear-gradient(90deg,#15120e,#b7893f,#15120e);
}
[data-testid="stSidebar"] { box-shadow: 12px 0 35px rgba(50,32,8,.06); }

@media (max-width: 900px) {
  .luxury-hero { padding:30px 26px; min-height:280px; background-position:68% center; }
  .luxury-hero:before { content:""; position:absolute; inset:0; background:rgba(255,255,255,.40); }
  .luxury-title { font-size:42px; }
}

/* Responsive shell: one codebase for desktop, iPad and phones. */
.mobile-nav-shell { display:none; }
[id] { scroll-margin-top: 86px; }
[data-testid="stPlotlyChart"],
[data-testid="stPlotlyChart"] > div,
.js-plotly-plot,
.plot-container,
.svg-container { max-width:100% !important; }
[data-testid="stDataFrame"] { max-width:100%; }

@media (max-width: 900px) {
  .block-container {
    max-width:100%; padding:0.85rem 1rem 2.25rem; overflow-x:hidden;
  }
  [data-testid="stSidebar"] { width:min(88vw, 350px) !important; }
  .mobile-nav-shell {
    display:block; position:sticky; top:0.35rem; z-index:999;
    margin:0 0 0.9rem; padding:0.48rem;
    border:1px solid rgba(183,137,63,.30); border-radius:14px;
    background:rgba(255,253,249,.94); backdrop-filter:blur(12px);
    -webkit-backdrop-filter:blur(12px); box-shadow:0 8px 28px rgba(35,24,10,.10);
  }
  .mobile-nav {
    display:flex; gap:0.45rem; overflow-x:auto; overscroll-behavior-x:contain;
    scrollbar-width:none; -webkit-overflow-scrolling:touch; white-space:nowrap;
  }
  .mobile-nav::-webkit-scrollbar { display:none; }
  .mobile-nav a,
  .mobile-nav a:visited {
    flex:0 0 auto; min-height:44px; display:inline-flex; align-items:center;
    color:#f5ead8 !important; text-decoration:none !important;
    border:1px solid #3a2b16; border-radius:10px;
    background:linear-gradient(135deg,#0d0b08 0%,#251c12 100%);
    padding:0.58rem 0.82rem; font-size:0.82rem; font-weight:750;
    box-shadow:0 5px 14px rgba(34,24,9,.12);
    transition:transform .16s ease,border-color .16s ease,color .16s ease,box-shadow .16s ease;
  }
  .mobile-nav a:hover,
  .mobile-nav a:active {
    color:#ffe2a8 !important; border-color:#d4a95c;
    background:linear-gradient(135deg,#25190d 0%,#4b3217 100%);
    box-shadow:0 8px 20px rgba(183,137,63,.24); transform:translateY(-1px);
  }

  [data-testid="stHorizontalBlock"] {
    flex-wrap:wrap !important; gap:0.85rem !important; align-items:stretch !important;
  }
  [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
    min-width:0 !important;
  }
  /* KPI/filter rows with 3+ columns become a comfortable 2-column grid on iPad. */
  [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3)) > [data-testid="stColumn"] {
    flex:1 1 calc(50% - 0.5rem) !important;
    width:calc(50% - 0.5rem) !important;
    min-width:260px !important;
  }
  .brand-card, .upload-panel, .analysis-panel, .section-divider { border-radius:14px; }
  .executive-banner { padding:21px 20px; border-radius:15px; }
  .executive-title { font-size:29px; }
  .section-divider { margin:28px 0 14px; padding:15px 17px; }
  .section-divider-title { font-size:25px; }
  .section-title { font-size:27px; }
  .kpi-card { min-height:108px; padding:15px; }
  .kpi-value { font-size:clamp(20px, 4vw, 28px); }
  [data-testid="stPlotlyChart"] { width:100% !important; overflow:visible !important; }
  [data-baseweb="tab-list"] {
    overflow-x:auto !important; flex-wrap:nowrap !important; scrollbar-width:none;
    -webkit-overflow-scrolling:touch;
  }
  [data-baseweb="tab-list"]::-webkit-scrollbar { display:none; }
  [data-baseweb="tab"] { flex:0 0 auto !important; min-width:max-content; }
  div[data-baseweb="select"] > div, input, textarea { min-height:44px; }
}

@media (max-width: 820px) {
  /* Two-column chart groups stack in iPad portrait so labels stay readable. */
  [data-testid="stHorizontalBlock"]:not(:has(> [data-testid="stColumn"]:nth-child(3))) {
    flex-direction:column !important;
  }
  [data-testid="stHorizontalBlock"]:not(:has(> [data-testid="stColumn"]:nth-child(3))) > [data-testid="stColumn"] {
    flex:1 1 100% !important; width:100% !important; min-width:0 !important;
  }
  .about-grid { grid-template-columns:1fr; }
  .luxury-hero { min-height:auto; padding:26px 22px; border-radius:18px; }
  .luxury-title { font-size:38px; }
  .luxury-copy { font-size:15px; line-height:1.55; }
}

@media (max-width: 600px) {
  .block-container { padding:0.65rem 0.72rem 1.8rem; }
  [data-testid="stHorizontalBlock"] { flex-direction:column !important; }
  [data-testid="stHorizontalBlock"] > [data-testid="stColumn"],
  [data-testid="stHorizontalBlock"]:has(> [data-testid="stColumn"]:nth-child(3)) > [data-testid="stColumn"] {
    flex:1 1 100% !important; width:100% !important; min-width:0 !important;
  }
  .luxury-hero { padding:22px 18px; margin-bottom:15px; }
  .luxury-title { font-size:32px; line-height:1.08; }
  .luxury-eyebrow { font-size:10px; letter-spacing:.13em; }
  .luxury-copy { font-size:14px; }
  .luxury-badges { gap:7px; margin-top:16px; }
  .luxury-badge { padding:7px 9px; font-size:11px; }
  .executive-banner { padding:18px 16px; }
  .executive-title { font-size:25px; }
  .executive-copy { font-size:13px; }
  .section-divider { padding:14px; margin:23px 0 12px; }
  .section-divider-title { font-size:22px; }
  .section-divider-copy { font-size:12px; }
  .section-title { font-size:24px; }
  .kpi-card { min-height:96px; }
  .kpi-value { font-size:23px; }
  .analysis-panel { padding:15px; }
  .analysis-panel-title { font-size:18px; }
  [data-testid="stDataFrame"] { overflow-x:auto !important; -webkit-overflow-scrolling:touch; }
  [data-testid="stSegmentedControl"] [data-baseweb="button-group"] {
    flex-wrap:nowrap !important; overflow-x:auto !important; padding-bottom:4px;
    -webkit-overflow-scrolling:touch; scrollbar-width:none;
  }
  [data-testid="stSegmentedControl"] [data-baseweb="button-group"]::-webkit-scrollbar { display:none; }
  [data-testid="stSegmentedControl"] button { flex:0 0 auto !important; white-space:nowrap !important; }
}

/* Tabs and expanders follow the same neutral/gold language. */
[data-baseweb="tab-list"] { gap:6px; }
[data-baseweb="tab"] {
  min-height:44px; border-radius:10px; color:#4e4030 !important;
  border:1px solid #d8c8ad; background:#fffdf9; padding:0 14px;
}
[data-baseweb="tab"][aria-selected="true"] {
  color:#ffe2a8 !important; border-color:#d4a95c;
  background:linear-gradient(135deg,#1a140d 0%,#4a3218 100%);
}
[data-testid="stExpander"] { border-color:#e3d3b8 !important; border-radius:12px !important; }

@media (max-width: 600px) {
  .brand-card { padding:17px 16px; }
  .brand-title { font-size:31px; line-height:1.08; }
  .brand-subtitle { font-size:13px; line-height:1.45; }
  [data-testid="stFileUploader"] section { padding:0.75rem !important; }
  [data-testid="stDataFrame"] > div { max-width:100% !important; overflow-x:auto !important; }
  [data-testid="stDownloadButton"] button,
  div.stButton > button,
  [data-testid="stFormSubmitButton"] button { width:100% !important; }
  .st-key-global_fx_compact [data-testid="stHorizontalBlock"] {
    flex-direction:column !important; flex-wrap:nowrap !important; gap:.55rem !important; align-items:stretch !important;
  }
  .st-key-global_fx_compact [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
    width:100% !important; min-width:0 !important; flex:1 1 100% !important;
  }
  .st-key-global_fx_compact [data-testid="stNumberInput"],
  .st-key-global_fx_compact [data-baseweb="input"] { width:100% !important; max-width:none !important; }
  .fx-compact-title { font-size:12px; }
  .fx-compact-value { font-size:11px; white-space:normal; overflow-wrap:anywhere; }
  .st-key-global_metal_filter { padding:14px 12px 12px; }
  .st-key-global_metal_filter [data-testid="stPills"] [data-baseweb="button-group"] {
    grid-template-columns:1fr !important;
  }
  .global-metal-filter-note b { font-size:18px; }
}

/* Report mode switch and comparison cards. */
[data-testid="stSegmentedControl"] { margin: 0.35rem 0 1rem; }
[data-testid="stSegmentedControl"] button { min-height: 44px; font-weight: 700; }
.compare-upload-card {
  border:1px solid var(--line); border-radius:16px; background:rgba(255,255,255,.94);
  padding:16px 18px; margin-bottom:10px; box-shadow:0 8px 24px rgba(34,24,9,.035);
}
.compare-period-title { font-family:Georgia,serif; font-size:24px; color:#21180d; margin-bottom:4px; }
.compare-period-copy { color:var(--muted); font-size:13px; }
.delta-positive { color:#2f6d3b; font-weight:700; }
.delta-negative { color:#9b3d36; font-weight:700; }
.delta-neutral { color:#6c6c6c; font-weight:700; }

/* High-contrast global FX input on the main page. */
[data-testid="stNumberInput"] input {
  color:#17130f !important;
  background:#ffffff !important;
  font-weight:700 !important;
  -webkit-text-fill-color:#17130f !important;
}
[data-testid="stNumberInput"] button {
  color:#ffffff !important;
  background:linear-gradient(135deg,#d9af67,#b97928) !important;
  border-color:#a76a18 !important;
}
[data-testid="stNumberInput"] button * { color:#ffffff !important; }

/* Compact site-wide FX control. */
.st-key-global_fx_compact {
  margin:.2rem 0 .85rem; padding:.52rem .72rem .40rem;
  border:1px solid rgba(183,137,63,.34); border-radius:12px;
  background:linear-gradient(135deg,rgba(255,253,249,.96),rgba(246,235,215,.92));
  box-shadow:0 8px 22px rgba(62,40,10,.055);
}
.st-key-global_fx_compact [data-testid="stHorizontalBlock"] {
  gap:1rem !important; align-items:center !important; overflow:visible !important;
}
.st-key-global_fx_compact [data-testid="stNumberInput"] { margin:0 !important; width:100% !important; min-width:0 !important; }
.st-key-global_fx_compact [data-testid="stNumberInput"] input {
  min-height:38px !important; height:38px !important; padding-top:.25rem !important; padding-bottom:.25rem !important;
}
.fx-compact-title { color:#3b2b16; font-size:13px; line-height:1.2; font-weight:800; }
.fx-compact-value { color:#8c5d1d; font-size:12px; line-height:1.35; margin-top:2px; white-space:normal; overflow-wrap:anywhere; }


/* One visual sidebar system for General, Comparison, Baserow and Sonu. */
[data-testid="stSidebar"] {
  --sidebar-nav-bg:linear-gradient(135deg,#181006 0%,#35230f 58%,#4a3014 100%);
  --sidebar-nav-hover:linear-gradient(135deg,#241707 0%,#503316 58%,#68431b 100%);
  --sidebar-nav-current:linear-gradient(135deg,#c5903b 0%,#9a641f 55%,#74430f 100%);
  --sidebar-nav-border:rgba(207,151,60,.58);
  --sidebar-nav-text:#fff8ec;
}
[data-testid="stSidebar"] [data-testid="stImage"] {
  margin:0 0 1rem !important;
}
[data-testid="stSidebar"] [data-testid="stImage"] img {
  display:block; width:100%; border-radius:11px;
  border:1px solid rgba(183,137,63,.14);
  box-shadow:0 12px 28px rgba(0,0,0,.22);
}
.sidebar-product-header {
  margin:0 0 1.15rem; padding:0 0 1rem;
  border-bottom:1px solid rgba(183,137,63,.18);
}
.sidebar-suite-title {
  color:#f7efe2; font-size:14px; line-height:1.25; font-weight:800;
}
.sidebar-module-title {
  margin-top:5px; color:#f2cf8c; font-family:Georgia,serif;
  font-size:20px; line-height:1.12; font-weight:700;
  overflow-wrap:anywhere;
}
.sidebar-version {
  margin-top:8px; color:#a99a84; font-size:12px; line-height:1.3;
}
.sidebar-navigation-title {
  margin:0 0 .72rem; color:#cdbb9b; font-size:11px; line-height:1.3;
  font-weight:800; letter-spacing:.08em; text-transform:uppercase;
}
/* HTML anchor navigation and Streamlit button navigation intentionally share
   the same geometry, alignment, palette and interaction states. */
.sidebar-nav-item,
.sidebar-nav-item:visited,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button {
  box-sizing:border-box !important; width:100% !important; min-height:44px !important;
  display:flex !important; align-items:center !important; justify-content:center !important;
  margin:0 0 7px !important; padding:.68rem .78rem !important;
  border:1px solid var(--sidebar-nav-border) !important; border-radius:10px !important;
  background:var(--sidebar-nav-bg) !important;
  color:var(--sidebar-nav-text) !important; text-decoration:none !important;
  text-align:center !important; font-size:.91rem !important; line-height:1.25 !important;
  font-weight:750 !important; box-shadow:0 6px 16px rgba(0,0,0,.19) !important;
  transition:transform .16s ease,border-color .16s ease,color .16s ease,
             box-shadow .16s ease,background .16s ease,opacity .16s ease !important;
}
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton {
  margin:0 !important;
}
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button p,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button span {
  color:inherit !important; width:100%; text-align:center !important;
}
.sidebar-nav-item:hover,
.sidebar-nav-item:focus,
.sidebar-nav-item:active,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button:hover,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button:focus {
  color:#ffffff !important; text-decoration:none !important;
  border-color:#e1b15f !important; background:var(--sidebar-nav-hover) !important;
  box-shadow:0 9px 22px rgba(183,137,63,.25) !important;
  transform:translateY(-1px) !important; outline:none !important;
}
.sidebar-nav-item.is-current,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button[kind="primary"] {
  color:#ffffff !important; border-color:#efc578 !important;
  background:var(--sidebar-nav-current) !important;
  box-shadow:inset 0 0 0 1px rgba(255,255,255,.10),0 9px 22px rgba(116,67,15,.34) !important;
}
.sidebar-nav-item.is-disabled,
.sidebar-nav-item.is-disabled:hover,
.sidebar-nav-item.is-disabled:focus,
.sidebar-nav-item.is-disabled:active,
[data-testid="stSidebar"] [class*="st-key-sidebar_navigation_controls"] div.stButton > button:disabled {
  opacity:.62 !important; cursor:not-allowed !important; transform:none !important;
  color:#d8cbb8 !important; border-color:rgba(183,137,63,.30) !important;
  background:linear-gradient(135deg,#100c07 0%,#21170d 100%) !important;
  box-shadow:0 4px 11px rgba(0,0,0,.11) !important;
}
.sidebar-status {
  min-height:48px; display:flex; align-items:center; gap:9px;
  margin:1rem 0 .55rem; padding:.72rem .78rem; border-radius:10px;
  border:1px solid rgba(183,137,63,.24); background:rgba(255,255,255,.035);
  color:#ddd0bd; font-size:12px; line-height:1.35; font-weight:700;
}
.sidebar-status-dot {
  flex:0 0 auto; width:8px; height:8px; border-radius:999px;
  background:#9e8e77; box-shadow:0 0 0 4px rgba(158,142,119,.10);
}
.sidebar-status-success { border-color:rgba(91,157,102,.34); background:rgba(41,104,53,.15); color:#d7ecd9; }
.sidebar-status-success .sidebar-status-dot { background:#67bd76; box-shadow:0 0 0 4px rgba(103,189,118,.12); }
.sidebar-status-warning { border-color:rgba(212,169,92,.42); background:rgba(151,99,16,.15); color:#f2d59f; }
.sidebar-status-warning .sidebar-status-dot { background:#d4a95c; box-shadow:0 0 0 4px rgba(212,169,92,.12); }
.sidebar-status-error { border-color:rgba(184,76,65,.42); background:rgba(122,39,31,.16); color:#f0c1bc; }
.sidebar-status-error .sidebar-status-dot { background:#d07167; box-shadow:0 0 0 4px rgba(208,113,103,.12); }
.sidebar-source { color:#9f907b; font-size:11px; line-height:1.4; margin:0 0 .85rem; }
.sidebar-action-separator { height:1px; margin:.85rem 0 1rem; background:rgba(183,137,63,.18); }
.sidebar-footer {
  margin:1.05rem 0 .3rem; padding-top:.9rem; border-top:1px solid rgba(183,137,63,.16);
  color:#7f725f; font-size:10px; line-height:1.35;
}
.mobile-nav-item,
.mobile-nav-item:visited {
  flex:0 0 auto; min-height:44px; display:inline-flex; align-items:center;
  color:#f5ead8 !important; text-decoration:none !important;
  border:1px solid #3a2b16; border-radius:10px;
  background:linear-gradient(135deg,#0d0b08 0%,#251c12 100%);
  padding:.58rem .82rem; font-size:.82rem; font-weight:750;
  box-shadow:0 5px 14px rgba(34,24,9,.12);
}
.mobile-nav-item:hover,
.mobile-nav-item:active,
.mobile-nav-item.is-current {
  color:#ffe2a8 !important; border-color:#d4a95c;
  background:linear-gradient(135deg,#25190d 0%,#4b3217 100%);
  box-shadow:0 8px 20px rgba(183,137,63,.24);
}
.mobile-nav-item.is-disabled {
  opacity:.43; cursor:not-allowed; box-shadow:none;
}


/* 1.11.1 compact analytics workspaces and collapsed report settings. */
.executive-banner-compact { padding:18px 22px; margin-bottom:14px; }
.executive-banner-compact .executive-title { font-size:30px; }
.comparison-period-strip {
  display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:14px;
  margin:4px 0 14px; padding:14px 16px; border:1px solid var(--line);
  border-radius:14px; background:rgba(255,255,255,.92);
}
.comparison-period-strip > div:not(.comparison-period-arrow) { display:flex; flex-direction:column; gap:4px; }
.comparison-period-strip > div:last-child { text-align:right; }
.comparison-period-strip b { font-family:Georgia,serif; font-size:18px; color:#241a0f; }
.comparison-period-strip span { color:var(--muted); font-size:12px; }
.comparison-period-arrow { color:var(--gold); font-size:24px; font-weight:800; }
.comparison-metric-card {
  min-height:158px; padding:16px; border:1px solid var(--line); border-radius:14px;
  background:rgba(255,255,255,.96); box-shadow:0 8px 24px rgba(34,24,9,.04);
}
.comparison-metric-label { color:var(--muted); font-size:11px; font-weight:800; letter-spacing:.06em; text-transform:uppercase; }
.comparison-metric-values { display:grid; grid-template-columns:1fr auto 1fr; align-items:center; gap:8px; margin-top:14px; }
.comparison-metric-values > div:not(.comparison-arrow) { min-width:0; }
.comparison-metric-values span { display:block; color:#8a8176; font-size:10px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.comparison-metric-values b { display:block; margin-top:5px; color:#18130d; font-family:Georgia,serif; font-size:19px; overflow-wrap:anywhere; }
.comparison-arrow { color:var(--gold); font-weight:800; }
.comparison-metric-delta { margin-top:13px; padding-top:10px; border-top:1px solid #eee7dc; color:#6f4b16; font-size:12px; font-weight:800; }
.about-hero {
  padding:26px 28px; margin:4px 0 18px; border-radius:18px; color:#fff8eb;
  background:linear-gradient(135deg,#15110b 0%,#2f2111 58%,#5a3a16 100%);
  box-shadow:0 18px 42px rgba(48,31,9,.15);
}
.about-hero-title { margin:5px 0 7px; font-family:Georgia,serif; font-size:42px; }
.about-hero-copy { max-width:780px; color:#e2d4bd; font-size:15px; line-height:1.6; }
.about-grid-compact { grid-template-columns:repeat(3,minmax(0,1fr)); }
.updates-preview { padding:4px 0 12px; }
.updates-scroll-standalone { max-height:62vh; padding:16px; border:1px solid var(--line); border-radius:14px; background:rgba(255,255,255,.9); }
.st-key-standard_workspace [data-baseweb="button-group"],
.st-key-comparison_workspace [data-baseweb="button-group"],
.st-key-about_workspace [data-baseweb="button-group"] {
  display:flex !important; flex-wrap:nowrap !important; overflow-x:auto !important;
  -webkit-overflow-scrolling:touch; scrollbar-width:none; padding-bottom:4px;
}
.st-key-standard_workspace [data-baseweb="button-group"]::-webkit-scrollbar,
.st-key-comparison_workspace [data-baseweb="button-group"]::-webkit-scrollbar,
.st-key-about_workspace [data-baseweb="button-group"]::-webkit-scrollbar { display:none; }
.st-key-standard_workspace button,
.st-key-comparison_workspace button,
.st-key-about_workspace button { flex:0 0 auto !important; white-space:nowrap !important; }

@media (max-width: 900px) {
  .about-grid-compact { grid-template-columns:repeat(2,minmax(0,1fr)); }
  .comparison-metric-card { min-height:145px; }
}
@media (max-width: 600px) {
  .luxury-hero { padding:17px 15px; }
  .luxury-title { font-size:28px; }
  .luxury-copy { font-size:13px; line-height:1.45; }
  .luxury-divider { margin:11px 0; }
  .luxury-badges { display:none; }
  .executive-banner-compact { padding:15px; }
  .executive-banner-compact .executive-title { font-size:23px; }
  .comparison-period-strip { grid-template-columns:1fr; gap:7px; text-align:left; }
  .comparison-period-strip > div:last-child { text-align:left; }
  .comparison-period-arrow { transform:rotate(90deg); width:max-content; }
  .comparison-metric-card { min-height:auto; }
  .comparison-metric-values b { font-size:17px; }
  .about-grid-compact { grid-template-columns:1fr; }
  .about-hero { padding:20px 17px; }
  .about-hero-title { font-size:32px; }
}


/* Analitika 2.0 — unified visual system. Business logic is intentionally untouched. */
:root {
  --gold:#b9822e; --gold-deep:#845318; --gold-pale:#f6ead3;
  --ink:#18140f; --muted:#71695f; --line:#e8e0d4; --paper:#f7f4ef;
  --surface:#ffffff; --success:#39714c; --warning:#9a681e; --danger:#a23c37;
  --radius-sm:10px; --radius-md:16px; --radius-lg:22px;
  --shadow-soft:0 10px 30px rgba(49,34,15,.055);
}
html, body, [data-testid="stAppViewContainer"] { background:#f8f6f2 !important; }
[data-testid="stAppViewContainer"] > .main { background:
  radial-gradient(circle at 92% 2%, rgba(190,145,72,.065), transparent 24%),
  linear-gradient(180deg,#fbfaf8 0%,#f7f4ef 100%) !important; }
.block-container { max-width:1560px; padding-top:1.1rem; padding-bottom:3.2rem; }

/* No legacy black sidebar in any 2.0 workspace. */
[data-testid="stSidebar"],
[data-testid="stExpandSidebarButton"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapseButton"] { display:none !important; visibility:hidden !important; }
[data-testid="stHeader"], [data-testid="stToolbar"] { height:0 !important; min-height:0 !important; }

/* Executive module header: compact, intelligent and restrained. */
.luxury-hero {
  min-height:auto; padding:25px 29px 24px; margin:0 0 16px; border-radius:var(--radius-lg);
  border:1px solid rgba(183,137,63,.25);
  background:
    radial-gradient(circle at 92% 8%, rgba(190,145,72,.12), transparent 25%),
    linear-gradient(135deg,#fffefa 0%,#f8f1e5 100%);
  box-shadow:var(--shadow-soft);
}
.luxury-hero:after { background:linear-gradient(110deg,rgba(183,137,63,.045),transparent 48%); }
.luxury-hero-content { max-width:920px; }
.luxury-eyebrow { color:#946322; font-size:11px; letter-spacing:.19em; }
.luxury-title { font-size:clamp(38px,4vw,52px); margin:8px 0 8px; letter-spacing:-.025em; }
.luxury-divider { width:58px; height:2px; margin:13px 0; }
.luxury-copy { max-width:850px; color:#5f574d; font-size:15px; line-height:1.62; }
.luxury-badges { margin-top:17px; gap:8px; }
.luxury-badge { padding:7px 11px; background:rgba(255,255,255,.7); border-color:rgba(183,137,63,.24); color:#684818; }

/* One consistent navigation language. */
[data-testid="stSegmentedControl"] [data-baseweb="button-group"] { gap:7px !important; }
[data-testid="stSegmentedControl"] button,
[data-testid="stSegmentedControl"] [role="radio"],
button[data-testid="stBaseButton-segmented_control"],
button[kind="segmented_control"],
div[data-baseweb="button-group"] button {
  min-height:43px !important; border:1px solid #dfd4c4 !important;
  background:#fff !important; color:#51483d !important; text-shadow:none !important;
  box-shadow:0 3px 10px rgba(45,31,13,.035) !important;
}
[data-testid="stSegmentedControl"] button *,
[data-testid="stSegmentedControl"] [role="radio"] *,
button[data-testid="stBaseButton-segmented_control"] *,
button[kind="segmented_control"] *,
div[data-baseweb="button-group"] button * { color:#51483d !important; }
[data-testid="stSegmentedControl"] button:hover,
[data-testid="stSegmentedControl"] [role="radio"]:hover,
button[data-testid="stBaseButton-segmented_control"]:hover,
button[kind="segmented_control"]:hover,
div[data-baseweb="button-group"] button:hover {
  border-color:#c89a55 !important; background:#fcf7ee !important; transform:translateY(-1px);
}
[data-testid="stSegmentedControl"] button[aria-pressed="true"],
[data-testid="stSegmentedControl"] [role="radio"][aria-checked="true"],
button[data-testid="stBaseButton-segmented_controlActive"],
button[kind="segmented_controlActive"],
[data-testid="stSegmentedControl"] button[data-active="true"],
div[data-baseweb="button-group"] button[aria-pressed="true"] {
  border-color:#9b671f !important;
  background:linear-gradient(135deg,#c9984a 0%,#a96d20 100%) !important;
  color:#fff !important; box-shadow:0 7px 18px rgba(128,81,20,.18) !important;
}
[data-testid="stSegmentedControl"] button[aria-pressed="true"] *,
[data-testid="stSegmentedControl"] [role="radio"][aria-checked="true"] *,
button[data-testid="stBaseButton-segmented_controlActive"] *,
button[kind="segmented_controlActive"] *,
[data-testid="stSegmentedControl"] button[data-active="true"] *,
div[data-baseweb="button-group"] button[aria-pressed="true"] * { color:#fff !important; }

/* Clear action hierarchy. */
div.stButton > button,
div.stDownloadButton > button,
[data-testid="stFileUploader"] button {
  min-height:44px !important; border-radius:var(--radius-sm) !important;
  border:1px solid #d8c9b4 !important; background:#fff !important; color:#62451b !important;
  text-shadow:none !important; box-shadow:0 4px 13px rgba(45,31,13,.045) !important;
}
div.stButton > button *, div.stDownloadButton > button *, [data-testid="stFileUploader"] button * { color:#62451b !important; }
div.stButton > button:hover, div.stDownloadButton > button:hover, [data-testid="stFileUploader"] button:hover {
  border-color:#b9822e !important; background:#fffaf2 !important; transform:translateY(-1px); box-shadow:0 7px 18px rgba(80,52,15,.09) !important;
}
div.stButton > button[kind="primary"],
[data-testid="stFormSubmitButton"] button,
[data-testid="stFormSubmitButton"] button[kind="primary"] {
  border:1px solid #9d6418 !important;
  background:linear-gradient(135deg,#cda154 0%,#a96d20 100%) !important;
  color:#fff !important; text-shadow:0 1px 1px rgba(60,35,5,.18) !important;
  box-shadow:0 8px 20px rgba(128,81,20,.19) !important;
}
div.stButton > button[kind="primary"] *, [data-testid="stFormSubmitButton"] button * { color:#fff !important; }

/* Cards, controls and tables share one geometry. */
[data-testid="stMetric"], .kpi-card, .comparison-metric-card, .analysis-panel,
.about-card, .wh-metric, .wh-stock-card, .sonu-data-card {
  border:1px solid var(--line) !important; border-radius:var(--radius-md) !important;
  background:rgba(255,255,255,.94) !important; box-shadow:var(--shadow-soft) !important;
}
[data-testid="stMetric"] { padding:15px 16px; min-height:104px; }
[data-testid="stMetricLabel"] { color:#7a7064; font-size:12px; letter-spacing:.02em; }
[data-testid="stMetricValue"] { color:#201a13; font-family:Georgia,serif; letter-spacing:-.02em; }
[data-testid="stExpander"] { border:1px solid var(--line) !important; border-radius:14px !important; background:rgba(255,255,255,.78) !important; box-shadow:0 5px 18px rgba(45,31,13,.035); overflow:hidden; }
[data-testid="stExpander"] summary { min-height:48px; padding:.25rem .25rem; }
[data-testid="stForm"] { border:1px solid var(--line); border-radius:var(--radius-md); padding:17px; background:rgba(255,255,255,.82); }
[data-testid="stDataFrame"] { border:1px solid var(--line); border-radius:14px; overflow:hidden; background:#fff; box-shadow:0 6px 20px rgba(45,31,13,.035); }
[data-testid="stDataFrame"] [role="columnheader"] { background:#f6f0e6 !important; color:#433727 !important; font-weight:750 !important; }
[data-testid="stFileUploader"] section { border:1px dashed #caa86d; border-radius:var(--radius-md); background:rgba(255,255,255,.76); padding:1rem; }
input, textarea, div[data-baseweb="select"] > div { border-radius:10px !important; }

/* Quiet section dividers for every module. */
.section-divider, .warehouse-section-heading {
  margin:31px 0 14px !important; padding:15px 18px !important; border-radius:14px !important;
  border:1px solid rgba(183,137,63,.2) !important;
  background:linear-gradient(90deg,rgba(246,234,211,.78),rgba(255,255,255,.88)) !important;
  box-shadow:none !important;
}
.section-divider-title, .warehouse-section-title { font-size:26px !important; }
.executive-banner, .sonu-ai-brief { border-radius:18px !important; box-shadow:var(--shadow-soft) !important; }

/* Report identity and empty states. */
.report-context {
  display:flex; align-items:center; gap:11px; margin:11px 0 14px; padding:11px 13px;
  border:1px solid #dfd5c7; border-radius:13px; background:rgba(255,255,255,.88);
}
.report-context-dot { width:9px; height:9px; flex:0 0 auto; border-radius:50%; background:#4d865e; box-shadow:0 0 0 4px rgba(77,134,94,.10); }
.report-context-copy { min-width:0; display:flex; flex-direction:column; gap:2px; }
.report-context-copy strong { color:#30271d; font-size:13px; }
.report-context-copy span { color:#7a7064; font-size:12px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }

/* Product overview. */
.product-flow { display:flex; align-items:center; justify-content:center; flex-wrap:wrap; gap:8px; margin:14px 0 22px; padding:13px 15px; border:1px solid var(--line); border-radius:14px; background:rgba(255,255,255,.78); }
.product-flow span { padding:6px 9px; border-radius:9px; background:#f6ead3; color:#5f4218; font-size:12px; font-weight:750; }
.product-flow b { color:#b9822e; font-size:16px; }

/* Consistent feedback colours. */
[data-testid="stAlert"] { border-radius:13px !important; border-width:1px !important; box-shadow:none !important; }
hr { margin:1.2rem 0; }

@media (max-width:900px) {
  .block-container { padding:0.85rem .9rem 2.4rem; }
  .luxury-hero { padding:22px 21px; border-radius:18px; }
  .luxury-title { font-size:38px; }
  .luxury-copy { max-width:680px; }
  [data-testid="stMetric"] { min-height:96px; }
}
@media (max-width:600px) {
  .block-container { padding:.65rem .68rem 1.9rem; }
  .luxury-hero { padding:18px 16px; margin-bottom:12px; border-radius:16px; }
  .luxury-eyebrow { font-size:9px; letter-spacing:.16em; }
  .luxury-title { font-size:29px; margin:6px 0; }
  .luxury-copy { font-size:13px; line-height:1.5; }
  .luxury-divider { margin:9px 0; width:44px; }
  .luxury-badges { display:none; }
  .report-context { padding:10px 11px; }
  .product-flow { justify-content:flex-start; overflow-x:auto; flex-wrap:nowrap; }
  .product-flow span, .product-flow b { flex:0 0 auto; }
  .section-divider, .warehouse-section-heading { margin:24px 0 12px !important; padding:13px 14px !important; }
  .section-divider-title, .warehouse-section-title { font-size:23px !important; }
  [data-testid="stMetric"] { min-height:auto; padding:13px; }
}

</style>
"""


st.markdown(_css(), unsafe_allow_html=True)


def money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ")


def analytics_fx_rate() -> float:
    """Compatibility wrapper for the single site-wide VND/USD rate."""
    return get_vnd_per_usd()


def to_usd(value: float) -> float:
    return vnd_to_usd(value)


def usd_money(value: float) -> str:
    return f"${money(to_usd(value))}"



def is_monetary_column(name: str) -> bool:
    normalized = str(name).strip().casefold()
    return (
        normalized.startswith("выручка")
        or normalized.startswith("средняя стоимость")
        or normalized.startswith("δ выручки")
        or normalized.startswith("δ средней стоимости")
        or normalized.startswith("Δ выручки".casefold())
        or normalized.startswith("Δ средней стоимости".casefold())
    )


def pct(value: float) -> str:
    return f"{value:.2%}".replace(".", ",")


def kpi_card(label: str, value: str, note: str = "") -> None:
    st.markdown(
        f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div><div class="kpi-note">{note}</div></div>',
        unsafe_allow_html=True,
    )


def leader_kpi_card(label: str, name: str, metric: str) -> None:
    """Render a leader card with the result and metric at readable prominence."""
    st.markdown(
        f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{name}</div>'
        f'<div class="kpi-leader-metric">{metric}</div></div>',
        unsafe_allow_html=True,
    )


def base_store_name(name: str) -> str:
    return name.split(" — ")[0]


def is_tourist_flow_store(name: str) -> bool:
    """Stores excluded only from the retail-network leader ranking."""
    normalized = "".join(str(name).upper().split())
    return normalized == "OUTLET" or normalized.startswith("63")


def retail_leader_summary(store_summary: pd.DataFrame) -> pd.DataFrame:
    """Retail stores used for revenue/quantity leaders; keeps the full report intact."""
    if store_summary.empty or "Магазин" not in store_summary.columns:
        return store_summary.copy()
    retail = store_summary.loc[
        ~store_summary["Магазин"].astype(str).map(is_tourist_flow_store)
    ].copy()
    return retail if not retail.empty else store_summary.copy()


def segment_totals(store) -> dict[str, dict[str, float]]:
    result: dict[str, dict[str, float]] = {}
    for segment in SEG_ORDER:
        q, a = totals_for(store, seg=segment)
        result[segment] = {"qty": int(q), "amount": float(a)}
    return result


def network_summary(stores: Iterable) -> pd.DataFrame:
    rows = []
    for store in stores:
        segs = segment_totals(store)
        row = {
            "Магазин": base_store_name(store.name),
            "Период": store.period_text(),
            "Количество": store.total_qty,
            "Выручка": store.total_amount,
            "Средняя стоимость": store.total_amount / store.total_qty if store.total_qty else 0,
        }
        for seg in SEG_ORDER:
            row[f"{SEGMENT_LABELS[seg]} — шт. %"] = segs[seg]["qty"] / store.total_qty if store.total_qty else 0
            row[f"{SEGMENT_LABELS[seg]} — продажи %"] = segs[seg]["amount"] / store.total_amount if store.total_amount else 0
        rows.append(row)
    return pd.DataFrame(rows)


def network_segment_summary(stores: Iterable) -> pd.DataFrame:
    """Compact network-level segment mix for the executive brief."""
    rows: list[dict] = []
    stores = list(stores)
    total_qty = sum(int(store.total_qty) for store in stores)
    total_sales = sum(float(store.total_amount) for store in stores)
    for segment in SEG_ORDER:
        qty = 0
        sales = 0.0
        for store in stores:
            current_qty, current_sales = totals_for(store, seg=segment)
            qty += int(current_qty)
            sales += float(current_sales)
        rows.append({
            "Сегмент": SEGMENT_LABELS[segment],
            "Количество": qty,
            "Выручка": sales,
            "Средняя стоимость": sales / qty if qty else 0,
            "% количества": qty / total_qty if total_qty else 0,
            "% выручки": sales / total_sales if total_sales else 0,
        })
    return pd.DataFrame(rows)


def executive_store_summary(stores: Iterable) -> pd.DataFrame:
    """One-row-per-store management table used in the operational brief."""
    stores = list(stores)
    total_sales = sum(float(store.total_amount) for store in stores)
    rows: list[dict] = []
    for store in stores:
        segments = segment_totals(store)
        leader_segment = max(SEG_ORDER, key=lambda segment: segments[segment]["amount"])
        leader_sales = float(segments[leader_segment]["amount"])
        rows.append({
            "Магазин": base_store_name(store.name),
            "Выручка": float(store.total_amount),
            "% выручки сети": float(store.total_amount) / total_sales if total_sales else 0,
            "Количество": int(store.total_qty),
            "Средняя стоимость": float(store.total_amount) / int(store.total_qty) if store.total_qty else 0,
            "Главный сегмент": SEGMENT_LABELS[leader_segment],
            "% главного сегмента": leader_sales / float(store.total_amount) if store.total_amount else 0,
        })
    if not rows:
        return pd.DataFrame(columns=[
            "Магазин", "Выручка", "% выручки сети", "Количество",
            "Средняя стоимость", "Главный сегмент", "% главного сегмента",
        ])
    return pd.DataFrame(rows).sort_values("Выручка", ascending=False).reset_index(drop=True)


def executive_insights(
    stores: list[StoreData],
    store_summary: pd.DataFrame,
    segment_summary: pd.DataFrame,
    supplier_df: pd.DataFrame,
) -> list[str]:
    """Generate factual, decision-oriented observations without forecasting."""
    if store_summary.empty:
        return []

    lines: list[str] = []
    total_sales = float(store_summary["Выручка"].sum())

    retail_summary = retail_leader_summary(store_summary)
    revenue_leader = retail_summary.sort_values("Выручка", ascending=False).iloc[0]
    retail_total_sales = float(retail_summary["Выручка"].sum())
    retail_share = float(revenue_leader["Выручка"]) / retail_total_sales if retail_total_sales else 0
    lines.append(
        f"Лидер розничной сети по выручке — {revenue_leader['Магазин']}: "
        f"{usd_money(float(revenue_leader['Выручка']))}, "
        f"или {pct(retail_share)} выручки розничной сети."
    )

    top_three_share = float(store_summary.head(3)["Выручка"].sum()) / total_sales if total_sales else 0
    lines.append(f"Три крупнейших магазина формируют {pct(top_three_share)} выручки сети.")

    if not segment_summary.empty:
        segment_leader = segment_summary.sort_values("Выручка", ascending=False).iloc[0]
        lines.append(
            f"Главный сегмент сети — {segment_leader['Сегмент']}: "
            f"{pct(float(segment_leader['% выручки']))} выручки."
        )

    avg_leader = retail_summary.sort_values("Средняя стоимость", ascending=False).iloc[0]
    lines.append(
        f"Самая высокая средняя стоимость проданного изделия — в {avg_leader['Магазин']}: "
        f"{usd_money(float(avg_leader['Средняя стоимость']))}."
    )

    concentration_leader = store_summary.sort_values("% главного сегмента", ascending=False).iloc[0]
    lines.append(
        f"Наибольшая концентрация на одном сегменте — в {concentration_leader['Магазин']}: "
        f"{concentration_leader['Главный сегмент']} дает "
        f"{pct(float(concentration_leader['% главного сегмента']))} выручки магазина."
    )

    if supplier_has_meaningful_detail(supplier_df):
        suppliers = supplier_summary(supplier_df)
        supplier_leader = suppliers.iloc[0]
        top_supplier_share = float(suppliers.head(3)["% выручки"].sum())
        lines.append(
            f"Лидер среди поставщиков — {supplier_leader['Поставщик']} "
            f"({pct(float(supplier_leader['% выручки']))}); топ-3 поставщика дают "
            f"{pct(top_supplier_share)} выручки."
        )

    return lines[:6]


def render_executive_brief(
    stores: list[StoreData],
    summary_df: pd.DataFrame,
    supplier_df: pd.DataFrame,
) -> None:
    """Short first-screen summary without repeating detailed analytics below."""
    store_summary = executive_store_summary(stores)
    segment_summary = network_segment_summary(stores)
    total_qty = int(summary_df["Количество"].sum())
    total_sales = float(summary_df["Выручка"].sum())
    average_item = total_sales / total_qty if total_qty else 0
    periods = sorted(set(summary_df["Период"].astype(str).tolist())) if "Период" in summary_df.columns else []
    period_label = periods[0] if len(periods) == 1 else f"{len(periods)} периода"

    st.markdown(
        '<div class="executive-banner executive-banner-compact"><div class="executive-banner-content">'
        '<div class="executive-eyebrow">ОБЩАЯ СВОДКА</div>'
        '<div class="executive-title">Результаты сети коротко</div>'
        '<div class="executive-copy">Основные цифры и фактические акценты. Подробности открываются в рабочем пространстве ниже.</div>'
        '</div></div>',
        unsafe_allow_html=True,
    )

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        kpi_card("Период", period_label)
    with k2:
        kpi_card("Выручка сети", usd_money(total_sales))
    with k3:
        kpi_card("Продано", f"{money(total_qty)} шт.")
    with k4:
        kpi_card("Средняя стоимость", usd_money(average_item), "выручка ÷ количество")
    with k5:
        kpi_card("Магазинов", str(len(stores)))

    if not store_summary.empty:
        retail_summary = retail_leader_summary(store_summary)
        revenue_leader = retail_summary.sort_values("Выручка", ascending=False).iloc[0]
        qty_leader = retail_summary.sort_values("Количество", ascending=False).iloc[0]
        avg_leader = retail_summary.sort_values("Средняя стоимость", ascending=False).iloc[0]
        segment_leader = segment_summary.sort_values("Выручка", ascending=False).iloc[0]
        l1, l2, l3, l4 = st.columns(4)
        with l1:
            leader_kpi_card(
                "Лидер розничной сети по выручке",
                escape(str(revenue_leader["Магазин"])),
                usd_money(float(revenue_leader["Выручка"])),
            )
        with l2:
            leader_kpi_card(
                "Лидер розничной сети по количеству",
                escape(str(qty_leader["Магазин"])),
                f"{money(float(qty_leader['Количество']))} шт.",
            )
        with l3:
            leader_kpi_card(
                "Самая высокая средняя стоимость",
                escape(str(avg_leader["Магазин"])),
                usd_money(float(avg_leader["Средняя стоимость"])),
            )
        with l4:
            leader_kpi_card(
                "Главный сегмент по выручке",
                escape(str(segment_leader["Сегмент"])),
                f"{pct(float(segment_leader['% выручки']))} выручки сети",
            )

    insight_panel(
        "Что важно",
        executive_insights(stores, store_summary, segment_summary, supplier_df)[:4],
    )

def segment_bar(df: pd.DataFrame, segment: str) -> go.Figure:
    qty_key = f"{SEGMENT_LABELS[segment]} — шт. %"
    sales_key = f"{SEGMENT_LABELS[segment]} — продажи %"
    fig = go.Figure()
    fig.add_bar(
        x=df["Магазин"], y=df[qty_key] * 100, name="Шт. %",
        marker_color=SEGMENT_COLORS[segment], text=[pct(v) for v in df[qty_key]], textposition="outside",
        hovertemplate="%{x}<br>Количество: %{y:.2f}%<extra></extra>",
    )
    fig.add_bar(
        x=df["Магазин"], y=df[sales_key] * 100, name="Продажи %",
        marker_color=LIGHT_COLORS[segment], text=[pct(v) for v in df[sales_key]], textposition="outside",
        hovertemplate="%{x}<br>Выручка: %{y:.2f}%<extra></extra>",
    )
    fig.update_layout(
        title=SEGMENT_LABELS[segment].upper(), barmode="group", height=380,
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=30, r=20, t=55, b=35), legend=dict(orientation="h", y=1.06),
        yaxis=dict(title="%", range=[0, 105], gridcolor="#ece8e1"),
        xaxis=dict(title=""), font=dict(family="Arial", color="#1c1813"),
    )
    return fig


def donut(labels: list[str], values: list[float], title: str, colors: list[str] | None = None, monetary: bool = False) -> go.Figure:
    # Outside labels need real breathing room in Streamlit columns.
    # `automargin` lets Plotly expand the drawable area instead of clipping callouts.
    display_values = [to_usd(value) for value in values] if monetary else values
    hover_value = "$%{value:,.0f}" if monetary else "%{value:,.2f}"
    pie_kwargs = {
        "labels": labels,
        "values": display_values,
        "hole": .58,
        "textinfo": "label+percent",
        "textposition": "auto",
        "automargin": True,
        "sort": False,
        "insidetextorientation": "horizontal",
        "hovertemplate": f"%{{label}}<br>{hover_value}<br>%{{percent}}<extra></extra>",
    }
    if colors:
        pie_kwargs["marker"] = dict(colors=colors)
    fig = go.Figure(go.Pie(**pie_kwargs))
    fig.update_traces(textfont=dict(size=11), outsidetextfont=dict(size=11))
    fig.update_layout(
        title=title, height=430, showlegend=False,
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=85, r=85, t=60, b=55),
        font=dict(family="Arial", color="#1c1813"),
    )
    return fig


def horizontal_bar(df: pd.DataFrame, label_col: str, value_col: str, title: str, suffix: str = "") -> go.Figure:
    clean = df[df[value_col] > 0].copy().sort_values(value_col, ascending=True)
    monetary = is_monetary_column(value_col)
    display_values = clean[value_col].astype(float) / analytics_fx_rate() if monetary else clean[value_col]
    labels = [f"${money(v)}" if monetary else f"{money(v)}{suffix}" for v in display_values]
    max_value = float(display_values.max()) if not clean.empty else 0.0

    # Reserve extra x-axis space for labels printed outside the bars.
    # Longer numbers receive a little more headroom.
    longest_label = max((len(label) for label in labels), default=0)
    headroom = 1.30 if longest_label >= 12 else 1.22
    x_range = [0, max_value * headroom] if max_value > 0 else None

    fig = go.Figure(go.Bar(
        x=display_values, y=clean[label_col], orientation="h",
        marker_color="#b7893f", text=labels, textposition="outside",
        cliponaxis=False, textfont=dict(size=11),
        hovertemplate="%{y}<br>%{text}<extra></extra>",
    ))
    fig.update_layout(
        title=title, height=max(330, 42 * len(clean) + 100),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=135, t=55, b=35),
        xaxis=dict(gridcolor="#ece8e1", range=x_range, automargin=True),
        yaxis=dict(title="", automargin=True),
    )
    return fig


def stone_dataframe(store) -> pd.DataFrame:
    rows = []
    for seg in SEG_ORDER:
        _, seg_amount = totals_for(store, seg=seg)
        for stone in STONE_ORDERS[seg]:
            q, a = totals_for(store, seg, stone)
            rows.append({
                "Сегмент": SEGMENT_LABELS[seg], "Камень": stone,
                "Количество": q, "% количества магазина": q / store.total_qty if store.total_qty else 0,
                "Выручка": a, "% выручки магазина": a / store.total_amount if store.total_amount else 0,
                "Средняя стоимость": a / q if q else 0,
                "% выручки сегмента": a / seg_amount if seg_amount else 0,
            })
    return pd.DataFrame(rows)


def product_dataframe(store, segment: str | None = None, stone: str | None = None) -> pd.DataFrame:
    rows: list[dict] = []
    stone_qty, stone_amount = totals_for(store, segment, stone) if segment and stone else (0, 0)
    for (seg, stone_name), products in store.data.items():
        if segment and seg != segment:
            continue
        if stone and stone_name != stone:
            continue
        for product, vals in products.items():
            qty = int(vals.get("qty", 0))
            amount = float(vals.get("amount", 0))
            if qty == 0 and amount == 0:
                continue
            rows.append({
                "Сегмент": SEGMENT_LABELS.get(seg, seg),
                "Камень": stone_name,
                "Номенклатурная группа": PRODUCT_LABELS.get(product, product),
                "Код группы": product,
                "Количество": qty,
                "Выручка": amount,
                "% количества магазина": qty / store.total_qty if store.total_qty else 0,
                "% выручки магазина": amount / store.total_amount if store.total_amount else 0,
                "% количества камня": qty / stone_qty if stone_qty else 0,
                "% выручки камня": amount / stone_amount if stone_amount else 0,
                "Средняя стоимость": amount / qty if qty else 0,
            })
    if not rows:
        return pd.DataFrame(columns=[
            "Сегмент", "Камень", "Номенклатурная группа", "Код группы", "Количество", "Выручка",
            "% количества магазина", "% выручки магазина", "% количества камня", "% выручки камня",
            "Средняя стоимость",
        ])
    order_map = {PRODUCT_LABELS.get(p, p): idx for idx, p in enumerate(PRODUCT_ORDER)}
    df = pd.DataFrame(rows)
    df["_order"] = df["Номенклатурная группа"].map(order_map).fillna(999)
    return df.sort_values(["_order", "Номенклатурная группа"]).drop(columns="_order")


def cross_store_product_dataframe(stores: list, segment: str, stone: str, product_label: str) -> pd.DataFrame:
    rows = []
    for store in stores:
        df = product_dataframe(store, segment, stone)
        selected = df[df["Номенклатурная группа"] == product_label]
        qty = int(selected["Количество"].sum()) if not selected.empty else 0
        amount = float(selected["Выручка"].sum()) if not selected.empty else 0
        rows.append({
            "Магазин": base_store_name(store.name),
            "Количество": qty,
            "Выручка": amount,
            "Средняя стоимость": amount / qty if qty else 0,
            "% количества магазина": qty / store.total_qty if store.total_qty else 0,
            "% выручки магазина": amount / store.total_amount if store.total_amount else 0,
        })
    return pd.DataFrame(rows)


def formatted_table(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare sortable tables and convert monetary values from VND to USD.

    Column keys stay unchanged so Streamlit keeps sorting the column selected by
    the user. The visible USD suffix is applied through column configuration.
    """
    display = df.copy()
    if "Код группы" in display.columns:
        display = display.drop(columns="Код группы")

    for col in list(display.columns):
        if is_monetary_column(str(col)):
            display[col] = pd.to_numeric(display[col], errors="coerce").fillna(0) / analytics_fx_rate()
    return display


def table_column_config(df: pd.DataFrame) -> dict:
    """Apply display formatting while preserving numeric sorting semantics."""
    config: dict = {}
    for col in df.columns:
        name = str(col)
        if name.startswith("%") or name.endswith(" %") or name.startswith("Δ %"):
            config[col] = st.column_config.NumberColumn(format="percent")
        elif name == "Количество" or name.startswith("Количество ·") or name == "Δ количества":
            config[col] = st.column_config.NumberColumn(format="localized", step=1)
        elif is_monetary_column(name):
            config[col] = st.column_config.NumberColumn(label=f"{name}, USD", format="localized", step=1)
    return config


def data_table(df: pd.DataFrame, *, key: str | None = None) -> None:
    """Render an interactive sortable table with true numeric columns."""
    display = formatted_table(df)
    st.dataframe(
        display,
        width="stretch",
        hide_index=True,
        key=key,
        column_config=table_column_config(display),
    )




def comparison_period_info(stores: list[StoreData]) -> tuple[str, object | None, object | None]:
    periods = [period for store in stores for period in store.periods]
    if not periods:
        return "Период не найден", None, None
    start = min(period[0] for period in periods)
    end = max(period[1] for period in periods)
    return f"{start:%d.%m.%Y} - {end:%d.%m.%Y}", start, end


def stores_fact_dataframe(stores: list[StoreData]) -> pd.DataFrame:
    """Normalize StoreData into a flat fact table for cross-period filtering."""
    rows: list[dict] = []
    for store in stores:
        store_name = base_store_name(store.name)
        for (segment, stone), products in store.data.items():
            for product, values in products.items():
                qty = int(values.get("qty", 0))
                sales = float(values.get("amount", 0))
                if qty == 0 and sales == 0:
                    continue
                rows.append({
                    "Магазин": store_name,
                    "Сегмент": SEGMENT_LABELS.get(segment, segment),
                    "Код сегмента": segment,
                    "Камень": stone,
                    "Номенклатурная группа": PRODUCT_LABELS.get(product, product),
                    "Код группы": product,
                    "Количество": qty,
                    "Выручка": sales,
                })
    columns = [
        "Магазин", "Сегмент", "Код сегмента", "Камень",
        "Номенклатурная группа", "Код группы", "Количество", "Выручка",
    ]
    return pd.DataFrame(rows, columns=columns)


def outlet_direction_frame(store: StoreData | None, direction: str = "GIFT TT") -> pd.DataFrame:
    """Return one comparison row for a separate OUTLET direction.

    GIFT TT is not a jewelry segment and must never appear in segment, stone or
    product filters. It is compared only against the same direction inside the
    OUTLET store comparison.
    """
    values = store.extras.get(direction, {}) if store is not None else {}
    qty = int(values.get("qty", 0) or 0)
    sales = float(values.get("amount", 0) or 0)
    return pd.DataFrame([{
        "Направление": direction,
        "Количество": qty,
        "Выручка": sales,
        "Средняя стоимость": sales / qty if qty else 0,
    }])


def aggregate_metrics(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=group_cols + ["Количество", "Выручка", "Средняя стоимость"])
    grouped = df.groupby(group_cols, as_index=False, dropna=False).agg(
        Количество=("Количество", "sum"),
        Выручка=("Выручка", "sum"),
    )
    grouped["Средняя стоимость"] = grouped["Выручка"] / grouped["Количество"].replace(0, pd.NA)
    grouped["Средняя стоимость"] = grouped["Средняя стоимость"].fillna(0)
    return grouped


def compare_metric_frames(
    first: pd.DataFrame,
    second: pd.DataFrame,
    keys: list[str],
    metrics: tuple[str, ...] = ("Количество", "Выручка", "Средняя стоимость"),
) -> pd.DataFrame:
    """Outer-join two periods and keep every numeric column truly sortable."""
    first_cols = keys + [metric for metric in metrics if metric in first.columns]
    second_cols = keys + [metric for metric in metrics if metric in second.columns]
    left = first[first_cols].copy() if not first.empty else pd.DataFrame(columns=first_cols)
    right = second[second_cols].copy() if not second.empty else pd.DataFrame(columns=second_cols)
    left = left.rename(columns={metric: f"{metric} · Период 1" for metric in metrics if metric in left.columns})
    right = right.rename(columns={metric: f"{metric} · Период 2" for metric in metrics if metric in right.columns})
    result = left.merge(right, on=keys, how="outer")

    for metric in metrics:
        first_col = f"{metric} · Период 1"
        second_col = f"{metric} · Период 2"
        if first_col not in result.columns:
            result[first_col] = 0.0
        if second_col not in result.columns:
            result[second_col] = 0.0
        result[first_col] = pd.to_numeric(result[first_col], errors="coerce").fillna(0)
        result[second_col] = pd.to_numeric(result[second_col], errors="coerce").fillna(0)
        delta_name = {
            "Количество": "Δ количества",
            "Выручка": "Δ выручки",
            "Средняя стоимость": "Δ средней стоимости",
        }.get(metric, f"Δ {metric.lower()}")
        pct_name = {
            "Количество": "Δ количества %",
            "Выручка": "Δ выручки %",
            "Средняя стоимость": "Δ средней стоимости %",
        }.get(metric, f"Δ {metric.lower()} %")
        result[delta_name] = result[second_col] - result[first_col]
        denominator = result[first_col].abs().replace(0, pd.NA)
        result[pct_name] = (result[delta_name] / denominator).astype("Float64")

    return result


def comparison_totals(stores: list[StoreData]) -> dict[str, float]:
    qty = sum(int(store.total_qty) for store in stores)
    sales = sum(float(store.total_amount) for store in stores)
    return {"Количество": qty, "Выручка": sales, "Средняя стоимость": sales / qty if qty else 0}


def delta_text(
    first: float,
    second: float,
    *,
    suffix: str = "",
    percent: bool = True,
    monetary: bool = False,
) -> str:
    delta = second - first
    sign = "+" if delta > 0 else ""
    if monetary:
        absolute = f"{sign}${money(to_usd(delta))}"
    else:
        absolute = f"{sign}{money(delta)}{suffix}"
    if not percent or first == 0:
        return absolute
    relative = delta / abs(first)
    return f"{absolute} · {sign}{pct(relative)}"


def comparison_bar(
    first: pd.DataFrame,
    second: pd.DataFrame,
    category: str,
    value: str,
    title: str,
    first_label: str,
    second_label: str,
) -> go.Figure:
    merged = first[[category, value]].merge(
        second[[category, value]], on=category, how="outer", suffixes=("_1", "_2")
    ).fillna(0)
    merged = merged.sort_values(f"{value}_2", ascending=True)
    monetary = is_monetary_column(value)
    first_values = merged[f"{value}_1"].astype(float) / analytics_fx_rate() if monetary else merged[f"{value}_1"]
    second_values = merged[f"{value}_2"].astype(float) / analytics_fx_rate() if monetary else merged[f"{value}_2"]
    max_value = max(float(first_values.max() or 0), float(second_values.max() or 0))
    first_text = [f"${money(v)}" if monetary else money(v) for v in first_values]
    second_text = [f"${money(v)}" if monetary else money(v) for v in second_values]
    hover_prefix = "$" if monetary else ""
    fig = go.Figure()
    fig.add_bar(
        x=first_values, y=merged[category], orientation="h", name=first_label,
        marker_color="#d8c3a0", text=first_text,
        textposition="outside", cliponaxis=False,
        hovertemplate=f"%{{y}}<br>{first_label}: {hover_prefix}%{{x:,.0f}}<extra></extra>",
    )
    fig.add_bar(
        x=second_values, y=merged[category], orientation="h", name=second_label,
        marker_color="#b7893f", text=second_text,
        textposition="outside", cliponaxis=False,
        hovertemplate=f"%{{y}}<br>{second_label}: {hover_prefix}%{{x:,.0f}}<extra></extra>",
    )
    fig.update_layout(
        title=title, barmode="group", height=max(360, 54 * len(merged) + 110),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=145, t=60, b=35),
        xaxis=dict(gridcolor="#ece8e1", range=[0, max_value * 1.28 if max_value else 1], fixedrange=True),
        yaxis=dict(title="", automargin=True, fixedrange=True),
        legend=dict(orientation="h", y=1.08),
    )
    return fig


def jewelry_detail_scope(detail: pd.DataFrame) -> pd.DataFrame:
    """Keep the main jewelry network and exclude auxiliary OUTLET directions."""
    if detail.empty or "Магазин" not in detail.columns:
        return detail.copy()
    return detail.loc[~detail["Магазин"].isin({"GIFT TT", "CAFE"})].copy()


def metal_purity_summary(detail: pd.DataFrame) -> pd.DataFrame:
    keys = ["Группа металла", "Проба"]
    detail = jewelry_detail_scope(detail)
    if detail.empty:
        return pd.DataFrame(columns=keys + ["Количество", "Выручка", "Средняя стоимость"])
    return aggregate_metrics(detail, keys)


def metal_comparison_chart(
    first_detail: pd.DataFrame,
    second_detail: pd.DataFrame,
    first_label: str,
    second_label: str,
) -> go.Figure:
    """One responsive figure with pieces and USD totals by purity."""
    first = metal_purity_summary(first_detail)
    second = metal_purity_summary(second_detail)
    keys = ["Группа металла", "Проба"]
    merged = first.merge(second, on=keys, how="outer", suffixes=("_1", "_2")).fillna(0)
    if merged.empty:
        merged = pd.DataFrame({
            "Группа металла": [], "Проба": [],
            "Количество_1": [], "Количество_2": [], "Выручка_1": [], "Выручка_2": [],
        })
    order = {name: index for index, name in enumerate(METAL_GROUPS)}
    merged["_group_order"] = merged["Группа металла"].map(order).fillna(len(order))
    merged = merged.sort_values(["_group_order", "Проба"])
    labels = [f"{group} · {purity}" for group, purity in zip(merged["Группа металла"], merged["Проба"])]
    qty_first = pd.to_numeric(merged.get("Количество_1", 0), errors="coerce").fillna(0)
    qty_second = pd.to_numeric(merged.get("Количество_2", 0), errors="coerce").fillna(0)
    sales_first = pd.to_numeric(merged.get("Выручка_1", 0), errors="coerce").fillna(0) / analytics_fx_rate()
    sales_second = pd.to_numeric(merged.get("Выручка_2", 0), errors="coerce").fillna(0) / analytics_fx_rate()

    fig = make_subplots(
        rows=2,
        cols=1,
        subplot_titles=("Продано по пробам, шт.", "Продано по пробам, USD"),
        vertical_spacing=0.2,
    )
    fig.add_bar(
        x=labels, y=qty_first, name=first_label, marker_color="#d8c3a0",
        text=[money(value) for value in qty_first], textposition="outside",
        row=1, col=1,
    )
    fig.add_bar(
        x=labels, y=qty_second, name=second_label, marker_color="#b7893f",
        text=[money(value) for value in qty_second], textposition="outside",
        row=1, col=1,
    )
    fig.add_bar(
        x=labels, y=sales_first, name=first_label, marker_color="#d8c3a0",
        text=[f"${money(value)}" for value in sales_first], textposition="outside",
        showlegend=False, row=2, col=1,
    )
    fig.add_bar(
        x=labels, y=sales_second, name=second_label, marker_color="#b7893f",
        text=[f"${money(value)}" for value in sales_second], textposition="outside",
        showlegend=False, row=2, col=1,
    )
    fig.update_layout(
        barmode="group",
        height=max(650, 58 * max(len(labels), 1) + 430),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=25, r=30, t=70, b=90),
        legend=dict(orientation="h", y=1.08),
    )
    fig.update_xaxes(tickangle=-25, automargin=True, fixedrange=True)
    fig.update_yaxes(gridcolor="#ece8e1", fixedrange=True, rangemode="tozero")
    return fig


def render_comparison_metal_section(
    first_detail: pd.DataFrame,
    second_detail: pd.DataFrame,
    first_label: str,
    second_label: str,
) -> None:
    first_detail = jewelry_detail_scope(first_detail)
    second_detail = jewelry_detail_scope(second_detail)
    if first_detail.empty and second_detail.empty:
        st.info("После выбранного фильтра данных по пробам нет.")
        return
    locked_plotly_chart(
        metal_comparison_chart(first_detail, second_detail, first_label, second_label),
        width="stretch",
        key="comparison_metal_purity_chart",
    )
    detail_keys = [
        "Группа металла", "Проба", "Сегмент", "Камень", "Номенклатурная группа",
    ]
    first_table = aggregate_metrics(first_detail, detail_keys)
    second_table = aggregate_metrics(second_detail, detail_keys)
    comparison = annotate_change_status(compare_metric_frames(first_table, second_table, detail_keys), 3)
    sort_column = "Выручка · Период 2" if "Выручка · Период 2" in comparison.columns else detail_keys[0]
    with st.expander("Полная таблица по пробам, камням и группам", expanded=False):
        data_table(comparison.sort_values(sort_column, ascending=False), key="comparison_metal_detail_table")


def render_comparison_period_cards(
    title: str,
    first_metrics: dict[str, float],
    second_metrics: dict[str, float],
    first_label: str,
    second_label: str,
) -> None:
    st.markdown(f"### {title}")
    left, right = st.columns(2)
    for column, label, values in [
        (left, first_label, first_metrics),
        (right, second_label, second_metrics),
    ]:
        with column:
            st.markdown(f'<div class="compare-period-title">{escape(label)}</div>', unsafe_allow_html=True)
            a, b, c = st.columns(3)
            with a:
                kpi_card("Выручка", usd_money(values.get("Выручка", 0)))
            with b:
                kpi_card("Количество", f"{money(values.get('Количество', 0))} шт.")
            with c:
                kpi_card("Средняя стоимость", usd_money(values.get("Средняя стоимость", 0)))


def render_comparison_summary(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
) -> None:
    first_totals = comparison_totals(stores_first)
    second_totals = comparison_totals(stores_second)
    render_comparison_period_cards("Сеть целиком", first_totals, second_totals, first_label, second_label)

    d1, d2, d3 = st.columns(3)
    with d1:
        kpi_card("Изменение выручки", delta_text(first_totals["Выручка"], second_totals["Выручка"], monetary=True))
    with d2:
        kpi_card("Изменение количества", delta_text(first_totals["Количество"], second_totals["Количество"], suffix=" шт."))
    with d3:
        kpi_card(
            "Изменение средней стоимости",
            delta_text(first_totals["Средняя стоимость"], second_totals["Средняя стоимость"], monetary=True),
        )

    first_store = network_summary(stores_first)
    second_store = network_summary(stores_second)
    store_compare = compare_metric_frames(first_store, second_store, ["Магазин"])
    st.markdown("### Сравнение магазинов")
    locked_plotly_chart(
        comparison_bar(
            first_store, second_store, "Магазин", "Выручка",
            "Выручка по магазинам: два периода", first_label, second_label,
        ),
        width="stretch",
        key="comparison_network_store_chart",
    )
    data_table(store_compare.sort_values("Выручка · Период 2", ascending=False), key="comparison_store_table")

    first_segment = network_segment_summary(stores_first)
    second_segment = network_segment_summary(stores_second)
    segment_compare = compare_metric_frames(first_segment, second_segment, ["Сегмент"])
    st.markdown("### Сегменты сети")
    data_table(segment_compare.sort_values("Выручка · Период 2", ascending=False), key="comparison_segment_table")


@st.fragment
def render_comparison_store_fragment(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
) -> None:
    names = sorted({base_store_name(store.name) for store in stores_first + stores_second})
    selected = st.selectbox("Выберите магазин", names, key="comparison_store_select")
    first_store = next((store for store in stores_first if base_store_name(store.name) == selected), None)
    second_store = next((store for store in stores_second if base_store_name(store.name) == selected), None)

    def metrics(store: StoreData | None) -> dict[str, float]:
        if store is None:
            return {"Количество": 0, "Выручка": 0, "Средняя стоимость": 0}
        return {
            "Количество": int(store.total_qty),
            "Выручка": float(store.total_amount),
            "Средняя стоимость": float(store.total_amount) / int(store.total_qty) if store.total_qty else 0,
        }

    render_comparison_period_cards(selected, metrics(first_store), metrics(second_store), first_label, second_label)

    first_segments = network_segment_summary([first_store]) if first_store else pd.DataFrame()
    second_segments = network_segment_summary([second_store]) if second_store else pd.DataFrame()
    segment_compare = compare_metric_frames(first_segments, second_segments, ["Сегмент"])
    data_table(segment_compare, key="comparison_selected_store_segments")

    if selected == "OUTLET":
        st.markdown("#### Отдельные направления OUTLET")
        st.caption("GIFT TT не относится к сегментам и сравнивается только с GIFT TT второго периода.")
        gift_compare = compare_metric_frames(
            outlet_direction_frame(first_store, "GIFT TT"),
            outlet_direction_frame(second_store, "GIFT TT"),
            ["Направление"],
        )
        data_table(gift_compare, key="comparison_outlet_gift_tt")


@st.fragment
def render_comparison_interactive_fragment(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
) -> None:
    first_facts = stores_fact_dataframe(stores_first)
    second_facts = stores_fact_dataframe(stores_second)
    combined = pd.concat([first_facts, second_facts], ignore_index=True)
    if combined.empty:
        st.info("В отчетах нет данных для интерактивного сравнения.")
        return

    store_options = ["Все магазины"] + sorted(combined["Магазин"].dropna().astype(str).unique().tolist())
    segment_options = ["Все сегменты"] + sorted(combined["Сегмент"].dropna().astype(str).unique().tolist())
    c1, c2 = st.columns(2)
    with c1:
        selected_store = st.selectbox("Магазин", store_options, key="comparison_interactive_store")
    with c2:
        selected_segment = st.selectbox("Сегмент", segment_options, key="comparison_interactive_segment")

    filtered_combined = combined.copy()
    if selected_store != "Все магазины":
        filtered_combined = filtered_combined[filtered_combined["Магазин"] == selected_store]
    if selected_segment != "Все сегменты":
        filtered_combined = filtered_combined[filtered_combined["Сегмент"] == selected_segment]

    stone_options = ["Все камни"] + sorted(filtered_combined["Камень"].dropna().astype(str).unique().tolist())
    selected_stone = st.selectbox("Камень / группа камней", stone_options, key="comparison_interactive_stone")
    if selected_stone != "Все камни":
        filtered_combined = filtered_combined[filtered_combined["Камень"] == selected_stone]

    product_options = ["Все номенклатурные группы"] + sorted(
        filtered_combined["Номенклатурная группа"].dropna().astype(str).unique().tolist()
    )
    selected_product = st.selectbox(
        "Номенклатурная группа", product_options, key="comparison_interactive_product"
    )

    def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
        result = df.copy()
        if selected_store != "Все магазины":
            result = result[result["Магазин"] == selected_store]
        if selected_segment != "Все сегменты":
            result = result[result["Сегмент"] == selected_segment]
        if selected_stone != "Все камни":
            result = result[result["Камень"] == selected_stone]
        if selected_product != "Все номенклатурные группы":
            result = result[result["Номенклатурная группа"] == selected_product]
        return result

    first_filtered = apply_filters(first_facts)
    second_filtered = apply_filters(second_facts)

    def total_metrics(df: pd.DataFrame) -> dict[str, float]:
        qty = float(df["Количество"].sum()) if not df.empty else 0
        sales = float(df["Выручка"].sum()) if not df.empty else 0
        return {"Количество": qty, "Выручка": sales, "Средняя стоимость": sales / qty if qty else 0}

    render_comparison_period_cards(
        "Выбранный срез", total_metrics(first_filtered), total_metrics(second_filtered), first_label, second_label
    )

    if selected_store == "Все магазины":
        group = ["Магазин"]
    elif selected_segment == "Все сегменты":
        group = ["Сегмент"]
    elif selected_stone == "Все камни":
        group = ["Камень"]
    else:
        group = ["Номенклатурная группа"]
    first_grouped = aggregate_metrics(first_filtered, group)
    second_grouped = aggregate_metrics(second_filtered, group)
    comparison = compare_metric_frames(first_grouped, second_grouped, group)
    data_table(comparison, key="comparison_interactive_table")


@st.fragment
def render_comparison_supplier_fragment(
    first_supplier_df: pd.DataFrame,
    second_supplier_df: pd.DataFrame,
    first_label: str,
    second_label: str,
) -> None:
    if not supplier_has_meaningful_detail(first_supplier_df) and not supplier_has_meaningful_detail(second_supplier_df):
        st.info("В этих выгрузках нет полноценной детализации по поставщикам.")
        return

    first_summary = supplier_summary(first_supplier_df)
    second_summary = supplier_summary(second_supplier_df)
    comparison = annotate_change_status(compare_metric_frames(first_summary, second_summary, ["Поставщик"]), 3)
    with st.expander("Полная таблица поставщиков", expanded=False):
        data_table(comparison.sort_values("Выручка · Период 2", ascending=False), key="comparison_supplier_table")

    names = sorted(set(first_summary.get("Поставщик", pd.Series(dtype=str)).astype(str)) |
                   set(second_summary.get("Поставщик", pd.Series(dtype=str)).astype(str)))
    if not names:
        return
    selected = st.selectbox("Поставщик", names, key="comparison_supplier_select")

    def supplier_metrics(summary: pd.DataFrame) -> dict[str, float]:
        row = summary[summary["Поставщик"] == selected]
        if row.empty:
            return {"Количество": 0, "Выручка": 0, "Средняя стоимость": 0}
        current = row.iloc[0]
        return {
            "Количество": float(current["Количество"]),
            "Выручка": float(current["Выручка"]),
            "Средняя стоимость": float(current["Средняя стоимость"]),
        }

    render_comparison_period_cards(
        selected, supplier_metrics(first_summary), supplier_metrics(second_summary), first_label, second_label
    )

    if "Магазин" in first_supplier_df.columns or "Магазин" in second_supplier_df.columns:
        first_detail = first_supplier_df[first_supplier_df["Поставщик"] == selected]
        second_detail = second_supplier_df[second_supplier_df["Поставщик"] == selected]
        first_stores = aggregate_metrics(first_detail, ["Магазин"])
        second_stores = aggregate_metrics(second_detail, ["Магазин"])
        by_store = compare_metric_frames(first_stores, second_stores, ["Магазин"])
        st.markdown("#### Поставщик по магазинам")
        data_table(by_store, key="comparison_supplier_store_table")



def period_days(start, end) -> int | None:
    if start is None or end is None:
        return None
    return max(1, int((end - start).days) + 1)


def comparison_metric_card(
    label: str,
    first_value: str,
    second_value: str,
    delta: str,
    first_label: str,
    second_label: str,
) -> None:
    tone = "delta-negative" if str(delta).lstrip().startswith("-") else "delta-positive" if str(delta).lstrip().startswith("+") else "delta-neutral"
    st.markdown(
        '<div class="comparison-metric-card">'
        f'<div class="comparison-metric-label">{escape(label)}</div>'
        '<div class="comparison-metric-values">'
        f'<div><span>{escape(first_label)}</span><b>{escape(first_value)}</b></div>'
        f'<div class="comparison-arrow">→</div>'
        f'<div><span>{escape(second_label)}</span><b>{escape(second_value)}</b></div>'
        '</div>'
        f'<div class="comparison-metric-delta {tone}">{escape(delta)}</div>'
        '</div>',
        unsafe_allow_html=True,
    )


def render_comparison_kpi_strip(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
    first_start,
    first_end,
    second_start,
    second_end,
) -> tuple[int | None, int | None]:
    first = comparison_totals(stores_first)
    second = comparison_totals(stores_second)
    first_days = period_days(first_start, first_end)
    second_days = period_days(second_start, second_end)
    first_daily = first["Выручка"] / first_days if first_days else 0
    second_daily = second["Выручка"] / second_days if second_days else 0

    st.markdown(
        '<div class="comparison-period-strip">'
        f'<div><b>{escape(first_label)}</b><span>{first_days or "—"} дней</span></div>'
        '<div class="comparison-period-arrow">→</div>'
        f'<div><b>{escape(second_label)}</b><span>{second_days or "—"} дней</span></div>'
        '</div>',
        unsafe_allow_html=True,
    )
    cols = st.columns(4)
    with cols[0]:
        comparison_metric_card(
            "Выручка",
            usd_money(first["Выручка"]),
            usd_money(second["Выручка"]),
            delta_text(first["Выручка"], second["Выручка"], monetary=True),
            first_label,
            second_label,
        )
    with cols[1]:
        comparison_metric_card(
            "Количество",
            f"{money(first['Количество'])} шт.",
            f"{money(second['Количество'])} шт.",
            delta_text(first["Количество"], second["Количество"], suffix=" шт."),
            first_label,
            second_label,
        )
    with cols[2]:
        comparison_metric_card(
            "Средняя стоимость",
            usd_money(first["Средняя стоимость"]),
            usd_money(second["Средняя стоимость"]),
            delta_text(first["Средняя стоимость"], second["Средняя стоимость"], monetary=True),
            first_label,
            second_label,
        )
    with cols[3]:
        comparison_metric_card(
            "Выручка в день",
            usd_money(first_daily),
            usd_money(second_daily),
            delta_text(first_daily, second_daily, monetary=True),
            first_label,
            second_label,
        )

    if first_days and second_days and first_days != second_days:
        st.warning(
            f"Периоды разной длительности: {first_days} и {second_days} дней. "
            "Для корректной оценки динамики ориентируйтесь также на выручку и количество в день."
        )
    return first_days, second_days


def annotate_change_status(frame: pd.DataFrame, threshold_percent: int = 0) -> pd.DataFrame:
    result = frame.copy()
    first_col = "Выручка · Период 1"
    second_col = "Выручка · Период 2"
    if first_col not in result.columns or second_col not in result.columns:
        result["Статус"] = "Без данных"
        return result
    threshold = max(0, threshold_percent) / 100

    def status(row) -> str:
        first = float(row[first_col])
        second = float(row[second_col])
        if first == 0 and second > 0:
            return "Новая группа"
        if first > 0 and second == 0:
            return "Исчезла из продаж"
        if first == 0 and second == 0:
            return "Без продаж"
        change = (second - first) / abs(first)
        if change >= threshold and second > first:
            return "Рост"
        if change <= -threshold and second < first:
            return "Снижение"
        return "Без существенного изменения"

    result["Статус"] = result.apply(status, axis=1)
    return result


def change_bar(frame: pd.DataFrame, key_col: str, title: str, positive: bool) -> go.Figure:
    data = frame.copy()
    data = data.loc[data["Δ выручки"] > 0] if positive else data.loc[data["Δ выручки"] < 0]
    data = data.nlargest(5, "Δ выручки") if positive else data.nsmallest(5, "Δ выручки")
    data = data.sort_values("Δ выручки", ascending=True)
    values = data["Δ выручки"].astype(float) / analytics_fx_rate()
    labels = data[key_col].astype(str)
    fig = go.Figure(go.Bar(
        x=values,
        y=labels,
        orientation="h",
        marker_color="#4f8a5b" if positive else "#b6534f",
        text=[f"{value:+,.0f} $".replace(",", " ") for value in values],
        textposition="outside",
        cliponaxis=False,
        hovertemplate="%{y}<br>%{x:+,.0f} $<extra></extra>",
    ))
    max_abs = max([abs(float(v)) for v in values], default=1)
    fig.update_layout(
        title=title,
        height=max(300, 48 * max(len(data), 1) + 120),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=115, t=60, b=30),
        xaxis=dict(gridcolor="#ece8e1", range=[-max_abs * 1.25 if not positive else 0, max_abs * 1.25 if positive else 0]),
        yaxis=dict(title="", automargin=True),
        showlegend=False,
    )
    return fig


def comparison_driver_frames(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    supplier_first: pd.DataFrame,
    supplier_second: pd.DataFrame,
) -> dict[str, tuple[pd.DataFrame, pd.DataFrame, str]]:
    first_facts = stores_fact_dataframe(stores_first)
    second_facts = stores_fact_dataframe(stores_second)
    frames: dict[str, tuple[pd.DataFrame, pd.DataFrame, str]] = {
        "Магазины": (network_summary(stores_first), network_summary(stores_second), "Магазин"),
        "Камни": (aggregate_metrics(first_facts, ["Камень"]), aggregate_metrics(second_facts, ["Камень"]), "Камень"),
        "Номенклатурные группы": (
            aggregate_metrics(first_facts, ["Номенклатурная группа"]),
            aggregate_metrics(second_facts, ["Номенклатурная группа"]),
            "Номенклатурная группа",
        ),
    }
    if supplier_has_meaningful_detail(supplier_first) or supplier_has_meaningful_detail(supplier_second):
        frames["Поставщики"] = (supplier_summary(supplier_first), supplier_summary(supplier_second), "Поставщик")
    return frames


@st.fragment
def render_comparison_drivers_fragment(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    supplier_first: pd.DataFrame,
    supplier_second: pd.DataFrame,
) -> None:
    frames = comparison_driver_frames(stores_first, stores_second, supplier_first, supplier_second)
    left_control, right_control = st.columns(2)
    with left_control:
        selected = st.selectbox("Что анализировать", list(frames), key="comparison_driver_dimension")
    with right_control:
        threshold = st.select_slider(
            "Порог существенного изменения",
            options=[0, 3, 5, 10],
            value=3,
            format_func=lambda value: f"{value}%",
            key="comparison_driver_threshold",
        )
    first, second, key_col = frames[selected]
    comparison = compare_metric_frames(first, second, [key_col])
    comparison = annotate_change_status(comparison, threshold)

    chart_left, chart_right = st.columns(2)
    with chart_left:
        locked_plotly_chart(
            change_bar(comparison, key_col, "Главные источники роста", True),
            width="stretch",
            key=f"driver_growth_{selected}_{threshold}",
        )
    with chart_right:
        locked_plotly_chart(
            change_bar(comparison, key_col, "Главные источники снижения", False),
            width="stretch",
            key=f"driver_decline_{selected}_{threshold}",
        )

    new_count = int((comparison["Статус"] == "Новая группа").sum())
    lost_count = int((comparison["Статус"] == "Исчезла из продаж").sum())
    st.caption(f"Новых: {new_count} · исчезнувших из продаж: {lost_count} · порог: {threshold}%")
    with st.expander("Полная таблица драйверов", expanded=False):
        data_table(comparison.sort_values("Δ выручки", ascending=False), key="comparison_drivers_full_table")


def render_comparison_overview_detail(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
) -> None:
    first_store = network_summary(stores_first)
    second_store = network_summary(stores_second)
    store_compare = annotate_change_status(compare_metric_frames(first_store, second_store, ["Магазин"]), 3)

    ranking = store_compare.assign(
        _max=store_compare[["Выручка · Период 1", "Выручка · Период 2"]].max(axis=1)
    ).nlargest(10, "_max")
    first_top = ranking[["Магазин", "Выручка · Период 1"]].rename(columns={"Выручка · Период 1": "Выручка"})
    second_top = ranking[["Магазин", "Выручка · Период 2"]].rename(columns={"Выручка · Период 2": "Выручка"})
    locked_plotly_chart(
        comparison_bar(
            first_top,
            second_top,
            "Магазин",
            "Выручка",
            "Топ-10 магазинов: два периода",
            first_label,
            second_label,
        ),
        width="stretch",
        key="comparison_overview_store_chart",
    )

    first_segment = network_segment_summary(stores_first)
    second_segment = network_segment_summary(stores_second)
    segment_compare = compare_metric_frames(first_segment, second_segment, ["Сегмент"])
    first_shares = first_segment[["Сегмент", "% количества", "% выручки"]].rename(columns={
        "% количества": "% количества · Период 1",
        "% выручки": "% выручки · Период 1",
    })
    second_shares = second_segment[["Сегмент", "% количества", "% выручки"]].rename(columns={
        "% количества": "% количества · Период 2",
        "% выручки": "% выручки · Период 2",
    })
    segment_compare = segment_compare.merge(first_shares, on="Сегмент", how="left").merge(
        second_shares, on="Сегмент", how="left"
    )
    segment_compare = annotate_change_status(segment_compare, 3)
    st.markdown("#### Взвешенная структура сегментов")
    st.caption("Доли каждого периода рассчитываются от итогов всей сети.")
    data_table(segment_compare.sort_values("Выручка · Период 2", ascending=False), key="comparison_segment_weighted")

    with st.expander("Полное сравнение магазинов", expanded=False):
        data_table(store_compare.sort_values("Выручка · Период 2", ascending=False), key="comparison_store_full_table")


@st.fragment
def render_comparison_stones_groups_fragment(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    first_label: str,
    second_label: str,
) -> None:
    first_facts = stores_fact_dataframe(stores_first)
    second_facts = stores_fact_dataframe(stores_second)
    dimensions = {
        "Сегменты": "Сегмент",
        "Камни": "Камень",
        "Номенклатурные группы": "Номенклатурная группа",
    }
    selected_dimension = st.segmented_control(
        "Уровень детализации",
        tuple(dimensions),
        default="Камни",
        key="comparison_stone_group_dimension",
    ) or "Камни"
    key_col = dimensions[selected_dimension]
    first = aggregate_metrics(first_facts, [key_col])
    second = aggregate_metrics(second_facts, [key_col])
    comparison = annotate_change_status(compare_metric_frames(first, second, [key_col]), 3)

    top = comparison.assign(
        _max=comparison[["Выручка · Период 1", "Выручка · Период 2"]].max(axis=1)
    ).nlargest(12, "_max")
    first_top = top[[key_col, "Выручка · Период 1"]].rename(columns={"Выручка · Период 1": "Выручка"})
    second_top = top[[key_col, "Выручка · Период 2"]].rename(columns={"Выручка · Период 2": "Выручка"})
    locked_plotly_chart(
        comparison_bar(
            first_top,
            second_top,
            key_col,
            "Выручка",
            f"{selected_dimension}: два периода",
            first_label,
            second_label,
        ),
        width="stretch",
        key=f"comparison_stone_group_chart_{selected_dimension}",
    )

    names = comparison[key_col].dropna().astype(str).tolist()
    if names:
        selected_name = st.selectbox(key_col, names, key="comparison_stone_group_selected")
        row = comparison.loc[comparison[key_col].astype(str) == selected_name].iloc[0]
        render_comparison_period_cards(
            selected_name,
            {
                "Количество": float(row["Количество · Период 1"]),
                "Выручка": float(row["Выручка · Период 1"]),
                "Средняя стоимость": float(row["Средняя стоимость · Период 1"]),
            },
            {
                "Количество": float(row["Количество · Период 2"]),
                "Выручка": float(row["Выручка · Период 2"]),
                "Средняя стоимость": float(row["Средняя стоимость · Период 2"]),
            },
            first_label,
            second_label,
        )
        st.caption(f"Статус: {row['Статус']}")

    with st.expander(f"Полная таблица: {selected_dimension.lower()}", expanded=False):
        data_table(comparison.sort_values("Выручка · Период 2", ascending=False), key="comparison_stone_group_full")

def render_comparison_report(
    stores_first: list[StoreData],
    stores_second: list[StoreData],
    supplier_first: pd.DataFrame,
    supplier_second: pd.DataFrame,
    first_label: str,
    second_label: str,
    first_start=None,
    first_end=None,
    second_start=None,
    second_end=None,
) -> None:
    st.markdown('<div id="comparison-workspace"></div>', unsafe_allow_html=True)
    section_divider(
        "Сравнение периодов",
        "Короткий итог всегда остается сверху; ниже открывается только выбранный рабочий раздел.",
        "СРАВНИТЕЛЬНЫЙ АНАЛИЗ",
    )
    render_comparison_kpi_strip(
        stores_first,
        stores_second,
        first_label,
        second_label,
        first_start,
        first_end,
        second_start,
        second_end,
    )

    first_totals = comparison_totals(stores_first)
    second_totals = comparison_totals(stores_second)
    revenue_delta = second_totals["Выручка"] - first_totals["Выручка"]
    direction = "выросла" if revenue_delta > 0 else "снизилась" if revenue_delta < 0 else "не изменилась"
    volume_effect = (second_totals["Количество"] - first_totals["Количество"]) * first_totals["Средняя стоимость"]
    value_effect = second_totals["Количество"] * (
        second_totals["Средняя стоимость"] - first_totals["Средняя стоимость"]
    )
    insight_panel(
        "Итог одним абзацем",
        [
            f"Выручка {direction}: {delta_text(first_totals['Выручка'], second_totals['Выручка'], monetary=True)}.",
            f"Количество: {delta_text(first_totals['Количество'], second_totals['Количество'], suffix=' шт.')}.",
            f"Средняя стоимость: {delta_text(first_totals['Средняя стоимость'], second_totals['Средняя стоимость'], monetary=True)}.",
            f"Расчётный вклад количества — {delta_text(0, volume_effect, percent=False, monetary=True)}, вклад средней стоимости и структуры — {delta_text(0, value_effect, percent=False, monetary=True)}.",
        ],
    )

    options = (
        "Итог изменений",
        "Драйверы",
        "Магазины",
        "Камни и группы",
        "Металлы и пробы",
        "Поставщики",
        "Исследование данных",
    )
    selected = st.segmented_control(
        "Раздел сравнения",
        options,
        default=options[0],
        key="comparison_workspace",
    ) or options[0]

    if selected == "Итог изменений":
        render_comparison_overview_detail(stores_first, stores_second, first_label, second_label)
    elif selected == "Драйверы":
        render_comparison_drivers_fragment(
            stores_first,
            stores_second,
            supplier_first,
            supplier_second,
        )
    elif selected == "Магазины":
        render_comparison_store_fragment(stores_first, stores_second, first_label, second_label)
    elif selected == "Камни и группы":
        render_comparison_stones_groups_fragment(stores_first, stores_second, first_label, second_label)
    elif selected == "Металлы и пробы":
        render_comparison_metal_section(supplier_first, supplier_second, first_label, second_label)
    elif selected == "Поставщики":
        render_comparison_supplier_fragment(supplier_first, supplier_second, first_label, second_label)
    else:
        render_comparison_interactive_fragment(stores_first, stores_second, first_label, second_label)


def section_divider(title: str, subtitle: str = "", kicker: str = "ANALITIKA") -> None:
    st.markdown(
        f'<div class="section-divider">'
        f'<div class="section-divider-kicker">{kicker}</div>'
        f'<div class="section-divider-title">{title}</div>'
        f'<div class="section-divider-copy">{subtitle}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def insight_panel(title: str, lines: list[str]) -> None:
    clean = [line for line in lines if line]
    if not clean:
        return
    body = "".join(f'<div class="analysis-line">• {line}</div>' for line in clean)
    st.markdown(
        f'<div class="analysis-panel"><div class="analysis-panel-title">{title}</div>{body}</div>',
        unsafe_allow_html=True,
    )


def network_conclusions(summary_df: pd.DataFrame) -> list[str]:
    """Network observations with retail-only leaders and weighted segment shares."""
    if summary_df.empty:
        return []
    lines: list[str] = []
    retail = retail_leader_summary(summary_df)
    leader = retail.sort_values("Выручка", ascending=False).iloc[0]
    lines.append(f"Лидер розничной сети по выручке — {leader['Магазин']}: {usd_money(leader['Выручка'])}.")
    qty_leader = retail.sort_values("Количество", ascending=False).iloc[0]
    lines.append(f"Больше всего изделий в розничной сети продано в {qty_leader['Магазин']} — {money(qty_leader['Количество'])} шт.")
    total_sales = float(summary_df["Выручка"].sum())
    weighted_segments: dict[str, float] = {}
    for seg in SEG_ORDER:
        col = f"{SEGMENT_LABELS[seg]} — продажи %"
        if col in summary_df.columns and total_sales:
            weighted_segments[SEGMENT_LABELS[seg]] = float((summary_df[col] * summary_df["Выручка"]).sum()) / total_sales
    if weighted_segments:
        seg_name, seg_share = max(weighted_segments.items(), key=lambda item: item[1])
        lines.append(f"Доминирующий сегмент сети — {seg_name}: {pct(seg_share)} общей выручки.")
    return lines[:4]


def interactive_conclusions(store, segment: str, stone: str, product_df: pd.DataFrame, selected_product: str, qty: int, sales: float) -> list[str]:
    lines: list[str] = []
    lines.append(f"Текущий фильтр: {base_store_name(store.name)} → {SEGMENT_LABELS[segment]} → {stone}.")
    if selected_product != "Все номенклатурные группы":
        lines.append(f"Группа «{selected_product}» формирует {pct(sales / store.total_amount if store.total_amount else 0)} выручки магазина и {pct(qty / store.total_qty if store.total_qty else 0)} количества.")
    elif not product_df.empty:
        top = product_df.sort_values("Выручка", ascending=False).iloc[0]
        lines.append(f"Лидер по выручке внутри {stone} — «{top['Номенклатурная группа']}»: {pct(float(top['% выручки камня']))}.")
        if len(product_df) > 1:
            low = product_df[product_df["Количество"] > 0].sort_values("Выручка", ascending=True)
            if not low.empty:
                row = low.iloc[0]
                lines.append(f"Минимальная представленность — «{row['Номенклатурная группа']}»: {money(row['Количество'])} шт.")
    avg = sales / qty if qty else 0
    if avg:
        lines.append(f"Средняя стоимость в выбранном срезе — {usd_money(avg)}.")
    return lines[:4]


def supplier_conclusions(df: pd.DataFrame, summary: pd.DataFrame) -> list[str]:
    if df.empty or summary.empty:
        return []
    lines: list[str] = []
    leader = summary.iloc[0]
    lines.append(f"Лидер среди поставщиков — {leader['Поставщик']}: {pct(float(leader['% выручки']))} общей выручки.")
    qty_leader = summary.sort_values("Количество", ascending=False).iloc[0]
    lines.append(f"По количеству лидирует {qty_leader['Поставщик']} — {money(qty_leader['Количество'])} шт.")
    if "Магазин" in df.columns:
        coverage = df.groupby("Поставщик")["Магазин"].nunique().sort_values(ascending=False)
        if not coverage.empty:
            lines.append(f"Самое широкое покрытие у {coverage.index[0]} — {int(coverage.iloc[0])} магазинов.")
    return lines[:4]


def conclusions(store, all_stores: list) -> list[str]:
    lines: list[str] = []
    seg = segment_totals(store)
    if store.total_amount:
        leader = max(SEG_ORDER, key=lambda x: seg[x]["amount"])
        share = seg[leader]["amount"] / store.total_amount
        lines.append(f"Основную выручку формирует {SEGMENT_LABELS[leader]} — {pct(share)}.")
    network_avg = sum(s.total_amount for s in all_stores) / max(1, sum(s.total_qty for s in all_stores))
    store_avg = store.total_amount / store.total_qty if store.total_qty else 0
    if network_avg:
        delta = store_avg / network_avg - 1
        direction = "выше" if delta >= 0 else "ниже"
        lines.append(f"Средняя стоимость изделия {direction} средней по сети на {pct(abs(delta))}.")
    top_stones = [(stone, totals_for(store, "TOP STONES", stone)[1]) for stone in TOP_ORDER]
    top_stones = [x for x in top_stones if x[1] > 0]
    if top_stones:
        name, amount = max(top_stones, key=lambda x: x[1])
        top_total = seg["TOP STONES"]["amount"]
        lines.append(f"Лидер внутри Top Stones — {name}: {pct(amount / top_total if top_total else 0)} выручки сегмента.")
        products = product_dataframe(store, "TOP STONES", name)
        if not products.empty:
            product = products.sort_values("Выручка", ascending=False).iloc[0]
            lines.append(
                f"В {name} основную выручку дает группа «{product['Номенклатурная группа']}» — "
                f"{pct(float(product['% выручки камня']))}."
            )
    return lines[:4]


def interactive_explorer(store, all_stores: list, namespace: str = "interactive") -> None:
    st.caption("Выберите сегмент → камень → номенклатурную группу. Данные и диаграммы перестроятся сразу.")

    f1, f2, f3 = st.columns(3)
    with f1:
        selected_segment = st.selectbox(
            "Сегмент",
            SEG_ORDER,
            format_func=lambda s: SEGMENT_LABELS[s],
            key=f"{namespace}_segment_{base_store_name(store.name)}",
        )

    available_stones = [
        stone for stone in STONE_ORDERS[selected_segment]
        if totals_for(store, selected_segment, stone)[0] or totals_for(store, selected_segment, stone)[1]
    ]
    if not available_stones:
        available_stones = STONE_ORDERS[selected_segment]
    with f2:
        selected_stone = st.selectbox(
            "Камень / группа камней",
            available_stones,
            key=f"{namespace}_stone_{base_store_name(store.name)}",
        )

    product_df = product_dataframe(store, selected_segment, selected_stone)
    product_options = ["Все номенклатурные группы"] + product_df["Номенклатурная группа"].drop_duplicates().tolist()
    with f3:
        selected_product = st.selectbox(
            "Номенклатурная группа",
            product_options,
            key=f"{namespace}_product_{base_store_name(store.name)}",
        )

    if product_df.empty:
        st.info("В выбранной группе нет продаж за этот период.")
        return

    stone_qty, stone_sales = totals_for(store, selected_segment, selected_stone)
    if selected_product == "Все номенклатурные группы":
        selected_qty = stone_qty
        selected_sales = stone_sales
        context_note = f"Итого по {selected_stone}"
    else:
        selected_rows = product_df[product_df["Номенклатурная группа"] == selected_product]
        selected_qty = int(selected_rows["Количество"].sum())
        selected_sales = float(selected_rows["Выручка"].sum())
        context_note = f"{selected_stone} → {selected_product}"

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1: kpi_card("Количество", f"{money(selected_qty)} шт.", context_note)
    with k2: kpi_card("Выручка", usd_money(selected_sales), context_note)
    with k3: kpi_card("Средняя стоимость", usd_money(selected_sales / selected_qty if selected_qty else 0))
    with k4: kpi_card("% количества магазина", pct(selected_qty / store.total_qty if store.total_qty else 0))
    with k5: kpi_card("% выручки магазина", pct(selected_sales / store.total_amount if store.total_amount else 0))

    if selected_product == "Все номенклатурные группы":
        left, right = st.columns(2)
        with left:
            locked_plotly_chart(
                horizontal_bar(product_df, "Номенклатурная группа", "Количество", f"{selected_stone}: количество по группам", " шт."),
                width="stretch",
            )
        with right:
            locked_plotly_chart(
                horizontal_bar(product_df, "Номенклатурная группа", "Выручка", f"{selected_stone}: выручка по группам"),
                width="stretch",
            )
        data_table(product_df, key="interactive_product_table")
    else:
        comparison = cross_store_product_dataframe(all_stores, selected_segment, selected_stone, selected_product)
        left, right = st.columns(2)
        with left:
            locked_plotly_chart(
                horizontal_bar(comparison, "Магазин", "Количество", f"{selected_product}: количество по магазинам", " шт."),
                width="stretch",
            )
        with right:
            locked_plotly_chart(
                horizontal_bar(comparison, "Магазин", "Выручка", f"{selected_product}: выручка по магазинам"),
                width="stretch",
            )
        st.markdown("#### Сравнение выбранной группы по сети")
        data_table(comparison, key="interactive_comparison_table")

    insight_panel(
        "Аналитика по выбранным параметрам",
        interactive_conclusions(store, selected_segment, selected_stone, product_df, selected_product, selected_qty, selected_sales),
    )


def store_view(store, all_stores: list) -> None:
    st.markdown(f'<div class="section-title">Магазин {base_store_name(store.name)}</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1: kpi_card("Выручка", usd_money(store.total_amount))
    with c2: kpi_card("Продано изделий", money(store.total_qty) + " шт.")
    with c3: kpi_card("Средняя стоимость", usd_money(store.total_amount / store.total_qty if store.total_qty else 0))
    with c4:
        network_sales = sum(s.total_amount for s in all_stores)
        kpi_card("Доля в выручке сети", pct(store.total_amount / network_sales if network_sales else 0))

    seg = segment_totals(store)
    labels = [SEGMENT_LABELS[s] for s in SEG_ORDER]
    colors = [SEGMENT_COLORS[s] for s in SEG_ORDER]
    a, b = st.columns(2)
    with a:
        locked_plotly_chart(
            donut(labels, [seg[s]["amount"] for s in SEG_ORDER], "Структура продаж", colors, monetary=True),
            width="stretch",
            key=f"store_sales_structure_{base_store_name(store.name)}",
        )
    with b:
        locked_plotly_chart(
            donut(labels, [seg[s]["qty"] for s in SEG_ORDER], "Структура количества", colors),
            width="stretch",
            key=f"store_qty_structure_{base_store_name(store.name)}",
        )

    detail_options = ["Все камни", "Все номенклатурные группы", "Top Stones", "Pearls", "Other Stones"]
    detail_mode = st.segmented_control(
        "Детализация магазина",
        detail_options,
        default="Все камни",
        key="store_detail_mode",
    ) or "Все камни"

    data = stone_dataframe(store)
    if detail_mode == "Все камни":
        data_table(data, key="store_stone_table")
    elif detail_mode == "Все номенклатурные группы":
        data_table(product_dataframe(store), key="store_product_table")
    else:
        segment_lookup = {
            "Top Stones": "TOP STONES",
            "Pearls": "PEARLS",
            "Other Stones": "COLORED STONES",
        }
        seg_code = segment_lookup[detail_mode]
        subset = data[data["Сегмент"] == detail_mode]
        x1, x2 = st.columns(2)
        with x1:
            locked_plotly_chart(
                donut(subset["Камень"].tolist(), subset["Количество"].tolist(), f"{detail_mode}: количество"),
                width="stretch",
                key=f"store_detail_qty_{base_store_name(store.name)}_{seg_code}",
            )
        with x2:
            locked_plotly_chart(
                donut(subset["Камень"].tolist(), subset["Выручка"].tolist(), f"{detail_mode}: выручка", monetary=True),
                width="stretch",
                key=f"store_detail_sales_{base_store_name(store.name)}_{seg_code}",
            )
        st.markdown("#### Номенклатурные группы сегмента")
        data_table(product_dataframe(store, seg_code), key=f"store_segment_table_{seg_code}")

    if base_store_name(store.name) == "OUTLET" and store.extras:
        st.markdown("### Дополнительные подразделения OUTLET")
        cols = st.columns(2)
        for idx, name in enumerate(["GIFT TT", "CAFE"]):
            values = store.extras.get(name, {"qty": 0, "amount": 0})
            avg = values["amount"] / values["qty"] if values["qty"] else 0
            with cols[idx]:
                st.markdown(f"**{name}**")
                a, b, c = st.columns(3)
                with a: kpi_card("Выручка", usd_money(values["amount"]))
                with b: kpi_card("Количество", f"{money(values['qty'])} шт.")
                with c: kpi_card("Средняя стоимость", usd_money(avg))

    insight_panel("Аналитика по магазину", conclusions(store, all_stores))


def is_supplier_report(path: Path) -> bool:
    """Detect the supplier hierarchy export by its first header rows."""
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        ws = wb.active
        header = " ".join(str(ws.cell(r, 1).value or "") for r in range(1, 7)).upper()
        return "ПОСТАВЩИК" in header and "НОМЕНКЛАТУРНАЯ ГРУППА" in header
    finally:
        wb.close()


def normalize_purity_label(value: object) -> str:
    """Return one stable user-facing purity label, including blank values."""
    text = " ".join(str(value or "").strip().split())
    return text if text else "Не указано"


def classify_metal_group(purity: object) -> str:
    """Map a report purity to Silver, Gold/Platinum or Other.

    The grouping intentionally follows the business rule used by comparison:
    every AU or PT variant is precious-gold/platinum, every recognizable
    silver/925 variant is silver, and all remaining or blank values are Other.
    """
    text = normalize_purity_label(purity).upper().replace("Ё", "Е")
    compact = re.sub(r"[^A-ZА-Я0-9]+", "", text)
    if compact in {"НЕУКАЗАНО", "OTHER0", "OTHER", "0"}:
        return "Другое"
    if "AU" in compact or "GOLD" in compact or "ЗОЛОТ" in compact:
        return "Золото и платина"
    if "PT" in compact or "PLATIN" in compact or "ПЛАТИН" in compact:
        return "Золото и платина"
    if (
        "925" in compact
        or compact.startswith("AG")
        or "SILVER" in compact
        or "СЕРЕБ" in compact
    ):
        return "Серебро"
    return "Другое"


def parse_supplier_report_with_period(path: Path) -> tuple[pd.DataFrame, tuple | None]:
    """Parse the current hierarchical sales report in one workbook pass.

    Current format: Store → Stone → Purity → Product group. The header can still
    mention Supplier even when 1C does not render a supplier leaf. Older exports
    without the Purity level remain supported by the legacy supplier branch.
    Returns are ignored because only the Sold columns H/I are read.
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    rows: list[dict] = []
    try:
        ws = wb.active
        period = extract_period(ws)
        hierarchy_header = str(ws.cell(4, 1).value or "")
        header_upper = hierarchy_header.upper()
        has_store_dimension = "МАГАЗИН" in header_upper
        has_purity_dimension = "ПРОБА" in header_upper

        current_store: str | None = None
        current_stone: str | None = None
        current_purity = "Не указано"
        current_product: str | None = None
        skip_store_section = False

        stone_indent = 2 if has_store_dimension else 0
        second_indent = stone_indent + 2
        third_indent = second_indent + 2

        for row in range(7, ws.max_row + 1):
            cell = ws.cell(row, 1)
            text = " ".join(str(cell.value or "").strip().split())
            upper = text.upper()
            indent = int(cell.alignment.indent or 0)

            if text and (upper in {"ИТОГО", "ИТОГО:", "ПОСТАВЩИКИ"} or upper.startswith("ОТЧЕТ")):
                continue

            if has_store_dimension and indent == 0 and cell.font.bold and text:
                normalized = normalize_store_from_report(text)
                current_store = normalized
                skip_store_section = normalized is None
                current_stone = None
                current_purity = "Не указано"
                current_product = None
                continue

            if has_store_dimension and skip_store_section:
                continue

            # Blank hierarchy labels are meaningful in the 1C export. They are
            # normalized instead of skipped so their product rows are retained.
            if indent == stone_indent and (has_store_dimension or not cell.font.bold):
                current_stone = text or "Other"
                current_purity = "Не указано"
                current_product = None
                continue

            if has_purity_dimension:
                if current_stone and indent == second_indent:
                    current_purity = normalize_purity_label(text)
                    current_product = None
                    continue

                if current_stone and indent >= third_indent:
                    product = norm_product(text or "Other")
                    if product.upper() in SKIP_PRODUCTS:
                        continue
                    qty = int(round(float(ws.cell(row, 8).value or 0)))
                    amount = float(ws.cell(row, 9).value or 0)
                    if qty == 0 and amount == 0:
                        continue
                    segment, stone, rule = classify(current_stone)
                    rows.append({
                        "Магазин": current_store if has_store_dimension else "Сеть",
                        "Поставщик": "Other",
                        "Проба": current_purity,
                        "Группа металла": classify_metal_group(current_purity),
                        "Сегмент": SEGMENT_LABELS.get(segment, segment),
                        "Код сегмента": segment,
                        "Камень": stone,
                        "Исходный камень": current_stone,
                        "Номенклатурная группа": PRODUCT_LABELS.get(product, product),
                        "Код группы": product,
                        "Количество": qty,
                        "Выручка": amount,
                        "Правило": rule,
                    })
                    continue

            # Backward-compatible hierarchy without purity:
            # Store → Stone → Product group → Supplier.
            if current_stone and indent == second_indent:
                current_product = norm_product(text or "Other")
                continue

            supplier_indent = third_indent
            is_supplier = (
                current_stone
                and current_product
                and indent >= supplier_indent
                and not cell.font.bold
            )
            if not is_supplier:
                continue

            if current_product.upper() in SKIP_PRODUCTS:
                continue
            qty = int(round(float(ws.cell(row, 8).value or 0)))
            amount = float(ws.cell(row, 9).value or 0)
            if qty == 0 and amount == 0:
                continue
            segment, stone, rule = classify(current_stone)
            supplier_name = text.strip()
            if supplier_name.upper() in {
                "", "СЕТЬ", "NETWORK", "NONE", "NAN", "UNKNOWN",
                "НЕ УКАЗАН", "БЕЗ ПОСТАВЩИКА",
            }:
                supplier_name = "Other"
            rows.append({
                "Магазин": current_store if has_store_dimension else "Сеть",
                "Поставщик": supplier_name,
                "Проба": "Не указано",
                "Группа металла": "Другое",
                "Сегмент": SEGMENT_LABELS.get(segment, segment),
                "Код сегмента": segment,
                "Камень": stone,
                "Исходный камень": current_stone,
                "Номенклатурная группа": PRODUCT_LABELS.get(current_product, current_product),
                "Код группы": current_product,
                "Количество": qty,
                "Выручка": amount,
                "Правило": rule,
            })
    finally:
        wb.close()

    columns = [
        "Магазин", "Поставщик", "Проба", "Группа металла",
        "Сегмент", "Код сегмента", "Камень", "Исходный камень",
        "Номенклатурная группа", "Код группы", "Количество", "Выручка", "Правило",
    ]
    return pd.DataFrame(rows, columns=columns), period


def parse_supplier_report(path: Path) -> pd.DataFrame:
    detail, _ = parse_supplier_report_with_period(path)
    return detail


def supplier_units_from_detail(
    detail: pd.DataFrame,
    period: tuple | None,
    file_name: str,
) -> dict[str, StoreData]:
    """Convert already-parsed supplier rows into StoreData without reopening Excel."""
    if detail.empty:
        return {}

    stores: dict[str, StoreData] = {}
    touched: set[str] = set()
    for row in detail.to_dict("records"):
        store_name = str(row["Магазин"])
        if store_name in {"GIFT TT", "CAFE"}:
            outlet = stores.setdefault("OUTLET", StoreData("OUTLET"))
            outlet.extras[store_name]["qty"] += int(row["Количество"])
            outlet.extras[store_name]["amount"] += float(row["Выручка"])
            touched.add("OUTLET")
            continue
        store = stores.setdefault(store_name, StoreData(store_name))
        touched.add(store_name)
        store.add(
            row["Код сегмента"], row["Камень"], row["Код группы"],
            int(row["Количество"]), float(row["Выручка"]),
            str(row["Исходный камень"]), str(row["Правило"]),
        )
    for name in touched:
        stores[name].add_period(period, file_name)
    return stores


def supplier_report_units(path: Path) -> dict[str, StoreData]:
    detail, period = parse_supplier_report_with_period(path)
    return supplier_units_from_detail(detail, period, path.name)

def filter_metal_groups(detail: pd.DataFrame, selected_groups: Iterable[str]) -> pd.DataFrame:
    """Filter parsed sales rows by the globally selected metal groups."""
    if detail.empty or "Группа металла" not in detail.columns:
        return detail.copy()
    selected = {str(value) for value in selected_groups}
    return detail.loc[detail["Группа металла"].isin(selected)].copy()


def period_tuple_from_stores(stores: list[StoreData]) -> tuple | None:
    periods = [period for store in stores for period in store.periods]
    if not periods:
        return None
    return min(period[0] for period in periods), max(period[1] for period in periods)


def rebuild_filtered_stores(
    detail: pd.DataFrame,
    original_stores: list[StoreData],
    period: tuple | None,
    file_name: str,
) -> list[StoreData]:
    """Rebuild every StoreData object after a global metal filter.

    Empty stores are kept with zero metrics so period-to-period tables preserve
    the full network layout even when a selected metal group has no sales in one
    store or one period.
    """
    rebuilt = supplier_units_from_detail(detail, period, file_name)
    ordered_names: list[str] = []
    for original in original_stores:
        name = base_store_name(original.name)
        if name not in ordered_names:
            ordered_names.append(name)
    for name in rebuilt:
        if name not in ordered_names:
            ordered_names.append(name)
    result: list[StoreData] = []
    for name in ordered_names:
        current = rebuilt.get(name)
        if current is None:
            current = StoreData(name)
            current.add_period(period, file_name)
        result.append(current)
    return result


def detected_filter_state_key(mode: str) -> str:
    return f"global_filter_detected::{mode}"


def selected_metal_groups() -> tuple[str, ...]:
    """Return the one global metal selection shared by every workspace."""
    selected = st.session_state.get("global_metal_groups", list(METAL_GROUPS))
    return tuple(str(value) for value in (selected or []))


def sync_detected_filter_values(mode: str, values: Iterable[object]) -> None:
    """Persist exact purity/material labels and refresh once so the top control shows them."""
    normalized = tuple(sorted({normalize_purity_label(value) for value in values}))
    key = detected_filter_state_key(mode)
    previous = tuple(st.session_state.get(key, ()))
    if normalized != previous:
        st.session_state[key] = normalized
        st.rerun()


def render_metal_filter_control(mode: str) -> tuple[str, ...]:
    """Render the same visible global metal filter in every application mode."""
    st.markdown('<div id="global-metal-filter"></div>', unsafe_allow_html=True)
    detected_purities = st.session_state.get(detected_filter_state_key(mode), ())
    with st.container(key="global_metal_filter"):
        st.markdown(
            '<div class="global-metal-filter-note">'
            '<b>Металл и пробы</b>'
            '<span>Фильтр общий для выбранного раздела. Отключенная группа полностью исключается из KPI, количества, суммы, таблиц и диаграмм, если в источнике есть поле пробы или материала.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        selected = st.pills(
            "Группы металла",
            list(METAL_GROUPS),
            selection_mode="multi",
            default=list(METAL_GROUPS),
            key="global_metal_groups",
            help="Серебро — варианты 925/Ag; золото и платина — AU и PT; другое — Other 0, пустая проба, латунь, сталь и остальные материалы.",
            width="stretch",
            label_visibility="collapsed",
        )
        selected_tuple = tuple(str(value) for value in (selected or []))
        if selected_tuple:
            st.markdown(
                '<div class="global-metal-filter-active"><b>В отчете:</b> '
                + escape(", ".join(selected_tuple))
                + '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.warning("Выберите хотя бы одну группу металла.")

        purity_values = sorted({normalize_purity_label(value) for value in detected_purities})
        if purity_values:
            chips = "".join(
                f'<span class="detected-purity-chip">{escape(value)}</span>'
                for value in purity_values
            )
            st.markdown(
                '<div class="detected-purities"><b>Распознано в текущих данных:</b><br>'
                + chips
                + '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div class="detected-purities">После подключения данных здесь появятся конкретные пробы или материалы из источника.</div>',
                unsafe_allow_html=True,
            )
        return selected_tuple



def supplier_has_meaningful_detail(df: pd.DataFrame) -> bool:
    """Return True only when the report contains real supplier differentiation."""
    if df.empty or "Поставщик" not in df.columns:
        return False
    summary = supplier_summary(df)
    if summary.empty:
        return False
    names = summary["Поставщик"].astype(str).str.strip()
    real = summary.loc[names.str.casefold() != "other"]
    return not real.empty

def supplier_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Поставщик", "Количество", "Выручка", "Средняя стоимость", "% количества", "% выручки"])
    df = df.copy()
    df["Поставщик"] = df["Поставщик"].fillna("Other").astype(str).str.strip()
    df.loc[df["Поставщик"].str.upper().isin({"", "СЕТЬ", "NETWORK", "NONE", "NAN", "UNKNOWN", "НЕ УКАЗАН", "БЕЗ ПОСТАВЩИКА"}), "Поставщик"] = "Other"
    result = df.groupby("Поставщик", as_index=False).agg(
        Количество=("Количество", "sum"),
        Выручка=("Выручка", "sum"),
    )
    total_qty = float(result["Количество"].sum())
    total_sales = float(result["Выручка"].sum())
    result["Средняя стоимость"] = result["Выручка"] / result["Количество"].replace(0, pd.NA)
    result["Средняя стоимость"] = result["Средняя стоимость"].fillna(0)
    result["% количества"] = result["Количество"] / total_qty if total_qty else 0
    result["% выручки"] = result["Выручка"] / total_sales if total_sales else 0
    return result.sort_values("Выручка", ascending=False)


SUPPLIER_PIE_MIN_SHARE = 0.045


def supplier_pie_data(summary: pd.DataFrame, share_col: str) -> tuple[list[str], list[float]]:
    """Collapse suppliers below 4.5% into Other for supplier pie charts only.

    The detailed table and horizontal charts keep the original suppliers unchanged.
    Each pie is grouped independently by its own metric (revenue or quantity).
    """
    if summary.empty or share_col not in summary.columns:
        return [], []

    labels: list[str] = []
    values: list[float] = []
    other_value = 0.0

    for _, row in summary.iterrows():
        label = str(row["Поставщик"]).strip() or "Other"
        value = float(row[share_col])
        if label.casefold() == "other" or value < SUPPLIER_PIE_MIN_SHARE:
            other_value += value
        else:
            labels.append(label)
            values.append(value)

    if other_value > 0:
        labels.append("Other")
        values.append(other_value)

    return labels, values


def supplier_view(df: pd.DataFrame) -> None:
    st.caption("Общая аналитика по сети из выгрузки «Камень → Номенклатурная группа → Поставщик».")
    if df.empty:
        st.info("Загрузите выгрузку с поставщиками на странице «Главная».")
        return

    summary = supplier_summary(df)
    total_qty = int(df["Количество"].sum())
    total_sales = float(df["Выручка"].sum())
    c1, c2, c3, c4 = st.columns(4)
    with c1: kpi_card("Поставщиков", str(summary["Поставщик"].nunique()))
    with c2: kpi_card("Продано изделий", f"{money(total_qty)} шт.")
    with c3: kpi_card("Выручка", usd_money(total_sales))
    with c4: kpi_card("Средняя стоимость", usd_money(total_sales / total_qty if total_qty else 0))

    revenue_labels, revenue_values = supplier_pie_data(summary, "% выручки")
    quantity_labels, quantity_values = supplier_pie_data(summary, "% количества")

    left, right = st.columns(2)
    with left:
        locked_plotly_chart(donut(revenue_labels, revenue_values, "Доля поставщиков по выручке"), width="stretch")
    with right:
        locked_plotly_chart(donut(quantity_labels, quantity_values, "Доля поставщиков по количеству"), width="stretch")

    left2, right2 = st.columns(2)
    with left2:
        locked_plotly_chart(horizontal_bar(summary.head(15), "Поставщик", "Выручка", "Топ поставщиков по выручке"), width="stretch")
    with right2:
        locked_plotly_chart(horizontal_bar(summary.head(15), "Поставщик", "Количество", "Топ поставщиков по количеству", " шт."), width="stretch")

    with st.expander("Полная таблица поставщиков", expanded=False):
        data_table(summary, key="supplier_summary_table")

    supplier_names = summary["Поставщик"].tolist()
    selected = st.selectbox("Выберите поставщика", supplier_names, key="supplier_selected")
    detail = df[df["Поставщик"] == selected].copy()
    selected_qty = int(detail["Количество"].sum())
    selected_sales = float(detail["Выручка"].sum())
    a, b, c, d = st.columns(4)
    with a: kpi_card("Поставщик", selected)
    with b: kpi_card("Количество", f"{money(selected_qty)} шт.")
    with c: kpi_card("Выручка", usd_money(selected_sales))
    with d: kpi_card("Средняя стоимость", usd_money(selected_sales / selected_qty if selected_qty else 0))

    by_segment = detail.groupby("Сегмент", as_index=False).agg(
        Количество=("Количество", "sum"), Выручка=("Выручка", "sum")
    ).sort_values("Выручка", ascending=False)
    by_product = detail.groupby("Номенклатурная группа", as_index=False).agg(
        Количество=("Количество", "sum"), Выручка=("Выручка", "sum")
    ).sort_values("Выручка", ascending=False)
    by_stone = detail.groupby(["Сегмент", "Камень"], as_index=False).agg(
        Количество=("Количество", "sum"), Выручка=("Выручка", "sum")
    ).sort_values("Выручка", ascending=False)
    by_store = detail.groupby("Магазин", as_index=False).agg(
        Количество=("Количество", "sum"), Выручка=("Выручка", "sum")
    ).sort_values("Выручка", ascending=False) if "Магазин" in detail.columns else pd.DataFrame()

    if not by_store.empty and by_store["Магазин"].nunique() > 1:
        st.markdown("#### По магазинам")
        locked_plotly_chart(horizontal_bar(by_store, "Магазин", "Выручка", f"{selected}: выручка по магазинам"), width="stretch")
        data_table(by_store, key="supplier_store_table")

    seg_l, seg_r = st.columns(2)
    with seg_l:
        locked_plotly_chart(donut(by_segment["Сегмент"].tolist(), by_segment["Выручка"].tolist(), f"{selected}: сегменты по выручке", monetary=True), width="stretch")
    with seg_r:
        locked_plotly_chart(donut(by_segment["Сегмент"].tolist(), by_segment["Количество"].tolist(), f"{selected}: сегменты по количеству"), width="stretch")

    l, r = st.columns(2)
    with l:
        locked_plotly_chart(horizontal_bar(by_product, "Номенклатурная группа", "Выручка", f"{selected}: номенклатурные группы"), width="stretch")
    with r:
        locked_plotly_chart(horizontal_bar(by_stone.head(20), "Камень", "Выручка", f"{selected}: камни"), width="stretch")

    table_mode = st.segmented_control(
        "Таблица детализации поставщика",
        ["Сегменты", "Номенклатурные группы", "Камни", "Полная детализация"],
        default="Сегменты",
        key="supplier_table_mode",
    ) or "Сегменты"
    if table_mode == "Сегменты":
        table_df = by_segment
    elif table_mode == "Номенклатурные группы":
        table_df = by_product
    elif table_mode == "Камни":
        table_df = by_stone
    else:
        table_df = detail
    data_table(table_df, key="supplier_detail_table")

    if "Магазин" in df.columns and df["Магазин"].nunique() > 1:
        st.caption("Доступен полный разрез: поставщик × магазин × камень × номенклатурная группа.")

    insight_panel("Аналитика по поставщикам", supplier_conclusions(df, summary))


def _merge_units(target: dict[str, StoreData], incoming: dict[str, StoreData]) -> None:
    for name, source in incoming.items():
        dest = target.setdefault(name, StoreData(name))
        dest.periods.extend(source.periods)
        dest.files.extend(source.files)
        for (segment, stone), products in source.data.items():
            for product, vals in products.items():
                dest.add(segment, stone, product, int(vals.get("qty", 0)), float(vals.get("amount", 0)), stone, "merged")
        for extra_name, vals in source.extras.items():
            dest.extras[extra_name]["qty"] += int(vals.get("qty", 0))
            dest.extras[extra_name]["amount"] += float(vals.get("amount", 0))


def parse_uploads(uploaded_files):
    """Parse all uploaded workbooks once and return stores, errors and supplier detail."""
    errors: list[tuple[str, str]] = []
    supplier_frames: list[pd.DataFrame] = []
    stores: dict[str, StoreData] = {}

    with tempfile.TemporaryDirectory(prefix="analitika_parse_") as temp_dir:
        normal_paths: list[Path] = []
        supplier_paths: list[Path] = []

        for uploaded in uploaded_files:
            original_name = Path(str(uploaded.name)).name or "report.xlsx"
            suffix = Path(original_name).suffix.lower() or ".xlsx"
            digest = hashlib.sha256(uploaded.getvalue()).hexdigest()[:16]
            path = Path(temp_dir) / f"{digest}{suffix}"
            path.write_bytes(uploaded.getvalue())
            try:
                if is_supplier_report(path):
                    supplier_paths.append(path)
                else:
                    normal_paths.append(path)
            except Exception as exc:
                errors.append((uploaded.name, str(exc)))

        if normal_paths:
            normal_stores, normal_errors = build_report_units(normal_paths)
            _merge_units(stores, normal_stores)
            errors.extend(normal_errors)

        for path in supplier_paths:
            try:
                detail, period = parse_supplier_report_with_period(path)
                if not detail.empty:
                    supplier_frames.append(detail)
                    _merge_units(stores, supplier_units_from_detail(detail, period, path.name))
            except Exception as exc:
                errors.append((path.name, str(exc)))

    if supplier_frames:
        supplier_df = pd.concat(supplier_frames, ignore_index=True, copy=False)
        supplier_df["Поставщик"] = supplier_df["Поставщик"].fillna("Other").astype(str).str.strip()
        service_values = {"", "СЕТЬ", "NETWORK", "NONE", "NAN", "UNKNOWN", "НЕ УКАЗАН", "БЕЗ ПОСТАВЩИКА"}
        supplier_df.loc[supplier_df["Поставщик"].str.upper().isin(service_values), "Поставщик"] = "Other"
    else:
        supplier_df = pd.DataFrame()

    return stores, errors, supplier_df


def cache_payloads(uploaded_files: list[StoredUpload]) -> tuple[tuple[str, bytes], ...]:
    """Immutable payload accepted by the shared report cache."""
    return tuple((item.name, item.getvalue()) for item in uploaded_files)


@st.cache_resource(show_spinner=False)
def excel_parse_lock() -> threading.Lock:
    """One process-wide lock for memory-intensive Excel parsing.

    The lock itself is a thread-safe global resource. Parsed business data is
    never stored here; it remains isolated in each user session.
    """
    return threading.Lock()


@st.cache_resource(scope="session", ttl=900, max_entries=3, show_spinner=False)
def parse_report_bundle(payloads: tuple[tuple[str, bytes], ...]):
    """Reuse parsed data only inside the current browser session.

    The bundle contains mutable StoreData objects and pandas DataFrames. Keeping
    them in a global resource cache makes the same instances available to
    different users and Streamlit threads. A session-scoped cache prevents
    cross-user access while still avoiding repeated Excel parsing on reruns.
    """
    uploads = [StoredUpload(name, data) for name, data in payloads]
    # Excel parsing and workbook normalization are the heaviest operations in
    # the app. Community Cloud runs sessions in parallel threads, so serialize
    # this short stage to prevent two users from doubling the native-memory
    # peak at the same moment.
    with excel_parse_lock():
        return parse_uploads(uploads)


def guide_sections(path: Path) -> dict[str, str]:
    try:
        text = path.read_text(encoding="utf-8")
    except OSError:
        return {"Руководство недоступно": "Файл USER_GUIDE.md не найден."}
    parts = re.split(r"(?m)^##\s+", text)
    sections: dict[str, str] = {}
    intro = parts[0].strip()
    if intro:
        intro = re.sub(r"(?m)^#\s+.*$", "", intro, count=1).strip()
        sections["Быстрый старт"] = intro
    for part in parts[1:]:
        lines = part.strip().splitlines()
        if not lines:
            continue
        sections[lines[0].strip()] = "\n".join(lines[1:]).strip()
    return sections or {"Руководство": text}


def latest_updates_html(changelog_path: Path, limit: int = 3) -> str:
    try:
        text = changelog_path.read_text(encoding="utf-8")
    except OSError:
        return '<div class="about-step">История обновлений недоступна.</div>'
    sections = re.split(r"(?m)^##\s+", text)[1:limit + 1]
    cards: list[str] = []
    for section in sections:
        lines = [line.strip() for line in section.strip().splitlines()]
        if not lines:
            continue
        bullets = [re.sub(r"^-\s*", "", line) for line in lines[1:] if line.startswith("-")]
        cards.append(
            '<div class="about-step">'
            f'<b>{escape(lines[0])}</b><br>'
            f'{escape(" ".join(bullets))}'
            '</div>'
        )
    return "".join(cards)


def render_about() -> None:
    """Open the product overview directly on the five business capabilities."""
    st.markdown('<div id="about"></div>', unsafe_allow_html=True)
    st.markdown("## Возможности")
    st.markdown(
        f'<div class="about-grid about-grid-compact">{feature_cards_html()}</div>',
        unsafe_allow_html=True,
    )
    st.markdown("### Для руководителя")
    st.markdown(
        "Система показывает результаты сети, причины изменений, потребность в пополнении "
        "и состояние складских операций. Каждый показатель можно раскрыть до магазина, "
        "категории, камня, поставщика или SKU."
    )


def render_user_guide() -> None:
    path = Path(__file__).with_name("USER_GUIDE.md")
    sections = guide_sections(path)
    if st.session_state.get("user_guide_chapter") not in sections:
        st.session_state["user_guide_chapter"] = next(iter(sections))
    selected = st.selectbox("Глава руководства", list(sections), key="user_guide_chapter")
    st.markdown(f"## {selected}")
    st.markdown(sections[selected])
    pdf_path = Path(__file__).with_name("Analitika_USER_GUIDE.pdf")
    try:
        payload = pdf_path.read_bytes()
    except OSError:
        payload = b""
    if payload:
        st.download_button(
            "Скачать красиво оформленное руководство в PDF",
            data=payload,
            file_name="Analitika_USER_GUIDE.pdf",
            mime="application/pdf",
            type="primary",
            width="stretch",
        )
    else:
        st.warning("PDF-версия руководства временно недоступна.")


def render_release_history() -> None:
    st.markdown("## История обновлений")
    st.caption("Список формируется автоматически из CHANGELOG.md.")
    st.markdown(
        '<div class="updates-scroll updates-scroll-standalone" tabindex="0">'
        + release_history_html(Path(__file__).with_name("CHANGELOG.md"))
        + '</div>',
        unsafe_allow_html=True,
    )


def render_about_mode() -> None:
    options = ("О программе", "Руководство", "История обновлений")
    if st.session_state.get("about_workspace") not in options:
        st.session_state["about_workspace"] = options[0]
    selected = st.segmented_control(
        "Документация",
        options,
        key="about_workspace",
    ) or st.session_state.get("about_workspace", options[0])
    if selected == "Руководство":
        render_user_guide()
    elif selected == "История обновлений":
        render_release_history()
    else:
        render_about()
    st.caption(f"Analitika Web {APP_VERSION} · Princess Jewelry · Developed by Vladimir Panasyan")


@st.fragment
def render_store_fragment(stores: list[StoreData]) -> None:
    """Rerun only the store block when its controls change."""
    try:
        store_names = [base_store_name(store.name) for store in stores]
        chosen = st.selectbox("Выберите магазин", store_names, index=0, key="store_page_select")
        chosen_store = next(store for store in stores if base_store_name(store.name) == chosen)
        store_view(chosen_store, stores)
    finally:
        gc.collect()


@st.fragment
def render_interactive_fragment(stores: list[StoreData]) -> None:
    """Isolate interactive filters from the rest of the dashboard."""
    try:
        store_names = [base_store_name(store.name) for store in stores]
        chosen_interactive = st.selectbox(
            "Магазин для интерактивного анализа",
            store_names,
            index=0,
            key="interactive_store_select",
        )
        interactive_store = next(
            store for store in stores if base_store_name(store.name) == chosen_interactive
        )
        interactive_explorer(interactive_store, stores, namespace="main_interactive")
    finally:
        gc.collect()


@st.fragment
def render_supplier_fragment(supplier_df: pd.DataFrame) -> None:
    """Rerun supplier controls without rebuilding summary and store charts."""
    try:
        if not supplier_has_meaningful_detail(supplier_df):
            st.info(
                "В текущей выгрузке нет полноценной детализации по поставщикам. "
                "Раздел автоматически заполнится в отчете, где поставщики указаны отдельными значениями."
            )
        else:
            supplier_view(supplier_df)
    finally:
        gc.collect()


MODE_GUIDE_CHAPTERS = {
    "Обычный отчет": "1. Общий анализ продаж",
    "Сравнение периодов": "2. Сравнение периодов",
    "Сувениры и касты на складе": "6. Склад Baserow",
    "Заказ Sonu": "5. Заказ Sonu",
    "Заказ поставщику": "4. Заказ поставщику",
}


def _open_user_guide(chapter: str | None = None) -> None:
    st.session_state["report_mode"] = "О программе"
    st.session_state["about_workspace"] = "Руководство"
    if chapter:
        st.session_state["user_guide_chapter"] = chapter


def render_report_settings(mode: str) -> None:
    """Keep the shared exchange-rate and purity controls in one compact block."""
    if mode not in {
        "Обычный отчет",
        "Сравнение периодов",
        "Сувениры и касты на складе",
        "Заказ Sonu",
        "Заказ поставщику",
    }:
        return
    with st.expander("⚙️ Курс и пробы", expanded=False):
        if mode == "Сравнение периодов":
            st.caption(
                "Один набор проб и один курс применяются симметрично к обоим периодам. "
                "После изменения сравнительные показатели пересчитываются автоматически."
            )
        elif mode == "Заказ поставщику":
            st.caption(
                "Заказ рассчитывается в штуках, поэтому курс не изменяет рекомендации. "
                "Выбор проб сохраняется единым для сайта и применяется там, где входной файл содержит поле «Проба»."
            )
        elif mode == "Сувениры и касты на складе":
            st.caption(
                "Пробы и группы металла ограничивают складские позиции. Курс хранится единым для сайта; "
                "остатки в штуках от него не зависят."
            )
        else:
            st.caption(
                "Пробы ограничивают данные выбранного раздела, а курс используется для денежных показателей в USD. "
                "После изменения доступные расчёты обновляются автоматически."
            )
        render_metal_filter_control(mode)
        st.divider()
        render_global_fx_control()


def render_mode_help_page(mode: str) -> None:
    """Show the full logic of the active module from the shared user guide."""
    chapter = MODE_GUIDE_CHAPTERS.get(mode)
    if not chapter:
        st.info("Для этого раздела отдельная инструкция не требуется.")
        return
    sections = guide_sections(Path(__file__).with_name("USER_GUIDE.md"))
    body = sections.get(chapter)
    st.markdown(f"## Как работать: {mode}")
    if body:
        st.markdown(body)
    else:
        st.warning("Глава руководства временно недоступна.")
    st.button(
        "Открыть полное руководство",
        key=f"open_full_guide::{mode}",
        on_click=_open_user_guide,
        args=(chapter,),
        width="stretch",
    )


def render_mode_workspace_tab(mode: str) -> bool:
    """Return True for the working tab and render instructions otherwise."""
    if mode == "О программе":
        return True
    options = ("Работа", "Как с этим работать")
    key = "mode_workspace_view::" + hashlib.sha1(mode.encode("utf-8")).hexdigest()[:10]
    if st.session_state.get(key) not in options:
        st.session_state[key] = options[0]
    selected = st.segmented_control(
        "Вкладка раздела",
        options,
        key=key,
        width="stretch",
        label_visibility="collapsed",
    ) or options[0]
    if selected == "Как с этим работать":
        render_mode_help_page(mode)
        return False
    return True


def render_standard_overview(
    stores: list[StoreData],
    summary_df: pd.DataFrame,
) -> None:
    """Compressed network overview; detailed tables stay collapsed."""
    store_summary = executive_store_summary(stores)
    segment_summary = network_segment_summary(stores)

    left, right = st.columns(2)
    with left:
        locked_plotly_chart(
            horizontal_bar(
                store_summary.head(10).sort_values("Выручка", ascending=True),
                "Магазин",
                "Выручка",
                "Топ-10 магазинов по выручке",
            ),
            width="stretch",
            key="standard_overview_store_revenue",
        )
    with right:
        locked_plotly_chart(
            donut(
                segment_summary["Сегмент"].tolist(),
                segment_summary["Выручка"].tolist(),
                "Взвешенная структура выручки сети",
                [SEGMENT_COLORS[segment] for segment in SEG_ORDER],
                monetary=True,
            ),
            width="stretch",
            key="standard_overview_segment_mix",
        )

    st.markdown("#### Структура сети по сегментам")
    st.caption("Доли рассчитаны от общей выручки и общего количества сети, а не как среднее процентов магазинов.")
    data_table(segment_summary, key="standard_weighted_segment_table")

    with st.expander("Полная таблица по магазинам", expanded=False):
        data_table(summary_df, key="standard_network_summary_full")


def render_stones_workspace(stores: list[StoreData]) -> None:
    """Network stone and product-group workspace with compact defaults."""
    facts = stores_fact_dataframe(stores)
    if facts.empty:
        st.info("В отчете нет данных по камням и номенклатурным группам.")
        return

    segment_options = [SEGMENT_LABELS[segment] for segment in SEG_ORDER]
    selected_segment = st.segmented_control(
        "Сегмент",
        segment_options,
        default=segment_options[0],
        key="standard_stones_segment",
    ) or segment_options[0]
    scoped = facts.loc[facts["Сегмент"] == selected_segment].copy()
    stone_summary = aggregate_metrics(scoped, ["Камень"]).sort_values("Выручка", ascending=False)
    if stone_summary.empty:
        st.info("В выбранном сегменте нет продаж.")
        return

    top = stone_summary.head(12)
    left, right = st.columns(2)
    with left:
        locked_plotly_chart(
            horizontal_bar(
                top.sort_values("Выручка", ascending=True),
                "Камень",
                "Выручка",
                "Камни по выручке",
            ),
            width="stretch",
            key=f"standard_stone_revenue_{selected_segment}",
        )
    with right:
        locked_plotly_chart(
            horizontal_bar(
                top.sort_values("Количество", ascending=True),
                "Камень",
                "Количество",
                "Камни по количеству",
                " шт.",
            ),
            width="stretch",
            key=f"standard_stone_qty_{selected_segment}",
        )

    selected_stone = st.selectbox(
        "Камень / группа камней",
        stone_summary["Камень"].astype(str).tolist(),
        key="standard_stone_select",
    )
    product_summary = aggregate_metrics(
        scoped.loc[scoped["Камень"].astype(str) == str(selected_stone)],
        ["Номенклатурная группа"],
    ).sort_values("Выручка", ascending=False)
    st.markdown(f"#### {selected_stone}: номенклатурные группы")
    data_table(product_summary, key="standard_stone_product_table")

    with st.expander("Полная таблица камней выбранного сегмента", expanded=False):
        data_table(stone_summary, key="standard_stone_summary_full")


def render_standard_workspace(
    stores: list[StoreData],
    summary_df: pd.DataFrame,
    supplier_df: pd.DataFrame,
) -> None:
    st.markdown('<div id="workspace"></div>', unsafe_allow_html=True)
    section_divider(
        "Рабочее пространство",
        "Открывается только выбранный раздел — меньше прокрутки и быстрее работа на iPad и смартфоне.",
        "АНАЛИЗ ПРОДАЖ",
    )
    options = (
        "Обзор",
        "Магазины",
        "Камни и группы",
        "Поставщики",
        "Исследование данных",
    )
    selected = st.segmented_control(
        "Раздел анализа",
        options,
        default=options[0],
        key="standard_workspace",
    ) or options[0]

    if selected == "Обзор":
        render_standard_overview(stores, summary_df)
    elif selected == "Магазины":
        render_store_fragment(stores)
    elif selected == "Камни и группы":
        render_stones_workspace(stores)
    elif selected == "Поставщики":
        render_supplier_fragment(supplier_df)
    else:
        render_interactive_fragment(stores)

HERO_CONTENT = {
    "Обычный отчет": {
        "eyebrow": "АНАЛИТИКА ПРОДАЖ",
        "title": "Общий анализ продаж",
        "copy": "Сводный обзор продаж по сети: результаты магазинов, структура ассортимента, ключевые категории и точки роста.",
        "badges": ("Сеть и магазины", "Ассортимент", "Структура продаж"),
    },
    "Сравнение периодов": {
        "eyebrow": "ДИНАМИКА БИЗНЕСА",
        "title": "Сравнение периодов",
        "copy": "Сопоставление двух периодов по ключевым показателям, магазинам и ассортименту с выделением основных причин роста и снижения.",
        "badges": ("Динамика продаж", "Драйверы изменений", "Сравнение структуры"),
    },
    "Сувениры и касты на складе": {
        "eyebrow": "СКЛАДСКИЕ ОПЕРАЦИИ",
        "title": "Склад Baserow",
        "copy": "Единое рабочее пространство для контроля остатков, приёмки товаров, передачи в бухгалтерию и обработки поставок.",
        "badges": ("Остатки", "Приёмка", "Складские операции"),
    },
    "Заказ Sonu": {
        "eyebrow": "УПРАВЛЕНИЕ АССОРТИМЕНТОМ",
        "title": "Заказ Sonu",
        "copy": "Анализ продаж ассортимента Sonu и подготовка рекомендаций по пополнению с разбивкой по изделиям, камням и типам браслетов.",
        "badges": ("Ассортимент Sonu", "Продажи и остатки", "Рекомендации"),
    },
    "Заказ поставщику": {
        "eyebrow": "ПЛАНИРОВАНИЕ ПОСТАВОК",
        "title": "Заказ поставщику",
        "copy": "Расчёт потребности в товарах с учётом продаж, остатков, приоритетных магазинов и правил ассортиментного пополнения.",
        "badges": ("Рекомендации", "Остатки и продажи", "Формирование заказа"),
    },
    "О программе": {
        "eyebrow": "PRINCESS JEWELRY",
        "title": "Analitika",
        "copy": "Единая система для анализа продаж, управления заказами поставщикам и контроля складских операций Princess Jewelry.",
        "badges": ("Возможности", "Руководство", "История обновлений"),
    },
}


def render_hero(mode: str) -> None:
    """Render a practical module-specific header instead of one generic slogan."""
    content = HERO_CONTENT.get(mode, HERO_CONTENT["Обычный отчет"])
    badges = "".join(
        f'<span class="luxury-badge">{escape(str(label))}</span>'
        for label in content["badges"]
    )
    st.markdown('<div id="upload"></div>', unsafe_allow_html=True)
    st.markdown(
        '<section class="luxury-hero">'
        '<div class="luxury-hero-content">'
        f'<div class="luxury-eyebrow">{escape(str(content["eyebrow"]))}</div>'
        f'<div class="luxury-title">{escape(str(content["title"]))}</div>'
        '<div class="luxury-divider"></div>'
        f'<div class="luxury-copy">{escape(str(content["copy"]))}</div>'
        f'<div class="luxury-badges">{badges}</div>'
        '</div></section>',
        unsafe_allow_html=True,
    )


def render_report_context(title: str, details: str, *, tone: str = "ready") -> None:
    """Compact identity bar shown after a file or data source is ready."""
    st.markdown(
        f'<div class="report-context report-context-{escape(tone)}">'
        '<div class="report-context-dot"></div>'
        '<div class="report-context-copy">'
        f'<strong>{escape(title)}</strong><span>{escape(details)}</span>'
        '</div></div>',
        unsafe_allow_html=True,
    )




def render_standard_report_mode() -> None:
    uploaded_files = st.file_uploader(
        "Загрузите общую выгрузку Excel",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="upload_widget",
    )
    persist_uploads(uploaded_files)
    active_files = saved_uploads()

    if not active_files:
        # Требования к файлу и логика раздела находятся во вкладке «Как с этим работать».
        st.stop()

    file_names = ", ".join(item.name for item in active_files)
    loaded_col, action_col = st.columns([3, 1], vertical_alignment="center")
    with loaded_col:
        render_report_context("Отчёт готов к анализу", file_names)
    with action_col:
        if st.button("Загрузить другой отчёт", key="replace_report_inline", width="stretch"):
            clear_saved_uploads()
            st.rerun()

    with st.spinner("Обрабатываем отчет..."):
        stores_dict, errors, supplier_df = parse_report_bundle(cache_payloads(active_files))

    if errors:
        st.warning("Некоторые файлы не удалось обработать:\n" + "\n".join(f"• {name}: {error}" for name, error in errors))
    stores = list(stores_dict.values())

    has_metal_data = not supplier_df.empty and {"Проба", "Группа металла"}.issubset(supplier_df.columns)
    if has_metal_data:
        sync_detected_filter_values("Обычный отчет", supplier_df["Проба"].tolist())
        selected_metals = selected_metal_groups()
        if not selected_metals:
            st.error("Оставьте включенной хотя бы одну группу металла.")
            st.stop()
        supplier_df = filter_metal_groups(supplier_df, selected_metals)
        stores = rebuild_filtered_stores(
            supplier_df, stores, period_tuple_from_stores(stores), file_names
        )
    else:
        st.warning(
            "В загруженном файле нет уровня «Проба». Фильтр показан, но для пересчета обычного отчета нужен новый единый формат."
        )

    summary_df = network_summary(stores)
    if summary_df.empty or "Количество" not in summary_df.columns:
        st.error("В файле не найдены строки продаж. Проверьте структуру выгрузки.")
        st.stop()

    st.markdown('<div id="executive"></div>', unsafe_allow_html=True)
    render_executive_brief(stores, summary_df, supplier_df)
    render_standard_workspace(stores, summary_df, supplier_df)


def render_comparison_mode() -> None:
    ready = bool(st.session_state.get("comparison_ready"))

    if not ready:
        with st.form("comparison_upload_form", clear_on_submit=False):
            left, right = st.columns(2)
            with left:
                st.markdown('<div class="compare-period-title">Период 1</div>', unsafe_allow_html=True)
                first_file = st.file_uploader(
                    "Первый отчет Excel",
                    type=["xlsx", "xlsm"],
                    accept_multiple_files=False,
                    key="comparison_upload_1",
                )
            with right:
                st.markdown('<div class="compare-period-title">Период 2</div>', unsafe_allow_html=True)
                second_file = st.file_uploader(
                    "Второй отчет Excel",
                    type=["xlsx", "xlsm"],
                    accept_multiple_files=False,
                    key="comparison_upload_2",
                )

            submitted = st.form_submit_button(
                "Запустить сравнительный анализ",
                type="primary",
                width="stretch",
            )

        if submitted:
            if first_file is None or second_file is None:
                st.error("Загрузите оба отчета перед запуском сравнительного анализа.")
            elif st.session_state.get("comparison_processing"):
                st.info("Сравнение уже запускается. Подождите несколько секунд.")
            else:
                st.session_state["comparison_processing"] = True
                try:
                    persist_comparison_upload(1, first_file)
                    persist_comparison_upload(2, second_file)
                    st.session_state["comparison_ready"] = True
                finally:
                    st.session_state["comparison_processing"] = False
                st.rerun()

        st.stop()

    first_saved = saved_comparison_upload(1)
    second_saved = saved_comparison_upload(2)
    both_loaded = first_saved is not None and second_saved is not None

    if not both_loaded:
        clear_comparison_uploads()
        st.warning("Файлы сравнения не сохранились. Загрузите оба отчета повторно.")
        st.rerun()

    render_report_context(
        "Два периода готовы к сравнению",
        f"{first_saved.name}  ↔  {second_saved.name}",
    )
    if st.button("Загрузить другие периоды", key="replace_comparison_inline", width="stretch"):
        clear_comparison_uploads()
        st.rerun()

    with st.spinner("Сопоставляем два периода..."):
        first_stores_dict, first_errors, first_supplier_df = parse_report_bundle(single_upload_payload(first_saved))
        second_stores_dict, second_errors, second_supplier_df = parse_report_bundle(single_upload_payload(second_saved))

    errors = [("Период 1 · " + name, error) for name, error in first_errors] + [
        ("Период 2 · " + name, error) for name, error in second_errors
    ]
    if errors:
        st.warning("Некоторые данные не удалось обработать:\n" + "\n".join(f"• {name}: {error}" for name, error in errors))

    stores_first = list(first_stores_dict.values())
    stores_second = list(second_stores_dict.values())
    if not stores_first or not stores_second:
        st.error("Один из отчетов не содержит распознаваемых строк продаж.")
        st.stop()

    first_label, first_start, first_end = comparison_period_info(stores_first)
    second_label, second_start, second_end = comparison_period_info(stores_second)
    if first_start is not None and second_start is not None and first_start == second_start and first_end == second_end:
        st.error("В двух файлах указан одинаковый период. Для сравнения нужны отчеты за разные периоды.")
        st.stop()

    if first_start is not None and second_start is not None and first_start > second_start:
        stores_first, stores_second = stores_second, stores_first
        first_supplier_df, second_supplier_df = second_supplier_df, first_supplier_df
        first_label, second_label = second_label, first_label
        first_start, second_start = second_start, first_start
        first_end, second_end = second_end, first_end

    has_metal_data = all(
        not detail.empty and {"Проба", "Группа металла"}.issubset(detail.columns)
        for detail in (first_supplier_df, second_supplier_df)
    )
    if has_metal_data:
        sync_detected_filter_values(
            "Сравнение периодов",
            pd.concat([first_supplier_df["Проба"], second_supplier_df["Проба"]], ignore_index=True).tolist(),
        )
        selected_metals = selected_metal_groups()
        if not selected_metals:
            st.error("Оставьте включенной хотя бы одну группу металла.")
            st.stop()
        first_supplier_df = filter_metal_groups(first_supplier_df, selected_metals)
        second_supplier_df = filter_metal_groups(second_supplier_df, selected_metals)
        stores_first = rebuild_filtered_stores(
            first_supplier_df,
            stores_first,
            period_tuple_from_stores(stores_first),
            first_label,
        )
        stores_second = rebuild_filtered_stores(
            second_supplier_df,
            stores_second,
            period_tuple_from_stores(stores_second),
            second_label,
        )
    else:
        st.warning(
            "В одном из файлов нет уровня «Проба». Анализ построен по всем данным без фильтра металла. "
            "Для полноценного сравнения используйте новый единый формат отчета."
        )

    render_comparison_report(
        stores_first,
        stores_second,
        first_supplier_df,
        second_supplier_df,
        first_label,
        second_label,
        first_start,
        first_end,
        second_start,
        second_end,
    )


def render_warehouse_mode() -> None:
    render_warehouse_dashboard(selected_metal_groups())



def render_sonu_mode() -> None:
    render_sonu_order_dashboard(selected_metal_groups())


def render_supplier_order_mode() -> None:
    render_supplier_order_dashboard()

def main() -> None:
    if st.session_state.get("report_mode") not in REPORT_MODES:
        st.session_state["report_mode"] = "Обычный отчет"
    active_mode = str(st.session_state["report_mode"])
    render_hero(active_mode)
    mode = st.segmented_control(
        "Режим отчета",
        list(REPORT_MODES),
        key="report_mode",
    ) or active_mode

    if not render_mode_workspace_tab(mode):
        return
    render_report_settings(mode)

    if mode == "Сравнение периодов":
        render_comparison_mode()
    elif mode == "Сувениры и касты на складе":
        render_warehouse_mode()
    elif mode == "Заказ Sonu":
        render_sonu_mode()
    elif mode == "Заказ поставщику":
        render_supplier_order_mode()
    elif mode == "О программе":
        render_about_mode()
    else:
        render_standard_report_mode()


if __name__ == "__main__":
    main()
