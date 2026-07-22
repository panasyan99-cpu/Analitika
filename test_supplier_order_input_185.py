from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from src.order_workflow import (
    GREEN_STONES_GROUP,
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    build_supplier_excel,
    canonical_stone,
    infer_ntr2,
    order_stone_bucket,
    parse_order_workbook,
    store_uploaded_workbook,
)


def _item(stone: str, sku: str = "SKU-GA") -> OrderItem:
    return OrderItem(
        row=11,
        set_id="Green set",
        sku=sku,
        stone=stone,
        group="Ring",
        sales=5,
        stock_63=0,
        stock_20=0,
        stores={"NTR1": 1, "NTR2": 2},
        total_stock=3,
        working_stock=3,
        ntr2_stock=2,
        ntr2_calculated=False,
        tvp_raw=0,
    )


def test_actual_ntr2_is_used_without_reconstruction() -> None:
    value, calculated, warning = infer_ntr2(
        total=9,
        store_values={"NTR1": 1, "NTR2": 4, "AB": 2, "63": 0, "20": 0},
        has_actual_ntr2=True,
    )
    assert value == 4
    assert calculated is False
    assert warning == "Сумма магазинов отличается от «Всего» на 2 шт."


def test_arbitrary_input_filename_is_preserved(tmp_path, monkeypatch) -> None:
    import src.order_workflow as workflow

    monkeypatch.setattr(workflow, "UPLOAD_DIR", tmp_path)
    payload = b"arbitrary workbook payload"
    path, digest = store_uploaded_workbook("YJ stones July final.xlsm", payload)
    assert path.name == f"{digest}.xlsm"
    assert path.read_bytes() == payload


def test_green_stones_is_navigation_only_and_concrete_stone_is_normalized() -> None:
    assert canonical_stone("Green Agat") == "Green Agate"
    assert canonical_stone("Green Stones", "ABC-GA-001") == "Green Agate"
    assert canonical_stone("Green Stones", "ABC-CD-001") == "Chrome Diopside"
    assert order_stone_bucket("Green Agat") == GREEN_STONES_GROUP
    assert order_stone_bucket("Chrome Diopside") == GREEN_STONES_GROUP


def test_supplier_excel_never_exports_green_stones_for_known_green_agate() -> None:
    item = _item("Green Stones", "ABC-GA-001")
    parsed = ParsedOrderWorkbook(
        source_name="Any report name.xlsx",
        source_hash="hash",
        upload_path="unused.xlsx",
        period="",
        supplier="Y&J",
        store_columns=("NTR1", "NTR2"),
        has_actual_ntr2=True,
        items=(item,),
    )
    draft = OrderDraft(
        source_hash="hash",
        source_name=parsed.source_name,
        mode=ORDER_MODE_STONES,
        orders={item.key: 5},
    )
    payload = build_supplier_excel(parsed, (item,), draft)
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook["Order"]
    assert sheet["C2"].value == "Green Agate"
    assert sheet["C2"].value != GREEN_STONES_GROUP


def test_parser_accepts_arbitrary_filename_and_actual_ntr2_column(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "Продажи товаров за период 01.01.2026 - 31.03.2026"
    sheet["A4"] = "Поставщик(и): Y&J"
    sheet["E7"] = "Продажи за период"
    sheet["G7"] = "Остатки"
    sheet["O7"] = "ТВП"
    sheet["G8"] = "63"
    sheet["H8"] = "20"
    sheet["I8"] = "NTR1"
    sheet["J8"] = "NTR2"
    sheet["K8"] = "AB"
    sheet["N8"] = "Всего"
    sheet["A11"] = "Зелёные модели"
    sheet["A12"] = "ABC-GA-001"
    sheet["B12"] = "Green Agat"
    sheet["C12"] = "Ring"
    sheet["E12"] = 5
    sheet["G12"] = 0
    sheet["H12"] = 0
    sheet["I12"] = 1
    sheet["J12"] = 4
    sheet["K12"] = 2
    sheet["N12"] = 9
    sheet["O12"] = 0

    path = tmp_path / "YJ stones 2026 final.xlsx"
    workbook.save(path)
    parsed = parse_order_workbook(path)

    assert parsed.source_name == "YJ stones 2026 final.xlsx"
    assert parsed.has_actual_ntr2 is True
    assert parsed.warnings == ()
    assert parsed.items[0].ntr2_stock == 4
    assert parsed.items[0].ntr2_calculated is False
    assert parsed.items[0].stone == "Green Agat"
