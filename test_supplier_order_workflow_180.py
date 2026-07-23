from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    CATEGORY_MEDIUM,
    CATEGORY_TOP,
    CATEGORY_WEAK,
    CATEGORY_ZERO,
    ORDER_MODE_PEARLS,
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    build_order_sets,
    build_supplier_excel,
    classify_set,
    infer_ntr2,
    item_in_mode,
    load_draft,
    parse_order_workbook,
    ring_validation,
    save_draft,
    suggested_order_quantity,
)


def item(
    sku: str,
    *,
    set_id: str = "Set# 1",
    stone: str = "RUBY",
    group: str = "Ring",
    sales: int = 0,
    total: int = 0,
    stock_63: int = 0,
    stock_20: int = 0,
    tvp: int = 0,
    stock_tt: int = 0,
    row: int = 12,
) -> OrderItem:
    return OrderItem(
        row=row,
        set_id=set_id,
        sku=sku,
        stone=stone,
        group=group,
        sales=sales,
        stock_63=stock_63,
        stock_20=stock_20,
        stores={"63": stock_63, "20": stock_20},
        total_stock=total,
        working_stock=max(0, total - stock_63 - stock_20),
        ntr2_stock=0,
        ntr2_calculated=True,
        tvp_raw=tvp,
        stock_tt=stock_tt,
    )


def test_set_category_uses_strongest_item_not_sum():
    weak_total_seven = [item("A", sales=2), item("B", sales=2, row=13), item("C", sales=2, row=14), item("D", sales=1, row=15)]
    category, driver, maximum, _ = classify_set(weak_total_seven)
    assert category == CATEGORY_WEAK
    assert maximum == 2
    assert driver == "A"

    category, _, maximum, _ = classify_set([item("A", sales=1), item("B", sales=4, row=13)])
    assert category == CATEGORY_MEDIUM
    assert maximum == 4

    category, _, maximum, _ = classify_set([item("A", sales=5), item("B", sales=0, row=13)])
    assert category == CATEGORY_TOP
    assert maximum == 5


def test_zero_sets_are_split_by_working_stock():
    category, _, _, segment = classify_set([item("A", sales=0, total=2)])
    assert category == CATEGORY_ZERO
    assert segment == "Нулевые с остатком"
    category, _, _, segment = classify_set([item("A", sales=0, total=0)])
    assert segment == "0/0 — не было остатка"


def test_mode_split_and_exclusions():
    assert item_in_mode(item("S", stone="RUBY"), ORDER_MODE_STONES)
    assert not item_in_mode(item("P", stone="FRESH WATER PEARL - WHITE"), ORDER_MODE_STONES)
    assert item_in_mode(item("P", stone="FRESH WATER PEARL - WHITE"), ORDER_MODE_PEARLS)
    assert not item_in_mode(item("P", stone="FRESH WATER ROUND PEARL - WHITE"), ORDER_MODE_PEARLS)
    assert not item_in_mode(item("P", stone="SEA PEARL WHITE"), ORDER_MODE_PEARLS)
    assert not item_in_mode(item("S", stone="BLUE SAPPHIRE HIGH QUALITY"), ORDER_MODE_STONES)
    assert not item_in_mode(item("S", stone="RS5=BSHQ-U113"), ORDER_MODE_STONES)
    assert not item_in_mode(item("S", stone="EMERALD HIGH QUALITY"), ORDER_MODE_STONES)


def test_positive_tvp_collapses_set_but_negative_tvp_keeps_error_visible():
    positive = build_order_sets([item("A", sales=5, tvp=3)], ORDER_MODE_STONES)[0]
    assert positive.has_positive_tvp
    assert not positive.has_negative_tvp

    mixed = build_order_sets([item("A", sales=5, tvp=3), item("B", sales=0, tvp=-1, row=13)], ORDER_MODE_STONES)[0]
    assert mixed.has_positive_tvp
    assert mixed.has_negative_tvp


def test_ntr2_is_inferred_until_real_column_exists():
    value, calculated, warning = infer_ntr2(5, {"63": 0, "20": 0, "NTR1": 1, "AB": 2}, False)
    assert value == 2
    assert calculated is True
    assert warning is None

    value, calculated, warning = infer_ntr2(5, {"63": 0, "20": 0, "NTR1": 1, "AB": 2, "NTR2": 2}, True)
    assert value == 2
    assert calculated is False
    assert warning is None


def test_recommendation_uses_internal_three_month_rate_and_positive_tvp_blocks_auto_order():
    assert suggested_order_quantity(item("A", sales=2, total=0, stock_tt=1)) == 3
    assert suggested_order_quantity(item("A", sales=6, total=6, stock_tt=1)) == 0
    assert suggested_order_quantity(item("A", sales=6, total=4, stock_tt=1)) == 4
    assert suggested_order_quantity(item("A", sales=7, total=1, stock_tt=1)) == 5
    assert suggested_order_quantity(item("A", sales=7, total=1, tvp=3, stock_tt=1)) == 0
    assert suggested_order_quantity(item("A", sales=7, total=1, tvp=-10, stock_tt=1)) == 5


def test_tt_rule_uses_set_ratio_without_report_speed_field():
    assert suggested_order_quantity(item("A", group="Earrings", sales=4, total=6, stock_tt=0)) == 3
    assert suggested_order_quantity(item("A", group="Ring", sales=4, total=3, stock_tt=0)) == 3
    assert suggested_order_quantity(item("A", group="Pendant", sales=5, total=2, stock_tt=0)) == 4


def _make_report(path: Path, include_ntr2: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"
    ws["A1"] = "Продажи товаров за период 20.04.2026 - 20.07.2026"
    ws["A4"] = "Поставщик(и): Y&J"
    ws["E7"] = "Продажи за период"
    ws["F7"] = "Скорость продаж (в день)"
    ws["G7"] = "Остатки"
    tvp_col = "P" if include_ntr2 else "O"
    total_col = "O" if include_ntr2 else "N"
    ws[f"{tvp_col}7"] = "ТВП"
    ws["A8"] = "Наименование товара"
    ws["D8"] = "Картинка"
    ws["E8"] = "Общие"
    ws["F8"] = "Общие"
    store_headers = ["63", "20", "AB", "NTR1", "Swandor", "Outlet", "Stock TT"]
    if include_ntr2:
        store_headers.append("NTR2")
    for index, header in enumerate(store_headers, start=7):
        ws.cell(8, index).value = header
    ws[f"{total_col}8"] = "Всего"

    ws["A11"] = "Set# 00001"
    ws["A12"] = "ER-1"
    ws["B12"] = "RUBY"
    ws["C12"] = "Earrings"
    ws["E12"] = 5
    ws["G12"] = 0
    ws["H12"] = 1
    ws["I12"] = 1
    ws["J12"] = 1
    ws["K12"] = 0
    ws["L12"] = 0
    ws["M12"] = 0
    if include_ntr2:
        ws["N12"] = 2
    ws[f"{total_col}12"] = 5
    ws[f"{tvp_col}12"] = 3

    ws["A13"] = "RG-1"
    ws["B13"] = "RUBY"
    ws["C13"] = "Ring"
    ws["E13"] = 1
    ws["G13"] = 0
    ws["H13"] = 0
    ws["I13"] = 0
    ws["J13"] = 0
    ws["K13"] = 0
    ws["L13"] = 0
    ws["M13"] = 0
    if include_ntr2:
        ws["N13"] = 0
    ws[f"{total_col}13"] = 0
    ws[f"{tvp_col}13"] = -1
    wb.save(path)


def test_parser_reads_current_report_and_infers_ntr2(tmp_path: Path):
    path = tmp_path / "order.xlsx"
    _make_report(path, include_ntr2=False)
    parsed = parse_order_workbook(path)
    assert parsed.period == "20.04.2026 - 20.07.2026"
    assert parsed.supplier == "Y&J"
    assert parsed.has_actual_ntr2 is False
    assert parsed.store_columns == ("63", "20", "AB", "NTR1", "Swandor", "Outlet", "Stock TT")
    first, second = parsed.items
    assert first.ntr2_stock == 2
    assert first.working_stock == 4  # Total 5 minus hidden store 20 = 1.
    assert first.stock_tt == 0
    assert first.tvp_raw == 3
    assert second.tvp_raw == -1
    assert any("Ошибка ТВП" in message for message in second.errors)


def test_parser_uses_future_real_ntr2_column(tmp_path: Path):
    path = tmp_path / "order_ntr2.xlsx"
    _make_report(path, include_ntr2=True)
    parsed = parse_order_workbook(path)
    assert parsed.has_actual_ntr2 is True
    assert parsed.items[0].ntr2_stock == 2
    assert parsed.items[0].ntr2_calculated is False


def test_ring_allocation_and_export(tmp_path: Path):
    ring = item("RG-1", sales=5, total=2)
    earrings = item("ER-1", group="Earrings", sales=5, total=0, row=13)
    assert ring.is_ring is True
    assert earrings.is_ring is False
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders = {ring.key: 5, earrings.key: 3}
    draft.sizes = {ring.key: {"18": 2, "19": 2, "20": 1}}
    draft.stock_checked = {ring.key: True}
    assert ring_validation(ring, draft) == (5, 5, True, True)

    parsed = ParsedOrderWorkbook(
        source_name="source.xlsx",
        source_hash="hash",
        upload_path=str(tmp_path / "empty.xlsx"),
        period="",
        supplier="",
        store_columns=(),
        has_actual_ntr2=False,
        items=(ring, earrings),
    )
    # A minimal valid XLSX zip is enough because these fixture items have no photos.
    empty = Workbook()
    empty.save(parsed.upload_path)
    payload = build_supplier_excel(parsed, [ring, earrings], draft)
    result = load_workbook(io.BytesIO(payload))
    ws = result["Order"]
    assert [ws.cell(1, col).value for col in range(1, 7)] == [
        "Фото", "Артикул", "Камень", "Группа", "Количество к заказу", "Размеры"
    ]
    assert ws["B2"].value == "RG-1"
    assert ws["E2"].value == 5
    assert ws["F2"].value == "18 × 2; 19 × 2; 20 × 1"
    assert ws["F3"].value in (None, "")


def test_draft_is_persisted_in_sqlite(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(workflow, "DRAFT_DB", tmp_path / "drafts.sqlite3")
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders = {"A": 5}
    saved_at = save_draft(draft)
    restored = load_draft("hash", "source.xlsx", ORDER_MODE_STONES)
    assert saved_at
    assert restored.orders == {"A": 5}


def test_same_set_is_split_by_normalized_stone_before_category_and_tvp():
    sets = build_order_sets(
        [
            item("BS-RING", set_id="Set# 2622", stone="Blue Sapphire", sales=0, total=2, tvp=0, row=12),
            item("BS-PENDANT", set_id="Set# 2622", stone="BLUE SAPPHIRE", group="Pendant", sales=5, total=0, tvp=0, row=13),
            item("RUBY-RING", set_id="Set# 2622", stone="Ruby", sales=8, total=1, tvp=-1, row=14),
            item("RUBY-PENDANT", set_id="Set# 2622", stone="RUBY", group="Pendant", sales=1, total=1, tvp=3, row=15),
        ],
        ORDER_MODE_STONES,
    )

    assert len(sets) == 2
    by_stone = {order_set.stone: order_set for order_set in sets}

    sapphire = by_stone["Blue Sapphire"]
    assert sapphire.set_id == "Set# 2622"
    assert [row.sku for row in sapphire.items] == ["BS-RING", "BS-PENDANT"]
    assert sapphire.category == CATEGORY_TOP
    assert sapphire.driver_sku == "BS-PENDANT"
    assert sapphire.has_positive_tvp is False
    assert sapphire.has_negative_tvp is False

    ruby = by_stone["Ruby"]
    assert [row.sku for row in ruby.items] == ["RUBY-RING", "RUBY-PENDANT"]
    assert ruby.category == CATEGORY_TOP
    assert ruby.driver_sku == "RUBY-RING"
    assert ruby.has_positive_tvp is True
    assert ruby.has_negative_tvp is True

    assert sapphire.key != ruby.key
