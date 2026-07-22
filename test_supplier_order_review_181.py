from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw

import src.order_workflow as workflow
from src.order_workflow import (
    CATEGORY_TOP,
    CATEGORY_WEAK,
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    build_order_sets,
    canonical_stone,
    parse_order_workbook,
    validate_draft_payload,
)


def _item(sku: str, *, sales: int, row: int, tvp: int = 0) -> OrderItem:
    return OrderItem(
        row=row,
        set_id="Без комплекта",
        sku=sku,
        stone="LAP",
        group="Ring",
        sales=sales,
        stock_63=0,
        stock_20=0,
        stores={},
        total_stock=0,
        working_stock=0,
        ntr2_stock=0,
        ntr2_calculated=True,
        tvp_raw=tvp,
        ungrouped=True,
    )


def test_order_stone_aliases_match_shared_report_vocabulary():
    assert canonical_stone("Lapis Lazuli") == "Lapis Lazurite"
    assert canonical_stone("Lapiz") == "Lapis Lazurite"
    assert canonical_stone("LAP") == "Lapis Lazurite"
    assert canonical_stone("WHO") == "White Howlite"
    assert canonical_stone("GAM") == "Green Amethyst"
    assert canonical_stone("DJ") == "Dalmatian Jasper"
    assert canonical_stone("OB") == "Obsidian"
    assert canonical_stone("RJ") == "Red Jasper"
    assert canonical_stone("GA") == "Green Agate"
    assert canonical_stone("Black Agate") == "Agate"
    assert canonical_stone("London BT") == "London Topaz"
    assert canonical_stone("Swiss BT") == "Swiss Topaz"


def test_legacy_auto_seeded_draft_is_reset_to_zero():
    draft = validate_draft_payload(
        {
            "version": 1,
            "source_hash": "hash",
            "source_name": "Заказ.xlsx",
            "mode": ORDER_MODE_STONES,
            "orders": {"A": 5, "B": 10},
            "sizes": {"A": {"18": 5}},
        }
    )
    assert draft.version == workflow.DRAFT_VERSION
    assert draft.orders == {}
    assert draft.sizes == {}


def test_new_items_are_seeded_with_zero_not_recommendation(monkeypatch):
    saved = []
    monkeypatch.setattr(workflow, "_save_session_draft", lambda draft: saved.append(dict(draft.orders)))
    item = OrderItem(
        row=12,
        set_id="Set# 1",
        sku="RG-1",
        stone="Ruby",
        group="Ring",
        sales=8,
        stock_63=0,
        stock_20=0,
        stores={},
        total_stock=0,
        working_stock=0,
        ntr2_stock=0,
        ntr2_calculated=True,
        tvp_raw=0,
    )
    order_set = build_order_sets([item], ORDER_MODE_STONES)[0]
    draft = OrderDraft(source_hash="hash", source_name="Заказ.xlsx", mode=ORDER_MODE_STONES)
    workflow._seed_defaults(draft, [order_set])
    assert draft.orders[item.key] == 0
    assert saved


def test_ungrouped_items_do_not_promote_each_other_and_tvp_is_separate():
    sets = build_order_sets(
        [
            _item("TOP", sales=5, row=20),
            _item("WEAK", sales=1, row=21),
            _item("TVP", sales=5, row=22, tvp=3),
        ],
        ORDER_MODE_STONES,
    )
    virtual = [order_set for order_set in sets if order_set.is_ungrouped]
    assert len(virtual) == 3
    assert any(order_set.category == CATEGORY_TOP and not order_set.has_positive_tvp and [i.sku for i in order_set.items] == ["TOP"] for order_set in virtual)
    assert any(order_set.category == CATEGORY_WEAK and [i.sku for i in order_set.items] == ["WEAK"] for order_set in virtual)
    assert any(order_set.category == CATEGORY_TOP and order_set.has_positive_tvp and [i.sku for i in order_set.items] == ["TVP"] for order_set in virtual)


def _make_visual_match_report(path: Path, image_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TDSheet"
    ws["A1"] = "Продажи товаров за период 20.04.2026 - 20.07.2026"
    ws["A4"] = "Поставщик(и): Y&J"
    ws["E7"] = "Продажи за период"
    ws["G7"] = "Остатки"
    ws["O7"] = "ТВП"
    ws["A8"] = "Наименование товара"
    ws["D8"] = "Картинка"
    for index, header in enumerate(["63", "20", "AB", "NTR1", "Swandor", "Outlet", "Stock TT"], start=7):
        ws.cell(8, index).value = header
    ws["N8"] = "Всего"

    ws["A11"] = "Set# 00001"
    ws["A12"] = "RG-OLD"
    ws["B12"] = "Lapis Lazuli"
    ws["C12"] = "Ring"
    ws["E12"] = 5
    ws["L12"] = 3
    ws["N12"] = 4
    ws.add_image(XLImage(str(image_path)), "D12")

    ws["A13"] = "<Без комплекта>"
    ws["A14"] = "RG-NEW"
    ws["B14"] = "LAP"
    ws["C14"] = "Ring"
    ws["E14"] = 3
    ws["N14"] = 0
    ws.add_image(XLImage(str(image_path)), "D14")
    wb.save(path)


def test_parser_recognizes_ungrouped_section_tt_and_exact_visual_match(tmp_path: Path):
    image_path = tmp_path / "ring.png"
    image = Image.new("RGB", (160, 120), "white")
    draw = ImageDraw.Draw(image)
    draw.ellipse((45, 20, 115, 90), outline="black", width=8)
    draw.rectangle((72, 5, 88, 30), fill="red")
    image.save(image_path)

    report = tmp_path / "order.xlsx"
    _make_visual_match_report(report, image_path)
    parsed = parse_order_workbook(report)
    grouped, ungrouped = parsed.items
    assert grouped.stock_tt == 3
    assert ungrouped.ungrouped is True
    assert ungrouped.visual_match_set_id == "Set# 00001"
    assert ungrouped.visual_match_sku == "RG-OLD"
    assert ungrouped.visual_match_status == "confirmed"


def test_saved_workspace_reopens_report_and_selected_positions(tmp_path: Path, monkeypatch):
    runtime = tmp_path / "runtime"
    monkeypatch.setattr(workflow, "UPLOAD_DIR", runtime / "uploads")
    monkeypatch.setattr(workflow, "DRAFT_DB", runtime / "order_drafts.sqlite3")
    workflow.cached_parse_order_workbook.clear()

    image_path = tmp_path / "ring.png"
    Image.new("RGB", (80, 80), "white").save(image_path)
    report_path = tmp_path / "Заказ.xlsx"
    _make_visual_match_report(report_path, image_path)

    stored_path, digest = workflow.store_uploaded_workbook(report_path.name, report_path.read_bytes())
    parsed = workflow.cached_parse_order_workbook(str(stored_path), report_path.name, digest)
    item = parsed.items[0]
    draft = OrderDraft(source_hash=digest, source_name=report_path.name, mode=ORDER_MODE_STONES)
    draft.orders[item.key] = 5
    workflow.save_draft(draft)

    workspaces = workflow.list_saved_order_workspaces()
    assert len(workspaces) == 1
    workspace = workspaces[0]
    assert workspace.source_name == report_path.name
    assert workspace.selected_positions == 1
    assert workspace.total_quantity == 5
    assert workspace.preferred_mode == ORDER_MODE_STONES

    restored = workflow.load_saved_order_workspace(workspace)
    assert restored.source_hash == digest
    assert restored.source_name == report_path.name
    assert len(restored.items) == 2
