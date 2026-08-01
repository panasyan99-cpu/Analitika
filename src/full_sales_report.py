from __future__ import annotations

from collections import deque
from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from .report import classify, extract_period, norm_product, normalize_store_from_report


FULL_HIERARCHY_FIELDS = (
    "МАГАЗИН",
    "МЕНЕДЖЕР",
    "ТОВАР",
    "КАМЕНЬ/ВСТАВКА",
    "ПРОБА",
    "НОМЕНКЛАТУРНАЯ ГРУППА",
)


@dataclass(frozen=True)
class _ReportRow:
    number: int
    text: str
    indent: int
    bold: bool
    values: tuple[object, ...]


def _clean_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _normalized(value: object) -> str:
    return _clean_text(value).upper().replace("Ё", "Е")


def _number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def normalize_purity_label(value: object) -> str:
    text = _clean_text(value)
    return text if text else "Не указано"


def classify_metal_group(purity: object) -> str:
    text = normalize_purity_label(purity).upper().replace("Ё", "Е")
    compact = re.sub(r"[^A-ZА-Я0-9]+", "", text)
    if compact in {"НЕУКАЗАНО", "OTHER0", "OTHER", "0"}:
        return "Другое"
    if "AU" in compact or "GOLD" in compact or "ЗОЛОТ" in compact:
        return "Золото и платина"
    if "PT" in compact or "PLATIN" in compact or "ПЛАТИН" in compact:
        return "Золото и платина"
    if "925" in compact or compact.startswith("AG") or "SILVER" in compact or "СЕРЕБ" in compact:
        return "Серебро"
    return "Другое"


def hierarchy_header(path: Path) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return _clean_text(ws.cell(4, 1).value)
    finally:
        wb.close()


def is_full_sales_report(path: Path) -> bool:
    header = _normalized(hierarchy_header(path))
    return all(field in header for field in FULL_HIERARCHY_FIELDS)


def _metric_columns(ws) -> dict[str, int | None]:
    active_group = ""
    found: dict[tuple[str, str], int] = {}
    for column in range(1, ws.max_column + 1):
        top = _normalized(ws.cell(4, column).value)
        if top:
            active_group = top
        bottom = _normalized(ws.cell(5, column).value)
        if bottom:
            found[(active_group, bottom)] = column

    def pick(group: str, *bottom_fragments: str) -> int | None:
        group = _normalized(group)
        fragments = tuple(_normalized(item) for item in bottom_fragments)
        for (candidate_group, candidate_bottom), column in found.items():
            if group not in candidate_group:
                continue
            if all(fragment in candidate_bottom for fragment in fragments):
                return column
        return None

    shipped_qty = pick("Отгружено", "Кол-во")
    shipped_amount = pick("Отгружено", "Сумма")
    sold_qty = pick("Продано", "Кол-во")
    sold_amount = pick("Продано", "Сумма")
    return_qty = pick("Возврат", "Кол-во")
    return_amount = pick("Возврат", "Сумма")

    # The report's exact quantity is the shipped quantity. In weighted rows the
    # "Продано" quantity is rounded to whole pieces, while revenue is unchanged.
    return {
        "qty": shipped_qty or sold_qty,
        "amount": shipped_amount or sold_amount,
        "return_qty": return_qty,
        "return_amount": return_amount,
    }


def _row_from_cells(number: int, cells: tuple[object, ...]) -> _ReportRow:
    first = cells[0]
    alignment = getattr(first, "alignment", None)
    font = getattr(first, "font", None)
    return _ReportRow(
        number=number,
        text=_clean_text(getattr(first, "value", None)),
        indent=int(getattr(alignment, "indent", 0) or 0),
        bold=bool(getattr(font, "bold", False)),
        values=tuple(getattr(cell, "value", None) for cell in cells),
    )


def _value(row: _ReportRow, column: int | None) -> float:
    if not column or column < 1 or column > len(row.values):
        return 0.0
    return _number(row.values[column - 1])


def _is_leaf(row: _ReportRow, first_child: _ReportRow, second_child: _ReportRow) -> bool:
    if not row.bold or row.indent not in {3, 4, 5}:
        return False
    return (
        not first_child.bold
        and not second_child.bold
        and first_child.indent == row.indent + 4
        and second_child.indent == row.indent + 3
    )


def parse_full_sales_report_with_period(path: Path) -> tuple[pd.DataFrame, tuple | None]:
    """Parse the current full 1C sales hierarchy exactly once at SKU level.

    Expected hierarchy:
    Store -> Manager -> Category -> Branch -> Nomenclature group -> SKU -> Stone -> Purity.

    Only SKU/service leaves are retained. Every subtotal row is ignored, so the
    sum of parsed rows must equal the 1C grand total before data is returned.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        if not all(field in _normalized(ws.cell(4, 1).value) for field in FULL_HIERARCHY_FIELDS):
            raise ValueError("Файл не соответствует полному формату 1С «Магазин → Менеджер → Товар → Камень → Проба → Номенклатурная группа».")

        period = extract_period(ws)
        columns = _metric_columns(ws)
        if columns["qty"] is None or columns["amount"] is None:
            raise ValueError("В заголовке не найдены колонки «Отгружено: Кол-во» и «Отгружено: Сумма прод.».")

        row_iter = (
            _row_from_cells(number, cells)
            for number, cells in enumerate(
                ws.iter_rows(min_row=7, max_col=ws.max_column, values_only=False),
                start=7,
            )
        )
        window: deque[_ReportRow] = deque()
        for _ in range(3):
            try:
                window.append(next(row_iter))
            except StopIteration:
                break

        current_store_raw = ""
        current_store: str | None = None
        current_manager = "Не указан"
        current_category = "Не указано"
        current_branch = "Не указано"
        current_group = ""
        rows: list[dict] = []
        grand_total: dict[str, float] | None = None

        while window:
            current = window[0]
            first_child = window[1] if len(window) > 1 else _ReportRow(0, "", -1, False, ())
            second_child = window[2] if len(window) > 2 else _ReportRow(0, "", -1, False, ())
            upper = _normalized(current.text)

            if upper.startswith("ИТОГО"):
                grand_total = {
                    "qty": _value(current, columns["qty"]),
                    "amount": _value(current, columns["amount"]),
                    "return_qty": _value(current, columns["return_qty"]),
                    "return_amount": _value(current, columns["return_amount"]),
                }
            elif _is_leaf(current, first_child, second_child):
                raw_stone = first_child.text or "Other"
                purity = normalize_purity_label(second_child.text)
                segment, stone, rule = classify(raw_stone)

                if current.indent == 3:
                    raw_group = current.text or "Без названия"
                    branch = "Услуги"
                elif current.indent == 4:
                    raw_group = current_branch if current_branch != "Не указано" else current_category
                    branch = current_branch
                else:
                    raw_group = current_group or current_branch or current_category
                    branch = current_branch

                qty = _value(current, columns["qty"])
                amount = _value(current, columns["amount"])
                return_qty = _value(current, columns["return_qty"])
                return_amount = _value(current, columns["return_amount"])
                if current_store is not None and (qty != 0 or amount != 0 or return_qty != 0 or return_amount != 0):
                    product_code = norm_product(raw_group or "Other")
                    rows.append({
                        "Магазин": current_store,
                        "Исходный магазин": current_store_raw,
                        "Продавец": current_manager or "Не указан",
                        "Категория": current_category or "Не указано",
                        "Подгруппа": branch or "Не указано",
                        "Товар": current.text or "Без названия",
                        "Проба": purity,
                        "Группа металла": classify_metal_group(purity),
                        "Сегмент": {
                            "TOP STONES": "Top Stones",
                            "PEARLS": "Pearls",
                            "COLORED STONES": "Other Stones",
                        }.get(segment, segment),
                        "Код сегмента": segment,
                        "Камень": stone,
                        "Исходный камень": raw_stone,
                        "Номенклатурная группа": raw_group or "Other",
                        "Код группы": product_code,
                        "Количество": qty,
                        "Выручка": amount,
                        "Возврат количество": return_qty,
                        "Возврат сумма": return_amount,
                        "Остаток": 0.0,
                        "Дата остатка": period[1] if period else None,
                        "Есть остаток": False,
                        "Правило": rule,
                        "Источник": path.name,
                    })
            elif current.text:
                # State rows are identified only after leaf detection because a
                # SKU can use the same indent/bold style as category or branch.
                if current.indent == 0 and current.bold:
                    current_store_raw = current.text
                    current_store = normalize_store_from_report(current.text)
                    current_manager = "Не указан"
                    current_category = "Не указано"
                    current_branch = "Не указано"
                    current_group = ""
                elif current.indent == 2 and not current.bold:
                    current_manager = current.text or "Не указан"
                    current_category = "Не указано"
                    current_branch = "Не указано"
                    current_group = ""
                elif current.indent == 4 and current.bold:
                    current_category = current.text
                    current_branch = "Не указано"
                    current_group = ""
                elif current.indent == 5 and current.bold:
                    current_branch = current.text
                    current_group = ""
                elif current.indent == 6 and current.bold:
                    current_group = current.text

            window.popleft()
            try:
                window.append(next(row_iter))
            except StopIteration:
                pass

        detail = pd.DataFrame(rows)
        if detail.empty:
            raise ValueError("В полном отчёте не найдены SKU-строки продаж.")

        parsed = {
            "qty": float(detail["Количество"].sum()),
            "amount": float(detail["Выручка"].sum()),
            "return_qty": float(detail["Возврат количество"].sum()),
            "return_amount": float(detail["Возврат сумма"].sum()),
        }
        if grand_total is None:
            raise ValueError("В отчёте не найдена итоговая строка «Итого».")

        differences = {
            key: parsed[key] - grand_total[key]
            for key in parsed
            if abs(parsed[key] - grand_total[key]) > (0.1 if "qty" in key else 0.5)
        }
        if differences:
            readable = ", ".join(
                f"{key}: распознано {parsed[key]:,.3f}, в 1С {grand_total[key]:,.3f}"
                for key in differences
            )
            raise ValueError(f"Контроль итогов не пройден ({readable}). Графики не построены, чтобы не показывать неверные данные.")

        detail.attrs["source_totals"] = grand_total
        detail.attrs["parsed_totals"] = parsed
        return detail, period
    finally:
        wb.close()

SUPPLIER_SUMMARY_COLUMNS = (
    "Магазин", "Поставщик", "Проба", "Группа металла",
    "Сегмент", "Код сегмента", "Камень", "Исходный камень",
    "Номенклатурная группа", "Код группы", "Количество", "Выручка",
    "Возврат количество", "Возврат сумма", "Остаток", "Дата остатка",
    "Есть остаток", "Правило",
)


def is_supplier_summary_report(path: Path) -> bool:
    """Detect the compact 1C hierarchy grouped only by supplier.

    Detection is intentionally limited to the hierarchy header (row 4). The
    report filter ``Поставщик(и):`` in row 2 must never classify a full sales
    workbook as a supplier report.
    """
    header = _normalized(hierarchy_header(path))
    detailed_fields = ("МАГАЗИН", "КАМЕНЬ", "ПРОБА", "НОМЕНКЛАТУРНАЯ ГРУППА")
    return "ПОСТАВЩИК" in header and not any(field in header for field in detailed_fields)


def parse_supplier_summary_report_with_period(path: Path) -> tuple[pd.DataFrame, tuple | None]:
    """Parse a compact 1C supplier summary without counting its subtotal row.

    Typical rows are: unnamed supplier (optional), bold ``Поставщики`` subtotal,
    named supplier leaves and final ``Итого``. Sales use the ``Отгружено`` block;
    returns are retained separately.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        header = _normalized(ws.cell(4, 1).value)
        if "ПОСТАВЩИК" not in header:
            raise ValueError("В строке иерархии отчёта не найден уровень «Поставщик».")
        period = extract_period(ws)
        columns = _metric_columns(ws)
        if columns["qty"] is None or columns["amount"] is None:
            raise ValueError("В отчёте поставщиков не найдены колонки «Отгружено: Кол-во» и «Сумма прод.»." )

        rows: list[dict] = []
        grand_total: dict[str, float] | None = None
        for number, cells in enumerate(
            ws.iter_rows(min_row=7, max_col=ws.max_column, values_only=False),
            start=7,
        ):
            row = _row_from_cells(number, cells)
            upper = _normalized(row.text)
            qty = _value(row, columns["qty"])
            amount = _value(row, columns["amount"])
            return_qty = _value(row, columns["return_qty"])
            return_amount = _value(row, columns["return_amount"])

            if upper.startswith("ИТОГО"):
                grand_total = {
                    "qty": qty,
                    "amount": amount,
                    "return_qty": return_qty,
                    "return_amount": return_amount,
                }
                continue
            if upper in {"ПОСТАВЩИКИ", "ПОСТАВЩИК"}:
                continue
            if qty == 0 and amount == 0 and return_qty == 0 and return_amount == 0:
                continue

            supplier = row.text or "Не указан"
            rows.append({
                "Магазин": "Сеть",
                "Поставщик": supplier,
                "Проба": "Не указано",
                "Группа металла": "Другое",
                "Сегмент": "Other Stones",
                "Код сегмента": "COLORED STONES",
                "Камень": "Other Colored Stones",
                "Исходный камень": "Other",
                "Номенклатурная группа": "Other",
                "Код группы": "Other",
                "Количество": qty,
                "Выручка": amount,
                "Возврат количество": return_qty,
                "Возврат сумма": return_amount,
                "Остаток": 0.0,
                "Дата остатка": period[1] if period else None,
                "Есть остаток": False,
                "Правило": "supplier summary",
            })

        detail = pd.DataFrame(rows, columns=SUPPLIER_SUMMARY_COLUMNS)
        if detail.empty:
            raise ValueError("В отчёте поставщиков не найдены строки поставщиков.")
        if grand_total is None:
            raise ValueError("В отчёте поставщиков не найдена итоговая строка «Итого».")

        parsed = {
            "qty": float(detail["Количество"].sum()),
            "amount": float(detail["Выручка"].sum()),
            "return_qty": float(detail["Возврат количество"].sum()),
            "return_amount": float(detail["Возврат сумма"].sum()),
        }
        differences = {
            key: parsed[key] - grand_total[key]
            for key in parsed
            if abs(parsed[key] - grand_total[key]) > (0.1 if "qty" in key else 0.5)
        }
        if differences:
            readable = ", ".join(
                f"{key}: распознано {parsed[key]:,.3f}, в 1С {grand_total[key]:,.3f}"
                for key in differences
            )
            raise ValueError(f"Контроль итогов поставщиков не пройден ({readable}).")
        detail.attrs["source_totals"] = grand_total
        detail.attrs["parsed_totals"] = parsed
        return detail, period
    finally:
        wb.close()

