from __future__ import annotations

import json
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Product, canonical_stone_name, normalize_stone_names

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
RULES_PATH = Path(__file__).with_name("stone_suffixes.json")

CATEGORIES = ["", "Брелоки", "Подвески", "Серьги", "Кольца", "Браслеты", "Ожерелья", "Аксессуары"]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _yes(value: Any) -> bool:
    return _text(value).lower() in {"да", "yes", "true", "1", "получено", "checked"}


def _int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(" ", "").replace(",", ".")))
    except (TypeError, ValueError):
        return default


def _float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def normalize_description(text: str) -> str:
    result = (text or "").strip().lstrip("/")
    result = re.sub(
        r"\bS\.?\s*Steel\b|\bStainless\s+steel\b|\b304\s+Stainless\s+steel\b",
        "Steel",
        result,
        flags=re.I,
    )
    result = re.sub(r"\bZinc\s+Alloy\b|\bZinc\b|\bAlloy\b", "Brass", result, flags=re.I)
    result = re.sub(r"\s*/\s*", " / ", result)
    result = re.sub(r"\s*\+\s*", " + ", result)
    return re.sub(r"\s{2,}", " ", result).strip(" /")


def _prefix(sku: str) -> str:
    match = re.match(r"^[A-Z]+", sku.upper())
    return match.group(0) if match else ""


def detect_category(sku: str, description: str) -> str:
    lower = description.lower()
    if "earring" in lower:
        return "Серьги"
    if "necklace" in lower:
        return "Ожерелья"
    if any(token in lower for token in ("accessories", "metal tape", "plastic accessories")):
        return "Аксессуары"
    prefix = _prefix(sku)
    if prefix == "FXK":
        return "Брелоки"
    if prefix in {"FXE", "FBE", "FAE", "KE"}:
        return "Серьги"
    if prefix in {"FXR", "FBR", "KR"}:
        return "Кольца"
    if prefix in {"FXN", "FBN", "FAN", "XN", "RSN", "SFN"}:
        return "Ожерелья"
    if prefix in {"FXBR", "FBBR", "FABR", "BBRB", "XBR", "FLBR", "KH"}:
        return "Браслеты"
    if prefix in {"AN", "FXP", "KP"}:
        return "Подвески"
    return ""


def detect_materials(description: str) -> str:
    lower = description.lower()
    rules = [
        ("Silver", ("серебро", "silver 925", "sterling silver")),
        ("Steel", ("s.steel", "stainless steel", "304 stainless", "steel")),
        ("Brass", ("zinc", "alloy", "brass")),
        ("Plastic", ("plastic",)),
        ("Metal tape", ("metal tape",)),
        ("Aluminum", ("aluminum", "aluminium")),
        ("Leather", ("leather",)),
        ("Cord", ("cord", "rope", "string")),
        ("Textile", ("textile", "fabric")),
        ("Glass", ("glass",)),
        ("Wood", ("wood",)),
        ("Resin", ("resin",)),
    ]
    values = [name for name, tokens in rules if any(token in lower for token in tokens)]
    return "; ".join(dict.fromkeys(values))


STONE_RULES = [
    ("Rose Quartz", ("rose quartz",)),
    ("Smoky Quartz", ("smoky quartz",)),
    ("Picture Jasper", ("picture jasper",)),
    ("Amazonite", ("amazonite",)),
    ("Amethyst", ("amethyst",)),
    ("Tourmaline", ("tourmaline",)),
    ("Lapis Lazurite", ("lapis lazurite", "lapis lazuli", "lapis")),
    ("Labradorite", ("labradorite",)),
    ("Fluorite", ("green fluorite", "fluorite")),
    ("Turquoise", ("turquoise",)),
    ("Agate", (
        "black agate", "green agate", "blue lace agate", "blue agate",
        "red agate", "yellow agate", "white agate", "agate",
    )),
    ("Onyx", ("black onyx", "onyx")),
    ("Howlite", ("white howlite", "howlite")),
    ("Aventurine", ("red aventurine", "green aventurine", "aventurine")),
    ("Obsidian", ("obsidian",)),
    ("Tiger Eye", ("tiger eye", "tiger's eye")),
    ("Moonstone", ("moonstone",)),
    ("Opalite", ("opalite",)),
    ("Sodalite", ("sodalite",)),
    ("Garnet", ("garnet",)),
    ("Pearl", ("pearl",)),
    ("Zircon", ("zircon", " cz ")),
    ("Quartz", ("quartz",)),
    ("Jasper", ("jasper",)),
]

DESCRIPTION_COLOR_RULES = [
    ("Чёрный", ("black agate", "black onyx", "black stone")),
    ("Зелёный", ("green agate", "green fluorite", "green aventurine")),
    ("Красный", ("red agate", "red aventurine", "red jasper")),
    ("Жёлтый", ("yellow agate",)),
    ("Белый", ("white agate", "white howlite")),
    ("Синий", ("blue agate", "blue lace agate", "lapis")),
    ("Розовый", ("pink agate", "rose quartz")),
    ("Фиолетовый", ("purple agate", "amethyst")),
]

STONE_COLORS = {
    "Amethyst": "Фиолетовый",
    "Green Amethyst": "Зелёный",
    "Amazonite": "Голубой",
    "Labradorite": "Серый",
    "Green Fluorite": "Зелёный",
    "Fluorite": "Зелёный",
    "Rose Quartz": "Розовый",
    "Red Aventurine": "Красный",
    "Green Aventurine": "Зелёный",
    "Aventurine": "Зелёный",
    "Tourmaline": "Мультиколор",
    "Lapis Lazurite": "Синий",
    "Turquoise": "Бирюзовый",
    "Picture Jasper": "Коричневый",
    "Dalmatian Jasper": "Белый; Чёрный",
    "Red Jasper": "Красный",
    "Jasper": "Коричневый",
    "Green Agate": "Зелёный",
    "Black Agate": "Чёрный",
    "Agate": "Натуральный",
    "White Howlite": "Белый",
    "Black Onyx": "Чёрный",
    "Onyx": "Чёрный",
    "Obsidian": "Чёрный",
    "Tiger Eye": "Коричневый",
    "Moonstone": "Белый",
    "Opalite": "Белый",
    "Sodalite": "Синий",
    "Garnet": "Красный",
    "Pearl": "Белый",
    "Zircon": "Прозрачный",
    "Multistone": "Мультиколор",
}


def _suffix_rules() -> list[dict[str, str]]:
    try:
        return json.loads(RULES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def detect_stone(sku: str, description: str) -> tuple[str, str, str]:
    suffix = sku.upper().rsplit("-", 1)[-1] if "-" in sku else ""
    for rule in sorted(_suffix_rules(), key=lambda row: len(str(row.get("code", ""))), reverse=True):
        if suffix == str(rule.get("code", "")).upper().strip():
            return (
                normalize_stone_names(_text(rule.get("stone"))),
                _text(rule.get("color")),
                _text(rule.get("confidence")),
            )

    lower = f" {description.lower()} "
    stones: list[str] = []
    for stone, tokens in STONE_RULES:
        if any(token in lower for token in tokens):
            canonical = canonical_stone_name(stone)
            if canonical and canonical not in stones:
                stones.append(canonical)

    colors: list[str] = []
    for color, tokens in DESCRIPTION_COLOR_RULES:
        if any(token in lower for token in tokens) and color not in colors:
            colors.append(color)

    for stone in stones:
        color = STONE_COLORS.get(stone, "")
        for value in [part.strip() for part in color.split(";") if part.strip()]:
            if value not in colors:
                colors.append(value)

    return "; ".join(stones), "; ".join(colors), "По описанию" if stones else "Не распознано"


def _merge_products(products: list[Product]) -> list[Product]:
    merged: OrderedDict[str, Product] = OrderedDict()
    actual_totals: dict[str, int] = {}
    for product in products:
        sku = product.sku.strip()
        if not sku:
            continue
        if sku not in merged:
            merged[sku] = Product(**product.to_dict())
            actual_totals[sku] = product.actual_qty or 0
            continue
        target = merged[sku]
        target.qty_document += product.qty_document
        actual_totals[sku] += product.actual_qty or 0
        target.boxes = ", ".join(
            dict.fromkeys(
                part.strip()
                for value in [
                    str(target.boxes or ""),
                    str(product.boxes or ""),
                ]
                for part in value.split(",")
                if part.strip()
            )
        )
        for attr in ["description", "category", "material", "stone", "color", "image_path", "recognition"]:
            if not getattr(target, attr) and getattr(product, attr):
                setattr(target, attr, getattr(product, attr))
        if target.unit_weight_kg is None:
            target.unit_weight_kg = product.unit_weight_kg
        target.comment = "; ".join(dict.fromkeys(x for x in [target.comment, product.comment] if x))
        target.checked = target.checked and product.checked

    result = list(merged.values())
    for index, product in enumerate(result, start=1):
        product.number = index
        actual = actual_totals.get(product.sku, 0)
        product.received = actual > 0
        product.actual_manual = actual if product.received and actual != product.qty_document else None
    return result


def parse_master_xlsx(path: Path, image_dir: Path) -> list[Product]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {_text(sheet.cell(1, col).value).lower(): col for col in range(1, sheet.max_column + 1)}

    aliases = {
        "number": ["№", "номер"],
        "boxes": ["коробки", "номера коробок"],
        "sku": ["артикул", "sku"],
        "qty": ["по документу", "кол-во по документу", "количество по документу"],
        "category": ["категория"],
        "material": ["материал"],
        "stone": ["камень"],
        "color": ["цвет"],
        "weight": ["вес 1 шт.", "вес 1 шт. (кг)"],
        "description": ["описание"],
        "received": ["получено"],
        "manual": ["факт вручную"],
        "actual": ["факт", "кол-во фактически"],
        "comment": ["комментарий"],
        "checked": ["проверено"],
        "recognition": ["распознавание"],
    }

    def find(names: list[str]) -> Optional[int]:
        for name in names:
            col = headers.get(name.lower())
            if col:
                return col
        return None

    cols = {key: find(names) for key, names in aliases.items()}
    if not cols["sku"] or not cols["qty"]:
        raise ValueError("Файл не похож на мастер-каталог: нет колонок «Артикул» и «По документу».")

    row_images: dict[int, str] = {}
    image_dir.mkdir(parents=True, exist_ok=True)
    for index, image in enumerate(sheet._images, start=1):
        try:
            row = int(image.anchor._from.row) + 1
            extension = str(getattr(image, "format", "png") or "png").lower()
            target = image_dir / f"master_row_{row}_{index}.{extension}"
            target.write_bytes(image._data())
            row_images.setdefault(row, str(target))
        except Exception:
            continue

    products: list[Product] = []
    for row in range(2, sheet.max_row + 1):
        sku = _text(sheet.cell(row, cols["sku"]).value)
        qty = _int(sheet.cell(row, cols["qty"]).value)
        if not sku or qty <= 0:
            continue
        received_value = sheet.cell(row, cols["received"]).value if cols["received"] else None
        actual_value = sheet.cell(row, cols["actual"]).value if cols["actual"] else None
        manual_value = sheet.cell(row, cols["manual"]).value if cols["manual"] else None
        actual = _int(actual_value, -1)
        manual = _int(manual_value, -1)
        received = _yes(received_value) or actual > 0 or manual > 0
        actual_total = actual if actual >= 0 else (manual if manual >= 0 else (qty if received else 0))
        products.append(
            Product(
                number=_int(sheet.cell(row, cols["number"]).value, row - 1) if cols["number"] else row - 1,
                boxes=_text(sheet.cell(row, cols["boxes"]).value) if cols["boxes"] else "",
                sku=sku,
                qty_document=qty,
                description=_text(sheet.cell(row, cols["description"]).value) if cols["description"] else "",
                category=_text(sheet.cell(row, cols["category"]).value) if cols["category"] else "",
                material=_text(sheet.cell(row, cols["material"]).value) if cols["material"] else "",
                stone=normalize_stone_names(
                    _text(sheet.cell(row, cols["stone"]).value) if cols["stone"] else ""
                ),
                color=_text(sheet.cell(row, cols["color"]).value) if cols["color"] else "",
                unit_weight_kg=_float(sheet.cell(row, cols["weight"]).value) if cols["weight"] else None,
                image_path=row_images.get(row, ""),
                received=actual_total > 0,
                actual_manual=actual_total if actual_total > 0 and actual_total != qty else None,
                comment=_text(sheet.cell(row, cols["comment"]).value) if cols["comment"] else "",
                checked=_yes(sheet.cell(row, cols["checked"]).value) if cols["checked"] else False,
                recognition=_text(sheet.cell(row, cols["recognition"]).value) if cols["recognition"] else "",
            )
        )
    return _merge_products(products)


def _column_index(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref).group(0)
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - 64
    return value - 1


def parse_supplier_xlsx(path: Path, image_dir: Path) -> list[Product]:
    ns_main = f"{{{MAIN_NS}}}"
    products: list[Product] = []
    image_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(f"{ns_main}si"):
                shared.append("".join(node.text or "" for node in item.iter(f"{ns_main}t")))

        sheet = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
        values_by_row: dict[int, list[Any]] = {}
        for row_node in sheet.findall(f".//{ns_main}row"):
            row_index = int(row_node.attrib["r"]) - 1
            values: list[Any] = [None] * 13
            for cell in row_node.findall(f"{ns_main}c"):
                column = _column_index(cell.attrib["r"])
                if column >= 13:
                    continue
                cell_type = cell.attrib.get("t")
                value_node = cell.find(f"{ns_main}v")
                inline = cell.find(f"{ns_main}is")
                value: Any = None
                if cell_type == "s" and value_node is not None:
                    value = shared[int(value_node.text)]
                elif cell_type == "inlineStr" and inline is not None:
                    value = "".join(node.text or "" for node in inline.iter(f"{ns_main}t"))
                elif value_node is not None:
                    raw = value_node.text
                    try:
                        numeric = float(raw)
                        value = int(numeric) if numeric.is_integer() else numeric
                    except ValueError:
                        value = raw
                values[column] = value
            values_by_row[row_index] = values

        row_images: dict[int, str] = {}
        rel_path = "xl/drawings/_rels/drawing1.xml.rels"
        drawing_path = "xl/drawings/drawing1.xml"
        if rel_path in archive.namelist() and drawing_path in archive.namelist():
            rel_root = ET.fromstring(archive.read(rel_path))
            rels = {node.attrib["Id"]: node.attrib["Target"].replace("../", "xl/") for node in rel_root}
            drawing = ET.fromstring(archive.read(drawing_path))
            ns = {"xdr": XDR_NS, "a": A_NS}
            for index, anchor in enumerate(list(drawing), start=1):
                from_node = anchor.find("xdr:from", ns)
                picture = anchor.find("xdr:pic", ns)
                if from_node is None or picture is None:
                    continue
                row_index = int(from_node.find("xdr:row", ns).text)
                blip = picture.find(".//a:blip", ns)
                relationship = blip.attrib[f"{{{REL_NS}}}embed"]
                media = rels[relationship]
                target = image_dir / f"supplier_row_{row_index + 1}_{index}{Path(media).suffix.lower()}"
                target.write_bytes(archive.read(media))
                row_images[row_index] = str(target)

        generated = 130
        for row_index in sorted(values_by_row):
            row = values_by_row[row_index]
            line, boxes, sku_value, qty_value, description_value, weight_value = (
                row[1], row[2], row[4], row[5], row[7], row[12]
            )
            qty = _int(qty_value)
            if qty <= 0 or not (sku_value or description_value):
                continue
            sku = _text(sku_value)
            description = _text(description_value)
            if sku.upper().startswith("CBOX") or "inset card" in sku.lower() or "paper box" in description.lower():
                continue
            if not sku:
                lower = description.lower()
                if "metal tape" in lower:
                    sku = "NO-SKU-METAL-TAPE"
                elif "plastic" in lower:
                    sku = "NO-SKU-PLASTIC-ACCESSORIES"
                elif "earring" in lower:
                    sku = "NO-SKU-SILVER-EARRINGS"
                elif "necklace" in lower:
                    sku = "NO-SKU-SILVER-NECKLACE"
                elif "accessories" in lower:
                    sku = "NO-SKU-SILVER-ACCESSORIES"
                else:
                    sku = f"NO-SKU-{generated}"
                    generated += 1
            stone, color, recognition = detect_stone(sku, description)
            weight = _float(weight_value)
            products.append(
                Product(
                    number=_int(line, len(products) + 1),
                    boxes=_text(boxes),
                    sku=sku,
                    qty_document=qty,
                    description=normalize_description(description),
                    category=detect_category(sku, description),
                    material=detect_materials(description),
                    stone=stone,
                    color=color,
                    unit_weight_kg=round(weight / qty, 4) if weight is not None and qty else None,
                    image_path=row_images.get(row_index, ""),
                    recognition=recognition,
                )
            )
    return _merge_products(products)


def _safe_sheet_dimensions(sheet: Any) -> tuple[int, int]:
    """
    Returns worksheet dimensions even when an XLSX producer omitted the
    standard dimension metadata.

    Some valid Excel files return max_row/max_column as None in read-only
    mode. openpyxl can still calculate the real used range.
    """
    max_row = getattr(sheet, "max_row", None)
    max_column = getattr(sheet, "max_column", None)

    if not isinstance(max_row, int) or not isinstance(max_column, int):
        try:
            sheet.calculate_dimension(force=True)
            max_row = getattr(sheet, "max_row", None)
            max_column = getattr(sheet, "max_column", None)
        except Exception:
            pass

    if not isinstance(max_row, int) or max_row < 1:
        max_row = 1
    if not isinstance(max_column, int) or max_column < 1:
        # Header detection never needs more than the first 30 columns.
        # Iterating the first row is a final fallback for unusual XLSX files.
        try:
            first_row = next(
                sheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                (),
            )
            max_column = max(1, len(first_row))
        except Exception:
            max_column = 30

    return max_row, max_column


def load_products(path: Path, image_dir: Path) -> list[Product]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
    )
    try:
        sheet = workbook[workbook.sheetnames[0]]
        _, max_column = _safe_sheet_dimensions(sheet)
        headers = {
            _text(sheet.cell(1, column).value).lower()
            for column in range(1, min(max_column, 30) + 1)
        }
    finally:
        workbook.close()

    if "артикул" in headers and (
        "по документу" in headers
        or "кол-во по документу" in headers
        or "количество по документу" in headers
    ):
        return parse_master_xlsx(path, image_dir)
    return parse_supplier_xlsx(path, image_dir)


def _excel_option_width(values: list[str], minimum: int = 12, maximum: int = 28) -> int:
    longest = 0
    for value in values:
        for part in re.split(r"\s*(?:;|,|\||\+)\s*", str(value or "")):
            longest = max(longest, len(part.strip()))
    return max(minimum, min(maximum, longest + 4))


def export_master(path: Path, products: list[Product]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Мастер-каталог"
    headers = [
        "№", "Коробки", "Фото", "Артикул", "По документу", "Категория",
        "Материал", "Камень", "Цвет", "Вес 1 шт.", "Описание", "Получено",
        "Факт вручную", "Факт", "Расхождение", "Статус", "Комментарий",
        "Проверено", "Распознавание", "Название", "Серебряная категория",
        "Серебро 925", "Покрытие", "Размер", "Единица учёта", "Продаётся отдельно",
        "Оригинальное название", "Вес партии, г", "Серебро RMB/г", "Работа RMB/г",
        "Цена RMB/г", "Сумма RMB", "Курс USD/RMB", "CIF, %", "Закупка USD/ед.",
        "Продажа USD при импорте", "Курс USD/VND при импорте",
        "Коэффициент при импорте", "Продажа VND при импорте",
    ]
    sheet.append(headers)
    dark = PatternFill("solid", fgColor="1F4E78")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = dark
        cell.font = white
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [
        6,
        max(14, min(22, max([len(str(p.boxes or "")) for p in products] + [8]) + 3)),
        20,
        max(18, min(30, max([len(p.sku) for p in products] + [8]) + 3)),
        15,
        _excel_option_width([p.category for p in products], 14, 24),
        _excel_option_width([p.material for p in products], 14, 26),
        _excel_option_width([p.stone for p in products], 14, 26),
        _excel_option_width([p.color for p in products], 12, 22),
        14,
        34,
        12,
        14,
        12,
        12,
        20,
        28,
        12,
        20, 30, 24, 14, 14, 16, 16, 18, 16, 28, 16, 16, 16, 16, 18, 16, 12, 18, 20, 22, 20, 22,
    ]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "D2"

    for row_index, product in enumerate(products, start=2):
        actual = product.actual_qty
        values = [
            product.number, product.boxes, "", product.sku, product.qty_document,
            product.category, product.material, product.stone, product.color,
            product.unit_weight_kg, product.description, "Да" if product.received else "Нет",
            product.actual_manual, actual, product.variance, product.status, product.comment,
            "Да" if product.checked else "Нет", product.recognition,
            product.name, product.silver_category, "Да" if product.silver_925 else "Нет",
            product.plating, product.size, product.unit_label,
            "Да" if product.sellable else "Нет", product.original_name,
            product.total_weight_g, product.silver_rmb_per_g, product.labour_rmb_per_g,
            product.price_rmb_per_g, product.amount_rmb, product.usd_rmb_rate,
            product.cif_percent, product.purchase_usd_per_unit, product.invoice_sale_usd,
            product.invoice_usd_vnd_rate, product.invoice_coefficient, product.invoice_sale_vnd,
        ]
        sheet.append(values)
        sheet.row_dimensions[row_index].height = 92
        for cell in sheet[row_index]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        if product.image_path and Path(product.image_path).exists():
            try:
                image = XLImage(product.image_path)
                image.width = 105
                image.height = 88
                sheet.add_image(image, f"C{row_index}")
            except Exception:
                pass
    workbook.save(path)
