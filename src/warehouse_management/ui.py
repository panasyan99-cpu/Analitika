from __future__ import annotations

from dataclasses import asdict, replace
import base64
import json
from html import escape
from io import BytesIO
from pathlib import Path
import shutil
from datetime import datetime, timedelta
import tempfile
from typing import Any, Iterable
from urllib.parse import urljoin
import uuid

import pandas as pd
import streamlit as st

from src.auth import can_write
from openpyxl import load_workbook
from PIL import Image, ImageOps

from .client import WarehouseClient, WarehouseClientError, as_int, link_ids, select_text
from .models import Product, SupplySummary
from .packing import CATEGORIES, export_master, load_products
from .service import WarehouseService, WarehouseServiceError
from .schema import BaserowSchemaManager, WarehouseSchemaError, SUPPLY_LINES_TABLE_NAME
from .silver import (
    SILVER_CATEGORIES,
    SILVER_DEFAULT_COEFFICIENT,
    SILVER_DEFAULT_USD_VND,
    is_silver_invoice,
    parse_silver_invoice,
    refresh_calculated_silver_prices,
    silver_sale_vnd,
)


WORKSPACES = (
    "Главная",
    "Товары",
    "Поставки",
    "Передача",
    "История",
)

SUPPLY_WORKSPACES = ("Реестр", "Новая поставка", "Приёмка")
HISTORY_WORKSPACES = ("Операции",)

WAREHOUSE_PHOTO_SIZE = 320
WAREHOUSE_TABLE_ROW_HEIGHT = 138
WAREHOUSE_PAGE_SIZE_OPTIONS = (10, 20, 30)

WAREHOUSE_MANAGEMENT_CSS = """
<style>
.wm-shell {border:1px solid rgba(183,137,63,.22);border-radius:20px;padding:18px 20px;
 background:linear-gradient(135deg,#fffdf8 0%,#f8f1e4 58%,#fff 100%);margin:.25rem 0 1rem;box-shadow:0 10px 30px rgba(70,51,25,.05)}
.wm-kicker {text-transform:uppercase;letter-spacing:.13em;font-size:11px;font-weight:800;color:#a47732;margin-bottom:5px}
.wm-shell-title {font-family:Georgia,serif;font-size:31px;line-height:1.08;color:#19140e;margin:0 0 5px}
.wm-shell-copy {color:#6f6253;font-size:14px;max-width:850px}
.wm-page-head {margin:.2rem 0 1rem}
.wm-page-head h2 {font-family:Georgia,serif;font-size:28px;line-height:1.12;color:#19140e;margin:0 0 4px}
.wm-page-head p {margin:0;color:#786b5d;font-size:14px}
.wm-context {border:1px solid rgba(183,137,63,.24);border-radius:13px;padding:11px 14px;background:#fffaf1;margin:.35rem 0 1rem;color:#5e5140}
.wm-good {border-left:4px solid #3a7d51;background:#f4fbf6;padding:11px 13px;border-radius:9px}
.wm-warning {border-left:4px solid #b7893f;background:#fffaf1;padding:11px 13px;border-radius:9px}
.wm-danger {border-left:4px solid #aa3939;background:#fff6f6;padding:11px 13px;border-radius:9px}
.wm-photo-placeholder {min-height:178px;border:1px dashed rgba(183,137,63,.35);border-radius:14px;display:flex;align-items:center;justify-content:center;background:#fbfaf7;color:#8a8175}
.wm-product-card {border:1px solid rgba(183,137,63,.20);border-radius:14px;padding:11px 12px;background:#fff;margin:-2px 0 8px;min-height:116px}
.wm-product-card .sku {font-family:Georgia,serif;font-size:19px;font-weight:700;color:#171411;margin-bottom:5px}
.wm-product-card .meta {font-size:12px;color:#71685c;min-height:34px;line-height:1.35}
.wm-product-card .stock {font-size:13px;font-weight:800;margin-top:7px}
.wm-silver-price {margin:8px 0;padding:8px 9px;border-radius:10px;background:#fff8e8;border:1px solid rgba(183,137,63,.32);font-size:13px;line-height:1.5;color:#46351f}
.wm-silver-price span {color:#7c6a53;font-size:12px}
.wm-stock-ok {color:#315d43}.wm-stock-low {color:#9b651c}.wm-stock-zero {color:#9c3535}
.wm-stepper {display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin:.4rem 0 1rem}
.wm-step {border:1px solid #e8dfd0;border-radius:12px;padding:9px 10px;background:#fff;color:#796d5f;font-size:12px}
.wm-step b {display:block;color:#2a241d;font-size:13px;margin-bottom:2px}
.wm-step.active {border-color:#b7893f;background:#fff9ed;box-shadow:inset 0 0 0 1px rgba(183,137,63,.16)}
.wm-supply-card {border:1px solid rgba(183,137,63,.21);border-radius:15px;padding:13px 14px;background:#fff;margin-bottom:8px}
.wm-supply-card .title {font-family:Georgia,serif;font-weight:700;font-size:19px;color:#211a12}
.wm-supply-card .sub {color:#786d60;font-size:12px;margin:3px 0 10px}
.wm-status {display:inline-block;border-radius:999px;padding:3px 8px;font-size:11px;font-weight:800;margin-top:5px}
.wm-status-good {background:#eaf6ee;color:#2f6744}.wm-status-warn {background:#fff2d9;color:#8a5a18}.wm-status-neutral {background:#f1efeb;color:#6c6258}
.wm-empty {border:1px dashed #d9cfbf;border-radius:14px;padding:24px;text-align:center;color:#756a5e;background:#fcfbf8}
.wm-toolbar-note {font-size:12px;color:#84786b;padding-top:8px}
.wm-edit-hint {border:1px solid rgba(183,137,63,.42);background:#fff8e8;border-radius:12px;padding:10px 12px;margin:.4rem 0 .8rem;color:#664919;font-weight:700}
.wm-row-meta {font-size:13px;color:#75695b;line-height:1.45}
.wm-row-sku {font-family:Georgia,serif;font-size:20px;font-weight:800;color:#211a12;margin-bottom:5px}
div[data-testid="stNumberInput"] input {background:#fff8df !important;border:2px solid #c58b2b !important;border-radius:10px !important;font-weight:800 !important;font-size:18px !important;color:#2b2115 !important;}
div[data-testid="stNumberInput"] label p {font-weight:800 !important;color:#7b5418 !important;}
@media (max-width:900px){.wm-stepper{grid-template-columns:repeat(2,minmax(0,1fr))}.wm-shell-title{font-size:27px}}
@media (max-width:640px){.wm-shell{padding:14px}.wm-shell-title{font-size:24px}.wm-page-head h2{font-size:24px}.wm-photo-placeholder{min-height:135px}.wm-stepper{grid-template-columns:1fr}}
</style>
"""


def _clear_cache(*, photos: bool = False) -> None:
    """Clear only warehouse caches; do not invalidate analysis/SONU/order caches."""
    try:
        from src.warehouse import fetch_image_bytes, fetch_table_rows

        fetch_table_rows.clear()
        if photos:
            fetch_image_bytes.clear()
    except Exception:
        pass
    try:
        _remote_thumbnail_data_uri.clear()
    except Exception:
        pass



_PENDING_WIDGET_STATE_KEY = "_warehouse_pending_widget_state"


def _queue_widget_state(*, state: Any | None = None, **updates: Any) -> None:
    """Queue widget values for the next rerun.

    Streamlit forbids changing a widget-owned session-state key after that
    widget has already been instantiated in the current run.  Warehouse
    navigation buttons can be rendered below the primary navigation widget,
    so they must write to a neutral pending key and apply the values at the
    very beginning of the next rerun.
    """
    target = st.session_state if state is None else state
    pending = dict(target.get(_PENDING_WIDGET_STATE_KEY, {}) or {})
    pending.update(updates)
    target[_PENDING_WIDGET_STATE_KEY] = pending


def _apply_pending_widget_state(*, state: Any | None = None) -> None:
    target = st.session_state if state is None else state
    pending = dict(target.pop(_PENDING_WIDGET_STATE_KEY, {}) or {})
    allowed = {
        "warehouse_workspace": set(WORKSPACES),
        "warehouse_supply_workspace": set(SUPPLY_WORKSPACES),
        "warehouse_history_workspace": set(HISTORY_WORKSPACES),
        "warehouse_catalog_mode": {"Каталог", "Управление"},
        "warehouse_catalog_action": {"Добавить", "Редактировать", "Удалить / деактивировать"},
    }
    for key, value in pending.items():
        valid_values = allowed.get(key)
        if valid_values is not None and value not in valid_values:
            continue
        target[key] = value


def _resolved_config(config: Any) -> Any:
    current_id = int(getattr(config, "supply_lines_table_id", 0) or 0)
    runtime_id = int(st.session_state.get("warehouse_supply_lines_table_id", 0) or 0)
    if current_id or runtime_id:
        return replace(config, supply_lines_table_id=current_id or runtime_id)
    try:
        discovered = WarehouseClient(config).discover_table_id(SUPPLY_LINES_TABLE_NAME)
    except WarehouseClientError:
        discovered = 0
    if discovered:
        st.session_state["warehouse_supply_lines_table_id"] = discovered
        return replace(config, supply_lines_table_id=discovered)
    return config


def _service(config: Any) -> WarehouseService:
    resolved = _resolved_config(config)
    return WarehouseService(WarehouseClient(resolved))


def _silver_price_settings() -> tuple[int, float]:
    """Warehouse-local silver pricing controls; independent from report settings."""
    st.session_state.setdefault("warehouse_silver_usd_vnd", SILVER_DEFAULT_USD_VND)
    st.session_state.setdefault("warehouse_silver_coefficient", SILVER_DEFAULT_COEFFICIENT)
    with st.container(border=True):
        st.markdown("### Цена серебра")
        st.caption(
            "Закупка в USD фиксируется из поставки. Продажа в VND пересчитывается только "
            "по текущему курсу и коэффициенту; история поставок не изменяется."
        )
        columns = st.columns([1, 1, 2])
        usd_vnd = int(
            columns[0].number_input(
                "Курс USD/VND",
                min_value=1_000,
                max_value=100_000,
                step=100,
                key="warehouse_silver_usd_vnd",
            )
        )
        coefficient = float(
            columns[1].number_input(
                "Коэффициент",
                min_value=0.1,
                max_value=100.0,
                step=0.1,
                format="%.2f",
                key="warehouse_silver_coefficient",
            )
        )
        columns[2].markdown(
            '<div class="wm-context" style="margin-top:1.6rem">'
            '<b>Продажа VND</b> = закупка USD × курс × коэффициент. '
            'Результат округляется вверх до 1 000 VND.</div>',
            unsafe_allow_html=True,
        )
    return usd_vnd, coefficient


def _ensure_silver_schema(config: Any) -> bool:
    """Create additive silver fields right before the first silver import."""
    resolved = _resolved_config(config)
    supply_lines_id = int(getattr(resolved, "supply_lines_table_id", 0) or 0)
    if not supply_lines_id:
        st.error("Сначала должна быть доступна таблица «Позиции поставок».")
        return False
    email = str(getattr(resolved, "email", "") or "").strip()
    password = str(getattr(resolved, "password", "") or "")
    if not email or not password:
        st.error("Автоматические реквизиты Baserow не настроены в приватной конфигурации.")
        return False
    try:
        manager = BaserowSchemaManager(resolved.base_url, email, password)
        created = manager.ensure_silver_fields(
            components_table_id=int(resolved.components_table_id),
            supply_lines_table_id=supply_lines_id,
            supplies_table_id=int(resolved.supplies_table_id),
        )
        st.session_state["warehouse_silver_schema_ready"] = True
        st.session_state["warehouse_silver_schema_created"] = created
        return True
    except Exception as exc:
        st.error(f"Не удалось подготовить поля серебра в Baserow: {exc}")
        return False


def _auto_prepare_safe_schema(config: Any, *, force: bool = False) -> WarehouseService | None:
    service = _service(config)
    if service.has_supply_lines:
        return service
    if not can_write():
        return None

    attempted_key = "warehouse_schema_auto_attempted"
    if st.session_state.get(attempted_key) and not force:
        return None
    st.session_state[attempted_key] = True
    resolved = service.config
    email = str(getattr(resolved, "email", "") or "").strip()
    password = str(getattr(resolved, "password", "") or "")
    if not email or not password:
        st.session_state["warehouse_schema_auto_error"] = (
            "Рабочие данные Baserow не настроены в Streamlit Secrets."
        )
        return None

    try:
        with st.spinner("Проверяем безопасную структуру склада..."):
            manager = BaserowSchemaManager(resolved.base_url, email, password)
            report = manager.ensure_and_migrate(
                database_id=int(resolved.database_id),
                souvenirs_table_id=int(resolved.souvenirs_table_id),
                components_table_id=int(resolved.components_table_id),
                operations_table_id=int(resolved.operations_table_id),
                supplies_table_id=int(resolved.supplies_table_id),
            )
        st.session_state["warehouse_supply_lines_table_id"] = int(report.table_id)
        st.session_state["warehouse_schema_report"] = report.to_dict()
        st.session_state.pop("warehouse_schema_auto_error", None)
        _clear_cache(photos=True)
        return _service(config)
    except WarehouseSchemaError as exc:
        st.session_state["warehouse_schema_auto_error"] = str(exc)
        return None


def _require_safe_schema(config: Any) -> WarehouseService | None:
    if not can_write():
        st.info("Режим просмотра: создание и изменение складских данных недоступно.")
        return None
    service = _service(config)
    if service.has_supply_lines:
        return service
    service = _auto_prepare_safe_schema(config)
    if service and service.has_supply_lines:
        return service

    message = str(
        st.session_state.get("warehouse_schema_auto_error")
        or "Безопасная таблица «Позиции поставок» пока недоступна."
    )
    st.error(message)
    if st.button("Повторить автоматическую настройку", key="warehouse_retry_schema"):
        st.session_state.pop("warehouse_schema_auto_attempted", None)
        _auto_prepare_safe_schema(config, force=True)
        st.rerun()
    return None


def _safe_action(action, *, success: str | None = None) -> Any:
    try:
        result = action()
    except (WarehouseClientError, WarehouseServiceError, ValueError, OSError) as exc:
        st.error(str(exc))
        return None
    _clear_cache()
    if success:
        st.success(success)
    # Preserve a truthy success signal for service methods that intentionally
    # return None after completing a write operation.
    return True if result is None else result


def _summary_options(summaries: list[SupplySummary]) -> tuple[list[str], dict[str, SupplySummary]]:
    mapping = {
        f"{item.supply_id} · {item.status or 'Без статуса'} · принято {item.qty_received:,} · ожидается {item.qty_waiting:,}": item
        for item in summaries
    }
    return list(mapping), mapping


def _photo_url(value: Any, base_url: str, *, size: str = "small") -> str:
    """Return a Baserow image URL, preferring a lightweight thumbnail."""
    files = value if isinstance(value, list) else [value] if value else []
    if not files or not isinstance(files[0], dict):
        return ""
    item = files[0]
    thumbnails = item.get("thumbnails") or {}
    orders = {
        "small": ("small", "card_cover", "tiny", "large"),
        "large": ("large", "card_cover", "small", "tiny"),
    }
    for name in orders.get(size, orders["small"]):
        candidate = thumbnails.get(name)
        if isinstance(candidate, dict) and candidate.get("url"):
            return urljoin(str(base_url).rstrip("/") + "/", str(candidate["url"]).strip())
    url = str(item.get("url") or "").strip()
    return urljoin(str(base_url).rstrip("/") + "/", url) if url else ""


def _item_photo_url(item: Any, config: Any, *, size: str = "small") -> str:
    return _photo_url(getattr(item, "photo", None), str(config.base_url), size=size)


def _row_photo_url(row: dict[str, Any], config: Any, *, size: str = "small") -> str:
    return _photo_url(row.get("Фото"), str(config.base_url), size=size)


def _local_thumbnail_data_uri(path_value: str, modified_ns: int = 0) -> str:
    del modified_ns
    path = Path(str(path_value or ""))
    if not path.exists():
        return ""
    try:
        with Image.open(path) as source:
            image = ImageOps.exif_transpose(source).convert("RGB")
            image.thumbnail((WAREHOUSE_PHOTO_SIZE, WAREHOUSE_PHOTO_SIZE), getattr(Image, "Resampling", Image).LANCZOS)
            canvas = Image.new("RGB", (WAREHOUSE_PHOTO_SIZE + 12, WAREHOUSE_PHOTO_SIZE + 12), "white")
            canvas.paste(image, (((WAREHOUSE_PHOTO_SIZE + 12) - image.width) // 2, ((WAREHOUSE_PHOTO_SIZE + 12) - image.height) // 2))
            output = BytesIO()
            canvas.save(output, format="JPEG", quality=76, optimize=True)
        encoded = base64.b64encode(output.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{encoded}"
    except Exception:
        return ""


@st.cache_data(ttl=1800, show_spinner=False, max_entries=2048)
def _remote_thumbnail_data_uri(url: str, token: str) -> str:
    """Fetch private Baserow media server-side and expose a compact browser-safe thumbnail."""
    if not url:
        return ""
    from src.warehouse import fetch_image_bytes

    raw = fetch_image_bytes(url, token)
    if not raw:
        return ""
    try:
        with Image.open(BytesIO(raw)) as source:
            image = ImageOps.exif_transpose(source).convert("RGB")
            image.thumbnail((WAREHOUSE_PHOTO_SIZE, WAREHOUSE_PHOTO_SIZE), getattr(Image, "Resampling", Image).LANCZOS)
            canvas = Image.new("RGB", (WAREHOUSE_PHOTO_SIZE + 12, WAREHOUSE_PHOTO_SIZE + 12), "white")
            canvas.paste(image, (((WAREHOUSE_PHOTO_SIZE + 12) - image.width) // 2, ((WAREHOUSE_PHOTO_SIZE + 12) - image.height) // 2))
            output = BytesIO()
            canvas.save(output, format="JPEG", quality=78, optimize=True)
        encoded = base64.b64encode(output.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{encoded}"
    except Exception:
        return ""


def _item_photo_data_uri(item: Any, config: Any) -> str:
    return _remote_thumbnail_data_uri(_item_photo_url(item, config, size="large"), str(config.token))


def _row_photo_data_uri(row: dict[str, Any], config: Any) -> str:
    return _remote_thumbnail_data_uri(_row_photo_url(row, config, size="large"), str(config.token))


def _format_vnd(value: int | float | None) -> str:
    return f"{int(value or 0):,}".replace(",", " ") + " VND"


def _catalog_dataframe(items: list[Any], config: Any) -> pd.DataFrame:
    usd_vnd = int(st.session_state.get("warehouse_silver_usd_vnd", SILVER_DEFAULT_USD_VND))
    coefficient = float(st.session_state.get("warehouse_silver_coefficient", SILVER_DEFAULT_COEFFICIENT))
    rows: list[dict[str, Any]] = []
    for item in items:
        purchase = item.purchase_usd_per_unit if item.silver_925 else None
        sale = silver_sale_vnd(purchase, usd_vnd, coefficient) if item.silver_925 else None
        rows.append(
            {
                "Фото": _item_photo_data_uri(item, config),
                "Артикул": item.sku,
                "Название": item.name,
                "Раздел": "Серебро 925" if item.silver_925 else item.section,
                "Группа": item.silver_category if item.silver_925 else item.category,
                "Покрытие": item.plating,
                "Размер": item.size,
                "Единица": item.unit_label,
                "Остаток": item.balance,
                "Минимум": item.min_balance,
                "Закупка USD": purchase,
                "Продажа VND": sale,
                "Назначение": "Отдельная продажа" if item.sellable else "Для производства" if item.silver_925 else "",
                "Материал": item.material,
                "Камень": item.stone,
                "Цвет": item.color,
                "Коробки": item.boxes,
                "row_id": item.row_id,
            }
        )
    return pd.DataFrame(rows)

def _render_item_photo(item: Any, config: Any, *, width: int | str = "stretch") -> None:
    data_uri = _item_photo_data_uri(item, config)
    if data_uri:
        st.image(data_uri, width=width)
    else:
        st.markdown('<div class="wm-photo-placeholder">Нет фотографии</div>', unsafe_allow_html=True)



def _page_slice(records: list[Any], *, key: str, default_size: int = 10) -> list[Any]:
    """Return only the visible page so photos are fetched lazily."""
    if not records:
        return []
    controls = st.columns([1, 1, 4])
    size_options = list(WAREHOUSE_PAGE_SIZE_OPTIONS)
    default_index = size_options.index(default_size) if default_size in size_options else 0
    page_size = int(
        controls[0].selectbox(
            "Строк на странице",
            size_options,
            index=default_index,
            key=f"{key}_page_size",
        )
    )
    page_count = max(1, (len(records) + page_size - 1) // page_size)
    page_key = f"{key}_page"
    current_state = as_int(st.session_state.get(page_key), 1)
    if current_state < 1 or current_state > page_count:
        st.session_state[page_key] = 1
    page = int(
        controls[1].selectbox(
            "Страница",
            list(range(1, page_count + 1)),
            format_func=lambda value: f"{value} из {page_count}",
            key=page_key,
        )
    )
    controls[2].caption(
        f"Показано {(page - 1) * page_size + 1}–{min(page * page_size, len(records))} "
        f"из {len(records)}. Фото загружаются только для этой страницы."
    )
    start = (page - 1) * page_size
    return records[start : start + page_size]


def _record_photo_data_uri(record: dict[str, Any], config: Any) -> str:
    if record.get("item") is not None:
        return _item_photo_data_uri(record["item"], config)
    if record.get("row") is not None:
        return _row_photo_data_uri(record["row"], config)
    return str(record.get("photo") or "")


def _render_quantity_editor(
    records: list[dict[str, Any]],
    config: Any,
    *,
    draft_key: str,
    page_key: str,
    revision: int,
    quantity_label: str,
    minimum: int = 0,
) -> dict[str, int]:
    """Render large-photo rows with a strict per-row quantity maximum."""
    existing = dict(st.session_state.get(draft_key, {}) or {})
    draft: dict[str, int] = {}
    for record in records:
        record_key = str(record["id"])
        maximum_raw = record.get("maximum")
        maximum = as_int(maximum_raw) if maximum_raw is not None else None
        initial = as_int(record.get("initial"), minimum)
        value = as_int(existing.get(record_key, initial), initial)
        value = max(value, minimum)
        if maximum is not None:
            value = min(value, maximum)
        draft[record_key] = value
    st.session_state[draft_key] = draft

    st.markdown(
        '<div class="wm-edit-hint">Золотые поля справа редактируются. '
        'Сайт физически не позволит указать количество выше доступного максимума.</div>',
        unsafe_allow_html=True,
    )
    visible = _page_slice(records, key=page_key, default_size=10)
    for record in visible:
        record_key = str(record["id"])
        maximum_raw = record.get("maximum")
        maximum = as_int(maximum_raw) if maximum_raw is not None else None
        with st.container(border=True):
            photo_col, info_col, qty_col = st.columns([1.15, 2.55, 1.35])
            with photo_col:
                data_uri = _record_photo_data_uri(record, config)
                if data_uri:
                    st.image(data_uri, width=190)
                else:
                    st.markdown(
                        '<div class="wm-photo-placeholder">Нет фотографии</div>',
                        unsafe_allow_html=True,
                    )
            with info_col:
                st.markdown(
                    f'<div class="wm-row-sku">{escape(str(record.get("sku") or ""))}</div>',
                    unsafe_allow_html=True,
                )
                meta = record.get("meta") or []
                st.markdown(
                    '<div class="wm-row-meta">'
                    + "<br>".join(escape(str(value)) for value in meta if str(value).strip())
                    + "</div>",
                    unsafe_allow_html=True,
                )
            with qty_col:
                if maximum is not None:
                    st.caption(f"Доступный максимум: {maximum:,} шт.")
                widget_key = f"{draft_key}_qty_{record_key}_{revision}"
                if widget_key in st.session_state:
                    current = as_int(st.session_state[widget_key], draft[record_key])
                    current = max(current, minimum)
                    if maximum is not None:
                        current = min(current, maximum)
                    st.session_state[widget_key] = current
                kwargs: dict[str, Any] = {
                    "label": quantity_label,
                    "min_value": int(minimum),
                    "value": int(draft[record_key]),
                    "step": 1,
                    "key": widget_key,
                    "help": "Это итоговое количество по данной позиции.",
                }
                if maximum is not None:
                    kwargs["max_value"] = int(maximum)
                value = int(st.number_input(**kwargs))
                draft[record_key] = value
                st.session_state[draft_key] = dict(draft)
    return draft


def _parse_optional_weight(value: Any) -> float | None:
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        result = float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError("Введите вес числом, например 125,40.") from exc
    if result < 0:
        raise ValueError("Вес не может быть отрицательным.")
    return result


def _render_weight_receiving_editor(
    records: list[dict[str, Any]],
    config: Any,
    *,
    draft_key: str,
    page_key: str,
    revision: int,
) -> dict[int, dict[str, Any]]:
    """Render clean-weight inputs and return confirmed current quantities."""
    existing = dict(st.session_state.get(draft_key, {}) or {})
    st.markdown(
        '<div class="wm-warning"><b>Режим для уже пришедшего товара.</b><br>'
        'По каждой позиции введите чистый остаток без пакета и тары. Сайт примет полное '
        'количество по инвойсу, а разницу оформит расходом «Использовано до постановки на учёт».</div>',
        unsafe_allow_html=True,
    )
    visible = _page_slice(records, key=page_key, default_size=10)
    for record in visible:
        record_key = str(record["id"])
        row = record.get("row") or {}
        document = as_int(record.get("maximum"))
        unit_weight_g = float(row.get("_unit_weight_g") or 0.0)
        unit_label = str(row.get("_unit_label") or "шт.")
        entry = dict(existing.get(record_key, {}) or {})
        with st.container(border=True):
            photo_col, info_col, weight_col = st.columns([1.05, 2.35, 1.65])
            with photo_col:
                data_uri = _record_photo_data_uri(record, config)
                if data_uri:
                    st.image(data_uri, width=180)
                else:
                    st.markdown(
                        '<div class="wm-photo-placeholder">Нет фотографии</div>',
                        unsafe_allow_html=True,
                    )
            with info_col:
                st.markdown(
                    f'<div class="wm-row-sku">{escape(str(record.get("sku") or ""))}</div>',
                    unsafe_allow_html=True,
                )
                meta = list(record.get("meta") or [])
                meta.append(f"Средний вес: {unit_weight_g:.6f} г/{unit_label}")
                st.markdown(
                    '<div class="wm-row-meta">'
                    + "<br>".join(escape(str(value)) for value in meta if str(value).strip())
                    + "</div>",
                    unsafe_allow_html=True,
                )
            with weight_col:
                weight_key = f"{draft_key}_weight_{record_key}_{revision}"
                weight_text = st.text_input(
                    "Чистый остаток, г",
                    value=str(entry.get("weight_text") or ""),
                    key=weight_key,
                    placeholder="Например: 125,40",
                    help="Вес товара без пакета, коробки, бирки и другой тары.",
                )
                entry["weight_text"] = weight_text
                try:
                    weight_g = _parse_optional_weight(weight_text)
                except ValueError as exc:
                    st.error(str(exc))
                    weight_g = None
                if weight_g is None:
                    st.caption("Введите вес — количество рассчитается автоматически.")
                    entry.pop("quantity", None)
                elif unit_weight_g <= 0:
                    st.error("В поставке нет среднего веса единицы. Используйте приёмку по количеству.")
                    entry.pop("quantity", None)
                else:
                    raw_quantity = weight_g / unit_weight_g
                    estimated = WarehouseService.estimate_quantity_from_weight(
                        weight_g,
                        unit_weight_g,
                        maximum=document,
                    )
                    if raw_quantity > document + 0.5:
                        st.warning(
                            f"По весу получается около {raw_quantity:,.1f}, но по документу максимум {document:,}."
                        )
                    st.metric("Расчётный остаток", f"{estimated:,} {unit_label}")
                    expected_weight = estimated * unit_weight_g
                    st.caption(
                        f"Округление: {raw_quantity:,.2f} → {estimated:,}; "
                        f"отклонение {abs(weight_g - expected_weight):.4f} г."
                    )
                    override_key = f"{draft_key}_override_{record_key}_{revision}"
                    override = st.checkbox(
                        "Исправить количество вручную",
                        value=bool(entry.get("manual_override", False)),
                        key=override_key,
                    )
                    entry["manual_override"] = override
                    final_quantity = estimated
                    if override:
                        manual_key = f"{draft_key}_manual_{record_key}_{revision}"
                        manual_value = min(
                            max(as_int(entry.get("manual_qty"), estimated), 0),
                            document,
                        )
                        final_quantity = int(
                            st.number_input(
                                "Подтверждённый остаток",
                                min_value=0,
                                max_value=int(document),
                                value=int(manual_value),
                                step=1,
                                key=manual_key,
                            )
                        )
                        entry["manual_qty"] = final_quantity
                    entry["quantity"] = final_quantity
                existing[record_key] = entry

    # Keep values from other pages and rebuild the complete submit payload.
    st.session_state[draft_key] = existing
    measurements: dict[int, dict[str, Any]] = {}
    for record in records:
        record_key = str(record["id"])
        entry = dict(existing.get(record_key, {}) or {})
        try:
            weight_g = _parse_optional_weight(entry.get("weight_text"))
        except ValueError:
            continue
        if weight_g is None:
            continue
        row = record.get("row") or {}
        document = as_int(record.get("maximum"))
        unit_weight_g = float(row.get("_unit_weight_g") or 0.0)
        if unit_weight_g <= 0:
            continue
        estimated = WarehouseService.estimate_quantity_from_weight(
            weight_g,
            unit_weight_g,
            maximum=document,
        )
        final_quantity = (
            min(max(as_int(entry.get("manual_qty"), estimated), 0), document)
            if bool(entry.get("manual_override"))
            else estimated
        )
        measurements[int(record["id"])] = {
            "weight_g": weight_g,
            "quantity": final_quantity,
            "estimated": estimated,
        }
    return measurements


def _render_catalog_cards(items: list[Any], config: Any, key: str) -> None:
    if not items:
        st.markdown('<div class="wm-empty">По выбранным фильтрам позиций нет.</div>', unsafe_allow_html=True)
        return
    usd_vnd = int(st.session_state.get("warehouse_silver_usd_vnd", SILVER_DEFAULT_USD_VND))
    coefficient = float(st.session_state.get("warehouse_silver_coefficient", SILVER_DEFAULT_COEFFICIENT))
    page_size = st.segmented_control(
        "Карточек на странице", [6, 12, 18], default=12, key=f"{key}_page_size"
    ) or 12
    page_count = max(1, (len(items) + int(page_size) - 1) // int(page_size))
    page = st.selectbox(
        "Страница",
        list(range(1, page_count + 1)),
        format_func=lambda value: f"{value} из {page_count}",
        key=f"{key}_page",
    )
    current = items[(int(page) - 1) * int(page_size): int(page) * int(page_size)]
    for start in range(0, len(current), 3):
        columns = st.columns(3)
        for column, item in zip(columns, current[start:start + 3]):
            with column:
                with st.container(border=True):
                    _render_item_photo(item, config)
                    title = item.name or item.sku
                    details = " · ".join(
                        part
                        for part in (
                            item.silver_category if item.silver_925 else item.category,
                            item.plating if item.silver_925 else item.material,
                            item.size if item.silver_925 else item.stone,
                            item.color,
                        )
                        if part
                    ) or "Характеристики не указаны"
                    stock_class = "wm-stock-zero" if item.balance <= 0 else "wm-stock-low" if item.balance <= 15 else "wm-stock-ok"
                    stock_text = "Нет в наличии" if item.balance <= 0 else "Заканчивается" if item.balance <= 15 else "В наличии"
                    silver_html = ""
                    if item.silver_925:
                        purchase = float(item.purchase_usd_per_unit or 0.0)
                        sale = silver_sale_vnd(purchase, usd_vnd, coefficient)
                        purpose = "Продаётся отдельно" if item.sellable else "Для производства"
                        silver_html = (
                            f'<div class="wm-silver-price"><b>Закупка:</b> ${purchase:,.6f} / {escape(item.unit_label or "шт.")}<br>'
                            f'<b>Продажа:</b> {_format_vnd(sale)}<br><span>{escape(purpose)}</span></div>'
                        )
                    st.markdown(
                        '<div class="wm-product-card">'
                        f'<div class="sku">{escape(title)}</div>'
                        f'<div class="meta">{escape(item.sku)} · {escape(details)}</div>'
                        f'{silver_html}'
                        f'<div class="stock {stock_class}">{stock_text} · {int(item.balance):,} {escape(item.unit_label or "шт.")} · минимум {int(item.min_balance):,}</div>'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    if can_write() and st.button("Открыть карточку", key=f"{key}_open_{item.row_id}", width="stretch"):
                        _queue_widget_state(
                            warehouse_catalog_mode="Управление",
                            warehouse_catalog_action="Редактировать",
                            warehouse_catalog_selected_id=int(item.row_id),
                        )
                        st.rerun()

def _page_header(title: str, copy: str) -> None:
    st.markdown(
        '<div class="wm-page-head">'
        f'<h2>{escape(title)}</h2>'
        f'<p>{escape(copy)}</p>'
        '</div>',
        unsafe_allow_html=True,
    )


def _navigate(workspace: str, subpage: str | None = None) -> None:
    updates: dict[str, Any] = {"warehouse_workspace": workspace}
    if workspace == "Поставки" and subpage:
        updates["warehouse_supply_workspace"] = subpage
    if workspace == "История" and subpage:
        updates["warehouse_history_workspace"] = subpage
    _queue_widget_state(**updates)


def _workflow(active: int = 0) -> None:
    steps = (
        ("1", "Создать поставку", "Загрузить Master и проверить фото"),
        ("2", "Принять товар", "Указать фактическое количество"),
        ("3", "Передать", "Отправить выбранное в бухгалтерию"),
        ("4", "Проверить историю", "Контролировать остатки и операции"),
    )
    cards = []
    for index, (number, title, copy) in enumerate(steps, start=1):
        css = "wm-step active" if active == index else "wm-step"
        cards.append(f'<div class="{css}"><b>{number}. {escape(title)}</b>{escape(copy)}</div>')
    st.markdown('<div class="wm-stepper">' + ''.join(cards) + '</div>', unsafe_allow_html=True)


def _supply_status_class(status: str, waiting: int) -> str:
    value = str(status or "").casefold()
    if waiting <= 0 or "полностью" in value:
        return "wm-status-good"
    if "част" in value or waiting > 0:
        return "wm-status-warn"
    return "wm-status-neutral"


def _reset_supply_draft() -> None:
    runtime = st.session_state.pop("warehouse_supply_runtime_dir", None)
    st.session_state.pop("warehouse_supply_products", None)
    st.session_state.pop("warehouse_supply_editor", None)
    st.session_state.pop("warehouse_supply_file", None)
    st.session_state.pop("warehouse_supply_meta", None)
    if runtime:
        shutil.rmtree(str(runtime), ignore_errors=True)


def _selected_summary(summaries: list[SupplySummary], row_id: int | None) -> SupplySummary | None:
    return next((item for item in summaries if int(item.row_id) == int(row_id or 0)), None)


def render_overview(config: Any, selected_metal_groups: Iterable[str]) -> None:
    from src.warehouse import filter_warehouse_bundle, load_bundle, render_attention, render_overview

    _page_header("Главная склада", "Текущие остатки, проблемные позиции и быстрый переход к ежедневным операциям.")
    _workflow(0)
    if can_write():
        quick = st.columns(3)
        with quick[0]:
            with st.container(border=True):
                st.markdown("### Новая поставка")
                st.caption("Загрузить Packing List или Master, проверить фото и создать поставку.")
                st.button("Добавить поставку", type="primary", width="stretch", on_click=_navigate, args=("Поставки", "Новая поставка"), key="warehouse_home_new")
        with quick[1]:
            with st.container(border=True):
                st.markdown("### Приёмка")
                st.caption("Принять всё ожидаемое или указать фактическое количество по строкам.")
                st.button("Перейти к приёмке", width="stretch", on_click=_navigate, args=("Поставки", "Приёмка"), key="warehouse_home_receive")
        with quick[2]:
            with st.container(border=True):
                st.markdown("### Бухгалтерия")
                st.caption("Выбрать поставку и передать максимально доступное количество.")
                st.button("Передать товар", width="stretch", on_click=_navigate, args=("Передача",), key="warehouse_home_transfer")
    else:
        st.markdown('<div class="wm-context">Открыт режим просмотра. Остатки, поставки и история доступны без возможности изменить данные.</div>', unsafe_allow_html=True)

    st.divider()
    with st.spinner("Загружаем актуальный склад..."):
        bundle = load_bundle(config)
    selected = tuple(str(value) for value in selected_metal_groups)
    if selected:
        bundle = filter_warehouse_bundle(bundle, selected)
    render_overview(bundle)
    with st.expander("Позиции, требующие внимания", expanded=False):
        render_attention(bundle)


def render_catalog(config: Any) -> None:
    _page_header(
        "Товары",
        "Сувенирка, обычные комплектующие и серебро 925 разделены. Цена серебра пересчитывается по верхним настройкам склада.",
    )
    service = _service(config)
    section_col, mode_col = st.columns([1.4, 1])
    with section_col:
        display_section = st.segmented_control(
            "Раздел",
            ["Сувенирка", "Комплектующие", "Серебро 925"],
            default="Сувенирка",
            key="warehouse_catalog_manage_section",
        ) or "Сувенирка"
    actual_section = "Комплектующие" if display_section == "Серебро 925" else display_section
    with mode_col:
        mode_options = ["Каталог", "Управление"] if can_write() else ["Каталог"]
        if st.session_state.get("warehouse_catalog_mode") not in mode_options:
            st.session_state["warehouse_catalog_mode"] = "Каталог"
        mode = st.segmented_control(
            "Режим", mode_options, default="Каталог", key="warehouse_catalog_mode"
        ) or "Каталог"
    show_archive = st.checkbox(
        "Показать архивные карточки",
        value=False,
        key=f"warehouse_catalog_archive_{display_section}",
    )
    with st.spinner("Читаем каталог Baserow..."):
        all_items = service.catalog(actual_section, include_inactive=show_archive)
    if display_section == "Серебро 925":
        items = [item for item in all_items if item.silver_925]
    elif display_section == "Комплектующие":
        items = [item for item in all_items if not item.silver_925]
    else:
        items = all_items

    if mode == "Каталог":
        if display_section == "Серебро 925":
            filter_cols = st.columns([2.4, 1.2, 1.25, 1, 1])
        else:
            filter_cols = st.columns([2.6, 1, 1.2, 1])
        query = filter_cols[0].text_input(
            "Поиск",
            placeholder="Артикул, название, группа, покрытие или размер",
            key=f"warehouse_catalog_search_{display_section}",
        ).strip().casefold()
        status = filter_cols[1].selectbox(
            "Остаток", ["Все", "Есть", "Мало", "Нет"], key=f"warehouse_catalog_status_{display_section}"
        )
        categories = sorted(
            {
                item.silver_category if display_section == "Серебро 925" else item.category
                for item in items
                if (item.silver_category if display_section == "Серебро 925" else item.category)
            }
        )
        category = filter_cols[2].selectbox(
            "Группа" if display_section == "Серебро 925" else "Категория",
            ["Все", *categories],
            key=f"warehouse_catalog_category_{display_section}",
        )
        if display_section == "Серебро 925":
            purpose = filter_cols[3].selectbox(
                "Назначение",
                ["Все", "Для производства", "Отдельная продажа"],
                key="warehouse_catalog_silver_purpose",
            )
            view = filter_cols[4].selectbox(
                "Вид", ["Карточки", "Таблица"], key=f"warehouse_catalog_view_{display_section}"
            )
        else:
            purpose = "Все"
            view = filter_cols[3].selectbox(
                "Вид", ["Карточки", "Таблица"], key=f"warehouse_catalog_view_{display_section}"
            )
        filtered = []
        for item in items:
            if query and query not in item.search_text.casefold():
                continue
            if status == "Есть" and item.balance <= 0:
                continue
            if status == "Мало" and not (0 < item.balance <= 15):
                continue
            if status == "Нет" and item.balance > 0:
                continue
            item_category = item.silver_category if display_section == "Серебро 925" else item.category
            if category != "Все" and item_category != category:
                continue
            if purpose == "Для производства" and item.sellable:
                continue
            if purpose == "Отдельная продажа" and not item.sellable:
                continue
            filtered.append(item)
        metrics = st.columns(4)
        metrics[0].metric("Найдено SKU", len(filtered))
        metrics[1].metric("С фотографией", sum(bool(_item_photo_url(item, config)) for item in filtered))
        metrics[2].metric("Заканчиваются", sum(0 < item.balance <= 15 for item in filtered))
        metrics[3].metric("Нет в наличии", sum(item.balance <= 0 for item in filtered))
        if view == "Карточки":
            _render_catalog_cards(filtered, config, f"warehouse_catalog_cards_{display_section}")
        else:
            visible_items = _page_slice(
                filtered, key=f"warehouse_catalog_table_{display_section}", default_size=20
            )
            table = _catalog_dataframe(visible_items, config).drop(columns=["row_id"], errors="ignore")
            st.dataframe(
                table,
                width="stretch",
                hide_index=True,
                height=min(720, max(220, len(visible_items) * WAREHOUSE_TABLE_ROW_HEIGHT)),
                row_height=WAREHOUSE_TABLE_ROW_HEIGHT,
                column_config={
                    "Фото": st.column_config.ImageColumn("Фото", width="large"),
                    "Остаток": st.column_config.NumberColumn("Остаток", format="localized"),
                    "Минимум": st.column_config.NumberColumn("Минимум", format="localized"),
                    "Закупка USD": st.column_config.NumberColumn("Закупка USD", format="$ %.6f"),
                    "Продажа VND": st.column_config.NumberColumn("Продажа VND", format="localized"),
                },
            )
        return

    action = st.segmented_control(
        "Действие",
        ["Добавить", "Редактировать", "Удалить / деактивировать"],
        default="Редактировать" if items else "Добавить",
        key="warehouse_catalog_action",
    )
    if action == "Добавить":
        st.markdown(
            '<div class="wm-context">Создайте постоянную карточку товара. Остаток появится только после операции прихода.</div>',
            unsafe_allow_html=True,
        )
        with st.form(f"warehouse_add_catalog_{display_section}", clear_on_submit=True):
            c1, c2 = st.columns(2)
            sku = c1.text_input("Артикул *")
            category = c2.text_input("Категория")
            material = c1.text_input("Материал", value="Silver" if display_section == "Серебро 925" else "")
            stone = c2.text_input("Камни")
            color = c1.text_input("Цвет")
            minimum = c1.number_input("Минимальный остаток", min_value=1, value=10, step=1)
            photo = c2.file_uploader("Фото", type=["jpg", "jpeg", "png", "webp"], key=f"warehouse_add_photo_{display_section}")
            comment = st.text_area("Комментарий")
            submitted = st.form_submit_button("Создать карточку", type="primary", width="stretch")
        if submitted:
            temp_path = None
            if photo is not None:
                suffix = Path(photo.name).suffix or ".jpg"
                temp_path = Path(tempfile.gettempdir()) / f"warehouse-photo-{uuid.uuid4().hex}{suffix}"
                temp_path.write_bytes(photo.getvalue())
            result = _safe_action(
                lambda: service.add_catalog_item(
                    section=actual_section,
                    sku=sku,
                    category=category,
                    material=material,
                    stone=stone,
                    color=color,
                    boxes="",
                    minimum=int(minimum),
                    comment=comment,
                    photo_path=temp_path,
                )
            )
            if temp_path:
                temp_path.unlink(missing_ok=True)
            if result is not None:
                st.success(f"Карточка {sku} создана.")
                st.rerun()
        return

    if not items:
        st.info("Каталог пуст.")
        return
    labels = {f"{item.sku} · {item.name or item.category} · остаток {item.balance}": item for item in items}
    selected_id = int(st.session_state.get("warehouse_catalog_selected_id", 0) or 0)
    label_values = list(labels)
    default_index = next((i for i, label in enumerate(label_values) if labels[label].row_id == selected_id), 0)
    label = st.selectbox(
        "Карточка", label_values, index=default_index, key=f"warehouse_manage_item_{display_section}_{action}"
    )
    item = labels[label]
    st.session_state["warehouse_catalog_selected_id"] = int(item.row_id)
    preview, editor = st.columns([1, 2])
    with preview:
        _render_item_photo(item, config)
        st.markdown(f"**{item.name or item.sku}**")
        st.caption(f"{item.sku} · остаток {item.balance} {item.unit_label} · минимум {item.min_balance}")
        if item.silver_925:
            purchase = float(item.purchase_usd_per_unit or 0)
            sale = silver_sale_vnd(
                purchase,
                int(st.session_state.get("warehouse_silver_usd_vnd", SILVER_DEFAULT_USD_VND)),
                float(st.session_state.get("warehouse_silver_coefficient", SILVER_DEFAULT_COEFFICIENT)),
            )
            st.info(f"Закупка: ${purchase:,.6f} / {item.unit_label}\n\nПродажа: {_format_vnd(sale)}")
    with editor:
        if action == "Редактировать":
            with st.form(f"warehouse_edit_form_{display_section}_{item.row_id}"):
                c1, c2 = st.columns(2)
                category = c1.text_input("Категория", value=item.category)
                material = c2.text_input("Материал", value=item.material)
                stone = c1.text_input("Камни", value=item.stone)
                color = c2.text_input("Цвет", value=item.color)
                minimum = c2.number_input("Минимальный остаток", min_value=1, value=max(int(item.min_balance), 1))
                extra_payload: dict[str, Any] = {}
                if item.silver_925:
                    name = c1.text_input("Название", value=item.name)
                    silver_category = c2.selectbox(
                        "Группа серебра",
                        list(SILVER_CATEGORIES),
                        index=list(SILVER_CATEGORIES).index(item.silver_category) if item.silver_category in SILVER_CATEGORIES else 0,
                    )
                    plating = c1.text_input("Покрытие", value=item.plating)
                    size = c2.text_input("Размер", value=item.size)
                    unit_label = c1.text_input("Единица учёта", value=item.unit_label)
                    purchase = c2.number_input(
                        "Закупка USD/ед.", min_value=0.0, value=float(item.purchase_usd_per_unit or 0), format="%.6f"
                    )
                    sellable = c1.checkbox("Продаётся отдельно", value=item.sellable)
                    extra_payload = {
                        "Название": name,
                        "Серебряная категория": silver_category,
                        "Покрытие": plating,
                        "Размер": size,
                        "Единица учёта": unit_label,
                        "Закупка USD/ед.": purchase,
                        "Продаётся отдельно": sellable,
                        "Серебро 925": True,
                    }
                replacement_photo = st.file_uploader(
                    "Заменить фотографию",
                    type=["jpg", "jpeg", "png", "webp"],
                    key=f"warehouse_edit_photo_{display_section}_{item.row_id}",
                )
                comment = st.text_area("Комментарий", value=str((item.raw or {}).get("Комментарий") or ""))
                saved = st.form_submit_button("Сохранить изменения", type="primary", width="stretch")
            if saved:
                photo_path = None
                if replacement_photo is not None:
                    suffix = Path(replacement_photo.name).suffix or ".jpg"
                    photo_path = Path(tempfile.gettempdir()) / f"warehouse-photo-{uuid.uuid4().hex}{suffix}"
                    photo_path.write_bytes(replacement_photo.getvalue())
                payload = {
                    "Категория": category or None,
                    "Материал": material,
                    "Камень": stone,
                    "Цвет": color,
                    "Минимальный остаток": int(minimum),
                    "Комментарий": comment,
                    **extra_payload,
                }
                result = _safe_action(
                    lambda: service.update_catalog_item(actual_section, item.row_id, payload, photo_path=photo_path)
                )
                if photo_path:
                    photo_path.unlink(missing_ok=True)
                if result:
                    st.success(f"Карточка {item.sku} обновлена.")
                    st.rerun()
        else:
            st.markdown(
                '<div class="wm-warning">Карточка с операциями не удаляется физически — она деактивируется, чтобы сохранить историю.</div>',
                unsafe_allow_html=True,
            )
            confirmation = st.text_input(
                f"Для подтверждения введите артикул {item.sku}",
                key=f"warehouse_delete_confirmation_{display_section}_{item.row_id}",
            )
            if st.button(
                "Удалить или деактивировать",
                type="primary",
                disabled=confirmation.strip() != item.sku,
                key=f"warehouse_delete_button_{display_section}_{item.row_id}",
            ):
                result = _safe_action(lambda: service.deactivate_or_delete_catalog_item(actual_section, item.row_id))
                if result == "deleted":
                    st.success("Карточка удалена: по ней не было операций.")
                elif result == "deactivated":
                    st.success("Карточка деактивирована, история сохранена.")
                st.rerun()

def _supply_table(summaries: list[SupplySummary]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Поставка": item.supply_id,
                "Дата": item.date,
                "Поставщик": item.supplier,
                "Статус": item.status,
                "SKU": item.sku_total,
                "По документу": item.qty_document,
                "Принято": item.qty_received,
                "Ожидается": item.qty_waiting,
            }
            for item in summaries
        ]
    )


def render_supplies(config: Any) -> None:
    _page_header("Реестр поставок", "Быстро найдите поставку, посмотрите прогресс и откройте товары с фотографиями.")
    service = _service(config)
    with st.spinner("Загружаем реестр поставок..."):
        summaries = service.supply_summaries()
    if not summaries:
        st.markdown('<div class="wm-empty">В Baserow пока нет связанных поставок.</div>', unsafe_allow_html=True)
        return

    selected = _selected_summary(summaries, st.session_state.get("warehouse_selected_supply_id"))
    if selected is None:
        filters = st.columns([2, 1, 1])
        query = filters[0].text_input("Поиск", placeholder="Номер или поставщик", key="warehouse_supply_search").strip().casefold()
        statuses = sorted({item.status for item in summaries if item.status})
        status = filters[1].selectbox("Статус", ["Все", *statuses], key="warehouse_supply_status_filter")
        only_open = filters[2].toggle("Только незавершённые", value=False, key="warehouse_supply_only_open")
        current = [item for item in summaries if (not query or query in f"{item.supply_id} {item.supplier}".casefold()) and (status == "Все" or item.status == status) and (not only_open or item.qty_waiting > 0)]
        metrics = st.columns(4)
        metrics[0].metric("Поставок", len(current))
        metrics[1].metric("По документу", f"{sum(x.qty_document for x in current):,}")
        metrics[2].metric("Принято", f"{sum(x.qty_received for x in current):,}")
        metrics[3].metric("Ожидается", f"{sum(x.qty_waiting for x in current):,}")
        if not current:
            st.markdown('<div class="wm-empty">Поставок по выбранным фильтрам нет.</div>', unsafe_allow_html=True)
            return
        for start in range(0, len(current), 2):
            columns = st.columns(2)
            for column, supply in zip(columns, current[start:start + 2]):
                with column:
                    with st.container(border=True):
                        status_class = _supply_status_class(supply.status, supply.qty_waiting)
                        progress = 1.0 if supply.qty_document <= 0 else min(max(supply.qty_received / supply.qty_document, 0.0), 1.0)
                        st.markdown(
                            '<div class="wm-supply-card">'
                            f'<div class="title">{escape(supply.supply_id)}</div>'
                            f'<div class="sub">{escape(supply.date or "Без даты")} · {escape(supply.supplier or "Поставщик не указан")}</div>'
                            f'<span class="wm-status {status_class}">{escape(supply.status or "Без статуса")}</span>'
                            '</div>', unsafe_allow_html=True,
                        )
                        st.progress(progress)
                        small = st.columns(3)
                        small[0].metric("SKU", supply.sku_total)
                        small[1].metric("Принято", supply.qty_received)
                        small[2].metric("Осталось", supply.qty_waiting)
                        if st.button("Открыть поставку", key=f"warehouse_open_supply_{supply.row_id}", width="stretch"):
                            st.session_state["warehouse_selected_supply_id"] = int(supply.row_id)
                            st.rerun()
        with st.expander("Показать все поставки таблицей", expanded=False):
            st.dataframe(_supply_table(current), width="stretch", hide_index=True, height=360)
        return

    top = st.columns([1, 4, 1])
    if top[0].button("← К реестру", key="warehouse_supply_back"):
        st.session_state.pop("warehouse_selected_supply_id", None)
        st.rerun()
    top[1].markdown(f"### {selected.supply_id}")
    if selected.qty_waiting > 0 and can_write():
        top[2].button("Принять", type="primary", width="stretch", on_click=_navigate, args=("Поставки", "Приёмка"), key=f"warehouse_supply_to_receiving_{selected.row_id}")
        st.session_state["warehouse_receiving_supply_id"] = int(selected.row_id)
    summary_cols = st.columns(4)
    summary_cols[0].metric("SKU", selected.sku_total)
    summary_cols[1].metric("По документу", selected.qty_document)
    summary_cols[2].metric("Принято", selected.qty_received)
    summary_cols[3].metric("Ожидается", selected.qty_waiting)
    rows = service.supply_products(selected)
    visible_rows = _page_slice(
        rows,
        key=f"warehouse_supply_detail_{selected.row_id}",
        default_size=20,
    )
    detail = pd.DataFrame([{
        "Фото": _row_photo_data_uri(row, config), "Артикул": row.get("Артикул"), "Коробки": row.get("_boxes"),
        "По документу": as_int(row.get("_document")), "Принято": as_int(row.get("_received")),
        "Ожидается": max(as_int(row.get("_document")) - as_int(row.get("_received")), 0),
        "Остаток": as_int(row.get("Остаток")), "row_id": int(row["id"]),
    } for row in visible_rows])
    st.dataframe(
        detail.drop(columns=["row_id"], errors="ignore"),
        width="stretch",
        hide_index=True,
        height=min(720, max(220, len(detail) * WAREHOUSE_TABLE_ROW_HEIGHT)),
        row_height=WAREHOUSE_TABLE_ROW_HEIGHT,
        column_config={"Фото": st.column_config.ImageColumn("Фото", width="large")},
    )
    silver_rows = [row for row in rows if bool(row.get("_silver_925"))]
    if silver_rows:
        with st.expander("Развёрнутая цена серебра по этой поставке", expanded=True):
            st.caption(
                "Это исторические значения из инвойса. Изменение верхнего курса и коэффициента "
                "не переписывает эту таблицу."
            )
            price_table = pd.DataFrame(
                [
                    {
                        "SKU": row.get("Артикул"),
                        "Название": row.get("_line_name") or row.get("Название"),
                        "Оригинал": row.get("_original_name"),
                        "Группа": row.get("_silver_category"),
                        "Покрытие": row.get("_plating"),
                        "Размер": row.get("_size"),
                        "Единица": row.get("_unit_label"),
                        "Количество": as_int(row.get("_document")),
                        "Вес партии, г": row.get("_total_weight_g"),
                        "Вес единицы, г": row.get("_unit_weight_g"),
                        "Серебро RMB/г": row.get("_silver_rmb_per_g"),
                        "Работа RMB/г": row.get("_labour_rmb_per_g"),
                        "Цена RMB/г": row.get("_price_rmb_per_g"),
                        "Сумма RMB": row.get("_amount_rmb"),
                        "USD/RMB": row.get("_usd_rmb_rate"),
                        "CIF, %": row.get("_cif_percent"),
                        "Закупка USD/ед.": row.get("_purchase_usd"),
                        "Продажа USD при импорте": row.get("_invoice_sale_usd"),
                        "USD/VND при импорте": row.get("_invoice_usd_vnd"),
                        "Коэффициент при импорте": row.get("_invoice_coefficient"),
                        "Продажа VND при импорте": row.get("_invoice_sale_vnd"),
                    }
                    for row in silver_rows
                ]
            )
            st.dataframe(price_table, width="stretch", hide_index=True, height=600)
    if can_write():
        with st.expander("Исправить поставку", expanded=False):
            all_detail = pd.DataFrame([{
                "Артикул": row.get("Артикул"),
                "Принято": as_int(row.get("_received")),
                "row_id": int(row["id"]),
            } for row in rows])
            waiting = all_detail.loc[all_detail["Принято"] <= 0] if not all_detail.empty else all_detail
            remove_skus = st.multiselect("Убрать непринятые позиции", waiting["Артикул"].tolist() if not waiting.empty else [], key=f"warehouse_remove_waiting_{selected.row_id}")
            if st.button("Убрать выбранные из поставки", disabled=not remove_skus, key=f"warehouse_remove_waiting_button_{selected.row_id}"):
                ids = all_detail.loc[all_detail["Артикул"].isin(remove_skus), "row_id"].astype(int).tolist()
                removed = _safe_action(lambda: service.remove_waiting_from_supply(selected, ids))
                if removed is not None:
                    st.success(f"Убрано позиций: {removed}")
                    st.rerun()
            st.divider()
            st.caption("Полностью удалить можно только пустую поставку без приёмки.")
            confirm = st.text_input(f"Для удаления введите {selected.supply_id}", key=f"warehouse_delete_supply_confirm_{selected.row_id}")
            if st.button("Удалить пустую поставку", disabled=confirm.strip() != selected.supply_id or selected.qty_received > 0, key=f"warehouse_delete_supply_{selected.row_id}"):
                result = _safe_action(lambda: service.delete_empty_supply(selected))
                if result:
                    st.session_state.pop("warehouse_selected_supply_id", None)
                    st.success("Пустая поставка удалена.")
                    st.rerun()


def _runtime_dir() -> Path:
    root = Path(".runtime") / "warehouse_uploads"
    root.mkdir(parents=True, exist_ok=True)
    cutoff = datetime.now() - timedelta(hours=24)
    for child in root.iterdir():
        try:
            if child.is_dir() and datetime.fromtimestamp(child.stat().st_mtime) < cutoff:
                shutil.rmtree(child, ignore_errors=True)
        except OSError:
            continue
    return root


def _parse_uploaded_supply(uploaded: Any, service: WarehouseService) -> tuple[list[Product], Path, dict[str, Any]]:
    session_dir = _runtime_dir() / uuid.uuid4().hex
    image_dir = session_dir / "images"
    image_dir.mkdir(parents=True, exist_ok=True)
    path = session_dir / Path(uploaded.name).name
    data = uploaded.getvalue()
    if len(data) > 150 * 1024 * 1024:
        raise ValueError("Размер файла превышает лимит 150 МБ.")
    path.write_bytes(data)
    if is_silver_invoice(path):
        products, meta = parse_silver_invoice(path, image_dir)
        skus = service.next_silver_skus(len(products))
        for product, sku in zip(products, skus):
            product.sku = sku
        metadata = meta.to_dict()
        metadata["source_name"] = Path(uploaded.name).name
        metadata["section"] = "Комплектующие"
        return products, session_dir, metadata
    products = load_products(path, image_dir)
    return products, session_dir, {
        "source_type": "standard",
        "source_name": Path(uploaded.name).name,
    }

def _products_editor_frame(products: list[Product]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Фото": _local_thumbnail_data_uri(
                    product.image_path,
                    Path(product.image_path).stat().st_mtime_ns if product.image_path and Path(product.image_path).exists() else 0,
                ),
                "№": product.number,
                "Артикул": product.sku,
                "Название": product.name or product.description,
                "Коробки": product.boxes,
                "По документу": product.qty_document,
                "Категория": product.category,
                "Серебряная категория": product.silver_category,
                "Материал": product.material,
                "Камень": product.stone,
                "Цвет": product.color,
                "Покрытие": product.plating,
                "Размер": product.size,
                "Единица": product.unit_label,
                "Продаётся отдельно": bool(product.sellable),
                "Закупка USD/ед.": product.purchase_usd_per_unit,
                "Получено сейчас": bool(product.received),
                "Факт": product.actual_manual if product.actual_manual is not None else product.qty_document if product.received else 0,
                "Комментарий": product.comment,
                "image_path": product.image_path,
                "description": product.description,
                "unit_weight_kg": product.unit_weight_kg,
                "silver_925": product.silver_925,
                "original_name": product.original_name,
                "total_weight_g": product.total_weight_g,
                "silver_rmb_per_g": product.silver_rmb_per_g,
                "labour_rmb_per_g": product.labour_rmb_per_g,
                "price_rmb_per_g": product.price_rmb_per_g,
                "amount_rmb": product.amount_rmb,
                "usd_rmb_rate": product.usd_rmb_rate,
                "cif_percent": product.cif_percent,
                "invoice_sale_usd": product.invoice_sale_usd,
                "invoice_usd_vnd_rate": product.invoice_usd_vnd_rate,
                "invoice_coefficient": product.invoice_coefficient,
                "invoice_sale_vnd": product.invoice_sale_vnd,
            }
            for product in products
        ]
    )


def _optional_float(value: Any) -> float | None:
    if value is None or pd.isna(value) or str(value).strip() == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _products_from_editor(frame: pd.DataFrame) -> list[Product]:
    products: list[Product] = []
    for _, row in frame.iterrows():
        received = bool(row.get("Получено сейчас", False))
        actual = as_int(row.get("Факт")) if received else None
        products.append(
            Product(
                number=as_int(row.get("№")),
                boxes=str(row.get("Коробки") or ""),
                sku=str(row.get("Артикул") or "").strip(),
                qty_document=as_int(row.get("По документу")),
                description=str(row.get("description") or row.get("Название") or ""),
                category=str(row.get("Категория") or ""),
                material=str(row.get("Материал") or ""),
                stone=str(row.get("Камень") or ""),
                color=str(row.get("Цвет") or ""),
                unit_weight_kg=_optional_float(row.get("unit_weight_kg")),
                image_path=str(row.get("image_path") or ""),
                received=received,
                actual_manual=actual,
                comment=str(row.get("Комментарий") or ""),
                name=str(row.get("Название") or ""),
                silver_category=str(row.get("Серебряная категория") or ""),
                silver_925=bool(row.get("silver_925", False)),
                plating=str(row.get("Покрытие") or ""),
                size=str(row.get("Размер") or ""),
                unit_label=str(row.get("Единица") or "шт."),
                sellable=bool(row.get("Продаётся отдельно", False)),
                original_name=str(row.get("original_name") or ""),
                total_weight_g=_optional_float(row.get("total_weight_g")),
                silver_rmb_per_g=_optional_float(row.get("silver_rmb_per_g")),
                labour_rmb_per_g=_optional_float(row.get("labour_rmb_per_g")),
                price_rmb_per_g=_optional_float(row.get("price_rmb_per_g")),
                amount_rmb=_optional_float(row.get("amount_rmb")),
                usd_rmb_rate=_optional_float(row.get("usd_rmb_rate")),
                cif_percent=_optional_float(row.get("cif_percent")),
                purchase_usd_per_unit=_optional_float(row.get("Закупка USD/ед.")),
                invoice_sale_usd=_optional_float(row.get("invoice_sale_usd")),
                invoice_usd_vnd_rate=as_int(row.get("invoice_usd_vnd_rate")) or None,
                invoice_coefficient=_optional_float(row.get("invoice_coefficient")),
                invoice_sale_vnd=as_int(row.get("invoice_sale_vnd")) or None,
            )
        )
    return products

def render_new_supply(config: Any) -> None:
    _page_header(
        "Новая поставка",
        "Загрузите Packing List, Master или серебряный Invoice. Файл серебра распознаётся автоматически вместе с ценами и фотографиями.",
    )
    service = _require_safe_schema(config)
    if service is None:
        return
    raw_products = st.session_state.get("warehouse_supply_products", [])
    metadata = dict(st.session_state.get("warehouse_supply_meta", {}) or {})
    is_silver = metadata.get("source_type") == "silver_invoice"
    _workflow(1 if not raw_products else 2)

    if not raw_products:
        with st.container(border=True):
            st.markdown("### 1. Загрузите файл поставки")
            st.caption(
                "Поддерживаются XLSX и XLSM до 150 МБ. Для серебряного инвойса сайт автоматически "
                "извлекает позиции, фотографии, серебро 925, покрытия, размеры и цены. "
                "После проверки создаётся ожидаемая поставка; фактическая приёмка выполняется отдельно."
            )
            uploaded = st.file_uploader(
                "Файл поставки",
                type=["xlsx", "xlsm"],
                key="warehouse_supply_file",
                label_visibility="collapsed",
            )
            action_cols = st.columns([1, 2])
            if action_cols[0].button(
                "Разобрать файл",
                type="primary",
                width="stretch",
                disabled=uploaded is None,
                key="warehouse_parse_supply",
            ):
                try:
                    with st.spinner("Извлекаем строки, цены и фотографии..."):
                        products, session_dir, parsed_meta = _parse_uploaded_supply(uploaded, service)
                    old = st.session_state.get("warehouse_supply_runtime_dir")
                    if old:
                        shutil.rmtree(str(old), ignore_errors=True)
                    st.session_state["warehouse_supply_runtime_dir"] = str(session_dir)
                    st.session_state["warehouse_supply_products"] = [product.to_dict() for product in products]
                    st.session_state["warehouse_supply_meta"] = parsed_meta
                    st.session_state.pop("warehouse_new_supply_id", None)
                    st.session_state.pop("warehouse_new_supply_id_input", None)
                    st.success(
                        f"Распознано: {len(products)} SKU"
                        + (" · серебро 925 · цены сохранены" if parsed_meta.get("source_type") == "silver_invoice" else "")
                    )
                    st.rerun()
                except Exception as exc:
                    st.error(f"Не удалось разобрать файл: {exc}")
            action_cols[1].caption(
                "Оригинал остаётся первичным документом. Для сайта фотографии автоматически обрезаются и уменьшаются, чтобы склад не зависал."
            )
        return

    products = [Product.from_dict(item) for item in raw_products]
    usd_vnd = int(st.session_state.get("warehouse_silver_usd_vnd", SILVER_DEFAULT_USD_VND))
    coefficient = float(st.session_state.get("warehouse_silver_coefficient", SILVER_DEFAULT_COEFFICIENT))

    if is_silver:
        source_variant = str(metadata.get("source_variant") or "")
        if source_variant == "supplier_raw_2026_06_30":
            recognized_note = (
                "Все 14 позиций классифицированы по согласованным группам, фотографии восстановлены, "
                "цены сохранены. Сейчас создаётся только ожидаемая поставка; остатки не изменятся."
            )
        else:
            recognized_note = (
                "Все товары создаются в «Комплектующих» как серебро 925. Пусеты учитываются парами; "
                "цепочка №11 помечена как товар для отдельной продажи. Сейчас создаётся только ожидаемая поставка."
            )
        st.markdown(
            '<div class="wm-good"><b>Серебряный Invoice распознан.</b> ' + recognized_note + '</div>',
            unsafe_allow_html=True,
        )
        price_info = st.columns(5)
        price_info[0].metric("Позиций", len(products))
        price_info[1].metric("По документу", f"{sum(product.qty_document for product in products):,}")
        price_info[2].metric("Вес", f"{sum(product.total_weight_g or 0 for product in products):,.2f} г")
        price_info[3].metric("Сумма", f"{sum(product.amount_rmb or 0 for product in products):,.2f} RMB")
        price_info[4].metric("Фото", sum(bool(product.image_path) for product in products))
        if metadata.get("purchase_price_calculated"):
            st.info(
                "В исходном файле указаны закупочные цены в RMB: серебро/г, работа/г, цена/г и сумма строки. "
                "Закупка USD/ед. рассчитана без дополнительной надбавки: Сумма RMB ÷ "
                f"{float(metadata.get('usd_rmb') or 6.71):.4f} ÷ количество. Продажная цена сейчас: "
                f"закупка USD × {coefficient:g} × {usd_vnd:,} VND, с округлением вверх до 1 000 VND. "
                "Закупку USD можно поправить до создания поставки."
            )
    else:
        summary = st.columns(5)
        summary[0].metric("SKU", len(products))
        summary[1].metric("С фото", sum(bool(product.image_path and Path(product.image_path).exists()) for product in products))
        summary[2].metric("По документу", sum(product.qty_document for product in products))
        summary[3].metric("Принимается", sum(product.actual_qty or 0 for product in products))
        summary[4].metric("Ожидается", sum(product.waiting_qty for product in products))

    control = st.columns([1, 1, 3])
    if control[0].button("Загрузить другой файл", width="stretch", key="warehouse_supply_reset"):
        _reset_supply_draft()
        st.rerun()
    control[2].caption(
        "Создание поставки не означает приёмку. Все строки по умолчанию ожидаются. "
        "Когда товар приедет, откройте «Приёмка → По поставке» и внесите фактическое количество после пересчёта."
    )

    frame = _products_editor_frame(products)
    if is_silver:
        frame["Продажа VND сейчас"] = [
            silver_sale_vnd(value, usd_vnd, coefficient)
            for value in frame["Закупка USD/ед."].tolist()
        ]
        visible_columns = [
            "Фото", "№", "Артикул", "Название", "Серебряная категория", "Покрытие", "Размер",
            "Единица", "По документу", "Закупка USD/ед.", "Продажа VND сейчас",
            "Продаётся отдельно", "Получено сейчас", "Факт", "Коробки", "Комментарий",
        ]
        disabled_columns = ["Фото", "№", "Продажа VND сейчас"]
        if not metadata.get("purchase_price_calculated"):
            disabled_columns.append("Закупка USD/ед.")
        column_config = {
            "Фото": st.column_config.ImageColumn("Фото", width="large"),
            "Серебряная категория": st.column_config.SelectboxColumn(
                "Группа серебра", options=list(SILVER_CATEGORIES), required=True
            ),
            "По документу": st.column_config.NumberColumn("По документу", min_value=0, step=1),
            "Закупка USD/ед.": st.column_config.NumberColumn("Закупка USD", format="$ %.6f"),
            "Продажа VND сейчас": st.column_config.NumberColumn("Продажа VND", format="localized"),
            "Факт": st.column_config.NumberColumn("Факт", min_value=0, step=1),
            "Получено сейчас": st.column_config.CheckboxColumn("Получено сейчас"),
            "Продаётся отдельно": st.column_config.CheckboxColumn("Продаётся отдельно"),
        }
    else:
        visible_columns = [
            "Фото", "№", "Артикул", "Коробки", "По документу", "Категория", "Материал",
            "Камень", "Цвет", "Получено сейчас", "Факт", "Комментарий",
        ]
        disabled_columns = ["Фото"]
        column_config = {
            "Фото": st.column_config.ImageColumn("Фото", width="large"),
            "Категория": st.column_config.SelectboxColumn("Категория", options=CATEGORIES),
            "По документу": st.column_config.NumberColumn("По документу", min_value=0, step=1),
            "Факт": st.column_config.NumberColumn("Факт", min_value=0, step=1),
            "Получено сейчас": st.column_config.CheckboxColumn("Получено сейчас"),
        }

    edited = st.data_editor(
        frame,
        column_order=visible_columns,
        hide_index=True,
        width="stretch",
        height=650,
        row_height=WAREHOUSE_TABLE_ROW_HEIGHT,
        num_rows="fixed",
        disabled=disabled_columns,
        key="warehouse_supply_editor",
        column_config=column_config,
    )
    updated_products = _products_from_editor(edited)
    if is_silver and metadata.get("purchase_price_calculated"):
        refresh_calculated_silver_prices(
            updated_products,
            usd_vnd=usd_vnd,
            coefficient=coefficient,
        )
    st.session_state["warehouse_supply_products"] = [product.to_dict() for product in updated_products]

    issues: list[str] = []
    issues += [f"Строка {p.number}: нет артикула" for p in updated_products if not p.sku]
    issues += [f"{p.sku or p.number}: количество по документу равно 0" for p in updated_products if p.qty_document <= 0]
    if is_silver:
        issues += [f"{p.sku}: не указана закупка USD" for p in updated_products if not p.purchase_usd_per_unit]
        issues += [f"{p.sku}: не указана группа серебра" for p in updated_products if not p.silver_category]
    sku_rows: dict[str, list[int]] = {}
    for product in updated_products:
        if product.sku:
            sku_rows.setdefault(product.sku.casefold(), []).append(product.number)
    for sku_key, row_numbers in sku_rows.items():
        if len(row_numbers) > 1:
            shown = next(product.sku for product in updated_products if product.sku.casefold() == sku_key)
            issues.append(f"SKU {shown} повторяется в строках {', '.join(map(str, row_numbers))}")
    no_photo = [p.sku for p in updated_products if not (p.image_path and Path(p.image_path).exists())]
    with st.expander(f"Проверка данных · ошибок {len(issues)} · без фото {len(no_photo)}", expanded=bool(issues)):
        if issues:
            st.error("\n".join(f"• {item}" for item in issues[:50]))
        else:
            st.success("Обязательные поля заполнены.")
        if no_photo:
            st.warning("Без фотографии: " + ", ".join(no_photo[:40]))

    st.markdown("### 3. Подтвердите поставку")
    default_prefix = "SIL" if is_silver else "SUP"
    default_supply_id = st.session_state.setdefault(
        "warehouse_new_supply_id", service.next_supply_id(prefix=default_prefix)
    )
    info1, info2 = st.columns(2)
    supply_id = info1.text_input(
        "Номер поставки *", value=default_supply_id, key="warehouse_new_supply_id_input"
    )
    if is_silver:
        info2.markdown("**Раздел:** Комплектующие → Серебро 925")
        section = "Комплектующие"
    else:
        section = info2.segmented_control(
            "Тип поставки",
            ["Сувенирка", "Комплектующие"],
            default="Сувенирка",
            key="warehouse_new_supply_section",
        ) or "Сувенирка"

    supplier_default = str(metadata.get("supplier") or "")
    invoice_default = str(metadata.get("source_name") or "")
    supplier = info2.text_input("Поставщик", value=supplier_default, key="warehouse_new_supply_supplier")
    invoice = info1.text_input("Invoice", value=invoice_default, key="warehouse_new_supply_invoice")
    comment_default = (
        f"Серебро 925 · invoice {metadata.get('invoice_date') or ''} · товар ожидается"
        if is_silver
        else ""
    )
    comment = info2.text_input("Комментарий", value=comment_default, key="warehouse_new_supply_comment")

    master_buffer = BytesIO()
    with tempfile.TemporaryDirectory(prefix="analitika-master-") as temp_dir:
        master_path = Path(temp_dir) / "Master.xlsx"
        export_master(master_path, updated_products)
        master_buffer.write(master_path.read_bytes())
    download_col, create_col = st.columns([1, 2])
    download_col.download_button(
        "Скачать проверенный Master",
        data=master_buffer.getvalue(),
        file_name=f"{supply_id or 'Master'}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    can_create = bool(supply_id.strip()) and bool(updated_products) and not issues
    create_label = "Создать поставку в Baserow"
    if create_col.button(
        create_label,
        type="primary",
        width="stretch",
        disabled=not can_create,
    ):
        spinner_text = "Создаём карточки, позиции ожидаемой поставки и сохраняем цены..."
        with st.spinner(spinner_text):
            if is_silver and not _ensure_silver_schema(config):
                return
            command_id = st.session_state.setdefault(
                "warehouse_new_supply_command_id", f"IMPORT-{uuid.uuid4().hex}"
            )
            result = _safe_action(
                lambda: service.create_supply_from_products(
                    supply_id=supply_id,
                    supplier=supplier,
                    invoice=invoice,
                    comment=comment,
                    products=updated_products,
                    section=section,
                    command_id=command_id,
                )
            )
        if result:
            _reset_supply_draft()
            st.success(
                f"Поставка {result['supply_id']} создана. SKU: {result['sku']}; "
                f"принято: {result['received']} шт.; ожидается: {result['waiting']} шт."
            )
            if result.get("failed_photos"):
                st.warning("Не удалось загрузить фото: " + ", ".join(result["failed_photos"][:30]))
            st.session_state.pop("warehouse_new_supply_command_id", None)
            st.session_state.pop("warehouse_new_supply_id", None)
            _queue_widget_state(
                warehouse_workspace="Поставки",
                warehouse_supply_workspace="Реестр",
            )
            st.rerun()

def render_receiving(config: Any) -> None:
    _page_header("Приёмка", "Выберите поставку, проверьте крупные фотографии и укажите фактически полученное количество.")
    service = _require_safe_schema(config)
    if service is None:
        return
    mode = st.segmented_control(
        "Способ",
        ["По поставке", "Ручной приход"],
        default="По поставке",
        key="warehouse_receiving_mode",
    ) or "По поставке"

    if mode == "Ручной приход":
        section = st.segmented_control(
            "Раздел",
            ["Сувенирка", "Комплектующие"],
            default="Сувенирка",
            key="warehouse_manual_receipt_section",
        ) or "Сувенирка"
        items = service.catalog(section)
        labels = {f"{item.sku} · остаток {item.balance}": item for item in items}
        selected_labels = st.multiselect(
            "Найдите и выберите товары",
            list(labels),
            key="warehouse_manual_receipt_items",
        )
        records = [
            {
                "id": labels[label].row_id,
                "item": labels[label],
                "sku": labels[label].sku,
                "maximum": None,
                "initial": 1,
                "meta": [
                    f"Текущий остаток: {labels[label].balance:,} шт.",
                    "Ручной приход увеличит остаток после проведения.",
                ],
            }
            for label in selected_labels
        ]
        if not records:
            st.markdown('<div class="wm-empty">Выберите хотя бы один товар.</div>', unsafe_allow_html=True)
            return
        draft_key = f"warehouse_manual_receipt_draft_{section}"
        revision = as_int(st.session_state.get(f"{draft_key}_revision"), 0)
        draft = _render_quantity_editor(
            records,
            config,
            draft_key=draft_key,
            page_key=f"{draft_key}_rows",
            revision=revision,
            quantity_label="Принять, шт.",
            minimum=1,
        )
        comment = st.text_input(
            "Основание и комментарий *",
            placeholder="Например: товар без Packing List / возврат / инвентаризация",
            key="warehouse_manual_receipt_comment",
        )
        total = sum(draft.values())
        action = st.columns([1, 2])
        action[0].metric("К приходу", f"{total:,} шт.")
        if action[1].button(
            "Провести ручной приход",
            type="primary",
            width="stretch",
            disabled=total <= 0 or not comment.strip(),
        ):
            quantities = {int(key): int(value) for key, value in draft.items() if int(value) > 0}
            command_id = st.session_state.setdefault(
                "warehouse_manual_receipt_command",
                f"CMD-REC-{uuid.uuid4().hex}",
            )
            result = _safe_action(
                lambda: service.manual_operation(
                    operation_type="Приход",
                    section=section,
                    quantities=quantities,
                    comment=comment,
                    command_id=command_id,
                )
            )
            if result:
                st.success(f"Проведено: {result['batch_id']} · {result['quantity']} шт.")
                st.session_state.pop("warehouse_manual_receipt_command", None)
                st.session_state.pop(draft_key, None)
                st.session_state[f"{draft_key}_revision"] = revision + 1
                st.rerun()
        return

    summaries = [item for item in service.supply_summaries() if item.qty_waiting > 0]
    if not summaries:
        st.markdown('<div class="wm-empty">Нет поставок с ожидаемым количеством.</div>', unsafe_allow_html=True)
        return
    options, mapping = _summary_options(summaries)
    preferred_id = int(st.session_state.get("warehouse_receiving_supply_id", 0) or 0)
    default_index = next((i for i, label in enumerate(options) if mapping[label].row_id == preferred_id), 0)
    selected_label = st.selectbox(
        "Поставка",
        options,
        index=default_index,
        key="warehouse_receiving_supply",
    )
    supply = mapping[selected_label]
    st.session_state["warehouse_receiving_supply_id"] = int(supply.row_id)
    supply_metrics = st.columns(4)
    supply_metrics[0].metric("SKU", supply.sku_total)
    supply_metrics[1].metric("По документу", supply.qty_document)
    supply_metrics[2].metric("Уже принято", supply.qty_received)
    supply_metrics[3].metric("Ожидается", supply.qty_waiting)

    rows = service.supply_products(supply)
    records: list[dict[str, Any]] = []
    for row in rows:
        waiting = max(as_int(row.get("_document")) - as_int(row.get("_received")), 0)
        if waiting <= 0:
            continue
        line_id = as_int(row.get("_line_id")) or int(row["id"])
        records.append(
            {
                "id": line_id,
                "row": row,
                "sku": row.get("Артикул"),
                "maximum": waiting,
                "initial": 0,
                "meta": [
                    f"{row.get('_line_name') or row.get('Название') or ''}",
                    f"По документу: {as_int(row.get('_document')):,} {row.get('_unit_label') or 'шт.'}",
                    f"Уже принято: {as_int(row.get('_received')):,} {row.get('_unit_label') or 'шт.'}",
                    f"Осталось принять: {waiting:,} {row.get('_unit_label') or 'шт.'}",
                    f"Коробки: {row.get('_boxes') or 'не указаны'}",
                ],
            }
        )
    if not records:
        st.info("В поставке больше нет ожидаемых позиций.")
        return

    receiving_calculation = st.segmented_control(
        "Как определить количество",
        ["По количеству", "По весу — товар уже в работе"],
        default="По количеству",
        key=f"warehouse_receiving_calculation_{supply.row_id}",
        help=(
            "Обычная поставка принимается по количеству. Режим по весу нужен только тогда, "
            "когда поставка уже была получена полностью и часть товара успели использовать до постановки на учёт."
        ),
    ) or "По количеству"

    if receiving_calculation == "По весу — товар уже в работе":
        if not st.session_state.get("warehouse_weight_schema_ready"):
            with st.spinner("Проверяем поля весовой приёмки в Baserow..."):
                if not _ensure_silver_schema(config):
                    return
            st.session_state["warehouse_weight_schema_ready"] = True
        weight_records: list[dict[str, Any]] = []
        blocked_partial: list[str] = []
        blocked_weight: list[str] = []
        for row in rows:
            document = as_int(row.get("_document"))
            received = as_int(row.get("_received"))
            if document <= received:
                continue
            if received > 0:
                blocked_partial.append(str(row.get("Артикул") or ""))
                continue
            if float(row.get("_unit_weight_g") or 0.0) <= 0:
                blocked_weight.append(str(row.get("Артикул") or ""))
                continue
            line_id = as_int(row.get("_line_id")) or int(row["id"])
            weight_records.append(
                {
                    "id": line_id,
                    "row": row,
                    "sku": row.get("Артикул"),
                    "maximum": document,
                    "meta": [
                        f"{row.get('_line_name') or row.get('Название') or ''}",
                        f"По инвойсу: {document:,} {row.get('_unit_label') or 'шт.'}",
                        f"Вес партии по инвойсу: {float(row.get('_total_weight_g') or 0.0):,.4f} г",
                        f"Коробки: {row.get('_boxes') or 'не указаны'}",
                    ],
                }
            )
        if blocked_partial:
            st.warning(
                "Уже частично принятые строки нельзя переводить в весовой режим: "
                + ", ".join(blocked_partial[:20])
            )
        if blocked_weight:
            st.warning(
                "Нет среднего веса единицы; примите по количеству: "
                + ", ".join(blocked_weight[:20])
            )
        if not weight_records:
            st.info("В этой поставке нет строк, доступных для первичной приёмки по весу.")
            return

        weight_draft_key = f"warehouse_weight_receiving_draft_{supply.row_id}"
        weight_revision_key = f"{weight_draft_key}_revision"
        weight_revision = as_int(st.session_state.get(weight_revision_key), 0)
        top_weight = st.columns([1, 3])
        if top_weight[0].button(
            "Очистить веса",
            width="stretch",
            key=f"warehouse_weight_clear_{supply.row_id}",
        ):
            st.session_state.pop(weight_draft_key, None)
            st.session_state[weight_revision_key] = weight_revision + 1
            st.rerun()
        top_weight[1].caption(
            "Можно обработать позиции частями: строки без введённого веса останутся в ожидании. Нулевой вес означает, что текущий остаток равен нулю."
        )
        measurements = _render_weight_receiving_editor(
            weight_records,
            config,
            draft_key=weight_draft_key,
            page_key=f"warehouse_weight_receiving_rows_{supply.row_id}",
            revision=weight_revision,
        )
        received_total = 0
        current_total = 0
        for record in weight_records:
            measurement = measurements.get(int(record["id"]))
            if not measurement:
                continue
            received_total += as_int(record.get("maximum"))
            current_total += as_int(measurement.get("quantity"))
        written_off_total = max(received_total - current_total, 0)
        weight_metrics = st.columns(4)
        weight_metrics[0].metric("Заполнено SKU", len(measurements))
        weight_metrics[1].metric("Полный приход", f"{received_total:,}")
        weight_metrics[2].metric("Текущий остаток", f"{current_total:,}")
        weight_metrics[3].metric("Использовано до учёта", f"{written_off_total:,}")
        confirmed = st.checkbox(
            "Подтверждаю: выбранные позиции поставщик привёз полностью, а разницу нужно оформить расходом «Использовано до постановки на учёт».",
            key=f"warehouse_weight_confirm_{supply.row_id}",
        )
        if st.button(
            "Провести приёмку по весу",
            type="primary",
            width="stretch",
            disabled=not measurements or not confirmed,
            key=f"warehouse_weight_submit_{supply.row_id}",
        ):
            command_key = f"warehouse_weight_receiving_command_{supply.row_id}"
            command_id = st.session_state.setdefault(command_key, f"CMD-RECW-{uuid.uuid4().hex}")
            result = _safe_action(
                lambda: service.receive_existing_supply_by_weight(
                    supply,
                    measurements,
                    command_id=command_id,
                )
            )
            if result:
                st.success(
                    f"Приёмка {result['batch_id']}: полный приход {result['received']:,}; "
                    f"текущий остаток {result['current']:,}; списано как использованное {result['written_off']:,}."
                )
                st.session_state.pop(weight_draft_key, None)
                st.session_state[weight_revision_key] = weight_revision + 1
                st.session_state.pop(command_key, None)
                st.rerun()
        return

    draft_key = f"warehouse_receiving_draft_{supply.row_id}"
    revision_key = f"{draft_key}_revision"
    revision = as_int(st.session_state.get(revision_key), 0)
    current_draft = dict(st.session_state.get(draft_key, {}) or {})
    for record in records:
        key = str(record["id"])
        current_draft[key] = min(max(as_int(current_draft.get(key), 0), 0), as_int(record["maximum"]))
    st.session_state[draft_key] = current_draft

    buttons = st.columns([1, 1, 3])
    if buttons[0].button(
        "Принять всё",
        type="primary",
        width="stretch",
        key=f"warehouse_receiving_all_{supply.row_id}",
    ):
        st.session_state[draft_key] = {str(record["id"]): as_int(record["maximum"]) for record in records}
        st.session_state[revision_key] = revision + 1
        st.rerun()
    if buttons[1].button(
        "Очистить",
        width="stretch",
        key=f"warehouse_receiving_clear_{supply.row_id}",
    ):
        st.session_state[draft_key] = {str(record["id"]): 0 for record in records}
        st.session_state[revision_key] = revision + 1
        st.rerun()
    buttons[2].caption("Золотое поле справа — итоговое количество по каждой позиции.")

    draft = _render_quantity_editor(
        records,
        config,
        draft_key=draft_key,
        page_key=f"warehouse_receiving_rows_{supply.row_id}",
        revision=revision,
        quantity_label="Принять сейчас, шт.",
        minimum=0,
    )
    total = sum(draft.values())
    footer = st.columns([1, 2])
    footer[0].metric("К приёмке", f"{total:,} шт.")
    if footer[1].button(
        "Провести приёмку",
        type="primary",
        width="stretch",
        disabled=total <= 0,
        key=f"warehouse_receive_submit_{supply.row_id}",
    ):
        quantities = {int(key): int(value) for key, value in draft.items() if int(value) > 0}
        command_key = f"warehouse_receiving_command_{supply.row_id}"
        command_id = st.session_state.setdefault(command_key, f"CMD-REC-{uuid.uuid4().hex}")
        result = _safe_action(lambda: service.receive_supply(supply, quantities, command_id=command_id))
        if result:
            st.success(f"Приёмка {result['batch_id']}: {result['sku']} SKU, {result['quantity']} шт.")
            st.session_state.pop(draft_key, None)
            st.session_state[revision_key] = revision + 1
            st.session_state.pop(command_key, None)
            st.rerun()

def _read_transfer_excel(data: bytes) -> list[tuple[str, int]]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip().casefold() for value in rows[0]]
    sku_index = next((i for i, value in enumerate(headers) if value in {"артикул", "sku", "item", "код"}), None)
    qty_index = next((i for i, value in enumerate(headers) if value in {"количество", "qty", "quantity", "шт"}), None)
    missing = []
    if sku_index is None:
        missing.append("Артикул / SKU")
    if qty_index is None:
        missing.append("Количество / Qty")
    if missing:
        raise ValueError("В Excel отсутствуют обязательные колонки: " + ", ".join(missing))
    result: list[tuple[str, int]] = []
    for row in rows[1:]:
        sku = str(row[sku_index] or "").strip() if sku_index < len(row) else ""
        quantity = as_int(row[qty_index]) if qty_index < len(row) else 0
        if sku and quantity > 0:
            result.append((sku, quantity))
    return result


def render_transfer(config: Any) -> None:
    _page_header(
        "Передача в бухгалтерию",
        "Выберите способ передачи. Количество по каждой позиции редактируется в выделенном золотом поле и не может превышать доступный остаток.",
    )
    _workflow(3)
    service = _require_safe_schema(config)
    if service is None:
        return

    mode = st.segmented_control(
        "Способ передачи",
        ["По поставке — рекомендуемый", "По отдельным SKU", "Из Excel"],
        default="По поставке — рекомендуемый",
        key="warehouse_transfer_mode",
    ) or "По поставке — рекомендуемый"

    if mode == "По поставке — рекомендуемый":
        summaries = [item for item in service.supply_summaries() if item.qty_received > 0]
        if not summaries:
            st.markdown('<div class="wm-empty">Нет поставок с принятым товаром.</div>', unsafe_allow_html=True)
            return
        options, mapping = _summary_options(summaries)
        supply = mapping[st.selectbox("Выберите поставку", options, key="warehouse_transfer_supply")]
        info = st.columns(4)
        info[0].metric("SKU", supply.sku_total)
        info[1].metric("Принято", supply.qty_received)
        info[2].metric("Ожидается приёмка", supply.qty_waiting)
        info[3].metric("Статус", supply.status or "—")

        rows = service.supply_products(supply)
        already = {} if service.has_supply_lines else service.transferred_by_supply(supply.supply_id)
        records: list[dict[str, Any]] = []
        for row in rows:
            row_id = int(row["id"])
            transferred = as_int(row.get("_transferred")) if service.has_supply_lines else already.get(row_id, 0)
            received = as_int(row.get("_received"))
            stock = as_int(row.get("Остаток"))
            maximum = min(max(received - transferred, 0), max(stock, 0))
            if maximum <= 0:
                continue
            line_id = as_int(row.get("_line_id")) or row_id
            records.append(
                {
                    "id": line_id,
                    "row": row,
                    "sku": row.get("Артикул"),
                    "maximum": maximum,
                    "initial": maximum,
                    "meta": [
                        f"Принято из этой поставки: {received:,} шт.",
                        f"Уже передано: {transferred:,} шт.",
                        f"Текущий складской остаток: {stock:,} шт.",
                        f"Коробки: {row.get('_boxes') or 'не указаны'}",
                    ],
                    "stock": stock,
                    "product_row_id": row_id,
                }
            )
        if not records:
            st.info("По этой поставке больше нет доступного количества для передачи.")
            return

        draft_key = f"warehouse_transfer_draft_{supply.row_id}"
        revision_key = f"{draft_key}_revision"
        revision = as_int(st.session_state.get(revision_key), 0)
        current_draft = dict(st.session_state.get(draft_key, {}) or {})
        for record in records:
            key = str(record["id"])
            current_draft[key] = min(
                max(as_int(current_draft.get(key), as_int(record["maximum"])), 0),
                as_int(record["maximum"]),
            )
        st.session_state[draft_key] = current_draft

        controls = st.columns([1, 1, 3])
        if controls[0].button(
            "Выбрать максимум",
            type="primary",
            width="stretch",
            key=f"warehouse_transfer_max_{supply.row_id}",
        ):
            st.session_state[draft_key] = {str(record["id"]): as_int(record["maximum"]) for record in records}
            st.session_state[revision_key] = revision + 1
            st.rerun()
        if controls[1].button(
            "Очистить",
            width="stretch",
            key=f"warehouse_transfer_clear_{supply.row_id}",
        ):
            st.session_state[draft_key] = {str(record["id"]): 0 for record in records}
            st.session_state[revision_key] = revision + 1
            st.rerun()
        controls[2].caption("Все позиции изначально выбраны по максимуму. Уменьшайте только нужные строки.")

        draft = _render_quantity_editor(
            records,
            config,
            draft_key=draft_key,
            page_key=f"warehouse_transfer_rows_{supply.row_id}",
            revision=revision,
            quantity_label="Передать, шт.",
            minimum=0,
        )
        total = sum(draft.values())
        attention_records = [
            record
            for record in records
            if draft.get(str(record["id"]), 0) > 0
            and as_int(record.get("stock")) - draft.get(str(record["id"]), 0) <= 15
        ]
        if attention_records:
            with st.expander(
                f"После передачи требуют внимания: {len(attention_records)} позиций",
                expanded=False,
            ):
                preview_rows = attention_records[:20]
                show = pd.DataFrame(
                    [
                        {
                            "Фото": _record_photo_data_uri(record, config),
                            "Артикул": record["sku"],
                            "Остаток": record["stock"],
                            "Передать": draft.get(str(record["id"]), 0),
                            "Останется": as_int(record["stock"]) - draft.get(str(record["id"]), 0),
                        }
                        for record in preview_rows
                    ]
                )
                st.dataframe(
                    show,
                    hide_index=True,
                    width="stretch",
                    height=min(520, max(180, len(show) * WAREHOUSE_TABLE_ROW_HEIGHT)),
                    row_height=WAREHOUSE_TABLE_ROW_HEIGHT,
                    column_config={"Фото": st.column_config.ImageColumn("Фото", width="large")},
                )
                if len(attention_records) > len(preview_rows):
                    st.caption(f"Показаны первые {len(preview_rows)} из {len(attention_records)} позиций.")
        comment = st.text_input(
            "Комментарий",
            value=f"Поставка {supply.supply_id}",
            key=f"warehouse_transfer_comment_{supply.row_id}",
        )
        footer = st.columns([1, 2])
        footer[0].metric("К передаче", f"{total:,} шт.")
        if footer[1].button(
            "Передать выбранное",
            type="primary",
            width="stretch",
            disabled=total <= 0,
            key=f"warehouse_transfer_submit_{supply.row_id}",
        ):
            quantities = {int(key): int(value) for key, value in draft.items() if int(value) > 0}
            command_key = f"warehouse_transfer_command_{supply.row_id}"
            command_id = st.session_state.setdefault(command_key, f"CMD-ACC-{uuid.uuid4().hex}")
            result = _safe_action(
                lambda: service.transfer_supply(
                    supply,
                    quantities,
                    comment=comment,
                    command_id=command_id,
                )
            )
            if result:
                st.success(f"Передача {result['batch_id']}: {result['sku']} SKU, {result['quantity']} шт.")
                st.session_state.pop(command_key, None)
                st.session_state.pop(draft_key, None)
                st.session_state[revision_key] = revision + 1
                st.rerun()
        return

    if mode == "По отдельным SKU":
        section = st.segmented_control(
            "Раздел",
            ["Сувенирка", "Комплектующие"],
            default="Сувенирка",
            key="warehouse_manual_transfer_section",
        ) or "Сувенирка"
        items = [item for item in service.catalog(section) if item.balance > 0]
        labels = {f"{item.sku} · остаток {item.balance}": item for item in items}
        selected_labels = st.multiselect(
            "Найдите и выберите товары",
            list(labels),
            key="warehouse_manual_transfer_items",
        )
        records = [
            {
                "id": labels[label].row_id,
                "item": labels[label],
                "sku": labels[label].sku,
                "maximum": labels[label].balance,
                "initial": min(1, labels[label].balance),
                "meta": [
                    f"Доступный остаток: {labels[label].balance:,} шт.",
                    "Укажите итоговое количество в золотом поле справа.",
                ],
            }
            for label in selected_labels
        ]
        if not records:
            st.markdown('<div class="wm-empty">Выберите товары для передачи.</div>', unsafe_allow_html=True)
            return
        draft_key = f"warehouse_manual_transfer_draft_{section}"
        revision = as_int(st.session_state.get(f"{draft_key}_revision"), 0)
        draft = _render_quantity_editor(
            records,
            config,
            draft_key=draft_key,
            page_key=f"{draft_key}_rows",
            revision=revision,
            quantity_label="Передать, шт.",
            minimum=0,
        )
        comment = st.text_input("Комментарий", key="warehouse_manual_transfer_comment")
        total = sum(draft.values())
        footer = st.columns([1, 2])
        footer[0].metric("К передаче", f"{total:,} шт.")
        if footer[1].button(
            "Провести передачу",
            type="primary",
            width="stretch",
            disabled=total <= 0,
            key="warehouse_manual_transfer_submit",
        ):
            quantities = {int(key): int(value) for key, value in draft.items() if int(value) > 0}
            command_id = st.session_state.setdefault(
                "warehouse_manual_transfer_command",
                f"CMD-ACC-{uuid.uuid4().hex}",
            )
            result = _safe_action(
                lambda: service.manual_operation(
                    operation_type="Передача в бухгалтерию",
                    section=section,
                    quantities=quantities,
                    comment=comment,
                    command_id=command_id,
                )
            )
            if result:
                st.success(f"Передача {result['batch_id']}: {result['quantity']} шт.")
                st.session_state.pop("warehouse_manual_transfer_command", None)
                st.session_state.pop(draft_key, None)
                st.session_state[f"{draft_key}_revision"] = revision + 1
                st.rerun()
        return

    st.caption("Excel должен содержать столбцы SKU/Артикул и Количество.")
    uploaded = st.file_uploader("Файл передачи", type=["xlsx", "xlsm"], key="warehouse_transfer_excel")
    section = st.segmented_control(
        "Раздел для Excel",
        ["Сувенирка", "Комплектующие"],
        default="Сувенирка",
        key="warehouse_transfer_excel_section",
    ) or "Сувенирка"
    if uploaded is None:
        st.markdown('<div class="wm-empty">Загрузите Excel-файл со списком товаров.</div>', unsafe_allow_html=True)
        return
    try:
        requested = _read_transfer_excel(uploaded.getvalue())
    except ValueError as exc:
        st.error(str(exc))
        return
    catalog = {item.sku.casefold(): item for item in service.catalog(section)}
    records: list[dict[str, Any]] = []
    missing: list[str] = []
    limited: list[str] = []
    unavailable: list[str] = []
    for sku, quantity in requested:
        item = catalog.get(sku.casefold())
        if item is None:
            missing.append(sku)
            continue
        if item.balance <= 0:
            unavailable.append(item.sku)
            continue
        initial = min(max(quantity, 0), item.balance)
        if quantity > item.balance:
            limited.append(f"{item.sku}: запрошено {quantity}, доступно {item.balance}")
        records.append(
            {
                "id": item.row_id,
                "item": item,
                "sku": item.sku,
                "maximum": item.balance,
                "initial": initial,
                "meta": [
                    f"Запрошено в Excel: {quantity:,} шт.",
                    f"Доступный остаток: {item.balance:,} шт.",
                    "Значение автоматически ограничено остатком.",
                ],
            }
        )
    if missing:
        st.warning("Не найдены: " + ", ".join(missing[:30]))
    if unavailable:
        st.warning("Нет в наличии: " + ", ".join(unavailable[:30]))
    if limited:
        st.warning("Количество автоматически уменьшено до доступного остатка:\n" + "\n".join(f"• {item}" for item in limited[:30]))
    if not records:
        st.info("В файле нет позиций, доступных для передачи.")
        return

    file_signature = f"{uploaded.name}_{len(uploaded.getvalue())}_{section}"
    draft_key = f"warehouse_excel_transfer_draft_{abs(hash(file_signature))}"
    revision = as_int(st.session_state.get(f"{draft_key}_revision"), 0)
    draft = _render_quantity_editor(
        records,
        config,
        draft_key=draft_key,
        page_key=f"{draft_key}_rows",
        revision=revision,
        quantity_label="Передать, шт.",
        minimum=0,
    )
    total = sum(draft.values())
    footer = st.columns([1, 2])
    footer[0].metric("К передаче", f"{total:,} шт.")
    if footer[1].button(
        "Провести Excel-передачу",
        type="primary",
        width="stretch",
        disabled=total <= 0,
    ):
        quantities = {int(key): int(value) for key, value in draft.items() if int(value) > 0}
        command_id = st.session_state.setdefault(
            "warehouse_excel_transfer_command",
            f"CMD-ACC-{uuid.uuid4().hex}",
        )
        result = _safe_action(
            lambda: service.manual_operation(
                operation_type="Передача в бухгалтерию",
                section=section,
                quantities=quantities,
                comment=f"Excel {uploaded.name}",
                command_id=command_id,
            )
        )
        if result:
            st.success(f"Передача {result['batch_id']}: {result['quantity']} шт.")
            st.session_state.pop("warehouse_excel_transfer_command", None)
            st.session_state.pop(draft_key, None)
            st.rerun()

def render_operations(config: Any) -> None:
    from src.warehouse import normalize_operations

    _page_header("История операций", "Фильтруйте движения склада и создавайте обратную корректировку без удаления истории.")
    service = _service(config)
    raw = service.client.list_rows(config.operations_table_id)
    frame = normalize_operations(raw)
    filters = st.columns([1, 1, 2])
    operation_types = ["Все", *sorted(frame["Тип операции"].dropna().astype(str).unique().tolist())] if not frame.empty else ["Все"]
    selected_type = filters[0].selectbox("Тип", operation_types, key="warehouse_operations_type")
    period = filters[1].selectbox("Показать", ["Все", "Последние 100", "Последние 500"], key="warehouse_operations_limit")
    query = filters[2].text_input("Поиск", placeholder="SKU, Batch ID или поставка", key="warehouse_operations_query").strip().casefold()
    current = frame.copy()
    if selected_type != "Все": current = current.loc[current["Тип операции"] == selected_type]
    if query: current = current.loc[current.astype(str).apply(lambda row: query in " ".join(row).casefold(), axis=1)]
    if period != "Все": current = current.head(int(period.split()[1]))
    metrics = st.columns(4)
    metrics[0].metric("Операций", len(current))
    metrics[1].metric("Приход", int(current.loc[current["Тип операции"].eq("Приход"), "Количество"].sum()) if not current.empty and "Количество" in current else 0)
    metrics[2].metric("Передано", int(current.loc[current["Тип операции"].eq("Передача в бухгалтерию"), "Количество"].sum()) if not current.empty and "Количество" in current else 0)
    metrics[3].metric("Корректировки", int(current["Тип операции"].isin(["Возврат", "Расход", "Корректировка"]).sum()) if not current.empty else 0)
    if "Дата" in current.columns: current["Дата"] = current["Дата"].dt.strftime("%d.%m.%Y %H:%M")
    st.dataframe(current, width="stretch", hide_index=True, height=590)
    if can_write():
        with st.expander("Создать корректировку операции", expanded=False):
            selectable = [row for row in raw if as_int(row.get("Количество")) > 0]
            labels = {f"{row.get('Batch ID') or row.get('id')} · {select_text(row.get('Тип операции'))} · {row.get('Операция') or ''}": row for row in selectable}
            if not labels:
                st.info("Нет операций для корректировки.")
            else:
                label = st.selectbox("Операция", list(labels), key="warehouse_correction_operation")
                operation = labels[label]
                available = service.correction_available(operation)
                st.caption(f"Доступно к корректировке: {available} шт.")
                if available <= 0:
                    st.info("Операция уже скорректирована полностью.")
                else:
                    quantity = st.number_input("Количество для отмены", min_value=1, max_value=available, value=available)
                    comment = st.text_input("Причина корректировки *", key="warehouse_correction_comment")
                    confirm = st.checkbox("Подтверждаю создание обратной операции", key="warehouse_correction_confirm")
                    if st.button("Создать корректировку", type="primary", disabled=not confirm or not comment.strip()):
                        command_id = st.session_state.setdefault("warehouse_correction_command", f"CMD-COR-{uuid.uuid4().hex}")
                        result = _safe_action(lambda: service.correct_operation(
                            operation, quantity=int(quantity), comment=comment, command_id=command_id
                        ))
                        if result:
                            st.success(f"Корректировка создана: {result['batch_id']}; осталось {result['remaining']} шт.")
                            st.session_state.pop("warehouse_correction_command", None)
                            st.rerun()


def render_supply_hub(config: Any) -> None:
    options = list(SUPPLY_WORKSPACES) if can_write() else ["Реестр"]
    if st.session_state.get("warehouse_supply_workspace") not in options:
        st.session_state["warehouse_supply_workspace"] = "Реестр"
    current = st.segmented_control(
        "Раздел поставок",
        options,
        key="warehouse_supply_workspace",
        label_visibility="collapsed",
    ) or "Реестр"
    if current == "Реестр":
        render_supplies(config)
    elif current == "Новая поставка":
        render_new_supply(config)
    else:
        render_receiving(config)


def render_history_hub(config: Any) -> None:
    # «Обслуживание» удалено из интерфейса. Схема Baserow проверяется и
    # восстанавливается автоматически серверным рабочим аккаунтом.
    render_operations(config)


def render_warehouse_workspace(config: Any, selected_metal_groups: Iterable[str]) -> None:
    # Загружается только выбранный раздел; технически загружается только выбранный раздел, остальные рабочие пространства не выполняются.
    """Render one lazy, task-oriented warehouse workspace inside Analitika."""
    # Apply queued navigation before any warehouse widget owns its session-state key.
    _apply_pending_widget_state()
    st.markdown(WAREHOUSE_MANAGEMENT_CSS, unsafe_allow_html=True)
    st.markdown(
        '<div class="wm-shell">'
        '<div class="wm-kicker">Princess Warehouse Online</div>'
        '<div class="wm-shell-title">Сувениры и касты на складе</div>'
        '<div class="wm-shell-copy">Фото товаров, поставки, приёмка, передача в бухгалтерию и история операций — в одном рабочем блоке.</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    _silver_price_settings()
    if can_write() and not int(getattr(_resolved_config(config), "supply_lines_table_id", 0) or 0):
        _auto_prepare_safe_schema(config)
    st.session_state.setdefault("warehouse_workspace", "Главная")
    workspace_options = list(WORKSPACES) if can_write() else ["Главная", "Товары", "Поставки", "История"]
    if st.session_state.get("warehouse_workspace") not in workspace_options:
        st.session_state["warehouse_workspace"] = "Главная"
    current = st.segmented_control(
        "Раздел склада",
        workspace_options,
        key="warehouse_workspace",
        label_visibility="collapsed",
    ) or "Главная"
    toolbar = st.columns([1, 1.45, 4.55])
    if toolbar[0].button("Обновить данные", key="warehouse_workspace_refresh", width="stretch"):
        _clear_cache(); st.rerun()
    if can_write() and toolbar[1].button(
        "Актуализировать Baserow",
        key="warehouse_workspace_reconcile_baserow",
        width="stretch",
        help=(
            "Сверяет строки поставок и их статусы с проведёнными операциями приёмки, "
            "обновляет связи товаров и удаляет только пустые карточки, оставшиеся от "
            "удалённых непринятых позиций. Новые приходы не создаются."
        ),
    ):
        with st.spinner("Сверяем Baserow с актуальными поставками и приёмками..."):
            report = _safe_action(lambda: _service(config).synchronize_baserow_from_documents())
        if isinstance(report, dict):
            st.success(
                "Baserow актуализирован: "
                f"строк поставок обновлено — {report['lines_updated']}, "
                f"поставок — {report['supplies_updated']}, "
                f"связей товаров — {report['catalog_relinked']}, "
                f"пустых карточек удалено — {report['catalog_deleted']}, "
                f"исторических карточек деактивировано — {report['catalog_deactivated']}."
            )
    toolbar[2].markdown(
        '<div class="wm-toolbar-note">«Обновить данные» перечитывает экран. '
        '«Актуализировать Baserow» исправляет только расхождения между текущими поставками, '
        'проведёнными приёмками и карточками товаров; новые операции не создаёт.</div>',
        unsafe_allow_html=True,
    )
    if current == "Главная":
        render_overview(config, selected_metal_groups)
    elif current == "Товары":
        render_catalog(config)
    elif current == "Поставки":
        render_supply_hub(config)
    elif current == "Передача":
        render_transfer(config)
    else:
        render_history_hub(config)


