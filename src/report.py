from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from collections import defaultdict
import re, zipfile, shutil, json
from datetime import datetime
import calendar
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont
import tempfile
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU

INPUT_DIR = Path.cwd()
OUTPUT = INPUT_DIR/'StoneReport_final_v4.xlsx'
PROJECT_DIR = INPUT_DIR/'StoneReport_v4_project'
ZIP_OUT = INPUT_DIR/'StoneReport_v4_project.zip'

PRODUCT_ORDER = ['Earrings','Ring','Pendant','Bracelet','Necklace','Brooch','Pearl Necklace','Pearl Bracelet','Pearl Chain','Stone','Other']
SKIP_PRODUCTS = {'CHAIN','MATERIALS'}
TOP_ORDER = ['Blue Sapphire','Ruby','Moissanite','London Topaz','Swiss Topaz','Other Topaz','Green Stones']
COLORED_ORDER = ['Black Stones','Quartz Group','Amethyst','Garnet','Agate','Other Colored Stones']
PEARL_ORDER = ['Sea Pearl','Round White Freshwater Pearl','White Freshwater Pearl','Colored Freshwater Pearl','Baroque Pearl']
SEG_ORDER = ['TOP STONES','PEARLS','COLORED STONES']
SEG_COLORS = {'TOP STONES':'7030A0','PEARLS':'FFC000','COLORED STONES':'548235'}
COLORS = {'TITLE':'1F4E78','TOTAL':'E2F0D9','SUBTOTAL':'D9EAD3','WHITE':'FFFFFF','GRID':'C8C8C8','WARN':'FCE4D6'}

STORE_PATTERNS = [
    # The detector is intentionally tolerant: AB1, AB_NEW, NTR11, NTR 1 New, etc.
    (re.compile(r'(^|[^A-Z0-9])NTR\s*2[0-9]*([^A-Z0-9]|$)', re.I), 'NTR2'),
    (re.compile(r'(^|[^A-Z0-9])NTR\s*1[0-9]*([^A-Z0-9]|$)', re.I), 'NTR1'),
    (re.compile(r'(^|[^A-Z0-9])NTR[0-9]*([^A-Z0-9]|$)', re.I), 'NTR1'),
    (re.compile(r'(^|[^A-Z0-9])AB[0-9]*([^A-Z0-9]|$)', re.I), 'AB'),
    (re.compile(r'(^|[^A-Z0-9])TT[0-9]*([^A-Z0-9]|$)|ALL\s*SALES\s*TT', re.I), 'TT'),
    (re.compile(r'(^|[^A-Z0-9])SCR[0-9]*([^A-Z0-9]|$)', re.I), 'SCR'),
    (re.compile(r'(^|[^A-Z0-9])20[0-9]*([^A-Z0-9]|$)', re.I), '20'),
    (re.compile(r'(^|[^A-Z0-9])63(\.\d+|\s*\d+)?([^A-Z0-9]|$)|MUI\s*NE', re.I), '63'),
]

PRODUCT_MAP = {
    'EARRINGS':'Earrings','EARRING':'Earrings','RING':'Ring','PENDANT':'Pendant',
    'BRACELET':'Bracelet','NECKLACE':'Necklace','BROOCH':'Brooch','PEARL NECKLACE':'Pearl Necklace',
    'PEARL BRACELET':'Pearl Bracelet','PEARL CHAIN':'Pearl Chain','STONE':'Stone','OTHER':'Other'
}

COLOR_WORDS = ['ROSE','PINK','GRAY','GREY','BLACK','PURPLE','PEACH','LAVENDER','GOLD','GOLDEN','BROWN','CHOCOLATE','MULTI','MULTICOLOR','YELLOW','BLUE','GREEN','ORANGE','RED']
SEA_WORDS = ['SEA PEARL','AKOYA','TAHITI','TAHITIAN','SOUTH SEA','MABE','GALATEA','FACETED','KESHI']

# explicit fallback aliases. Exact raw names from reports are also written to RULES automatically.
MANUAL_RULES = [
    # TOP STONES
    ('MOISSANITE','TOP STONES','Moissanite','Any moissanite / moissanite mix has absolute priority'),
    ('MOISANITE','TOP STONES','Moissanite','Typo'),
    ('MOSSANITE','TOP STONES','Moissanite','Typo'),
    ('MUSSONITE','TOP STONES','Moissanite','Typo'),
    ('SAPPHIRE','TOP STONES','Blue Sapphire','Any sapphire color/quality/treatment'),
    ('SAPPHRIE','TOP STONES','Blue Sapphire','Typo'),
    ('RUBY','TOP STONES','Ruby','Any ruby unless moissanite is present'),
    ('LONDON TOPAZ','TOP STONES','London Topaz','London topaz synonyms'),
    ('LONDON BLUE TOPAZ','TOP STONES','London Topaz','London topaz synonyms'),
    ('LONDON BT','TOP STONES','London Topaz','London topaz synonyms'),
    ('LONDON BLUE BT','TOP STONES','London Topaz','London topaz synonyms'),
    ('LONDON BLUE T','TOP STONES','London Topaz','London topaz synonyms'),
    ('SWISS TOPAZ','TOP STONES','Swiss Topaz','Swiss topaz synonyms'),
    ('SWISS BLUE TOPAZ','TOP STONES','Swiss Topaz','Swiss topaz synonyms'),
    ('SWISS BT','TOP STONES','Swiss Topaz','Swiss topaz synonyms'),
    ('SWISS BLUE BT','TOP STONES','Swiss Topaz','Swiss topaz synonyms'),
    ('SWISS BLUE T','TOP STONES','Swiss Topaz','Swiss topaz synonyms'),
    ('LON BT','TOP STONES','London Topaz','London topaz abbreviation'),
    ('LOND BT','TOP STONES','London Topaz','London topaz abbreviation'),
    ('LONDON B T','TOP STONES','London Topaz','London topaz abbreviation'),
    ('SWIS TOPAZ','TOP STONES','Swiss Topaz','Swiss typo'),
    ('SWISSTOPAZ','TOP STONES','Swiss Topaz','Swiss typo'),
    ('SWISS B T','TOP STONES','Swiss Topaz','Swiss topaz abbreviation'),
    ('SIVS TOPAZ','TOP STONES','Swiss Topaz','Typo / phonetic Swiss'),
    ('VISTOPAZ','TOP STONES','Swiss Topaz','Typo / phonetic Swiss'),
    ('WHITE TOPAZ','TOP STONES','Other Topaz','Other Topaz'),
    ('BLUE TOPAZ','TOP STONES','Other Topaz','Other Topaz'),
    ('SKY BLUE TOPAZ','TOP STONES','Other Topaz','Other Topaz'),
    ('SKY TOPAZ','TOP STONES','Other Topaz','Other Topaz'),
    ('MLBT','TOP STONES','Other Topaz','Multi / blue topaz abbreviation'),
    ('MULTI BLUE TOPAZ','TOP STONES','Other Topaz','Other Topaz'),
    ('MULTI BT','TOP STONES','Other Topaz','Other Topaz'),
    ('BT','TOP STONES','Other Topaz','Blue Topaz abbreviation when not London/Swiss'),

    ('WHIT TOPAZ','TOP STONES','Other Topaz','Typo: White Topaz'),
    ('WHITETOPAZ','TOP STONES','Other Topaz','Typo: White Topaz'),
    ('LONODN BT','TOP STONES','London Topaz','Typo: London BT'),
    ('LONODN TOPAZ','TOP STONES','London Topaz','Typo: London Topaz'),
    ('SAPPPHIRE','TOP STONES','Blue Sapphire','Typo: Sapphire'),
    ('EMERALD','TOP STONES','Green Stones','Green stones group'),
    ('CREATED EMERALD','TOP STONES','Green Stones','Green stones group'),
    ('CREATE EMERALD','TOP STONES','Green Stones','Green stones group'),
    ('CREATED EMEALD','TOP STONES','Green Stones','Typo'),
    ('EMREAL','TOP STONES','Green Stones','Typo'),
    ('CHROME DIOPSIDE','TOP STONES','Green Stones','Green stones group'),
    ('CHROME DIOPOSIDE','TOP STONES','Green Stones','Typo'),
    ('DIOPSIDE','TOP STONES','Green Stones','Green stones group'),
    ('GREEN AGATE','TOP STONES','Green Stones','Green stones group'),
    ('GREEN AGAT','TOP STONES','Green Stones','Green stones group'),
    ('PERIDOT','TOP STONES','Green Stones','Green stones group'),
    # PEARLS
    ('BAROQUE','PEARLS','Baroque Pearl','Any baroque pearl'),
    ('AKOYA','PEARLS','Sea Pearl','Sea pearl group'),
    ('TAHITI','PEARLS','Sea Pearl','Sea pearl group'),
    ('TAHITIAN','PEARLS','Sea Pearl','Sea pearl group'),
    ('SOUTH SEA','PEARLS','Sea Pearl','Sea pearl group'),
    ('SEA PEARL','PEARLS','Sea Pearl','Sea pearl group'),
    ('MABE','PEARLS','Sea Pearl','Sea pearl group'),
    ('GALATEA','PEARLS','Sea Pearl','Sea pearl group'),
    ('FACETED','PEARLS','Sea Pearl','Sea pearl group'),
    ('FRESHWATER PEARL ROUND WHITE','PEARLS','Round White Freshwater Pearl','White round freshwater'),
    ('FRESHWATER PEARL WHITE','PEARLS','White Freshwater Pearl','White freshwater'),
    ('FRESHWATER PEARL ROSE','PEARLS','Colored Freshwater Pearl','Colored freshwater'),
    ('FRESHWATER PEARL PINK','PEARLS','Colored Freshwater Pearl','Colored freshwater'),
    ('FRESHWATER PEARL GRAY','PEARLS','Colored Freshwater Pearl','Colored freshwater'),
    ('FRESHWATER PEARL GREY','PEARLS','Colored Freshwater Pearl','Colored freshwater'),
    ('FRESHWATER PEARL BLACK','PEARLS','Colored Freshwater Pearl','Colored freshwater'),
    ('FRESHWATER PEARL ROUND BLACK','PEARLS','Colored Freshwater Pearl','Colored freshwater; round ignored for colored'),
    ('FRESHWATER PEARL GRAY ROUND','PEARLS','Colored Freshwater Pearl','Colored freshwater; round ignored for colored'),
    ('FRESHWATER PEARL ROUND GRAY','PEARLS','Colored Freshwater Pearl','Colored freshwater; round ignored for colored'),
    ('FRESHWATER PEARL ROUND ROSE','PEARLS','Colored Freshwater Pearl','Colored freshwater; round ignored for colored'),
    ('WHITE PEARL','PEARLS','White Freshwater Pearl','White freshwater'),
    # OTHER STONES
    ('SMOKY','OTHER STONES','Rauch Topaz','Smoky/Honey/Rauch are combined'),
    ('SMOKEY','OTHER STONES','Rauch Topaz','Smoky/Honey/Rauch are combined'),
    ('HONEY','OTHER STONES','Rauch Topaz','Smoky/Honey/Rauch are combined'),
    ('RAUCH','OTHER STONES','Rauch Topaz','Smoky/Honey/Rauch are combined'),
    ('GARNET','OTHER STONES','Garnet','Garnet/Rhodolite/Granada combined'),
    ('RHODOLITE','OTHER STONES','Garnet','Garnet/Rhodolite/Granada combined'),
    ('RODOLITE','OTHER STONES','Garnet','Garnet/Rhodolite/Granada combined'),
    ('GRANADA','OTHER STONES','Garnet','Garnet/Rhodolite/Granada combined'),
    ('GRANATE','OTHER STONES','Garnet','Garnet/Rhodolite/Granada combined'),
    ('PADALITE','OTHER STONES','Garnet','User synonym / possible typo'),
    ('ALL CZ','OTHER STONES','Color CZ','All CZ / CZ / zirconia combined'),
    ('CUBIC ZIRCONIA','OTHER STONES','Color CZ','All CZ / CZ / zirconia combined'),
    ('ZIRCONIA','OTHER STONES','Color CZ','All CZ / CZ / zirconia combined'),
    ('ZIRCON','OTHER STONES','Color CZ','All CZ / CZ / zirconia combined'),
    ('CZ','OTHER STONES','Color CZ','All CZ / CZ / zirconia combined'),
    ('BLACK ONYX','OTHER STONES','Onyx','All onyx combined'),
    ('ONYX','OTHER STONES','Onyx','All onyx combined'),
    ('MATT ONYX','OTHER STONES','Onyx','All onyx combined'),
    ('MATTE ONYX','OTHER STONES','Onyx','All onyx combined'),
    ('PICTURE JASPER','OTHER STONES','Jasper','All jasper combined'),
    ('JASPER','OTHER STONES','Jasper','All jasper combined'),

    ('ABALONE','OTHER STONES','Abalone','Other stones exact group'),
    ('HELIOTIS','OTHER STONES','Abalone','Heliotis / abalone shell group'),
    ('MOP','OTHER STONES','Mother of Pearl','Mother of Pearl group'),
    ('MOTHER OF PEARL','OTHER STONES','Mother of Pearl','Mother of Pearl group'),
    ('AGATE','OTHER STONES','Agate','All non-green agate combined'),
    ('BLACK AGATE','OTHER STONES','Agate','All non-green agate combined'),
    ('RED AGATE','OTHER STONES','Agate','All non-green agate combined'),
    ('AMBER','OTHER STONES','Amber','Other stones exact group'),
    ('AMETHYST','OTHER STONES','Amethyst','All amethyst combined'),
    ('GREEN AMETHYST','OTHER STONES','Amethyst','All amethyst combined'),
    ('AMMOLITE','OTHER STONES','Ammolite','Other stones exact group'),
    ('APATITE','OTHER STONES','Apatite','Other stones exact group'),
    ('AQUAMARINE','OTHER STONES','Aquamarine','Other stones exact group'),
    ('BISMUTH','OTHER STONES','Bismuth','Other stones exact group'),
    ('BLACK DIAMOND','OTHER STONES','Diamond','All diamond combined'),
    ('DIAMOND','OTHER STONES','Diamond','All diamond combined'),
    ('BLACK SPINEL','OTHER STONES','Spinel','All spinel combined'),
    ('SPINEL','OTHER STONES','Spinel','All spinel combined'),
    ('CARNELIAN','OTHER STONES','Carnelian','Other stones exact group'),
    ('CHALCEDONY','OTHER STONES','Chalcedony','Other stones exact group'),
    ('CHRYSOPRASE','OTHER STONES','Chrysoprase','Not Green Stones per user'),
    ('CITRINE','OTHER STONES','Citrine','Other stones exact group'),
    ('CORAL','OTHER STONES','Coral','Other stones exact group'),
    ('CORONDUM','OTHER STONES','Corundum','Other stones exact group'),
    ('FLUORITE','OTHER STONES','Fluorite','Other stones exact group'),
    ('QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('WHITE QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('GREEN QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('LEMON QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('ROSE QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('RUTILE QUARTZ','OTHER STONES','Quartz','Generic quartz group'),
    ('TANZANITE QUARTZ','OTHER STONES','Quartz','Tanzanite quartz as quartz group'),
    ('TANZNITE QUARTZ','OTHER STONES','Quartz','Typo: Tanzanite quartz'),
    ('GREY PARL','PEARLS','Colored Freshwater Pearl','Typo: grey pearl'),
    ('HEMATITE','OTHER STONES','Hematite','Other stones exact group'),
    ('HYPERSTHENE','OTHER STONES','Hypersthene','Other stones exact group'),
    ('IOLITE','OTHER STONES','Iolite','All iolite combined'),
    ('IHOLIT','OTHER STONES','Iolite','Typo / phonetic iolite'),
    ('JADE','OTHER STONES','Jade','Other stones exact group'),
    ('KYN','OTHER STONES','Kyanite','KYN as kyanite'),
    ('KYANITE','OTHER STONES','Kyanite','Other stones exact group'),
    ('LABRADORITE','OTHER STONES','Labradorite','Other stones exact group'),
    ('LAPISE','OTHER STONES','Lapis','Lapis typo/group'),
    ('LAPIS','OTHER STONES','Lapis','Lapis group'),
    ('LAZURITE','OTHER STONES','Lapis','Lazurite/Lapis group'),
    ('LARIMAR','OTHER STONES','Larimar','Other stones exact group'),
    ('MALACHITE','OTHER STONES','Malachite','Other stones exact group'),
    ('METEORITE','OTHER STONES','Meteorite','Other stones exact group'),
    ('MOONSTONE','OTHER STONES','Moonstone','Other stones exact group'),
    ('MORGANITE','OTHER STONES','Morganite','Other stones exact group'),
    ('OPAL','OTHER STONES','Opal','Other stones exact group'),
    ('PRENITE','OTHER STONES','Prehnite','Typo / Prehnite group'),
    ('PREHNITE','OTHER STONES','Prehnite','Other stones exact group'),
    ('PYRITE','OTHER STONES','Pyrite','Other stones exact group'),
    ('RUBELLITE','OTHER STONES','Rubellite','Other stones exact group'),
    ('SEMI','OTHER STONES','Semi Precious Mix','Semi / multi semi group'),
    ('MULTI SEMI','OTHER STONES','Semi Precious Mix','Semi / multi semi group'),
    ('SMONEY','OTHER STONES','Smoney','Other stones exact group'),
    ('SULTANITE','OTHER STONES','Sultanite','Other stones exact group'),
    ('SUN STONE','OTHER STONES','Sunstone','Other stones exact group'),
    ('SYNTHETIC Y5','OTHER STONES','Synthetic Stone','Synthetic group'),
    ('SYN.Y3','OTHER STONES','Synthetic Stone','Synthetic group'),
    ('TANZANITE','OTHER STONES','Tanzanite','Other stones exact group'),
    ('TERAHERTZ','OTHER STONES','Terahertz','Other stones exact group'),
    ('TIGER EYE','OTHER STONES','Tiger Eye','Other stones exact group'),
    ('TOURMALINE','OTHER STONES','Tourmaline','Other stones exact group'),
    ('TSAVORITE','OTHER STONES','Tsavorite','Other stones exact group'),
    ('TURQUOISE','OTHER STONES','Turquoise','Other stones exact group'),
    ('WHITE HOWLITE','OTHER STONES','Howlite','Howlite group'),
    ('HOWLITE','OTHER STONES','Howlite','Howlite group'),
    ('COLOR D','OTHER STONES','Color D','Other stones exact group'),
    ('MYSTIC TOPAZ','OTHER STONES','Mystic','Mystic is not Topaz for this report'),
    ('MYSTIC MB','OTHER STONES','Mystic','Mystic is Other Stones'),
    ('MYST MB','OTHER STONES','Mystic','Mystic is Other Stones'),
    ('MYSTIC QUARTZ','OTHER STONES','Mystic','Mystic is Other Stones'),
]

def detect_store(path: Path) -> str:
    name = path.stem.upper().replace('_', ' ').replace('-', ' ').replace('.', ' ')
    compact = re.sub(r'[^A-Z0-9]', '', name)

    # Explicit compact checks catch names like AB1.xlsx and NTR11.xlsx.
    # NTR2 must be checked before NTR1.
    if compact.startswith('NTR2'):
        return 'NTR2'
    if compact.startswith('NTR1') or compact == 'NTR':
        return 'NTR1'
    if compact.startswith('AB'):
        return 'AB'
    if compact.startswith('TT') or 'ALLSALESTT' in compact:
        return 'TT'
    if compact.startswith('SCR'):
        return 'SCR'
    if compact.startswith('20'):
        return '20'
    if compact.startswith('63') or 'MUINE' in compact:
        return '63'

    for pat, store in STORE_PATTERNS:
        if pat.search(name):
            return store
    raise ValueError(f'Не удалось определить магазин по имени файла: {path.name}')

def norm_product(text: str) -> str:
    return PRODUCT_MAP.get(' '.join(str(text).strip().split()).upper(), str(text).strip().title())

def clean_stone(raw: str) -> str:
    t = str(raw).upper()
    for a,b in [('LAB-CREATED','LAB CREATED'),('FRESH WATER','FRESHWATER'),('FWP','FRESHWATER PEARL'),('F/W','FRESHWATER'),('-', ' '),('/', ' '),(',', ' '),(';',' ')]:
        t=t.replace(a,b)
    t = re.sub(r'\bNATURAL\b','',t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def title_name(t):
    out = t.title().replace('Cz','CZ').replace('Mlbt','MLBT').replace('Bt','BT')
    return out

def classify(raw: str):
    t = clean_stone(raw)
    # Absolute priority: any moissanite mix is Moissanite.
    if re.search(r'MO+I?S+A?N+I?T|MOSSANIT|MUSSONIT', t): return 'TOP STONES','Moissanite','moissanite priority'
    if re.search(r'SAPP+H?I?R?E|SAPPPHIRE|SAPPHRIE|SAPPHIRE', t): return 'TOP STONES','Blue Sapphire','any sapphire'
    if 'RUBY' in t and 'ZOISITE' not in t and 'CIOSITE' not in t: return 'TOP STONES','Ruby','ruby'
    if ('LONDON' in t or 'LONODN' in t or 'LOND' in t) and ('TOPAZ' in t or re.search(r'\bBT\b|BLUE T', t)): return 'TOP STONES','London Topaz','London topaz'
    if ('SWISS' in t or 'SWIS' in t or 'SIVS' in t or 'VISTOPAZ' in t) and ('TOPAZ' in t or re.search(r'\bBT\b|BLUE T', t)): return 'TOP STONES','Swiss Topaz','Swiss topaz'
    if 'MYST' not in t and (any(x in t for x in ['WHITE TOPAZ','WHIT TOPAZ','WHITETOPAZ','WHITE BT','WHITE BLUE TOPAZ','BLUE TOPAZ','SKY BLUE TOPAZ','SKY TOPAZ','MLBT','MULTI BLUE TOPAZ','MULTI BT']) or re.fullmatch(r'BT', t)):
        return 'TOP STONES','Other Topaz','other topaz'
    if any(x in t for x in ['CREATED EMERALD','CREATE EMERALD','CREATED EMEALD','EMERALD','EMREAL','CHROME DIOPSIDE','CHROME DIOPOSIDE','DIOPSIDE','GREEN AGATE','GREEN AGAT','PERIDOT']):
        return 'TOP STONES','Green Stones','green stones'
    # Pearls. Colored freshwater wins over ROUND.
    if 'GREY PARL' in t: return 'PEARLS','Colored Freshwater Pearl','grey pearl typo'
    if 'PEARL' in t or 'PARL' in t or any(w in t for w in ['AKOYA','TAHITI','TAHITIAN','SOUTH SEA','MABE']):
        if 'BAROQUE' in t: return 'PEARLS','Baroque Pearl','baroque'
        if any(w in t for w in SEA_WORDS): return 'PEARLS','Sea Pearl','sea pearl'
        if any(w in t for w in COLOR_WORDS if w != 'WHITE'): return 'PEARLS','Colored Freshwater Pearl','colored freshwater'
        if 'ROUND' in t: return 'PEARLS','Round White Freshwater Pearl','round white freshwater'
        return 'PEARLS','White Freshwater Pearl','white freshwater'
    # Colored Stones: fixed compact business groups.
    if any(x in t for x in ['BLACK SPINEL','SPINEL','ONYX','OBSIDIAN','BLACK AGATE']):
        return 'COLORED STONES','Black Stones','black stones group'
    if 'AMETHYST' in t:
        return 'COLORED STONES','Amethyst','amethyst group'
    if any(x in t for x in ['MYST','CITRINE','LEMON QUARTZ','ROSE QUARTZ','GREEN QUARTZ','WHITE QUARTZ','QUARTZ','SMOKY','SMOKEY','HONEY','RAUCH']):
        return 'COLORED STONES','Quartz Group','quartz group'
    if any(x in t for x in ['RODOLITE','RHODOLITE','GARNET','GRANADA','GRANATE','PADALITE','ALMANDINE','PYROPE']):
        return 'COLORED STONES','Garnet','garnet group'
    if 'AGATE' in t or 'AGAT' in t:
        return 'COLORED STONES','Agate','non-green agate'
    return 'COLORED STONES','Other Colored Stones','other colored stones'

@dataclass
class StoreData:
    name: str
    periods: list[tuple[datetime, datetime, str]] = field(default_factory=list)
    files: list[str] = field(default_factory=list)
    data: dict = field(default_factory=lambda: defaultdict(lambda: defaultdict(lambda: {'qty':0,'amount':0.0})))
    raw_map: dict = field(default_factory=lambda: defaultdict(lambda: {'qty':0,'amount':0.0,'segment':'','column':'','clean':'','rule':''}))
    total_qty: int = 0
    total_amount: float = 0.0
    extras: dict = field(default_factory=lambda: defaultdict(lambda: {'qty':0,'amount':0.0}))
    def add_period(self, period, file_name):
        self.files.append(file_name)
        if period: self.periods.append((*period, file_name))
    def period_text(self):
        if not self.periods: return 'Period not found in source report'
        s=min(p[0] for p in self.periods); e=max(p[1] for p in self.periods)
        if len({(p[0],p[1]) for p in self.periods}) == 1:
            return f'{s:%d.%m.%Y} - {e:%d.%m.%Y}'
        return f'{s:%d.%m.%Y} - {e:%d.%m.%Y} (combined from {len(self.periods)} files)'
    def add(self, segment, stone, product, qty, amount, raw, rule):
        self.data[(segment, stone)][product]['qty'] += qty
        self.data[(segment, stone)][product]['amount'] += amount
        self.total_qty += qty; self.total_amount += amount
        r=self.raw_map[raw]
        r['qty'] += qty; r['amount'] += amount; r['segment']=segment; r['column']=stone; r['clean']=clean_stone(raw); r['rule']=rule

def to_int(v):
    try: return 0 if v is None else int(round(float(v)))
    except Exception: return 0

def to_float(v):
    try: return 0.0 if v is None else float(v)
    except Exception: return 0.0

def extract_period(ws):
    """Extract exact dates or Russian month/year range from the report header."""
    month_map = {
        'ЯНВАРЬ':1,'ЯНВАРЯ':1,'ФЕВРАЛЬ':2,'ФЕВРАЛЯ':2,'МАРТ':3,'МАРТА':3,
        'АПРЕЛЬ':4,'АПРЕЛЯ':4,'МАЙ':5,'МАЯ':5,'ИЮНЬ':6,'ИЮНЯ':6,
        'ИЮЛЬ':7,'ИЮЛЯ':7,'АВГУСТ':8,'АВГУСТА':8,'СЕНТЯБРЬ':9,'СЕНТЯБРЯ':9,
        'ОКТЯБРЬ':10,'ОКТЯБРЯ':10,'НОЯБРЬ':11,'НОЯБРЯ':11,'ДЕКАБРЬ':12,'ДЕКАБРЯ':12,
    }
    for r in range(1,10):
        for c in range(1,5):
            val=ws.cell(r,c).value
            if not val: continue
            text=str(val).strip()
            m=re.search(r'(\d{2}\.\d{2}\.\d{4})\s*[-–—]\s*(\d{2}\.\d{2}\.\d{4})', text)
            if m:
                return datetime.strptime(m.group(1),'%d.%m.%Y'), datetime.strptime(m.group(2),'%d.%m.%Y')
            upper=text.upper().replace('Ё','Е')
            m=re.search(r'([А-Я]+)\s+(20\d{2})\s*Г?\.?\s*[-–—]\s*([А-Я]+)\s+(20\d{2})', upper)
            if m and m.group(1) in month_map and m.group(3) in month_map:
                sm, sy = month_map[m.group(1)], int(m.group(2))
                em, ey = month_map[m.group(3)], int(m.group(4))
                return datetime(sy,sm,1), datetime(ey,em,calendar.monthrange(ey,em)[1])
    return None

TARGET_STORE_ALIASES = {
    'AB':'AB', 'NTR1':'NTR1', 'NTR2':'NTR2', 'SCR':'SCR', 'TT':'TT', '20':'20',
    '63NDC-RETAIL':'63', '63NDC-TIMINGS':'63', '63NDC-TIMING':'63',
    '63-RETAIL':'63', '63-TIMINGS':'63', '63':'63',
}

def normalize_store_from_report(value: str):
    """Normalize a store section from the consolidated 1C export.

    Known outlets keep stable names, while future retail stores are preserved
    instead of being silently dropped. Service sections are handled separately.
    """
    t=' '.join(str(value).strip().upper().replace('_',' ').split())
    compact=re.sub(r'[^A-ZА-Я0-9]','',t)
    if ('GIFT' in compact or 'GIFTS' in compact) and ('TT' in compact or 'ТТ' in compact): return 'GIFT TT'
    if compact in {'CAFE','КАФЕ'} or compact.startswith('CAFE') or compact.startswith('КАФЕ'): return 'CAFE'
    if '63NDC' in compact or compact.startswith('63TIM') or compact.startswith('63RETAIL'):
        return '63'
    if compact.startswith('NTR2'): return 'NTR2'
    if compact.startswith('NTR1'): return 'NTR1'
    if compact == 'AB' or compact.startswith('ABRETAIL'): return 'AB'
    if compact == 'SCR' or compact.startswith('SCRRETAIL'): return 'SCR'
    if compact == 'TT' or compact.startswith('TTRETAIL'): return 'OUTLET'
    if compact == '20' or compact.startswith('20RETAIL'): return '20'
    # Ignore common non-retail service/warehouse sections.
    if any(x in compact for x in ['RECEP','RECEPTION','STOCK','WAREHOUSE','СКЛАД','PRINCESSHANG']):
        return None
    # Future stores are not hardcoded: retain a cleaned short label.
    cleaned=re.sub(r'(RETAIL|SHOP|STORE|МАГАЗИН)','',t).strip(' -_')
    return cleaned[:24] if cleaned else None

def is_consolidated_report(ws) -> bool:
    for r in range(1,min(ws.max_row,12)+1):
        text=str(ws.cell(r,1).value or '').upper()
        if 'МАГАЗИН' in text and 'КАМЕНЬ' in text and 'НОМЕНКЛАТУРНАЯ ГРУППА' in text:
            return True
    return False

def parse_consolidated_file(path: Path) -> list[StoreData]:
    """Parse one 1C export containing all stores.

    Returns are ignored. TT becomes the OUTLET report. GIFT TT and CAFE are
    accumulated as compact auxiliary metrics on the OUTLET sheet only.
    """
    wb=load_workbook(path, data_only=True)
    ws=wb.active
    period=extract_period(ws)
    stores: dict[str, StoreData] = {}
    current_store=None
    current_aux=None
    current_stone=None
    outlet_extras=defaultdict(lambda: {'qty':0,'amount':0.0})
    for row in range(1, ws.max_row+1):
        cell=ws.cell(row,1); value=cell.value
        if value is None: continue
        text=str(value).strip()
        if not text: continue
        style_id=cell.style_id
        if style_id == 65:
            detected=normalize_store_from_report(text)
            current_stone=None
            current_aux=detected if detected in {'GIFT TT','CAFE'} else None
            current_store=None if current_aux else detected
            if current_aux:
                outlet_extras[current_aux]['qty'] = to_int(ws.cell(row,8).value)
                outlet_extras[current_aux]['amount'] = to_float(ws.cell(row,9).value)
            if current_store and current_store not in stores:
                sd=StoreData(current_store); sd.base_store=current_store
                sd.add_period(period, path.name); stores[current_store]=sd
            continue
        if style_id == 66:
            current_stone=text
            continue
        if style_id != 67 or not current_stone:
            continue
        qty=to_int(ws.cell(row,8).value); amount=to_float(ws.cell(row,9).value)
        if qty==0 and amount==0: continue
        if current_aux:
            # Auxiliary totals are taken from the store header row to avoid double counting.
            continue
        if not current_store: continue
        product=norm_product(text)
        if product.upper() in SKIP_PRODUCTS: continue
        seg, stone, rule=classify(current_stone)
        stores[current_store].add(seg,stone,product,qty,amount,current_stone,rule)
    wb.close()
    if 'OUTLET' in stores:
        for k,v in outlet_extras.items(): stores['OUTLET'].extras[k]=dict(v)
    return list(stores.values())

def preview_source(path: Path) -> tuple[str, tuple[datetime,datetime] | None]:
    wb=load_workbook(path, data_only=True, read_only=False)
    ws=wb.active
    period=extract_period(ws)
    if is_consolidated_report(ws):
        names=[]
        for row in range(1,ws.max_row+1):
            c=ws.cell(row,1)
            if c.style_id==65 and c.value:
                name=normalize_store_from_report(c.value)
                if name and name not in {'GIFT TT','CAFE'} and name not in names: names.append(name)
        wb.close()
        return ('Все магазины: ' + ', '.join(names)) if names else 'Все магазины', period
    wb.close()
    return detect_store(path), period

def parse_file(path: Path) -> StoreData:
    sd=StoreData(detect_store(path))
    wb=load_workbook(path, data_only=True)
    try:
        ws=wb.active
        sd.add_period(extract_period(ws), path.name)
        current_stone=None
        for row in range(1, ws.max_row+1):
            cell=ws.cell(row,1); value=cell.value
            if value is None: continue
            text=str(value).strip()
            if not text: continue
            upper=text.upper()
            if upper.startswith('ИТОГО') or upper.startswith('TOTAL') or 'VLADIMIR PANASIAN' in upper: continue
            if 'КАМЕНЬ' in upper or 'НОМЕНКЛАТУРНАЯ ГРУППА' in upper or upper in {'ПОСТАВЩИК(И):','ТОВАР(Ы):'} or upper.startswith('ОТЧЕТ О ПРОДАЖАХ'):
                continue
            indent=cell.alignment.indent or 0
            qty=to_int(ws.cell(row,8).value); amount=to_float(ws.cell(row,9).value)
            if indent < 1:
                # Stone row. Skip report-level product total rows before first real stone.
                if upper in PRODUCT_MAP:
                    continue
                current_stone=text
                continue
            if not current_stone: continue
            product=norm_product(text)
            if product.upper() in SKIP_PRODUCTS: continue
            if qty==0 and amount==0: continue
            seg, stone, rule=classify(current_stone)
            sd.add(seg,stone,product,qty,amount,current_stone,rule)
    finally:
        wb.close()
    return sd

def combine_stores(files):
    stores={}; errors=[]
    for p in files:
        try: sd=parse_file(p)
        except Exception as e:
            errors.append((p.name,str(e))); continue
        target=stores.setdefault(sd.name, StoreData(sd.name))
        for per in sd.periods: target.periods.append(per)
        target.files.extend(sd.files)
        for key, products in sd.data.items():
            for prod, vals in products.items():
                # use pseudo raw for aggregate not needed
                target.data[key][prod]['qty'] += vals['qty']; target.data[key][prod]['amount'] += vals['amount']
                target.total_qty += vals['qty']; target.total_amount += vals['amount']
        for raw, vals in sd.raw_map.items():
            r=target.raw_map[raw]
            r['qty']+=vals['qty']; r['amount']+=vals['amount']; r['segment']=vals['segment']; r['column']=vals['column']; r['clean']=vals['clean']; r['rule']=vals['rule']
    return stores, errors

def totals_for(store, seg=None, stone=None):
    q=0; a=0.0
    for (s, st), products in store.data.items():
        if seg and s!=seg: continue
        if stone and st!=stone: continue
        for vals in products.values(): q += vals['qty']; a += vals['amount']
    return q,a

def ordered_columns(store):
    # Fixed layout for every store; empty groups remain visible as zero columns.
    return ([('TOP STONES', n) for n in TOP_ORDER] +
            [('PEARLS', n) for n in PEARL_ORDER] +
            [('COLORED STONES', n) for n in COLORED_ORDER])

def safe_sheet_title(name): return re.sub(r'[\\/*?:\[\]]',' ',name)[:31]

def style_cell(cell,border,alignment=None,font=None,fill=None,numfmt=None):
    cell.border=border
    if alignment: cell.alignment=alignment
    if font: cell.font=font
    if fill: cell.fill=fill
    if numfmt: cell.number_format=numfmt

def add_rules_sheet(wb, stores, border, white, fill_title, center, left):
    ws=wb.create_sheet('RULES')
    headers=['RAW NAME / ALIAS','CLEANED','SEGMENT','COLUMN','RULE TYPE','QTY IN SAMPLE','SALES IN SAMPLE']
    ws.append(headers)
    for c in range(1,len(headers)+1): style_cell(ws.cell(1,c),border,center,white,fill_title)
    # manual alias rows first
    for alias,seg,col,note in MANUAL_RULES:
        ws.append([alias, clean_stone(alias), seg, col, 'manual alias: '+note, None, None])
    # exact names found in uploaded reports
    all_raw={}
    for store in stores.values():
        for raw,vals in store.raw_map.items():
            if raw not in all_raw:
                all_raw[raw]=vals.copy()
            else:
                all_raw[raw]['qty']+=vals['qty']; all_raw[raw]['amount']+=vals['amount']
    for raw in sorted(all_raw):
        vals=all_raw[raw]
        ws.append([raw, vals['clean'], vals['segment'], vals['column'], 'exact found: '+vals['rule'], vals['qty'], vals['amount']])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            style_cell(cell,border,left if cell.column in [1,2,5] else center)
            if cell.column==7: cell.number_format='# ##0'
    for c,w in enumerate([34,34,18,24,44,14,16],1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False
    return ws

def add_unknown_sheet(wb, stores, border, white, fill_title, center, left):
    ws=wb.create_sheet('UNKNOWN STONES')
    headers=['RAW NAME','CLEANED','PROPOSED SEGMENT','PROPOSED COLUMN','QTY','SALES','COMMENT']
    ws.append(headers)
    for c in range(1,len(headers)+1): style_cell(ws.cell(1,c),border,center,white,fill_title)
    rows=[]
    for store in stores.values():
        for raw, vals in store.raw_map.items():
            if vals['rule'].startswith('fallback: own name'):
                rows.append([raw, vals['clean'], vals['segment'], vals['column'], vals['qty'], vals['amount'], 'Review and add exact rule if needed'])
    if not rows:
        ws.append(['No unknown stones in this sample','','','','','',''])
    else:
        for row in sorted(rows, key=lambda r:r[0]): ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            style_cell(cell,border,left if cell.column in [1,2,7] else center)
            if cell.column==6: cell.number_format='# ##0'
    for c,w in enumerate([34,34,18,24,12,16,32],1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False

def build_report(stores, output):
    wb=Workbook(); ws=wb.active; ws.title='SUMMARY'
    thin=Side(style='thin', color=COLORS['GRID']); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    bold=Font(bold=True); white=Font(bold=True,color='FFFFFF')
    center=Alignment(horizontal='center',vertical='center',wrap_text=True); left=Alignment(horizontal='left',vertical='center',wrap_text=True)
    fill_title=PatternFill('solid', fgColor=COLORS['TITLE']); fill_total=PatternFill('solid', fgColor=COLORS['TOTAL'])
    fill_warn=PatternFill('solid', fgColor=COLORS['WARN'])
    seg_fills={seg:PatternFill('solid', fgColor=SEG_COLORS[seg]) for seg in SEG_ORDER}

    headers=['Store','Period','Files','Total PCS','Total Sales','Top Stones PCS %','Top Stones Sales %','Pearls PCS %','Pearls Sales %','Other Stones PCS %','Other Stones Sales %']
    ws.append(headers)
    for c in range(1,len(headers)+1): style_cell(ws.cell(1,c),border,center,white,fill_title)
    for store in [stores[k] for k in sorted(stores)]:
        row=[store.name, store.period_text(), ', '.join(store.files), store.total_qty, store.total_amount]
        for seg in SEG_ORDER:
            q,a=totals_for(store,seg=seg)
            row += [q/store.total_qty if store.total_qty else 0, a/store.total_amount if store.total_amount else 0]
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            style_cell(cell,border,left if cell.column in [1,2,3] else center)
            if cell.column==5: cell.number_format='# ##0'
            if cell.column>=6: cell.number_format='0.00%'
    for i,w in enumerate([12,34,40,12,16,18,18,16,16,20,20],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False

    for store in [stores[k] for k in sorted(stores)]:
        ws=wb.create_sheet(safe_sheet_title(store.name)); cols=ordered_columns(store); ncols=max(5,1+len(cols)*2)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
        ws.cell(1,1).value=f'{store.name} — Stone Sales Report'
        style_cell(ws.cell(1,1),border,center,Font(bold=True,size=16,color='FFFFFF'),fill_title)
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
        ws.cell(2,1).value=f'Report period: {store.period_text()}'
        style_cell(ws.cell(2,1),border,left,Font(bold=True,size=11),fill_total)
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=ncols)
        ws.cell(3,1).value=f'Source files: {", ".join(store.files)}'
        style_cell(ws.cell(3,1),border,left,Font(italic=True,size=10),None)
        ws.cell(4,1).value='Product / Total'; style_cell(ws.cell(4,1),border,center,white,fill_title)
        col=2; i=0
        while i < len(cols):
            seg=cols[i][0]; start=col
            while i < len(cols) and cols[i][0]==seg:
                i+=1; col+=2
            end=col-1
            ws.merge_cells(start_row=4,start_column=start,end_row=4,end_column=end)
            ws.cell(4,start).value=seg.title(); style_cell(ws.cell(4,start),border,center,white,seg_fills[seg])
            for cc in range(start,end+1): style_cell(ws.cell(4,cc),border,center,white,seg_fills[seg])
        col=2
        for seg, stone in cols:
            ws.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
            ws.cell(5,col).value=stone
            for cc in [col,col+1]: style_cell(ws.cell(5,cc),border,center,white,seg_fills[seg])
            ws.cell(6,col).value='PCS'; ws.cell(6,col+1).value='Sales'
            for cc in [col,col+1]: style_cell(ws.cell(6,cc),border,center,white,seg_fills[seg])
            col+=2
        row=7
        for product in PRODUCT_ORDER:
            has_any=any(store.data[k].get(product,{}).get('qty',0) or store.data[k].get(product,{}).get('amount',0) for k in cols)
            if product in ['Stone','Other'] and not has_any: continue
            ws.cell(row,1).value=product; style_cell(ws.cell(row,1),border,left,bold)
            col=2
            for key in cols:
                vals=store.data[key].get(product, {'qty':0,'amount':0.0})
                ws.cell(row,col).value=vals['qty'] or None; ws.cell(row,col+1).value=vals['amount'] or None
                style_cell(ws.cell(row,col),border,center)
                style_cell(ws.cell(row,col+1),border,center,None,None,'# ##0')
                col+=2
            row+=1
        for label,kind in [('TOTAL PCS','qty'),('TOTAL SALES','amount'),('% OF STORE PCS','qty_pct'),('% OF STORE SALES','amount_pct')]:
            ws.cell(row,1).value=label; style_cell(ws.cell(row,1),border,left,bold,fill_total)
            col=2
            for seg,stone in cols:
                q,a=totals_for(store,seg,stone)
                if kind=='qty': values=[q,None]; fmts=['0','0']
                elif kind=='amount': values=[None,a]; fmts=['# ##0','# ##0']
                elif kind=='qty_pct': values=[q/store.total_qty if store.total_qty else 0,None]; fmts=['0.00%','0.00%']
                else: values=[None,a/store.total_amount if store.total_amount else 0]; fmts=['0.00%','0.00%']
                for idx,cc in enumerate([col,col+1]):
                    ws.cell(row,cc).value=values[idx]; style_cell(ws.cell(row,cc),border,center,bold,fill_total,fmts[idx])
                col+=2
            row+=1
        row+=2
        ws.cell(row,1).value='Segment analysis'; ws.cell(row,1).font=Font(bold=True,size=13); row+=1
        for c,h in enumerate(['Segment','PCS','Sales','PCS % of store','Sales % of store'],1):
            ws.cell(row,c).value=h; style_cell(ws.cell(row,c),border,center,white,fill_title)
        row+=1
        for seg in SEG_ORDER:
            q,a=totals_for(store,seg=seg); vals=[seg.title(),q,a,q/store.total_qty if store.total_qty else 0,a/store.total_amount if store.total_amount else 0]
            for c,v in enumerate(vals,1):
                ws.cell(row,c).value=v; style_cell(ws.cell(row,c),border,left if c==1 else center,white if c==1 else None,seg_fills[seg] if c==1 else None)
                if c==3: ws.cell(row,c).number_format='# ##0'
                if c>=4: ws.cell(row,c).number_format='0.00%'
            row+=1
        ws.freeze_panes='B7'; ws.sheet_view.showGridLines=False
        ws.column_dimensions['A'].width=24
        for c in range(2,ncols+1): ws.column_dimensions[get_column_letter(c)].width=12 if c%2==0 else 15
        for r in range(1,row+1): ws.row_dimensions[r].height=22

    add_rules_sheet(wb, stores, border, white, fill_title, center, left)
    wb.save(output)

def write_project():
    if PROJECT_DIR.exists(): shutil.rmtree(PROJECT_DIR)
    (PROJECT_DIR/'reports').mkdir(parents=True); (PROJECT_DIR/'src').mkdir()
    for f in INPUT_DIR.glob('*.xlsx'):
        if f.name.lower() in ['20.xlsx','63.1.xlsx','63.2.xlsx','ab.xlsx','ntr1.xlsx','ntr2.xlsx','scr.xlsx','tt.xlsx']:
            shutil.copy(f, PROJECT_DIR/'reports'/f.name)
    main_py = '''from pathlib import Path\nfrom src.report import run\n\nBASE = Path(__file__).resolve().parent\nrun(input_dir=BASE / "reports", output_file=BASE / "StoneReport_final.xlsx")\nprint("Готово: StoneReport_final.xlsx")\n'''
    report_py=Path(__file__).read_text(encoding='utf-8')
    report_py=report_py.replace("INPUT_DIR = Path.cwd()", "INPUT_DIR = Path.cwd()")
    report_py += "\n\ndef run(input_dir: Path, output_file: Path):\n    files=[p for p in input_dir.iterdir() if p.suffix.lower() in ('.xlsx','.xls') and not p.name.startswith('~$')]\n    stores, errors = combine_stores(files)\n    if errors:\n        print('Ошибки определения/чтения файлов:')\n        for name, err in errors: print(name, err)\n    if not stores:\n        raise RuntimeError('Не найдено подходящих отчетов в папке reports')\n    build_report(stores, output_file)\n    return output_file\n"
    (PROJECT_DIR/'main.py').write_text(main_py,encoding='utf-8')
    (PROJECT_DIR/'src'/'__init__.py').write_text('',encoding='utf-8')
    (PROJECT_DIR/'src'/'report.py').write_text(report_py,encoding='utf-8')
    (PROJECT_DIR/'README.txt').write_text('''StoneReport v4\n\nКак пользоваться:\n1. Положите Excel-отчеты магазинов в папку reports.\n2. Названия файлов могут быть свободными: AB New.xlsx, AB Jewelry Sales Report.xlsx, NTR 1 New.xlsx, All Sales TT.xlsx.\n3. Запустите main.py.\n4. На выходе появится StoneReport_final.xlsx.\n\nЧто внутри:\n- SUMMARY с периодом каждого магазина и процентами PCS/Sales.\n- Листы магазинов с периодом отчета сверху.\n- RULES: большой список соответствий RAW -> Column.\n- UNKNOWN STONES убран: все названия сразу классифицируются в нужные колонки.\n\nКлючевые правила:\n- Moissanite имеет абсолютный приоритет, даже если в названии есть Ruby/CZ/Sapphire.\n- Все Sapphire -> Blue Sapphire.\n- Green Stones: Emerald, Created Emerald, Chrome Diopside, Green Agate, Peridot.\n- Mystic Topaz / Mystic MB -> Other Stones, не Topaz.\n- Onyx variants -> Onyx. Jasper variants -> Jasper.\n- 63.1 и 63.2 объединяются в магазин 63.\n''',encoding='utf-8')
    if ZIP_OUT.exists(): ZIP_OUT.unlink()
    with zipfile.ZipFile(ZIP_OUT,'w',zipfile.ZIP_DEFLATED) as z:
        for p in PROJECT_DIR.rglob('*'):
            z.write(p, p.relative_to(PROJECT_DIR.parent))

if __name__=='__main__':
    files=[p for p in INPUT_DIR.glob('*.xlsx') if p.name.lower() in ['20.xlsx','63.1.xlsx','63.2.xlsx','ab.xlsx','ntr1.xlsx','ntr2.xlsx','scr.xlsx','tt.xlsx']]
    stores, errors=combine_stores(files)
    print('stores', sorted(stores), 'errors', errors)
    for s in sorted(stores):
        sd=stores[s]
        print(s, sd.period_text(), sd.total_qty, sd.total_amount, len(sd.data), 'raw',len(sd.raw_map))
    build_report(stores, OUTPUT)
    write_project()
    print('created', OUTPUT, ZIP_OUT)


def run(input_dir: Path, output_file: Path):
    files=[p for p in input_dir.iterdir() if p.suffix.lower() in ('.xlsx','.xls') and not p.name.startswith('~$')]
    stores, errors = combine_stores(files)
    if errors:
        print('Ошибки определения/чтения файлов:')
        for name, err in errors: print(name, err)
    if not stores:
        raise RuntimeError('Не найдено подходящих отчетов в папке reports')
    build_report(stores, output_file)
    return output_file

# ---- EXE / GUI entry points ----
def run_files(files: list[Path], output_file: Path):
    files = [Path(p) for p in files if Path(p).suffix.lower() in ('.xlsx', '.xls') and not Path(p).name.startswith('~$')]
    stores, errors = combine_stores(files)
    if errors:
        details = '\n'.join([f'{name}: {err}' for name, err in errors])
        raise RuntimeError('Ошибки определения/чтения файлов:\n' + details)
    if not stores:
        raise RuntimeError('Не найдено подходящих Excel-отчетов')
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    build_report(stores, output_file)
    return output_file


def run(input_dir: Path, output_file: Path):
    input_dir = Path(input_dir)
    files = [p for p in input_dir.iterdir() if p.suffix.lower() in ('.xlsx', '.xls') and not p.name.startswith('~$')]
    return run_files(files, output_file)


# ============================================================
# Analitika 1.1.0: separate periods, comparisons and charts
# ============================================================

def _period_key(sd: StoreData):
    if sd.periods:
        return sd.periods[0][0], sd.periods[0][1]
    return None, None


def _period_short(sd: StoreData) -> str:
    start, end = _period_key(sd)
    if start and end:
        if start.year == end.year and start.month == end.month and start.day == 1:
            return start.strftime('%m.%Y')
        return f'{start:%d.%m.%y}-{end:%d.%m.%y}'
    return Path(sd.files[0]).stem[:18] if sd.files else 'period'


def _merge_store_data(target: StoreData, source: StoreData):
    target.periods.extend(source.periods)
    target.files.extend(source.files)
    for key, products in source.data.items():
        for prod, vals in products.items():
            target.data[key][prod]['qty'] += vals['qty']
            target.data[key][prod]['amount'] += vals['amount']
            target.total_qty += vals['qty']
            target.total_amount += vals['amount']
    for raw, vals in source.raw_map.items():
        r = target.raw_map[raw]
        r['qty'] += vals['qty']; r['amount'] += vals['amount']
        r['segment'] = vals['segment']; r['column'] = vals['column']
        r['clean'] = vals['clean']; r['rule'] = vals['rule']
    for k,v in source.extras.items():
        target.extras[k]['qty'] += v['qty']; target.extras[k]['amount'] += v['amount']


def build_report_units(files):
    """Build report units from either legacy one-store files or the new all-stores export.
    One consolidated file creates separate units for every store inside it.
    Multiple periods remain separate and are compared on COMPARE sheets.
    """
    parsed=[]; errors=[]
    for p in files:
        p=Path(p)
        try:
            wb=load_workbook(p, data_only=True, read_only=False)
            ws=wb.active
            consolidated=is_consolidated_report(ws)
            header_text=' '.join(str(ws.cell(rr,1).value or '') for rr in range(1,min(ws.max_row,7)+1)).upper()
            supplier_only=('ПОСТАВЩИК' in header_text and 'НОМЕНКЛАТУРНАЯ ГРУППА' in header_text and 'МАГАЗИН' not in header_text)
            wb.close()
            if supplier_only:
                # This file is consumed by the web Suppliers module, not by store-sales parsing.
                continue
            if consolidated:
                parsed.extend(parse_consolidated_file(p))
            else:
                sd=parse_file(p); sd.base_store=sd.name; parsed.append(sd)
        except Exception as e:
            errors.append((p.name, str(e)))

    # Legacy 63.1/63.2 files with the same period are merged. In a consolidated
    # export both 63 sections are already accumulated into one StoreData.
    units=[]; used=set()
    for i,sd in enumerate(parsed):
        if i in used: continue
        if getattr(sd,'base_store',sd.name)=='63':
            start,end=_period_key(sd)
            merged=StoreData('63'); merged.base_store='63'; _merge_store_data(merged,sd); used.add(i)
            for j,other in enumerate(parsed[i+1:],start=i+1):
                if j in used: continue
                if getattr(other,'base_store',other.name)=='63' and _period_key(other)==(start,end):
                    # Merge only pieces originating from the same source file, or legacy 63.1/63.2 pair.
                    same_source=bool(set(sd.files)&set(other.files))
                    legacy_pair=all(re.search(r'63[ ._-]?[12]', f, re.I) for f in (sd.files+other.files))
                    if same_source or legacy_pair:
                        _merge_store_data(merged,other); used.add(j)
            units.append(merged)
        else:
            units.append(sd); used.add(i)

    counts=defaultdict(int); result={}
    for sd in units:
        base=getattr(sd,'base_store',sd.name)
        label=f'{base} — {_period_short(sd)}'
        counts[label]+=1
        if counts[label]>1: label=f'{label} ({counts[label]})'
        sd.name=label; result[label]=sd
    return result,errors


def _add_summary_charts(wb):
    ws=wb['SUMMARY']
    n=ws.max_row
    if n < 2: return
    # Revenue chart
    chart=BarChart(); chart.type='col'; chart.style=10
    chart.title='Sales by report and period'; chart.y_axis.title='Sales'; chart.x_axis.title='Store / period'
    chart.height=8; chart.width=18
    chart.add_data(Reference(ws,min_col=5,min_row=1,max_row=n), titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n))
    ws.add_chart(chart,'M2')
    # Quantity chart
    qty=BarChart(); qty.type='col'; qty.style=11
    qty.title='Quantity by report and period'; qty.y_axis.title='PCS'; qty.x_axis.title='Store / period'
    qty.height=8; qty.width=18
    qty.add_data(Reference(ws,min_col=4,min_row=1,max_row=n), titles_from_data=True)
    qty.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n))
    ws.add_chart(qty,'M18')
    # Segment revenue share (100% stacked)
    seg=BarChart(); seg.type='bar'; seg.grouping='stacked'; seg.overlap=100; seg.style=12
    seg.title='Revenue mix by segment'; seg.x_axis.title='Share'; seg.y_axis.title='Store / period'
    seg.height=10; seg.width=18
    seg.add_data(Reference(ws,min_col=7,max_col=11,min_row=1,max_row=n), titles_from_data=True, from_rows=False)
    # only sales pct columns 7,9,11 are desired; openpyxl cannot non-contiguous easily, rebuild series manually
    seg.series=[]
    for col in (7,9,11):
        data=Reference(ws,min_col=col,min_row=1,max_row=n)
        seg.add_data(data,titles_from_data=True)
    seg.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n))
    ws.add_chart(seg,'M34')
    for col in range(13,31): ws.column_dimensions[get_column_letter(col)].width=12


def _add_store_charts(wb, stores):
    for label, store in stores.items():
        title=safe_sheet_title(label)
        if title not in wb.sheetnames: continue
        ws=wb[title]
        # Locate segment analysis header
        header_row=None
        for row in range(1,ws.max_row+1):
            if ws.cell(row,1).value=='Segment':
                header_row=row; break
        if not header_row: continue
        start=header_row+1; end=min(start+2,ws.max_row)
        pie=PieChart(); pie.title='Sales share by segment'; pie.height=8; pie.width=10
        pie.add_data(Reference(ws,min_col=3,min_row=header_row,max_row=end),titles_from_data=True)
        pie.set_categories(Reference(ws,min_col=1,min_row=start,max_row=end))
        pie.dataLabels=DataLabelList(); pie.dataLabels.showPercent=True
        ws.add_chart(pie,'G'+str(header_row))


def _comparison_sheet(wb, base_store, reports):
    if len(reports)<2: return
    reports=sorted(reports,key=lambda s: (_period_key(s)[0] or datetime.min, s.name))
    name=safe_sheet_title(f'COMPARE {base_store}')
    ws=wb.create_sheet(name)
    headers=['Period','Files','PCS','Sales','Δ PCS','Δ PCS %','Δ Sales','Δ Sales %','Top Stones Sales %','Pearls Sales %','Other Stones Sales %']
    ws.append(headers)
    for i,sd in enumerate(reports, start=2):
        prev=reports[i-3] if i>2 else None
        vals=[sd.period_text(),', '.join(sd.files),sd.total_qty,sd.total_amount]
        if prev:
            dq=sd.total_qty-prev.total_qty; da=sd.total_amount-prev.total_amount
            vals += [dq, dq/prev.total_qty if prev.total_qty else 0, da, da/prev.total_amount if prev.total_amount else 0]
        else: vals += [None,None,None,None]
        for seg in SEG_ORDER:
            _,a=totals_for(sd,seg=seg); vals.append(a/sd.total_amount if sd.total_amount else 0)
        ws.append(vals)
    thin=Side(style='thin',color=COLORS['GRID']); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    fill=PatternFill('solid',fgColor=COLORS['TITLE']); white=Font(bold=True,color='FFFFFF'); center=Alignment(horizontal='center',vertical='center',wrap_text=True)
    for c in range(1,len(headers)+1): style_cell(ws.cell(1,c),border,center,white,fill)
    for row in ws.iter_rows(min_row=2):
        for cell in row: style_cell(cell,border,center)
        for c in (4,7): row[c-1].number_format='# ##0'
        for c in (6,8,9,10,11): row[c-1].number_format='0.00%'
    for c,w in enumerate([25,38,12,16,12,12,16,12,18,16,20],1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False
    n=ws.max_row
    line=LineChart(); line.title=f'{base_store}: sales by period'; line.y_axis.title='Sales'; line.x_axis.title='Period'; line.height=8; line.width=16
    line.add_data(Reference(ws,min_col=4,min_row=1,max_row=n),titles_from_data=True); line.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n)); ws.add_chart(line,'M2')
    qty=LineChart(); qty.title=f'{base_store}: quantity by period'; qty.y_axis.title='PCS'; qty.x_axis.title='Period'; qty.height=8; qty.width=16
    qty.add_data(Reference(ws,min_col=3,min_row=1,max_row=n),titles_from_data=True); qty.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n)); ws.add_chart(qty,'M18')
    mix=LineChart(); mix.title=f'{base_store}: segment mix'; mix.y_axis.title='Share'; mix.x_axis.title='Period'; mix.height=9; mix.width=16
    mix.add_data(Reference(ws,min_col=9,max_col=11,min_row=1,max_row=n),titles_from_data=True); mix.set_categories(Reference(ws,min_col=1,min_row=2,max_row=n)); ws.add_chart(mix,'M34')


def build_report_v110(stores, output):
    # Use the proven v4 writer, then enrich the workbook.
    build_report(stores, output)
    wb=load_workbook(output)
    _add_summary_charts(wb)
    _add_store_charts(wb, stores)
    grouped=defaultdict(list)
    for sd in stores.values(): grouped[getattr(sd,'base_store',sd.name.split(' — ')[0])].append(sd)
    for store,reports in sorted(grouped.items()): _comparison_sheet(wb,store,reports)
    # Put comparison sheets after SUMMARY.
    summary=wb['SUMMARY']; wb._sheets.remove(summary); wb._sheets.insert(0,summary)
    wb.save(output)


def run_files(files: list[Path], output_file: Path):
    files=[Path(p) for p in files if Path(p).suffix.lower() in ('.xlsx','.xls') and not Path(p).name.startswith('~$')]
    stores,errors=build_report_units(files)
    if errors:
        details='\n'.join([f'{name}: {err}' for name,err in errors])
        raise RuntimeError('Ошибки определения/чтения файлов:\n'+details)
    if not stores: raise RuntimeError('Не найдено подходящих Excel-отчетов')
    output_file=Path(output_file); output_file.parent.mkdir(parents=True,exist_ok=True)
    build_report_v110(stores,output_file)
    return output_file


def run(input_dir: Path, output_file: Path):
    input_dir=Path(input_dir)
    files=[p for p in input_dir.iterdir() if p.suffix.lower() in ('.xlsx','.xls') and not p.name.startswith('~$')]
    return run_files(files,output_file)

# ============================================================
# Analitika 1.1.3 RC: Russian executive dashboards
# ============================================================

def _ru_segment(seg: str) -> str:
    return {'TOP STONES':'TOP STONES','PEARLS':'PEARLS','COLORED STONES':'COLORED STONES'}.get(seg, seg)


def _segment_totals(store):
    out={}
    for seg in SEG_ORDER:
        q,a=totals_for(store,seg=seg); out[seg]={'qty':q,'amount':a}
    return out


def _apply_chart_colors(chart, colors):
    try:
        from openpyxl.chart.series import DataPoint
        chart.series[0].data_points=[DataPoint(idx=i, graphPr={'solidFill':c}) for i,c in enumerate(colors)]
    except Exception:
        pass


def _make_pie(ws, title, data_col, header_row, first_row, last_row, anchor):
    chart=PieChart(); chart.title=title; chart.height=8.2; chart.width=11.5
    chart.add_data(Reference(ws,min_col=data_col,min_row=header_row,max_row=last_row),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=1,min_row=first_row,max_row=last_row))
    chart.dataLabels=DataLabelList(); chart.dataLabels.showCatName=True; chart.dataLabels.showPercent=True
    chart.dataLabels.showVal=False; chart.legend=None
    _apply_chart_colors(chart,[SEG_COLORS[x] for x in SEG_ORDER])
    ws.add_chart(chart,anchor)


def _conclusions(store, network_avg_price=None, network_seg_sales=None):
    seg=_segment_totals(store); lines=[]
    if store.total_amount:
        leader=max(SEG_ORDER,key=lambda x:seg[x]['amount'])
        share=seg[leader]['amount']/store.total_amount
        lines.append(f"Основную выручку формирует {_ru_segment(leader)} — {share:.1%}.")
    if network_seg_sales and store.total_amount:
        top_share=seg['TOP STONES']['amount']/store.total_amount
        delta=top_share-network_seg_sales.get('TOP STONES',0)
        if abs(delta)>=0.03:
            direction='выше' if delta>0 else 'ниже'
            lines.append(f"Доля TOP STONES {direction} средней по сети на {abs(delta):.1%}.")
    tq,ta=totals_for(store,'TOP STONES','Moissanite')
    top_amount=seg['TOP STONES']['amount']
    if ta and store.total_amount:
        lines.append(f"Moissanite: {ta/store.total_amount:.1%} продаж магазина и {ta/top_amount:.1%} категории TOP STONES.")
    _,ga=totals_for(store,'TOP STONES','Green Stones')
    if store.total_amount and ga/store.total_amount<0.02:
        lines.append(f"Green Stones имеют минимальную долю — {ga/store.total_amount:.1%}.")
    avg=store.total_amount/store.total_qty if store.total_qty else 0
    if network_avg_price:
        delta=avg/network_avg_price-1
        lines.append(f"Средняя стоимость изделия {'выше' if delta>=0 else 'ниже'} средней по сети на {abs(delta):.1%}.")
    return lines[:5]


def build_executive_report(stores, output):
    wb=Workbook(); summary=wb.active; summary.title='SUMMARY'
    thin=Side(style='thin',color='D9D9D9'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    center=Alignment(horizontal='center',vertical='center',wrap_text=True); left=Alignment(horizontal='left',vertical='center',wrap_text=True)
    white=Font(color='FFFFFF',bold=True); bold=Font(bold=True); title_font=Font(size=22,bold=True,color='132451')
    fills={seg:PatternFill('solid',fgColor=SEG_COLORS[seg]) for seg in SEG_ORDER}
    navy=PatternFill('solid',fgColor='132451'); pale=PatternFill('solid',fgColor='F3F6FA'); total_fill=PatternFill('solid',fgColor='DDEBF7')
    stores_list=[stores[k] for k in sorted(stores)]
    total_qty=sum(s.total_qty for s in stores_list); total_sales=sum(s.total_amount for s in stores_list)
    periods=[p for s in stores_list for p in s.periods]
    period_text='не определён'
    if periods: period_text=f"{min(p[0] for p in periods):%d.%m.%Y} — {max(p[1] for p in periods):%d.%m.%Y}"
    sources=sorted({f for s in stores_list for f in s.files})
    avg_price=total_sales/total_qty if total_qty else 0
    # Summary title and meta
    summary.merge_cells('A1:C1'); summary['A1']='SUMMARY'; summary['A1'].font=title_font; summary['A1'].alignment=left
    summary['A2']='Период отчёта:'; summary['B2']=period_text; summary['A3']='Источник данных:'; summary['B3']=', '.join(sources)
    summary['A2'].font=summary['A3'].font=bold
    # KPI cards
    cards=[('D1','E2','МАГАЗИНОВ',len(stores_list),'7030A0'),('F1','G2','ВСЕГО ШТ.',total_qty,'4472C4'),('H1','I2','ОБЩАЯ ВЫРУЧКА',total_sales,'548235'),('J1','K2','СРЕДНЯЯ СТОИМОСТЬ',avg_price,'BF7B00')]
    for tl,br,label,val,color in cards:
        min_col=summary[tl].column; min_row=summary[tl].row; max_col=summary[br].column; max_row=summary[br].row
        summary.merge_cells(start_row=min_row,start_column=min_col,end_row=max_row,end_column=max_col)
        c=summary.cell(min_row,min_col); c.value=f"{label}\n{val:,.0f}".replace(',',' '); c.font=Font(size=14,bold=True,color=color); c.alignment=center; c.fill=PatternFill('solid',fgColor='F8F9FC'); c.border=border
    # Main table
    r=6
    headers=['Магазин','Период','Источник','Шт.','Продажи (VND)','TOP STONES\nШт. %','TOP STONES\nПродажи %','PEARLS\nШт. %','PEARLS\nПродажи %','COLORED STONES\nШт. %','COLORED STONES\nПродажи %']
    for c,h in enumerate(headers,1):
        cell=summary.cell(r,c,h); cell.font=white; cell.alignment=center; cell.border=border
        if c<=5: cell.fill=navy
        elif c<=7: cell.fill=fills['TOP STONES']
        elif c<=9: cell.fill=fills['PEARLS']
        else: cell.fill=fills['COLORED STONES']
    data_start=r+1
    for st in stores_list:
        vals=[st.name.split(' — ')[0],st.period_text(),', '.join(st.files),st.total_qty,st.total_amount]
        seg=_segment_totals(st)
        for sg in SEG_ORDER:
            vals += [seg[sg]['qty']/st.total_qty if st.total_qty else 0, seg[sg]['amount']/st.total_amount if st.total_amount else 0]
        r+=1
        for c,v in enumerate(vals,1):
            cell=summary.cell(r,c,v); cell.border=border; cell.alignment=left if c<=3 else center
            if c in (4,5): cell.number_format='# ##0'
            if c>=6: cell.number_format='0.00%'
            if c in (6,7): cell.fill=PatternFill('solid',fgColor='F1E7F8')
            elif c in (8,9): cell.fill=PatternFill('solid',fgColor='FFF4CC')
            elif c in (10,11): cell.fill=PatternFill('solid',fgColor='E6F0DF')
    # total row
    r+=1; summary.cell(r,1,'ИТОГО'); summary.cell(r,4,total_qty); summary.cell(r,5,total_sales)
    seg_all={sg:{'qty':sum(_segment_totals(s)[sg]['qty'] for s in stores_list),'amount':sum(_segment_totals(s)[sg]['amount'] for s in stores_list)} for sg in SEG_ORDER}
    c=6
    for sg in SEG_ORDER:
        summary.cell(r,c,seg_all[sg]['qty']/total_qty if total_qty else 0); summary.cell(r,c+1,seg_all[sg]['amount']/total_sales if total_sales else 0); c+=2
    for c in range(1,12):
        cell=summary.cell(r,c); cell.font=bold; cell.fill=total_fill; cell.border=border; cell.alignment=center
        if c in (4,5): cell.number_format='# ##0'
        if c>=6: cell.number_format='0.00%'
    table_end=r
    chart_row=max(15,table_end+3)
    # Stable dashboard charts are rendered as PNG images instead of native Excel charts.
    # Native openpyxl charts can shift or overlap depending on the Excel version,
    # display scaling and local font metrics. Embedded images remain pixel-perfect.
    chart_tmp = Path(tempfile.mkdtemp(prefix='analitika_charts_'))
    chart_anchors = {'TOP STONES':'A15', 'PEARLS':'E15', 'COLORED STONES':'I15'}
    light_colors={'TOP STONES':'B78ED2','PEARLS':'FFE08A','COLORED STONES':'A9D18E'}
    seg_cols={'TOP STONES':(6,7),'PEARLS':(8,9),'COLORED STONES':(10,11)}

    def _font(size, bold=False):
        candidates = [
            'C:/Windows/Fonts/arialbd.ttf' if bold else 'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/calibrib.ttf' if bold else 'C:/Windows/Fonts/calibri.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]
        for candidate in candidates:
            try:
                return ImageFont.truetype(candidate, size)
            except Exception:
                pass
        return ImageFont.load_default()

    def _draw_segment_chart(segment, first_col, second_col, path):
        width, height = 410, 300
        img = Image.new('RGB', (width, height), 'white')
        d = ImageDraw.Draw(img)
        title_font_img = _font(20, True)
        label_font = _font(12, False)
        value_font = _font(11, True)
        small_font = _font(10, False)
        main = '#' + SEG_COLORS[segment]
        light = '#' + light_colors[segment]
        # Card border and title
        d.rounded_rectangle((2,2,width-3,height-3), radius=14, outline=main, width=2, fill='white')
        bbox=d.textbbox((0,0), segment, font=title_font_img)
        d.text(((width-(bbox[2]-bbox[0]))/2, 13), segment, fill=main, font=title_font_img)
        # Legend
        d.rectangle((125,47,137,59), fill=main); d.text((142,45),'Шт. %',fill='#222222',font=small_font)
        d.rectangle((215,47,227,59), fill=light); d.text((232,45),'Продажи %',fill='#222222',font=small_font)
        stores_names=[str(summary.cell(rr,1).value) for rr in range(data_start, table_end)]
        pcs=[float(summary.cell(rr,first_col).value or 0) for rr in range(data_start, table_end)]
        sales=[float(summary.cell(rr,second_col).value or 0) for rr in range(data_start, table_end)]
        plot_left, plot_right, plot_top, plot_bottom = 42, width-15, 75, 245
        # horizontal grid 0..100%
        for tick in range(0,101,20):
            y=plot_bottom-(plot_bottom-plot_top)*tick/100
            d.line((plot_left,y,plot_right,y), fill='#D9D9D9', width=1)
            d.text((8,y-7),f'{tick}%',fill='#555555',font=small_font)
        n=max(1,len(stores_names)); group_w=(plot_right-plot_left)/n
        bar_w=min(20,group_w*0.26)
        for i,name in enumerate(stores_names):
            center_x=plot_left+group_w*(i+0.5)
            for j,(val,color) in enumerate(((pcs[i],main),(sales[i],light))):
                x0=center_x + (-bar_w-2 if j==0 else 2)
                x1=x0+bar_w
                y0=plot_bottom-(plot_bottom-plot_top)*max(0,min(1,val))
                d.rectangle((x0,y0,x1,plot_bottom), fill=color)
                txt=f'{val:.2%}'.replace('.',',')
                tb=d.textbbox((0,0),txt,font=value_font)
                d.text(((x0+x1-(tb[2]-tb[0]))/2,max(plot_top,y0-16)),txt,fill=main,font=value_font)
            nb=d.textbbox((0,0),name,font=label_font)
            d.text((center_x-(nb[2]-nb[0])/2,plot_bottom+8),name,fill='#222222',font=label_font)
        img.save(path, quality=95)

    for sg in SEG_ORDER:
        first,last=seg_cols[sg]
        png=chart_tmp/(sg.lower().replace(' ','_')+'.png')
        _draw_segment_chart(sg, first, last, png)
        xl_img=XLImage(str(png))
        xl_img.width=410; xl_img.height=300
        summary.add_image(xl_img, chart_anchors[sg])
    # Segment analysis total: keep it directly beneath the chart row.
    sr=37
    summary.cell(sr,1,'СТРУКТУРА СЕГМЕНТОВ (ИТОГО)').font=Font(size=13,bold=True,color='132451')
    hdr=sr+1
    cols=1
    for sg in SEG_ORDER:
        summary.merge_cells(start_row=hdr,start_column=cols,end_row=hdr,end_column=cols+3)
        c=summary.cell(hdr,cols,sg); c.fill=fills[sg]; c.font=white; c.alignment=center
        for j,h in enumerate(['Шт.','Продажи (VND)','Шт. %','Продажи %']):
            cc=summary.cell(hdr+1,cols+j,h); cc.fill=fills[sg]; cc.font=white; cc.alignment=center; cc.border=border
        vals=[seg_all[sg]['qty'],seg_all[sg]['amount'],seg_all[sg]['qty']/total_qty if total_qty else 0,seg_all[sg]['amount']/total_sales if total_sales else 0]
        for j,v in enumerate(vals):
            cc=summary.cell(hdr+2,cols+j,v); cc.border=border; cc.alignment=center; cc.font=bold
            cc.number_format='0.00%' if j>=2 else '# ##0'
        cols+=4
    summary.sheet_view.showGridLines=False; summary.freeze_panes='A7'
    widths=[14,28,28,12,18,15,16,15,16,18,20]
    for i,w in enumerate(widths,1): summary.column_dimensions[get_column_letter(i)].width=w

    # Network references for store conclusions
    network_avg=avg_price
    network_seg={sg:(seg_all[sg]['amount']/total_sales if total_sales else 0) for sg in SEG_ORDER}
    # Store sheets
    for st in stores_list:
        base=st.name.split(' — ')[0]
        ws=wb.create_sheet(safe_sheet_title(st.name))
        ws.merge_cells('A1:H1'); ws['A1']=f'МАГАЗИН: {base}'; ws['A1'].font=title_font; ws['A1'].alignment=left
        ws['A2']='Период отчёта:'; ws['B2']=st.period_text(); ws['D2']='Источник данных:'; ws['E2']=', '.join(st.files)
        ws['A2'].font=ws['D2'].font=bold
        # KPI cards 3 only
        av=st.total_amount/st.total_qty if st.total_qty else 0
        kpis=[('A4','B5','ПРОДАЖИ',st.total_amount,'7030A0'),('C4','D5','ПРОДАНО ИЗДЕЛИЙ',st.total_qty,'548235'),('E4','F5','СРЕДНЯЯ СТОИМОСТЬ',av,'BF7B00')]
        for tl,br,label,val,color in kpis:
            a=ws[tl]; b=ws[br]; ws.merge_cells(start_row=a.row,start_column=a.column,end_row=b.row,end_column=b.column)
            c=ws.cell(a.row,a.column); c.value=f"{label}\n{val:,.0f}".replace(',',' '); c.font=Font(size=13,bold=True,color=color); c.alignment=center; c.border=border; c.fill=pale
        # AI conclusions to the right
        ws.merge_cells('J4:N4'); ws['J4']='ВЫВОДЫ'; ws['J4'].font=Font(size=14,bold=True,color='132451'); ws['J4'].alignment=left
        for i,line in enumerate(_conclusions(st,network_avg,network_seg),5):
            ws.merge_cells(start_row=i,start_column=10,end_row=i,end_column=14)
            ws.cell(i,10).value='• '+line; ws.cell(i,10).alignment=Alignment(wrap_text=True,vertical='top'); ws.cell(i,10).fill=pale
        # Vertical stone table
        row=8
        headers=['Сегмент','Камень','Продажи (VND)','% от продаж магазина','Количество (шт.)','% от количества магазина','Средняя стоимость изделия','% продаж внутри сегмента']
        for c,h in enumerate(headers,1):
            cell=ws.cell(row,c,h); cell.fill=navy; cell.font=white; cell.alignment=center; cell.border=border
        row+=1
        for sg,names in [('TOP STONES',TOP_ORDER),('PEARLS',PEARL_ORDER),('COLORED STONES',COLORED_ORDER)]:
            sg_amount=_segment_totals(st)[sg]['amount']
            for name in names:
                q,a=totals_for(st,sg,name); vals=[sg,name,a,a/st.total_amount if st.total_amount else 0,q,q/st.total_qty if st.total_qty else 0,a/q if q else 0,a/sg_amount if sg_amount else 0]
                for c,v in enumerate(vals,1):
                    cell=ws.cell(row,c,v); cell.border=border; cell.alignment=left if c==2 else center
                    if c==1: cell.fill=fills[sg]; cell.font=white
                    elif sg=='TOP STONES': cell.fill=PatternFill('solid',fgColor='F5EEF9')
                    elif sg=='PEARLS': cell.fill=PatternFill('solid',fgColor='FFF9E6')
                    else: cell.fill=PatternFill('solid',fgColor='EEF5EA')
                    if c in (3,5,7): cell.number_format='# ##0'
                    if c in (4,6,8): cell.number_format='0.00%'
                row+=1
        # Total
        vals=['ИТОГО','',st.total_amount,1,st.total_qty,1,av,1]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row,c,v); cell.fill=navy; cell.font=white; cell.border=border; cell.alignment=center
            if c in (3,5,7): cell.number_format='# ##0'
            if c in (4,6,8): cell.number_format='0.00%'
        table_end=row
        # Segment analysis source for charts
        seg_header=table_end+3
        ws.cell(seg_header,1,'Сегмент'); ws.cell(seg_header,2,'Количество'); ws.cell(seg_header,3,'Продажи')
        for c in range(1,4): ws.cell(seg_header,c).font=white; ws.cell(seg_header,c).fill=navy; ws.cell(seg_header,c).border=border
        for i,sg in enumerate(SEG_ORDER,seg_header+1):
            q,a=totals_for(st,seg=sg); ws.cell(i,1,sg); ws.cell(i,2,q); ws.cell(i,3,a)
            for c in range(1,4): ws.cell(i,c).border=border
            ws.cell(i,1).fill=fills[sg]; ws.cell(i,1).font=white; ws.cell(i,2).number_format='# ##0'; ws.cell(i,3).number_format='# ##0'
        _make_pie(ws,'СТРУКТУРА ПРОДАЖ',3,seg_header,seg_header+1,seg_header+3,'E'+str(seg_header))
        _make_pie(ws,'СТРУКТУРА КОЛИЧЕСТВА',2,seg_header,seg_header+1,seg_header+3,'J'+str(seg_header))
        # OUTLET auxiliary blocks only
        if base=='OUTLET' and st.extras:
            outrow=seg_header+18
            ws.cell(outrow,1,'ДОПОЛНИТЕЛЬНЫЕ ПОДРАЗДЕЛЕНИЯ OUTLET').font=Font(size=13,bold=True,color='132451')
            outrow+=1
            for name in ('GIFT TT','CAFE'):
                v=st.extras.get(name,{'qty':0,'amount':0.0}); avg=v['amount']/v['qty'] if v['qty'] else 0
                ws.cell(outrow,1,name); ws.cell(outrow,2,'Продажи'); ws.cell(outrow,3,v['amount']); ws.cell(outrow,4,'Количество'); ws.cell(outrow,5,v['qty']); ws.cell(outrow,6,'Средняя стоимость'); ws.cell(outrow,7,avg)
                for c in range(1,8): ws.cell(outrow,c).border=border; ws.cell(outrow,c).alignment=center
                ws.cell(outrow,1).font=bold; ws.cell(outrow,1).fill=PatternFill('solid',fgColor='E4DFEC' if name=='GIFT TT' else 'FCE4D6')
                ws.cell(outrow,3).number_format=ws.cell(outrow,5).number_format=ws.cell(outrow,7).number_format='# ##0'
                outrow+=2
        # Footer
        fr=max(ws.max_row+3,seg_header+18)
        ws.merge_cells(start_row=fr,start_column=1,end_row=fr,end_column=14); ws.cell(fr,1).value='Analitika 1.1.5 RC  |  Princess Jewelry  |  Разработка: Vladimir Panasyan'; ws.cell(fr,1).font=Font(color='FFFFFF',bold=True); ws.cell(fr,1).fill=navy; ws.cell(fr,1).alignment=center
        widths=[18,28,18,20,18,22,22,20,3,18,18,18,18,18]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes='A9'; ws.sheet_view.showGridLines=False
    # Rules as actual encountered mapping
    rules=wb.create_sheet('ПРАВИЛА')
    rh=['Исходное название','Очищенное название','Сегмент','Итоговая группа','Количество','Продажи']
    for c,h in enumerate(rh,1): rules.cell(1,c,h).fill=navy; rules.cell(1,c).font=white; rules.cell(1,c).border=border; rules.cell(1,c).alignment=center
    combined={}
    for st in stores_list:
        for raw,v in st.raw_map.items():
            key=raw; x=combined.setdefault(key,dict(v));
            if x is not v: x['qty']+=v['qty']; x['amount']+=v['amount']
    for r,(raw,v) in enumerate(sorted(combined.items()),2):
        vals=[raw,v['clean'],v['segment'],v['column'],v['qty'],v['amount']]
        for c,val in enumerate(vals,1): rules.cell(r,c,val).border=border
        rules.cell(r,5).number_format=rules.cell(r,6).number_format='# ##0'
    for i,w in enumerate([36,36,20,26,14,18],1): rules.column_dimensions[get_column_letter(i)].width=w
    rules.freeze_panes='A2'; rules.sheet_view.showGridLines=False
    output=Path(output); output.parent.mkdir(parents=True,exist_ok=True); wb.save(output)
    try:
        shutil.rmtree(chart_tmp, ignore_errors=True)
    except Exception:
        pass


def _stabilize_workbook_layout(wb):
    """Make the workbook render consistently across Excel installations."""
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 85 if ws.title == 'SUMMARY' else 80
        ws.sheet_view.zoomScaleNormal = ws.sheet_view.zoomScale
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_format.defaultRowHeight = 19
        ws.sheet_format.defaultColWidth = 11
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        # A single font family prevents text metrics from changing between PCs.
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                current = cell.font
                cell.font = Font(
                    name='Arial',
                    size=current.sz or 10,
                    bold=current.b,
                    italic=current.i,
                    color=current.color,
                    underline=current.u,
                )
                if isinstance(cell.value, int):
                    cell.number_format = '# ##0'
                elif isinstance(cell.value, float) and '%' not in str(cell.number_format):
                    cell.number_format = '# ##0'
        # Keep headers and body rows visually stable.
        for r in range(1, ws.max_row + 1):
            if ws.row_dimensions[r].height is None:
                ws.row_dimensions[r].height = 20
        if ws.title == 'SUMMARY':
            for r,h in {1:34,2:22,3:22,6:28,7:24}.items():
                ws.row_dimensions[r].height=h
        elif ws.title != 'ПРАВИЛА' and not ws.title.startswith('COMPARE'):
            for r,h in {1:32,2:22,4:30,5:30,8:30}.items():
                ws.row_dimensions[r].height=h


def build_report_v110(stores, output):
    build_executive_report(stores, output)
    # Comparisons remain available for multiple periods.
    wb=load_workbook(output)
    grouped=defaultdict(list)
    for sd in stores.values(): grouped[getattr(sd,'base_store',sd.name.split(' — ')[0])].append(sd)
    for store,reports in sorted(grouped.items()): _comparison_sheet(wb,store,reports)
    _stabilize_workbook_layout(wb)
    wb.save(output)
