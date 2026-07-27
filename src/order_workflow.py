from __future__ import annotations

import hashlib
import io
import json
from html import escape
import math
import posixpath
import re
import sqlite3
import time
import uuid
from dataclasses import dataclass, field, replace
from contextlib import closing
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from functools import lru_cache
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile, ZIP_STORED

from PIL import Image, ImageChops, ImageOps, UnidentifiedImageError
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.navigation import NavigationItem, render_mobile_navigation, render_sidebar
from src.diagnostics import diagnostic_event, timed_operation
from src.order_persistence import (
    CloudStorageError,
    get_cloud_storage,
    get_cloud_storage_status,
    load_storage_config,
)


ORDER_MODE_STONES = "Камни"
ORDER_MODE_PEARLS = "Жемчуг"
ORDER_MODES = (ORDER_MODE_STONES, ORDER_MODE_PEARLS)

RECOMMENDATION_BASE = "Базовые рекомендации"
RECOMMENDATION_SEASONAL = "Сезонные рекомендации"
RECOMMENDATION_PROFILES = (RECOMMENDATION_BASE, RECOMMENDATION_SEASONAL)
DEFAULT_REPORT_MONTHS = 4

DELIVERY_STATUS_SENT = "sent"
DELIVERY_STATUS_APPROVED = "approved"
DELIVERY_STATUS_IN_PROGRESS = "in_progress"
DELIVERY_STATUS_RECEIVED = "received"
DELIVERY_STATUS_LABELS = {
    DELIVERY_STATUS_SENT: "Заказ отправлен",
    DELIVERY_STATUS_APPROVED: "Заказ согласован",
    DELIVERY_STATUS_IN_PROGRESS: "В работе",
    DELIVERY_STATUS_RECEIVED: "Получен",
}
DELIVERY_STATUSES = tuple(DELIVERY_STATUS_LABELS)
DELIVERY_DATE_FIELDS = {
    DELIVERY_STATUS_SENT: "sent_at",
    DELIVERY_STATUS_APPROVED: "approved_at",
    DELIVERY_STATUS_IN_PROGRESS: "in_progress_at",
    DELIVERY_STATUS_RECEIVED: "received_at",
}
DELIVERY_DATE_LABELS = {
    DELIVERY_STATUS_SENT: "Отправлен",
    DELIVERY_STATUS_APPROVED: "Согласован",
    DELIVERY_STATUS_IN_PROGRESS: "В работе",
    DELIVERY_STATUS_RECEIVED: "Получен",
}


def normalize_delivery_status(value: object, *, received: object = False) -> str:
    status = str(value or "").strip()
    if status in DELIVERY_STATUS_LABELS:
        return status
    return DELIVERY_STATUS_RECEIVED if bool(received) else DELIVERY_STATUS_SENT


def _date_only(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return raw[:10] if len(raw) >= 10 else raw


def normalize_delivery_dates(
    value: object,
    *,
    order_date: object = "",
    received_at: object = "",
    status: object = DELIVERY_STATUS_SENT,
    status_updated_at: object = "",
) -> dict[str, str]:
    raw = dict(value) if isinstance(value, dict) else {}
    result = {field: _date_only(raw.get(field, "")) for field in DELIVERY_DATE_FIELDS.values()}
    if not result["sent_at"]:
        result["sent_at"] = _date_only(order_date)
    normalized_status = normalize_delivery_status(status)
    active_field = DELIVERY_DATE_FIELDS[normalized_status]
    if not result[active_field]:
        fallback = received_at if normalized_status == DELIVERY_STATUS_RECEIVED else status_updated_at
        result[active_field] = _date_only(fallback)
    return result


def delivery_status_rank(status: object) -> int:
    return list(DELIVERY_STATUSES).index(normalize_delivery_status(status))


def validate_delivery_timeline(status: object, dates: dict[str, str]) -> None:
    normalized_status = normalize_delivery_status(status)
    required_statuses = DELIVERY_STATUSES[: delivery_status_rank(normalized_status) + 1]
    parsed: list[date] = []
    for stage in required_statuses:
        field = DELIVERY_DATE_FIELDS[stage]
        raw = _date_only(dates.get(field, ""))
        if not raw:
            raise ValueError(f"Укажите дату этапа «{DELIVERY_STATUS_LABELS[stage]}».")
        try:
            parsed.append(date.fromisoformat(raw))
        except ValueError as exc:
            raise ValueError(f"Некорректная дата этапа «{DELIVERY_STATUS_LABELS[stage]}».") from exc
    if any(current < previous for previous, current in zip(parsed, parsed[1:])):
        raise ValueError("Даты этапов должны идти по порядку: отправка → согласование → работа → получение.")


def delivery_history_text(dates: dict[str, str], status: object) -> str:
    pieces: list[str] = []
    active_rank = delivery_status_rank(status)
    for index, stage in enumerate(DELIVERY_STATUSES):
        raw = _date_only(dates.get(DELIVERY_DATE_FIELDS[stage], ""))
        if raw and index <= active_rank:
            try:
                formatted = date.fromisoformat(raw).strftime("%d.%m.%Y")
            except ValueError:
                formatted = raw
            pieces.append(f"{DELIVERY_DATE_LABELS[stage]}: {formatted}")
    return " · ".join(pieces)

# Lock catalogue supplied in ``backside butterfly for studs.xlsx``.
# English labels are written to the supplier Excel; Russian labels are shown
# only in the order interface.
EARRING_LOCKS: dict[str, tuple[str, str]] = {
    "A": ("English Lock", "Английский замок"),
    "B": ("Pin + Omega Clip", "Штифт с омега-зажимом"),
    "C": ("Hook", "Крючок"),
    "D": ("Post + Pin", "Пусета со штифтом"),
    "E": ("New Lock", "Новый замок"),
    "M": ("Casting Screw Stud", "Литая винтовая пусета"),
    "Q": ("Casting Small Stud — Flower", "Маленькая литая пусета «Цветок»"),
    "T": ("New Screw Stud", "Новая винтовая пусета"),
    "R": ("Casting Small Stud — Normal", "Маленькая литая обычная пусета"),
}

# These stones use the same full recommendation scale as pearls. All other
# coloured stones receive a one-unit reduction after the base recommendation
# has been calculated. Topaz and sapphire are matched by name below so every
# normalized variety is covered.
FULL_RECOMMENDATION_STONES = frozenset({
    "Ruby",
    "Moissanite",
    "Emerald",
    "Created Emerald",
    "Onyx",
    "Black Spinel",
    "Green Agate",
})

CATEGORY_TOP = "Топы продаж"
CATEGORY_MEDIUM = "Средние комплекты"
CATEGORY_WEAK = "Слабые комплекты"
CATEGORY_ZERO = "Нулевые комплекты"
CATEGORY_ORDER = (CATEGORY_TOP, CATEGORY_MEDIUM, CATEGORY_WEAK, CATEGORY_ZERO)
CATEGORY_SHORT = {
    CATEGORY_TOP: "Топы",
    CATEGORY_MEDIUM: "Средние",
    CATEGORY_WEAK: "Слабые",
    CATEGORY_ZERO: "Нулевые",
}
CATEGORY_TONE = {
    CATEGORY_TOP: "🔴",
    CATEGORY_MEDIUM: "🟠",
    CATEGORY_WEAK: "🟡",
    CATEGORY_ZERO: "⚪",
}

SAPPHIRE_ORDER_GROUP = "Sapphire"
RUBY_ORDER_GROUP = "Ruby"
MOISSANITE_ORDER_GROUP = "Moissanite"
TOPAZ_ORDER_GROUP = "Topaz"
GREEN_STONES_GROUP = "Green Stones"
OTHER_STONES_GROUP = "Other Stones"

STONE_ORDER_BUCKET_ORDER = (
    SAPPHIRE_ORDER_GROUP,
    RUBY_ORDER_GROUP,
    MOISSANITE_ORDER_GROUP,
    TOPAZ_ORDER_GROUP,
    GREEN_STONES_GROUP,
    OTHER_STONES_GROUP,
)
PEARL_ORDER_BUCKET_ORDER = ("White", "Grey", "Pink", "Black", "Baroque")

# Concrete values retained for analytics/recommendations. Supplier-order
# navigation intentionally uses only six business-facing stone sections.
GREEN_STONE_NAMES = frozenset({
    "Emerald",
    "Created Emerald",
    "Chrome Diopside",
    "Green Agate",
    "Peridot",
})
ORDER_GREEN_STONE_NAMES = frozenset({
    "Emerald",
    "Created Emerald",
    "Red Emerald",
    "Rhombium",
    "Chrome Diopside",
    "Green Agate",
    "Garnet",
    "Peridot",
})

# Kept as a compatibility alias for older tests/integrations. In the order UI
# all supported topaz variants now live under the single Topaz section.
OTHER_TOPAZ_GROUP = TOPAZ_ORDER_GROUP
OTHER_TOPAZ_NAMES = frozenset({
    "White Topaz",
    "Blue Topaz",
    "Sky Blue Topaz",
    "Multi Blue Topaz",
})
ORDER_TOPAZ_NAMES = frozenset({
    "London Topaz",
    "Swiss Topaz",
    "Azure Topaz",
    *OTHER_TOPAZ_NAMES,
})

DIRECT_ORDER_STONE_NAMES = frozenset({
    "Blue Sapphire",
    "Blue Sapphire High Quality",
    "Blue Sapphire Medium Quality",
    "Ruby",
    "Moissanite",
})
OTHER_STONE_NAMES = frozenset({
    "Abalone",
    "Mother Of Pearl",
    "Agate",
    "Amber",
    "Amethyst",
    "Green Amethyst",
    "Ammolite",
    "Apatite",
    "Aquamarine",
    "Bismuth",
    "Diamond",
    "Black Spinel",
    "Spinel",
    "Carnelian",
    "Chalcedony",
    "Chrysoprase",
    "Citrine",
    "Coral",
    "Corundum",
    "Fluorite",
    "Quartz",
    "White Quartz",
    "Green Quartz",
    "Lemon Quartz",
    "Rose Quartz",
    "Rutile Quartz",
    "Smoky",
    "Honey",
    "Mystic Quartz",
    "Garnet",
    "Rhodolite",
    "Onyx",
    "Obsidian",
    "Jasper",
    "Dalmatian Jasper",
    "Red Jasper",
    "Hematite",
    "Hypersthene",
    "Iolite",
    "Jade",
    "Kyanite",
    "Labradorite",
    "Lapis Lazurite",
    "Larimar",
    "Malachite",
    "Meteorite",
    "Moonstone",
    "Morganite",
    "Opal",
    "Prehnite",
    "Pyrite",
    "Rubellite",
    "Sultanite",
    "Sunstone",
    "Tanzanite",
    "Terahertz",
    "Tiger Eye",
    "Tourmaline",
    "Tsavorite",
    "Turquoise",
    "White Howlite",
})

UNRECOGNIZED_STONE = "Камень не распознан"

ANALYTICS_GROUP_ORDER = ("Earrings", "Ring", "Pendant", "Bracelet", "Necklace")
ANALYTICS_GROUP_LABELS = {
    "Earrings": "Серьги",
    "Ring": "Кольца",
    "Pendant": "Подвески",
    "Bracelet": "Браслеты",
    "Necklace": "Ожерелья",
    "Не указана": "Не указана",
}

TOP_STONE_ANALYTICS_FAMILIES: tuple[tuple[str, frozenset[str]], ...] = (
    (
        "Blue Sapphire — все вариации",
        frozenset({"Blue Sapphire", "Blue Sapphire High Quality", "Blue Sapphire Medium Quality"}),
    ),
    ("Ruby", frozenset({"Ruby"})),
    ("Moissanite", frozenset({"Moissanite"})),
    ("London Topaz", frozenset({"London Topaz"})),
    ("Swiss Topaz", frozenset({"Swiss Topaz"})),
    ("Other Topaz", OTHER_TOPAZ_NAMES),
    ("Green Stones", GREEN_STONE_NAMES),
)

QUARTZ_ANALYTICS_NAMES = frozenset({
    "Quartz",
    "White Quartz",
    "Green Quartz",
    "Lemon Quartz",
    "Rose Quartz",
    "Rutile Quartz",
    "Mystic Quartz",
    "Citrine",
    "Smoky",
    "Honey",
})
BLACK_STONE_ANALYTICS_NAMES = frozenset({"Onyx", "Black Spinel"})
GARNET_ANALYTICS_NAMES = frozenset({"Garnet", "Rhodolite"})
JASPER_ANALYTICS_NAMES = frozenset({"Jasper", "Dalmatian Jasper", "Red Jasper"})

COLORED_STONE_ANALYTICS_FAMILY_ORDER = (
    "Quartz Group",
    "Black Stones",
    "Garnet / Rhodolite",
    "Jasper",
    "CZ",
    "Other Colored Stones",
    "Unrecognized",
)

PEARL_ANALYTICS_FAMILY_ORDER = (
    "Sea Pearls",
    "Baroque Pearls",
    "White Freshwater",
    "Colored Freshwater",
    "Other Pearls",
)

RING_SIZES = tuple(range(15, 25))
DRAFT_VERSION = 6

ROOT_DIR = Path(__file__).resolve().parents[1]
RUNTIME_DIR = ROOT_DIR / "data" / "order_runtime"
UPLOAD_DIR = RUNTIME_DIR / "uploads"
DRAFT_DB = RUNTIME_DIR / "order_drafts.sqlite3"
ORDER_EXCLUSIONS_FILE = ROOT_DIR / "data" / "order_exclusions.json"

_XML_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XML_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_XML_PACKAGE_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_XML_DRAWING = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
_XML_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

# Goods that are mounted internally and therefore must not be ordered as ready items.
STONE_EXCLUSION_PATTERNS: tuple[str, ...] = (
    "BSHQ",
    "BSMQ",
    "BLUE SAPPHIRE HIGH QUALITY",
    "BLUE SAPPHIRE MEDIUM QUALITY",
    "BLUE SAPPHIRE HQ",
    "BLUE SAPPHIRE MQ",
    "EMERALD HIGH QUALITY",
    "EMERALD HQ",
)

SEA_PEARL_PATTERNS: tuple[str, ...] = (
    "SEA PEARL",
    "SOUTH SEA",
    "AKOYA",
    "TAHITI",
    "TAHITIAN",
    "GALATEA",
    "FACETED SEA",
)

ORDER_SECTIONS = (
    ("order-overview", "Сводка"),
    ("order-workspace", "Комплекты"),
    ("order-rings", "Размеры колец"),
    ("order-export", "Excel"),
)


@dataclass(frozen=True)
class OrderItem:
    row: int
    set_id: str
    sku: str
    stone: str
    group: str
    sales: int
    stock_63: int
    stock_20: int
    stores: dict[str, int]
    total_stock: int
    working_stock: int
    ntr2_stock: int
    ntr2_calculated: bool
    tvp_raw: int
    stock_tt: int = 0
    stock_tt_warehouse: int = 0
    stock_princess_hang: int = 0
    report_months: int = DEFAULT_REPORT_MONTHS
    eligible_store_count: int = 0
    image_path: str | None = None
    ungrouped: bool = False
    visual_match_set_id: str | None = None
    visual_match_sku: str | None = None
    visual_match_category: str | None = None
    visual_match_score: float = 0.0
    visual_match_status: str | None = None
    duplicate_sku: str | None = None
    duplicate_score: float = 0.0
    duplicate_status: str | None = None
    duplicate_reason: str | None = None
    duplicate_preferred: bool = False
    errors: tuple[str, ...] = ()

    @property
    def key(self) -> str:
        return f"{self.set_id}|{self.sku}|{self.row}"

    @property
    def positive_tvp(self) -> int:
        return max(0, self.tvp_raw)

    @property
    def display_stock(self) -> int:
        """Stock shown in the card: report total minus store 20 only.

        TT/Outlet and store 63 remain included and are shown separately as
        detail cubes. ``working_stock`` stays the internal recommendation
        value and must not be presented as the overall stock.
        """
        return max(0, self.total_stock - self.stock_20)

    @property
    def is_ring(self) -> bool:
        value = normalize_text(self.group)
        if value in {"RING", "RINGS", "КОЛЬЦО", "КОЛЬЦА"}:
            return True
        return bool(re.search(r"(^|[^A-ZА-Я])RINGS?($|[^A-ZА-Я])", value))

    @property
    def is_earrings(self) -> bool:
        return canonical_group(self.group) == "Earrings"

    @property
    def is_pendant(self) -> bool:
        return canonical_group(self.group) == "Pendant"


@dataclass(frozen=True)
class OrderSet:
    key: str
    set_id: str
    stone: str
    items: tuple[OrderItem, ...]
    category: str
    driver_sku: str
    max_sales: int
    has_positive_tvp: bool
    has_negative_tvp: bool
    zero_segment: str | None = None
    is_ungrouped: bool = False


@dataclass(frozen=True)
class OrderRecommendation:
    quantity: int
    reasons: tuple[str, ...] = ()
    blocked_by_tvp: bool = False
    rule: str = "none"
    transfers: tuple[str, ...] = ()


@dataclass(frozen=True)
class ParsedOrderWorkbook:
    source_name: str
    source_hash: str
    upload_path: str
    period: str
    supplier: str
    store_columns: tuple[str, ...]
    has_actual_ntr2: bool
    items: tuple[OrderItem, ...]
    report_months: int = DEFAULT_REPORT_MONTHS
    warnings: tuple[str, ...] = ()


@dataclass
class OrderDraft:
    source_hash: str
    source_name: str
    mode: str
    version: int = DRAFT_VERSION
    orders: dict[str, int] = field(default_factory=dict)
    sizes: dict[str, dict[str, int]] = field(default_factory=dict)
    stock_checked: dict[str, bool] = field(default_factory=dict)
    manual_edit: dict[str, bool] = field(default_factory=dict)
    limited_orders: dict[str, bool] = field(default_factory=dict)
    lock_changes: dict[str, str] = field(default_factory=dict)
    recommendation_profile: str = RECOMMENDATION_BASE
    stage: str = "order"
    selected_stone: str = ""
    status: str = "draft"
    created_at: str = ""
    updated_at: str = ""

    def touch(self) -> None:
        now = datetime.now().isoformat(timespec="seconds")
        if not self.created_at:
            self.created_at = now
        self.updated_at = now

    def as_payload(self) -> dict[str, Any]:
        self.touch()
        return {
            "version": self.version,
            "source_hash": self.source_hash,
            "source_name": self.source_name,
            "mode": self.mode,
            # Persist only meaningful values. A report can contain thousands of
            # rows; omitting zeros makes each cloud autosave small and fast.
            "orders": {str(k): int(v) for k, v in self.orders.items() if int(v) > 0},
            "sizes": {
                str(k): {str(size): int(qty) for size, qty in values.items() if int(qty) > 0}
                for k, values in self.sizes.items()
                if any(int(qty) > 0 for qty in values.values())
            },
            "stock_checked": {str(k): bool(v) for k, v in self.stock_checked.items() if bool(v)},
            "manual_edit": {str(k): bool(v) for k, v in self.manual_edit.items() if bool(v)},
            "limited_orders": {str(k): bool(v) for k, v in self.limited_orders.items() if bool(v)},
            "lock_changes": {
                str(k): str(v)
                for k, v in self.lock_changes.items()
                if str(v) in EARRING_LOCKS
            },
            "recommendation_profile": (
                self.recommendation_profile
                if self.recommendation_profile in RECOMMENDATION_PROFILES
                else RECOMMENDATION_BASE
            ),
            "stage": self.stage,
            "selected_stone": self.selected_stone,
            "status": "completed" if self.status == "completed" else "draft",
            "created_at": self.created_at,
            "updated_at": self.updated_at,
        }


@dataclass(frozen=True)
class ManualTransitOrder:
    """A lightweight manually entered supplier order used for delivery tracking."""

    order_id: str
    title: str
    order_date: str
    note: str = ""
    quantity: int = 0
    delivery_status: str = DELIVERY_STATUS_SENT
    delivery_dates: dict[str, str] = field(default_factory=dict)
    created_at: str = ""
    updated_at: str = ""
    received_at: str = ""
    status_updated_at: str = ""
    storage: str = "local"

    @property
    def received(self) -> bool:
        return normalize_delivery_status(self.delivery_status) == DELIVERY_STATUS_RECEIVED

    def as_payload(self) -> dict[str, Any]:
        now = datetime.now().isoformat(timespec="seconds")
        created_at = self.created_at or now
        delivery_status = normalize_delivery_status(self.delivery_status)
        delivery_dates = normalize_delivery_dates(
            self.delivery_dates,
            order_date=self.order_date,
            received_at=self.received_at,
            status=delivery_status,
            status_updated_at=self.status_updated_at,
        )
        active_field = DELIVERY_DATE_FIELDS[delivery_status]
        if not delivery_dates.get(active_field):
            delivery_dates[active_field] = _date_only(self.status_updated_at or now)
        validate_delivery_timeline(delivery_status, delivery_dates)
        received_at = delivery_dates.get("received_at", "") if delivery_status == DELIVERY_STATUS_RECEIVED else ""
        return {
            "schema_version": 3,
            "order_id": self.order_id,
            "title": self.title.strip(),
            "order_date": self.order_date,
            "note": self.note.strip(),
            "quantity": max(0, safe_int(self.quantity)),
            "delivery_status": delivery_status,
            "delivery_dates": delivery_dates,
            "status_updated_at": self.status_updated_at or now,
            # Compatibility field for data created by Analitika 2.0 before
            # the four-stage status workflow was introduced.
            "received": delivery_status == DELIVERY_STATUS_RECEIVED,
            "created_at": created_at,
            "updated_at": now,
            "received_at": received_at,
        }


@dataclass(frozen=True)
class SavedOrderWorkspace:
    """A resumable order workspace stored locally or in durable object storage."""

    source_hash: str
    source_name: str
    upload_path: str
    updated_at: str
    modes: tuple[str, ...]
    preferred_mode: str
    selected_positions: int
    total_quantity: int
    storage: str = "local"
    created_at: str = ""
    limited_positions: int = 0
    status: str = "draft"
    mode_details: dict[str, dict[str, Any]] = field(default_factory=dict)


# ---------------------------- pure business logic ----------------------------

@lru_cache(maxsize=1)
def load_order_exclusions() -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "stone_patterns": list(STONE_EXCLUSION_PATTERNS),
        "pearl_patterns": list(SEA_PEARL_PATTERNS),
        "exclude_round_pearl": True,
    }
    try:
        payload = json.loads(ORDER_EXCLUSIONS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError):
        return defaults
    if not isinstance(payload, dict):
        return defaults
    stone_patterns = payload.get("stone_patterns", defaults["stone_patterns"])
    pearl_patterns = payload.get("pearl_patterns", defaults["pearl_patterns"])
    return {
        "stone_patterns": [normalize_text(value) for value in stone_patterns if normalize_text(value)],
        "pearl_patterns": [normalize_text(value) for value in pearl_patterns if normalize_text(value)],
        "exclude_round_pearl": bool(payload.get("exclude_round_pearl", True)),
    }

def normalize_text(value: object) -> str:
    return " ".join(str(value or "").strip().upper().replace("Ё", "Е").split())


def safe_int(value: object) -> int:
    try:
        if value is None or str(value).strip() == "":
            return 0
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _display_stone_name(text: str) -> str:
    value = text.title()
    replacements = {
        "Cz": "CZ",
        "Bt": "BT",
        "Mlbt": "MLBT",
        "Hq": "HQ",
        "Mq": "MQ",
        "Mop": "MOP",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


STONE_EXACT_ALIASES: dict[str, str] = {
    "BS": "BLUE SAPPHIRE",
    "BSHQ": "BLUE SAPPHIRE HIGH QUALITY",
    "BSMQ": "BLUE SAPPHIRE MEDIUM QUALITY",
    "LBT": "LONDON TOPAZ",
    "SWBT": "SWISS TOPAZ",
    "WBT": "WHITE TOPAZ",
    "BT": "BLUE TOPAZ",
    "CE": "CREATED EMERALD",
    "CD": "CHROME DIOPSIDE",
    "EM": "EMERALD",
    "PERI": "PERIDOT",
    "AMST": "AMETHYST",
    "CIT": "CITRINE",
    "RDT": "RHODOLITE",
    "GARN": "GARNET",
    "APAT": "APATITE",
    "TANZ": "TANZANITE",
    "IOL": "IOLITE",
    "WHO": "WHITE HOWLITE",
    "GAM": "GREEN AMETHYST",
    "DJ": "DALMATIAN JASPER",
    "OB": "OBSIDIAN",
    "RJ": "RED JASPER",
    "GA": "GREEN AGATE",
    "GREEN AGAT": "GREEN AGATE",
    "LAP": "LAPIS LAZURITE",
    "AMA": "КАМЕНЬ НЕ РАСПОЗНАН",
    "FPW": "FRESHWATER PEARL WHITE",
    "FPC": "FRESHWATER PEARL COLORED",
    "TAH": "TAHITI PEARL",
    "SSP": "SOUTH SEA PEARL",
}


_GROUP_STONE_RULES: tuple[tuple[str, str], ...] = (
    (r"(?:^|[-_/\s])GA(?:$|[-_/\s])|GREEN AGATE|GREEN AGAT", "Green Agate"),
    (r"(?:^|[-_/\s])CD(?:$|[-_/\s])|CHROME DIOPSIDE|DIOPOSIDE", "Chrome Diopside"),
    (r"(?:^|[-_/\s])CE(?:$|[-_/\s])|CREATED EMERALD", "Created Emerald"),
    (r"(?:^|[-_/\s])PERI(?:$|[-_/\s])|PERIDOT", "Peridot"),
    (r"(?:^|[-_/\s])EM(?:$|[-_/\s])|EMERALD", "Emerald"),
    (r"WHITE TOPAZ|WHIT TOPAZ|WHITETOPAZ|(?:^|[-_/\s])WBT(?:$|[-_/\s])", "White Topaz"),
    (r"SKY BLUE TOPAZ|SKY TOPAZ", "Sky Blue Topaz"),
    (r"MULTI BLUE TOPAZ|MULTI BT|(?:^|[-_/\s])MLBT(?:$|[-_/\s])", "Multi Blue Topaz"),
    (r"BLUE TOPAZ|(?:^|[-_/\s])BT(?:$|[-_/\s])", "Blue Topaz"),
    (r"ABALONE|HELIOTIS", "Abalone"),
    (r"MOTHER OF PEARL|(?:^|[-_/\s])MOP(?:$|[-_/\s])", "Mother Of Pearl"),
    (r"BLACK SPINEL|(?:^|[-_/\s])BSP(?:$|[-_/\s])", "Black Spinel"),
    (r"SPINEL", "Spinel"),
    (r"BLACK ONYX|MATT ONYX|MATTE ONYX|ONYX", "Onyx"),
    (r"OBSIDIAN|(?:^|[-_/\s])OB(?:$|[-_/\s])", "Obsidian"),
    (r"GREEN AMETHYST|(?:^|[-_/\s])GAM(?:$|[-_/\s])", "Green Amethyst"),
    (r"AMETHYST|(?:^|[-_/\s])AMST(?:$|[-_/\s])", "Amethyst"),
    (r"MYSTIC TOPAZ", "Mystic Topaz"),
    (r"MYSTIC QUARTZ|MYSTIC MB|MYST MB|MYSTIC", "Mystic Quartz"),
    (r"CITRINE|(?:^|[-_/\s])CIT(?:$|[-_/\s])", "Citrine"),
    (r"SMOKY|SMOKEY|RAUCH", "Smoky"),
    (r"HONEY", "Honey"),
    (r"ROSE QUARTZ", "Rose Quartz"),
    (r"WHITE QUARTZ", "White Quartz"),
    (r"GREEN QUARTZ", "Green Quartz"),
    (r"LEMON QUARTZ", "Lemon Quartz"),
    (r"RUTILE QUARTZ", "Rutile Quartz"),
    (r"QUARTZ", "Quartz"),
    (r"RHODOLITE|RODOLITE|(?:^|[-_/\s])RDT(?:$|[-_/\s])", "Rhodolite"),
    (r"GARNET|GRANADA|GRANATE|(?:^|[-_/\s])GARN(?:$|[-_/\s])", "Garnet"),
    (r"PICTURE JASPER|DALMATIAN JASPER|(?:^|[-_/\s])DJ(?:$|[-_/\s])", "Dalmatian Jasper"),
    (r"RED JASPER|(?:^|[-_/\s])RJ(?:$|[-_/\s])", "Red Jasper"),
    (r"JASPER", "Jasper"),
    (r"LAPIS LAZURITE|LAPIS LAZULI|LAPIZ|LAZURITE|(?:^|[-_/\s])LAP(?:$|[-_/\s])", "Lapis Lazurite"),
    (r"WHITE HOWLITE|HOWLITE|(?:^|[-_/\s])WHO(?:$|[-_/\s])", "White Howlite"),
    (r"APATITE|(?:^|[-_/\s])APAT(?:$|[-_/\s])", "Apatite"),
    (r"TANZANITE|TANZNITE|(?:^|[-_/\s])TANZ(?:$|[-_/\s])", "Tanzanite"),
    (r"IOLITE|IHOLIT|(?:^|[-_/\s])IOL(?:$|[-_/\s])", "Iolite"),
    (r"OPAL", "Opal"),
    (r"AMBER", "Amber"),
    (r"MALACHITE", "Malachite"),
    (r"TIGER EYE", "Tiger Eye"),
    (r"TOURMALINE", "Tourmaline"),
    (r"TURQUOISE", "Turquoise"),
    (r"AQUAMARINE", "Aquamarine"),
    (r"MOONSTONE", "Moonstone"),
    (r"LABRADORITE", "Labradorite"),
    (r"KYANITE|(?:^|[-_/\s])KYN(?:$|[-_/\s])", "Kyanite"),
    (r"AGATE", "Agate"),
)


def _concrete_stone_from_group(combined: str, *, allowed: frozenset[str] | None = None) -> str:
    for pattern, concrete_name in _GROUP_STONE_RULES:
        if re.search(pattern, combined):
            return concrete_name if allowed is None or concrete_name in allowed else UNRECOGNIZED_STONE
    return UNRECOGNIZED_STONE


def canonical_stone(value: object, sku: object = "") -> str:
    """Normalize supplier stone names with the same aliases used in reports.

    The order workspace keeps meaningful stone distinctions (for example
    Citrine versus Smoky) while removing spelling variants and abbreviations.
    """
    text = normalize_text(value)
    sku_text = normalize_text(sku)
    if not text:
        text = sku_text
    if text in STONE_EXACT_ALIASES:
        return _display_stone_name(STONE_EXACT_ALIASES[text])

    replacements = (
        ("LAPIS LAZULI", "LAPIS LAZURITE"),
        ("LAPIZ", "LAPIS LAZURITE"),
        ("BLACK AGATE", "AGATE"),
        ("GREEN AGAT", "GREEN AGATE"),
        ("CREATED EMEARLD", "CREATED EMERALD"),
        ("CREATED EMEALD", "CREATED EMERALD"),
        ("CREATE EMERALD", "CREATED EMERALD"),
        ("EMREAL", "EMERALD"),
        ("WHITETOPAZ", "WHITE TOPAZ"),
        ("WHIT TOPAZ", "WHITE TOPAZ"),
        ("IHOLIT", "IOLITE"),
        ("GERY PEARL", "GREY PEARL"),
        ("GREY PARL", "GREY PEARL"),
        ("WHITEE PEARL", "WHITE PEARL"),
        ("WHITE  PEARL", "WHITE PEARL"),
        ("FRESH WATER", "FRESHWATER"),
        ("MYSTMB", "MYSTIC"),
        ("MYST MB", "MYSTIC"),
        ("MOISANITE", "MOISSANITE"),
        ("MOSSANITE", "MOISSANITE"),
        ("MUSSONITE", "MOISSANITE"),
        ("SAPPHRIE", "SAPPHIRE"),
        ("SAPPHIRE", "SAPPHIRE"),
    )
    for old, new in replacements:
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text).strip()
    if text == "LAZURITE":
        text = "LAPIS LAZURITE"

    # Priority and abbreviation rules mirror the report logic for mixed names.
    combined = f"{text} {sku_text}".strip()

    # Analytical groups are navigation-only. Recover the concrete stone from
    # the source text or SKU; unresolved rows must be reviewed manually and
    # never leave the system under a generic supplier-facing group name.
    if text in {"GREEN STONE", "GREEN STONES"}:
        return _concrete_stone_from_group(combined, allowed=GREEN_STONE_NAMES)
    if text in {"OTHER TOPAZ", "TOPAZ OTHER"}:
        return _concrete_stone_from_group(combined, allowed=OTHER_TOPAZ_NAMES)
    if text in {"OTHER STONE", "OTHER STONES"}:
        return _concrete_stone_from_group(combined, allowed=OTHER_STONE_NAMES)

    if re.search(r"MO+I?S+A?N+I?T|MOSSANIT|MUSSONIT", combined):
        return "Moissanite"
    if "BSHQ" in combined or "BLUE SAPPHIRE HIGH QUALITY" in combined or "BLUE SAPPHIRE HQ" in combined:
        return "Blue Sapphire High Quality"
    if "BSMQ" in combined or "BLUE SAPPHIRE MEDIUM QUALITY" in combined or "BLUE SAPPHIRE MQ" in combined:
        return "Blue Sapphire Medium Quality"
    if re.search(r"SAPP+H?I?R?E|SAPPPHIRE", combined):
        return "Blue Sapphire"
    if "RUBY" in combined and "ZOISITE" not in combined and "CIOSITE" not in combined:
        return "Ruby"
    if ("LONDON" in combined or re.search(r"(?:^|[-_/\s])LBT(?:$|[-_/\s])", combined)) and ("TOPAZ" in combined or "BT" in combined):
        return "London Topaz"
    if ("SWISS" in combined or "SWIS" in combined or re.search(r"(?:^|[-_/\s])SWBT(?:$|[-_/\s])", combined)) and ("TOPAZ" in combined or "BT" in combined):
        return "Swiss Topaz"
    if "AZURE TOPAZ" in combined or "AZUR TOPAZ" in combined:
        return "Azure Topaz"
    if "WHITE TOPAZ" in combined or re.search(r"(?:^|[-_/\s])WBT(?:$|[-_/\s])", combined):
        return "White Topaz"
    if "SKY BLUE TOPAZ" in combined or "SKY TOPAZ" in combined:
        return "Sky Blue Topaz"
    if "MULTI BLUE TOPAZ" in combined or "MULTI BT" in combined or re.search(r"(?:^|[-_/\s])MLBT(?:$|[-_/\s])", combined):
        return "Multi Blue Topaz"
    if "BLUE TOPAZ" in combined or re.search(r"(?:^|[-_/\s])BT(?:$|[-_/\s])", combined):
        return "Blue Topaz"
    if "CREATED EMERALD" in combined:
        return "Created Emerald"
    if "RED EMERALD" in combined:
        return "Red Emerald"
    if "RHOMBIUM" in combined:
        return "Rhombium"
    if "CHROME DIOPSIDE" in combined or "DIOPOSIDE" in combined:
        return "Chrome Diopside"
    if "EMERALD" in combined:
        return "Emerald"
    if "GREEN AMETHYST" in combined:
        return "Green Amethyst"
    if "AMETHYST" in combined:
        return "Amethyst"
    if "RHODOLITE" in combined or "RODOLITE" in combined:
        return "Rhodolite"
    if "GARNET" in combined:
        return "Garnet"
    if "CITRINE" in combined:
        return "Citrine"
    if "ROSE QUARTZ" in combined:
        return "Rose Quartz"
    if "WHITE QUARTZ" in combined:
        return "White Quartz"
    if "SMOKY" in combined or "SMOKEY" in combined:
        return "Smoky"
    if "HONEY" in combined:
        return "Honey"
    if "MYSTIC TOPAZ" in combined:
        return "Mystic Topaz"
    if "MYSTIC" in combined:
        return "Mystic Quartz"
    if "GREEN AGATE" in combined:
        return "Green Agate"
    if "AGATE" in combined:
        return "Agate"
    if "BLACK SPINEL" in combined:
        return "Black Spinel"
    if "ONYX" in combined:
        return "Onyx"
    if "OBSIDIAN" in combined:
        return "Obsidian"
    if "IOLITE" in combined:
        return "Iolite"
    if "TANZANITE" in combined or "TANZNITE" in combined:
        return "Tanzanite"
    if "PERIDOT" in combined:
        return "Peridot"
    if "OPAL" in combined:
        return "Opal"
    if "TOURMALINE" in combined:
        return "Tourmaline"

    if text in STONE_EXACT_ALIASES:
        text = STONE_EXACT_ALIASES[text]
    if text in {"КАМЕНЬ НЕ РАСПОЗНАН", "НЕ РАСПОЗНАНО (AMA)"}:
        return UNRECOGNIZED_STONE
    return _display_stone_name(text) if text else "Не указан"


def order_stone_bucket(value: object, sku: object = "") -> str:
    """Map every non-pearl material to one of six order sections.

    Concrete stone names remain unchanged in cards and Excel. This function is
    navigation-only: Sapphire, Ruby, Moissanite, Topaz, Green Stones and the
    catch-all Other Stones pool.
    """
    stone = canonical_stone(value, sku)
    if stone in {"Blue Sapphire", "Blue Sapphire High Quality", "Blue Sapphire Medium Quality"}:
        return SAPPHIRE_ORDER_GROUP
    if stone == "Ruby":
        return RUBY_ORDER_GROUP
    if stone == "Moissanite":
        return MOISSANITE_ORDER_GROUP
    if stone in ORDER_TOPAZ_NAMES:
        return TOPAZ_ORDER_GROUP
    if stone in ORDER_GREEN_STONE_NAMES:
        return GREEN_STONES_GROUP
    return OTHER_STONES_GROUP


def canonical_group(value: object) -> str:
    text = normalize_text(value)
    aliases = {
        "EARRING": "Earrings",
        "EARRINGS": "Earrings",
        "СЕРЬГИ": "Earrings",
        "RING": "Ring",
        "RINGS": "Ring",
        "КОЛЬЦО": "Ring",
        "КОЛЬЦА": "Ring",
        "PENDANT": "Pendant",
        "PENDANTS": "Pendant",
        "ПОДВЕСКА": "Pendant",
        "ПОДВЕСКИ": "Pendant",
        "BRACELET": "Bracelet",
        "BRACELETS": "Bracelet",
        "БРАСЛЕТ": "Bracelet",
        "БРАСЛЕТЫ": "Bracelet",
        "NECKLACE": "Necklace",
        "NECKLACES": "Necklace",
        "ОЖЕРЕЛЬЕ": "Necklace",
    }
    return aliases.get(text, text.title() if text else "Не указана")


def _analytics_group_label(value: object) -> str:
    group = canonical_group(value)
    return ANALYTICS_GROUP_LABELS.get(group, group)


def _ordered_group_totals(rows: Iterable[tuple[OrderItem, int]]) -> dict[str, int]:
    totals: dict[str, int] = {}
    for item, quantity in rows:
        amount = max(0, safe_int(quantity))
        if amount <= 0:
            continue
        group = canonical_group(item.group)
        totals[group] = totals.get(group, 0) + amount
    ordered: dict[str, int] = {}
    for group in ANALYTICS_GROUP_ORDER:
        if totals.get(group, 0) > 0:
            ordered[group] = totals.pop(group)
    for group in sorted(totals, key=lambda name: _analytics_group_label(name).casefold()):
        if totals[group] > 0:
            ordered[group] = totals[group]
    return ordered


def _analytics_summary(rows: Iterable[tuple[OrderItem, int]]) -> dict[str, Any]:
    materialized = [(item, max(0, safe_int(quantity))) for item, quantity in rows if safe_int(quantity) > 0]
    return {
        "total_quantity": sum(quantity for _item, quantity in materialized),
        "sku_count": len(materialized),
        "group_totals": _ordered_group_totals(materialized),
    }


def _stone_top_family(stone: str) -> str | None:
    for family, names in TOP_STONE_ANALYTICS_FAMILIES:
        if stone in names:
            return family
    return None


def _colored_stone_family(stone: str) -> str:
    normalized = normalize_text(stone)
    if stone in QUARTZ_ANALYTICS_NAMES:
        return "Quartz Group"
    if stone in BLACK_STONE_ANALYTICS_NAMES:
        return "Black Stones"
    if stone in GARNET_ANALYTICS_NAMES:
        return "Garnet / Rhodolite"
    if stone in JASPER_ANALYTICS_NAMES:
        return "Jasper"
    if "CZ" in normalized or "CUBIC ZIRCON" in normalized:
        return "CZ"
    if normalized in {"КАМЕНЬ НЕ РАСПОЗНАН", "НЕ РАСПОЗНАНО", "НЕ РАСПОЗНАНО (AMA)", "НЕ УКАЗАН", ""}:
        return "Unrecognized"
    return "Other Colored Stones"


def _pearl_analytics_family(stone: str) -> str:
    """Group completed pearl orders by business-facing colour families.

    White freshwater pearls stay together regardless of whether the source
    name contains ``ROUND``. Any explicit coloured freshwater variant wins
    over shape markers, so round pink/rose, grey and black pearls cannot leak
    into the white family.
    """
    normalized = normalize_text(stone)
    if any(token in normalized for token in ("SEA PEARL", "SOUTH SEA", "AKOYA", "TAHITI", "TAHITIAN", "GALATEA", "FACETED SEA")):
        return "Sea Pearls"
    if "BAROQUE" in normalized:
        return "Baroque Pearls"

    colored_tokens = (
        "COLORED",
        "COLOUR",
        "COLOR ",
        "PINK",
        "ROSE",
        "GREY",
        "GRAY",
        "DARK",
        "BLACK",
        "PURPLE",
        "PEACH",
    )
    is_pearl = "PEARL" in normalized
    is_freshwater = "FRESHWATER" in normalized or "FRESH WATER" in normalized

    if (is_freshwater or is_pearl) and any(token in normalized for token in colored_tokens):
        return "Colored Freshwater"
    if is_freshwater or (is_pearl and ("WHITE" in normalized or "ROUND" in normalized)):
        return "White Freshwater"
    return "Other Pearls"


def _analytics_family_payload(
    family: str,
    rows: Iterable[tuple[OrderItem, int]],
    *,
    include_sku_in_material_detection: bool = True,
) -> dict[str, Any]:
    materialized = tuple(rows)
    stone_rows: dict[str, list[tuple[OrderItem, int]]] = {}
    for item, quantity in materialized:
        stone = canonical_stone(item.stone, item.sku if include_sku_in_material_detection else "")
        stone_rows.setdefault(stone, []).append((item, quantity))
    stones = []
    for stone, values in sorted(stone_rows.items(), key=lambda pair: pair[0].casefold()):
        stones.append({"name": stone, **_analytics_summary(values)})
    return {"name": family, **_analytics_summary(materialized), "stones": stones}


def build_order_analytics(
    parsed: ParsedOrderWorkbook,
    draft: OrderDraft,
    mode: str,
) -> dict[str, Any]:
    """Build a persisted-order quantity breakdown without using recommendations.

    Quantities are summed in pieces, not SKU counts. Limited Order positions do
    not have an ordered quantity and are therefore kept outside the analytics.
    The saved draft is the only source of quantities, so later recommendation
    changes cannot rewrite completed-order history.
    """
    if mode not in ORDER_MODES:
        raise ValueError("Неизвестный тип заказа.")
    rows = [
        (item, max(0, safe_int(draft.orders.get(item.key, 0))))
        for item in parsed.items
        if item_in_analytics_mode(item, mode)
        and safe_int(draft.orders.get(item.key, 0)) > 0
        and not draft.limited_orders.get(item.key, False)
    ]
    result: dict[str, Any] = {
        "mode": mode,
        **_analytics_summary(rows),
        "limited_positions": sum(
            1
            for item in parsed.items
            if item_in_mode(item, mode) and bool(draft.limited_orders.get(item.key, False))
        ),
        "sections": [],
    }
    if mode == ORDER_MODE_STONES:
        section_rows: dict[str, dict[str, list[tuple[OrderItem, int]]]] = {
            "Топовые камни": {},
            "Цветные камни": {},
        }
        for item, quantity in rows:
            stone = canonical_stone(item.stone, item.sku)
            top_family = _stone_top_family(stone)
            section = "Топовые камни" if top_family else "Цветные камни"
            family = top_family or _colored_stone_family(stone)
            section_rows[section].setdefault(family, []).append((item, quantity))

        for section_name in ("Топовые камни", "Цветные камни"):
            families_raw = section_rows[section_name]
            if section_name == "Топовые камни":
                family_order = [name for name, _values in TOP_STONE_ANALYTICS_FAMILIES]
            else:
                family_order = list(COLORED_STONE_ANALYTICS_FAMILY_ORDER)
            families = [
                _analytics_family_payload(family, families_raw[family])
                for family in family_order
                if family in families_raw
            ]
            materialized = [row for family_rows in families_raw.values() for row in family_rows]
            result["sections"].append({
                "name": section_name,
                **_analytics_summary(materialized),
                "families": families,
            })
    else:
        family_rows: dict[str, list[tuple[OrderItem, int]]] = {}
        for item, quantity in rows:
            stone = canonical_stone(item.stone, "")
            family_rows.setdefault(_pearl_analytics_family(stone), []).append((item, quantity))
        families = [
            _analytics_family_payload(
                family,
                family_rows[family],
                include_sku_in_material_detection=False,
            )
            for family in PEARL_ANALYTICS_FAMILY_ORDER
            if family in family_rows
        ]
        result["sections"].append({
            "name": "Жемчуг",
            **_analytics_summary(rows),
            "families": families,
        })
    return result


def _compact_store_name(value: object) -> str:
    return re.sub(r"[^A-ZА-Я0-9]+", "", normalize_text(value))


def is_store_63(value: object) -> bool:
    """Match store 63 across labels used by supplier reports."""
    compact = _compact_store_name(value)
    return compact == "63" or compact.startswith("63NDC") or compact in {"631", "632"}


def is_store_20(value: object) -> bool:
    """Match the single excluded store across all report aliases.

    Business-wise ``20``, ``20NDC`` and ``Princess Hang`` are one and the
    same point.  Keeping the aliases in one predicate prevents the parser
    from subtracting the same store twice.
    """
    compact = _compact_store_name(value)
    return (
        compact == "20"
        or compact.startswith("20NDC")
        or compact in {"PRINCESSHANG", "PRINCESSHANGSTORE"}
    )


def is_tt_outlet_store(value: object) -> bool:
    text = normalize_text(value)
    if "STOCK" in text or "СКЛАД" in text:
        return False
    return text in {"OUTLET", "TT", "TT OUTLET", "OUTLET TT", "ТТ"}


def is_stock_tt_store(value: object) -> bool:
    """Return whether a store column is the transferable TT warehouse."""
    text = normalize_text(value)
    compact = _compact_store_name(value)
    return (
        ("STOCK" in text or "СКЛАД" in text)
        and ("TT" in text or "ТТ" in text)
    ) or compact in {"STOCKTT", "TTSTOCK", "СКЛАДТТ", "ТТСКЛАД"}


def is_ordinary_working_store(value: object) -> bool:
    """Stores whose last unit must remain on display and cannot be moved."""
    return not (
        is_store_20(value)
        or is_store_63(value)
        or is_tt_outlet_store(value)
        or is_stock_tt_store(value)
    )


def is_princess_hang_store(value: object) -> bool:
    """Compatibility alias for the unified store-20 predicate."""
    text = normalize_text(value)
    compact = re.sub(r"[^A-ZА-Я0-9]+", "", text)
    return compact in {"PRINCESSHANG", "PRINCESSHANGSTORE"} or ("PRINCESS" in text and "HANG" in text)


def resolve_store_20_stock(stores: dict[str, int]) -> tuple[int, str | None]:
    """Return one quantity for the 20/20NDC/Princess Hang point.

    Reports normally contain one of the aliases.  If a workbook contains
    more than one alias column, they are alternative representations of the
    same store and therefore must not be summed.  The largest non-negative
    value is used and a warning is returned when non-zero aliases disagree.
    """
    aliases = [(name, max(0, safe_int(qty))) for name, qty in stores.items() if is_store_20(name)]
    if not aliases:
        return 0, None
    quantity = max(qty for _, qty in aliases)
    non_zero = {qty for _, qty in aliases if qty > 0}
    warning = None
    if len(non_zero) > 1:
        details = ", ".join(f"{name}: {qty}" for name, qty in aliases)
        warning = (
            "Колонки 20/20NDC/Princess Hang обозначают один магазин, "
            f"но содержат разные значения ({details}). Использовано максимальное: {quantity}."
        )
    return quantity, warning


def canonical_store_values(store_values: dict[str, int]) -> dict[str, int]:
    """Collapse only the store-20 aliases for balance checks."""
    normalized: dict[str, int] = {}
    store_20, _ = resolve_store_20_stock(store_values)
    for name, value in store_values.items():
        if is_store_20(name):
            continue
        normalized[normalize_text(name)] = safe_int(value)
    if any(is_store_20(name) for name in store_values):
        normalized["20 / PRINCESS HANG"] = store_20
    return normalized


def is_pearl_name(value: object) -> bool:
    text = normalize_text(value)
    if "PEARL" in text or "PARL" in text:
        return True
    return text in {"FPW", "FPC", "TAH", "SSP"}


def is_excluded_pearl(value: object, sku: object = "") -> bool:
    text = normalize_text(value)
    canonical = normalize_text(canonical_stone(value, sku))
    combined = f"{text} {canonical} {normalize_text(sku)}".strip()
    settings = load_order_exclusions()
    if any(pattern in combined for pattern in settings["pearl_patterns"]):
        return True
    # Any round pearl is purchased outside this supplier order. Check both the
    # material description and SKU text so a visible ROUND marker cannot leak
    # into cards, recommendations, summaries, analytics or supplier Excel.
    if settings["exclude_round_pearl"] and "PEARL" in combined and "ROUND" in combined:
        return True
    return False


def pearl_order_bucket(value: object, sku: object = "") -> str | None:
    """Return one of the five purchasable pearl sections.

    Sea and Round pearls are excluded before classification. Generic or
    unresolved pearl labels are deliberately not invented: only White, Grey,
    Pink, Black and Baroque enter this supplier order.
    """
    if is_excluded_pearl(value, sku):
        return None
    canonical = normalize_text(canonical_stone(value, sku))
    combined = f"{normalize_text(value)} {canonical} {normalize_text(sku)}".strip()
    if "BAROQUE" in combined:
        return "Baroque"
    if "PINK" in combined or "ROSE" in combined:
        return "Pink"
    if "GREY" in combined or "GRAY" in combined:
        return "Grey"
    if "BLACK" in combined or "DARK" in combined:
        return "Black"
    if "WHITE" in combined or re.search(r"(?:^|[-_/\s])FPW(?:$|[-_/\s])", combined):
        return "White"
    return None


def order_set_navigation_bucket(order_set: OrderSet, mode: str) -> str | None:
    """Resolve the visible top-level section for an already-built set."""
    if mode == ORDER_MODE_PEARLS:
        for item in order_set.items:
            bucket = pearl_order_bucket(item.stone, item.sku)
            if bucket:
                return bucket
        return None
    return order_stone_bucket(order_set.stone)


def order_navigation_options(mode: str) -> tuple[str, ...]:
    """Return the exact business-approved section list for the order mode."""
    return PEARL_ORDER_BUCKET_ORDER if mode == ORDER_MODE_PEARLS else STONE_ORDER_BUCKET_ORDER


def is_excluded_stone(value: object) -> bool:
    text = normalize_text(value)
    return any(pattern in text for pattern in load_order_exclusions()["stone_patterns"])


def item_in_mode(item: OrderItem, mode: str) -> bool:
    # Bracelets are purchased from another supplier and must not enter this
    # order workspace or its Excel.
    if canonical_group(item.group) == "Bracelet":
        return False
    pearl = is_pearl_name(item.stone)
    if mode == ORDER_MODE_PEARLS:
        return pearl and pearl_order_bucket(item.stone, item.sku) is not None
    return (not pearl) and not is_excluded_stone(item.stone)


def item_in_analytics_mode(item: OrderItem, mode: str) -> bool:
    """Keep historical analytics compatible with older completed orders.

    Current pearl-order UI accepts only five explicit groups, while historical
    drafts may contain a generic Colored Freshwater label. Such rows remain in
    completed-order analytics, provided they are not Sea/Round exclusions.
    """
    if canonical_group(item.group) == "Bracelet":
        return False
    pearl = is_pearl_name(item.stone)
    if mode == ORDER_MODE_PEARLS:
        return pearl and not is_excluded_pearl(item.stone, item.sku)
    return (not pearl) and not is_excluded_stone(item.stone)


def classify_set(items: Iterable[OrderItem]) -> tuple[str, str, int, str | None]:
    materialized = tuple(items)
    if not materialized:
        return CATEGORY_ZERO, "", 0, "0/0"
    driver = max(materialized, key=lambda item: (item.sales, -item.row))
    maximum = max(0, int(driver.sales))
    if maximum >= 5:
        category = CATEGORY_TOP
    elif maximum >= 3:
        category = CATEGORY_MEDIUM
    elif maximum >= 1:
        category = CATEGORY_WEAK
    else:
        category = CATEGORY_ZERO
    zero_segment = None
    if category == CATEGORY_ZERO:
        zero_segment = "Нулевые с остатком" if any(item.working_stock > 0 for item in materialized) else "0/0 — не было остатка"
    return category, driver.sku, maximum, zero_segment


def order_set_material(item: OrderItem, mode: str) -> str:
    """Use the primary order material, never a secondary SKU stone marker."""
    if mode == ORDER_MODE_PEARLS:
        return pearl_order_bucket(item.stone, item.sku) or "Не указан"
    return canonical_stone(item.stone, item.sku)


def build_order_sets(items: Iterable[OrderItem], mode: str) -> tuple[OrderSet, ...]:
    # A supplier Set# may contain the same visual family in several stones.
    # The order workspace is stone-first, therefore a set must be split by the
    # normalized stone before category, TVP and error flags are calculated.
    # This prevents Ruby/London Topaz rows from leaking into Blue Sapphire and
    # from promoting or flagging the visible sapphire part of the set.
    normal_groups: dict[tuple[str, str], list[OrderItem]] = {}
    normal_order: list[tuple[str, str]] = []
    ungrouped_items: list[OrderItem] = []

    for item in items:
        if not item_in_mode(item, mode):
            continue
        if item.ungrouped or normalize_text(item.set_id) == "БЕЗ КОМПЛЕКТА":
            ungrouped_items.append(item)
            continue
        stone = order_set_material(item, mode)
        group_key = (stone, item.set_id)
        if group_key not in normal_groups:
            normal_groups[group_key] = []
            normal_order.append(group_key)
        normal_groups[group_key].append(item)

    result: list[OrderSet] = []
    for stone, set_id in normal_order:
        group_items = tuple(normal_groups[(stone, set_id)])
        category, driver_sku, max_sales, zero_segment = classify_set(group_items)
        has_positive = any(item.tvp_raw > 0 for item in group_items)
        has_negative = any(item.tvp_raw < 0 for item in group_items)
        result.append(OrderSet(
            key=f"{mode}|{stone}|{set_id}",
            set_id=set_id,
            stone=stone,
            items=group_items,
            category=category,
            driver_sku=driver_sku,
            max_sales=max_sales,
            has_positive_tvp=has_positive,
            has_negative_tvp=has_negative,
            zero_segment=zero_segment,
            is_ungrouped=False,
        ))

    # Standalone rows are not allowed to promote each other. They are first
    # classified one by one, then collected into one virtual "Без комплекта"
    # block per stone/category. TVP rows are kept in a separate virtual block
    # so one item in transit never hides all other standalone models.
    virtual_groups: dict[tuple[str, str, str | None, str], list[OrderItem]] = {}
    for item in ungrouped_items:
        category, _driver, _maximum, zero_segment = classify_set((item,))
        stone = order_set_material(item, mode)
        transit_bucket = "tvp" if item.tvp_raw > 0 else "regular"
        key = (stone, category, zero_segment, transit_bucket)
        virtual_groups.setdefault(key, []).append(item)

    category_rank = {category: index for index, category in enumerate(CATEGORY_ORDER)}
    for (stone, category, zero_segment, transit_bucket), grouped_items in sorted(
        virtual_groups.items(),
        key=lambda pair: (pair[0][0], category_rank[pair[0][1]], pair[0][2] or "", pair[0][3]),
    ):
        materialized = tuple(sorted(grouped_items, key=lambda item: item.row))
        driver = max(materialized, key=lambda item: (item.sales, -item.row))
        result.append(OrderSet(
            key=f"{mode}|ungrouped|{stone}|{category}|{zero_segment}|{transit_bucket}",
            set_id="Без комплекта",
            stone=stone,
            items=materialized,
            category=category,
            driver_sku=driver.sku,
            max_sales=driver.sales,
            has_positive_tvp=(transit_bucket == "tvp"),
            has_negative_tvp=any(item.tvp_raw < 0 for item in materialized),
            zero_segment=zero_segment,
            is_ungrouped=True,
        ))
    return tuple(result)



def monthly_sales_rate(item: OrderItem) -> float:
    """Convert actual report sales to a monthly rate.

    Supplier-order calculations use the approved fixed four-month sales
    window. The period caption from the workbook is informational only.
    """
    months = max(1, safe_int(getattr(item, "report_months", DEFAULT_REPORT_MONTHS)))
    return max(0, item.sales) / float(months)


def demand_until_arrival(item: OrderItem) -> int:
    """Expected sales during the two-month supplier lead time."""
    return int(math.ceil(monthly_sales_rate(item) * 2.0))


def ordinary_transferable_stock(item: OrderItem) -> int:
    """Units that may be moved from ordinary shops without emptying them."""
    total = 0
    for name, quantity in item.stores.items():
        if not is_ordinary_working_store(name):
            continue
        total += max(0, safe_int(quantity) - 1)

    # Old reports without an explicit NTR2 column reconstruct that store from
    # the total. Its last unit is protected exactly like every other shop.
    if item.ntr2_calculated and not any(normalize_text(name) == "NTR2" for name in item.stores):
        total += max(0, item.ntr2_stock - 1)
    return total


def transferable_stock(item: OrderItem) -> int:
    """All stock that can be sent to TT now.

    Every unit on Stock TT is transferable; ordinary stores keep their final
    display unit.
    """
    return ordinary_transferable_stock(item) + max(0, item.stock_tt_warehouse)


def _transfer_advice(item: OrderItem, target_tt: int = 3) -> tuple[str, ...]:
    shortage = max(0, target_tt - item.stock_tt)
    if shortage <= 0:
        return ()

    advice: list[str] = []
    from_warehouse = min(shortage, max(0, item.stock_tt_warehouse))
    if from_warehouse > 0:
        advice.append(f"Переместить {from_warehouse} шт. Stock TT → TT.")
        shortage -= from_warehouse

    if shortage > 0:
        ordinary = ordinary_transferable_stock(item)
        if ordinary > 0:
            advice.append(f"Переместить до {min(shortage, ordinary)} шт. из обычных магазинов → TT.")
    return tuple(advice)


def is_stud_item(item: OrderItem, order_set: OrderSet | None = None) -> bool:
    set_name = order_set.set_id if order_set is not None else item.set_id
    text = normalize_text(f"{item.group} {item.sku} {set_name}")
    return bool(re.search(r"PUS+ET|PUSET|STUD|ПУС+ЕТ", text))


def is_cross_item(item: OrderItem, order_set: OrderSet | None = None) -> bool:
    """Crosses are a special pendant assortment, identified by section text."""
    set_name = order_set.set_id if order_set is not None else item.set_id
    text = normalize_text(f"{set_name} {item.group}")
    return canonical_group(item.group) == "Pendant" and bool(
        re.search(r"КРЕСТ|CROSS|CRUCIFIX", text)
    )


def earring_lock_code(value: object) -> str | None:
    """Read the lock code immediately after the two-digit model year."""
    sku = re.sub(r"[^A-Z0-9]+", "", normalize_text(value))
    match = re.match(r"^S?KE\d{2}([A-Z])", sku)
    if not match:
        return None
    code = match.group(1)
    return code if code in EARRING_LOCKS else None


def earring_lock_export_label(code: object) -> str:
    normalized = normalize_text(code)
    if normalized not in EARRING_LOCKS:
        return ""
    english, _russian = EARRING_LOCKS[normalized]
    return f"{normalized} — {english}"


def _minimum_supplier_batch(quantity: int) -> tuple[int, tuple[str, ...]]:
    amount = max(0, safe_int(quantity))
    if amount <= 0:
        return 0, ()
    if amount == 1:
        return 0, ("Расчёт дал 1 шт.; такую автоматическую партию не формируем.",)
    if amount == 2:
        return 3, ("Расчёт дал 2 шт.; минимальная автоматическая партия — 3 шт.",)
    return amount, ()


def _finalize_recommendation_quantity(
    quantity: int,
    recommendation_profile: str = RECOMMENDATION_BASE,
) -> tuple[int, tuple[str, ...]]:
    """Apply seasonality once, then apply the global supplier minimum."""
    base = max(0, safe_int(quantity))
    notes: list[str] = []
    adjusted = base
    if recommendation_profile == RECOMMENDATION_SEASONAL and base > 0:
        adjusted = max(0, base - 1)
        notes.append(f"Сезонный режим уменьшил расчёт на 1: {base} → {adjusted}.")
    final, batch_notes = _minimum_supplier_batch(adjusted)
    notes.extend(batch_notes)
    return final, tuple(notes)


def _projected_total_stock(item: OrderItem) -> int:
    """Working stock expected at arrival, including TVP and two-month demand."""
    return max(0, item.working_stock + item.positive_tvp - demand_until_arrival(item))


def _group_items(order_set: OrderSet, group: str) -> tuple[OrderItem, ...]:
    return tuple(item for item in order_set.items if canonical_group(item.group) == group)


def _group_projected_stock(order_set: OrderSet, group: str) -> int:
    return sum(_projected_total_stock(item) for item in _group_items(order_set, group))


def _group_has_positive_tvp(order_set: OrderSet, group: str) -> bool:
    return any(item.positive_tvp > 0 for item in _group_items(order_set, group))


def _balance_target_item(item: OrderItem, order_set: OrderSet, group: str) -> bool:
    """Prevent one set-balance shortage from being recommended on every SKU."""
    candidates = [candidate for candidate in _group_items(order_set, group) if candidate.positive_tvp <= 0]
    if not candidates:
        return False
    selected = max(candidates, key=lambda candidate: (candidate.sales, -candidate.working_stock, -candidate.row))
    return selected.key == item.key


def _incoming_set_balance(item: OrderItem, order_set: OrderSet) -> tuple[int, str] | None:
    """Keep the existing TVP companion-balance safety net for real sets."""
    if order_set.is_ungrouped:
        return None

    group = canonical_group(item.group)
    if group == "Ring" and _group_has_positive_tvp(order_set, "Earrings"):
        if not _balance_target_item(item, order_set, "Ring"):
            return None
        projected_earrings = _group_projected_stock(order_set, "Earrings")
        ratio = 0.8 if order_set.category == CATEGORY_TOP else 0.6
        desired_rings = int(math.ceil(projected_earrings * ratio))
        projected_rings = _group_projected_stock(order_set, "Ring")
        if desired_rings > projected_rings:
            return desired_rings - projected_rings, "Выравнивание комплекта: серьги находятся в пути."

    if group == "Earrings" and _group_has_positive_tvp(order_set, "Ring"):
        if not _balance_target_item(item, order_set, "Earrings"):
            return None
        projected_rings = _group_projected_stock(order_set, "Ring")
        desired_earrings = 5 * int(math.ceil(projected_rings / 4.0)) if projected_rings > 0 else 0
        projected_earrings = _group_projected_stock(order_set, "Earrings")
        if desired_earrings > projected_earrings:
            return desired_earrings - projected_earrings, "Выравнивание комплекта: кольца находятся в пути."
    return None


def _earrings_tt_candidate(item: OrderItem, order_set: OrderSet) -> tuple[int, str, str] | None:
    sales = max(0, item.sales)
    tt = max(0, item.stock_tt)
    ordinary_move = ordinary_transferable_stock(item)

    if tt >= 3:
        return None

    if tt == 2:
        if sales >= 8:
            return 3, "TT = 2, но модель активно продаётся; сохраняем минимальный заказ 3.", "earrings_tt2_active"
        if 4 <= sales <= 7 and transferable_stock(item) <= 0:
            return 3, "TT = 2, продажи заметные, а перемещаемого запаса нет.", "earrings_tt2"
        return None

    if tt == 1:
        if sales >= 8:
            if ordinary_move > 0:
                return 3, "TT можно временно пополнить обычным перемещением; поставщику оставляем минимальный заказ.", "earrings_tt1_move"
            return 5, "TT = 1 и модель активно продаётся; нужен запас 5 на период ожидания.", "earrings_tt1_active"
        if sales >= 2:
            return 3, "TT = 1 и было не менее двух продаж; минимально заказываем 3.", "earrings_tt1"
        return None

    # TT == 0
    if sales >= 8:
        if ordinary_move > 0:
            return 3, "TT пустой, но есть переносимый запас из обычного магазина; заказываем минимум 3.", "earrings_tt0_move"
        return 5, "TT пустой и модель активно продаётся; формируем усиленный заказ 5.", "earrings_tt0_active"
    if sales >= 1:
        return 3, "TT пустой и у модели были продажи; минимально заказываем 3.", "earrings_tt0"
    if item.working_stock > 0:
        return 3, "TT пустой; модель представлена в сети, поэтому создаём минимальный запас 3.", "earrings_tt0_zero_sales_stock"
    return 5, "TT пустой и изделия нет в рабочей сети; создаём базовое наличие 5.", "earrings_tt0_zero"


def _ring_completeness_candidate(item: OrderItem, order_set: OrderSet) -> tuple[int, str, str] | None:
    tt = max(0, item.stock_tt)
    if tt >= 3:
        return None

    strong = order_set.category == CATEGORY_TOP or item.sales >= 5
    target = 4 if strong else 3

    if item.working_stock <= 0:
        if strong:
            network_target = max(target, max(0, item.eligible_store_count) + 2)
            return (
                network_target,
                "Колец нет в рабочей сети: для сильной модели заполняем рабочие магазины и 2 шт. в TT.",
                "ring_network_fill",
            )
        return 3, "Колец нет в рабочей сети: для слабой/средней модели достаточно минимального заказа 3.", "ring_tt_min"

    # Existing transferable shop stock may close the immediate TT gap. The
    # remaining shortage follows the desired 5:4 or 5:3 set ratio.
    effective_tt = tt + min(ordinary_transferable_stock(item), max(0, target - tt))
    shortage = max(0, target - effective_tt)
    if shortage <= 0:
        return None
    return shortage, (
        f"Комплектность колец: целевой уровень {target}, доступно для TT {effective_tt}."
    ), "ring_completeness"


def _pendant_tt_candidate(item: OrderItem, order_set: OrderSet) -> tuple[int, str, str] | None:
    sales = max(0, item.sales)
    tt = max(0, item.stock_tt)

    if tt >= 3:
        return None
    if tt == 0:
        if sales >= 1:
            return 3, "TT пустой и подвеска хотя бы раз продавалась; минимально заказываем 3.", "pendant_tt0"
        if order_set.category == CATEGORY_TOP:
            return 3, "Подвеска не продавалась, но комплект топовый; поддерживаем комплектность партией 3.", "pendant_top_set"
        return None
    if tt == 1:
        if sales >= 2:
            return 3, "TT = 1 и было не менее двух продаж; минимально заказываем 3.", "pendant_tt1"
        return None
    # TT == 2
    if sales >= 8:
        return 3, "TT = 2 и подвеска активно продаётся; сохраняем минимальный заказ 3.", "pendant_tt2_active"
    if 5 <= sales <= 7 and transferable_stock(item) <= 0:
        return 3, "TT = 2, продажи заметные, а перемещаемого запаса нет.", "pendant_tt2"
    return None


def _cross_target_candidate(item: OrderItem, order_set: OrderSet) -> tuple[int, str, str] | None:
    if not is_cross_item(item, order_set):
        return None
    if item.sales >= 5:
        shortage = max(0, 5 - item.working_stock)
        if shortage > 0:
            return shortage, "Крестик хорошо продаётся: добиваем рабочий запас до 5.", "cross_target_5"
    if item.sales >= 1 and item.working_stock > 0:
        shortage = max(0, 3 - item.working_stock)
        if shortage > 0:
            return shortage, "Крестик продавался: добиваем рабочий запас до 3.", "cross_target_3"
    return None


def _normal_demand_candidate(item: OrderItem) -> tuple[int, str, str] | None:
    demand = demand_until_arrival(item)
    shortage = max(0, demand - item.working_stock)
    if shortage <= 0:
        return None
    return (
        shortage,
        f"Спрос на 2 месяца: {demand}; рабочий остаток: {item.working_stock}; чистая потребность: {shortage}.",
        "demand",
    )


def build_order_recommendation(
    item: OrderItem,
    order_set: OrderSet,
    mode: str,
    recommendation_profile: str = RECOMMENDATION_BASE,
) -> OrderRecommendation:
    """Recommendation engine approved for release 1.10.1."""
    if item.duplicate_status == "suppress":
        return OrderRecommendation(
            0,
            (
                item.duplicate_reason
                or f"Есть очень похожая модель {item.duplicate_sku or ''}; позиция исключена как вероятный дубль.",
            ),
            False,
            "duplicate",
        )

    if item.positive_tvp > 0:
        return OrderRecommendation(
            quantity=0,
            reasons=(f"В пути уже {item.positive_tvp} шт.; этот SKU можно дозаказать только вручную.",),
            blocked_by_tvp=True,
            rule="tvp",
        )

    group = canonical_group(item.group)
    if group == "Bracelet":
        return OrderRecommendation(0, ("Браслеты закупаются у другого поставщика.",), False, "other_supplier")

    candidates: list[tuple[int, str, str]] = []

    demand_candidate = _normal_demand_candidate(item)
    if demand_candidate is not None:
        candidates.append(demand_candidate)

    incoming_balance = _incoming_set_balance(item, order_set)
    if incoming_balance is not None:
        quantity, reason = incoming_balance
        candidates.append((quantity, reason, "tvp_set_balance"))

    if group == "Earrings":
        special = _earrings_tt_candidate(item, order_set)
        if special is not None:
            candidates.append(special)
    elif group == "Ring":
        special = _ring_completeness_candidate(item, order_set)
        if special is not None:
            candidates.append(special)
    elif group == "Pendant":
        special = _pendant_tt_candidate(item, order_set)
        if special is not None:
            candidates.append(special)
        cross = _cross_target_candidate(item, order_set)
        if cross is not None:
            candidates.append(cross)

    if not candidates:
        return OrderRecommendation(0, ("Автоматическое пополнение сейчас не требуется.",), False, "none")

    base_quantity = max(quantity for quantity, _reason, _rule in candidates)
    reasons = tuple(dict.fromkeys(reason for quantity, reason, _rule in candidates if quantity > 0))
    priority = {
        "duplicate": 100,
        "tvp_set_balance": 90,
        "ring_network_fill": 80,
        "earrings_tt0_active": 75,
        "earrings_tt1_active": 74,
        "cross_target_5": 73,
        "cross_target_3": 72,
        "pendant_top_set": 70,
        "earrings_tt0_zero": 69,
        "earrings_tt0": 68,
        "ring_completeness": 67,
        "ring_tt_min": 66,
        "pendant_tt0": 65,
        "pendant_tt1": 64,
        "pendant_tt2_active": 63,
        "demand": 10,
    }
    rule = max(candidates, key=lambda row: (row[0], priority.get(row[2], 0)))[2]
    final_quantity, adjustment_reasons = _finalize_recommendation_quantity(
        base_quantity,
        recommendation_profile,
    )
    final_reasons = tuple(dict.fromkeys((*reasons, *adjustment_reasons)))

    transfer_target = 3 if group in {"Earrings", "Ring", "Pendant"} else 0
    transfers = _transfer_advice(item, transfer_target) if transfer_target else ()
    return OrderRecommendation(final_quantity, final_reasons, False, rule, transfers)


def suggested_order_quantity(
    item: OrderItem,
    order_set: OrderSet | None = None,
    mode: str = ORDER_MODE_STONES,
    recommendation_profile: str = RECOMMENDATION_BASE,
) -> int:
    """Compatibility wrapper used by tests and integrations."""
    if order_set is None:
        category, driver_sku, max_sales, zero_segment = classify_set((item,))
        order_set = OrderSet(
            key=f"compat|{item.key}",
            set_id=item.set_id,
            stone=canonical_stone(item.stone, item.sku),
            items=(item,),
            category=category,
            driver_sku=driver_sku,
            max_sales=max_sales,
            has_positive_tvp=item.tvp_raw > 0,
            has_negative_tvp=item.tvp_raw < 0,
            zero_segment=zero_segment,
            is_ungrouped=item.ungrouped,
        )
    return build_order_recommendation(
        item,
        order_set,
        mode,
        recommendation_profile,
    ).quantity


def infer_ntr2(total: int, store_values: dict[str, int], has_actual_ntr2: bool) -> tuple[int, bool, str | None]:
    normalized = canonical_store_values(store_values)
    if has_actual_ntr2:
        actual = normalized.get("NTR2", 0)
        delta = total - sum(normalized.values())
        warning = None if delta == 0 else f"Сумма магазинов отличается от «Всего» на {delta} шт."
        return max(0, actual), False, warning
    inferred = total - sum(normalized.values())
    if inferred < 0:
        return 0, True, f"Сумма магазинов превышает «Всего» на {abs(inferred)} шт."
    return inferred, True, None


@dataclass(frozen=True)
class _ImageSignature:
    digest: str
    dhash: int
    histogram: tuple[int, ...]
    aspect: float


def _normalized_image(payload: bytes) -> Image.Image | None:
    try:
        image = Image.open(io.BytesIO(payload)).convert("RGB")
    except (UnidentifiedImageError, OSError, ValueError):
        return None
    if image.width <= 1 or image.height <= 1:
        return None

    # Remove the mostly white supplier-photo margins before comparison.
    corners = [
        image.getpixel((0, 0)),
        image.getpixel((image.width - 1, 0)),
        image.getpixel((0, image.height - 1)),
        image.getpixel((image.width - 1, image.height - 1)),
    ]
    background = tuple(sum(pixel[channel] for pixel in corners) // len(corners) for channel in range(3))
    background_image = Image.new("RGB", image.size, background)
    difference = ImageChops.difference(image, background_image).convert("L")
    mask = difference.point(lambda value: 255 if value > 18 else 0)
    bbox = mask.getbbox()
    if bbox:
        left, top, right, bottom = bbox
        pad_x = max(2, int((right - left) * 0.05))
        pad_y = max(2, int((bottom - top) * 0.05))
        bbox = (
            max(0, left - pad_x),
            max(0, top - pad_y),
            min(image.width, right + pad_x),
            min(image.height, bottom + pad_y),
        )
        image = image.crop(bbox)

    image = ImageOps.contain(image, (128, 128), method=Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", (128, 128), (255, 255, 255))
    offset = ((128 - image.width) // 2, (128 - image.height) // 2)
    canvas.paste(image, offset)
    return canvas


def _flat_image_data(image: Image.Image):
    getter = getattr(image, "get_flattened_data", None)
    return getter() if callable(getter) else image.getdata()


def _image_signature(payload: bytes) -> _ImageSignature | None:
    image = _normalized_image(payload)
    if image is None:
        return None
    digest = hashlib.sha1(image.tobytes()).hexdigest()
    gray = image.convert("L").resize((17, 16), Image.Resampling.LANCZOS)
    pixels = list(_flat_image_data(gray))
    dhash = 0
    for row in range(16):
        base = row * 17
        for column in range(16):
            dhash = (dhash << 1) | int(pixels[base + column] > pixels[base + column + 1])

    small = image.resize((32, 32), Image.Resampling.BILINEAR)
    histogram = [0] * 64
    for red, green, blue in _flat_image_data(small):
        bucket = (red // 64) * 16 + (green // 64) * 4 + (blue // 64)
        histogram[min(63, bucket)] += 1
    return _ImageSignature(
        digest=digest,
        dhash=dhash,
        histogram=tuple(histogram),
        aspect=image.width / max(1, image.height),
    )


def _signature_similarity(left: _ImageSignature, right: _ImageSignature) -> float:
    if left.digest == right.digest:
        return 1.0
    hash_similarity = 1.0 - ((left.dhash ^ right.dhash).bit_count() / 256.0)
    histogram_similarity = sum(min(a, b) for a, b in zip(left.histogram, right.histogram)) / 1024.0
    aspect_similarity = min(left.aspect, right.aspect) / max(left.aspect, right.aspect)
    return 0.74 * hash_similarity + 0.21 * histogram_similarity + 0.05 * aspect_similarity


def _annotate_ungrouped_visual_matches(archive: ZipFile, items: list[OrderItem]) -> list[OrderItem]:
    """Attach conservative photo-match hints to items from <Без комплекта>."""
    grouped_items = [item for item in items if not item.ungrouped and item.image_path]
    ungrouped_items = [item for item in items if item.ungrouped and item.image_path]
    if not grouped_items or not ungrouped_items:
        return items

    needed_paths = {item.image_path for item in grouped_items + ungrouped_items if item.image_path}
    signatures: dict[str, _ImageSignature] = {}
    archive_names = set(archive.namelist())
    for image_path in needed_paths:
        if not image_path or image_path not in archive_names:
            continue
        signature = _image_signature(archive.read(image_path))
        if signature is not None:
            signatures[image_path] = signature

    candidates: dict[tuple[str, str], list[OrderItem]] = {}
    set_categories: dict[tuple[str, str], str] = {}
    by_stone_set: dict[tuple[str, str], list[OrderItem]] = {}
    for item in grouped_items:
        stone = canonical_stone(item.stone, item.sku)
        by_stone_set.setdefault((stone, item.set_id), []).append(item)
        key = (stone, canonical_group(item.group))
        candidates.setdefault(key, []).append(item)
    for stone_set, set_items in by_stone_set.items():
        set_categories[stone_set] = classify_set(set_items)[0]

    replacements: dict[str, OrderItem] = {}
    for item in ungrouped_items:
        signature = signatures.get(item.image_path or "")
        if signature is None:
            continue
        key = (canonical_stone(item.stone, item.sku), canonical_group(item.group))
        possible = [candidate for candidate in candidates.get(key, []) if candidate.image_path in signatures]
        if not possible:
            continue
        scored = sorted(
            ((_signature_similarity(signature, signatures[candidate.image_path or ""]), candidate) for candidate in possible),
            key=lambda pair: (-pair[0], pair[1].row),
        )
        best_score, best = scored[0]
        if best_score < 0.94:
            continue
        close_alternatives = {candidate.set_id for score, candidate in scored[1:4] if best_score - score <= 0.012}
        confirmed = best_score >= 0.965 and not close_alternatives
        replacements[item.key] = replace(
            item,
            visual_match_set_id=best.set_id,
            visual_match_sku=best.sku,
            visual_match_category=set_categories.get((canonical_stone(best.stone, best.sku), best.set_id)),
            visual_match_score=round(best_score, 3),
            visual_match_status="confirmed" if confirmed else "possible",
        )
    return [replacements.get(item.key, item) for item in items]


def _pendant_sku_parts(value: object) -> tuple[str, int | None, str]:
    """Return prefix, two-digit model year and normalized tail."""
    compact = re.sub(r"[^A-Z0-9]+", "", normalize_text(value))
    match = re.match(r"^(SKP|KP|P)(\d{2})(.*)$", compact)
    if not match:
        return "", None, compact
    prefix, year_text, tail = match.groups()
    try:
        year = 2000 + int(year_text)
    except ValueError:
        year = None
    return prefix, year, tail.lstrip("N")


def _annotate_pendant_duplicates(archive: ZipFile, items: list[OrderItem]) -> list[OrderItem]:
    """Find conservative old/new SKU duplicates for inactive crosses.

    The scan starts only when at least one cross has both zero sales and zero
    working stock. Different stones are never compared.
    """
    pendants = [
        item for item in items
        if canonical_group(item.group) == "Pendant" and is_cross_item(item)
    ]
    if len(pendants) < 2:
        return items

    archive_names = set(archive.namelist())
    signatures: dict[str, _ImageSignature] = {}
    for item in pendants:
        if not item.image_path or item.image_path not in archive_names:
            continue
        signature = _image_signature(archive.read(item.image_path))
        if signature is not None:
            signatures[item.image_path] = signature

    annotations: dict[str, tuple[int, float, OrderItem]] = {}

    def set_annotation(item: OrderItem, priority: int, score: float, updated: OrderItem) -> None:
        current = annotations.get(item.key)
        if current is None or (priority, score) > (current[0], current[1]):
            annotations[item.key] = (priority, score, updated)

    for index, left in enumerate(pendants):
        for right in pendants[index + 1:]:
            if canonical_stone(left.stone, left.sku) != canonical_stone(right.stone, right.sku):
                continue
            left_inactive = left.sales <= 0 and left.working_stock <= 0
            right_inactive = right.sales <= 0 and right.working_stock <= 0
            if not (left_inactive or right_inactive):
                continue

            _left_prefix, left_year, left_tail = _pendant_sku_parts(left.sku)
            _right_prefix, right_year, right_tail = _pendant_sku_parts(right.sku)
            same_tail = bool(left_tail and right_tail and left_tail == right_tail)

            visual_score = 0.0
            if left.image_path in signatures and right.image_path in signatures:
                visual_score = _signature_similarity(
                    signatures[left.image_path or ""],
                    signatures[right.image_path or ""],
                )
            if not same_tail and visual_score < 0.94:
                continue

            score = max(visual_score, 0.985 if same_tail else 0.0)
            left_active = left.sales > 0 or left.working_stock > 0
            right_active = right.sales > 0 or right.working_stock > 0

            preferred: OrderItem | None = None
            suppressed: OrderItem | None = None
            reason = ""

            if left_active != right_active:
                preferred = left if left_active else right
                suppressed = right if left_active else left
                reason = (
                    f"Есть очень похожая модель {preferred.sku} с продажами или остатком; "
                    "текущий SKU считаем вероятным старым дублем."
                )
            elif not left_active and not right_active and left_year and right_year and left_year != right_year:
                preferred = left if left_year > right_year else right
                suppressed = right if preferred is left else left
                reason = (
                    f"Обе похожие модели без продаж и остатков; выбрана более новая модель "
                    f"{preferred.sku} ({max(left_year, right_year)})."
                )

            if preferred is not None and suppressed is not None:
                set_annotation(
                    suppressed,
                    3,
                    score,
                    replace(
                        suppressed,
                        duplicate_sku=preferred.sku,
                        duplicate_score=round(score, 3),
                        duplicate_status="suppress",
                        duplicate_reason=reason,
                        duplicate_preferred=False,
                    ),
                )
                set_annotation(
                    preferred,
                    2,
                    score,
                    replace(
                        preferred,
                        duplicate_sku=suppressed.sku,
                        duplicate_score=round(score, 3),
                        duplicate_status="preferred",
                        duplicate_reason=(
                            f"Очень похожая модель {suppressed.sku} помечена как вероятный дубль. "
                            "В расчёте оставлен этот SKU."
                        ),
                        duplicate_preferred=True,
                    ),
                )
            else:
                review_reason = (
                    f"Есть очень похожая модель {right.sku if left is not right else left.sku}; "
                    "автоматически выбрать основную позицию нельзя."
                )
                set_annotation(
                    left,
                    1,
                    score,
                    replace(
                        left,
                        duplicate_sku=right.sku,
                        duplicate_score=round(score, 3),
                        duplicate_status="review",
                        duplicate_reason=review_reason,
                    ),
                )
                set_annotation(
                    right,
                    1,
                    score,
                    replace(
                        right,
                        duplicate_sku=left.sku,
                        duplicate_score=round(score, 3),
                        duplicate_status="review",
                        duplicate_reason=review_reason,
                    ),
                )

    replacements = {key: value[2] for key, value in annotations.items()}
    return [replacements.get(item.key, item) for item in items]


# ---------------------------- XLSX parser ------------------------------------

def _shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return ["".join(node.text or "" for node in si.iter(_XML_MAIN + "t")) for si in root.findall(_XML_MAIN + "si")]


def _cell_value(cell: ET.Element, strings: list[str]) -> str:
    value = cell.find(_XML_MAIN + "v")
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter(_XML_MAIN + "t"))
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        try:
            return strings[int(value.text)]
        except (ValueError, IndexError):
            return ""
    return value.text


def _read_sheet_rows(archive: ZipFile, sheet_path: str, strings: list[str]) -> dict[int, dict[str, str]]:
    """Stream worksheet rows without keeping the large XML tree in memory."""
    rows: dict[int, dict[str, str]] = {}
    with archive.open(sheet_path) as handle:
        for _event, element in ET.iterparse(handle, events=("end",)):
            if element.tag != _XML_MAIN + "row":
                continue
            row_number = int(element.attrib.get("r", "0") or 0)
            rows[row_number] = {
                _column_letter(cell.attrib.get("r", "")): _cell_value(cell, strings)
                for cell in element.findall(_XML_MAIN + "c")
            }
            element.clear()
    return rows


def _drawing_relationship_id(archive: ZipFile, sheet_path: str) -> str:
    with archive.open(sheet_path) as handle:
        for _event, element in ET.iterparse(handle, events=("end",)):
            if element.tag == _XML_MAIN + "drawing":
                return element.attrib.get(_XML_REL + "id", "")
            element.clear()
    return ""


def _column_letter(cell_ref: str) -> str:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    return match.group(1) if match else ""


def _workbook_sheet_path(archive: ZipFile) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    sheets = workbook.find(_XML_MAIN + "sheets")
    if sheets is None or not list(sheets):
        raise ValueError("В книге нет листов.")
    sheet = list(sheets)[0]
    rel_id = sheet.attrib[_XML_REL + "id"]
    target = relmap[rel_id].lstrip("/")
    return target if target.startswith("xl/") else "xl/" + target


def _sheet_relationships(archive: ZipFile, sheet_path: str) -> dict[str, str]:
    folder, filename = posixpath.split(sheet_path)
    rel_path = posixpath.join(folder, "_rels", filename + ".rels")
    if rel_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rel_path))
    result: dict[str, str] = {}
    for rel in root:
        target = rel.attrib.get("Target", "")
        resolved = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join(folder, target))
        result[rel.attrib.get("Id", "")] = resolved
    return result


def _image_index(archive: ZipFile, sheet_path: str) -> dict[int, str]:
    drawing_rel_id = _drawing_relationship_id(archive, sheet_path)
    if not drawing_rel_id:
        return {}
    sheet_rels = _sheet_relationships(archive, sheet_path)
    drawing_path = sheet_rels.get(drawing_rel_id)
    if not drawing_path or drawing_path not in archive.namelist():
        return {}

    folder, filename = posixpath.split(drawing_path)
    rel_path = posixpath.join(folder, "_rels", filename + ".rels")
    relmap: dict[str, str] = {}
    if rel_path in archive.namelist():
        root = ET.fromstring(archive.read(rel_path))
        for rel in root:
            target = rel.attrib.get("Target", "")
            resolved = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join(folder, target))
            relmap[rel.attrib.get("Id", "")] = resolved

    result: dict[int, str] = {}
    with archive.open(drawing_path) as handle:
        for _event, anchor in ET.iterparse(handle, events=("end",)):
            if anchor.tag not in {_XML_DRAWING + "twoCellAnchor", _XML_DRAWING + "oneCellAnchor"}:
                continue
            start = anchor.find(_XML_DRAWING + "from")
            picture = anchor.find(_XML_DRAWING + "pic")
            if start is not None and picture is not None:
                row_node = start.find(_XML_DRAWING + "row")
                blip = picture.find(".//" + _XML_A + "blip")
                if row_node is not None and row_node.text is not None and blip is not None:
                    media_path = relmap.get(blip.attrib.get(_XML_REL + "embed", ""))
                    if media_path:
                        # Drawing coordinates are zero-based; worksheet rows are one-based.
                        result.setdefault(int(row_node.text) + 1, media_path)
            anchor.clear()
    return result


def _extract_period_and_supplier(rows: dict[int, dict[str, str]]) -> tuple[str, str]:
    period = ""
    supplier = ""
    for row_number in range(1, 8):
        text = " ".join(rows.get(row_number, {}).values())
        if "Продажи товаров за период" in text:
            period = text.replace("Продажи товаров за период", "").strip()
        if "Поставщик(и):" in text:
            supplier = text.split("Поставщик(и):", 1)[-1].strip()
    return period, supplier


def report_month_count(period: object) -> int:
    """Supplier-order recommendations always use the approved four-month window.

    The caption is retained for display only. Different date formatting or a
    report exported on another computer must never silently change quantities.
    """
    _ = period
    return DEFAULT_REPORT_MONTHS


def _detect_columns(rows: dict[int, dict[str, str]]) -> tuple[str, str, list[str], dict[str, str]]:
    """Detect required supplier-report columns or fail with an explicit error.

    Earlier releases silently fell back to E/G/O/N when a header was renamed.
    That could produce plausible but incorrect orders. 1.10.5 refuses to
    calculate until the source workbook contains every required heading.
    """
    row7 = rows.get(7, {})
    row8 = rows.get(8, {})

    def find(mapping: dict[str, str], accepted: set[str]) -> str | None:
        return next((col for col, value in mapping.items() if normalize_text(value) in accepted), None)

    sales_col = find(row7, {"ПРОДАЖИ ЗА ПЕРИОД"})
    stock_start_col = find(row7, {"ОСТАТКИ", "ОСТАТОК"})
    tvp_col = find(row7, {"ТВП"})
    total_col = find(row8, {"ВСЕГО", "TOTAL"})
    missing: list[str] = []
    if not sales_col:
        missing.append("Продажи за период")
    if not stock_start_col:
        missing.append("Остатки")
    if not tvp_col:
        missing.append("ТВП")
    if not total_col:
        missing.append("Всего")
    if missing:
        found = [str(value).strip() for value in (*row7.values(), *row8.values()) if str(value).strip()]
        preview = ", ".join(found[:24]) or "заголовки отсутствуют"
        raise ValueError(
            "Не найдены обязательные колонки: " + ", ".join(missing) +
            ". Проверьте формат отчёта. Найденные заголовки: " + preview
        )

    def col_number(letter: str) -> int:
        result = 0
        for character in letter:
            result = result * 26 + (ord(character) - 64)
        return result

    assert stock_start_col is not None and total_col is not None
    stock_start_number = col_number(stock_start_col)
    total_number = col_number(total_col)
    if total_number <= stock_start_number:
        raise ValueError("Колонка «Всего» расположена раньше блока остатков. Проверьте структуру отчёта.")
    store_columns: list[str] = []
    store_names: dict[str, str] = {}
    for col, value in row8.items():
        number = col_number(col)
        name = " ".join(str(value or "").strip().split())
        if stock_start_number <= number < total_number and name:
            store_columns.append(col)
            store_names[col] = name
    if not store_columns:
        raise ValueError("Между колонками «Остатки» и «Всего» не найдены магазины.")
    assert sales_col is not None and tvp_col is not None
    return sales_col, tvp_col, store_columns, {**store_names, "__total__": total_col}


def parse_order_workbook(path: str | Path, source_name: str | None = None, source_hash: str | None = None) -> ParsedOrderWorkbook:
    workbook_path = Path(path)
    source_name = source_name or workbook_path.name
    if source_hash is None:
        digest = hashlib.sha256()
        with workbook_path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        source_hash = digest.hexdigest()

    parse_started = time.perf_counter()
    with ZipFile(workbook_path) as archive:
        sheet_path = _workbook_sheet_path(archive)
        strings = _shared_strings(archive)
        row_values = _read_sheet_rows(archive, sheet_path, strings)
        if not row_values:
            raise ValueError("В листе нет данных.")

        period, supplier = _extract_period_and_supplier(row_values)
        report_months = report_month_count(period)
        sales_col, tvp_col, store_cols, names = _detect_columns(row_values)
        total_col = names.pop("__total__")
        image_index = _image_index(archive, sheet_path)
        actual_ntr2 = any(normalize_text(name) == "NTR2" for name in names.values())

        current_set = ""
        in_ungrouped_section = False
        items: list[OrderItem] = []
        workbook_warnings: list[str] = []
        for row_number in sorted(row_values):
            if row_number < 11:
                continue
            values = row_values[row_number]
            first = str(values.get("A", "") or "").strip()
            normalized_first = normalize_text(first).strip("<>")
            if normalized_first.startswith("SET#"):
                current_set = first
                in_ungrouped_section = False
                continue
            if normalized_first == "БЕЗ КОМПЛЕКТА":
                current_set = "Без комплекта"
                in_ungrouped_section = True
                continue
            stone = str(values.get("B", "") or "").strip()
            group = str(values.get("C", "") or "").strip()
            # Every named section row starts a new supplier set, not only rows
            # whose title begins with Set#. The current workbook also contains
            # Russian family names such as «Короны BS» and «Кресты BS». Such
            # rows have a title in column A but no stone/group and must stop the
            # previous set immediately.
            if first and not stone and not group:
                current_set = first
                in_ungrouped_section = False
                continue
            if not first or not stone or not group:
                continue
            if not current_set:
                current_set = "Без комплекта"
                in_ungrouped_section = True

            stores = {names[col]: safe_int(values.get(col)) for col in store_cols if col in names}
            total = safe_int(values.get(total_col))
            stock_63 = sum(qty for name, qty in stores.items() if is_store_63(name))
            stock_20, store_20_warning = resolve_store_20_stock(stores)
            stock_tt = sum(qty for name, qty in stores.items() if is_tt_outlet_store(name))
            stock_tt_warehouse = sum(qty for name, qty in stores.items() if is_stock_tt_store(name))
            eligible_names = {
                normalize_text(name)
                for name in stores
                if is_ordinary_working_store(name)
            }
            if not actual_ntr2:
                eligible_names.add("NTR2")
            eligible_store_count = len(eligible_names)
            # Deprecated compatibility field. Princess Hang is not a separate
            # store: it is already included in the unified ``stock_20`` value.
            stock_princess_hang = 0
            ntr2, calculated, ntr2_warning = infer_ntr2(total, stores, actual_ntr2)
            working_raw = total - stock_63 - stock_20
            errors: list[str] = []
            if working_raw < 0:
                errors.append(f"Рабочий остаток отрицательный: {working_raw}")
            if store_20_warning:
                errors.append(store_20_warning)
            if ntr2_warning:
                errors.append(ntr2_warning)
            tvp = safe_int(values.get(tvp_col))
            if tvp < 0:
                errors.append(f"Ошибка ТВП: {tvp}")
            items.append(OrderItem(
                row=row_number,
                set_id=current_set,
                sku=first,
                stone=stone,
                group=group,
                sales=max(0, safe_int(values.get(sales_col))),
                stock_63=max(0, stock_63),
                stock_20=max(0, stock_20),
                stores=stores,
                total_stock=max(0, total),
                working_stock=max(0, working_raw),
                ntr2_stock=max(0, ntr2),
                ntr2_calculated=calculated,
                tvp_raw=tvp,
                stock_tt=max(0, stock_tt),
                stock_tt_warehouse=max(0, stock_tt_warehouse),
                stock_princess_hang=max(0, stock_princess_hang),
                report_months=report_months,
                eligible_store_count=eligible_store_count,
                image_path=image_index.get(row_number),
                ungrouped=in_ungrouped_section,
                errors=tuple(errors),
            ))

        if not items:
            raise ValueError("Не найдены строки изделий. Проверьте структуру отчёта.")
        items = _annotate_ungrouped_visual_matches(archive, items)
        items = _annotate_pendant_duplicates(archive, items)
        if not actual_ntr2:
            workbook_warnings.append("Колонки NTR2 пока нет: остаток NTR2 восстановлен как «Всего минус все явные магазины».")
        diagnostic_event(
            "supplier_order.parse_workbook",
            source_name=source_name,
            size_bytes=workbook_path.stat().st_size,
            items=len(items),
            duration_ms=round((time.perf_counter() - parse_started) * 1000, 1),
        )
        return ParsedOrderWorkbook(
            source_name=source_name,
            source_hash=source_hash,
            upload_path=str(workbook_path),
            period=period,
            supplier=supplier,
            store_columns=tuple(names.values()),
            has_actual_ntr2=actual_ntr2,
            items=tuple(items),
            report_months=report_months,
            warnings=tuple(workbook_warnings),
        )


def store_uploaded_workbook(name: str, payload: bytes) -> tuple[Path, str]:
    digest = hashlib.sha256(payload).hexdigest()
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(name).suffix.lower() or ".xlsx"
    target = UPLOAD_DIR / f"{digest}{suffix}"
    if not target.exists() or target.stat().st_size != len(payload):
        temporary = target.with_suffix(target.suffix + ".tmp")
        temporary.write_bytes(payload)
        temporary.replace(target)

    storage = get_cloud_storage()
    if storage is not None:
        # The workbook is immutable and addressed by its SHA-256. Re-uploading
        # exactly the same report therefore performs only a cheap existence
        # check, while a new report is safely copied to durable object storage.
        storage.save_workbook(digest, name, payload)
    return target, digest


@st.cache_resource(show_spinner=False, max_entries=6)
def cached_parse_order_workbook(path: str, source_name: str, source_hash: str) -> ParsedOrderWorkbook:
    return parse_order_workbook(path, source_name=source_name, source_hash=source_hash)


@st.cache_data(show_spinner=False, max_entries=12)
def load_visible_images(path: str, image_paths: tuple[str, ...]) -> dict[str, bytes]:
    result: dict[str, bytes] = {}
    if not image_paths:
        return result
    with ZipFile(path) as archive:
        names = set(archive.namelist())
        for image_path in image_paths:
            if image_path in names:
                result[image_path] = archive.read(image_path)
    return result


# ---------------------------- draft persistence ------------------------------

def _connect_drafts() -> sqlite3.Connection:
    DRAFT_DB.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DRAFT_DB, timeout=15)
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS order_drafts (
            draft_key TEXT PRIMARY KEY,
            source_hash TEXT NOT NULL,
            mode TEXT NOT NULL,
            payload TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS order_receipts (
            source_hash TEXT NOT NULL,
            mode TEXT NOT NULL,
            received INTEGER NOT NULL DEFAULT 0,
            received_at TEXT NOT NULL DEFAULT '',
            delivery_status TEXT NOT NULL DEFAULT 'sent',
            delivery_dates TEXT NOT NULL DEFAULT '{}',
            status_updated_at TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            PRIMARY KEY(source_hash, mode)
        )
        """
    )
    receipt_columns = {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(order_receipts)").fetchall()
    }
    if "delivery_status" not in receipt_columns:
        connection.execute(
            "ALTER TABLE order_receipts ADD COLUMN delivery_status TEXT NOT NULL DEFAULT 'sent'"
        )
    if "delivery_dates" not in receipt_columns:
        connection.execute(
            "ALTER TABLE order_receipts ADD COLUMN delivery_dates TEXT NOT NULL DEFAULT '{}'"
        )
    if "status_updated_at" not in receipt_columns:
        connection.execute(
            "ALTER TABLE order_receipts ADD COLUMN status_updated_at TEXT NOT NULL DEFAULT ''"
        )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS manual_transit_orders (
            order_id TEXT PRIMARY KEY,
            payload TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    return connection


def draft_key(source_hash: str, mode: str) -> str:
    return hashlib.sha256(f"{source_hash}|{mode}".encode("utf-8")).hexdigest()


def validate_draft_payload(payload: object) -> OrderDraft:
    if not isinstance(payload, dict):
        raise ValueError("Черновик должен быть JSON-объектом.")
    mode = str(payload.get("mode", ""))
    if mode not in ORDER_MODES:
        raise ValueError("В черновике не указан корректный тип заказа.")

    payload_version = max(1, safe_int(payload.get("version", 1)))
    if payload_version == 1:
        # Version 1 automatically prefilled recommendations. The new workflow
        # starts every item from zero, therefore legacy auto-seeded quantities
        # must not silently appear in the final Excel.
        return OrderDraft(
            source_hash=str(payload.get("source_hash", "")),
            source_name=str(payload.get("source_name", "")),
            mode=mode,
            version=DRAFT_VERSION,
        )

    orders = {str(k): max(0, safe_int(v)) for k, v in dict(payload.get("orders", {})).items()}
    sizes: dict[str, dict[str, int]] = {}
    for key, values in dict(payload.get("sizes", {})).items():
        if isinstance(values, dict):
            sizes[str(key)] = {str(size): max(0, safe_int(qty)) for size, qty in values.items() if str(size) in {str(x) for x in RING_SIZES}}
    return OrderDraft(
        source_hash=str(payload.get("source_hash", "")),
        source_name=str(payload.get("source_name", "")),
        mode=mode,
        version=DRAFT_VERSION,
        orders=orders,
        sizes=sizes,
        stock_checked={str(k): bool(v) for k, v in dict(payload.get("stock_checked", {})).items()},
        manual_edit={str(k): bool(v) for k, v in dict(payload.get("manual_edit", {})).items()},
        limited_orders={str(k): bool(v) for k, v in dict(payload.get("limited_orders", {})).items()},
        lock_changes={
            str(k): str(v)
            for k, v in dict(payload.get("lock_changes", {})).items()
            if str(v) in EARRING_LOCKS
        },
        recommendation_profile=(
            str(payload.get("recommendation_profile", RECOMMENDATION_BASE))
            if str(payload.get("recommendation_profile", RECOMMENDATION_BASE)) in RECOMMENDATION_PROFILES
            else RECOMMENDATION_BASE
        ),
        stage=str(payload.get("stage", "order")) if str(payload.get("stage", "order")) in {"order", "rings"} else "order",
        selected_stone=str(payload.get("selected_stone", "")),
        status="completed" if str(payload.get("status", "draft")) == "completed" else "draft",
        created_at=str(payload.get("created_at", "")),
        updated_at=str(payload.get("updated_at", "")),
    )


def _load_local_draft_payload(source_hash: str, mode: str) -> dict[str, Any] | None:
    key = draft_key(source_hash, mode)
    try:
        with closing(_connect_drafts()) as connection:
            row = connection.execute("SELECT payload FROM order_drafts WHERE draft_key = ?", (key,)).fetchone()
    except sqlite3.Error:
        return None
    if not row:
        return None
    try:
        raw = json.loads(row[0])
    except (TypeError, json.JSONDecodeError):
        return None
    return raw if isinstance(raw, dict) else None


def load_draft(source_hash: str, source_name: str, mode: str) -> OrderDraft:
    # Durable cloud storage is the source of truth after a redeploy or when the
    # order is opened from another computer. SQLite remains an immediate local
    # cache and a fallback when object storage is temporarily unavailable.
    raw: dict[str, Any] | None = None
    storage = get_cloud_storage()
    if storage is not None:
        try:
            raw = storage.load_draft(source_hash, mode)
        except CloudStorageError:
            raw = None
    if raw is None:
        raw = _load_local_draft_payload(source_hash, mode)
    if raw is not None:
        try:
            draft = validate_draft_payload(raw)
            draft.source_hash = source_hash
            draft.source_name = source_name
            draft.mode = mode
            return draft
        except (ValueError, TypeError):
            pass
    return OrderDraft(source_hash=source_hash, source_name=source_name, mode=mode)


def _save_draft_locally(payload: dict[str, Any]) -> None:
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    key = draft_key(str(payload.get("source_hash", "")), str(payload.get("mode", "")))
    with closing(_connect_drafts()) as connection:
        connection.execute(
            """
            INSERT INTO order_drafts(draft_key, source_hash, mode, payload, updated_at)
            VALUES(?, ?, ?, ?, ?)
            ON CONFLICT(draft_key) DO UPDATE SET
                payload = excluded.payload,
                updated_at = excluded.updated_at
            """,
            (
                key,
                str(payload.get("source_hash", "")),
                str(payload.get("mode", "")),
                serialized,
                str(payload.get("updated_at", "")),
            ),
        )
        connection.commit()


def save_draft(draft: OrderDraft, *, sync_cloud: bool = True) -> str:
    """Persist a draft locally and optionally flush it to durable cloud storage."""
    payload = draft.as_payload()
    _save_draft_locally(payload)
    if sync_cloud:
        storage = get_cloud_storage()
        if storage is not None:
            with timed_operation("supplier_order.cloud_save", mode=draft.mode):
                storage.save_draft(payload)
    return draft.updated_at


def _local_receipt_status(source_hash: str, mode: str) -> dict[str, Any]:
    try:
        with closing(_connect_drafts()) as connection:
            row = connection.execute(
                """
                SELECT received, received_at, delivery_status, delivery_dates, status_updated_at, updated_at
                FROM order_receipts
                WHERE source_hash = ? AND mode = ?
                """,
                (source_hash, mode),
            ).fetchone()
    except sqlite3.Error:
        return {
            "delivery_status": DELIVERY_STATUS_SENT,
            "status_updated_at": "",
            "received": False,
            "received_at": "",
            "updated_at": "",
        }
    if not row:
        return {
            "delivery_status": DELIVERY_STATUS_SENT,
            "status_updated_at": "",
            "received": False,
            "received_at": "",
            "updated_at": "",
        }
    delivery_status = normalize_delivery_status(row[2], received=row[0])
    try:
        raw_dates = json.loads(str(row[3] or "{}"))
    except json.JSONDecodeError:
        raw_dates = {}
    delivery_dates = normalize_delivery_dates(
        raw_dates,
        received_at=row[1],
        status=delivery_status,
        status_updated_at=row[4],
    )
    return {
        "delivery_status": delivery_status,
        "delivery_dates": delivery_dates,
        "status_updated_at": str(row[4] or ""),
        "received": delivery_status == DELIVERY_STATUS_RECEIVED,
        "received_at": str(row[1] or ""),
        "updated_at": str(row[5] or ""),
    }


def _save_local_delivery_status(
    source_hash: str,
    mode: str,
    delivery_status: str,
    delivery_dates: dict[str, str],
    received_at: str,
    status_updated_at: str,
    updated_at: str,
) -> None:
    normalized_status = normalize_delivery_status(delivery_status)
    dates = normalize_delivery_dates(
        delivery_dates,
        received_at=received_at,
        status=normalized_status,
        status_updated_at=status_updated_at,
    )
    with closing(_connect_drafts()) as connection:
        connection.execute(
            """
            INSERT INTO order_receipts(
                source_hash, mode, received, received_at,
                delivery_status, delivery_dates, status_updated_at, updated_at
            )
            VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(source_hash, mode) DO UPDATE SET
                received = excluded.received,
                received_at = excluded.received_at,
                delivery_status = excluded.delivery_status,
                delivery_dates = excluded.delivery_dates,
                status_updated_at = excluded.status_updated_at,
                updated_at = excluded.updated_at
            """,
            (
                source_hash,
                mode,
                int(normalized_status == DELIVERY_STATUS_RECEIVED),
                received_at,
                normalized_status,
                json.dumps(dates, ensure_ascii=False, separators=(",", ":")),
                status_updated_at,
                updated_at,
            ),
        )
        connection.commit()


def _save_local_receipt_status(source_hash: str, mode: str, received: bool, received_at: str, updated_at: str) -> None:
    """Compatibility helper for tests and cached data from older builds."""
    status = DELIVERY_STATUS_RECEIVED if bool(received) else DELIVERY_STATUS_SENT
    dates = normalize_delivery_dates({}, received_at=received_at, status=status, status_updated_at=updated_at)
    _save_local_delivery_status(source_hash, mode, status, dates, received_at, updated_at, updated_at)


def set_order_delivery_status(
    source_hash: str,
    mode: str,
    delivery_status: str,
    *,
    status_date: str = "",
    delivery_dates: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Set one of four dated operational statuses for a completed suborder."""
    normalized_status = normalize_delivery_status(delivery_status)
    now = datetime.now().isoformat(timespec="seconds")
    current = _local_receipt_status(source_hash, mode)
    dates = normalize_delivery_dates(
        delivery_dates if delivery_dates is not None else current.get("delivery_dates", {}),
        received_at=current.get("received_at", ""),
        status=current.get("delivery_status", DELIVERY_STATUS_SENT),
        status_updated_at=current.get("status_updated_at", ""),
    )
    active_field = DELIVERY_DATE_FIELDS[normalized_status]
    dates[active_field] = _date_only(status_date) or dates.get(active_field) or date.today().isoformat()
    for later_status in DELIVERY_STATUSES[delivery_status_rank(normalized_status) + 1:]:
        dates[DELIVERY_DATE_FIELDS[later_status]] = ""
    validate_delivery_timeline(normalized_status, dates)
    received_at = dates.get("received_at", "") if normalized_status == DELIVERY_STATUS_RECEIVED else ""
    _save_local_delivery_status(source_hash, mode, normalized_status, dates, received_at, now, now)
    storage = get_cloud_storage()
    if storage is not None:
        details = storage.set_mode_delivery_status(
            source_hash,
            mode,
            normalized_status,
            status_date=dates[active_field],
            delivery_dates=dates,
        )
        cloud_status = normalize_delivery_status(
            details.get("delivery_status", ""),
            received=details.get("received", False),
        )
        cloud_dates = normalize_delivery_dates(
            details.get("delivery_dates", {}),
            order_date=details.get("created_at", ""),
            received_at=details.get("received_at", ""),
            status=cloud_status,
            status_updated_at=details.get("status_updated_at", now),
        )
        _save_local_delivery_status(
            source_hash,
            mode,
            cloud_status,
            cloud_dates,
            str(details.get("received_at", "")),
            str(details.get("status_updated_at", now)),
            str(details.get("updated_at", now)),
        )
        return details
    return {
        "delivery_status": normalized_status,
        "delivery_dates": dates,
        "status_updated_at": now,
        "received": normalized_status == DELIVERY_STATUS_RECEIVED,
        "received_at": received_at,
        "updated_at": now,
    }


def set_order_received(source_hash: str, mode: str, received: bool) -> dict[str, Any]:
    """Compatibility wrapper for the former checkbox-based status control."""
    status = DELIVERY_STATUS_RECEIVED if bool(received) else DELIVERY_STATUS_SENT
    if not received:
        return set_order_delivery_status(source_hash, mode, status)
    current = _local_receipt_status(source_hash, mode)
    today = date.today().isoformat()
    dates = normalize_delivery_dates(
        current.get("delivery_dates", {}),
        received_at=current.get("received_at", ""),
        status=current.get("delivery_status", DELIVERY_STATUS_SENT),
        status_updated_at=current.get("status_updated_at", ""),
    )
    for stage in DELIVERY_STATUSES:
        dates[DELIVERY_DATE_FIELDS[stage]] = dates.get(DELIVERY_DATE_FIELDS[stage]) or today
    return set_order_delivery_status(
        source_hash,
        mode,
        status,
        status_date=dates["received_at"],
        delivery_dates=dates,
    )



def _manual_order_from_payload(payload: object, *, storage: str = "local") -> ManualTransitOrder | None:
    if not isinstance(payload, dict):
        return None
    order_id = str(payload.get("order_id", "")).strip()
    title = str(payload.get("title", "")).strip()
    if not order_id or not title:
        return None
    return ManualTransitOrder(
        order_id=order_id,
        title=title,
        order_date=str(payload.get("order_date", "")),
        note=str(payload.get("note", "")),
        quantity=max(0, safe_int(payload.get("quantity", 0))),
        delivery_status=normalize_delivery_status(
            payload.get("delivery_status", ""),
            received=payload.get("received", False),
        ),
        delivery_dates=normalize_delivery_dates(
            payload.get("delivery_dates", {}),
            order_date=payload.get("order_date", ""),
            received_at=payload.get("received_at", ""),
            status=payload.get("delivery_status", DELIVERY_STATUS_SENT),
            status_updated_at=payload.get("status_updated_at", ""),
        ),
        created_at=str(payload.get("created_at", "")),
        updated_at=str(payload.get("updated_at", "")),
        received_at=str(payload.get("received_at", "")),
        status_updated_at=str(payload.get("status_updated_at", "")),
        storage=storage,
    )


def _save_manual_order_locally(payload: dict[str, Any]) -> None:
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    with closing(_connect_drafts()) as connection:
        connection.execute(
            """
            INSERT INTO manual_transit_orders(order_id, payload, updated_at)
            VALUES(?, ?, ?)
            ON CONFLICT(order_id) DO UPDATE SET
                payload = excluded.payload,
                updated_at = excluded.updated_at
            """,
            (str(payload.get("order_id", "")), serialized, str(payload.get("updated_at", ""))),
        )
        connection.commit()


def save_manual_transit_order(order: ManualTransitOrder) -> ManualTransitOrder:
    payload = order.as_payload()
    _save_manual_order_locally(payload)
    storage = get_cloud_storage()
    if storage is not None:
        payload = storage.save_manual_order(payload)
        _save_manual_order_locally(payload)
    normalized = _manual_order_from_payload(payload, storage="cloud" if storage is not None else "local")
    if normalized is None:
        raise ValueError("Не удалось сохранить ручной заказ.")
    return normalized


def list_manual_transit_orders() -> tuple[ManualTransitOrder, ...]:
    merged: dict[str, ManualTransitOrder] = {}
    try:
        with closing(_connect_drafts()) as connection:
            rows = connection.execute(
                "SELECT payload FROM manual_transit_orders ORDER BY updated_at DESC"
            ).fetchall()
    except sqlite3.Error:
        rows = []
    for (serialized,) in rows:
        try:
            payload = json.loads(serialized)
        except (TypeError, json.JSONDecodeError):
            continue
        order = _manual_order_from_payload(payload, storage="local")
        if order is not None:
            merged[order.order_id] = order

    storage = get_cloud_storage()
    if storage is not None:
        try:
            cloud_rows = storage.list_manual_orders()
        except CloudStorageError:
            cloud_rows = ()
        for payload in cloud_rows:
            order = _manual_order_from_payload(dict(payload), storage="cloud")
            if order is None:
                continue
            current = merged.get(order.order_id)
            if current is None or order.updated_at >= current.updated_at:
                merged[order.order_id] = order
                _save_manual_order_locally(dict(payload))
    return tuple(
        sorted(
            merged.values(),
            key=lambda row: (row.order_date, row.updated_at),
            reverse=True,
        )
    )


def set_manual_transit_order_status(
    order: ManualTransitOrder,
    delivery_status: str,
    *,
    status_date: str = "",
    delivery_dates: dict[str, str] | None = None,
) -> ManualTransitOrder:
    normalized_status = normalize_delivery_status(delivery_status)
    now = datetime.now().isoformat(timespec="seconds")
    dates = normalize_delivery_dates(
        delivery_dates if delivery_dates is not None else order.delivery_dates,
        order_date=order.order_date,
        received_at=order.received_at,
        status=order.delivery_status,
        status_updated_at=order.status_updated_at,
    )
    active_field = DELIVERY_DATE_FIELDS[normalized_status]
    dates[active_field] = _date_only(status_date) or dates.get(active_field) or date.today().isoformat()
    for later_status in DELIVERY_STATUSES[delivery_status_rank(normalized_status) + 1:]:
        dates[DELIVERY_DATE_FIELDS[later_status]] = ""
    validate_delivery_timeline(normalized_status, dates)
    updated = replace(
        order,
        delivery_status=normalized_status,
        delivery_dates=dates,
        received_at=dates.get("received_at", "") if normalized_status == DELIVERY_STATUS_RECEIVED else "",
        status_updated_at=now,
    )
    return save_manual_transit_order(updated)


def set_manual_transit_order_received(order: ManualTransitOrder, received: bool) -> ManualTransitOrder:
    """Compatibility wrapper for the former checkbox-based manual order status."""
    status = DELIVERY_STATUS_RECEIVED if bool(received) else DELIVERY_STATUS_SENT
    if not received:
        return set_manual_transit_order_status(order, status)
    today = date.today().isoformat()
    dates = normalize_delivery_dates(
        order.delivery_dates,
        order_date=order.order_date,
        received_at=order.received_at,
        status=order.delivery_status,
        status_updated_at=order.status_updated_at,
    )
    for stage in DELIVERY_STATUSES:
        dates[DELIVERY_DATE_FIELDS[stage]] = dates.get(DELIVERY_DATE_FIELDS[stage]) or today
    return set_manual_transit_order_status(
        order,
        status,
        status_date=dates["received_at"],
        delivery_dates=dates,
    )


def delete_manual_transit_order(order_id: str) -> None:
    storage = get_cloud_storage()
    if storage is not None:
        storage.delete_manual_order(order_id)
    with closing(_connect_drafts()) as connection:
        connection.execute("DELETE FROM manual_transit_orders WHERE order_id = ?", (order_id,))
        connection.commit()


def purge_order_workspaces_except(source_hash: str) -> tuple[int, int]:
    """Keep previous workspaces as recoverable history.

    A new workbook has a different SHA-256 and therefore receives an isolated
    draft automatically. Earlier versions physically deleted every other draft
    here, which made recovery impossible. Version 1.8.8 clears only active
    widget/session state; persisted local and cloud workspaces are retained
    until the user explicitly deletes them.
    """
    _ = source_hash
    return 0, 0

def _find_uploaded_workbook(source_hash: str) -> Path | None:
    """Return the persisted source workbook for a draft hash, if it still exists."""

    if not source_hash or not UPLOAD_DIR.exists():
        return None
    candidates = sorted(
        (
            path
            for path in UPLOAD_DIR.glob(f"{source_hash}.*")
            if path.is_file() and not path.name.endswith(".tmp")
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _list_local_saved_order_workspaces() -> tuple[SavedOrderWorkspace, ...]:
    """List resumable reports together with their locally cached selections."""

    try:
        with closing(_connect_drafts()) as connection:
            rows = connection.execute(
                "SELECT source_hash, mode, payload, updated_at FROM order_drafts ORDER BY updated_at DESC"
            ).fetchall()
    except sqlite3.Error:
        return ()

    grouped: dict[str, dict[str, Any]] = {}
    for source_hash, mode, serialized, updated_at in rows:
        workbook_path = _find_uploaded_workbook(str(source_hash))
        if workbook_path is None:
            continue
        try:
            draft = validate_draft_payload(json.loads(serialized))
        except (ValueError, TypeError, json.JSONDecodeError):
            continue
        normalized_mode = str(mode) if str(mode) in ORDER_MODES else draft.mode
        selected_positions = sum(max(0, safe_int(value)) > 0 for value in draft.orders.values())
        total_quantity = sum(max(0, safe_int(value)) for value in draft.orders.values())
        limited_positions = sum(bool(value) for value in draft.limited_orders.values())
        draft_updated = str(updated_at or draft.updated_at or "")
        draft_created = str(draft.created_at or draft_updated)
        receipt = _local_receipt_status(str(source_hash), normalized_mode)
        details = {
            "created_at": draft_created,
            "updated_at": max(draft_updated, str(receipt.get("updated_at", ""))),
            "selected_positions": selected_positions,
            "total_quantity": total_quantity,
            "limited_positions": limited_positions,
            "stage": draft.stage,
            "status": draft.status,
            "delivery_status": normalize_delivery_status(
                receipt.get("delivery_status", ""),
                received=receipt.get("received", False),
            ),
            "delivery_dates": normalize_delivery_dates(
                receipt.get("delivery_dates", {}),
                order_date=draft.updated_at or draft.created_at,
                received_at=receipt.get("received_at", ""),
                status=receipt.get("delivery_status", DELIVERY_STATUS_SENT),
                status_updated_at=receipt.get("status_updated_at", ""),
            ),
            "status_updated_at": str(receipt.get("status_updated_at", "")),
            "received": bool(receipt.get("received", False)),
            "received_at": str(receipt.get("received_at", "")),
        }
        record = grouped.setdefault(
            str(source_hash),
            {
                "source_name": draft.source_name or workbook_path.name,
                "upload_path": str(workbook_path),
                "created_at": draft_created,
                "updated_at": draft_updated,
                "preferred_mode": normalized_mode,
                "mode_details": {},
            },
        )
        record["mode_details"][normalized_mode] = details
        if draft_created and (not record["created_at"] or draft_created < record["created_at"]):
            record["created_at"] = draft_created
        if draft_updated >= str(record["updated_at"]):
            record["updated_at"] = draft_updated
            record["source_name"] = draft.source_name or record["source_name"]
            record["preferred_mode"] = normalized_mode

    result: list[SavedOrderWorkspace] = []
    for source_hash, values in grouped.items():
        mode_details = dict(values["mode_details"])
        modes = tuple(mode for mode in ORDER_MODES if mode in mode_details)
        selected_positions = sum(safe_int(row.get("selected_positions", 0)) for row in mode_details.values())
        total_quantity = sum(safe_int(row.get("total_quantity", 0)) for row in mode_details.values())
        limited_positions = sum(safe_int(row.get("limited_positions", 0)) for row in mode_details.values())
        statuses = [str(row.get("status", "draft")) for row in mode_details.values()]
        result.append(
            SavedOrderWorkspace(
                source_hash=source_hash,
                source_name=str(values["source_name"]),
                upload_path=str(values["upload_path"]),
                created_at=str(values["created_at"]),
                updated_at=str(values["updated_at"]),
                modes=modes,
                preferred_mode=str(values["preferred_mode"]),
                selected_positions=selected_positions,
                total_quantity=total_quantity,
                limited_positions=limited_positions,
                status="completed" if statuses and all(value == "completed" for value in statuses) else "draft",
                mode_details=mode_details,
                storage="local",
            )
        )
    return tuple(sorted(result, key=lambda workspace: workspace.updated_at, reverse=True))


def _cloud_workspace_from_manifest(manifest: dict[str, Any]) -> SavedOrderWorkspace | None:
    """Build a lightweight workspace from a cloud index row or manifest."""
    source_hash = str(manifest.get("source_hash", "")).strip()
    source_name = str(manifest.get("source_name", "")).strip() or f"{source_hash}.xlsx"
    workbook_key = str(manifest.get("workbook_key", "")).strip()
    if not source_hash or not workbook_key:
        return None
    drafts = manifest.get("drafts", {})
    if not isinstance(drafts, dict):
        drafts = {}
    mode_details: dict[str, dict[str, Any]] = {}
    preferred_mode = ORDER_MODE_STONES
    updated_at = str(manifest.get("updated_at", ""))
    created_at = str(manifest.get("created_at", ""))
    for mode in ORDER_MODES:
        details = drafts.get(mode)
        if not isinstance(details, dict):
            continue
        delivery_status = normalize_delivery_status(
            details.get("delivery_status", ""),
            received=details.get("received", False),
        )
        normalized = {
            "created_at": str(details.get("created_at", "")) or created_at,
            "updated_at": str(details.get("updated_at", "")) or updated_at,
            "selected_positions": max(0, safe_int(details.get("selected_positions", 0))),
            "total_quantity": max(0, safe_int(details.get("total_quantity", 0))),
            "limited_positions": max(0, safe_int(details.get("limited_positions", 0))),
            "stage": str(details.get("stage", "order")),
            "status": "completed" if str(details.get("status", "draft")) == "completed" else "draft",
            "delivery_status": delivery_status,
            "delivery_dates": normalize_delivery_dates(
                details.get("delivery_dates", {}),
                order_date=details.get("updated_at", "") or details.get("created_at", ""),
                received_at=details.get("received_at", ""),
                status=delivery_status,
                status_updated_at=details.get("status_updated_at", ""),
            ),
            "status_updated_at": str(details.get("status_updated_at", "")),
            "received": delivery_status == DELIVERY_STATUS_RECEIVED,
            "received_at": str(details.get("received_at", "")),
        }
        mode_details[mode] = normalized
        candidate_updated = normalized["updated_at"]
        if candidate_updated >= updated_at:
            preferred_mode = mode
            updated_at = candidate_updated or updated_at
    modes = tuple(mode for mode in ORDER_MODES if mode in mode_details)
    statuses = [str(row.get("status", "draft")) for row in mode_details.values()]
    status = "completed" if statuses and all(value == "completed" for value in statuses) else "draft"
    return SavedOrderWorkspace(
        source_hash=source_hash,
        source_name=source_name,
        upload_path="",
        created_at=created_at or updated_at,
        updated_at=updated_at,
        modes=modes,
        preferred_mode=preferred_mode,
        selected_positions=sum(safe_int(row.get("selected_positions", 0)) for row in mode_details.values()),
        total_quantity=sum(safe_int(row.get("total_quantity", 0)) for row in mode_details.values()),
        limited_positions=sum(safe_int(row.get("limited_positions", 0)) for row in mode_details.values()),
        status=status,
        mode_details=mode_details,
        storage="cloud",
    )


def list_saved_order_workspaces(
    *,
    refresh_cloud: bool = False,
    include_completed: bool = False,
) -> tuple[SavedOrderWorkspace, ...]:
    """Return the cloud order library, merged with a surviving local cache.

    Normal reads use the compact ``orders-index.json`` object. ``refresh_cloud``
    deliberately rebuilds that index from manifests and is wired to the UI's
    «Обновить список» button.
    """
    local_workspaces = _list_local_saved_order_workspaces()
    merged: dict[str, SavedOrderWorkspace] = {
        workspace.source_hash: workspace for workspace in local_workspaces
    }
    storage = get_cloud_storage()
    if storage is not None:
        for workspace in local_workspaces:
            path = Path(workspace.upload_path)
            if not path.exists():
                continue
            try:
                storage.save_workbook(workspace.source_hash, workspace.source_name, path.read_bytes())
                for mode in workspace.modes:
                    payload = _load_local_draft_payload(workspace.source_hash, mode)
                    if payload:
                        storage.save_draft(payload)
            except (CloudStorageError, OSError):
                continue
        try:
            rows = storage.list_order_index(refresh=refresh_cloud)
        except CloudStorageError:
            rows = ()
        for row in rows:
            workspace = _cloud_workspace_from_manifest(dict(row))
            if workspace is not None:
                merged[workspace.source_hash] = workspace
    values = merged.values()
    if not include_completed:
        values = (workspace for workspace in values if workspace.status != "completed")
    return tuple(sorted(values, key=lambda workspace: workspace.updated_at, reverse=True))


def _delete_local_order_workspace(source_hash: str) -> tuple[int, int]:
    """Delete SQLite drafts and every cached workbook for one source hash."""
    deleted_rows = 0
    deleted_files = 0
    try:
        with closing(_connect_drafts()) as connection:
            cursor = connection.execute("DELETE FROM order_drafts WHERE source_hash = ?", (source_hash,))
            deleted_rows = max(0, int(cursor.rowcount or 0))
            connection.execute("DELETE FROM order_receipts WHERE source_hash = ?", (source_hash,))
            connection.commit()
    except sqlite3.Error as exc:
        raise OSError(f"Не удалось очистить локальный черновик: {exc}") from exc
    if UPLOAD_DIR.exists():
        for path in UPLOAD_DIR.glob(f"{source_hash}.*"):
            if not path.is_file():
                continue
            try:
                path.unlink(missing_ok=True)
                deleted_files += 1
            except OSError as exc:
                raise OSError(f"Не удалось удалить локальный файл {path.name}: {exc}") from exc
    return deleted_rows, deleted_files


def delete_saved_order_workspace(workspace: SavedOrderWorkspace) -> tuple[int, int, int]:
    """Delete one order from Cloudflare, local cache and the current session."""
    cloud_deleted = 0
    storage = get_cloud_storage()
    if workspace.storage == "cloud" or storage is not None:
        if storage is None:
            raise CloudStorageError("Облачное хранилище недоступно — заказ не удалён.")
        cloud_deleted = len(storage.delete_workspace(workspace.source_hash))

    local_rows, local_files = _delete_local_order_workspace(workspace.source_hash)
    try:
        _clear_order_widget_state()
    except NameError:
        pass
    for key in list(st.session_state.keys()):
        text = str(key)
        if workspace.source_hash in text or text.startswith("supplier_order_delete_confirm::"):
            st.session_state.pop(key, None)
    st.session_state.pop("supplier_order_mode", None)
    active = st.session_state.get(ACTIVE_WORKSPACE_KEY)
    if isinstance(active, dict) and str(active.get("source_hash", "")) == workspace.source_hash:
        st.session_state.pop(ACTIVE_WORKSPACE_KEY, None)
        st.session_state.pop("supplier_order_upload", None)
    try:
        cached_parse_order_workbook.clear()
        load_visible_images.clear()
    except AttributeError:
        pass
    return cloud_deleted, local_rows, local_files


def load_saved_order_workspace(workspace: SavedOrderWorkspace) -> ParsedOrderWorkbook:
    path = Path(workspace.upload_path) if workspace.upload_path else _find_uploaded_workbook(workspace.source_hash)
    if path is None or not path.exists():
        storage = get_cloud_storage()
        if storage is None:
            raise FileNotFoundError("Сохранённый исходный отчёт не найден локально, а облачное хранилище не подключено.")
        path, _manifest = storage.restore_workbook(workspace.source_hash, UPLOAD_DIR)
    return cached_parse_order_workbook(
        str(path),
        workspace.source_name,
        workspace.source_hash,
    )


def draft_json_bytes(draft: OrderDraft) -> bytes:
    return json.dumps(draft.as_payload(), ensure_ascii=False, indent=2).encode("utf-8")


def full_order_backup_bytes(parsed: ParsedOrderWorkbook, draft: OrderDraft) -> bytes:
    """Portable emergency copy containing both the source report and draft."""
    source_path = Path(parsed.upload_path)
    if not source_path.exists():
        raise FileNotFoundError("Исходный Excel не найден для резервной копии.")
    output = io.BytesIO()
    safe_name = Path(parsed.source_name).name or f"{parsed.source_hash}.xlsx"
    with ZipFile(output, "w", compression=ZIP_STORED) as archive:
        archive.writestr(safe_name, source_path.read_bytes())
        archive.writestr("order_draft.json", draft_json_bytes(draft))
        archive.writestr(
            "backup_manifest.json",
            json.dumps(
                {
                    "schema_version": 1,
                    "source_hash": parsed.source_hash,
                    "source_name": parsed.source_name,
                    "mode": draft.mode,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                },
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8"),
        )
    return output.getvalue()


def import_draft_json(payload: bytes, expected_hash: str, mode: str) -> OrderDraft:
    try:
        raw = json.loads(payload.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Файл черновика не распознан.") from exc
    draft = validate_draft_payload(raw)
    if draft.mode != mode:
        raise ValueError(f"Этот черновик относится к заказу «{draft.mode}».")
    if draft.source_hash and draft.source_hash != expected_hash:
        raise ValueError("Черновик создан по другому исходному отчёту.")
    draft.source_hash = expected_hash
    save_draft(draft)
    return draft


# ---------------------------- Excel export -----------------------------------

def format_sizes(values: dict[str, int] | None) -> str:
    values = values or {}
    parts = [f"{size} × {safe_int(values.get(str(size), 0))}" for size in RING_SIZES if safe_int(values.get(str(size), 0)) > 0]
    return "; ".join(parts)


def build_supplier_excel(
    parsed: ParsedOrderWorkbook,
    selected_items: Iterable[OrderItem],
    draft: OrderDraft,
) -> bytes:
    items = [item for item in selected_items if draft.orders.get(item.key, 0) > 0 and not draft.limited_orders.get(item.key, False)]
    image_paths = tuple(sorted({item.image_path for item in items if item.image_path}))
    images = load_visible_images(parsed.upload_path, image_paths)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Order"
    headers = ["Photo", "SKU", "Stone", "Order Quantity", "Sizes", "Change Lock To"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1C1A17")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D2C4")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 28

    for row_index, item in enumerate(items, start=2):
        quantity = max(0, safe_int(draft.orders.get(item.key, 0)))
        sizes = format_sizes(draft.sizes.get(item.key)) if item.is_ring else ""
        lock_change = (
            earring_lock_export_label(draft.lock_changes.get(item.key, ""))
            if item.is_earrings
            else ""
        )
        sheet.append([
            "",
            item.sku,
            canonical_stone(item.stone, item.sku),
            quantity,
            sizes,
            lock_change,
        ])
        sheet.row_dimensions[row_index].height = 66
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E6E0D8"))
        image_data = images.get(item.image_path or "")
        if image_data:
            try:
                image = XLImage(io.BytesIO(image_data))
                image.width = 72
                image.height = 72
                sheet.add_image(image, f"A{row_index}")
            except Exception:
                sheet.cell(row_index, 1).value = "Фото не вставлено"

    widths = {"A": 14, "B": 27, "C": 28, "D": 22, "E": 38, "F": 32}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max(1, sheet.max_row)}"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_limited_order_excel(
    parsed: ParsedOrderWorkbook,
    selected_items: Iterable[OrderItem],
    draft: OrderDraft,
) -> bytes:
    items = [item for item in selected_items if draft.limited_orders.get(item.key, False)]
    image_paths = tuple(sorted({item.image_path for item in items if item.image_path}))
    images = load_visible_images(parsed.upload_path, image_paths)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Limited Order"
    headers = ["Фото", "Артикул", "Камень", "Группа", "Комплект", "Продажи", "Всего остаток", "TT", "63", "ТВП"]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="6B4F2B")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D2C4")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 28

    for row_index, item in enumerate(items, start=2):
        sheet.append([
            "", item.sku, canonical_stone(item.stone, item.sku), item.group, item.set_id,
            item.sales, item.display_stock, item.stock_tt, item.stock_63, item.tvp_raw,
        ])
        sheet.row_dimensions[row_index].height = 66
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E6E0D8"))
        image_data = images.get(item.image_path or "")
        if image_data:
            try:
                image = XLImage(io.BytesIO(image_data))
                image.width = 72
                image.height = 72
                sheet.add_image(image, f"A{row_index}")
            except Exception:
                sheet.cell(row_index, 1).value = "Фото не вставлено"

    widths = {"A": 14, "B": 27, "C": 27, "D": 18, "E": 23, "F": 12, "G": 18, "H": 10, "I": 10, "J": 10}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{max(1, sheet.max_row)}"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# ---------------------------- Streamlit UI -----------------------------------

def _draft_state_key(source_hash: str, mode: str) -> str:
    return f"supplier_order_draft::{source_hash}::{mode}"


ACTIVE_WORKSPACE_KEY = "supplier_order_active_workspace"


_ORDER_WIDGET_PREFIXES = (
    "supplier_order_draft::",
    "supplier_order_category::",
    "supplier_order_stone::",
    "supplier_order_focus_set::",
    "supplier_order_page::",
    "supplier_order_focus_page::",
    "supplier_order_active_mode::",
    "supplier_order_dirty::",
    "supplier_order_cloud_saved_at::",
    "supplier_order_output_signature::",
    "supplier_excel::",
    "limited_excel::",
    "order_qty::",
    "order_manual::",
    "limited_order::",
    "order_lock_change::",
    "supplier_order_recommendation_profile::",
    "ring_size::",
    "ring_stock_check::",
    "supplier_order_pending_widget_cleanup",
    "supplier_order_full_backup::",
    "prepare_full_backup::",
)


def _clear_order_widget_state() -> None:
    for key in list(st.session_state.keys()):
        if any(str(key).startswith(prefix) for prefix in _ORDER_WIDGET_PREFIXES):
            st.session_state.pop(key, None)
    st.session_state.pop("supplier_order_save_status", None)


def _activate_workspace(parsed: ParsedOrderWorkbook, mode: str | None = None) -> None:
    st.session_state[ACTIVE_WORKSPACE_KEY] = {
        "source_hash": parsed.source_hash,
        "source_name": parsed.source_name,
        "upload_path": parsed.upload_path,
    }
    if mode in ORDER_MODES:
        st.session_state["supplier_order_mode"] = mode


def _clear_active_workspace() -> None:
    st.session_state.pop(ACTIVE_WORKSPACE_KEY, None)
    st.session_state.pop("supplier_order_upload", None)


def _load_active_workspace() -> ParsedOrderWorkbook | None:
    raw = st.session_state.get(ACTIVE_WORKSPACE_KEY)
    if not isinstance(raw, dict):
        return None
    source_hash = str(raw.get("source_hash", ""))
    source_name = str(raw.get("source_name", ""))
    upload_path = str(raw.get("upload_path", ""))
    if not source_hash or not source_name or not upload_path or not Path(upload_path).exists():
        st.session_state.pop(ACTIVE_WORKSPACE_KEY, None)
        return None
    try:
        return cached_parse_order_workbook(upload_path, source_name, source_hash)
    except (OSError, ValueError, KeyError, BadZipFile):
        st.session_state.pop(ACTIVE_WORKSPACE_KEY, None)
        return None


def _get_session_draft(parsed: ParsedOrderWorkbook, mode: str) -> OrderDraft:
    key = _draft_state_key(parsed.source_hash, mode)
    if key not in st.session_state:
        st.session_state[key] = load_draft(parsed.source_hash, parsed.source_name, mode)
    draft = st.session_state[key]
    if not isinstance(draft, OrderDraft) or getattr(draft, "version", 1) < DRAFT_VERSION:
        draft = load_draft(parsed.source_hash, parsed.source_name, mode)
        st.session_state[key] = draft
    return draft


CLOUD_AUTOSAVE_INTERVAL_SECONDS = 12.0
ORDER_PAGE_SIZE = 10

def _draft_dirty_key(draft: OrderDraft) -> str:
    return f"supplier_order_dirty::{draft.source_hash}::{draft.mode}"

def _draft_cloud_time_key(draft: OrderDraft) -> str:
    return f"supplier_order_cloud_saved_at::{draft.source_hash}::{draft.mode}"


def _draft_output_state_signature(draft: OrderDraft) -> str:
    """Fingerprint only fields that affect generated supplier files."""
    payload = {
        "orders": {str(k): max(0, safe_int(v)) for k, v in draft.orders.items() if safe_int(v) > 0},
        "sizes": {
            str(k): {str(size): max(0, safe_int(qty)) for size, qty in values.items() if safe_int(qty) > 0}
            for k, values in draft.sizes.items()
            if any(safe_int(qty) > 0 for qty in values.values())
        },
        "limited_orders": {str(k): bool(v) for k, v in draft.limited_orders.items() if bool(v)},
        "lock_changes": {str(k): str(v) for k, v in draft.lock_changes.items() if str(v) in EARRING_LOCKS},
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _clear_generated_payloads(
    source_hash: str,
    mode: str,
    *,
    keep_keys: Iterable[str] = (),
    kinds: Iterable[str] = ("main", "limited", "backup"),
) -> int:
    """Release obsolete generated files while retaining the latest per kind."""
    keep = {str(key) for key in keep_keys}
    requested = {str(kind) for kind in kinds}
    prefix_by_kind = {
        "main": f"supplier_excel::{source_hash}::{mode}::",
        "limited": f"limited_excel::{source_hash}::{mode}::",
        "backup": f"supplier_order_full_backup::{source_hash}::{mode}",
    }
    prefixes = tuple(prefix for kind, prefix in prefix_by_kind.items() if kind in requested)
    removed = 0
    for key in list(st.session_state.keys()):
        text = str(key)
        if text in keep or not any(text.startswith(prefix) for prefix in prefixes):
            continue
        value = st.session_state.pop(key, None)
        if isinstance(value, (bytes, bytearray, memoryview)):
            removed += len(value)
    if removed:
        diagnostic_event(
            "supplier_order.generated_payloads_cleared",
            source_hash=source_hash,
            mode=mode,
            released_bytes=removed,
        )
    return removed


def _invalidate_stale_generated_payloads(draft: OrderDraft) -> None:
    state_key = f"supplier_order_output_signature::{draft.source_hash}::{draft.mode}"
    current = _draft_output_state_signature(draft)
    previous = str(st.session_state.get(state_key, ""))
    if previous and previous != current:
        _clear_generated_payloads(draft.source_hash, draft.mode)
    st.session_state[state_key] = current


def _save_session_draft(draft: OrderDraft, *, sync_cloud: bool = False) -> None:
    """Save instantly to the local cache; cloud writes are deliberately batched.

    Quantity, lock and size widgets must not wait for network latency. Explicit
    transitions, Excel preparation, completion and the manual save button call
    this function with ``sync_cloud=True``.
    """
    try:
        _invalidate_stale_generated_payloads(draft)
        saved_at = save_draft(draft, sync_cloud=sync_cloud)
        dirty_key = _draft_dirty_key(draft)
        if sync_cloud or get_cloud_storage() is None:
            st.session_state[dirty_key] = False
            st.session_state[_draft_cloud_time_key(draft)] = time.time()
            st.session_state["supplier_order_save_status"] = f"Сохранено: {saved_at.replace('T', ' ')}"
        else:
            if not st.session_state.get(dirty_key, False):
                st.session_state[_draft_cloud_time_key(draft)] = time.time()
            st.session_state[dirty_key] = True
            st.session_state["supplier_order_save_status"] = "Сохранено локально · облако обновится пакетно"
    except (sqlite3.Error, OSError, CloudStorageError) as exc:
        # Local SQLite is written before the cloud request. Keep the draft dirty
        # so the timed fragment retries the durable synchronization later.
        st.session_state[_draft_dirty_key(draft)] = True
        diagnostic_event("supplier_order.save_error", mode=draft.mode, error=str(exc))
        st.session_state["supplier_order_save_status"] = f"Локально сохранено, облако временно недоступно: {exc}"

def _flush_session_draft(draft: OrderDraft) -> None:
    _save_session_draft(draft, sync_cloud=True)

def _maybe_flush_session_draft(draft: OrderDraft) -> None:
    if not st.session_state.get(_draft_dirty_key(draft), False):
        return
    last = float(st.session_state.get(_draft_cloud_time_key(draft), 0.0) or 0.0)
    if time.time() - last >= CLOUD_AUTOSAVE_INTERVAL_SECONDS:
        _flush_session_draft(draft)


@st.fragment(run_every=CLOUD_AUTOSAVE_INTERVAL_SECONDS)
def _render_cloud_autosave_fragment(draft: OrderDraft) -> None:
    """Flush a dirty local draft even when the user stops interacting."""
    if st.session_state.get(_draft_dirty_key(draft), False):
        with timed_operation(
            "supplier_order.autosave_fragment",
            source_hash=draft.source_hash,
            mode=draft.mode,
        ):
            _flush_session_draft(draft)
    status = str(st.session_state.get("supplier_order_save_status", "Черновик ещё не сохранён"))
    st.caption(f"💾 {status}")


def _flush_workspace_session_drafts(source_hash: str) -> None:
    """Synchronize every loaded mode before leaving or replacing a workspace."""
    for mode in ORDER_MODES:
        draft = st.session_state.get(_draft_state_key(source_hash, mode))
        if isinstance(draft, OrderDraft) and st.session_state.get(_draft_dirty_key(draft), False):
            _flush_session_draft(draft)


def _flush_previous_mode_on_change(parsed: ParsedOrderWorkbook, current_mode: str) -> None:
    key = f"supplier_order_active_mode::{parsed.source_hash}"
    previous_mode = str(st.session_state.get(key, ""))
    if previous_mode in ORDER_MODES and previous_mode != current_mode:
        previous = st.session_state.get(_draft_state_key(parsed.source_hash, previous_mode))
        if isinstance(previous, OrderDraft) and st.session_state.get(_draft_dirty_key(previous), False):
            _flush_session_draft(previous)
    st.session_state[key] = current_mode


def _render_sidebar(parsed: ParsedOrderWorkbook | None, draft: OrderDraft | None) -> None:
    items = [NavigationItem(item_id=section_id, label=label, href=f"#{section_id}") for section_id, label in ORDER_SECTIONS]
    status = st.session_state.get("supplier_order_save_status", "Черновик ещё не сохранён")
    source = parsed.source_name if parsed else "Ожидается Excel-отчёт"
    result = render_sidebar(
        module_title="Заказ поставщику",
        navigation_title="Навигация заказа",
        items=items,
        status_text=status,
        status_tone="success" if str(status).startswith("Сохранено") else "neutral",
        source_text=source,
        action_label="Сохранить черновик" if draft else "Незавершённые заказы",
        action_key="supplier_order_manual_save" if draft else "supplier_order_sidebar_library",
    )
    if result.action_clicked:
        if draft:
            _flush_session_draft(draft)
        else:
            st.session_state["supplier_order_library_open"] = True
        st.rerun()
    render_mobile_navigation(items)



def _render_storage_status() -> bool:
    """Show whether orders survive refreshes, deployments and device changes."""
    status = get_cloud_storage_status()
    if status.available:
        st.success("☁️ Надёжное сохранение включено: Excel, выбранные позиции и размеры хранятся в облаке.")
        return True
    setup_hint = (
        "Добавьте раздел `[order_storage]` в Streamlit Secrets по примеру "
        "`.streamlit/secrets.toml.example`. До подключения облака локальный черновик "
        "может исчезнуть при новом деплое."
    )
    if status.configured:
        st.error(f"Облачное сохранение временно недоступно: {status.message}")
    else:
        st.error(f"Надёжное облачное сохранение ещё не настроено. {setup_hint}")
    return False


def _format_saved_order_time(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return "не указано"
    return text.replace("T", " ").replace("+00:00", " UTC")


def _saved_stage_label(value: str) -> str:
    return {
        "order": "Выбор количества",
        "rings": "Размеры колец / выгрузка",
    }.get(str(value), "Выбор количества")


def _workspace_status_label(details: dict[str, dict[str, Any]]) -> str:
    labels: list[str] = []
    for mode in ORDER_MODES:
        row = details.get(mode)
        if not isinstance(row, dict):
            state = "не начат"
        elif str(row.get("status", "draft")) == "completed":
            state = "завершён"
        else:
            state = "в работе"
        labels.append(f"{mode}: {state}")
    return " · ".join(labels)


def _render_analytics_metrics(summary: dict[str, Any], *, show_sku: bool = False) -> None:
    metrics: list[tuple[str, int]] = [("Всего, шт.", safe_int(summary.get("total_quantity", 0)))]
    group_totals = summary.get("group_totals", {})
    if isinstance(group_totals, dict):
        for group, quantity in group_totals.items():
            if safe_int(quantity) > 0:
                metrics.append((_analytics_group_label(group), safe_int(quantity)))
    if show_sku:
        metrics.append(("SKU", safe_int(summary.get("sku_count", 0))))
    for start in range(0, len(metrics), 4):
        chunk = metrics[start:start + 4]
        columns = st.columns(max(1, len(chunk)))
        for column, (label, value) in zip(columns, chunk):
            column.metric(label, value)


def _analytics_stone_rows(family: dict[str, Any]) -> list[dict[str, Any]]:
    stones = family.get("stones", [])
    if not isinstance(stones, list):
        return []
    used_groups: list[str] = []
    for stone in stones:
        totals = stone.get("group_totals", {}) if isinstance(stone, dict) else {}
        if not isinstance(totals, dict):
            continue
        for group in totals:
            if group not in used_groups:
                used_groups.append(group)
    ordered_groups = [group for group in ANALYTICS_GROUP_ORDER if group in used_groups]
    ordered_groups.extend(sorted((group for group in used_groups if group not in ANALYTICS_GROUP_ORDER), key=_analytics_group_label))
    rows: list[dict[str, Any]] = []
    for stone in stones:
        if not isinstance(stone, dict):
            continue
        totals = stone.get("group_totals", {})
        totals = totals if isinstance(totals, dict) else {}
        row: dict[str, Any] = {
            "Камень / вид жемчуга": str(stone.get("name", "")),
            "Всего, шт.": safe_int(stone.get("total_quantity", 0)),
        }
        for group in ordered_groups:
            row[_analytics_group_label(group)] = safe_int(totals.get(group, 0))
        rows.append(row)
    return rows


def _render_analytics_family(family: dict[str, Any]) -> None:
    name = str(family.get("name", "Группа"))
    total = safe_int(family.get("total_quantity", 0))
    st.markdown(f"#### {name} — {total} шт.")
    _render_analytics_metrics(family)
    stone_rows = _analytics_stone_rows(family)
    if len(stone_rows) > 1 or (stone_rows and str(stone_rows[0].get("Камень / вид жемчуга", "")) != name):
        st.dataframe(stone_rows, hide_index=True, width="stretch")


def _render_saved_order_analytics(workspace: SavedOrderWorkspace, mode: str) -> None:
    try:
        with st.spinner("Собираем аналитику сохранённого заказа..."):
            parsed = load_saved_order_workspace(workspace)
            draft = load_draft(workspace.source_hash, workspace.source_name, mode)
            analytics = build_order_analytics(parsed, draft, mode)
    except (OSError, ValueError, BadZipFile, CloudStorageError) as exc:
        st.error(f"Не удалось построить аналитику заказа: {exc}")
        return

    mode_label = "по камням" if mode == ORDER_MODE_STONES else "по жемчугу"
    st.markdown(f"### Информация по заказу {mode_label}")
    st.caption(
        "Все показатели рассчитаны в штуках по сохранённым количествам заказа. "
        "SKU показаны только справочно; позиции Limited Order в сумму не входят."
    )
    _render_analytics_metrics(analytics, show_sku=True)
    limited_positions = safe_int(analytics.get("limited_positions", 0))
    if limited_positions > 0:
        st.info(f"Отдельно в Limited Order: {limited_positions} SKU без количества в основном заказе.")
    if safe_int(analytics.get("total_quantity", 0)) <= 0:
        st.warning("В этом заказе пока нет сохранённых количеств для аналитики.")
        return

    sections = analytics.get("sections", [])
    if not isinstance(sections, list):
        return
    if mode == ORDER_MODE_STONES:
        for section in sections:
            if not isinstance(section, dict):
                continue
            title = f"{section.get('name', 'Раздел')} — {safe_int(section.get('total_quantity', 0))} шт."
            with st.expander(title, expanded=False):
                _render_analytics_metrics(section)
                families = section.get("families", [])
                if isinstance(families, list):
                    for family in families:
                        if isinstance(family, dict):
                            _render_analytics_family(family)
    else:
        section = sections[0] if sections and isinstance(sections[0], dict) else {}
        families = section.get("families", []) if isinstance(section, dict) else []
        if isinstance(families, list):
            for family in families:
                if not isinstance(family, dict):
                    continue
                title = f"{family.get('name', 'Вид жемчуга')} — {safe_int(family.get('total_quantity', 0))} шт."
                with st.expander(title, expanded=False):
                    _render_analytics_metrics(family)
                    stone_rows = _analytics_stone_rows(family)
                    if stone_rows:
                        st.dataframe(stone_rows, hide_index=True, width="stretch")


def _delivery_container_style(container_key: str, delivery_status: str) -> None:
    """Color one delivery card according to its operational status."""
    status = normalize_delivery_status(delivery_status)
    palette = {
        DELIVERY_STATUS_SENT: ("#fff8df", "#d9ad43", "rgba(173, 121, 11, 0.10)"),
        DELIVERY_STATUS_APPROVED: ("#fff8df", "#d9ad43", "rgba(173, 121, 11, 0.10)"),
        DELIVERY_STATUS_IN_PROGRESS: ("#fff8df", "#d9ad43", "rgba(173, 121, 11, 0.10)"),
        DELIVERY_STATUS_RECEIVED: ("#edf8ef", "#82b98a", "rgba(45, 114, 58, 0.10)"),
    }
    background, border, shadow = palette[status]
    st.markdown(
        f"""
        <style>
        .st-key-{container_key} {{
            background: {background};
            border-color: {border} !important;
            box-shadow: 0 8px 24px {shadow};
        }}
        .st-key-{container_key} [data-testid="stMetric"] {{
            background: rgba(255,255,255,0.64);
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _delivery_status_html(delivery_status: str) -> str:
    status = normalize_delivery_status(delivery_status)
    icons = {
        DELIVERY_STATUS_SENT: "●",
        DELIVERY_STATUS_APPROVED: "✓",
        DELIVERY_STATUS_IN_PROGRESS: "◆",
        DELIVERY_STATUS_RECEIVED: "✓",
    }
    return (
        f'<span class="delivery-status delivery-status-{status}">'
        f'{icons[status]} {escape(DELIVERY_STATUS_LABELS[status])}</span>'
    )


def _delivery_status_selectbox(
    *,
    value: str,
    key: str,
    label: str = "Статус заказа",
) -> str:
    current = normalize_delivery_status(value)
    return st.selectbox(
        label,
        options=list(DELIVERY_STATUSES),
        index=list(DELIVERY_STATUSES).index(current),
        format_func=lambda status: DELIVERY_STATUS_LABELS[status],
        key=key,
    )


def _format_manual_order_date(value: str) -> str:
    try:
        return datetime.fromisoformat(str(value)).strftime("%d.%m.%Y")
    except (TypeError, ValueError):
        return str(value or "Дата не указана")


def _date_input_value(value: object, fallback: date | None = None) -> date:
    raw = _date_only(value)
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return fallback or date.today()


def _stage_dates_for_editor(
    *,
    status: str,
    dates: dict[str, str],
    key_prefix: str,
) -> dict[str, str]:
    """Render all dates required by the selected status and return ISO values."""
    normalized_status = normalize_delivery_status(status)
    result = normalize_delivery_dates(dates, status=normalized_status)
    previous_value = _date_input_value(result.get("sent_at", ""))
    for stage in DELIVERY_STATUSES[: delivery_status_rank(normalized_status) + 1]:
        field = DELIVERY_DATE_FIELDS[stage]
        current_value = _date_input_value(result.get(field, ""), previous_value)
        selected = st.date_input(
            f"Дата: {DELIVERY_STATUS_LABELS[stage].lower()}",
            value=current_value,
            key=f"{key_prefix}::{field}",
        )
        result[field] = selected.isoformat()
        previous_value = selected
    return result


def _render_delivery_history(dates: dict[str, str], status: str) -> None:
    history = delivery_history_text(dates, status)
    if history:
        st.caption(history)
    else:
        st.caption("История этапов пока не заполнена.")


def _render_status_change_controls(
    *,
    current_status: str,
    current_dates: dict[str, str],
    key_prefix: str,
    save_callback,
) -> None:
    selected_status = _delivery_status_selectbox(
        value=current_status,
        key=f"{key_prefix}::status",
    )
    normalized_current = normalize_delivery_status(current_status)
    if selected_status != normalized_current:
        st.caption("Укажите даты этапов до выбранного статуса.")
        edited_dates = _stage_dates_for_editor(
            status=selected_status,
            dates=current_dates,
            key_prefix=f"{key_prefix}::change",
        )
        if st.button(
            "Сохранить статус",
            key=f"{key_prefix}::save_status",
            type="primary",
            width="stretch",
        ):
            try:
                validate_delivery_timeline(selected_status, edited_dates)
                save_callback(selected_status, edited_dates)
            except (CloudStorageError, OSError, sqlite3.Error, ValueError) as exc:
                st.error(f"Статус не сохранён: {exc}")
            else:
                st.rerun()

    with st.expander("Изменить даты этапов", expanded=False):
        edited_dates = _stage_dates_for_editor(
            status=normalized_current,
            dates=current_dates,
            key_prefix=f"{key_prefix}::history",
        )
        if st.button(
            "Сохранить даты",
            key=f"{key_prefix}::save_dates",
            width="stretch",
        ):
            try:
                validate_delivery_timeline(normalized_current, edited_dates)
                save_callback(normalized_current, edited_dates)
            except (CloudStorageError, OSError, sqlite3.Error, ValueError) as exc:
                st.error(f"Даты не сохранены: {exc}")
            else:
                st.rerun()


def _render_manual_transit_orders() -> None:
    st.markdown("### Заказы, добавленные вручную")
    with st.expander("＋ Добавить заказ вручную", expanded=False):
        title = st.text_input(
            "Название заказа",
            placeholder="Например: Жемчуг, камни или касты",
            key="manual_transit_title",
        )
        sent_col, qty_col = st.columns([1, 1])
        with sent_col:
            sent_date = st.date_input(
                "Дата отправки",
                value=date.today(),
                key="manual_transit_sent_date",
            )
        with qty_col:
            quantity = st.number_input(
                "Количество изделий",
                min_value=1,
                step=1,
                value=1,
                key="manual_transit_quantity",
            )
        delivery_status = st.selectbox(
            "Текущий статус",
            options=list(DELIVERY_STATUSES),
            index=0,
            format_func=lambda status: DELIVERY_STATUS_LABELS[status],
            key="manual_transit_status",
        )
        delivery_dates = {field: "" for field in DELIVERY_DATE_FIELDS.values()}
        delivery_dates["sent_at"] = sent_date.isoformat()
        previous_date = sent_date
        for stage in DELIVERY_STATUSES[1: delivery_status_rank(delivery_status) + 1]:
            field = DELIVERY_DATE_FIELDS[stage]
            selected_date = st.date_input(
                f"Дата: {DELIVERY_STATUS_LABELS[stage].lower()}",
                value=previous_date,
                key=f"manual_transit_{field}",
            )
            delivery_dates[field] = selected_date.isoformat()
            previous_date = selected_date
        note = st.text_area(
            "Комментарий",
            placeholder="Поставщик, ожидаемый срок или другая полезная информация",
            height=80,
            key="manual_transit_note",
        )
        if st.button(
            "Добавить в список",
            type="primary",
            width="stretch",
            key="manual_transit_add",
        ):
            try:
                if not str(title).strip():
                    raise ValueError("Введите название заказа.")
                if safe_int(quantity) <= 0:
                    raise ValueError("Количество изделий должно быть больше нуля.")
                validate_delivery_timeline(delivery_status, delivery_dates)
                order = ManualTransitOrder(
                    order_id=uuid.uuid4().hex,
                    title=str(title).strip(),
                    order_date=sent_date.isoformat(),
                    note=str(note).strip(),
                    quantity=max(1, safe_int(quantity)),
                    delivery_status=normalize_delivery_status(delivery_status),
                    delivery_dates=delivery_dates,
                )
                save_manual_transit_order(order)
            except (CloudStorageError, OSError, sqlite3.Error, ValueError) as exc:
                st.error(f"Не удалось сохранить заказ: {exc}")
            else:
                for key in list(st.session_state):
                    if str(key).startswith("manual_transit_"):
                        st.session_state.pop(key, None)
                st.success(f"Заказ добавлен со статусом «{DELIVERY_STATUS_LABELS[order.delivery_status]}».")
                st.rerun()

    manual_orders = list_manual_transit_orders()
    if not manual_orders:
        st.caption("Ручных записей пока нет.")
        return

    for order in manual_orders:
        dates = normalize_delivery_dates(
            order.delivery_dates,
            order_date=order.order_date,
            received_at=order.received_at,
            status=order.delivery_status,
            status_updated_at=order.status_updated_at,
        )
        container_key = f"manual_delivery_{order.order_id}"
        _delivery_container_style(container_key, order.delivery_status)
        with st.container(border=True, key=container_key):
            title_col, status_col = st.columns([4, 1])
            with title_col:
                st.markdown(f"**{escape(order.title)}**")
                details = f"{order.quantity} изделий"
                if order.note:
                    details += f" · {escape(order.note)}"
                st.caption(details)
                _render_delivery_history(dates, order.delivery_status)
            with status_col:
                st.markdown(_delivery_status_html(order.delivery_status), unsafe_allow_html=True)

            edit_col, delete_col = st.columns([3, 1])
            with edit_col:
                _render_status_change_controls(
                    current_status=order.delivery_status,
                    current_dates=dates,
                    key_prefix=f"manual_order::{order.order_id}",
                    save_callback=lambda status, edited, current=order: set_manual_transit_order_status(
                        current,
                        status,
                        status_date=edited[DELIVERY_DATE_FIELDS[status]],
                        delivery_dates=edited,
                    ),
                )
            with delete_col:
                if st.button(
                    "Удалить",
                    key=f"manual_order_delete::{order.order_id}",
                    width="stretch",
                ):
                    try:
                        delete_manual_transit_order(order.order_id)
                    except (CloudStorageError, OSError, sqlite3.Error) as exc:
                        st.error(f"Не удалось удалить запись: {exc}")
                    else:
                        st.rerun()


def _render_saved_order_library() -> None:
    st.markdown(
        """
        <style>
        .delivery-status {
            display:inline-flex; align-items:center; justify-content:center;
            padding:0.28rem 0.62rem; border-radius:999px; font-size:0.82rem;
            font-weight:700; white-space:nowrap;
        }
        .delivery-status-sent,
        .delivery-status-approved,
        .delivery-status-in_progress {background:#f3d982; color:#6b4b00;}
        .delivery-status-received {background:#bfe5c5; color:#185c27;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("## Заказы поставщику")

    controls_left, controls_right = st.columns([1, 1])
    with controls_left:
        refresh = st.button(
            "↻ Обновить список",
            key="supplier_order_library_refresh",
            width="stretch",
        )
    with controls_right:
        include_completed = st.toggle(
            "Показать завершённые",
            value=True,
            key="supplier_order_show_completed",
        )
    if refresh:
        for key in list(st.session_state):
            if str(key).startswith(("supplier_order_delivery::", "manual_order::")):
                st.session_state.pop(key, None)

    _render_manual_transit_orders()
    st.divider()
    st.markdown("### Заказы, сформированные в Analitika")

    with st.spinner("Получаем список заказов из облака..."):
        workspaces = list_saved_order_workspaces(
            refresh_cloud=refresh,
            include_completed=include_completed,
        )
    if not workspaces:
        st.info("Сохранённых заказов по выбранному фильтру нет.")
        return

    for index, workspace in enumerate(workspaces):
        confirm_key = f"supplier_order_delete_confirm::{workspace.source_hash}"
        with st.container(border=True):
            details = workspace.mode_details or {
                workspace.preferred_mode: {
                    "selected_positions": workspace.selected_positions,
                    "total_quantity": workspace.total_quantity,
                    "limited_positions": workspace.limited_positions,
                    "stage": "order",
                    "status": workspace.status,
                    "updated_at": workspace.updated_at,
                    "received": False,
                    "received_at": "",
                }
            }
            title_col, delete_col = st.columns([5, 1])
            with title_col:
                status_label = _workspace_status_label(details)
                storage_label = "☁️ Cloudflare R2" if workspace.storage == "cloud" else "локальный кэш"
                st.markdown(f"### {workspace.source_name}")
                st.caption(
                    f"{status_label} · {storage_label} · создан: {_format_saved_order_time(workspace.created_at)} · "
                    f"изменён: {_format_saved_order_time(workspace.updated_at)}"
                )
            with delete_col:
                if st.button(
                    "Удалить",
                    key=f"delete_supplier_order::{workspace.source_hash}::{index}",
                    width="stretch",
                ):
                    st.session_state[confirm_key] = True

            total_a, total_b, total_c = st.columns(3)
            total_a.metric("Всего в блоке, шт.", workspace.total_quantity)
            total_b.metric("SKU в блоке", workspace.selected_positions)
            total_c.metric("Limited Order, SKU", workspace.limited_positions)

            for mode in ORDER_MODES:
                row = details.get(mode)
                mode_exists = isinstance(row, dict)
                if not mode_exists:
                    row = {
                        "selected_positions": 0,
                        "total_quantity": 0,
                        "limited_positions": 0,
                        "stage": "order",
                        "status": "not_started",
                        "delivery_status": DELIVERY_STATUS_SENT,
                        "delivery_dates": {},
                        "received": False,
                        "received_at": "",
                    }
                mode_completed = mode_exists and str(row.get("status", "draft")) == "completed"
                delivery_status = normalize_delivery_status(
                    row.get("delivery_status", ""),
                    received=row.get("received", False),
                ) if mode_completed else DELIVERY_STATUS_SENT
                delivery_dates = normalize_delivery_dates(
                    row.get("delivery_dates", {}),
                    order_date=row.get("created_at", workspace.created_at),
                    received_at=row.get("received_at", ""),
                    status=delivery_status,
                    status_updated_at=row.get("status_updated_at", ""),
                ) if mode_completed else {}
                mode_slug = "stones" if mode == ORDER_MODE_STONES else "pearls"
                delivery_key = f"delivery_{workspace.source_hash[:16]}_{mode_slug}"
                if mode_completed:
                    _delivery_container_style(delivery_key, delivery_status)

                with st.container(border=True, key=delivery_key):
                    analytics_key = f"supplier_order_analytics_open::{workspace.source_hash}::{mode}"
                    info_col, status_col = st.columns([4, 1])
                    with info_col:
                        if not mode_exists:
                            mode_status = "не начат"
                        elif mode_completed:
                            mode_status = "завершён"
                        else:
                            mode_status = "черновик"
                        st.markdown(f"**Заказ: {mode}** · {mode_status}")
                        if mode_exists:
                            st.caption(
                                f"{safe_int(row.get('total_quantity', 0))} шт. · "
                                f"{safe_int(row.get('selected_positions', 0))} SKU · "
                                f"Limited: {safe_int(row.get('limited_positions', 0))} SKU · "
                                f"этап: {_saved_stage_label(str(row.get('stage', 'order')))}"
                            )
                        else:
                            st.caption("Можно начать позже из этого же исходного отчёта.")
                    with status_col:
                        if mode_completed:
                            st.markdown(_delivery_status_html(delivery_status), unsafe_allow_html=True)

                    if mode_completed:
                        _render_delivery_history(delivery_dates, delivery_status)
                        _render_status_change_controls(
                            current_status=delivery_status,
                            current_dates=delivery_dates,
                            key_prefix=f"supplier_order_delivery::{workspace.source_hash}::{mode}",
                            save_callback=lambda status, edited, source=workspace.source_hash, current_mode=mode: set_order_delivery_status(
                                source,
                                current_mode,
                                status,
                                status_date=edited[DELIVERY_DATE_FIELDS[status]],
                                delivery_dates=edited,
                            ),
                        )

                    analytics_col, action_col = st.columns([1.6, 1.2])
                    with analytics_col:
                        analytics_open = bool(st.session_state.get(analytics_key))
                        if st.button(
                            "Скрыть информацию" if analytics_open else "Информация по заказу",
                            key=f"toggle_supplier_order_analytics::{workspace.source_hash}::{mode}::{index}",
                            disabled=not mode_exists,
                            type="secondary",
                            width="stretch",
                        ):
                            st.session_state[analytics_key] = not analytics_open
                            st.rerun()
                    with action_col:
                        action_label = "Начать заказ"
                        if mode_exists:
                            action_label = "Открыть заказ" if mode_completed else "Продолжить заказ"
                        if st.button(
                            action_label,
                            key=f"resume_supplier_order::{workspace.source_hash}::{mode}::{index}",
                            type="primary" if mode_exists else "secondary",
                            width="stretch",
                        ):
                            try:
                                parsed = load_saved_order_workspace(workspace)
                            except (OSError, ValueError, BadZipFile, CloudStorageError) as exc:
                                st.error(f"Не удалось открыть сохранённый заказ: {exc}")
                            else:
                                _clear_order_widget_state()
                                _activate_workspace(parsed, mode)
                                st.session_state["supplier_order_library_open"] = False
                                st.rerun()

                    if mode_exists and bool(st.session_state.get(analytics_key)):
                        with st.container(border=True):
                            _render_saved_order_analytics(workspace, mode)

            if bool(st.session_state.get(confirm_key)):
                st.error(
                    f"Удалить заказ «{workspace.source_name}» полностью? Будут безвозвратно удалены "
                    "исходный Excel, оба черновика, Limited Order, резервные версии и запись облачного индекса."
                )
                confirm_col, cancel_col = st.columns(2)
                with confirm_col:
                    if st.button(
                        "Да, удалить отовсюду",
                        key=f"confirm_delete_supplier_order::{workspace.source_hash}",
                        type="primary",
                        width="stretch",
                    ):
                        try:
                            cloud_count, local_rows, local_files = delete_saved_order_workspace(workspace)
                        except (CloudStorageError, OSError) as exc:
                            st.error(f"Заказ не удалён полностью: {exc}")
                        else:
                            st.session_state["supplier_order_library_notice"] = (
                                f"Заказ «{workspace.source_name}» удалён: облачных объектов — {cloud_count}, "
                                f"локальных черновиков — {local_rows}, локальных файлов — {local_files}."
                            )
                            st.rerun()
                with cancel_col:
                    if st.button(
                        "Отмена",
                        key=f"cancel_delete_supplier_order::{workspace.source_hash}",
                        width="stretch",
                    ):
                        st.session_state.pop(confirm_key, None)
                        st.rerun()

def _render_upload() -> tuple[ParsedOrderWorkbook | None, bytes | None]:
    cloud_ready = _render_storage_status()
    notice = st.session_state.pop("supplier_order_library_notice", None)
    if notice:
        st.success(str(notice))

    active = _load_active_workspace()
    if active is not None:
        top_left, orders_col, change_col = st.columns([4, 1.25, 1])
        with top_left:
            st.markdown(
                '<div class="report-context"><div class="report-context-dot"></div><div class="report-context-copy">'
                f'<strong>Продолжаем сохранённый заказ</strong><span>{escape(active.source_name)}</span>'
                '</div></div>',
                unsafe_allow_html=True,
            )
        with orders_col:
            if st.button("Незавершённые заказы", key="supplier_order_open_library_active", width="stretch"):
                _flush_workspace_session_drafts(active.source_hash)
                _clear_active_workspace()
                st.session_state["supplier_order_library_open"] = True
                st.rerun()
        with change_col:
            if st.button("Новый заказ", key="supplier_order_change_report", width="stretch"):
                _flush_workspace_session_drafts(active.source_hash)
                _clear_active_workspace()
                st.session_state["supplier_order_library_open"] = False
                st.rerun()
        return active, None

    library_open = bool(st.session_state.get("supplier_order_library_open", False))
    library_col, new_col = st.columns(2)
    with library_col:
        if st.button(
            "☁️ Незавершённые заказы",
            key="supplier_order_toggle_library",
            type="primary" if library_open else "secondary",
            width="stretch",
        ):
            st.session_state["supplier_order_library_open"] = not library_open
            st.rerun()
    with new_col:
        if library_open and st.button("＋ Начать новый заказ", key="supplier_order_close_library", width="stretch"):
            st.session_state["supplier_order_library_open"] = False
            st.rerun()

    if library_open:
        _render_saved_order_library()
        st.divider()
        st.markdown("## Начать новый заказ")

    uploaded = st.file_uploader(
        "Загрузите отчёт для формирования заказа",
        type=["xlsx", "xlsm"],
        accept_multiple_files=False,
        key="supplier_order_upload",
    )
    if uploaded is None:
        # Формат выгрузки и вся логика расчёта описаны во вкладке «Как с этим работать».
        return None, None
    payload = bytes(uploaded.getvalue())
    storage_config = load_storage_config()
    if storage_config.required and not cloud_ready:
        st.error("Загрузка нового заказа заблокирована: обязательное облачное хранилище недоступно.")
        return None, None
    try:
        with st.spinner("Сохраняем исходный Excel в надёжное хранилище..."):
            path, digest = store_uploaded_workbook(Path(uploaded.name).name, payload)
    except CloudStorageError as exc:
        st.error(f"Excel не загружен: не удалось создать облачную копию. {exc}")
        return None, None
    try:
        with st.spinner("Читаем комплекты, остатки, ТВП и фотографии..."):
            parsed = cached_parse_order_workbook(str(path), Path(uploaded.name).name, digest)
    except (ValueError, BadZipFile, OSError) as exc:
        diagnostic_event("supplier_order.parse_error", source_name=Path(uploaded.name).name, error=str(exc))
        st.error(f"Отчёт не обработан: {exc}")
        return None, None
    purge_order_workspaces_except(parsed.source_hash)
    _clear_order_widget_state()
    _activate_workspace(parsed)
    st.session_state["supplier_order_library_open"] = False
    return parsed, payload


def _mode_sets(parsed: ParsedOrderWorkbook, mode: str) -> tuple[OrderSet, ...]:
    cache_key = f"supplier_order_sets::{parsed.source_hash}::{mode}"
    cached = st.session_state.get(cache_key)
    if isinstance(cached, tuple):
        return cached
    result = build_order_sets(parsed.items, mode)
    st.session_state[cache_key] = result
    return result


def _ordered_items(order_sets: Iterable[OrderSet]) -> list[OrderItem]:
    return [item for order_set in order_sets for item in order_set.items]


def _seed_defaults(draft: OrderDraft, order_sets: Iterable[OrderSet]) -> None:
    changed = False
    for item in _ordered_items(order_sets):
        if item.key not in draft.orders:
            # Recommendations remain visible as hints, but the actual order
            # always starts from zero and changes only after a user action.
            draft.orders[item.key] = 0
            changed = True
    if changed:
        _save_session_draft(draft)


def _category_reason(order_set: OrderSet) -> str:
    if order_set.is_ungrouped:
        return f"{len(order_set.items)} отдельных позиций. Каждая отнесена к категории по собственным продажам."
    if order_set.category == CATEGORY_ZERO:
        return "Все изделия комплекта имеют 0 продаж."
    return f"Категорию определил артикул {order_set.driver_sku}: продано {order_set.max_sales} шт."


def _order_input_key(item: OrderItem, mode: str, source_hash: str) -> str:
    return "order_qty::" + hashlib.sha1(f"v{DRAFT_VERSION}|{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest()


def _order_action_key(action: str, item: OrderItem, mode: str, source_hash: str) -> str:
    digest = hashlib.sha1(f"{action}|{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest()
    return f"order_action::{action}::{digest}"


def _clear_item_order_state(draft: OrderDraft, item: OrderItem) -> None:
    """Remove an item completely from the ordinary supplier order."""
    draft.orders[item.key] = 0
    draft.sizes.pop(item.key, None)
    draft.stock_checked.pop(item.key, None)
    draft.manual_edit.pop(item.key, None)
    draft.lock_changes.pop(item.key, None)


def _accept_recommendation_action(draft: OrderDraft, item_key: str, quantity: int) -> None:
    """Callback executed before Streamlit redraws only the order fragment."""
    with timed_operation("supplier_order.action", action="accept_recommendation", mode=draft.mode):
        draft.orders[item_key] = max(0, safe_int(quantity))
        draft.manual_edit.pop(item_key, None)
        _save_session_draft(draft)


def _enable_manual_quantity_action(draft: OrderDraft, item_key: str) -> None:
    with timed_operation("supplier_order.action", action="enable_manual_quantity", mode=draft.mode):
        draft.manual_edit[item_key] = True
        _save_session_draft(draft)


def _set_limited_order_action(draft: OrderDraft, item: OrderItem, enabled: bool) -> None:
    with timed_operation("supplier_order.action", action="limited_order", mode=draft.mode, enabled=enabled):
        if enabled:
            draft.limited_orders[item.key] = True
            _clear_item_order_state(draft, item)
        else:
            draft.limited_orders.pop(item.key, None)
        _save_session_draft(draft)


def _show_visual_match_action(draft: OrderDraft, item: OrderItem, mode: str) -> None:
    target_stone = order_stone_bucket(item.stone, item.sku)
    draft.selected_stone = target_stone
    if item.visual_match_category:
        st.session_state[f"supplier_order_category::{mode}::{target_stone}"] = item.visual_match_category
    st.session_state[f"supplier_order_focus_set::{mode}"] = item.visual_match_set_id
    _save_session_draft(draft)


def _remove_item_from_order_action(
    draft: OrderDraft,
    item: OrderItem,
    mode: str,
    source_hash: str,
) -> None:
    with timed_operation("supplier_order.action", action="remove_item", mode=draft.mode):
        _clear_item_order_state(draft, item)
        _queue_item_widget_cleanup(item, mode, source_hash)
        _save_session_draft(draft)


def _render_stock_metric(label: str, value: int, *, always: bool = False, compact_zero: bool = False) -> None:
    """Render stock value without hiding important zero values.

    Overall stock is always visible. TT/63 use a compact badge at zero and a
    warning badge for negative source values.
    """
    if always or value > 0:
        st.metric(label, value)
        return
    if value == 0 and compact_zero:
        st.markdown(
            f"<div style='font-size:.72rem;line-height:1.1;padding:.28rem .45rem;"
            f"border:1px solid rgba(128,128,128,.28);border-radius:.5rem;opacity:.72;"
            f"text-align:center'><span>{label}</span><br><b>0</b></div>",
            unsafe_allow_html=True,
        )
        return
    if value < 0:
        st.markdown(
            f"<div style='font-size:.72rem;line-height:1.1;padding:.28rem .45rem;"
            f"border:1px solid rgba(220,53,69,.55);border-radius:.5rem;"
            f"background:rgba(220,53,69,.08);text-align:center'><span>{label}</span><br>"
            f"<b>{value}</b> ⚠️</div>",
            unsafe_allow_html=True,
        )


def _lock_selector_key(item: OrderItem, mode: str, source_hash: str) -> str:
    digest = hashlib.sha1(f"lock|v{DRAFT_VERSION}|{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest()
    return f"order_lock_change::{digest}"


def _render_lock_change_control(
    item: OrderItem,
    draft: OrderDraft,
    mode: str,
    source_hash: str,
) -> bool:
    """Render a batched earring-lock selector without rerunning on selection."""
    if not item.is_earrings:
        return False

    current_code = earring_lock_code(item.sku)
    saved_code = draft.lock_changes.get(item.key, "")
    if saved_code == current_code or saved_code not in EARRING_LOCKS:
        saved_code = ""
        draft.lock_changes.pop(item.key, None)

    changed = False
    with st.popover("Заказать другой замок", width="stretch"):
        if current_code:
            english, russian = EARRING_LOCKS[current_code]
            st.success(f"Текущий: **{current_code} — {russian}**")
            st.caption(english)
        else:
            st.warning("Текущий замок по артикулу не определён.")

        choices = [""] + [code for code in EARRING_LOCKS if code != current_code]
        default_index = choices.index(saved_code) if saved_code in choices else 0
        with st.form(
            key="lock_form::" + hashlib.sha1(f"{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest(),
            border=False,
        ):
            selected = st.selectbox(
                "Выберите замок для заказа",
                choices,
                index=default_index,
                format_func=lambda code: (
                    "Не менять замок" if not code else f"{code} — {EARRING_LOCKS[code][1]}"
                ),
            )
            submitted = st.form_submit_button("Применить замок", width="stretch")
        if submitted and selected != saved_code:
            with timed_operation("supplier_order.action", action="change_lock", mode=draft.mode):
                if selected:
                    draft.lock_changes[item.key] = selected
                else:
                    draft.lock_changes.pop(item.key, None)
                _save_session_draft(draft)
            changed = True
            saved_code = selected
        if saved_code:
            st.info(f"В Excel: **{earring_lock_export_label(saved_code)}**", icon="🔧")
    return changed


def _render_item_row(
    item: OrderItem,
    order_set: OrderSet,
    image_data: bytes | None,
    draft: OrderDraft,
    mode: str,
    source_hash: str,
) -> bool:
    changed = False
    limited = bool(draft.limited_orders.get(item.key, False))
    recommendation = build_order_recommendation(
        item,
        order_set,
        mode,
        draft.recommendation_profile,
    )

    with st.container(border=True):
        photo, details, sales_col, stock_col, action_col = st.columns(
            [0.86, 1.68, 0.56, 1.42, 1.72],
            vertical_alignment="center",
        )
        with photo:
            if image_data:
                st.image(image_data, width="stretch")
            else:
                st.caption("Нет фото")
        with details:
            st.markdown(f"**{item.sku}**")
            st.caption(f"{canonical_stone(item.stone, item.sku)} · {item.group}")
            if item.is_earrings and sum(other.is_earrings for other in order_set.items) > 1:
                st.caption("Не единственные серьги в комплекте")
            if item.duplicate_status:
                duplicate_message = item.duplicate_reason or (
                    f"Есть очень похожая модель {item.duplicate_sku or ''} — возможно, дубль."
                )
                if item.duplicate_status == "suppress":
                    st.error(duplicate_message, icon="🔁")
                elif item.duplicate_status == "preferred":
                    st.success(duplicate_message, icon="🔁")
                else:
                    st.warning(duplicate_message, icon="🔁")
                if item.duplicate_score > 0:
                    st.caption(
                        f"Похожая модель: {item.duplicate_sku or 'не указана'} · "
                        f"сходство {item.duplicate_score:.0%}"
                    )
            if item.visual_match_set_id:
                match_title = "Найдено визуальное совпадение" if item.visual_match_status == "confirmed" else "Возможное визуальное совпадение"
                message = (
                    f"{match_title}: **{item.visual_match_set_id}** · "
                    f"{item.visual_match_sku or 'артикул не указан'} · "
                    f"сходство {item.visual_match_score:.0%}"
                )
                if item.visual_match_status == "confirmed":
                    st.success(message, icon="🔎")
                else:
                    st.warning(message, icon="🔎")
                st.button(
                    "Показать найденный комплект",
                    key="show_match::" + hashlib.sha1(f"{mode}|{item.key}".encode("utf-8")).hexdigest(),
                    width="stretch",
                    on_click=_show_visual_match_action,
                    args=(draft, item, mode),
                )
            for error in item.errors:
                st.error(error, icon="⚠️")
            if limited:
                st.warning("Limited Order: позиция исключена из обычного заказа.", icon="🔒")
            elif item.positive_tvp > 0:
                st.info(f"В пути: {item.positive_tvp} шт. Повторную автоматическую рекомендацию на этот SKU не даём.", icon="🚚")

        with sales_col:
            st.metric("Продажи", item.sales)
        with stock_col:
            stock_main, stock_tt, stock_63 = st.columns([1.15, 0.72, 0.72], vertical_alignment="center")
            with stock_main:
                _render_stock_metric("Общий остаток", item.display_stock, always=True)
            with stock_tt:
                _render_stock_metric("TT", item.stock_tt, compact_zero=True)
            with stock_63:
                _render_stock_metric("63", item.stock_63, compact_zero=True)

            if limited:
                st.warning("Limited Order", icon="🔒")
                st.button(
                    "Вернуть в обычный заказ",
                    key=_order_action_key("unlimited", item, mode, source_hash),
                    width="stretch",
                    on_click=_set_limited_order_action,
                    args=(draft, item, False),
                )
            else:
                control_columns = st.columns(2) if item.is_earrings else [st.container()]
                with control_columns[0]:
                    st.button(
                        "Limited Order",
                        key=_order_action_key("limited", item, mode, source_hash),
                        width="stretch",
                        on_click=_set_limited_order_action,
                        args=(draft, item, True),
                    )
                if item.is_earrings:
                    with control_columns[1]:
                        if _render_lock_change_control(item, draft, mode, source_hash):
                            changed = True

        with action_col:
            current = max(0, safe_int(draft.orders.get(item.key, 0)))
            manual_enabled = bool(draft.manual_edit.get(item.key, False))

            if limited:
                st.caption("Позиция исключена из обычного заказа.")
                return changed

            if recommendation.quantity > 0:
                reasons_html = "<br>".join(recommendation.reasons)
                st.markdown(
                    f"<div style='padding:.65rem .8rem;border-radius:.75rem;"
                    f"background:rgba(196,145,2,.10);border:1px solid rgba(196,145,2,.32);"
                    f"text-align:center'>"
                    f"<div style='font-size:.78rem;font-weight:600'>Рекомендуем заказать</div>"
                    f"<div style='font-size:2.2rem;line-height:1;font-weight:800;margin:.18rem 0'>{recommendation.quantity}</div>"
                    f"<div style='font-size:.72rem;opacity:.78'>{reasons_html}</div></div>",
                    unsafe_allow_html=True,
                )
            elif recommendation.blocked_by_tvp:
                st.caption("Можно только дозаказать вручную.")
            elif recommendation.rule != "none" and recommendation.reasons:
                st.caption(" ".join(recommendation.reasons))
            else:
                st.caption("Автоматический заказ не требуется.")

            for transfer in recommendation.transfers:
                st.info(transfer, icon="↔️")

            if manual_enabled:
                with st.form(
                    key="quantity_form::" + hashlib.sha1(f"{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest(),
                    border=False,
                ):
                    value = st.number_input(
                        "К заказу",
                        min_value=0,
                        max_value=999,
                        step=1,
                        value=current,
                    )
                    quantity_submitted = st.form_submit_button("Применить количество", width="stretch")
                value = max(0, safe_int(value))
                if quantity_submitted and value != current:
                    with timed_operation("supplier_order.action", action="manual_quantity", mode=draft.mode):
                        draft.orders[item.key] = value
                        _save_session_draft(draft)
                    current = value
                    changed = True
            else:
                st.metric("К заказу", current)

            if recommendation.quantity > 0 and not recommendation.blocked_by_tvp:
                st.button(
                    "Согласен с рекомендацией",
                    key=_order_action_key("accept", item, mode, source_hash),
                    type="primary",
                    width="stretch",
                    on_click=_accept_recommendation_action,
                    args=(draft, item.key, recommendation.quantity),
                )
                edit_label = "Изменить количество"
            elif recommendation.blocked_by_tvp:
                edit_label = "Дозаказать вручную"
            else:
                edit_label = "Добавить вручную"

            st.button(
                edit_label,
                key=_order_action_key("manual", item, mode, source_hash),
                width="stretch",
                on_click=_enable_manual_quantity_action,
                args=(draft, item.key),
            )

    return changed

def _render_set_card(
    order_set: OrderSet,
    images: dict[str, bytes],
    draft: OrderDraft,
    mode: str,
    source_hash: str,
) -> bool:
    changed = False
    icon = CATEGORY_TONE[order_set.category]
    focused = st.session_state.get(f"supplier_order_focus_set::{mode}") == order_set.set_id
    with st.container(border=True):
        if focused:
            st.info("Найденный визуально похожий комплект", icon="🔎")
        header_left, header_right = st.columns([4, 1])
        with header_left:
            st.markdown(f"### {order_set.set_id}")
            st.caption(_category_reason(order_set))
        with header_right:
            st.markdown(f"**{icon} {order_set.category}**")
            if order_set.has_negative_tvp:
                st.error("Ошибка ТВП", icon="⚠️")
        for item in order_set.items:
            changed = _render_item_row(item, order_set, images.get(item.image_path or ""), draft, mode, source_hash) or changed
    return changed


def _render_sets_group(sets: list[OrderSet], parsed: ParsedOrderWorkbook, draft: OrderDraft, mode: str, prefix: str) -> bool:
    if not sets:
        st.caption("Комплектов в этом сегменте нет.")
        return False
    ordered_sets = sorted(sets, key=lambda order_set: (order_set.is_ungrouped, order_set.items[0].row if order_set.items else 0))
    image_paths = tuple(sorted({item.image_path for order_set in ordered_sets for item in order_set.items if item.image_path}))
    images = load_visible_images(parsed.upload_path, image_paths)
    changed = False
    for index, order_set in enumerate(ordered_sets):
        st.markdown(f'<div id="{prefix}-{index}"></div>', unsafe_allow_html=True)
        changed = _render_set_card(order_set, images, draft, mode, parsed.source_hash) or changed
    return changed


def _render_category(category: str, sets: list[OrderSet], parsed: ParsedOrderWorkbook, draft: OrderDraft, mode: str) -> bool:
    changed = False
    if category == CATEGORY_ZERO:
        for segment in ("Нулевые с остатком", "0/0 — не было остатка"):
            segment_sets = [order_set for order_set in sets if order_set.zero_segment == segment]
            st.markdown(f"#### {segment}")
            changed = _render_category_segment(segment_sets, parsed, draft, mode, f"zero-{segment}") or changed
        return changed
    return _render_category_segment(sets, parsed, draft, mode, category)


def _render_category_segment(sets: list[OrderSet], parsed: ParsedOrderWorkbook, draft: OrderDraft, mode: str, prefix: str) -> bool:
    # TVP is controlled only by the global toggle. There is no secondary
    # expander, and relevant sets are rendered immediately with full context.
    return _render_sets_group(sets, parsed, draft, mode, "sets-" + re.sub(r"\W+", "-", prefix))


def _render_overview(parsed: ParsedOrderWorkbook, order_sets: tuple[OrderSet, ...], mode: str, draft: OrderDraft) -> None:
    st.markdown('<div id="order-overview"></div>', unsafe_allow_html=True)
    st.markdown("## Сводка заказа")
    items = _ordered_items(order_sets)
    excluded_count = len(parsed.items) - len(items)
    errors = sum(bool(item.errors) for item in items)
    positive_tvp_positions = sum(item.tvp_raw > 0 for item in items)
    positive_tvp_quantity = sum(item.positive_tvp for item in items)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Тип заказа", mode)
    c2.metric("Комплектов", len(order_sets))
    c3.metric("Изделий", len(items))
    c4.metric("Товаров в пути", positive_tvp_quantity, help=f"{positive_tvp_positions} позиций с ТВП больше нуля")
    c5.metric("Ошибки", errors)
    st.caption(f"Исключено или относится к другому типу заказа: {excluded_count} строк.")
    if parsed.period:
        st.caption(
            f"Период в подписи отчёта: {parsed.period} · утверждённый расчёт: 4 месяца · "
            f"горизонт заказа: 2 месяца"
        )
    st.caption(f"Режим рекомендаций: {draft.recommendation_profile}")
    if parsed.supplier:
        st.caption(f"Поставщик из отчёта: {parsed.supplier}")
    for warning in parsed.warnings:
        st.warning(warning)


def filter_order_sets_by_tvp(order_sets: Iterable[OrderSet], positive_only: bool) -> tuple[OrderSet, ...]:
    """Filter supplier sets by TVP while preserving the full set context.

    The toggle decides which *sets* are relevant: enabled keeps sets containing
    at least one row with TVP > 0; disabled keeps sets containing at least one
    row with TVP <= 0. Once a set is selected, all items of the same stone and
    set remain visible so earrings, rings and pendants are reviewed together.
    """
    result: list[OrderSet] = []
    for order_set in order_sets:
        matches = any(
            (item.tvp_raw > 0) if positive_only else (item.tvp_raw <= 0)
            for item in order_set.items
        )
        if matches:
            result.append(order_set)
    return tuple(result)


def order_quantity_summary(order_sets: Iterable[OrderSet], draft: OrderDraft) -> dict[str, int]:
    """Summarize ordered *units* and SKU positions by merchandise group.

    Limited Order rows are intentionally excluded because they do not enter
    the supplier Excel.  The function is pure so the UI and regression tests
    use the same definition of earrings/rings/pendants totals.
    """
    items = [
        item
        for item in _ordered_items(order_sets)
        if not draft.limited_orders.get(item.key, False)
        and max(0, safe_int(draft.orders.get(item.key, 0))) > 0
    ]
    result = {
        "earrings_qty": 0,
        "rings_qty": 0,
        "pendants_qty": 0,
        "other_qty": 0,
        "total_qty": 0,
        "sku_count": len(items),
    }
    for item in items:
        quantity = max(0, safe_int(draft.orders.get(item.key, 0)))
        group = canonical_group(item.group)
        if group == "Earrings":
            result["earrings_qty"] += quantity
        elif group == "Ring":
            result["rings_qty"] += quantity
        elif group == "Pendant":
            result["pendants_qty"] += quantity
        else:
            result["other_qty"] += quantity
        result["total_qty"] += quantity
    return result


def _render_quantity_summary(title: str, summary: dict[str, int]) -> None:
    """Render an unambiguous order summary where all group values are units."""
    st.markdown(f"**{title}**")
    columns = st.columns(5)
    columns[0].metric("Серьги, шт.", summary["earrings_qty"])
    columns[1].metric("Кольца, шт.", summary["rings_qty"])
    columns[2].metric("Подвески, шт.", summary["pendants_qty"])
    columns[3].metric("Всего, шт.", summary["total_qty"])
    columns[4].metric("SKU", summary["sku_count"])
    if summary["other_qty"] > 0:
        st.caption(f"Другие группы: {summary['other_qty']} шт.")


def _render_order_workspace(parsed: ParsedOrderWorkbook, order_sets: tuple[OrderSet, ...], draft: OrderDraft, mode: str) -> None:
    st.markdown('<div id="order-workspace"></div>', unsafe_allow_html=True)
    pearl_mode = mode == ORDER_MODE_PEARLS
    st.markdown("## Комплекты по жемчугу" if pearl_mode else "## Комплекты по камням")
    st.caption("Положительный ТВП учитывается автоматически: рекомендация блокируется, но ручной дозаказ остаётся доступен.")

    stones = list(order_navigation_options(mode))
    present_buckets = {
        bucket
        for order_set in order_sets
        if (bucket := order_set_navigation_bucket(order_set, mode)) is not None
    }
    if not present_buckets:
        st.warning("После исключений в выбранном типе заказа не осталось комплектов.")
        return
    first_present = next((bucket for bucket in stones if bucket in present_buckets), stones[0])
    default_bucket = draft.selected_stone if draft.selected_stone in stones else first_present
    default_index = stones.index(default_bucket)
    selected_stone = st.selectbox(
        "Тип жемчуга" if pearl_mode else "Группа камня",
        stones,
        index=default_index,
        key=f"supplier_order_stone::{mode}",
    )
    if selected_stone != draft.selected_stone:
        draft.selected_stone = selected_stone
        _save_session_draft(draft)

    stone_sets = [
        order_set for order_set in order_sets
        if order_set_navigation_bucket(order_set, mode) == selected_stone
    ]
    counts = {category: sum(order_set.category == category for order_set in stone_sets) for category in CATEGORY_ORDER}
    cols = st.columns(4)
    for column, category in zip(cols, CATEGORY_ORDER):
        column.metric(CATEGORY_SHORT[category], counts[category])

    selected_category = st.segmented_control(
        "Категория комплектов",
        list(CATEGORY_ORDER),
        default=CATEGORY_TOP,
        key=f"supplier_order_category::{mode}::{selected_stone}",
    ) or CATEGORY_TOP
    st.caption(
        f"{CATEGORY_TONE[selected_category]} {selected_category}: "
        f"{counts[selected_category]} комплектов. Продажи изделий внутри комплекта не суммируются."
    )
    category_sets = [order_set for order_set in stone_sets if order_set.category == selected_category]
    page_count = max(1, math.ceil(len(category_sets) / ORDER_PAGE_SIZE))
    page_digest = hashlib.sha1(
        f"{parsed.source_hash}|{mode}|{selected_stone}|{selected_category}".encode("utf-8")
    ).hexdigest()
    page_key = f"supplier_order_page::{page_digest}"
    focus_key = f"supplier_order_focus_page::{page_digest}"
    focused_set_id = str(st.session_state.get(f"supplier_order_focus_set::{mode}", ""))
    target_page = 1
    if focused_set_id:
        for index, order_set in enumerate(category_sets):
            if order_set.set_id == focused_set_id:
                target_page = index // ORDER_PAGE_SIZE + 1
                break
    if page_key not in st.session_state or st.session_state.get(focus_key) != focused_set_id:
        st.session_state[page_key] = target_page
        st.session_state[focus_key] = focused_set_id
    current_page = max(1, min(page_count, safe_int(st.session_state.get(page_key, 1)) or 1))
    st.session_state[page_key] = current_page
    if page_count > 1:
        current_page = st.selectbox(
            "Страница комплектов",
            list(range(1, page_count + 1)),
            key=page_key,
            format_func=lambda value: f"{value} из {page_count}",
        )
    start = (current_page - 1) * ORDER_PAGE_SIZE
    visible_sets = category_sets[start : start + ORDER_PAGE_SIZE]
    if category_sets:
        st.caption(
            f"Показаны комплекты {start + 1}–{start + len(visible_sets)} из {len(category_sets)}. "
            "На странице не больше 10 комплектов — фотографии и карточки обновляются быстрее."
        )
    changed = _render_category(selected_category, visible_sets, parsed, draft, mode)
    if changed:
        st.toast("Изменения автоматически сохранены", icon="💾")

    st.markdown("---")
    current_summary = order_quantity_summary(category_sets, draft)
    _render_quantity_summary(
        f"Выбранный раздел: {selected_stone} · {selected_category}",
        current_summary,
    )
    st.markdown("#### Итог по всему заказу")
    overall_summary = order_quantity_summary(order_sets, draft)
    _render_quantity_summary(f"Весь заказ · {mode}", overall_summary)

    active_items = [item for item in _ordered_items(order_sets) if not draft.limited_orders.get(item.key, False)]
    total_ordered = sum(max(0, draft.orders.get(item.key, 0)) for item in active_items)
    ordered_positions = sum(draft.orders.get(item.key, 0) > 0 for item in active_items)
    limited_positions = sum(bool(draft.limited_orders.get(item.key, False)) for item in _ordered_items(order_sets))
    st.markdown("---")
    left, middle, limited_col, right = st.columns([1.2, 1, 1, 1.6])
    left.metric("Заказано SKU", ordered_positions)
    middle.metric("Всего изделий, шт.", total_ordered)
    limited_col.metric("Limited Order, SKU", limited_positions)
    if right.button("Подтвердить количества и перейти к размерам", type="primary", width="stretch", disabled=ordered_positions == 0):
        draft.stage = "rings"
        _flush_session_draft(draft)
        st.rerun()

def _size_input_key(item: OrderItem, size: int, mode: str, source_hash: str) -> str:
    return "ring_size::" + hashlib.sha1(f"{source_hash}|{mode}|{item.key}|{size}".encode("utf-8")).hexdigest()


def _stock_check_key(item: OrderItem, mode: str, source_hash: str) -> str:
    return "ring_stock_check::" + hashlib.sha1(f"{source_hash}|{mode}|{item.key}".encode("utf-8")).hexdigest()


PENDING_ORDER_WIDGET_CLEANUP_KEY = "supplier_order_pending_widget_cleanup"


def _queue_item_widget_cleanup(item: OrderItem, mode: str, source_hash: str) -> None:
    """Clear stale quantity and ring-size widgets safely on the next rerun.

    Streamlit does not allow changing a widget's state after that widget has
    already been rendered in the current run.  The remove button is shown
    below the size widgets, so cleanup is queued and applied before any order
    widgets are rendered on the next run.
    """
    pending = {
        str(key)
        for key in st.session_state.get(PENDING_ORDER_WIDGET_CLEANUP_KEY, [])
        if str(key)
    }
    pending.add(_order_input_key(item, mode, source_hash))
    pending.add(_stock_check_key(item, mode, source_hash))
    pending.update(_size_input_key(item, size, mode, source_hash) for size in RING_SIZES)
    st.session_state[PENDING_ORDER_WIDGET_CLEANUP_KEY] = sorted(pending)


def _apply_pending_order_widget_cleanup() -> None:
    pending = st.session_state.pop(PENDING_ORDER_WIDGET_CLEANUP_KEY, [])
    if not isinstance(pending, (list, tuple, set)):
        return
    for key in pending:
        st.session_state.pop(str(key), None)


def normalize_copy_sku(value: object) -> str:
    """Return the exact SKU copied to the clipboard without edge whitespace."""
    return str(value or "").strip()


def copy_sku_html(value: object) -> str:
    """Build the isolated clipboard control used next to SKU text."""
    sku = normalize_copy_sku(value)
    safe_text = escape(sku)
    js_value = json.dumps(sku, ensure_ascii=False)
    return f"""
<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<style>
  html, body {{ margin: 0; padding: 0; background: transparent; overflow: hidden; }}
  .sku-copy-row {{
    min-height: 34px; display: flex; align-items: center; gap: 8px;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  }}
  .sku-copy-value {{
    color: #171411; font-size: 16px; line-height: 1.25; font-weight: 700;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }}
  .sku-copy-button {{
    flex: 0 0 auto; width: 30px; height: 30px; border-radius: 9px;
    border: 1px solid #d7b56d; background: #fffaf0; color: #815716;
    font-size: 16px; line-height: 1; cursor: pointer; padding: 0;
    display: inline-flex; align-items: center; justify-content: center;
  }}
  .sku-copy-button:hover {{ background: #f5e5bf; }}
  .sku-copy-button:focus-visible {{ outline: 2px solid #c68a27; outline-offset: 2px; }}
  .sku-copy-button.copied {{ background: #e7f5e9; border-color: #7fb288; color: #286335; }}
</style>
</head>
<body>
  <div class="sku-copy-row">
    <span class="sku-copy-value" title="{safe_text}">{safe_text}</span>
    <button id="copy-sku" class="sku-copy-button" type="button"
      title="Копировать артикул" aria-label="Копировать артикул {safe_text}">⧉</button>
  </div>
<script>
  const sku = {js_value};
  const button = document.getElementById('copy-sku');

  function fallbackCopy(text) {{
    const area = document.createElement('textarea');
    area.value = text;
    area.setAttribute('readonly', '');
    area.style.position = 'fixed';
    area.style.opacity = '0';
    document.body.appendChild(area);
    area.select();
    const ok = document.execCommand('copy');
    document.body.removeChild(area);
    if (!ok) throw new Error('copy failed');
  }}

  async function copySku() {{
    try {{
      if (navigator.clipboard && window.isSecureContext) {{
        await navigator.clipboard.writeText(sku);
      }} else {{
        fallbackCopy(sku);
      }}
      button.textContent = '✓';
      button.classList.add('copied');
      button.title = 'Артикул скопирован';
      setTimeout(() => {{
        button.textContent = '⧉';
        button.classList.remove('copied');
        button.title = 'Копировать артикул';
      }}, 1400);
    }} catch (error) {{
      button.textContent = '!';
      button.title = 'Не удалось скопировать';
      setTimeout(() => {{ button.textContent = '⧉'; button.title = 'Копировать артикул'; }}, 1600);
    }}
  }}

  button.addEventListener('click', copySku);
</script>
</body>
</html>
"""


def _render_copyable_sku(value: object) -> None:
    from streamlit.components.v1 import html as components_html

    components_html(copy_sku_html(value), height=38, scrolling=False)


def ring_validation(item: OrderItem, draft: OrderDraft) -> tuple[int, int, bool, bool]:
    quantity = max(0, draft.orders.get(item.key, 0))
    values = draft.sizes.get(item.key, {})
    allocated = sum(max(0, safe_int(values.get(str(size), 0))) for size in RING_SIZES)
    # Остаток показывается пользователю как информация и не блокирует выгрузку.
    # Четвёртое значение сохранено для совместимости с текущими вызовами.
    return quantity, allocated, allocated == quantity, True


def _render_ring_sizes(parsed: ParsedOrderWorkbook, order_sets: tuple[OrderSet, ...], draft: OrderDraft, mode: str) -> None:
    st.markdown('<div id="order-rings"></div>', unsafe_allow_html=True)
    st.markdown("## Размеры колец")
    st.caption("Заполните все размеры кольца и нажмите «Применить размеры». Изменение каждого поля отдельно больше не перезапускает сайт.")
    ordered_rings = [
        item for item in _ordered_items(order_sets)
        if item.is_ring and draft.orders.get(item.key, 0) > 0 and not draft.limited_orders.get(item.key, False)
    ]
    if not ordered_rings:
        st.info("В текущем заказе нет колец с положительным количеством.")
        return
    image_paths = tuple(sorted({item.image_path for item in ordered_rings if item.image_path}))
    images = load_visible_images(parsed.upload_path, image_paths)
    complete = 0

    for item in ordered_rings:
        quantity = max(0, draft.orders.get(item.key, 0))
        values = draft.sizes.setdefault(item.key, {})
        with st.container(border=True):
            left, right = st.columns([1, 4])
            with left:
                if item.image_path and item.image_path in images:
                    st.image(images[item.image_path], width="stretch")
                _render_copyable_sku(item.sku)
                st.caption(f"{canonical_stone(item.stone, item.sku)} · к заказу {quantity}")
                if item.working_stock > 0:
                    st.info(f"По этой позиции есть остаток: {item.working_stock} шт.", icon="ℹ️")
            with right:
                form_key = "ring_sizes_form::" + hashlib.sha1(
                    f"{parsed.source_hash}|{mode}|{item.key}".encode("utf-8")
                ).hexdigest()
                with st.form(form_key, border=False):
                    columns = st.columns(5)
                    entered_values: dict[str, int] = {}
                    for index, size in enumerate(RING_SIZES):
                        size_key = str(size)
                        current = max(0, safe_int(values.get(size_key, 0)))
                        with columns[index % 5]:
                            entered_values[size_key] = max(0, safe_int(st.number_input(
                                str(size), min_value=0, max_value=quantity, step=1, value=current
                            )))
                    apply_sizes = st.form_submit_button("Применить размеры", type="primary", width="stretch")
                if apply_sizes:
                    entered_total = sum(entered_values.values())
                    if entered_total > quantity:
                        st.error(f"Указано {entered_total}, но к заказу доступно только {quantity}.")
                    else:
                        with timed_operation("supplier_order.action", action="apply_ring_sizes", mode=draft.mode):
                            draft.sizes[item.key] = {key: value for key, value in entered_values.items() if value > 0}
                            _save_session_draft(draft)
                        st.toast("Размеры сохранены локально", icon="💾")
                        values = draft.sizes[item.key]

                requested, allocated, allocation_ok, _ = ring_validation(item, draft)
                if allocated > requested:
                    st.error(f"Распределено {allocated}, но к заказу доступно только {requested}.")
                elif allocated < requested:
                    st.warning(f"Распределено {allocated} из {requested} · осталось {requested - allocated}")
                else:
                    st.success(f"Распределено {allocated} из {requested}")

                st.button(
                    "Удалить из заказа",
                    key=_order_action_key("remove_from_order", item, mode, parsed.source_hash),
                    width="stretch",
                    on_click=_remove_item_from_order_action,
                    args=(draft, item, mode, parsed.source_hash),
                )
                if allocation_ok:
                    complete += 1
    st.caption(f"Размеры заполнены: {complete} из {len(ordered_rings)}")


def _export_readiness(order_sets: tuple[OrderSet, ...], draft: OrderDraft) -> tuple[bool, list[str]]:
    ordered_items = [item for item in _ordered_items(order_sets) if draft.orders.get(item.key, 0) > 0 and not draft.limited_orders.get(item.key, False)]
    reasons: list[str] = []
    if not ordered_items:
        reasons.append("В заказе нет изделий.")
    rings = [item for item in ordered_items if item.is_ring]
    incomplete = []
    for item in rings:
        _, _, allocation_ok, _ = ring_validation(item, draft)
        if not allocation_ok:
            incomplete.append(item.sku)
    if incomplete:
        reasons.append(f"Не завершены размеры для {len(incomplete)} колец.")
    return not reasons, reasons


def _draft_export_signature(draft: OrderDraft, *, limited: bool = False) -> str:
    payload = {
        "source_hash": draft.source_hash,
        "mode": draft.mode,
        "orders": draft.orders,
        "sizes": draft.sizes,
        "limited_orders": draft.limited_orders,
        "lock_changes": draft.lock_changes,
        "limited": limited,
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()

def _render_export(parsed: ParsedOrderWorkbook, order_sets: tuple[OrderSet, ...], draft: OrderDraft, mode: str) -> None:
    st.markdown('<div id="order-export"></div>', unsafe_allow_html=True)
    st.markdown("## Итоговый Excel")
    all_items = _ordered_items(order_sets)
    ordered_items = [
        item for item in all_items
        if draft.orders.get(item.key, 0) > 0 and not draft.limited_orders.get(item.key, False)
    ]
    limited_items = [item for item in all_items if draft.limited_orders.get(item.key, False)]
    ready, reasons = _export_readiness(order_sets, draft)
    total_quantity = sum(draft.orders.get(item.key, 0) for item in ordered_items)
    rings = [item for item in ordered_items if item.is_ring]
    ring_quantity = sum(max(0, safe_int(draft.orders.get(item.key, 0))) for item in rings)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("SKU", len(ordered_items))
    c2.metric("Изделий, шт.", total_quantity)
    c3.metric("Колец, шт.", ring_quantity)
    c4.metric("Limited Order, SKU", len(limited_items))

    if reasons:
        for reason in reasons:
            st.warning(reason)
        st.button("Скачать заказ в Excel", disabled=True, width="stretch")
    else:
        signature = _draft_export_signature(draft)
        payload_key = f"supplier_excel::{parsed.source_hash}::{mode}::{signature}"
        if st.button("Скачать заказ в Excel — подготовить файл", type="primary", width="stretch"):
            _flush_session_draft(draft)
            with st.spinner("Формируем Excel с фотографиями..."):
                with timed_operation("supplier_order.build_excel", mode=mode, kind="main"):
                    payload = build_supplier_excel(parsed, ordered_items, draft)
                _clear_generated_payloads(parsed.source_hash, mode, keep_keys=(payload_key,), kinds=("main",))
                st.session_state[payload_key] = payload
        payload = st.session_state.get(payload_key)
        if isinstance(payload, bytes):
            safe_mode = "stones" if mode == ORDER_MODE_STONES else "pearls"
            st.download_button(
                "Скачать подготовленный Excel",
                data=payload,
                file_name=f"supplier_order_{safe_mode}_{datetime.now().date().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                width="stretch",
            )
        else:
            st.caption("Excel строится только по кнопке и больше не пересобирается после каждого изменения заказа.")
        st.caption(
            "В основном файле: фото, SKU, камень, количество, размеры колец и выбранная замена замка. "
            "Заголовки Excel — на английском."
        )

    if limited_items:
        st.markdown("### Limited Order")
        st.caption("Отдельный внутренний список: рекомендации и обычные количества для этих изделий отключены.")
        limited_signature = _draft_export_signature(draft, limited=True)
        limited_key = f"limited_excel::{parsed.source_hash}::{mode}::{limited_signature}"
        if st.button("Подготовить Limited Order Excel", width="stretch"):
            _flush_session_draft(draft)
            with st.spinner("Формируем Limited Order Excel..."):
                with timed_operation("supplier_order.build_excel", mode=mode, kind="limited"):
                    payload = build_limited_order_excel(parsed, limited_items, draft)
                _clear_generated_payloads(parsed.source_hash, mode, keep_keys=(limited_key,), kinds=("limited",))
                st.session_state[limited_key] = payload
        limited_payload = st.session_state.get(limited_key)
        if isinstance(limited_payload, bytes):
            safe_mode = "stones" if mode == ORDER_MODE_STONES else "pearls"
            st.download_button(
                "Скачать Limited Order Excel",
                data=limited_payload,
                file_name=f"limited_order_{safe_mode}_{datetime.now().date().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

    st.divider()
    if draft.status == "completed":
        status_col, action_col = st.columns([4, 1])
        with status_col:
            mode_label = "по камням" if mode == ORDER_MODE_STONES else "по жемчугу"
            st.success(f"Заказ {mode_label} отмечен как завершённый. Второй тип заказа завершается независимо.")
        with action_col:
            if st.button("Вернуть в черновики", key=f"reopen_order::{parsed.source_hash}::{mode}", width="stretch"):
                draft.status = "draft"
                _flush_session_draft(draft)
                st.rerun()
    else:
        can_complete = bool(limited_items) or ready
        complete_label = "Завершить заказ по камням" if mode == ORDER_MODE_STONES else "Завершить заказ по жемчугу"
        if st.button(
            complete_label,
            key=f"complete_order::{parsed.source_hash}::{mode}",
            disabled=not can_complete,
            width="stretch",
        ):
            draft.status = "completed"
            _flush_session_draft(draft)
            st.rerun()
        if not can_complete:
            st.caption(
                "Этот тип заказа можно завершить после заполнения обязательных количеств и размеров "
                "или при наличии Limited Order. Заказ второго типа не влияет на эту кнопку."
            )


def _render_draft_tools(parsed: ParsedOrderWorkbook, draft: OrderDraft, mode: str) -> None:
    with st.expander("Резервная копия заказа", expanded=False):
        st.caption(
            "Основная защита — облачное автосохранение. Для аварийной копии можно скачать один ZIP; "
            "ручной импорт JSON в интерфейсе не используется."
        )
        safe_mode = "stones" if mode == ORDER_MODE_STONES else "pearls"
        backup_state_key = f"supplier_order_full_backup::{parsed.source_hash}::{mode}"
        if st.button(
            "Подготовить полный резерв: Excel + состояние заказа",
            key=f"prepare_full_backup::{mode}",
            width="stretch",
        ):
            try:
                with st.spinner("Собираем переносимую резервную копию..."):
                    st.session_state[backup_state_key] = full_order_backup_bytes(parsed, draft)
            except (OSError, FileNotFoundError) as exc:
                st.warning(f"Полную резервную копию сейчас создать нельзя: {exc}")
        full_backup = st.session_state.get(backup_state_key)
        if isinstance(full_backup, bytes):
            st.download_button(
                "Скачать подготовленный ZIP",
                data=full_backup,
                file_name=f"supplier_order_backup_{safe_mode}_{datetime.now().date().isoformat()}.zip",
                mime="application/zip",
                width="stretch",
            )


@st.fragment
def _render_order_stage_fragment(
    parsed: ParsedOrderWorkbook,
    order_sets: tuple[OrderSet, ...],
    draft: OrderDraft,
    mode: str,
) -> None:
    """Interactive order cards rerun without rebuilding the whole application."""
    with timed_operation(
        "supplier_order.order_fragment",
        source_hash=parsed.source_hash,
        mode=mode,
    ):
        _render_order_workspace(parsed, order_sets, draft, mode)
        _render_export(parsed, order_sets, draft, mode)


@st.fragment
def _render_ring_stage_fragment(
    parsed: ParsedOrderWorkbook,
    order_sets: tuple[OrderSet, ...],
    draft: OrderDraft,
    mode: str,
) -> None:
    """Ring-size forms and Excel readiness rerun as one small workspace."""
    with timed_operation(
        "supplier_order.ring_fragment",
        source_hash=parsed.source_hash,
        mode=mode,
    ):
        _render_ring_sizes(parsed, order_sets, draft, mode)
        _render_export(parsed, order_sets, draft, mode)


def render_supplier_order_dashboard() -> None:
    parsed, _ = _render_upload()
    if parsed is None:
        return

    mode_key = "supplier_order_mode"
    if st.session_state.get(mode_key) not in ORDER_MODES:
        st.session_state[mode_key] = ORDER_MODE_STONES
    mode = st.segmented_control(
        "Какой заказ формируем?",
        list(ORDER_MODES),
        key=mode_key,
    ) or ORDER_MODE_STONES
    _flush_previous_mode_on_change(parsed, mode)
    draft = _get_session_draft(parsed, mode)
    recommendation_profile = st.segmented_control(
        "Режим автоматических рекомендаций",
        list(RECOMMENDATION_PROFILES),
        default=(
            draft.recommendation_profile
            if draft.recommendation_profile in RECOMMENDATION_PROFILES
            else RECOMMENDATION_BASE
        ),
        key=f"supplier_order_recommendation_profile::{parsed.source_hash}::{mode}",
    ) or RECOMMENDATION_BASE
    if recommendation_profile != draft.recommendation_profile:
        draft.recommendation_profile = recommendation_profile
        _save_session_draft(draft)

    _apply_pending_order_widget_cleanup()
    status_col, save_col = st.columns([3, 1])
    with status_col:
        st.markdown(
            '<div class="report-context"><div class="report-context-dot"></div><div class="report-context-copy">'
            f'<strong>Заказ открыт</strong><span>{escape(parsed.source_name)} · '
            f'поставщик: {escape(parsed.supplier or "не указан")} · период: {escape(parsed.period or "не указан")}</span>'
            '</div></div>',
            unsafe_allow_html=True,
        )
    with save_col:
        if st.button(
            "Сохранить сейчас",
            key=f"supplier_order_manual_save_inline::{parsed.source_hash}::{mode}",
            width="stretch",
        ):
            _flush_session_draft(draft)
            st.toast("Черновик сохранён")
    _render_cloud_autosave_fragment(draft)

    order_sets = _mode_sets(parsed, mode)
    _seed_defaults(draft, order_sets)
    _render_draft_tools(parsed, draft, mode)
    _render_overview(parsed, order_sets, mode, draft)

    if draft.stage == "rings":
        if st.button("← Вернуться к количествам", width="stretch"):
            draft.stage = "order"
            _flush_session_draft(draft)
            st.rerun()
        _render_ring_stage_fragment(parsed, order_sets, draft, mode)
    else:
        _render_order_stage_fragment(parsed, order_sets, draft, mode)
