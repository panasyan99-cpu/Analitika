from __future__ import annotations

import gc
import io
import json
import math
import re
import threading
from datetime import datetime
from dataclasses import dataclass
from html import escape
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from src.currency import get_vnd_per_usd
from src.navigation import NavigationItem, render_mobile_navigation, render_sidebar
from src.report import COLORED_ORDER, PEARL_ORDER, TOP_ORDER, classify
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


# Plotly charts are view-only while preserving hover/tap tooltips.
LOCKED_CHART_CONFIG = {
    "displayModeBar": False,
    "displaylogo": False,
    "scrollZoom": False,
    "doubleClick": False,
    "showTips": False,
    "editable": False,
    "staticPlot": False,
    "responsive": True,
    "showAxisDragHandles": False,
    "showAxisRangeEntryBoxes": False,
}

CATEGORY_LABELS = {
    "Bracelet": "Браслеты",
    "Earrings": "Серьги",
    "Ring": "Кольца",
    "Necklace": "Ожерелья",
    "Pendant": "Подвески",
    "Brooch": "Броши",
    "Other": "Другое",
}

# The first Sonu workbook was reviewed visually. These SKUs are the stable
# photo-based seed. New/unseen models are handled by the user's 50/50 rule.
SLIDER_BRACELETS = {
    "SSNB-AFG018-BS", "SSNB-000042-BS", "SSNB-000022-BS", "SSNB-000028-EM",
    "SSNB-000033-RUBY", "SSNB-000034-BS", "SSNB-000044-SWBT", "SSNB-000046-BS",
    "SSNB-000047-BS", "SSNB-000049-BS", "SSNB-AFG016-MOIS", "XB-SN-000019-BS",
    "XB-SN-000021-TANZ-BS", "XB-SN-KSB015-RDT", "XB-SN-KSB016-RUBY",
    "XB-SN-KSB079-BS", "SSNB-000035-GARN", "SSNB-000038-TANZ", "SSNB-000041-BS",
    "SSNB-000045-BS", "SSNB-AFG019-CIT", "XB-SN-000018-BS", "XB-SN-000752-RUBY",
    "XB-SN-KSB092-BS", "XB-SN-S4381L-BS", "SSNB-000035-RUBY", "SSNB-AFG011-EM",
    "XB-SN-000020-BS", "BB000312-RUBY", "BR001229-AMBER", "SSNB-000005-BS",
    "SSNB-000019-RUBY", "SSNB-000027-APAT", "SSNB-000029-TANZ", "SSNB-000037-TANZ",
    "SSNB-000048-EM", "SSNB-AFG015-BS", "SSNB-AFG020-APAT", "SSNB-AFG022-BS",
    "XB-SN-000024-FPW-RUBY", "XB-SN-000026-EM", "XB-SN-000175-BS",
    "XB-SN-000176-EM", "XB-SN-KSB088-BS", "XB-SN-KSB096-BS", "XB-SN-KST176-IOL",
    "XB-SN-KST177-BS",
}

FULL_CIRCLE_BRACELETS = {
    "XB-SN-KSB075-MOIS", "XB-SN-KSB077-57-MOIS", "XB-SN-KSB090-EM",
    "SSNB-AFG038-BS", "XB-SN-KSB041-BS", "XB-SN-KSB067-BS", "XB-SN-KSB069-CIT",
    "XB-SN-KSB077-37-MOIS", "XB-SN-KSB077-53-MOIS", "XB-SN-KSB078-OPAL",
    "XB-SN-KSB086-RUBY", "XB-SN-KSB087-BS", "XB-SN-KSB089-EM", "XB-SN-KSB103-AMST",
    "XB-SN-KSB104-RUBY", "XB-SN-KSB105-BS", "XB-SN-KSB108-BS", "SSNB-AFG001-BS",
    "SSNB-AFG003-BS", "SSNB-AFG007-BS", "XB-SN-KSB099-BS", "XB-SN-KSB101-EM",
    "XB-SN-KSB054-BS", "BB108-LBT", "XB-SN-KSB050-BS", "XB-SN-KSB064-RUBY",
    "XB-SN-KSB109-RUBY", "GSBM-25-1-MOIS", "BB-000003-EM", "BB100117-MYSTIC",
    "SSNB-AFG025-EM", "SSNB-AFG026-BS", "SSNB-AFG031-BS", "SSNB-AFG039-BS",
    "SSNB-AFG040-BS", "SSNB-AFG048-BS", "SSNB-AFG049-BS", "SSNB-AFG050-BS",
    "SSNB-AFG051-BS", "XB-SN-KSB040-RUBY", "XB-SN-KSB077-34-MOIS",
    "XB-SN-KSB098-BS", "XB-SN-KSB106-BS", "XB-SN-S121-SWBT",
}


CENTERED_BRACELET_LABEL = "С затяжкой"
FULL_CIRCLE_BRACELET_LABEL = "Без затяжки (в круг)"
BRACELET_TYPE_ORDER = [CENTERED_BRACELET_LABEL, FULL_CIRCLE_BRACELET_LABEL]

BRACELET_OVERRIDE_FILE = (
    Path(__file__).resolve().parents[1] / "data" / "bracelet_classification_overrides.json"
)
BRACELET_CATALOG_FILE = (
    Path(__file__).resolve().parents[1] / "data" / "bracelet_classification_catalog.json"
)
BRACELET_FAMILY_KEY_PREFIX = "FAMILY::"
BRACELET_OVERRIDE_SESSION_KEY = "sonu_bracelet_manual_overrides"
BRACELET_REVIEW_OPEN_KEY = "sonu_bracelet_review_open"
BRACELET_REVIEW_MODE_KEY = "sonu_bracelet_review_mode"
BRACELET_REVIEW_INDEX_KEY = "sonu_bracelet_review_index"
BRACELET_REVIEW_DRAFT_KEY = "sonu_bracelet_review_draft"


def _normalize_bracelet_override(value: object) -> str | None:
    text = " ".join(str(value or "").strip().split()).lower()
    if text in {
        CENTERED_BRACELET_LABEL.lower(), "с затяжкой", "не полный круг",
        "неполный круг", "центральная композиция", "centered", "slider",
    }:
        return CENTERED_BRACELET_LABEL
    if text in {
        FULL_CIRCLE_BRACELET_LABEL.lower(), "без затяжки", "полный круг",
        "в круг", "full circle", "circle",
    }:
        return FULL_CIRCLE_BRACELET_LABEL
    return None


def _bracelet_family_override_key(family: object) -> str:
    value = " ".join(str(family or "").strip().upper().split())
    return f"{BRACELET_FAMILY_KEY_PREFIX}{value}" if value else ""


def _validated_bracelet_overrides(payload: object) -> dict[str, str]:
    """Accept legacy flat files plus v2 SKU/family decision backups."""
    flattened: dict[str, object] = {}
    if isinstance(payload, dict):
        if isinstance(payload.get("overrides"), dict):
            flattened.update(payload["overrides"])
        if isinstance(payload.get("sku_overrides"), dict):
            flattened.update(payload["sku_overrides"])
        if isinstance(payload.get("family_overrides"), dict):
            flattened.update({
                _bracelet_family_override_key(key): value
                for key, value in payload["family_overrides"].items()
            })
        if not flattened:
            flattened.update(payload)
    if not isinstance(flattened, dict):
        return {}
    result: dict[str, str] = {}
    for raw_key, value in flattened.items():
        normalized = _normalize_bracelet_override(value)
        key = str(raw_key or "").strip().upper()
        if key.startswith(BRACELET_FAMILY_KEY_PREFIX):
            family = key[len(BRACELET_FAMILY_KEY_PREFIX):].strip()
            key = _bracelet_family_override_key(family)
        if key and normalized:
            result[key] = normalized
    return result


def load_bracelet_catalog() -> dict[str, object]:
    """Load the reviewed Sonu seed catalog generated from sonunew.xlsx."""
    empty = {"sku_overrides": {}, "family_overrides": {}, "pending_families": {}, "metadata": {}}
    try:
        payload = json.loads(BRACELET_CATALOG_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return empty
    if not isinstance(payload, dict):
        return empty
    sku_overrides = _validated_bracelet_overrides({"sku_overrides": payload.get("sku_overrides", {})})
    family_flat = _validated_bracelet_overrides({"family_overrides": payload.get("family_overrides", {})})
    family_overrides = {
        key[len(BRACELET_FAMILY_KEY_PREFIX):]: value
        for key, value in family_flat.items()
        if key.startswith(BRACELET_FAMILY_KEY_PREFIX)
    }
    pending = payload.get("pending_families", {})
    if not isinstance(pending, dict):
        pending = {}
    metadata = payload.get("metadata", {})
    if not isinstance(metadata, dict):
        metadata = {}
    return {
        "sku_overrides": sku_overrides,
        "family_overrides": family_overrides,
        "pending_families": pending,
        "metadata": metadata,
    }


def load_bracelet_overrides() -> dict[str, str]:
    """Load durable manual bracelet decisions plus this browser session's changes."""
    stored: dict[str, str] = {}
    try:
        if BRACELET_OVERRIDE_FILE.exists():
            stored = _validated_bracelet_overrides(
                json.loads(BRACELET_OVERRIDE_FILE.read_text(encoding="utf-8"))
            )
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        stored = {}
    try:
        if st.runtime.exists():
            session_payload = st.session_state.get(BRACELET_OVERRIDE_SESSION_KEY, {})
            stored.update(_validated_bracelet_overrides(session_payload))
    except Exception:
        pass
    return stored


def bracelet_overrides_json(overrides: dict[str, str] | None = None) -> bytes:
    decisions = load_bracelet_overrides() if overrides is None else _validated_bracelet_overrides(overrides)
    sku_overrides = {
        key: value for key, value in decisions.items()
        if not key.startswith(BRACELET_FAMILY_KEY_PREFIX)
    }
    family_overrides = {
        key[len(BRACELET_FAMILY_KEY_PREFIX):]: value
        for key, value in decisions.items()
        if key.startswith(BRACELET_FAMILY_KEY_PREFIX)
    }
    payload = {
        "version": 2,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "sku_overrides": dict(sorted(sku_overrides.items())),
        "family_overrides": dict(sorted(family_overrides.items())),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def save_bracelet_overrides(
    decisions: dict[str, str],
    *,
    replace: bool = False,
) -> tuple[dict[str, str], bool, str]:
    """Save manual decisions in session and, when possible, to the app data file."""
    validated = _validated_bracelet_overrides(decisions)
    merged = {} if replace else load_bracelet_overrides()
    merged.update(validated)
    try:
        if st.runtime.exists():
            st.session_state[BRACELET_OVERRIDE_SESSION_KEY] = merged.copy()
    except Exception:
        pass

    persisted = False
    message = "Решения сохранены в текущем сеансе."
    try:
        BRACELET_OVERRIDE_FILE.parent.mkdir(parents=True, exist_ok=True)
        temporary = BRACELET_OVERRIDE_FILE.with_suffix(".tmp")
        temporary.write_bytes(bracelet_overrides_json(merged))
        temporary.replace(BRACELET_OVERRIDE_FILE)
        persisted = True
        message = "Решения сохранены в справочнике приложения."
    except OSError:
        message = (
            "Решения сохранены в текущем сеансе. Сервер не разрешил постоянную запись — "
            "скачайте резервный JSON и добавьте его в папку data репозитория."
        )
    return merged, persisted, message


def import_bracelet_overrides(file_bytes: bytes) -> tuple[dict[str, str], bool, str]:
    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Файл решений не распознан. Нужен JSON, выгруженный из Analitika.") from exc
    decisions = _validated_bracelet_overrides(payload)
    if not decisions:
        raise ValueError("В файле нет корректных решений по браслетам.")
    return save_bracelet_overrides(decisions)

# В первую группу также входят модели на кольцах с облегченной центральной
# композицией: камни сосредоточены преимущественно в середине, а не по всему кругу.
CENTERED_BRACELETS = SLIDER_BRACELETS


SONU_SECTIONS = [
    ("sonu-main-report", "Основной отчет"),
    ("sonu-bracelet-classification", "Классификация браслетов"),
    ("sonu-extra", "Общие выводы"),
]


STONE_GROUP_LABELS = {
    "TOP STONES": "Top Stones",
    "PEARLS": "Pearls",
    "COLORED STONES": "Other Stones",
}
STONE_GROUP_ORDER = ["Top Stones", "Pearls", "Other Stones"]
STONE_MEMBER_ORDER = {
    "Top Stones": TOP_ORDER,
    "Pearls": PEARL_ORDER,
    "Other Stones": COLORED_ORDER,
}

# Сокращения, используемые в артикулах и выгрузках Sonu. Порядок важен:
# он повторяет бизнес-приоритеты общей аналитики (Moissanite → Sapphire → Ruby...).
SONU_STONE_ALIASES: tuple[tuple[str, str, str], ...] = (
    (r"MOIS|MOISSANITE|MOISANITE", "MOISSANITE", "MOIS"),
    (r"SAPPHIRE|SAPPHRIE|(?:^|[-_/\s])BS(?:$|[-_/\s])", "SAPPHIRE", "BS"),
    (r"RUBY", "RUBY", "RUBY"),
    (r"LONDON|(?:^|[-_/\s])LBT(?:$|[-_/\s])", "LONDON TOPAZ", "LBT"),
    (r"SWISS|(?:^|[-_/\s])SWBT(?:$|[-_/\s])", "SWISS TOPAZ", "SWBT"),
    (r"WHITE TOPAZ|BLUE TOPAZ|(?:^|[-_/\s])W?BT(?:$|[-_/\s])", "BLUE TOPAZ", "BT"),
    (r"CREATED EMERALD|(?:^|[-_/\s])CE(?:$|[-_/\s])", "CREATED EMERALD", "CE"),
    (r"CHROME DIOPSIDE|(?:^|[-_/\s])CD(?:$|[-_/\s])", "CHROME DIOPSIDE", "CD"),
    (r"EMERALD|(?:^|[-_/\s])EM(?:$|[-_/\s])", "EMERALD", "EM"),
    (r"PERIDOT|(?:^|[-_/\s])PERI(?:$|[-_/\s])", "PERIDOT", "PERI"),
    (r"BAROQUE", "BAROQUE PEARL", "BAROQUE"),
    (r"AKOYA", "AKOYA", "AKOYA"),
    (r"TAHITI|TAHITIAN|(?:^|[-_/\s])TAH(?:$|[-_/\s])", "TAHITI", "TAH"),
    (r"SOUTH SEA|(?:^|[-_/\s])SSP(?:$|[-_/\s])", "SOUTH SEA PEARL", "SSP"),
    (r"FRESHWATER PEARL (?:PINK|ROSE|GRAY|GREY|BLACK)|(?:^|[-_/\s])FPC(?:$|[-_/\s])", "FRESHWATER PEARL PINK", "FPC"),
    (r"FRESHWATER|(?:^|[-_/\s])FPW(?:$|[-_/\s])|(?:^|[-_/\s])FWP(?:$|[-_/\s])", "FRESHWATER PEARL WHITE", "FPW"),
    (r"BLACK SPINEL|(?:^|[-_/\s])BSP(?:$|[-_/\s])", "BLACK SPINEL", "BSP"),
    (r"ONYX", "ONYX", "ONYX"),
    (r"OBSIDIAN", "OBSIDIAN", "OBSIDIAN"),
    (r"AMETHYST|(?:^|[-_/\s])AMST(?:$|[-_/\s])", "AMETHYST", "AMST"),
    (r"MYSTIC", "MYSTIC QUARTZ", "MYSTIC"),
    (r"CITRINE|(?:^|[-_/\s])CIT(?:$|[-_/\s])", "CITRINE", "CIT"),
    (r"SMOKY|SMOKEY|RAUCH|HONEY", "SMOKY QUARTZ", "SMOKY"),
    (r"QUARTZ", "QUARTZ", "QUARTZ"),
    (r"GARNET|(?:^|[-_/\s])GARN(?:$|[-_/\s])", "GARNET", "GARN"),
    (r"RHODOLITE|RODOLITE|(?:^|[-_/\s])RDT(?:$|[-_/\s])", "RHODOLITE", "RDT"),
    (r"GREEN AGATE", "GREEN AGATE", "GREEN AGATE"),
    (r"AGATE|(?:^|[-_/\s])AGAT(?:$|[-_/\s])", "AGATE", "AGATE"),
    (r"APATITE|(?:^|[-_/\s])APAT(?:$|[-_/\s])", "APATITE", "APAT"),
    (r"TANZANITE|(?:^|[-_/\s])TANZ(?:$|[-_/\s])", "TANZANITE", "TANZ"),
    (r"IOLITE|(?:^|[-_/\s])IOL(?:$|[-_/\s])", "IOLITE", "IOL"),
    (r"OPAL", "OPAL", "OPAL"),
    (r"AMBER", "AMBER", "AMBER"),
    (r"LAPIS", "LAPIS LAZURITE", "LAPIS"),
    (r"TURQUOISE", "TURQUOISE", "TURQUOISE"),
)


STONE_SUFFIX_TOKENS = {
    token
    for _, _, abbreviation in SONU_STONE_ALIASES
    for token in re.split(r"[-_/\s]+", abbreviation.upper())
    if token
}
STONE_SUFFIX_TOKENS.update({
    "BS", "RUBY", "MOIS", "EM", "AMST", "CIT", "APAT", "TANZ", "RDT",
    "SWBT", "LBT", "BT", "OPAL", "FPW", "FPC", "IOL", "GARN", "AMBER",
    "MYSTIC", "TOUR", "KYN", "PERI", "CD", "CE", "TAH", "SSP",
})


def _bracelet_model_key(sku: Any) -> str:
    """Return the design family, ignoring stone suffixes of the same bracelet model."""
    parts = [part for part in re.split(r"[-_/]+", str(sku or "").upper()) if part]
    while parts and parts[-1] in STONE_SUFFIX_TOKENS:
        parts.pop()
    return "-".join(parts)


CENTERED_MODEL_KEYS = {_bracelet_model_key(sku) for sku in CENTERED_BRACELETS}
FULL_CIRCLE_MODEL_KEYS = {_bracelet_model_key(sku) for sku in FULL_CIRCLE_BRACELETS}

SONU_METAL_GROUPS = ("Серебро", "Золото и платина", "Другое")


def classify_sonu_metal_group(purity: object) -> str:
    text = " ".join(str(purity or "").strip().split()).upper().replace("Ё", "Е")
    compact = re.sub(r"[^A-ZА-Я0-9]+", "", text)
    if "AU" in compact or "GOLD" in compact or "ЗОЛОТ" in compact or "PT" in compact or "PLATIN" in compact or "ПЛАТИН" in compact:
        return "Золото и платина"
    if "925" in compact or compact.startswith("AG") or "SILVER" in compact or "СЕРЕБ" in compact:
        return "Серебро"
    return "Другое"


def filter_sonu_metal_groups(frame: pd.DataFrame, selected_groups: Iterable[str]) -> pd.DataFrame:
    selected = {str(value) for value in selected_groups}
    result = frame.copy()
    result["Группа металла"] = result.get("Проба", pd.Series("", index=result.index)).map(classify_sonu_metal_group)
    return result.loc[result["Группа металла"].isin(selected)].reset_index(drop=True)


def _sync_detected_purities(values: Iterable[object]) -> None:
    normalized = tuple(sorted({" ".join(str(value or "").strip().split()) or "Не указано" for value in values}))
    key = "global_filter_detected::Заказ Sonu"
    if tuple(st.session_state.get(key, ())) != normalized:
        st.session_state[key] = normalized
        st.rerun()


def sonu_navigation_items(has_report: bool) -> list[NavigationItem]:
    """Return the complete Sonu menu even before a workbook is uploaded."""
    definitions = [
        ("sonu-upload", "Загрузка отчета", "#sonu-upload", True),
        *[(anchor, label, f"#{anchor}", has_report) for anchor, label in SONU_SECTIONS],
        ("sonu-export", "Полная выгрузка", "#sonu-export", has_report),
        ("about", "О программе", "#about", True),
    ]
    return [
        NavigationItem(
            item_id=item_id,
            label=label,
            href=href,
            enabled=enabled,
            current=(item_id == "sonu-upload" and not has_report),
        )
        for item_id, label, href, enabled in definitions
    ]


def _sonu_sidebar_navigation(
    has_report: bool,
    *,
    status_text: str | None = None,
    status_tone: str | None = None,
    action_label: str | None = None,
    action_key: str | None = None,
):
    """Render Sonu through the same sidebar shell as every other workspace."""
    return render_sidebar(
        module_title="Заказ Sonu",
        navigation_title="Навигация по отчету",
        items=sonu_navigation_items(has_report),
        status_text=status_text or ("Отчет Sonu загружен" if has_report else "Ожидается отчет Sonu"),
        status_tone=(status_tone or ("success" if has_report else "neutral")),
        source_text="Источник: Excel · отчет Sonu",
        action_label=action_label,
        action_key=action_key,
    )


def _sonu_mobile_navigation(has_report: bool) -> None:
    """Sticky touch navigation using the shared component and states."""
    render_mobile_navigation(sonu_navigation_items(has_report))


def _anchor(anchor: str) -> None:
    st.markdown(f'<div id="{anchor}" class="report-anchor"></div>', unsafe_allow_html=True)


def _safe_excel_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Remove binary image payloads and normalize missing values before Excel export."""
    result = frame.copy()
    if "Фото" in result.columns:
        result["Фото"] = result["Фото"].map(
            lambda value: "Есть" if isinstance(value, (bytes, bytearray)) else ""
        )
    return result.replace({pd.NA: None})


def _classify_sonu_stone(raw_stone: Any, sku: Any) -> tuple[str, str, str]:
    """Map Sonu names/SKU abbreviations to the shared business stone groups."""
    raw = str(raw_stone or "").strip()
    sku_text = str(sku or "").strip()
    combined = re.sub(r"\s+", " ", f"{raw} {sku_text}".upper()).strip()
    for pattern, expanded_name, abbreviation in SONU_STONE_ALIASES:
        if re.search(pattern, combined):
            segment, member, _ = classify(expanded_name)
            return STONE_GROUP_LABELS.get(segment, segment), member, abbreviation

    if raw and raw.casefold() not in {"не указан", "не указана", "none", "nan"}:
        segment, member, _ = classify(raw)
        return STONE_GROUP_LABELS.get(segment, segment), member, raw
    return "Other Stones", "Other Colored Stones", "Не определено"


def add_stone_classification(frame: pd.DataFrame) -> pd.DataFrame:
    """Add group/member columns without replacing the original stone name."""
    result = frame.copy()
    classified = result.apply(
        lambda row: _classify_sonu_stone(row.get("Камень"), row.get("SKU")), axis=1
    )
    result[["Группа камня", "Камень группы", "Сокращение"]] = pd.DataFrame(
        classified.tolist(), index=result.index
    )
    return result


def _first_image(values: pd.Series) -> bytes | None:
    return next((value for value in values if isinstance(value, (bytes, bytearray))), None)


def add_order_horizons(
    frame: pd.DataFrame,
    period_days: int,
    *,
    sales_column: str = "Продано за период",
    stock_column: str = "Остаток сети",
) -> pd.DataFrame:
    """Calculate how many pieces are needed for 30/45/90 days of coverage."""
    result = frame.copy()
    days = max(int(period_days or 0), 1)
    sales = pd.to_numeric(result.get(sales_column, 0), errors="coerce").fillna(0).clip(lower=0)
    stock = pd.to_numeric(result.get(stock_column, 0), errors="coerce").fillna(0).clip(lower=0)
    result["Продажи в день"] = sales / days
    for horizon in (30, 45, 90):
        required = sales * horizon / days
        result[f"Нужно на {horizon} дней"] = [
            max(int(math.ceil(float(target) - float(available))), 0)
            for target, available in zip(required, stock)
        ]
    return result


def network_sku_snapshot(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """Return one network-level row per SKU.

    The stock column in the final Sonu report already contains the total stock
    across the whole retail network and repeats on every store row of the same
    SKU. Sales are summed across stores, while stock is taken once.
    """
    enriched = add_stone_classification(frame)
    if enriched.empty:
        return pd.DataFrame()
    if "Остаток сети" not in enriched.columns:
        enriched["Остаток сети"] = 0.0
    aggregations: dict[str, tuple[str, Any]] = {
        "Категория": ("Категория", "first"),
        "Категория RU": ("Категория RU", "first"),
        "Камень": ("Камень", "first"),
        "Группа камня": ("Группа камня", "first"),
        "Камень группы": ("Камень группы", "first"),
        "Сокращение": ("Сокращение", "first"),
        "Проба": ("Проба", "first"),
        "Магазинов с продажами": ("Магазин", "nunique"),
        # Never sum this field: it is the same network stock repeated by store.
        "Остаток сети": ("Остаток сети", "max"),
        "Продано за период": ("Скорость продаж", "sum"),
        "Продажи VND": ("Продажи VND", "sum"),
    }
    if "Фото" in enriched.columns:
        aggregations["Фото"] = ("Фото", _first_image)
    grouped = enriched.groupby("SKU", as_index=False, dropna=False).agg(**aggregations)
    grouped["Продажи USD"] = grouped["Продажи VND"] / float(rate or 1)
    grouped["Средняя цена USD"] = pd.to_numeric(
        grouped["Продажи USD"] / grouped["Продано за период"].replace(0, pd.NA),
        errors="coerce",
    ).fillna(0.0)
    grouped = add_order_horizons(grouped, period_days)
    return grouped.sort_values(
        ["Продано за период", "Продажи USD", "SKU"], ascending=[False, False, True]
    ).reset_index(drop=True)



def _aggregate_sales_stock(
    sku: pd.DataFrame,
    group_columns: list[str],
    period_days: int,
) -> pd.DataFrame:
    """Aggregate one-row-per-SKU data into the user-facing sales/stock shape."""
    if sku.empty:
        return pd.DataFrame()
    working = sku.copy()
    working["_sold_sku"] = (pd.to_numeric(working["Продано за период"], errors="coerce").fillna(0) > 0).astype(int)
    working["_stock_sku"] = (pd.to_numeric(working["Остаток сети"], errors="coerce").fillna(0) > 0).astype(int)
    result = working.groupby(group_columns, as_index=False, dropna=False).agg(
        **{
            "Продано уникальных SKU": ("_sold_sku", "sum"),
            "Продано штук": ("Продано за период", "sum"),
            "Продано на сумму, USD": ("Продажи USD", "sum"),
            "Осталось уникальных SKU": ("_stock_sku", "sum"),
            "Всего штук": ("Остаток сети", "sum"),
        }
    )
    days = max(int(period_days or 0), 1)
    sold = pd.to_numeric(result["Продано штук"], errors="coerce").fillna(0).clip(lower=0)
    stock = pd.to_numeric(result["Всего штук"], errors="coerce").fillna(0).clip(lower=0)
    result["Продажи в день"] = sold / days
    result["Покрытие остатком, дней"] = [
        (float(available) / float(speed)) if float(speed) > 0 else math.nan
        for available, speed in zip(stock, result["Продажи в день"])
    ]
    for horizon in (30, 45, 90):
        demand = sold * horizon / days
        result[f"К заказу на {horizon} дней"] = [
            max(int(math.ceil(float(target) - float(available))), 0)
            for target, available in zip(demand, stock)
        ]
    result["Обеспеченность SKU, %"] = [
        min(float(stock_sku) / float(sold_sku), 2.0) if float(sold_sku) > 0 else 1.0
        for stock_sku, sold_sku in zip(
            result["Осталось уникальных SKU"], result["Продано уникальных SKU"]
        )
    ]
    return result


def _business_sort(frame: pd.DataFrame, *, include_category: bool = True) -> pd.DataFrame:
    if frame.empty:
        return frame
    result = frame.copy()
    group_order = {name: index for index, name in enumerate(STONE_GROUP_ORDER)}
    member_order = {
        (group, member): index
        for group, members in STONE_MEMBER_ORDER.items()
        for index, member in enumerate(members)
    }
    result["_group_order"] = result["Группа камня"].map(group_order).fillna(99)
    result["_member_order"] = [
        member_order.get((group, member), 99)
        for group, member in zip(result["Группа камня"], result["Камень группы"])
    ]
    order_columns = ["_group_order", "_member_order"]
    ascending = [True, True]
    if include_category and "Номенклатурная группа" in result.columns:
        order_columns.append("Номенклатурная группа")
        ascending.append(True)
    order_columns.extend(["Продано штук", "Продано на сумму, USD"])
    ascending.extend([False, False])
    return result.sort_values(order_columns, ascending=ascending).drop(
        columns=["_group_order", "_member_order"]
    ).reset_index(drop=True)


def network_assortment_summary(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """Large non-bracelet summary: stone group → stone → nomenclature group."""
    sku = network_sku_snapshot(frame, rate, period_days)
    if sku.empty:
        return pd.DataFrame()
    sku = sku.loc[sku["Категория"] != "Bracelet"].copy()
    if sku.empty:
        return pd.DataFrame()
    sku["Номенклатурная группа"] = sku["Категория RU"]
    result = _aggregate_sales_stock(
        sku,
        ["Группа камня", "Камень группы", "Номенклатурная группа"],
        period_days,
    )
    return _business_sort(result, include_category=True)


def network_group_summary(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """Chart-ready non-bracelet summary by the three business stone groups."""
    detail = network_assortment_summary(frame, rate, period_days)
    if detail.empty:
        return pd.DataFrame()
    numeric = [
        "Продано уникальных SKU", "Продано штук", "Продано на сумму, USD",
        "Осталось уникальных SKU", "Всего штук",
    ]
    result = detail.groupby("Группа камня", as_index=False, dropna=False)[numeric].sum()
    order = {name: index for index, name in enumerate(STONE_GROUP_ORDER)}
    result["_order"] = result["Группа камня"].map(order).fillna(99)
    return result.sort_values("_order").drop(columns="_order").reset_index(drop=True)


def complete_assortment_summary(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """One complete report for every assortment unit, including bracelet types."""
    sku = network_sku_snapshot(frame, rate, period_days)
    if sku.empty:
        return pd.DataFrame()
    sku = sku.copy()
    sku["Номенклатурная группа"] = sku["Категория RU"]
    bracelets = classify_bracelets(frame, rate, period_days)
    if not bracelets.empty:
        bracelet_types = bracelets.set_index("SKU")["Тип браслета"]
        mapped = sku["SKU"].map(bracelet_types)
        bracelet_mask = sku["Категория"] == "Bracelet"
        classified_mask = bracelet_mask & mapped.notna()
        pending_mask = bracelet_mask & mapped.isna()
        sku.loc[classified_mask, "Номенклатурная группа"] = (
            "Браслеты · " + mapped.loc[classified_mask].astype(str)
        )
        sku.loc[pending_mask, "Номенклатурная группа"] = "Браслеты · Требует классификации"
    result = _aggregate_sales_stock(
        sku,
        ["Группа камня", "Камень группы", "Номенклатурная группа"],
        period_days,
    )
    return _business_sort(result, include_category=True)


PRIORITY_ORDER = {
    "Очень нужно заказать": 1,
    "Нужно заказать": 2,
    "Желательно заказать": 3,
    "Плановое пополнение": 4,
    "Не критично": 5,
}


def _priority_for_row(row: pd.Series) -> str:
    sold = float(row.get("Продано штук", 0) or 0)
    if sold <= 0:
        return "Не критично"
    coverage = float(row.get("Покрытие остатком, дней", math.inf))
    sold_sku = float(row.get("Продано уникальных SKU", 0) or 0)
    stock_sku = float(row.get("Осталось уникальных SKU", 0) or 0)
    model_shortage = sold_sku > 0 and stock_sku < sold_sku * 0.5
    if coverage < 15 or stock_sku <= 0:
        return "Очень нужно заказать"
    if coverage < 30 or model_shortage:
        return "Нужно заказать"
    if coverage < 45:
        return "Желательно заказать"
    if coverage < 90:
        return "Плановое пополнение"
    return "Не критично"


def order_priority_report(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """Transparent AI-ready ranking based on sales speed, models and network stock."""
    report = complete_assortment_summary(frame, rate, period_days)
    if report.empty:
        return report
    report = report.copy()
    report["Приоритет заказа"] = report.apply(_priority_for_row, axis=1)
    report["_priority"] = report["Приоритет заказа"].map(PRIORITY_ORDER).fillna(99)
    return report.sort_values(
        ["_priority", "К заказу на 45 дней", "Продано штук", "Продано на сумму, USD"],
        ascending=[True, False, False, False],
    ).drop(columns="_priority").reset_index(drop=True)


def ai_sales_summary(report: pd.DataFrame, target_days: int) -> dict[str, Any]:
    """Create a compact management narrative from the transparent priority report."""
    if report.empty:
        return {
            "headline": "Нет данных для анализа.",
            "order_text": "Загрузите отчет с продажами и общим остатком сети.",
            "top_lines": [],
            "order_qty": 0,
            "urgent_count": 0,
        }
    sold_sku = int(report["Продано уникальных SKU"].sum())
    sold_qty = float(report["Продано штук"].sum())
    sales_usd = float(report["Продано на сумму, USD"].sum())
    stock_sku = int(report["Осталось уникальных SKU"].sum())
    stock_qty = float(report["Всего штук"].sum())
    order_column = f"К заказу на {int(target_days)} дней"
    order_qty = int(report[order_column].sum())
    urgent = report.loc[report["Приоритет заказа"] == "Очень нужно заказать"]
    candidates = report.loc[report[order_column] > 0].head(5)
    top_lines = [
        (
            f'{row["Камень группы"]} · {row["Номенклатурная группа"]}: '
            f'продано {_money(row["Продано штук"])} шт., остаток {_money(row["Всего штук"])} шт., '
            f'к заказу {_money(row[order_column])} шт.'
        )
        for _, row in candidates.iterrows()
    ]
    return {
        "headline": (
            f"За период продано {_money(sold_qty)} изделий по {_money(sold_sku)} уникальным SKU "
            f"на сумму ${_money(sales_usd)}. В сети осталось {_money(stock_qty)} изделий "
            f"по {_money(stock_sku)} SKU."
        ),
        "order_text": (
            f"Для покрытия продаж на {int(target_days)} дней расчетно требуется заказать "
            f"{_money(order_qty)} изделий. Позиций с самым высоким приоритетом: {len(urgent)}."
        ),
        "top_lines": top_lines,
        "order_qty": order_qty,
        "urgent_count": len(urgent),
    }


def stock_conflict_details(frame: pd.DataFrame) -> pd.DataFrame:
    """List SKUs whose repeated network stock is not identical across store rows."""
    if "Остаток сети" not in frame.columns or frame.empty:
        return pd.DataFrame(columns=["SKU", "Значения остатка", "Строк"])
    rows: list[dict[str, Any]] = []
    for sku, group in frame.groupby("SKU", dropna=False):
        values = sorted({float(value) for value in group["Остаток сети"].dropna().tolist()})
        if len(values) > 1:
            rows.append({
                "SKU": str(sku),
                "Значения остатка": ", ".join(_money(value) for value in values),
                "Строк": len(group),
            })
    return pd.DataFrame(rows)


def stock_conflict_skus(frame: pd.DataFrame) -> list[str]:
    """Backward-compatible compact conflict list used by tests and checks."""
    details = stock_conflict_details(frame)
    return details["SKU"].astype(str).tolist() if not details.empty else []


def stone_group_summary(frame: pd.DataFrame, rate: float) -> pd.DataFrame:
    enriched = add_stone_classification(frame)
    result = enriched.groupby("Группа камня", as_index=False, dropna=False).agg(
        Моделей=("SKU", "nunique"),
        Магазинов=("Магазин", "nunique"),
        **{
            "Скорость продаж": ("Скорость продаж", "sum"),
            "Продажи VND": ("Продажи VND", "sum"),
        },
    )
    result["Продажи USD"] = result["Продажи VND"] / rate
    result["Средняя цена USD"] = result["Продажи USD"] / result["Скорость продаж"].replace(0, pd.NA)
    result["Средняя цена USD"] = result["Средняя цена USD"].fillna(0)
    total_qty = float(result["Скорость продаж"].sum())
    total_sales = float(result["Продажи USD"].sum())
    result["Доля количества"] = result["Скорость продаж"] / total_qty if total_qty else 0.0
    result["Доля выручки"] = result["Продажи USD"] / total_sales if total_sales else 0.0
    order = {name: index for index, name in enumerate(STONE_GROUP_ORDER)}
    result["_order"] = result["Группа камня"].map(order).fillna(99)
    return result.drop(columns=["Продажи VND"]).sort_values("_order").drop(columns="_order").reset_index(drop=True)


def stone_member_summary(frame: pd.DataFrame, rate: float) -> pd.DataFrame:
    enriched = add_stone_classification(frame)

    def abbreviations(values: pd.Series) -> str:
        unique = sorted({str(value).strip() for value in values if str(value).strip()})
        return ", ".join(unique[:8])

    result = enriched.groupby(
        ["Группа камня", "Камень группы"], as_index=False, dropna=False
    ).agg(
        Сокращения=("Сокращение", abbreviations),
        Моделей=("SKU", "nunique"),
        Магазинов=("Магазин", "nunique"),
        **{
            "Скорость продаж": ("Скорость продаж", "sum"),
            "Продажи VND": ("Продажи VND", "sum"),
        },
    )
    result["Продажи USD"] = result["Продажи VND"] / rate
    result["Средняя цена USD"] = result["Продажи USD"] / result["Скорость продаж"].replace(0, pd.NA)
    result["Средняя цена USD"] = result["Средняя цена USD"].fillna(0)
    group_qty = result.groupby("Группа камня")["Скорость продаж"].transform("sum")
    group_sales = result.groupby("Группа камня")["Продажи USD"].transform("sum")
    result["Доля количества внутри группы"] = result["Скорость продаж"] / group_qty.replace(0, pd.NA)
    result["Доля выручки внутри группы"] = result["Продажи USD"] / group_sales.replace(0, pd.NA)
    result[["Доля количества внутри группы", "Доля выручки внутри группы"]] = result[
        ["Доля количества внутри группы", "Доля выручки внутри группы"]
    ].fillna(0)
    group_order = {name: index for index, name in enumerate(STONE_GROUP_ORDER)}
    member_order = {
        (group, member): index
        for group, members in STONE_MEMBER_ORDER.items()
        for index, member in enumerate(members)
    }
    result["_group_order"] = result["Группа камня"].map(group_order).fillna(99)
    result["_member_order"] = [
        member_order.get((group, member), 99)
        for group, member in zip(result["Группа камня"], result["Камень группы"])
    ]
    return result.drop(columns=["Продажи VND"]).sort_values(
        ["_group_order", "_member_order", "Продажи USD"], ascending=[True, True, False]
    ).drop(columns=["_group_order", "_member_order"]).reset_index(drop=True)


def stone_store_summary(frame: pd.DataFrame, rate: float) -> pd.DataFrame:
    enriched = add_stone_classification(frame)
    result = enriched.groupby(
        ["Магазин", "Группа камня"], as_index=False, dropna=False
    ).agg(
        Моделей=("SKU", "nunique"),
        **{
            "Скорость продаж": ("Скорость продаж", "sum"),
            "Продажи VND": ("Продажи VND", "sum"),
        },
    )
    result["Продажи USD"] = result["Продажи VND"] / rate
    return result.drop(columns="Продажи VND").sort_values(
        ["Магазин", "Продажи USD"], ascending=[True, False]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False, ttl=1800, max_entries=4)
def build_full_sonu_export(
    frame: pd.DataFrame,
    period: str,
    supplier: str,
    rate: float,
    overrides_signature: str = "",
) -> bytes:
    """Build the streamlined Sonu workbook around five merchandise groups."""
    period_days = _period_days(period)
    network_sku = network_sku_snapshot(frame, rate, period_days)
    section_tables = sonu_merchandise_tables(frame, rate, period_days)
    conflicts = stock_conflict_details(frame)
    source = add_usd_columns(add_stone_classification(frame), rate).sort_values(
        ["Скорость продаж", "Продажи USD"], ascending=False
    )
    category_overview = sonu_category_overview(section_tables)
    stone_category_overview = sonu_stone_category_overview(frame, rate, period_days)
    bracelet_classification_summary, bracelet_classification_detail = bracelet_classification_audit(
        frame, rate, period_days
    )
    ambiguous_bracelets = bracelet_classification_detail.loc[
        bracelet_classification_detail.get("Статус классификации", pd.Series(dtype=str)) == "Спорная модель"
    ].copy() if not bracelet_classification_detail.empty else pd.DataFrame()
    manual_bracelets = bracelet_classification_detail.loc[
        bracelet_classification_detail.get("Статус классификации", pd.Series(dtype=str)) == "Разобрано вручную"
    ].copy() if not bracelet_classification_detail.empty else pd.DataFrame()
    all_recommendations = []
    for section_name, table in section_tables.items():
        rec = sonu_order_recommendations(table)
        if not rec.empty:
            rec.insert(0, "Номенклатурная группа", section_name)
            all_recommendations.append(rec)
    recommendations = pd.concat(all_recommendations, ignore_index=True) if all_recommendations else pd.DataFrame()

    summary = pd.DataFrame([
        ("Период", period),
        ("Поставщик", supplier),
        ("Курс VND за 1 USD", rate),
        ("Продано уникальных SKU", int((network_sku["Продано за период"] > 0).sum()) if not network_sku.empty else 0),
        ("Продано за период, шт.", float(network_sku["Продано за период"].sum()) if not network_sku.empty else 0),
        ("Продажи, USD", float(network_sku["Продажи USD"].sum()) if not network_sku.empty else 0),
        ("SKU на остатке", int((network_sku["Остаток сети"] > 0).sum()) if not network_sku.empty else 0),
        ("Общий остаток сети, шт.", float(network_sku["Остаток сети"].sum()) if not network_sku.empty else 0),
        ("Расхождений сетевого остатка", len(conflicts)),
    ], columns=["Показатель", "Значение"])

    sheets = [("Сводка", summary), ("Камни и группы", stone_category_overview), ("Категории", category_overview)]
    for name, table in section_tables.items():
        sheets.append((name[:31], table))
    sheets.extend([
        ("Рекомендации", recommendations),
        ("Классификация браслетов", bracelet_classification_summary),
        ("Ручные решения", manual_bracelets),
        ("Спорные браслеты", ambiguous_bracelets),
        ("Все браслеты", bracelet_classification_detail),
        ("SKU сети", _user_facing_stone_columns(network_sku)),
        ("Контроль остатков", conflicts),
        ("Исходные данные", _user_facing_stone_columns(source)),
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, sheet_frame in sheets:
            _safe_excel_frame(sheet_frame).to_excel(writer, sheet_name=sheet_name[:31], index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="2B2115")
        header_font = Font(color="F6D899", bold=True)
        thin_gold = Side(style="thin", color="C59A52")
        border = Border(bottom=thin_gold)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.sheet_view.showGridLines = False
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            worksheet.row_dimensions[1].height = 30
            headers = {cell.column: str(cell.value or "") for cell in worksheet[1]}
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, bool) or not isinstance(cell.value, (int, float)):
                        continue
                    header = headers.get(cell.column, "")
                    if _is_percentage_column(header):
                        cell.number_format = "0%"
                    elif header == "Покрытие остатком, дней":
                        cell.number_format = "0.0"
                    else:
                        # No cents. Excel shows one separator for every three digits.
                        cell.value = round(float(cell.value))
                        cell.number_format = "# ##0"
            for idx, cell in enumerate(worksheet[1], start=1):
                max_len = len(str(cell.value or ""))
                for row_cell in worksheet[get_column_letter(idx)][1:]:
                    max_len = max(max_len, len(str(row_cell.value or "")))
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)
    output.seek(0)
    return output.getvalue()


@dataclass(frozen=True)
class SonuReport:
    data: pd.DataFrame
    period: str
    supplier: str
    bracelet_images: dict[str, bytes]


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = str(value).replace(" ", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0


def _clean_store(value: str) -> str:
    text = str(value or "").strip()
    compact = re.sub(r"\s+", "", text).upper()
    if compact.startswith("63"):
        return "63"
    for suffix in ("-RETAIL", "-TIMINGS", "-TIMING"):
        if compact.endswith(suffix):
            return text[: -len(suffix)].rstrip(" -")
    return text


def _extract_period(value: str) -> str:
    text = str(value or "")
    match = re.search(r"(\d{2}\.\d{2}\.\d{4})\s*[-–—]\s*(\d{2}\.\d{2}\.\d{4})", text)
    return f"{match.group(1)} – {match.group(2)}" if match else "Период не определён"


def _thumbnail(image_bytes: bytes, size: tuple[int, int] = (420, 320)) -> bytes:
    try:
        with Image.open(io.BytesIO(image_bytes)) as image:
            image = image.convert("RGB")
            image.thumbnail(size, Image.Resampling.LANCZOS)
            output = io.BytesIO()
            image.save(output, format="JPEG", quality=84, optimize=True)
            return output.getvalue()
    except Exception:
        return image_bytes


def parse_sonu_workbook(file_bytes: bytes) -> SonuReport:
    """Parse the final Sonu hierarchy with independent network stock in column L."""
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        sheet = workbook.active
        period = _extract_period(sheet.cell(1, 1).value)
        supplier_line = str(sheet.cell(2, 1).value or "").strip()
        supplier = supplier_line.split(":", 1)[-1].strip() if ":" in supplier_line else supplier_line
        if "SONU" not in supplier.upper():
            raise ValueError("Для раздела «Заказ Sonu» нужен отчет, отфильтрованный по поставщику Sonu.")

        hierarchy = str(sheet.cell(4, 1).value or "").upper()
        required = ("МАГАЗИН", "ТОВАР", "КАМЕНЬ", "ПРОБА", "НОМЕНКЛАТУРНАЯ ГРУППА")
        if not all(token in hierarchy for token in required):
            raise ValueError(
                "Неверная структура отчета. Нужны уровни: Магазин, Товар, Камень/вставка, "
                "Проба, Номенклатурная группа и Поставщик."
            )
        stock_header = str(sheet.cell(5, 12).value or "").strip().upper()
        if "ОСТАТОК" not in stock_header:
            raise ValueError(
                "В отчете отсутствует поле «Остаток». Загрузите новый формат Sonu, "
                "где в колонке L передается актуальный остаток SKU по всей сети."
            )

        rows: list[dict[str, Any]] = []
        row_to_index: dict[int, int] = {}
        context: dict[str, str] = {}
        for row_no in range(7, sheet.max_row + 1):
            value = sheet.cell(row_no, 1).value
            if value is None or str(value).strip() == "":
                continue
            level = int(sheet.row_dimensions[row_no].outlineLevel or 0)
            text = str(value).strip()
            if level == 0:
                context = {"store": _clean_store(text)}
            elif level == 1:
                context["division"] = text
            elif level == 2:
                context["metal_group"] = text
            elif level == 3:
                context["category"] = text
            elif level == 4:
                if not context.get("store") or not context.get("category"):
                    continue
                shipped = _number(sheet.cell(row_no, 4).value)
                sold = _number(sheet.cell(row_no, 8).value)
                sales = _number(sheet.cell(row_no, 9).value)
                network_stock = _number(sheet.cell(row_no, 12).value)
                item = {
                    "Магазин": context.get("store", ""),
                    "Раздел": context.get("division", ""),
                    "Металл": context.get("metal_group", ""),
                    "Категория": context.get("category", "Other"),
                    "Категория RU": CATEGORY_LABELS.get(
                        context.get("category", "Other"), context.get("category", "Other")
                    ),
                    "SKU": text,
                    "Камень": "Не указан",
                    "Проба": "Не указана",
                    "Отгружено": shipped,
                    "Скорость продаж": sold,
                    "Продажи VND": sales,
                    "Средняя цена VND": sales / sold if sold else _number(sheet.cell(row_no, 6).value),
                    "Остаток сети": network_stock,
                    "Фото": None,
                }
                rows.append(item)
                row_to_index[row_no] = len(rows) - 1
            elif level == 5 and rows:
                rows[-1]["Камень"] = text
            elif level == 6 and rows:
                rows[-1]["Проба"] = text

        if not rows:
            raise ValueError("В отчете не найдено товарных строк Sonu.")

        bracelet_rows = {
            row_no for row_no, idx in row_to_index.items() if rows[idx]["Категория"] == "Bracelet"
        }
        image_map: dict[str, bytes] = {}
        for image in getattr(sheet, "_images", []):
            try:
                anchor_row = int(image.anchor._from.row) + 1
                if anchor_row not in bracelet_rows:
                    continue
                idx = row_to_index.get(anchor_row)
                if idx is None:
                    continue
                sku = str(rows[idx]["SKU"]).strip().upper()
                data = _thumbnail(image._data())
                rows[idx]["Фото"] = data
                image_map.setdefault(sku, data)
            except Exception:
                continue

        frame = pd.DataFrame(rows)
        numeric_cols = [
            "Отгружено", "Скорость продаж", "Продажи VND", "Средняя цена VND", "Остаток сети"
        ]
        for column in numeric_cols:
            frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0).clip(lower=0)
        return SonuReport(frame, period, supplier or "Sonu", image_map)
    finally:
        workbook.close()
        gc.collect()


@st.cache_resource(show_spinner=False)
def _sonu_excel_lock() -> threading.Lock:
    return threading.Lock()


@st.cache_data(show_spinner=False, ttl=1800, max_entries=2)
def cached_parse_sonu(file_bytes: bytes) -> SonuReport:
    with _sonu_excel_lock():
        return parse_sonu_workbook(file_bytes)


def _usd(value: float, rate: float) -> float:
    return float(value) / float(rate) if rate else 0.0


def add_usd_columns(frame: pd.DataFrame, rate: float) -> pd.DataFrame:
    result = frame.copy()
    result["Продажи USD"] = result["Продажи VND"].map(lambda value: _usd(value, rate))
    result["Средняя цена USD"] = result["Средняя цена VND"].map(lambda value: _usd(value, rate))
    return result


def aggregate_sonu(frame: pd.DataFrame, group_columns: list[str], rate: float) -> pd.DataFrame:
    grouped = frame.groupby(group_columns, as_index=False, dropna=False).agg(
        Отгружено=("Отгружено", "sum"),
        **{
            "Скорость продаж": ("Скорость продаж", "sum"),
            "Продажи VND": ("Продажи VND", "sum"),
            "Моделей": ("SKU", "nunique"),
        },
    )
    grouped["Продажи USD"] = grouped["Продажи VND"] / rate
    grouped["Средняя цена USD"] = grouped["Продажи USD"] / grouped["Скорость продаж"].replace(0, pd.NA)
    grouped["Средняя цена USD"] = grouped["Средняя цена USD"].fillna(0)
    total_sold = float(grouped["Скорость продаж"].sum())
    grouped["Доля продаж"] = grouped["Скорость продаж"] / total_sold if total_sold else 0.0
    return grouped.drop(columns=["Продажи VND"])


def classify_bracelets(frame: pd.DataFrame, rate: float, period_days: int = 30) -> pd.DataFrame:
    """Classify bracelet SKUs using manual family decisions and the reviewed seed catalog."""
    grouped = network_sku_snapshot(frame, rate, period_days)
    if grouped.empty:
        return pd.DataFrame()
    grouped = grouped.loc[grouped["Категория"] == "Bracelet"].copy()
    if grouped.empty:
        return pd.DataFrame()

    grouped["_sku"] = grouped["SKU"].astype(str).str.upper().str.strip()
    grouped["Модельная семья"] = grouped["_sku"].map(_bracelet_model_key)
    grouped["Тип браслета"] = pd.NA
    grouped["Источник классификации"] = pd.NA

    manual_overrides = load_bracelet_overrides()
    manual_sku = {
        key: value for key, value in manual_overrides.items()
        if not key.startswith(BRACELET_FAMILY_KEY_PREFIX)
    }
    manual_family = {
        key[len(BRACELET_FAMILY_KEY_PREFIX):]: value
        for key, value in manual_overrides.items()
        if key.startswith(BRACELET_FAMILY_KEY_PREFIX)
    }

    # Human decisions are always first: exact SKU, then the whole model family.
    manual_types = grouped["_sku"].map(manual_sku)
    mask = manual_types.notna()
    grouped.loc[mask, "Тип браслета"] = manual_types.loc[mask]
    grouped.loc[mask, "Источник классификации"] = "Ручной выбор SKU"

    unresolved = grouped["Тип браслета"].isna()
    manual_family_types = grouped["Модельная семья"].map(manual_family)
    mask = unresolved & manual_family_types.notna()
    grouped.loc[mask, "Тип браслета"] = manual_family_types.loc[mask]
    grouped.loc[mask, "Источник классификации"] = "Ручной выбор семейства"

    catalog = load_bracelet_catalog()
    catalog_sku = catalog.get("sku_overrides", {})
    catalog_family = catalog.get("family_overrides", {})

    unresolved = grouped["Тип браслета"].isna()
    catalog_sku_types = grouped["_sku"].map(catalog_sku)
    mask = unresolved & catalog_sku_types.notna()
    grouped.loc[mask, "Тип браслета"] = catalog_sku_types.loc[mask]
    grouped.loc[mask, "Источник классификации"] = "Каталог SKU"

    unresolved = grouped["Тип браслета"].isna()
    catalog_family_types = grouped["Модельная семья"].map(catalog_family)
    mask = unresolved & catalog_family_types.notna()
    grouped.loc[mask, "Тип браслета"] = catalog_family_types.loc[mask]
    grouped.loc[mask, "Источник классификации"] = "Каталог семейства"

    # Legacy seeds remain as a safe fallback for older reports.
    unresolved = grouped["Тип браслета"].isna()
    exact_centered = unresolved & grouped["_sku"].isin(CENTERED_BRACELETS)
    exact_circle = unresolved & grouped["_sku"].isin(FULL_CIRCLE_BRACELETS)
    grouped.loc[exact_centered, "Тип браслета"] = CENTERED_BRACELET_LABEL
    grouped.loc[exact_centered, "Источник классификации"] = "Проверенная модель"
    grouped.loc[exact_circle, "Тип браслета"] = FULL_CIRCLE_BRACELET_LABEL
    grouped.loc[exact_circle, "Источник классификации"] = "Проверенная модель"

    unresolved = grouped["Тип браслета"].isna()
    family_centered = unresolved & grouped["Модельная семья"].isin(CENTERED_MODEL_KEYS)
    family_circle = unresolved & grouped["Модельная семья"].isin(FULL_CIRCLE_MODEL_KEYS)
    grouped.loc[family_centered, "Тип браслета"] = CENTERED_BRACELET_LABEL
    grouped.loc[family_centered, "Источник классификации"] = "Семейство модели"
    grouped.loc[family_circle, "Тип браслета"] = FULL_CIRCLE_BRACELET_LABEL
    grouped.loc[family_circle, "Источник классификации"] = "Семейство модели"

    # No hidden 50/50 split: genuinely borderline families stay pending until reviewed.
    unresolved = grouped["Тип браслета"].isna()
    grouped.loc[unresolved, "Источник классификации"] = "Требует классификации"

    type_order = {name: index for index, name in enumerate(BRACELET_TYPE_ORDER)}
    grouped["_type_order"] = grouped["Тип браслета"].map(type_order).fillna(99)
    return grouped.sort_values(
        ["_type_order", "Продано за период", "Продажи USD"], ascending=[True, False, False]
    ).drop(columns=["_sku", "_type_order"]).reset_index(drop=True)


def bracelet_classification_audit(frame: pd.DataFrame, rate: float, period_days: int = 30) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return transparent SKU-level results and family-level pending counts."""
    bracelets = classify_bracelets(frame, rate, period_days)
    summary_columns = [
        "Статус классификации", "Моделей", "Семейств",
        "С затяжкой", "Без затяжки (в круг)",
    ]
    detail_columns = [
        "SKU", "Модельная семья", "Камень группы", "Продано за период",
        "Продажи USD", "Остаток сети", "Тип браслета",
        "Источник классификации", "Статус классификации", "Пояснение",
    ]
    if bracelets.empty:
        return pd.DataFrame(columns=summary_columns), pd.DataFrame(columns=detail_columns)

    detail = bracelets.copy()
    detail["Статус классификации"] = detail["Источник классификации"].map({
        "Ручной выбор SKU": "Разобрано вручную",
        "Ручной выбор семейства": "Разобрано вручную",
        "Каталог SKU": "В справочнике",
        "Каталог семейства": "В справочнике",
        "Проверенная модель": "В справочнике",
        "Семейство модели": "В справочнике",
        "Требует классификации": "Требует классификации",
    }).fillna("Не определено")

    def explanation(row: pd.Series) -> str:
        source = row.get("Источник классификации")
        if source == "Ручной выбор SKU":
            return "Тип выбран вручную только для этого SKU."
        if source == "Ручной выбор семейства":
            return "Тип выбран вручную для модельной семьи и применён ко всем связанным SKU."
        if source in {"Каталог SKU", "Каталог семейства"}:
            return "Модель предварительно проверена по фотографии sonunew.xlsx и загружена в справочник."
        if source == "Проверенная модель":
            return "Модель есть в ранее проверенном справочнике."
        if source == "Семейство модели":
            return "Классификация унаследована от ранее проверенной модельной семьи."
        if source == "Требует классификации":
            return "Пограничная конструкция оставлена для ручной проверки по фотографии."
        return "Тип не определён."

    detail["Пояснение"] = detail.apply(explanation, axis=1)

    rows = []
    for status in ["В справочнике", "Разобрано вручную", "Требует классификации", "Не определено"]:
        current = detail.loc[detail["Статус классификации"] == status]
        if current.empty:
            continue
        rows.append({
            "Статус классификации": status,
            "Моделей": int(current["SKU"].nunique()),
            "Семейств": int(current["Модельная семья"].nunique()),
            "С затяжкой": int(current.loc[current["Тип браслета"] == CENTERED_BRACELET_LABEL, "SKU"].nunique()),
            "Без затяжки (в круг)": int(current.loc[current["Тип браслета"] == FULL_CIRCLE_BRACELET_LABEL, "SKU"].nunique()),
        })
    summary = pd.DataFrame(rows, columns=summary_columns)
    available = [column for column in detail_columns if column in detail.columns]
    detail = detail[available].sort_values(
        ["Статус классификации", "Продано за период", "Продажи USD", "SKU"],
        ascending=[True, False, False, True],
    ).reset_index(drop=True)
    return summary, detail


def _bracelet_review_rows(
    frame: pd.DataFrame,
    rate: float,
    period_days: int,
    *,
    mode: str,
) -> pd.DataFrame:
    """One review card per model family, never one card per stone variation."""
    bracelets = classify_bracelets(frame, rate, period_days)
    if bracelets.empty:
        return bracelets
    if mode == "manual":
        source_mask = bracelets["Источник классификации"].isin(
            ["Ручной выбор SKU", "Ручной выбор семейства"]
        )
    else:
        source_mask = bracelets["Источник классификации"] == "Требует классификации"
    current = bracelets.loc[source_mask].copy()
    if current.empty:
        return pd.DataFrame()

    rows: list[dict[str, Any]] = []
    for family, group in current.groupby("Модельная семья", dropna=False):
        candidates = group.copy()
        candidates["_has_photo"] = candidates["Фото"].map(
            lambda value: isinstance(value, (bytes, bytearray))
        )
        candidates = candidates.sort_values(
            ["_has_photo", "Продано за период", "Продажи USD"],
            ascending=[False, False, False],
        )
        representative = candidates.iloc[0]
        sku_values = sorted(group["SKU"].astype(str).str.upper().str.strip().unique().tolist())
        stone_values = sorted(
            value for value in group["Камень группы"].dropna().astype(str).unique().tolist() if value
        )
        selected_types = group["Тип браслета"].dropna().astype(str).unique().tolist()
        rows.append({
            "Модельная семья": str(family or "").strip().upper(),
            "SKU": str(representative.get("SKU", "")).strip().upper(),
            "SKU семьи": sku_values,
            "Моделей в семье": len(sku_values),
            "Камень группы": ", ".join(stone_values[:4]) + ("…" if len(stone_values) > 4 else ""),
            "Продано за период": float(group["Продано за период"].sum()),
            "Продажи USD": float(group["Продажи USD"].sum()),
            "Остаток сети": float(group["Остаток сети"].sum()),
            "Тип браслета": selected_types[0] if len(selected_types) == 1 else pd.NA,
            "Фото": representative.get("Фото"),
            "Источник классификации": representative.get("Источник классификации", ""),
        })
    return pd.DataFrame(rows).sort_values(
        ["Продано за период", "Продажи USD", "Модельная семья"],
        ascending=[False, False, True],
    ).reset_index(drop=True)


def _open_bracelet_review(mode: str) -> None:
    st.session_state[BRACELET_REVIEW_OPEN_KEY] = True
    st.session_state[BRACELET_REVIEW_MODE_KEY] = mode
    st.session_state[BRACELET_REVIEW_INDEX_KEY] = 0
    st.session_state[BRACELET_REVIEW_DRAFT_KEY] = (
        load_bracelet_overrides() if mode == "manual" else {}
    )


def _rerun_bracelet_dialog() -> None:
    try:
        st.rerun(scope="fragment")
    except Exception:
        st.rerun()


def _close_bracelet_review() -> None:
    st.session_state[BRACELET_REVIEW_OPEN_KEY] = False
    st.session_state[BRACELET_REVIEW_INDEX_KEY] = 0
    st.session_state[BRACELET_REVIEW_DRAFT_KEY] = {}


@st.dialog("Разобрать модели, требующие классификации", width="large")
def _bracelet_review_dialog(frame: pd.DataFrame, rate: float, period_days: int) -> None:
    mode = str(st.session_state.get(BRACELET_REVIEW_MODE_KEY, "pending"))
    rows = _bracelet_review_rows(frame, rate, period_days, mode=mode)
    if rows.empty:
        st.success(
            "Все модельные семьи классифицированы."
            if mode == "pending"
            else "Сохранённых ручных решений для текущего отчёта нет."
        )
        if st.button("Закрыть", width="stretch", key="bracelet_review_empty_close"):
            _close_bracelet_review()
            st.rerun()
        return

    index = min(max(int(st.session_state.get(BRACELET_REVIEW_INDEX_KEY, 0)), 0), len(rows) - 1)
    st.session_state[BRACELET_REVIEW_INDEX_KEY] = index
    row = rows.iloc[index]
    family = str(row.get("Модельная семья", "")).strip().upper()
    decision_key = _bracelet_family_override_key(family)
    sku = str(row.get("SKU", "")).strip().upper()
    family_skus = list(row.get("SKU семьи", []) or [])
    draft = dict(st.session_state.get(BRACELET_REVIEW_DRAFT_KEY, {}))
    saved = load_bracelet_overrides()
    if decision_key not in draft and mode == "manual" and decision_key in saved:
        draft[decision_key] = saved[decision_key]
        st.session_state[BRACELET_REVIEW_DRAFT_KEY] = draft
    selected = _normalize_bracelet_override(draft.get(decision_key))

    decision_keys = [_bracelet_family_override_key(value) for value in rows["Модельная семья"]]
    reviewed = sum(1 for key in decision_keys if key in draft)
    st.progress(
        reviewed / len(rows),
        text=f"Разобрано {reviewed} из {len(rows)} семей · семья {index + 1} из {len(rows)}",
    )

    image_column, details_column = st.columns([1.45, 1], vertical_alignment="top")
    with image_column:
        image = row.get("Фото")
        if isinstance(image, (bytes, bytearray)):
            st.image(image, width="stretch", caption=f"{family} · пример {sku}")
        else:
            st.markdown(
                '<div class="sonu-review-image-placeholder">'
                '<strong>Фотография не найдена</strong><span>Проверьте семейство по SKU</span></div>',
                unsafe_allow_html=True,
            )
    with details_column:
        st.markdown(f"### {escape(family)}")
        st.caption(
            f"Пример: {escape(sku)} · в семье {int(row.get('Моделей в семье', 1))} SKU. "
            "Решение применится ко всей модельной семье."
        )
        catalog_pending = load_bracelet_catalog().get("pending_families", {})
        pending_meta = catalog_pending.get(family, {}) if isinstance(catalog_pending, dict) else {}
        reason = str(pending_meta.get("reason", "Пограничная конструкция требует визуальной проверки."))
        st.info(reason)
        m1, m2 = st.columns(2)
        m1.metric("Продано", f"{_money(row.get('Продано за период', 0))} шт.")
        m2.metric("Остаток сети", f"{_money(row.get('Остаток сети', 0))} шт.")
        st.metric("Продажи", f"${_money(row.get('Продажи USD', 0))}")
        with st.expander(f"SKU модельной семьи · {len(family_skus)}"):
            st.write("\n".join(f"• {item}" for item in family_skus))
        st.markdown("**Выберите фактический тип браслета:**")

        if st.button(
            "С затяжкой / центральная композиция",
            width="stretch",
            type="primary" if selected == CENTERED_BRACELET_LABEL else "secondary",
            key=f"bracelet_choice_centered::{family}",
        ):
            draft[decision_key] = CENTERED_BRACELET_LABEL
            st.session_state[BRACELET_REVIEW_DRAFT_KEY] = draft
            _rerun_bracelet_dialog()
        if st.button(
            "Без затяжки / полный круг",
            width="stretch",
            type="primary" if selected == FULL_CIRCLE_BRACELET_LABEL else "secondary",
            key=f"bracelet_choice_circle::{family}",
        ):
            draft[decision_key] = FULL_CIRCLE_BRACELET_LABEL
            st.session_state[BRACELET_REVIEW_DRAFT_KEY] = draft
            _rerun_bracelet_dialog()
        if selected:
            st.success(f"Выбрано для всей семьи: {selected}")
        else:
            st.warning("Сделайте выбор, чтобы перейти к следующей семье.")

    previous_col, next_col = st.columns(2)
    with previous_col:
        if st.button(
            "← Предыдущая",
            width="stretch",
            disabled=index == 0,
            key=f"bracelet_review_previous::{index}",
        ):
            st.session_state[BRACELET_REVIEW_INDEX_KEY] = index - 1
            _rerun_bracelet_dialog()
    with next_col:
        if st.button(
            "Следующая →" if index < len(rows) - 1 else "Перейти к сохранению",
            width="stretch",
            disabled=selected is None,
            key=f"bracelet_review_next::{index}",
        ):
            if index < len(rows) - 1:
                st.session_state[BRACELET_REVIEW_INDEX_KEY] = index + 1
            _rerun_bracelet_dialog()

    st.divider()
    missing = [key for key in decision_keys if key not in draft]
    action_col, cancel_col = st.columns([1.35, 0.65])
    with action_col:
        if st.button(
            f"Сохранить решения ({len(decision_keys) - len(missing)}/{len(decision_keys)})",
            width="stretch",
            type="primary",
            disabled=bool(missing),
            key="bracelet_review_save_all",
        ):
            _, persisted, message = save_bracelet_overrides(
                {key: draft[key] for key in decision_keys}
            )
            st.session_state["bracelet_review_flash"] = (
                "success" if persisted else "warning", message
            )
            _close_bracelet_review()
            st.rerun()
    with cancel_col:
        if st.button("Закрыть без сохранения", width="stretch", key="bracelet_review_cancel"):
            _close_bracelet_review()
            st.rerun()


def _render_bracelet_classification_audit(frame: pd.DataFrame, rate: float, period_days: int) -> None:
    """Render catalog coverage and family-level guided review."""
    summary, detail = bracelet_classification_audit(frame, rate, period_days)
    st.markdown("## Классификация браслетов")
    st.caption(
        "Основной справочник сформирован по фотографиям sonunew.xlsx. "
        "Пограничные конструкции разбираются по модельным семьям: одно решение применяется ко всем каменным вариантам."
    )
    flash = st.session_state.pop("bracelet_review_flash", None)
    if flash:
        tone, message = flash
        getattr(st, tone, st.info)(message)
    if detail.empty:
        st.info("В отчёте нет браслетов для классификации.")
        return

    total = int(detail["SKU"].nunique())
    catalogued = int(detail.loc[detail["Статус классификации"] == "В справочнике", "SKU"].nunique())
    manual = detail.loc[detail["Статус классификации"] == "Разобрано вручную"].copy()
    pending = detail.loc[detail["Статус классификации"] == "Требует классификации"].copy()
    pending_sku = int(pending["SKU"].nunique())
    pending_families = int(pending["Модельная семья"].nunique())
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Всего SKU", total)
    c2.metric("В справочнике", catalogued)
    c3.metric("Разобрано вручную", int(manual["SKU"].nunique()))
    c4.metric("Требуют классификации", pending_sku)
    c5.metric("Семейств в разборе", pending_families)

    _table(summary, "sonu_bracelet_classification_summary")
    st.markdown("### Требуют классификации")
    if pending.empty:
        st.success("Все модельные семьи текущего отчёта классифицированы.")
    else:
        st.info(
            f"Осталось {pending_families} модельных семей ({pending_sku} SKU). "
            "В основном отчёте они временно не попадают ни в «С затяжкой», ни в «Без затяжки», "
            "пока вы не подтвердите тип по фотографии."
        )
        if st.button(
            f"Разобрать модели · {pending_families} семей / {pending_sku} SKU",
            type="primary",
            width="stretch",
            key="sonu_open_pending_bracelet_review",
        ):
            _open_bracelet_review("pending")
            st.rerun()

    if not manual.empty:
        manual_family_count = int(manual["Модельная семья"].nunique())
        if st.button(
            f"Изменить ручные решения · {manual_family_count} семей",
            width="stretch",
            key="sonu_edit_manual_bracelet_review",
        ):
            _open_bracelet_review("manual")
            st.rerun()

    with st.expander("Сохранение и резервная копия классификации"):
        st.caption(
            "Ручные решения сохраняются на уровне модельной семьи в "
            "data/bracelet_classification_overrides.json и имеют приоритет над встроенным каталогом. "
            "На Streamlit Cloud после нового деплоя локальная запись может быть сброшена — "
            "скачайте JSON и добавьте его в репозиторий."
        )
        overrides = load_bracelet_overrides()
        st.download_button(
            "Скачать резервную копию решений",
            data=bracelet_overrides_json(overrides),
            file_name="bracelet_classification_overrides.json",
            mime="application/json",
            width="stretch",
            key="sonu_download_bracelet_overrides",
            disabled=not overrides,
        )
        imported = st.file_uploader(
            "Загрузить сохранённые решения",
            type=["json"],
            accept_multiple_files=False,
            key="sonu_import_bracelet_overrides",
        )
        if imported is not None:
            import_marker = (imported.name, len(imported.getvalue()))
            if st.session_state.get("sonu_import_bracelet_overrides_marker") != import_marker:
                try:
                    imported_values, persisted, message = import_bracelet_overrides(imported.getvalue())
                except ValueError as exc:
                    st.error(str(exc))
                else:
                    st.session_state["sonu_import_bracelet_overrides_marker"] = import_marker
                    st.success(f"Импортировано решений: {len(imported_values)}. {message}")
                    st.rerun()

    with st.expander("Показать полную классификацию всех браслетов"):
        _table(detail, "sonu_all_bracelet_classification")

    if st.session_state.get(BRACELET_REVIEW_OPEN_KEY):
        _bracelet_review_dialog(frame, rate, period_days)


def bracelet_type_summary(bracelets: pd.DataFrame, period_days: int = 30) -> pd.DataFrame:
    """Comparison-ready summary for classified tightening and full-circle bracelets."""
    if bracelets.empty:
        return pd.DataFrame()
    bracelets = bracelets.loc[bracelets["Тип браслета"].isin(BRACELET_TYPE_ORDER)].copy()
    if bracelets.empty:
        return pd.DataFrame()
    result = _aggregate_sales_stock(bracelets, ["Тип браслета"], period_days)
    order = {name: index for index, name in enumerate(BRACELET_TYPE_ORDER)}
    result["_order"] = result["Тип браслета"].map(order).fillna(99)
    return result.sort_values("_order").drop(columns="_order").reset_index(drop=True)


def bracelet_stone_summary(bracelets: pd.DataFrame, period_days: int = 30) -> pd.DataFrame:
    """Stone summary inside each classified bracelet construction type."""
    if bracelets.empty:
        return pd.DataFrame()
    bracelets = bracelets.loc[bracelets["Тип браслета"].isin(BRACELET_TYPE_ORDER)].copy()
    if bracelets.empty:
        return pd.DataFrame()
    result = _aggregate_sales_stock(
        bracelets,
        ["Тип браслета", "Группа камня", "Камень группы"],
        period_days,
    )
    type_order = {name: index for index, name in enumerate(BRACELET_TYPE_ORDER)}
    result["_type_order"] = result["Тип браслета"].map(type_order).fillna(99)
    result = _business_sort(result, include_category=False)
    result["_type_order"] = result["Тип браслета"].map(type_order).fillna(99)
    return result.sort_values(
        ["_type_order", "Продано штук", "Продано на сумму, USD"],
        ascending=[True, False, False],
    ).drop(columns="_type_order").reset_index(drop=True)


SONU_CSS = """
<style>
.sonu-view-note { margin:10px 0 16px; color:#71685e; font-size:13px; line-height:1.5; }
.sonu-ai-brief { margin:14px 0 18px; padding:20px 22px; border-radius:18px; border:1px solid #dcc69e; background:linear-gradient(135deg,#fffaf0,#f5e5c2); box-shadow:0 12px 30px rgba(90,57,12,.08); }
.sonu-ai-label { color:#9a681e; font-size:11px; font-weight:850; letter-spacing:.11em; text-transform:uppercase; }
.sonu-ai-headline { color:#21180e; font-size:19px; font-weight:800; line-height:1.45; margin-top:8px; }
.sonu-ai-text { color:#65533d; font-size:14px; line-height:1.6; margin-top:8px; }
.sonu-data-card {
  border:1px solid #e6dccb; border-radius:17px; padding:16px 17px; margin:0 0 12px;
  background:linear-gradient(145deg,#fff,#fbf7f0); box-shadow:0 9px 24px rgba(38,27,12,.055);
  min-height:182px;
}
.sonu-data-card .eyebrow { color:#a77429; font-size:11px; font-weight:800; letter-spacing:.1em; text-transform:uppercase; }
.sonu-data-card .title { color:#17130f; font-size:18px; font-weight:800; line-height:1.25; margin:6px 0 5px; }
.sonu-data-card .meta { color:#71685e; font-size:12px; line-height:1.45; min-height:34px; }
.sonu-card-metrics { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; margin-top:13px; }
.sonu-card-metric { border-radius:11px; padding:9px 10px; background:rgba(183,137,63,.09); border:1px solid rgba(183,137,63,.16); }
.sonu-card-metric span { display:block; color:#857765; font-size:10px; text-transform:uppercase; letter-spacing:.05em; }
.sonu-card-metric strong { display:block; color:#2a2117; font-size:16px; margin-top:3px; }
.sonu-card-needs { margin-top:10px; padding-top:10px; border-top:1px solid #ece4d8; color:#6f6251; font-size:12px; line-height:1.55; }
.sonu-card-needs b { color:#9a681e; }
.sonu-review-image-placeholder { min-height:420px; border:1px dashed #d8c7aa; border-radius:18px; background:linear-gradient(145deg,#fffaf2,#f4ead8); display:flex; flex-direction:column; gap:8px; align-items:center; justify-content:center; color:#6f6251; text-align:center; padding:24px; }
.sonu-review-image-placeholder strong { color:#2a2117; font-size:18px; }
.sonu-review-image-placeholder span { color:#8a8073; font-size:13px; }
.sonu-card-image-placeholder { min-height:150px; border:1px dashed #d8c7aa; border-radius:14px; background:#faf7f2; display:flex; align-items:center; justify-content:center; color:#8a8073; margin-bottom:8px; }
@media (max-width:900px) { .sonu-data-card { min-height:160px; } }
@media (max-width:640px) {
  .sonu-review-image-placeholder { min-height:280px; }
  .sonu-data-card { min-height:auto; padding:14px; }
  .sonu-card-metrics { grid-template-columns:repeat(2,minmax(0,1fr)); }
}
</style>
"""


def _render_sonu_css() -> None:
    st.markdown(SONU_CSS, unsafe_allow_html=True)


def _metric_value(value: Any, *, money: bool = False) -> str:
    numeric = _number(value)
    return f"${_money(numeric)}" if money else _money(numeric)


def _card_html(
    *,
    eyebrow: str,
    title: str,
    meta: str,
    metrics: list[tuple[str, str]],
    needs: tuple[Any, Any, Any] | None = None,
) -> str:
    metric_html = "".join(
        f'<div class="sonu-card-metric"><span>{escape(label)}</span><strong>{escape(value)}</strong></div>'
        for label, value in metrics
    )
    needs_html = ""
    if needs is not None:
        needs_html = (
            '<div class="sonu-card-needs"><b>Потребность:</b> '
            f'30 дней — {escape(_money(needs[0]))} · '
            f'45 дней — {escape(_money(needs[1]))} · '
            f'90 дней — {escape(_money(needs[2]))}</div>'
        )
    return (
        '<div class="sonu-data-card">'
        f'<div class="eyebrow">{escape(eyebrow)}</div>'
        f'<div class="title">{escape(title)}</div>'
        f'<div class="meta">{escape(meta)}</div>'
        f'<div class="sonu-card-metrics">{metric_html}</div>'
        f'{needs_html}</div>'
    )


def _paged_frame(frame: pd.DataFrame, key: str, default_size: int = 12) -> pd.DataFrame:
    if frame.empty or len(frame) <= 18:
        return frame
    page_size = st.segmented_control(
        "Карточек на странице", [6, 12, 18], default=default_size, key=f"{key}_page_size"
    ) or default_size
    page_count = max(1, (len(frame) + int(page_size) - 1) // int(page_size))
    page = st.number_input("Страница", 1, page_count, 1, key=f"{key}_page")
    start = (int(page) - 1) * int(page_size)
    return frame.iloc[start:start + int(page_size)]


def _render_assortment_cards(frame: pd.DataFrame, key: str) -> None:
    if frame.empty:
        st.info("По выбранным фильтрам данных нет.")
        return
    current = _paged_frame(frame, key)
    for start in range(0, len(current), 3):
        columns = st.columns(3)
        for column, (_, row) in zip(columns, current.iloc[start:start + 3].iterrows()):
            with column:
                st.markdown(
                    _card_html(
                        eyebrow=str(row.get("Группа камня", "")),
                        title=f'{row.get("Камень группы", "")} · {row.get("Категория RU", "")}',
                        meta="Остаток и продажи суммированы по сети без дублей по магазинам.",
                        metrics=[
                            ("Продано SKU", _metric_value(row.get("Продано уникальных SKU"))),
                            ("Продано", f'{_metric_value(row.get("Продано штук"))} шт.'),
                            ("Продажи", _metric_value(row.get("Продано на сумму, USD"), money=True)),
                            ("SKU на остатке", _metric_value(row.get("Осталось уникальных SKU"))),
                            ("Остаток", f'{_metric_value(row.get("Всего штук"))} шт.'),
                        ],
                    ),
                    unsafe_allow_html=True,
                )


def _render_model_cards(frame: pd.DataFrame, key: str) -> None:
    if frame.empty:
        st.info("По выбранным фильтрам моделей нет.")
        return
    current = _paged_frame(frame, key)
    for start in range(0, len(current), 3):
        columns = st.columns(3)
        for column, (_, row) in zip(columns, current.iloc[start:start + 3].iterrows()):
            with column:
                image = row.get("Фото")
                if isinstance(image, (bytes, bytearray)):
                    st.image(image, width="stretch")
                elif "Фото" in current.columns:
                    st.markdown('<div class="sonu-card-image-placeholder">Нет фотографии</div>', unsafe_allow_html=True)
                bracelet_type = str(row.get("Тип браслета", "") or "").strip()
                meta_parts = [str(row.get("Категория RU", "")), str(row.get("Камень группы", ""))]
                if bracelet_type and bracelet_type.lower() != "nan":
                    meta_parts.insert(0, bracelet_type)
                meta = " · ".join(part for part in meta_parts if part and part.lower() != "nan")
                st.markdown(
                    _card_html(
                        eyebrow=str(row.get("Группа камня", "")),
                        title=str(row.get("SKU", "")),
                        meta=meta,
                        metrics=[
                            ("Остаток", f'{_metric_value(row.get("Остаток сети"))} шт.'),
                            ("Продано", f'{_metric_value(row.get("Продано за период"))} шт.'),
                            ("Магазинов", _metric_value(row.get("Магазинов с продажами"))),
                            ("Продажи", _metric_value(row.get("Продажи USD"), money=True)),
                        ],
                        needs=(
                            row.get("Нужно на 30 дней", 0),
                            row.get("Нужно на 45 дней", 0),
                            row.get("Нужно на 90 дней", 0),
                        ),
                    ),
                    unsafe_allow_html=True,
                )


def _render_simple_cards(
    frame: pd.DataFrame,
    key: str,
    *,
    eyebrow_column: str,
    title_column: str,
    meta_builder,
    metrics_builder,
    needs: bool = False,
) -> None:
    if frame.empty:
        st.info("По выбранным фильтрам данных нет.")
        return
    current = _paged_frame(frame, key)
    for start in range(0, len(current), 3):
        columns = st.columns(3)
        for column, (_, row) in zip(columns, current.iloc[start:start + 3].iterrows()):
            with column:
                need_values = None
                if needs:
                    need_values = (
                        row.get("Нужно на 30 дней", 0),
                        row.get("Нужно на 45 дней", 0),
                        row.get("Нужно на 90 дней", 0),
                    )
                st.markdown(
                    _card_html(
                        eyebrow=str(row.get(eyebrow_column, "")),
                        title=str(row.get(title_column, "")),
                        meta=str(meta_builder(row)),
                        metrics=metrics_builder(row),
                        needs=need_values,
                    ),
                    unsafe_allow_html=True,
                )


def _money(value: float, digits: int = 0) -> str:
    return f"{float(value):,.{digits}f}".replace(",", " ")


def _kpi(label: str, value: str, note: str = "") -> None:
    st.markdown(
        f"""
        <div style="border:1px solid #e8e0d4;border-radius:16px;background:#fff;padding:18px;min-height:115px;box-shadow:0 8px 24px rgba(38,27,12,.05)">
          <div style="font-size:12px;text-transform:uppercase;letter-spacing:.08em;color:#8b6a36;font-weight:700">{label}</div>
          <div style="font-size:25px;font-weight:750;color:#17130f;margin-top:8px">{value}</div>
          <div style="font-size:12px;color:#777;margin-top:5px">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _locked_chart(figure: go.Figure, key: str) -> None:
    figure.update_layout(
        dragmode=False,
        clickmode="event",
        hovermode="closest",
        legend_itemclick=False,
        legend_itemdoubleclick=False,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Arial", color="#1c1813"),
    )
    figure.update_xaxes(fixedrange=True, automargin=True)
    figure.update_yaxes(fixedrange=True, automargin=True)
    st.plotly_chart(figure, width="stretch", key=key, config=LOCKED_CHART_CONFIG)


def _horizontal_chart(frame: pd.DataFrame, label: str, metric: str, title: str, suffix: str = "", prefix: str = "") -> go.Figure:
    data = frame.loc[frame[metric] > 0].sort_values(metric, ascending=True).copy()
    labels = [f"{prefix}{_money(value)}{suffix}" for value in data[metric]]
    maximum = float(data[metric].max()) if not data.empty else 0.0
    fig = go.Figure(
        go.Bar(
            x=data[metric],
            y=data[label],
            orientation="h",
            marker_color="#b7893f",
            text=labels,
            textposition="outside",
            cliponaxis=False,
            hovertemplate="%{y}<br>%{text}<extra></extra>",
        )
    )
    fig.update_layout(
        title=title,
        height=max(340, len(data) * 40 + 100),
        margin=dict(l=20, r=130, t=55, b=35),
        xaxis=dict(range=[0, maximum * 1.3] if maximum else None, gridcolor="#ece8e1"),
        yaxis=dict(title=""),
    )
    return fig


def _user_facing_stone_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Use clear business terminology without changing internal calculations."""
    return frame.rename(columns={"Камень группы": "Вид камня"})


def _is_percentage_column(column: str) -> bool:
    name = str(column).strip().lower()
    return "%" in name or name.startswith("доля ") or "доля " in name


def _is_money_column(column: str) -> bool:
    """Return True for every user-facing USD field, including Total and averages."""
    return "USD" in str(column).upper()


def _rounded_sonu_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Round displayed Sonu measures while keeping percentages and day coverage precise."""
    result = frame.copy()
    for column in result.columns:
        if column == "Фото" or _is_percentage_column(column) or column == "Покрытие остатком, дней":
            continue
        if pd.api.types.is_numeric_dtype(result[column]):
            result[column] = pd.to_numeric(result[column], errors="coerce").round(0)
    return result


def _table(frame: pd.DataFrame, key: str) -> None:
    frame = _rounded_sonu_frame(_user_facing_stone_columns(frame))
    config: dict[str, Any] = {}
    for column in frame.columns:
        if column == "Фото":
            config[column] = st.column_config.ImageColumn("Фото", width="small")
        elif _is_percentage_column(column):
            config[column] = st.column_config.NumberColumn(format="percent")
        elif column == "Покрытие остатком, дней":
            config[column] = st.column_config.NumberColumn(format="%.1f")
        elif _is_money_column(column):
            # The header already states USD. Localized + step=1 removes cents and
            # uses the browser's three-digit thousands separator (spaces in RU locale).
            config[column] = st.column_config.NumberColumn(format="localized", step=1)
        elif pd.api.types.is_numeric_dtype(frame[column]):
            config[column] = st.column_config.NumberColumn(format="localized", step=1)
    st.dataframe(frame, width="stretch", hide_index=True, key=key, column_config=config)


def _persist_upload(uploaded_file) -> bytes | None:
    if uploaded_file is not None:
        current = bytes(uploaded_file.getvalue())
        previous = st.session_state.get("sonu_report_bytes")
        if previous != current:
            st.session_state["sonu_report_bytes"] = current
            st.session_state["sonu_report_name"] = uploaded_file.name
    return st.session_state.get("sonu_report_bytes")


def _period_days(period: str) -> int:
    match = re.search(r"(\d{2}\.\d{2}\.\d{4})\s*[–—-]\s*(\d{2}\.\d{2}\.\d{4})", str(period))
    if not match:
        return 30
    try:
        start = datetime.strptime(match.group(1), "%d.%m.%Y")
        end = datetime.strptime(match.group(2), "%d.%m.%Y")
    except ValueError:
        return 30
    return max((end - start).days + 1, 1)


def _render_store_section(frame: pd.DataFrame, rate: float, view_mode: str) -> None:
    summary = aggregate_sonu(frame, ["Магазин"], rate).sort_values("Скорость продаж", ascending=False)
    st.markdown("### Продажи по магазинам")
    st.caption("Остаток здесь не распределяется по магазинам: в исходном файле он уже общий по сети.")
    left, right = st.columns(2)
    with left:
        _locked_chart(
            _horizontal_chart(summary, "Магазин", "Скорость продаж", "Продано за отчетный период", " шт."),
            "sonu_store_qty",
        )
    with right:
        _locked_chart(
            _horizontal_chart(summary, "Магазин", "Продажи USD", "Продажи по магазинам", " $"),
            "sonu_store_sales",
        )
    if view_mode == "Таблица":
        _table(summary, "sonu_store_table")
    else:
        _render_simple_cards(
            summary,
            "sonu_store_cards",
            eyebrow_column="Магазин",
            title_column="Магазин",
            meta_builder=lambda row: "Фактические продажи за выбранный период",
            metrics_builder=lambda row: [
                ("Моделей", _metric_value(row.get("Моделей"))),
                ("Продано", f'{_metric_value(row.get("Скорость продаж"))} шт.'),
                ("Продажи", _metric_value(row.get("Продажи USD"), money=True)),
                ("Средняя цена", _metric_value(row.get("Средняя цена USD"), money=True)),
            ],
        )



SUMMARY_SORT_OPTIONS = {
    "Продано штук": "Продано штук",
    "Продано на сумму": "Продано на сумму, USD",
    "Продано уникальных SKU": "Продано уникальных SKU",
    "Осталось уникальных SKU": "Осталось уникальных SKU",
    "Всего штук на остатке": "Всего штук",
    "Группа камня": "Группа камня",
    "Вид камня": "Камень группы",
    "Номенклатурная группа": "Номенклатурная группа",
}


def _summary_controls(frame: pd.DataFrame, key: str, *, allow_category: bool = True) -> tuple[pd.DataFrame, str]:
    """Table/card switch plus user-configurable sorting for a business summary."""
    c1, c2, c3 = st.columns([1, 1.35, 0.85])
    with c1:
        view = st.segmented_control(
            "Вид представления", ["Таблица", "Карточки"], default="Таблица", key=f"{key}_view"
        ) or "Таблица"
    options = SUMMARY_SORT_OPTIONS.copy()
    if not allow_category:
        options.pop("Номенклатурная группа", None)
    with c2:
        sort_label = st.selectbox(
            "Сортировать по", list(options), index=0, key=f"{key}_sort"
        )
    with c3:
        direction = st.segmented_control(
            "Порядок", ["По убыванию", "По возрастанию"],
            default="По убыванию", key=f"{key}_direction"
        ) or "По убыванию"
    column = options[sort_label]
    return frame.sort_values(
        column,
        ascending=(direction == "По возрастанию"),
        na_position="last",
    ).reset_index(drop=True), view


def _render_business_cards(frame: pd.DataFrame, key: str, *, show_category: bool) -> None:
    if frame.empty:
        st.info("По выбранным условиям данных нет.")
        return
    current = _paged_frame(frame, key)
    for start in range(0, len(current), 3):
        columns = st.columns(3)
        for column, (_, row) in zip(columns, current.iloc[start:start + 3].iterrows()):
            with column:
                title = str(row.get("Камень группы", ""))
                meta = str(row.get("Номенклатурная группа", "")) if show_category else "Сводка по виду камня"
                st.markdown(
                    _card_html(
                        eyebrow=str(row.get("Группа камня", "")),
                        title=title,
                        meta=meta,
                        metrics=[
                            ("Продано SKU", _metric_value(row.get("Продано уникальных SKU"))),
                            ("Продано", f'{_metric_value(row.get("Продано штук"))} шт.'),
                            ("Продажи", _metric_value(row.get("Продано на сумму, USD"), money=True)),
                            ("SKU на остатке", _metric_value(row.get("Осталось уникальных SKU"))),
                            ("Остаток", f'{_metric_value(row.get("Всего штук"))} шт.'),
                        ],
                    ),
                    unsafe_allow_html=True,
                )


def _render_business_summary(frame: pd.DataFrame, key: str, *, show_category: bool) -> None:
    sorted_frame, view = _summary_controls(frame, key, allow_category=show_category)
    display_columns = ["Группа камня", "Камень группы"]
    if show_category:
        display_columns.append("Номенклатурная группа")
    display_columns.extend([
        "Продано уникальных SKU", "Продано штук", "Продано на сумму, USD",
        "Осталось уникальных SKU", "Всего штук",
    ])
    if view == "Таблица":
        _table(sorted_frame[display_columns], f"{key}_table")
    else:
        _render_business_cards(sorted_frame, f"{key}_cards", show_category=show_category)


def _priority_table(report: pd.DataFrame, target_days: int, key: str) -> None:
    order_column = f"К заказу на {int(target_days)} дней"
    columns = [
        "Приоритет заказа", "Группа камня", "Камень группы", "Номенклатурная группа",
        "Продано уникальных SKU", "Продано штук", "Продано на сумму, USD",
        "Осталось уникальных SKU", "Всего штук", "Покрытие остатком, дней", order_column,
    ]
    _table(report[columns], key)


def _render_ai_overview(frame: pd.DataFrame, rate: float, period_days: int, period: str) -> None:
    report = order_priority_report(frame, rate, period_days)
    sku = network_sku_snapshot(frame, rate, period_days)
    sold = float(sku["Продано за период"].sum()) if not sku.empty else 0.0
    sold_sku = int((sku["Продано за период"] > 0).sum()) if not sku.empty else 0
    stock = float(sku["Остаток сети"].sum()) if not sku.empty else 0.0
    stock_sku = int((sku["Остаток сети"] > 0).sum()) if not sku.empty else 0
    sales = float(sku["Продажи USD"].sum()) if not sku.empty else 0.0

    st.markdown("### Общий отчет Sonu")
    st.caption(
        "Продажи суммируются по всей сети, общий остаток каждого SKU берется один раз. "
        "В отчет включены все изделия; браслеты разделены по конструкции."
    )
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        _kpi("Период", period)
    with k2:
        _kpi("Продано SKU", _money(sold_sku), f"{_money(sold)} изделий")
    with k3:
        _kpi("Продажи", f"${_money(sales)}")
    with k4:
        _kpi("SKU на остатке", _money(stock_sku), f"{_money(stock)} изделий")
    with k5:
        avg = sales / sold if sold else 0.0
        _kpi("Средняя цена", f"${_money(avg)}")

    target_days = int(st.segmented_control(
        "Горизонт заказа", [30, 45, 90], default=45, key="sonu_ai_horizon",
        help="Сколько дней продаж должен покрыть текущий остаток вместе с новым заказом.",
    ) or 45)
    insight = ai_sales_summary(report, target_days)
    st.markdown(
        '<div class="sonu-ai-brief">'
        '<div class="sonu-ai-label">AI-аналитика продаж и заказа</div>'
        f'<div class="sonu-ai-headline">{escape(insight["headline"])}</div>'
        f'<div class="sonu-ai-text">{escape(insight["order_text"])}</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    if insight["top_lines"]:
        st.markdown("#### Главные позиции для заказа")
        for number, line in enumerate(insight["top_lines"], start=1):
            st.markdown(f"**{number}.** {line}")

    priority_filter = st.multiselect(
        "Показать приоритеты",
        list(PRIORITY_ORDER),
        default=list(PRIORITY_ORDER),
        key="sonu_priority_filter",
    )
    visible = report.loc[report["Приоритет заказа"].isin(priority_filter)].copy()
    st.markdown("#### Полный отчет по продажам, остаткам и приоритету заказа")
    st.caption(
        "Приоритет учитывает скорость продаж, покрытие остатком и количество моделей на остатке. "
        "Расчет прозрачен: до 15 дней — максимальный приоритет, 15–30 — высокий, 30–45 — средний, "
        "45–90 — плановый, от 90 дней — не критично."
    )
    _priority_table(visible, target_days, "sonu_ai_priority_table")


@st.fragment
def _render_network_section(frame: pd.DataFrame, rate: float, period_days: int) -> None:
    st.markdown("### Изделия без браслетов")
    st.caption(
        "Большая общая сводка построена в иерархии: группа камней → вид камня → "
        "номенклатурная группа. Остаток каждого SKU учитывается один раз по всей сети."
    )
    conflicts = stock_conflict_details(frame)
    if not conflicts.empty:
        st.warning(
            f"У {len(conflicts)} SKU найдены разные значения общего остатка. "
            "В расчет взято максимальное значение; подробности есть в Excel-листе «Контроль остатков»."
        )
    detail = network_assortment_summary(frame, rate, period_days)
    groups = network_group_summary(frame, rate, period_days)
    if detail.empty:
        st.info("В отчете нет изделий, кроме браслетов.")
        return

    left, right = st.columns(2)
    with left:
        _locked_chart(
            _horizontal_chart(
                groups, "Группа камня", "Продано штук", "Продано штук по группам камней", suffix=" шт."
            ),
            "sonu_non_bracelet_qty_groups",
        )
    with right:
        _locked_chart(
            _horizontal_chart(
                groups, "Группа камня", "Продано на сумму, USD", "Продажи по группам камней", prefix="$"
            ),
            "sonu_non_bracelet_sales_groups",
        )

    st.markdown("#### Сводка по камням и номенклатурным группам")
    _render_business_summary(detail, "sonu_non_bracelet_summary", show_category=True)


@st.fragment
def _render_bracelet_section(frame: pd.DataFrame, rate: float, period_days: int) -> None:
    bracelets = classify_bracelets(frame, rate, period_days)
    if bracelets.empty:
        st.info("В отчете нет браслетов.")
        return
    pending = bracelets.loc[bracelets["Тип браслета"].isna()].copy()
    classified = bracelets.loc[bracelets["Тип браслета"].isin(BRACELET_TYPE_ORDER)].copy()
    summary = bracelet_type_summary(classified, period_days)
    stones = bracelet_stone_summary(classified, period_days)

    st.markdown("### Браслеты: с затяжкой и без затяжки")
    st.caption(
        "В группу «С затяжкой» также входят облегченные модели на кольцах с камнями преимущественно "
        "в центре. «Без затяжки (в круг)» — модели с камнями по всей окружности."
    )
    if not pending.empty:
        st.warning(
            f"{pending['SKU'].nunique()} SKU из {pending['Модельная семья'].nunique()} модельных семей "
            "пока не включены в две группы. Завершите блок «Требуют классификации» ниже."
        )
    if summary.empty:
        st.info("Классифицированных браслетов для построения диаграмм пока нет.")
        return
    left, right = st.columns(2)
    with left:
        _locked_chart(
            _horizontal_chart(
                summary, "Тип браслета", "Продано штук", "Продано браслетов за период", suffix=" шт."
            ),
            "sonu_bracelet_qty_comparison",
        )
    with right:
        _locked_chart(
            _horizontal_chart(
                summary, "Тип браслета", "Продано на сумму, USD", "Продажи браслетов за период", prefix="$"
            ),
            "sonu_bracelet_sales_comparison",
        )

    for bracelet_type in BRACELET_TYPE_ORDER:
        type_stones = stones.loc[stones["Тип браслета"] == bracelet_type].copy()
        st.markdown(f"#### {bracelet_type}")
        if type_stones.empty:
            st.info("Для этого типа браслетов данных нет.")
            continue
        st.markdown("##### Виды камней")
        _render_business_summary(
            type_stones.drop(columns=["Тип браслета"]),
            f"sonu_bracelet_{_bracelet_model_key(bracelet_type)}",
            show_category=False,
        )


@st.fragment
def _render_models_section(frame: pd.DataFrame, rate: float, period_days: int, view_mode: str) -> None:
    data = network_sku_snapshot(frame, rate, period_days)
    bracelets = classify_bracelets(frame, rate, period_days)
    if not bracelets.empty:
        bracelet_types = bracelets.set_index("SKU")["Тип браслета"]
        data["Тип браслета"] = data["SKU"].map(bracelet_types).fillna("")
    else:
        data["Тип браслета"] = ""
    st.markdown("### Все SKU по сети")
    st.caption("Одна строка или карточка на SKU: общий остаток сети, суммарные продажи и потребность на 30/45/90 дней.")
    f1, f2, f3 = st.columns(3)
    with f1:
        categories = ["Все"] + sorted(data["Категория RU"].dropna().astype(str).unique().tolist())
        category = st.selectbox("Категория", categories, key="sonu_model_category")
    with f2:
        groups = ["Все"] + STONE_GROUP_ORDER
        stone_group = st.selectbox("Группа камня", groups, key="sonu_model_stone_group")
    with f3:
        search = st.text_input("Поиск по SKU или камню", key="sonu_model_search")
    available_members = sorted(data["Камень группы"].dropna().astype(str).unique().tolist())
    member = st.selectbox("Вид камня", ["Все"] + available_members, key="sonu_model_stone")

    if category != "Все":
        data = data.loc[data["Категория RU"] == category]
    if stone_group != "Все":
        data = data.loc[data["Группа камня"] == stone_group]
    if member != "Все":
        data = data.loc[data["Камень группы"] == member]
    if search:
        haystack = data[["SKU", "Камень", "Камень группы", "Сокращение"]].fillna("").astype(str).agg(" ".join, axis=1).str.casefold()
        data = data.loc[haystack.str.contains(re.escape(search.casefold()), regex=True)]

    if view_mode == "Таблица":
        columns = [
            "Фото", "SKU", "Категория RU", "Тип браслета", "Группа камня", "Камень группы",
            "Сокращение", "Проба", "Остаток сети", "Продано за период", "Магазинов с продажами",
            "Продажи USD", "Средняя цена USD", "Нужно на 30 дней", "Нужно на 45 дней", "Нужно на 90 дней",
        ]
        _table(data[columns], "sonu_models_table")
    else:
        _render_model_cards(data, "sonu_models_cards")



SONU_MAIN_SECTIONS = (
    ("Серьги", "Earrings"),
    ("Кольца", "Ring"),
    ("Подвески", "Pendant"),
    ("Браслеты не полный круг", CENTERED_BRACELET_LABEL),
    ("Браслеты полный круг", FULL_CIRCLE_BRACELET_LABEL),
)

SONU_MAIN_COLUMNS = [
    "Камень", "Кол-во уникальных SKU", "Продано изделий",
    "Общий Total продаж, USD", "SKU на остатке", "Остаток, шт.",
]


def _sonu_stone_table_from_sku(sku: pd.DataFrame) -> pd.DataFrame:
    if sku.empty:
        return pd.DataFrame(columns=SONU_MAIN_COLUMNS)
    work = sku.copy()
    work["_sold_sku"] = (pd.to_numeric(work["Продано за период"], errors="coerce").fillna(0) > 0).astype(int)
    work["_stock_sku"] = (pd.to_numeric(work["Остаток сети"], errors="coerce").fillna(0) > 0).astype(int)
    result = work.groupby("Камень группы", as_index=False, dropna=False).agg(
        **{
            "Кол-во уникальных SKU": ("_sold_sku", "sum"),
            "Продано изделий": ("Продано за период", "sum"),
            "Общий Total продаж, USD": ("Продажи USD", "sum"),
            "SKU на остатке": ("_stock_sku", "sum"),
            "Остаток, шт.": ("Остаток сети", "sum"),
        }
    ).rename(columns={"Камень группы": "Камень"})
    return result.sort_values(["Продано изделий", "Общий Total продаж, USD"], ascending=[False, False]).reset_index(drop=True)


def sonu_merchandise_tables(frame: pd.DataFrame, rate: float, period_days: int) -> dict[str, pd.DataFrame]:
    """Return the five user-facing Sonu tables requested by merchandise group."""
    sku = network_sku_snapshot(frame, rate, period_days)
    bracelets = classify_bracelets(frame, rate, period_days)
    tables = {}
    for title, category in SONU_MAIN_SECTIONS[:3]:
        tables[title] = _sonu_stone_table_from_sku(sku.loc[sku["Категория"] == category].copy())
    for title, bracelet_type in SONU_MAIN_SECTIONS[3:]:
        current = bracelets.loc[bracelets["Тип браслета"] == bracelet_type].copy() if not bracelets.empty else pd.DataFrame()
        tables[title] = _sonu_stone_table_from_sku(current)
    return tables


def sonu_order_recommendations(table: pd.DataFrame) -> pd.DataFrame:
    """Rank order actions without any day-based demand horizon."""
    if table.empty:
        return pd.DataFrame(columns=["Приоритет", "Камень", "Причина"])
    data = table.copy()
    sold = pd.to_numeric(data["Продано изделий"], errors="coerce").fillna(0)
    stock = pd.to_numeric(data["Остаток, шт."], errors="coerce").fillna(0)
    sold_sku = pd.to_numeric(data["Кол-во уникальных SKU"], errors="coerce").fillna(0)
    stock_sku = pd.to_numeric(data["SKU на остатке"], errors="coerce").fillna(0)
    positive = sold[sold > 0]
    high_sales = float(positive.quantile(.65)) if not positive.empty else 0.0
    rows = []
    for idx, row in data.iterrows():
        s, q, ss, qs = float(sold.loc[idx]), float(stock.loc[idx]), float(sold_sku.loc[idx]), float(stock_sku.loc[idx])
        breadth_low = ss > 0 and qs <= max(1.0, ss * .5)
        concentrated_stale = qs > 0 and breadth_low and q >= max(s * 1.5, qs * 3)
        demand_pressure = s >= max(high_sales, 1.0) and (q <= s or breadth_low)
        if demand_pressure and q <= max(s * .5, 2):
            priority, reason, rank = "Очень нужно заказать", "Высокие продажи при минимальном остатке и узком выборе моделей.", 1
        elif demand_pressure:
            priority, reason, rank = "Нужно заказать", "Продажи сильные, а остаток или количество доступных SKU уже ограничены.", 2
        elif concentrated_stale:
            priority, reason, rank = "Сначала обновить модели", "SKU на остатке мало, но штук много: запас сконцентрирован в непродающихся моделях.", 3
        elif s > 0 and q < s * 1.5:
            priority, reason, rank = "Желательно пополнить", "Текущий остаток ненамного превышает продажи за период.", 4
        else:
            priority, reason, rank = "Не критично", "Остаток и ширина ассортимента пока достаточны относительно продаж.", 5
        rows.append({"Приоритет": priority, "Камень": row["Камень"], "Причина": reason, "Продано изделий": s, "Остаток, шт.": q, "SKU на остатке": qs, "_rank": rank})
    return pd.DataFrame(rows).sort_values(["_rank", "Продано изделий", "Остаток, шт."], ascending=[True, False, True]).drop(columns="_rank").reset_index(drop=True)



def sonu_stone_category_overview(frame: pd.DataFrame, rate: float, period_days: int) -> pd.DataFrame:
    """Detailed network assortment view: stone group → stone → merchandise group."""
    sku = network_sku_snapshot(frame, rate, period_days)
    bracelets = classify_bracelets(frame, rate, period_days)
    parts: list[pd.DataFrame] = []

    for title, category in SONU_MAIN_SECTIONS[:3]:
        current = sku.loc[sku["Категория"] == category].copy()
        if not current.empty:
            current["Номенклатурная группа"] = title
            parts.append(current)

    for title, bracelet_type in SONU_MAIN_SECTIONS[3:]:
        current = bracelets.loc[bracelets["Тип браслета"] == bracelet_type].copy() if not bracelets.empty else pd.DataFrame()
        if not current.empty:
            current["Номенклатурная группа"] = title
            parts.append(current)
    if not bracelets.empty:
        pending = bracelets.loc[bracelets["Тип браслета"].isna()].copy()
        if not pending.empty:
            pending["Номенклатурная группа"] = "Браслеты требуют классификации"
            parts.append(pending)

    columns = [
        "Группа камней", "Камень", "Номенклатурная группа",
        "Продано уникальных SKU", "Продано изделий", "Продажи, USD",
        "Средняя цена изделия, USD", "SKU на остатке", "Остаток, шт.",
    ]
    if not parts:
        return pd.DataFrame(columns=columns)

    work = pd.concat(parts, ignore_index=True, sort=False)
    work["_sold_sku"] = (pd.to_numeric(work["Продано за период"], errors="coerce").fillna(0) > 0).astype(int)
    work["_stock_sku"] = (pd.to_numeric(work["Остаток сети"], errors="coerce").fillna(0) > 0).astype(int)
    work["Группа камня"] = work.get("Группа камня", "Other Stones").fillna("Other Stones")
    work["Камень группы"] = work.get("Камень группы", work.get("Камень", "Other")).fillna("Other")

    result = work.groupby(
        ["Группа камня", "Камень группы", "Номенклатурная группа"],
        as_index=False, dropna=False,
    ).agg(**{
        "Продано уникальных SKU": ("_sold_sku", "sum"),
        "Продано изделий": ("Продано за период", "sum"),
        "Продажи, USD": ("Продажи USD", "sum"),
        "SKU на остатке": ("_stock_sku", "sum"),
        "Остаток, шт.": ("Остаток сети", "sum"),
    }).rename(columns={"Группа камня": "Группа камней", "Камень группы": "Камень"})
    sold = pd.to_numeric(result["Продано изделий"], errors="coerce").fillna(0)
    sales = pd.to_numeric(result["Продажи, USD"], errors="coerce").fillna(0)
    result["Средняя цена изделия, USD"] = sales.div(sold.where(sold.ne(0), 1)).where(sold.ne(0), 0)
    order = {name: idx for idx, name in enumerate(STONE_GROUP_ORDER)}
    result["_group_order"] = result["Группа камней"].map(order).fillna(len(order))
    result = result.sort_values(
        ["_group_order", "Камень", "Продано изделий", "Номенклатурная группа"],
        ascending=[True, True, False, True],
    ).drop(columns="_group_order").reset_index(drop=True)
    return result[columns]

def sonu_category_overview(section_tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for name, table in section_tables.items():
        rows.append({
            "Номенклатурная группа": name,
            "Камней": int(table["Камень"].nunique()) if not table.empty else 0,
            "Продано уникальных SKU": int(table["Кол-во уникальных SKU"].sum()) if not table.empty else 0,
            "Продано изделий": float(table["Продано изделий"].sum()) if not table.empty else 0,
            "Продажи, USD": float(table["Общий Total продаж, USD"].sum()) if not table.empty else 0,
            "SKU на остатке": int(table["SKU на остатке"].sum()) if not table.empty else 0,
            "Остаток, шт.": float(table["Остаток, шт."].sum()) if not table.empty else 0,
        })
    return pd.DataFrame(rows)


def _render_sonu_recommendations(table: pd.DataFrame, key: str) -> None:
    recommendations = sonu_order_recommendations(table)
    actionable = recommendations.loc[recommendations["Приоритет"] != "Не критично"].head(6)
    st.markdown("##### Рекомендации к заказу")
    if actionable.empty:
        st.success("Критичных сигналов по продажам, остаткам и ширине SKU сейчас нет.")
        return
    _table(actionable[["Приоритет", "Камень", "Причина", "Продано изделий", "Остаток, шт.", "SKU на остатке"]], key)


def _render_sonu_group(title: str, table: pd.DataFrame, key: str) -> None:
    st.markdown(f"### {title}")
    st.caption("Остаток учитывается один раз на SKU по всей сети; продажи суммируются по всем магазинам.")
    if table.empty:
        st.info("В отчете нет данных по этой номенклатурной группе.")
        return
    _table(table[SONU_MAIN_COLUMNS], f"{key}_table")
    left, right = st.columns(2)
    with left:
        _locked_chart(_horizontal_chart(table, "Камень", "Продано изделий", "Продано по камням", suffix=" шт."), f"{key}_qty")
    with right:
        _locked_chart(_horizontal_chart(table, "Камень", "Общий Total продаж, USD", "Total продаж по камням", prefix="$"), f"{key}_sales")
    _render_sonu_recommendations(table, f"{key}_recommendations")


def _render_sonu_extra(section_tables: dict[str, pd.DataFrame], frame: pd.DataFrame, rate: float, period_days: int) -> None:
    st.markdown("## Общая картина ассортимента")
    st.caption("Развернутая сводка показывает каждый камень внутри номенклатурных групп: сколько моделей и изделий продано, среднюю цену и общий сетевой остаток.")
    detail = sonu_stone_category_overview(frame, rate, period_days)
    _table(detail, "sonu_stone_category_overview")
    overview = sonu_category_overview(section_tables)
    st.markdown("### Сводно по номенклатурным группам")
    left, right = st.columns(2)
    with left:
        _locked_chart(_horizontal_chart(overview, "Номенклатурная группа", "Продано изделий", "Продано по группам", suffix=" шт."), "sonu_category_qty")
    with right:
        _locked_chart(_horizontal_chart(overview, "Номенклатурная группа", "Продажи, USD", "Total продаж по группам", prefix="$"), "sonu_category_sales")
    all_rows = []
    for group, table in section_tables.items():
        rec = sonu_order_recommendations(table)
        if not rec.empty:
            rec.insert(0, "Номенклатурная группа", group)
            all_rows.append(rec)
    if all_rows:
        combined = pd.concat(all_rows, ignore_index=True)
        urgent = combined.loc[combined["Приоритет"].isin(["Очень нужно заказать", "Нужно заказать", "Сначала обновить модели"])].head(12)
        st.markdown("### Главные сигналы")
        if urgent.empty:
            st.success("Срочных сигналов по заказу или обновлению моделей не обнаружено.")
        else:
            _table(urgent[["Приоритет", "Номенклатурная группа", "Камень", "Причина"]], "sonu_main_signals")

def render_sonu_order_dashboard(selected_metal_groups: Iterable[str] = SONU_METAL_GROUPS) -> None:
    """Streamlit entry point for the streamlined five-group Sonu report."""
    _render_sonu_css()
    rate = get_vnd_per_usd()
    _anchor("sonu-upload")
    uploaded = st.file_uploader("Загрузите отчет Sonu", type=["xlsx", "xlsm"], accept_multiple_files=False, key="sonu_upload_widget", help="Остаток в файле — общий по сети и учитывается один раз на SKU.")
    file_bytes = _persist_upload(uploaded)
    if file_bytes is None:
        _sonu_sidebar_navigation(False); _sonu_mobile_navigation(False)
        st.info("Загрузите отчет Sonu. Все изделия этого поставщика анализируются как серебряный ассортимент; фильтр проб здесь не используется.")
        return
    try:
        with st.spinner("Разбираем продажи и общий остаток сети..."):
            report = cached_parse_sonu(file_bytes)
    except Exception as exc:
        navigation = _sonu_sidebar_navigation(False, status_text="Файл Sonu не распознан", status_tone="error", action_label="Удалить загруженный файл", action_key="sonu_clear_invalid")
        _sonu_mobile_navigation(False); st.error(str(exc))
        if navigation.action_clicked:
            st.session_state.pop("sonu_report_bytes", None); st.session_state.pop("sonu_report_name", None); st.session_state.pop("sonu_upload_widget", None); st.rerun()
        return
    navigation = _sonu_sidebar_navigation(True, action_label="Загрузить другой отчет", action_key="sonu_replace_report")
    _sonu_mobile_navigation(True)
    if navigation.action_clicked:
        st.session_state.pop("sonu_report_bytes", None); st.session_state.pop("sonu_report_name", None); st.session_state.pop("sonu_upload_widget", None); st.rerun()
    frame = report.data.copy()
    if "Проба" in frame.columns:
        _sync_detected_purities(frame["Проба"].tolist())
    selected = tuple(str(value) for value in selected_metal_groups)
    if not selected:
        st.error("Оставьте включенной хотя бы одну группу металла.")
        return
    frame = filter_sonu_metal_groups(frame, selected)
    if frame.empty:
        st.warning("После применения фильтра металла в отчете Sonu не осталось позиций.")
        return
    period_days = _period_days(report.period)
    st.caption(f"Файл: {st.session_state.get('sonu_report_name', 'Sonu.xlsx')} · Поставщик: {report.supplier} · Курс: 1 USD = {_money(rate)} VND · Остаток берется один раз на SKU по всей сети.")
    conflicts = stock_conflict_details(frame)
    if not conflicts.empty:
        st.warning(f"У {len(conflicts)} SKU обнаружены разные повторные значения сетевого остатка. В расчет взято максимальное значение.")
    section_tables = sonu_merchandise_tables(frame, rate, period_days)
    _anchor("sonu-main-report")
    st.markdown("## Основной отчет Sonu")
    st.caption("Пять последовательных блоков без разбивки по магазинам, пробам и горизонтам в днях.")
    keys = ["earrings", "rings", "pendants", "bracelets_centered", "bracelets_circle"]
    for (title, table), key in zip(section_tables.items(), keys):
        _render_sonu_group(title, table, f"sonu_{key}")
    _anchor("sonu-bracelet-classification")
    _render_bracelet_classification_audit(frame, rate, period_days)
    _anchor("sonu-extra")
    _render_sonu_extra(section_tables, frame, rate, period_days)
    _anchor("sonu-export")
    st.markdown("## Полная выгрузка Sonu")
    overrides_signature = bracelet_overrides_json(load_bracelet_overrides()).decode("utf-8")
    export_bytes = build_full_sonu_export(
        frame, report.period, report.supplier, rate, overrides_signature
    )
    safe_period = re.sub(r"[^0-9]+", "_", report.period).strip("_") or "period"
    st.download_button("Скачать полный отчет Sonu", data=export_bytes, file_name=f"Sonu_merchandise_report_{safe_period}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch", key="sonu_full_export", help="Пять товарных таблиц, рекомендации, общая сводка, SKU сети и контроль повторного остатка.")

