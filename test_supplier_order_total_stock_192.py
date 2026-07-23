from __future__ import annotations

import io
import inspect
from pathlib import Path

from openpyxl import Workbook, load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    build_limited_order_excel,
    parse_order_workbook,
)


def _item(*, total: int, stock20: int, stock63: int = 0, tt: int = 0) -> OrderItem:
    return OrderItem(
        row=12,
        set_id="Set# 1",
        sku="KE15E518B-BS",
        stone="BLUE SAPPHIRE",
        group="Earrings",
        sales=6,
        stock_63=stock63,
        stock_20=stock20,
        stores={"20": stock20, "63": stock63, "Outlet": tt},
        total_stock=total,
        working_stock=max(0, total - stock20 - stock63),
        ntr2_stock=0,
        ntr2_calculated=False,
        tvp_raw=0,
        stock_tt=tt,
    )


def test_display_stock_is_total_minus_store_20_only() -> None:
    item = _item(total=9, stock20=2, stock63=3, tt=1)
    assert item.display_stock == 7
    assert item.working_stock == 4


def test_example_ke15e518b_bs_shows_four_not_three() -> None:
    item = _item(total=4, stock20=0, stock63=1, tt=1)
    assert item.display_stock == 4
    assert item.stock_tt == 1
    assert item.stock_63 == 1
    assert item.working_stock == 3


def test_parser_accepts_20ndc_and_63ndc_labels(tmp_path: Path) -> None:
    path = tmp_path / "supplier.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["E7"] = "Продажи за период"
    sheet["G7"] = "Остатки"
    sheet["O7"] = "ТВП"
    for column, value in enumerate(["63NDC", "20NDC", "AB", "NTR1", "NTR2", "Outlet", "Stock TT", "Всего"], start=7):
        sheet.cell(8, column).value = value
    sheet["A11"] = "Set# 1"
    sheet["A12"] = "KE15E518B-BS"
    sheet["B12"] = "BLUE SAPPHIRE"
    sheet["C12"] = "Earrings"
    sheet["E12"] = 6
    sheet["G12"] = 1
    sheet["H12"] = 0
    sheet["I12"] = 0
    sheet["J12"] = 2
    sheet["K12"] = 0
    sheet["L12"] = 1
    sheet["M12"] = 0
    sheet["N12"] = 4
    sheet["O12"] = 0
    workbook.save(path)

    parsed = parse_order_workbook(path)
    item = parsed.items[0]
    assert item.total_stock == 4
    assert item.stock_20 == 0
    assert item.stock_63 == 1
    assert item.stock_tt == 1
    assert item.display_stock == 4
    assert item.working_stock == 3


def test_card_uses_display_stock_but_recommendations_keep_working_stock() -> None:
    row_source = inspect.getsource(workflow._render_item_row)
    recommendation_source = inspect.getsource(workflow.build_order_recommendation)
    assert '_render_stock_metric("Общий остаток", item.display_stock, always=True)' in row_source
    assert "item.working_stock" in recommendation_source


def test_limited_order_excel_exports_display_stock(tmp_path: Path) -> None:
    item = _item(total=9, stock20=2, stock63=3, tt=1)
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    parsed = ParsedOrderWorkbook(
        source_name="source.xlsx",
        source_hash="hash",
        upload_path=str(source),
        period="",
        supplier="",
        store_columns=(),
        has_actual_ntr2=True,
        items=(item,),
    )
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.limited_orders[item.key] = True

    workbook = load_workbook(io.BytesIO(build_limited_order_excel(parsed, [item], draft)))
    assert workbook["Limited Order"]["G2"].value == 7


def test_princess_hang_alias_is_the_same_as_store_20(tmp_path: Path) -> None:
    path = tmp_path / "princess-hang.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["E7"] = "Продажи за период"
    sheet["G7"] = "Остатки"
    sheet["N7"] = "ТВП"
    for column, value in enumerate(["63NDC", "Princess Hang", "AB", "NTR1", "NTR2", "Outlet", "Всего"], start=7):
        sheet.cell(8, column).value = value
    sheet["A11"] = "Set# 1"
    sheet["A12"] = "RG-PH"
    sheet["B12"] = "RUBY"
    sheet["C12"] = "Ring"
    sheet["E12"] = 6
    sheet["G12"] = 1
    sheet["H12"] = 2
    sheet["I12"] = 0
    sheet["J12"] = 2
    sheet["K12"] = 0
    sheet["L12"] = 1
    sheet["M12"] = 6
    sheet["N12"] = 0
    workbook.save(path)

    parsed = parse_order_workbook(path)
    item = parsed.items[0]
    assert item.total_stock == 6
    assert item.stock_20 == 2
    assert item.stock_princess_hang == 0
    assert item.display_stock == 4
    assert item.working_stock == 3


def test_store_20_alias_values_are_not_summed() -> None:
    quantity, warning = workflow.resolve_store_20_stock({"20NDC": 3, "Princess Hang": 3})
    assert quantity == 3
    assert warning is None


def test_conflicting_store_20_aliases_use_max_and_warn() -> None:
    quantity, warning = workflow.resolve_store_20_stock({"20": 1, "Princess Hang": 4})
    assert quantity == 4
    assert warning is not None
    assert "один магазин" in warning
