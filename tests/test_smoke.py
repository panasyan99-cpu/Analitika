from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

from src.report import run_files


class SmokeTest(unittest.TestCase):
    def test_consolidated_report(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "all_stores.xlsx"
            output = root / "result.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Отчет о продажах товаров за период Июль 2026 г. - Июль 2026 г."
            ws["A4"] = "Магазин; Камень/вставка; Номенклатурная группа"
            ws["H4"] = "Продано"
            ws["H5"] = "Кол-во"
            ws["I5"] = "Сумма"
            ws["A7"] = "AB"
            ws.row_dimensions[7].outlineLevel = 0
            ws["A8"] = "MOISSANITE + RUBY"
            ws.row_dimensions[8].outlineLevel = 1
            ws["A9"] = "Ring"
            ws.row_dimensions[9].outlineLevel = 2
            ws["H9"] = 2
            ws["I9"] = 1000000
            wb.save(source)

            run_files([source], output)
            result = load_workbook(output)
            self.assertIn("SUMMARY", result.sheetnames)
            self.assertIn("RULES", result.sheetnames)
            self.assertTrue(any(name.startswith("AB ") for name in result.sheetnames))
            result.close()


if __name__ == "__main__":
    unittest.main()
