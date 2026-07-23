from __future__ import annotations

import io
import inspect
from pathlib import Path

from openpyxl import Workbook, load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    CATEGORY_MEDIUM,
    CATEGORY_TOP,
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    OrderSet,
    ParsedOrderWorkbook,
    build_limited_order_excel,
    build_order_recommendation,
    build_supplier_excel,
    parse_order_workbook,
    validate_draft_payload,
)


def make_item(
    sku: str,
    *,
    group: str = "Ring",
    sales: int = 0,
    stock: int = 0,
    tt: int = 0,
    stock63: int = 0,
    tvp: int = 0,
    set_id: str = "Set# 1",
    row: int = 12,
) -> OrderItem:
    return OrderItem(
        row=row,
        set_id=set_id,
        sku=sku,
        stone="BLUE SAPPHIRE",
        group=group,
        sales=sales,
        stock_63=stock63,
        stock_20=0,
        stores={"TT": tt, "63": stock63},
        total_stock=stock + stock63,
        working_stock=stock,
        ntr2_stock=0,
        ntr2_calculated=False,
        tvp_raw=tvp,
        stock_tt=tt,
    )


def make_set(item: OrderItem, category: str = CATEGORY_MEDIUM, *others: OrderItem) -> OrderSet:
    items = (item, *others)
    return OrderSet(
        key=f"Blue Sapphire|{item.set_id}",
        set_id=item.set_id,
        stone="Blue Sapphire",
        items=items,
        category=category,
        driver_sku=item.sku,
        max_sales=max(x.sales for x in items),
        has_positive_tvp=any(x.tvp_raw > 0 for x in items),
        has_negative_tvp=any(x.tvp_raw < 0 for x in items),
    )


def test_two_month_recommendation_and_explicit_tvp_block():
    enough = make_item("RG-ENOUGH", sales=6, stock=6, tt=1)
    low = make_item("RG-LOW", sales=6, stock=4, tt=1)
    transit = make_item("RG-TVP", sales=9, stock=0, tt=0, tvp=5)

    assert build_order_recommendation(enough, make_set(enough), ORDER_MODE_STONES).quantity == 0
    assert build_order_recommendation(low, make_set(low), ORDER_MODE_STONES).quantity == 4
    rec = build_order_recommendation(transit, make_set(transit, CATEGORY_TOP), ORDER_MODE_STONES)
    assert rec.quantity == 0
    assert rec.blocked_by_tvp is True
    assert "5" in rec.reasons[0]


def test_tt_balance_and_stud_rules():
    earrings = make_item("ER-SET", group="Earrings", sales=5, stock=6, tt=2)
    ring = make_item("RG-SET", group="Ring", sales=1, stock=2, tt=0, row=13)
    order_set = make_set(earrings, CATEGORY_TOP, ring)
    rec = build_order_recommendation(ring, order_set, ORDER_MODE_STONES)
    assert rec.quantity == 4
    assert any("баланс" in reason.lower() or "TT" in reason for reason in rec.reasons)

    stud = make_item("ER-STUD", group="Stud Earrings", sales=4, stock=0, tt=0, set_id="Пусеты BS")
    stud_rec = build_order_recommendation(stud, make_set(stud, CATEGORY_MEDIUM), ORDER_MODE_STONES)
    assert stud_rec.quantity == 10
    assert stud_rec.rule == "studs"


def test_princess_hang_is_removed_from_visible_working_stock(tmp_path: Path):
    path = tmp_path / "any-name.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["E7"] = "Продажи за период"
    ws["G7"] = "Остатки"
    ws["M7"] = "ТВП"
    headers = ["63", "20", "TT", "Princess Hang", "NTR2"]
    for index, value in enumerate(headers, start=7):
        ws.cell(8, index).value = value
    ws["L8"] = "Всего"
    ws["A11"] = "Set# 1"
    ws["A12"] = "RG-1"
    ws["B12"] = "RUBY"
    ws["C12"] = "Ring"
    ws["E12"] = 6
    ws["G12"] = 2  # 63
    ws["H12"] = 1  # 20
    ws["I12"] = 3  # TT
    ws["J12"] = 4  # Princess Hang
    ws["K12"] = 2  # NTR2
    ws["L12"] = 12
    ws["M12"] = 0
    wb.save(path)

    parsed = parse_order_workbook(path)
    item = parsed.items[0]
    assert parsed.source_name == "any-name.xlsx"
    assert item.stock_princess_hang == 4
    assert item.working_stock == 5  # 12 - 63(2) - 20(1) - Princess Hang(4)
    assert item.stock_tt == 3
    assert item.stock_63 == 2


def test_limited_order_persists_and_has_separate_excel(tmp_path: Path):
    item = make_item("RG-LIMITED", sales=5, stock=1, tt=0)
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders[item.key] = 5
    draft.manual_edit[item.key] = True
    draft.limited_orders[item.key] = True
    payload = draft.as_payload()
    restored = validate_draft_payload(payload)
    assert restored.limited_orders[item.key] is True
    assert restored.manual_edit[item.key] is True

    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    parsed = ParsedOrderWorkbook(
        source_name="source.xlsx",
        source_hash="hash",
        upload_path=str(source),
        period="",
        supplier="",
        store_columns=(),
        has_actual_ntr2=True,
        items=(item,),
    )
    main = load_workbook(io.BytesIO(build_supplier_excel(parsed, [item], draft)))
    assert main["Order"].max_row == 1
    limited = load_workbook(io.BytesIO(build_limited_order_excel(parsed, [item], draft)))
    sheet = limited["Limited Order"]
    assert sheet["B2"].value == "RG-LIMITED"
    assert sheet["G2"].value == 1


def test_ui_has_no_tvp_toggle_or_ntr2_caption_and_uses_three_stock_cubes():
    workspace_source = inspect.getsource(workflow._render_order_workspace)
    row_source = inspect.getsource(workflow._render_item_row)
    assert "st.toggle" not in workspace_source
    assert "NTR2:" not in row_source
    assert '"Всего остаток"' in row_source
    assert '"TT"' in row_source
    assert '"63"' in row_source
    assert "Согласен с рекомендацией" in row_source
    assert "Limited Order" in row_source
