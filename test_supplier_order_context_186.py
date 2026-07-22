from __future__ import annotations

import io
import inspect
from pathlib import Path

from openpyxl import load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    CATEGORY_TOP,
    ORDER_MODE_STONES,
    OTHER_STONES_GROUP,
    OTHER_TOPAZ_GROUP,
    UNRECOGNIZED_STONE,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    build_order_sets,
    build_supplier_excel,
    canonical_stone,
    filter_order_sets_by_tvp,
    order_stone_bucket,
    purge_order_workspaces_except,
)


def _item(
    row: int,
    sku: str,
    *,
    set_id: str = "Set# 00847",
    stone: str = "Blue Sapphire",
    group: str = "Ring",
    sales: int = 0,
    tvp: int = 0,
) -> OrderItem:
    return OrderItem(
        row=row,
        set_id=set_id,
        sku=sku,
        stone=stone,
        group=group,
        sales=sales,
        stock_63=0,
        stock_20=0,
        stores={},
        total_stock=0,
        working_stock=0,
        ntr2_stock=0,
        ntr2_calculated=True,
        tvp_raw=tvp,
    )


def test_top_set_displays_earrings_even_when_ring_drives_category_and_tvp_differs() -> None:
    ring = _item(1, "SKR17N368B-BSS", group="Ring", sales=5, tvp=0)
    earrings = _item(2, "SKE17E368B-BSS", group="Earrings", sales=0, tvp=3)
    order_set = build_order_sets((ring, earrings), ORDER_MODE_STONES)[0]
    assert order_set.category == CATEGORY_TOP
    assert order_set.driver_sku == ring.sku
    assert [item.sku for item in order_set.items] == [ring.sku, earrings.sku]
    assert [item.sku for item in filter_order_sets_by_tvp((order_set,), False)[0].items] == [ring.sku, earrings.sku]
    assert [item.sku for item in filter_order_sets_by_tvp((order_set,), True)[0].items] == [ring.sku, earrings.sku]


def test_tvp_secondary_expander_is_removed() -> None:
    source = inspect.getsource(workflow._render_category_segment)
    assert "st.expander" not in source
    assert "Есть товар в пути" not in source


def test_new_source_purges_old_workbook_and_old_selected_draft(tmp_path: Path, monkeypatch) -> None:
    runtime = tmp_path / "runtime"
    monkeypatch.setattr(workflow, "UPLOAD_DIR", runtime / "uploads")
    monkeypatch.setattr(workflow, "DRAFT_DB", runtime / "order_drafts.sqlite3")

    old_path, old_hash = workflow.store_uploaded_workbook("old.xlsx", b"old-report")
    new_path, new_hash = workflow.store_uploaded_workbook("new.xlsx", b"new-report")
    old_draft = OrderDraft(
        source_hash=old_hash,
        source_name="old.xlsx",
        mode=ORDER_MODE_STONES,
        orders={"old-item": 9},
        sizes={"old-item": {"18": 5, "19": 4}},
        stock_checked={"old-item": True},
    )
    new_draft = OrderDraft(
        source_hash=new_hash,
        source_name="new.xlsx",
        mode=ORDER_MODE_STONES,
        orders={"new-item": 0},
    )
    workflow.save_draft(old_draft)
    workflow.save_draft(new_draft)

    removed_drafts, removed_files = purge_order_workspaces_except(new_hash)
    assert removed_drafts == 1
    assert removed_files == 1
    assert not old_path.exists()
    assert new_path.exists()
    assert workflow.load_draft(old_hash, "old.xlsx", ORDER_MODE_STONES).orders == {}
    assert workflow.load_draft(new_hash, "new.xlsx", ORDER_MODE_STONES).orders == {"new-item": 0}


def test_widget_keys_are_isolated_by_source_hash() -> None:
    item = _item(1, "SKU")
    assert workflow._order_input_key(item, ORDER_MODE_STONES, "hash-a") != workflow._order_input_key(item, ORDER_MODE_STONES, "hash-b")
    assert workflow._size_input_key(item, 18, ORDER_MODE_STONES, "hash-a") != workflow._size_input_key(item, 18, ORDER_MODE_STONES, "hash-b")
    assert workflow._stock_check_key(item, ORDER_MODE_STONES, "hash-a") != workflow._stock_check_key(item, ORDER_MODE_STONES, "hash-b")


def test_all_analytical_groups_resolve_to_concrete_supplier_stones() -> None:
    assert canonical_stone("Green Stones", "ABC-GA-1") == "Green Agate"
    assert canonical_stone("Other Topaz", "ABC-MLBT-1") == "Multi Blue Topaz"
    assert canonical_stone("Other Topaz", "ABC-WBT-1") == "White Topaz"
    assert canonical_stone("Other Stones", "ABC-LAP-1") == "Lapis Lazurite"
    assert canonical_stone("Other Stones", "ABC-AMST-1") == "Amethyst"
    assert canonical_stone("Other Stones", "ABC-UNKNOWN-1") == UNRECOGNIZED_STONE
    assert order_stone_bucket("Other Topaz", "ABC-WBT-1") == OTHER_TOPAZ_GROUP
    assert order_stone_bucket("Other Stones", "ABC-LAP-1") == OTHER_STONES_GROUP


def test_supplier_excel_exports_concrete_other_topaz_and_other_stones() -> None:
    topaz = _item(1, "ABC-WBT-1", stone="Other Topaz")
    lapis = _item(2, "ABC-LAP-1", stone="Other Stones", group="Pendant")
    parsed = ParsedOrderWorkbook(
        source_name="supplier-any-name.xlsx",
        source_hash="hash",
        upload_path="unused.xlsx",
        period="",
        supplier="Y&J",
        store_columns=(),
        has_actual_ntr2=True,
        items=(topaz, lapis),
    )
    draft = OrderDraft(
        source_hash="hash",
        source_name=parsed.source_name,
        mode=ORDER_MODE_STONES,
        orders={topaz.key: 5, lapis.key: 5},
    )
    payload = build_supplier_excel(parsed, (topaz, lapis), draft)
    book = load_workbook(io.BytesIO(payload), data_only=True)
    sheet = book["Order"]
    assert sheet["C2"].value == "White Topaz"
    assert sheet["C3"].value == "Lapis Lazurite"
    assert sheet["C2"].value != OTHER_TOPAZ_GROUP
    assert sheet["C3"].value != OTHER_STONES_GROUP
