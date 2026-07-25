from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

import src.order_workflow as workflow
from src.order_workflow import (
    ORDER_MODE_PEARLS,
    ORDER_MODE_STONES,
    OrderDraft,
    OrderItem,
    ParsedOrderWorkbook,
    _export_readiness,
    build_limited_order_excel,
    build_order_sets,
    build_supplier_excel,
)


def _item(*, stone: str, group: str = "Ring", stock: int = 3) -> OrderItem:
    return OrderItem(
        row=12,
        set_id="Set# 1",
        sku="SKU-1",
        stone=stone,
        group=group,
        sales=5,
        stock_63=0,
        stock_20=0,
        stores={},
        total_stock=stock,
        working_stock=stock,
        ntr2_stock=0,
        ntr2_calculated=False,
        tvp_raw=0,
        stock_tt=0,
    )


def _parsed(tmp_path: Path, item: OrderItem) -> ParsedOrderWorkbook:
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    return ParsedOrderWorkbook(
        source_name=source.name,
        source_hash="hash",
        upload_path=str(source),
        period="",
        supplier="Y&J",
        store_columns=(),
        has_actual_ntr2=True,
        items=(item,),
    )


def test_main_stones_and_pearls_export_has_english_header_without_product_group(tmp_path: Path) -> None:
    for mode, stone in ((ORDER_MODE_STONES, "Ruby"), (ORDER_MODE_PEARLS, "Fresh Water Pearl - White")):
        item = _item(stone=stone)
        draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=mode)
        draft.orders[item.key] = 3
        draft.sizes[item.key] = {"18": 1, "19": 1, "20": 1}
        workbook = load_workbook(io.BytesIO(build_supplier_excel(_parsed(tmp_path, item), [item], draft)))
        sheet = workbook["Order"]
        assert [sheet.cell(1, column).value for column in range(1, 7)] == [
            "Photo", "SKU", "Stone", "Order Quantity", "Sizes", "Change Lock To"
        ]
        assert sheet.max_column == 6
        assert sheet["D2"].value == 3
        assert sheet["E2"].value == "18 × 1; 19 × 1; 20 × 1"
        assert "Ring" not in [sheet.cell(2, column).value for column in range(1, 7)]


def test_limited_order_export_contract_is_unchanged(tmp_path: Path) -> None:
    item = _item(stone="Ruby")
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.limited_orders[item.key] = True
    workbook = load_workbook(io.BytesIO(build_limited_order_excel(_parsed(tmp_path, item), [item], draft)))
    sheet = workbook["Limited Order"]
    assert [sheet.cell(1, column).value for column in range(1, 11)] == [
        "Фото", "Артикул", "Камень", "Группа", "Комплект",
        "Продажи", "Всего остаток", "TT", "63", "ТВП",
    ]
    assert sheet["D2"].value == "Ring"


def test_existing_stock_is_notice_only_and_does_not_block_export() -> None:
    item = _item(stone="Ruby", stock=4)
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders[item.key] = 3
    draft.sizes[item.key] = {"18": 1, "19": 1, "20": 1}
    draft.stock_checked = {}
    order_sets = build_order_sets((item,), ORDER_MODE_STONES)
    ready, reasons = _export_readiness(order_sets, draft)
    assert ready is True
    assert reasons == []

    source = Path(workflow.__file__).read_text(encoding="utf-8")
    assert 'st.info(f"По этой позиции есть остаток: {item.working_stock} шт."' in source
    assert "С остатком сверился" not in source
    assert "Не подтверждена сверка с остатком" not in source


def test_earring_lock_change_is_exported(tmp_path: Path) -> None:
    item = _item(stone="Ruby", group="Earrings")
    item = workflow.replace(item, sku="SKE17A004")
    draft = OrderDraft(source_hash="hash", source_name="source.xlsx", mode=ORDER_MODE_STONES)
    draft.orders[item.key] = 3
    draft.lock_changes[item.key] = "C"
    workbook = load_workbook(io.BytesIO(build_supplier_excel(_parsed(tmp_path, item), [item], draft)))
    assert workbook["Order"]["F2"].value == "C — Hook"
