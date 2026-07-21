from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.sonu import (
    _is_money_column,
    _money,
    _rounded_sonu_frame,
    build_full_sonu_export,
    cached_parse_sonu,
)


def test_grouped_number_format_has_no_cents():
    assert _money(34) == "34"
    assert _money(34300.49) == "34 300"
    assert _money(34400200.4) == "34 400 200"


def test_all_usd_column_variants_are_detected_and_rounded():
    frame = pd.DataFrame({
        "Общий Total продаж, USD": [42638.6692],
        "Продажи, USD": [36796.5589],
        "Средняя цена изделия, USD": [501.9392],
        "Продано изделий": [471.0],
    })
    rounded = _rounded_sonu_frame(frame)
    assert all(_is_money_column(column) for column in frame.columns[:3])
    assert rounded.iloc[0].to_dict() == {
        "Общий Total продаж, USD": 42639.0,
        "Продажи, USD": 36797.0,
        "Средняя цена изделия, USD": 502.0,
        "Продано изделий": 471.0,
    }


def test_sonu_excel_uses_integer_grouped_format():
    path = Path("/mnt/data/sonu(2).xlsx")
    if not path.exists():
        return
    report = cached_parse_sonu(path.read_bytes())
    payload = build_full_sonu_export(report.data, report.period, report.supplier, 26300)
    workbook = load_workbook(BytesIO(payload), data_only=True)
    sheet = workbook["Камни и группы"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for header in ("Продажи, USD", "Средняя цена изделия, USD", "Продано изделий", "Остаток, шт."):
        column = headers[header]
        cells = [sheet.cell(row, column) for row in range(2, sheet.max_row + 1) if sheet.cell(row, column).value is not None]
        assert cells
        assert all(cell.number_format == "# ##0" for cell in cells)
        assert all(float(cell.value).is_integer() for cell in cells)
