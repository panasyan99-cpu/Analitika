from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.full_sales_report import is_full_sales_report, parse_full_sales_report_with_period
from src.report import normalize_store_from_report


def _row(ws, number, text, indent, *, bold, qty=0, amount=0, sold_qty=None, returns=0, return_amount=0):
    cell = ws.cell(number, 1, text)
    cell.alignment = Alignment(indent=indent)
    cell.font = Font(bold=bold)
    ws.cell(number, 4, qty)
    ws.cell(number, 7, amount)
    ws.cell(number, 8, qty if sold_qty is None else sold_qty)
    ws.cell(number, 9, amount)
    ws.cell(number, 10, returns)
    ws.cell(number, 11, return_amount)


def _full_book() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Отчет о продажах товаров за период Июль 2026 г."
    ws["A2"] = "Поставщик(и):"
    ws["A4"] = "Магазин; Менеджер; Товар; Камень/вставка; Проба; Номенклатурная группа"
    ws["D4"] = "Отгружено"
    ws["H4"] = "Продано"
    ws["J4"] = "Возврат"
    ws["D5"] = "Кол-во (ед.)"
    ws["G5"] = "Сумма прод."
    ws["H5"] = "Кол-во"
    ws["I5"] = "Сумма"
    ws["J5"] = "Кол-во"
    ws["K5"] = "Сумма"
    return wb


def test_full_report_is_not_confused_with_supplier_filter(tmp_path: Path):
    wb = _full_book()
    ws = wb.active
    _row(ws, 7, "Princess-Hang", 0, bold=True, qty=1.5, amount=10_000)
    _row(ws, 8, "Anna", 2, bold=False, qty=1.5, amount=10_000)
    _row(ws, 9, "Ювелирные", 4, bold=True, qty=1.5, amount=10_000)
    _row(ws, 10, "Silver", 5, bold=True, qty=1.5, amount=10_000)
    _row(ws, 11, "Earrings", 6, bold=True, qty=1.5, amount=10_000)
    _row(ws, 12, "SKU-1", 5, bold=True, qty=1.5, amount=10_000, sold_qty=2)
    _row(ws, 13, "BLUE SAPPHIRE", 9, bold=False, qty=1.5, amount=10_000)
    _row(ws, 14, "B 925", 8, bold=False, qty=1.5, amount=10_000)
    _row(ws, 15, "Итого:", 0, bold=True, qty=1.5, amount=10_000)
    path = tmp_path / "full.xlsx"
    wb.save(path)

    assert is_full_sales_report(path)
    detail, period = parse_full_sales_report_with_period(path)
    assert period[0].strftime("%d.%m.%Y") == "01.07.2026"
    assert period[1].strftime("%d.%m.%Y") == "31.07.2026"
    assert detail["Количество"].sum() == 1.5  # shipped qty, not rounded Sold qty
    assert detail["Выручка"].sum() == 10_000
    assert detail.iloc[0]["Магазин"] == "20"
    assert detail.iloc[0]["Продавец"] == "Anna"
    assert detail.iloc[0]["Категория"] == "Ювелирные"
    assert detail.iloc[0]["Подгруппа"] == "Silver"
    assert detail.iloc[0]["Номенклатурная группа"] == "Earrings"


def test_blank_sku_is_retained_and_totals_are_checked(tmp_path: Path):
    wb = _full_book()
    ws = wb.active
    _row(ws, 7, "Cafe", 0, bold=True, qty=3, amount=90_000)
    _row(ws, 8, "Cafe", 2, bold=False, qty=3, amount=90_000)
    _row(ws, 9, "Продукты", 4, bold=True, qty=3, amount=90_000)
    _row(ws, 10, None, 3, bold=True, qty=3, amount=90_000)
    _row(ws, 11, None, 7, bold=False, qty=3, amount=90_000)
    _row(ws, 12, "OTHER 0", 6, bold=False, qty=3, amount=90_000)
    _row(ws, 13, "Итого:", 0, bold=True, qty=3, amount=90_000)
    path = tmp_path / "blank-sku.xlsx"
    wb.save(path)

    detail, _ = parse_full_sales_report_with_period(path)
    assert len(detail) == 1
    assert detail.iloc[0]["Товар"] == "Без названия"
    assert detail.iloc[0]["Продавец"] == "Cafe"
    assert detail["Выручка"].sum() == 90_000


def test_princess_hang_is_store_20():
    assert normalize_store_from_report("Princess-Hang") == "20"


def test_compact_supplier_summary_uses_shipped_columns_and_skips_subtotal(tmp_path: Path):
    from src.full_sales_report import (
        is_supplier_summary_report,
        parse_supplier_summary_report_with_period,
    )

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Отчет о продажах товаров за период 01.07.2026 - 30.07.2026"
    ws["A4"] = "Поставщик"
    ws["B4"] = "Отгружено"
    ws["F4"] = "Возврат"
    ws["B5"] = "Кол-во (ед.)"
    ws["E5"] = "Сумма прод."
    ws["F5"] = "Кол-во"
    ws["G5"] = "Сумма"

    ws["B7"] = 2
    ws["E7"] = 20_000
    ws["A8"] = "Поставщики"
    ws["B8"] = 3
    ws["E8"] = 30_000
    ws["A9"] = "Own production"
    ws["B9"] = 1
    ws["E9"] = 10_000
    ws["F9"] = 1
    ws["G9"] = 1_000
    ws["A10"] = "Taiwan"
    ws["B10"] = 2
    ws["E10"] = 20_000
    ws["A11"] = "Итого:"
    ws["B11"] = 5
    ws["E11"] = 50_000
    ws["F11"] = 1
    ws["G11"] = 1_000

    path = tmp_path / "suppliers.xlsx"
    wb.save(path)

    assert is_supplier_summary_report(path)
    detail, period = parse_supplier_summary_report_with_period(path)
    assert period[0].strftime("%d.%m.%Y") == "01.07.2026"
    assert set(detail["Поставщик"]) == {"Не указан", "Own production", "Taiwan"}
    assert detail["Количество"].sum() == 5
    assert detail["Выручка"].sum() == 50_000
    assert detail["Возврат количество"].sum() == 1
    assert detail["Возврат сумма"].sum() == 1_000
