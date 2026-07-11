from __future__ import annotations

import json
import os
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None

from openpyxl import load_workbook

from src.report import run_files, preview_source
from src.updater import UpdateError, check_for_update, download_installer, launch_installer

APP_NAME = "Аналитика"
APP_VERSION = "1.1.6"
COMPANY = "Princess Jewelry"
DEVELOPER = "Vladimir Panasyan"

BG = "#070707"
SIDEBAR = "#050505"
PANEL = "#111111"
CARD = "#181818"
CARD2 = "#202020"
BORDER = "#34302a"
GOLD = "#D7A441"
GOLD_DARK = "#9A6F1F"
TEXT = "#F5F5F5"
MUTED = "#B8B8B8"
GREEN = "#85D37A"
RED = "#E57373"


def app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_path(relative: str) -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / relative  # type: ignore[attr-defined]
    return app_base_dir() / relative


def user_data_dir() -> Path:
    """Persistent user files, outside the installed application folder."""
    documents = Path(os.environ.get("USERPROFILE", str(Path.home()))) / "Documents"
    base = documents / "Analitika"
    for name in ("Output", "History", "Settings", "Logs", "Updates", "Reports"):
        (base / name).mkdir(parents=True, exist_ok=True)
    return base


def settings_path() -> Path:
    return user_data_dir() / "Settings" / "settings.json"


def load_settings() -> dict:
    defaults = {"check_updates": True, "channel": "stable"}
    path = settings_path()
    try:
        if path.exists():
            saved = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                defaults.update(saved)
    except Exception:
        pass
    if defaults.get("channel") not in {"stable", "rc"}:
        defaults["channel"] = "stable"
    defaults["check_updates"] = bool(defaults.get("check_updates", True))
    return defaults


def save_settings(settings: dict) -> None:
    path = settings_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def error_log_path() -> Path:
    return user_data_dir() / "Logs" / "analitika_error.log"


def write_error_log(err: str) -> Path:
    p = error_log_path()
    with p.open("a", encoding="utf-8") as f:
        f.write("\n" + "=" * 90 + "\n")
        f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
        f.write(err + "\n")
    return p


def default_output_path() -> Path:
    base = user_data_dir() / "Output"
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    return base / f"Analitika_{stamp}.xlsx"


def period_to_text(period):
    if not period:
        return "не найден"
    start, end = period
    return f"{start:%d.%m.%Y} — {end:%d.%m.%Y}"


def preview_excel(path: Path) -> tuple[str, str]:
    """Preview both consolidated all-store exports and legacy store files.

    The filename is not used for consolidated reports: stores are detected from
    the workbook contents.
    """
    store, period = preview_source(path)
    return store, period_to_text(period)


class AnalitikaApp(tk.Tk):
    def __init__(self, initial_files: list[Path] | None = None):
        super().__init__()
        self.title(f"{APP_NAME} {APP_VERSION}")
        self.geometry("1280x760")
        self.minsize(1100, 680)
        self.configure(bg=BG)

        self.files: list[Path] = []
        self.file_meta: dict[Path, tuple[str, str, str]] = {}
        self.output_path: Path = default_output_path()
        self.last_output: Path | None = None
        self.is_running = False
        self._images = []
        self.settings = load_settings()

        self._load_assets()
        self._setup_style()
        self._build_ui()
        if self.settings.get("check_updates", True):
            self.after(1800, self._check_updates)
        self._set_app_icon()

        if initial_files:
            self.add_files(initial_files)

    def _load_assets(self):
        self.logo_image = None
        self.logo_small = None
        logo = resource_path("assets/logo.png")
        try:
            if logo.exists() and Image and ImageTk:
                img = Image.open(logo).convert("RGBA")
                img.thumbnail((240, 190), Image.LANCZOS)
                self.logo_image = ImageTk.PhotoImage(img)
                self._images.append(self.logo_image)

                img2 = Image.open(logo).convert("RGBA")
                img2.thumbnail((42, 42), Image.LANCZOS)
                self.logo_small = ImageTk.PhotoImage(img2)
                self._images.append(self.logo_small)
            elif logo.exists():
                self.logo_image = tk.PhotoImage(file=str(logo))
                self._images.append(self.logo_image)
        except Exception:
            self.logo_image = None
            self.logo_small = None

    def _set_app_icon(self):
        ico = resource_path("assets/analitika.ico")
        try:
            if ico.exists():
                self.iconbitmap(str(ico))
        except Exception:
            pass

    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            "Analitika.Treeview",
            rowheight=32,
            font=("Segoe UI", 10),
            background="#151515",
            fieldbackground="#151515",
            foreground=TEXT,
            bordercolor=BORDER,
            borderwidth=0,
        )
        style.configure(
            "Analitika.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#292929",
            foreground=TEXT,
            borderwidth=1,
            relief="flat",
        )
        style.map(
            "Analitika.Treeview",
            background=[("selected", "#3A2A10")],
            foreground=[("selected", TEXT)],
        )
        style.configure("TProgressbar", troughcolor="#303030", background=GOLD, bordercolor="#303030")

    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main()

    def _build_sidebar(self):
        sidebar = tk.Frame(self, bg=SIDEBAR, width=250, highlightbackground="#221a0e", highlightthickness=1)
        sidebar.grid(row=0, column=0, sticky="ns")
        sidebar.grid_propagate(False)
        sidebar.grid_rowconfigure(6, weight=1)

        logo_frame = tk.Frame(sidebar, bg=SIDEBAR)
        logo_frame.grid(row=0, column=0, sticky="ew", padx=16, pady=(26, 20))
        if self.logo_image:
            tk.Label(logo_frame, image=self.logo_image, bg=SIDEBAR).pack()
        else:
            tk.Label(logo_frame, text=COMPANY.upper(), bg=SIDEBAR, fg=TEXT, font=("Georgia", 15)).pack()

        self._side_button(sidebar, "▰  Аналитика", active=True, command=self.show_main_info).grid(row=1, column=0, padx=18, pady=(18, 8), sticky="ew")
        self._side_button(sidebar, "ⓘ  Об отчёте", command=self.show_about).grid(row=2, column=0, padx=18, pady=8, sticky="ew")
        self._side_button(sidebar, "⚙  Настройки", command=self.show_settings).grid(row=3, column=0, padx=18, pady=8, sticky="ew")

        bottom = tk.Frame(sidebar, bg=SIDEBAR, highlightbackground=GOLD_DARK, highlightthickness=1)
        bottom.grid(row=7, column=0, sticky="ew", padx=26, pady=24)
        tk.Label(bottom, text=f"{APP_NAME} {APP_VERSION}", bg=SIDEBAR, fg=GOLD, font=("Segoe UI", 10, "bold")).pack(pady=(14, 4))
        tk.Label(bottom, text="© Princess Jewelry", bg=SIDEBAR, fg=TEXT, font=("Segoe UI", 9)).pack()
        tk.Label(bottom, text="Все права защищены", bg=SIDEBAR, fg=MUTED, font=("Segoe UI", 9)).pack(pady=(2, 4))
        tk.Label(bottom, text=f"Разработка: {DEVELOPER}", bg=SIDEBAR, fg=GOLD, font=("Segoe UI", 8, "bold")).pack(pady=(0, 14))

    def _side_button(self, parent, text: str, active: bool = False, command=None):
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg="#201809" if active else SIDEBAR,
            fg=GOLD if active else MUTED,
            activebackground="#2a200c",
            activeforeground=GOLD,
            font=("Segoe UI", 11, "bold" if active else "normal"),
            padx=18,
            pady=14,
            anchor="w",
            relief="flat",
            bd=0,
            cursor="hand2",
        )

    def _build_main(self):
        main = tk.Frame(self, bg=BG)
        main.grid(row=0, column=1, sticky="nsew")
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(3, weight=1)

        header = tk.Frame(main, bg="#0d0d0d", height=118, highlightbackground="#252525", highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        header.grid_columnconfigure(1, weight=1)
        if self.logo_small:
            tk.Label(header, image=self.logo_small, bg="#0d0d0d").grid(row=0, column=0, rowspan=2, padx=(28, 14), pady=22, sticky="w")
        tk.Label(header, text="Аналитика", bg="#0d0d0d", fg=TEXT, font=("Segoe UI", 27, "bold")).grid(row=0, column=1, padx=(0, 20), pady=(24, 0), sticky="w")
        tk.Label(header, text="Формирование отчета по камням и жемчугу", bg="#0d0d0d", fg=GOLD, font=("Segoe UI", 13, "bold")).grid(row=1, column=1, padx=(0, 20), pady=(0, 22), sticky="w")
        tk.Label(header, text="◆  ◆  ◆", bg="#0d0d0d", fg=GOLD, font=("Segoe UI", 30)).grid(row=0, column=2, rowspan=2, padx=40, sticky="e")

        top_cards = tk.Frame(main, bg=BG)
        top_cards.grid(row=1, column=0, sticky="ew", padx=28, pady=22)
        top_cards.grid_columnconfigure(0, weight=1)
        top_cards.grid_columnconfigure(1, weight=1)
        top_cards.grid_columnconfigure(2, weight=0)
        self._select_card(top_cards, "▣", "Выбрать файлы", "Выберите Excel-файлы отчетов", "Выбрать файлы", self.select_files).grid(row=0, column=0, padx=(0, 10), sticky="ew")
        self._select_card(top_cards, "▣", "Выбрать папку", "Загрузить все файлы из Reports", "Выбрать папку", self.select_folder).grid(row=0, column=1, padx=10, sticky="ew")
        self._rules_card(top_cards).grid(row=0, column=2, padx=(10, 0), sticky="nsew")

        table_card = self._card(main)
        table_card.grid(row=2, column=0, sticky="nsew", padx=28, pady=(0, 16))
        table_card.grid_columnconfigure(0, weight=1)
        table_card.grid_rowconfigure(1, weight=1)
        self.table_title = tk.Label(table_card, text="Выбранные файлы (0)", bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold"))
        self.table_title.grid(row=0, column=0, padx=16, pady=(14, 8), sticky="w")
        cols = ("file", "store", "period", "rows", "status")
        self.tree = ttk.Treeview(table_card, columns=cols, show="headings", selectmode="extended", style="Analitika.Treeview", height=7)
        self.tree.heading("file", text="Файл")
        self.tree.heading("store", text="Магазин")
        self.tree.heading("period", text="Период")
        self.tree.heading("rows", text="Строк")
        self.tree.heading("status", text="Статус")
        self.tree.column("file", width=330, anchor="w")
        self.tree.column("store", width=100, anchor="center")
        self.tree.column("period", width=190, anchor="center")
        self.tree.column("rows", width=80, anchor="center")
        self.tree.column("status", width=120, anchor="center")
        self.tree.grid(row=1, column=0, padx=14, pady=(0, 0), sticky="nsew")
        self.tree.tag_configure("ok", foreground=GREEN)
        self.tree.tag_configure("bad", foreground=RED)
        self.summary_var = tk.StringVar(value="Файлов: 0     Магазинов: 0     Строк: 0     Период: —")
        tk.Label(table_card, textvariable=self.summary_var, bg=CARD, fg=MUTED, font=("Segoe UI", 10)).grid(row=2, column=0, padx=16, pady=12, sticky="w")

        bottom = tk.Frame(main, bg=BG)
        bottom.grid(row=3, column=0, sticky="nsew", padx=28, pady=(0, 20))
        bottom.grid_columnconfigure(0, weight=3)
        bottom.grid_columnconfigure(1, weight=2)
        bottom.grid_rowconfigure(0, weight=1)
        self._build_output_panel(bottom).grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self._build_log_panel(bottom).grid(row=0, column=1, sticky="nsew", padx=(10, 0))

    def _card(self, parent):
        return tk.Frame(parent, bg=CARD, highlightbackground=BORDER, highlightthickness=1)

    def _select_card(self, parent, icon, title, subtitle, btn_text, command):
        f = self._card(parent)
        f.grid_columnconfigure(1, weight=1)
        tk.Label(f, text=icon, bg=CARD, fg=GOLD, font=("Segoe UI", 28)).grid(row=0, column=0, rowspan=3, padx=(22, 16), pady=22)
        tk.Label(f, text=title, bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold")).grid(row=0, column=1, padx=(0, 18), pady=(22, 0), sticky="w")
        tk.Label(f, text=subtitle, bg=CARD, fg=MUTED, font=("Segoe UI", 9)).grid(row=1, column=1, padx=(0, 18), pady=(3, 14), sticky="w")
        self._button(f, btn_text, command).grid(row=2, column=1, padx=(0, 18), pady=(0, 18), sticky="w")
        return f

    def _rules_card(self, parent):
        f = self._card(parent)
        f.configure(width=260)
        f.grid_propagate(False)
        tk.Label(f, text="Готовый отчет", bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=16, pady=(16, 6))
        tk.Label(f, text="✓  Rules встроены", bg=CARD, fg=GREEN, font=("Segoe UI", 9)).pack(anchor="w", padx=16, pady=(0, 10))
        self.top_run_btn = self._button(f, "▶  Сформировать", self.generate_report, primary=True)
        self.top_run_btn.pack(anchor="w", fill="x", padx=16, pady=(0, 8))
        self._button(f, "Очистить список", self.clear_files).pack(anchor="w", fill="x", padx=16, pady=(0, 8))
        self._button(f, "Открыть Output", self.open_output_folder).pack(anchor="w", fill="x", padx=16, pady=(0, 16))
        return f

    def _build_output_panel(self, parent):
        f = self._card(parent)
        f.grid_columnconfigure(0, weight=1)
        tk.Label(f, text="Параметры сохранения отчёта", bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, padx=16, pady=(14, 12), sticky="w")
        tk.Label(f, text="Итоговый файл", bg=CARD, fg=MUTED, font=("Segoe UI", 9)).grid(row=1, column=0, padx=16, sticky="w")
        self.output_var = tk.StringVar(value=str(self.output_path))
        entry = tk.Entry(f, textvariable=self.output_var, bg="#111111", fg=TEXT, insertbackground=TEXT, relief="flat", font=("Segoe UI", 10))
        entry.grid(row=2, column=0, padx=16, pady=(5, 12), sticky="ew", ipady=8)
        self._button(f, "Обзор", self.select_output).grid(row=2, column=1, padx=(0, 16), pady=(5, 12), sticky="e")
        self.run_btn = self._button(f, "▶  Сформировать отчёт", self.generate_report, primary=True, big=True)
        self.run_btn.grid(row=3, column=0, padx=16, pady=(4, 16), sticky="w")
        self.status_var = tk.StringVar(value="Готов к работе")
        tk.Label(f, textvariable=self.status_var, bg=CARD, fg=TEXT, font=("Segoe UI", 10)).grid(row=3, column=0, padx=(275, 16), pady=(4, 16), sticky="w")
        self.progress = ttk.Progressbar(f, mode="indeterminate")
        self.progress.grid(row=4, column=0, columnspan=2, padx=16, pady=(0, 16), sticky="ew")
        return f

    def _build_log_panel(self, parent):
        f = self._card(parent)
        f.grid_columnconfigure(0, weight=1)
        f.grid_rowconfigure(1, weight=1)
        tk.Label(f, text="Лог", bg=CARD, fg=TEXT, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, padx=16, pady=(14, 8), sticky="w")
        self.log = tk.Text(f, height=8, bg="#0B0B0B", fg="#CFE8C8", insertbackground=TEXT, relief="flat", font=("Consolas", 9), wrap="word")
        self.log.grid(row=1, column=0, padx=16, pady=(0, 10), sticky="nsew")
        actions = tk.Frame(f, bg=CARD)
        actions.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 14))
        self._button(actions, "Открыть отчёт", self.open_last_output).pack(side="left", padx=(0, 10))
        self._button(actions, "Открыть Output", self.open_output_folder).pack(side="left")
        self._log("Готов к работе")
        return f

    def _button(self, parent, text, command, primary=False, big=False):
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg=GOLD if primary else "#151515",
            fg="#0a0a0a" if primary else TEXT,
            activebackground="#F0C15A" if primary else "#262626",
            activeforeground="#0a0a0a" if primary else TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=GOLD_DARK,
            font=("Segoe UI", 11 if big else 9, "bold" if primary else "normal"),
            padx=18 if big else 12,
            pady=12 if big else 7,
            cursor="hand2",
        )

    def _log(self, text: str):
        stamp = datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{stamp}] {text}\n")
        self.log.see("end")


    def _check_updates(self, manual: bool = False):
        def worker():
            try:
                info = check_for_update(
                    APP_VERSION, app_base_dir(), channel=self.settings.get("channel", "stable")
                )
                if info:
                    self.after(0, self._offer_update, info)
                elif manual:
                    self.after(0, lambda: messagebox.showinfo(
                        "Обновления", f"Установлена актуальная версия {APP_VERSION}."
                    ))
            except UpdateError as exc:
                if manual:
                    self.after(0, lambda: messagebox.showwarning("Обновления", str(exc)))
                else:
                    self.after(0, lambda: self._log(f"Проверка обновлений недоступна: {exc}"))
            except Exception as exc:
                self.after(0, lambda: self._log(f"Ошибка проверки обновлений: {exc}"))
        threading.Thread(target=worker, daemon=True).start()

    def _offer_update(self, info):
        notes = (info.get("notes") or "Исправления и улучшения.").strip()[:1000]
        release_kind = "тестовая версия" if info.get("prerelease") else "стабильное обновление"
        text = (
            f"Доступно {release_kind} Аналитики {info.get('version')}.\n\n"
            f"{notes}\n\n"
            "Скачать обновление и подготовить его к установке?"
        )
        if not info.get("download_url"):
            messagebox.showinfo(
                "Обновление Аналитики",
                text + "\n\nУстановщик не прикреплен к релизу. Откройте страницу релиза вручную.",
            )
            return
        if not messagebox.askyesno("Обновление Аналитики", text):
            return

        def worker():
            try:
                self.after(0, lambda: self.status_var.set("Скачиваю обновление..."))
                installer = download_installer(
                    info["download_url"],
                    user_data_dir() / "Updates",
                    expected_size=int(info.get("asset_size") or 0),
                    filename=info.get("asset_name"),
                )
                self.after(0, self._update_ready, installer, info.get("version"))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror(
                    APP_NAME, f"Не удалось подготовить обновление:\n{exc}"
                ))
                self.after(0, lambda: self.status_var.set("Готов к работе"))
        threading.Thread(target=worker, daemon=True).start()

    def _update_ready(self, installer: Path, version: str | None):
        self.status_var.set("Обновление готово к установке")
        if messagebox.askyesno(
            "Обновление готово",
            f"Версия {version or ''} скачана.\n\n"
            "Установить обновление сейчас? Приложение будет закрыто и запущено снова.",
        ):
            try:
                launch_installer(installer, silent=True)
                self.after(400, self.destroy)
            except Exception as exc:
                messagebox.showerror(APP_NAME, f"Не удалось запустить установщик:\n{exc}")

    def show_main_info(self):
        messagebox.showinfo(APP_NAME, "Главный экран: выберите Excel-файлы или папку с отчетами, затем нажмите золотую кнопку 'Сформировать'.")

    def show_about(self):
        messagebox.showinfo(
            "Об отчёте",
            "Аналитика формирует отчет по камням и жемчугу.\n\n"
            "Что делает программа:\n"
            "• определяет магазины по содержимому общей выгрузки;\n"
            "• читает период отчета;\n"
            "• группирует камни и жемчуг по правилам Princess Jewelry;\n"
            "• создает SUMMARY, отдельные листы найденных магазинов и COMPARE;\n"
            "• строит диаграммы выручки, количества и структуры продаж;\n"
            "• сохраняет итоговый Excel в выбранный путь.\n\n"
            f"Версия: {APP_VERSION}\nКанал: {self.settings.get('channel', 'stable').upper()}\n"
            f"Разработка: {DEVELOPER}"
        )

    def show_settings(self):
        win = tk.Toplevel(self)
        win.title("Настройки — Аналитика")
        win.geometry("570x430")
        win.resizable(False, False)
        win.configure(bg=BG)
        win.transient(self)
        win.grab_set()

        tk.Label(win, text="Настройки", bg=BG, fg=TEXT, font=("Segoe UI", 19, "bold")).pack(anchor="w", padx=28, pady=(24, 8))

        auto_var = tk.BooleanVar(value=self.settings.get("check_updates", True))
        channel_var = tk.StringVar(value=self.settings.get("channel", "stable"))

        tk.Checkbutton(
            win, text="Проверять обновления автоматически при запуске",
            variable=auto_var, bg=BG, fg=TEXT, selectcolor=CARD,
            activebackground=BG, activeforeground=TEXT, font=("Segoe UI", 10),
        ).pack(anchor="w", padx=28, pady=(8, 12))

        tk.Label(win, text="Канал обновлений", bg=BG, fg=GOLD, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=28)
        tk.Radiobutton(
            win, text="Стабильный — только проверенные релизы", value="stable",
            variable=channel_var, bg=BG, fg=TEXT, selectcolor=CARD,
            activebackground=BG, activeforeground=TEXT, font=("Segoe UI", 10),
        ).pack(anchor="w", padx=42, pady=(8, 2))
        tk.Radiobutton(
            win, text="Тестовый (RC) — предварительные и стабильные релизы", value="rc",
            variable=channel_var, bg=BG, fg=TEXT, selectcolor=CARD,
            activebackground=BG, activeforeground=TEXT, font=("Segoe UI", 10),
        ).pack(anchor="w", padx=42, pady=(2, 14))

        tk.Label(
            win,
            text=f"Версия: {APP_VERSION}\nПользовательские данные: {user_data_dir()}",
            bg=BG, fg=MUTED, justify="left", wraplength=500, font=("Segoe UI", 9),
        ).pack(anchor="w", padx=28, pady=(0, 16))

        def apply_settings():
            self.settings = {
                "check_updates": bool(auto_var.get()),
                "channel": channel_var.get(),
            }
            save_settings(self.settings)
            self._log(f"Настройки сохранены. Канал: {self.settings['channel'].upper()}")
            messagebox.showinfo(APP_NAME, "Настройки сохранены.", parent=win)

        buttons = tk.Frame(win, bg=BG)
        buttons.pack(fill="x", padx=28, pady=8)
        self._button(buttons, "Сохранить", apply_settings, primary=True).pack(side="left", padx=(0, 10))
        self._button(buttons, "Проверить обновления", lambda: self._check_updates(manual=True)).pack(side="left", padx=(0, 10))
        self._button(buttons, "Открыть папку данных", lambda: self.open_path(user_data_dir())).pack(side="left")

        tk.Label(
            win,
            text="Stable подходит руководителям. RC предназначен для тестирования новых возможностей. Отчеты и настройки не удаляются при обновлении.",
            bg=BG, fg=GREEN, wraplength=500, justify="left", font=("Segoe UI", 9),
        ).pack(anchor="w", padx=28, pady=(18, 0))

    def clear_files(self):
        if self.is_running:
            return
        self.files.clear()
        self.file_meta.clear()
        self.refresh_list()
        self._log("Список файлов очищен")

    def select_files(self):
        selected = filedialog.askopenfilenames(
            title="Выберите Excel-отчеты",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        self.add_files([Path(x) for x in selected])

    def select_folder(self):
        start = user_data_dir() / "Reports"
        start.mkdir(exist_ok=True)
        folder = filedialog.askdirectory(title="Выберите папку с отчетами", initialdir=str(start))
        if not folder:
            return
        p = Path(folder)
        files = [x for x in p.iterdir() if x.is_file() and x.suffix.lower() in {".xlsx", ".xls"} and not x.name.startswith("~$")]
        if not files:
            messagebox.showinfo(APP_NAME, "В выбранной папке нет Excel-файлов.")
            return
        self.add_files(files)

    def add_files(self, paths: list[Path]):
        added = 0
        for path in paths:
            path = Path(path)
            if path.suffix.lower() not in {".xlsx", ".xls"} or path.name.startswith("~$"):
                continue
            if path in self.files:
                continue
            try:
                store, period = preview_excel(path)
                status = "Готов"
            except Exception as e:
                store, period, status = "—", "—", "Ошибка"
                self._log(f"Не удалось распознать {path.name}: {e}")
            self.files.append(path)
            self.file_meta[path] = (store, period, status)
            added += 1
        self.refresh_list()
        if added:
            self._log(f"Файлы загружены: {added}")
            self._log("Нажмите золотую кнопку 'Сформировать' справа сверху")

    def refresh_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        stores = set()
        total_rows = 0
        periods = []
        for idx, path in enumerate(self.files, start=1):
            store, period, status = self.file_meta.get(path, ("—", "—", "—"))
            if store != "—":
                stores.add(store)
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                rows = wb.active.max_row
                wb.close()
            except Exception:
                rows = 0
            total_rows += rows
            if period not in {"—", "не найден"}:
                periods.append(period)
            tag = "ok" if status == "Готов" else "bad"
            self.tree.insert("", "end", iid=str(path), values=(path.name, store, period, f"{rows:,}".replace(",", " "), f"✓ {status}" if status == "Готов" else status), tags=(tag,))
        self.table_title.configure(text=f"Выбранные файлы ({len(self.files)})")
        period_summary = periods[0] if periods and len(set(periods)) == 1 else ("разные" if periods else "—")
        self.summary_var.set(f"Файлов: {len(self.files)}     Магазинов: {len(stores)}     Строк: {total_rows:,}     Период: {period_summary}".replace(",", " "))
        self.status_var.set(f"Выбрано файлов: {len(self.files)}")

    def select_output(self):
        selected = filedialog.asksaveasfilename(
            title="Куда сохранить итоговый отчет",
            defaultextension=".xlsx",
            initialdir=str(self.output_path.parent),
            initialfile=self.output_path.name,
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if selected:
            self.output_path = Path(selected)
            self.output_var.set(str(self.output_path))
            self.refresh_list()

    def generate_report(self):
        if self.is_running:
            return
        if not self.files:
            messagebox.showwarning(APP_NAME, "Сначала выберите Excel-отчеты.")
            return
        self.output_path = Path(self.output_var.get())
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.is_running = True
        self.run_btn.configure(state="disabled", text="Формирую...")
        if hasattr(self, "top_run_btn"):
            self.top_run_btn.configure(state="disabled", text="Формирую...")
        self.progress.start(12)
        self.status_var.set("Формирую отчет...")
        self._log("Запуск формирования отчета")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            output = run_files(self.files, self.output_path)
            self.after(0, self._success, output)
        except Exception:
            self.after(0, self._failure, traceback.format_exc())

    def _success(self, output: Path):
        self.is_running = False
        self.last_output = output
        self.progress.stop()
        self.run_btn.configure(state="normal", text="▶  Сформировать отчёт")
        if hasattr(self, "top_run_btn"):
            self.top_run_btn.configure(state="normal", text="▶  Сформировать")
        self.status_var.set(f"Готово: {output.name}")
        self._log(f"Отчет создан: {output}")
        messagebox.showinfo(APP_NAME, f"Отчет успешно создан:\n{output}")

    def _failure(self, err: str):
        self.is_running = False
        self.progress.stop()
        self.run_btn.configure(state="normal", text="▶  Сформировать отчёт")
        if hasattr(self, "top_run_btn"):
            self.top_run_btn.configure(state="normal", text="▶  Сформировать")
        self.status_var.set("Ошибка формирования отчета")
        p = write_error_log(err)
        self._log("Ошибка:\n" + err[-1500:])
        messagebox.showerror(APP_NAME, f"Не удалось сформировать отчет.\n\nЛог сохранен:\n{p}\n\n{err[-1800:]}")

    def open_last_output(self):
        if self.last_output and self.last_output.exists():
            self.open_path(self.last_output)
        else:
            messagebox.showinfo(APP_NAME, "Готовый отчет еще не создан.")

    def open_output_folder(self):
        folder = self.output_path.parent
        folder.mkdir(parents=True, exist_ok=True)
        self.open_path(folder)

    @staticmethod
    def open_path(path: Path):
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(path))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                os.system(f'open "{path}"')
            else:
                os.system(f'xdg-open "{path}"')
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Не удалось открыть:\n{path}\n\n{e}")


def main():
    initial = [Path(a) for a in sys.argv[1:] if Path(a).suffix.lower() in {".xlsx", ".xls"}]
    app = AnalitikaApp(initial_files=initial)
    app.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        err = traceback.format_exc()
        try:
            p = write_error_log(err)
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Аналитика — ошибка запуска", f"Программа не смогла запуститься.\n\nЛог ошибки сохранен:\n{p}\n\n{err[-2000:]}")
            root.destroy()
        except Exception:
            print(err)
        raise
